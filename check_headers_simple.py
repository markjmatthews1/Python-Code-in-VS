#!/usr/bin/env python3
"""
Check Excel Headers
Simple script to check current Excel file headers
"""

import pandas as pd
import openpyxl
import os

def check_headers():
    print("=== CHECKING EXCEL HEADERS ===\n")
    
    # Check Trade Tracker file headers
    try:
        wb = openpyxl.load_workbook('Bryan Perry Transactions.xlsx')
        ws = wb['Open_Trades_2025']
        print('📊 TRADE TRACKER HEADERS (Open_Trades_2025):')
        headers = []
        for col in range(1, ws.max_column + 1):
            header = ws.cell(row=1, column=col).value
            if header:
                headers.append(str(header))
            else:
                break
        for i, h in enumerate(headers, 1):
            print(f'  {i:2d}. {h}')
        
        print(f"\n  📈 Total columns: {len(headers)}")
        print(f"  📊 Data rows: {ws.max_row - 1}")
        wb.close()
    except Exception as e:
        print(f'❌ Trade tracker error: {e}')

    print()

    # Check dividend file headers  
    try:
        div_file = r'dividend_tracker\DividendTrackerApp\outputs\Dividends_2025.xlsx'
        if os.path.exists(div_file):
            wb = openpyxl.load_workbook(div_file)
            print('📊 DIVIDEND FILE SHEETS:')
            for sheet in wb.sheetnames:
                print(f'\n  📋 Sheet: {sheet}')
                ws = wb[sheet]
                headers = []
                for col in range(1, min(20, ws.max_column + 1)):
                    header = ws.cell(row=1, column=col).value
                    if header:
                        headers.append(str(header))
                if headers:
                    print(f'     Headers: {headers}')
                    print(f'     Dimensions: {ws.max_row} rows × {ws.max_column} columns')
            wb.close()
        else:
            print(f'📊 Dividend file not found at: {div_file}')
    except Exception as e:
        print(f'❌ Dividend file error: {e}')

if __name__ == "__main__":
    check_headers()
