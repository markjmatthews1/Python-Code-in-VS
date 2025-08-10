"""
repair_excel_header.py

This script will repair the header row (row 2) of the '2025 Results' sheet in 'Bryan Perry Transactions.xlsx'.
It will overwrite the header row with the correct column names, ensuring all columns are present and in the correct order.
"""

from openpyxl import load_workbook

EXCEL_FILENAME = "Bryan Perry Transactions.xlsx"
SHEET_NAME = "2025 Results"
HEADER_ROW = 2

CORRECT_HEADER = [
    "Stock Purchase Type", "Expiration Date", "Strike price", "Stock ticker", "Date of transaction",
    "Number of shares +/-", "Cost", "Total investment", "Assigned/exercised Y/N", "Call Value", "Put Value",
    "Put cash requirement", "Notes", "Close Date", "Close Price", "Realized profit loss", "P&L %", "Notes2",
    "Current Price", "Current P/L"
]

def repair_header():
    wb = load_workbook(EXCEL_FILENAME)
    ws = wb[SHEET_NAME]

    # Unmerge any merged cells in the header row
    merged_ranges = list(ws.merged_cells.ranges)
    for merged_range in merged_ranges:
        if merged_range.min_row == HEADER_ROW or merged_range.max_row == HEADER_ROW:
            ws.unmerge_cells(str(merged_range))

    # Clear all cells in the header row (row 2)
    max_col = ws.max_column
    for col_idx in range(1, max_col + 1):
        ws.cell(row=HEADER_ROW, column=col_idx, value=None)

    # Write the correct header
    for col_idx, col_name in enumerate(CORRECT_HEADER, start=1):
        ws.cell(row=HEADER_ROW, column=col_idx, value=col_name)

    # Remove any extra columns beyond the expected header
    if max_col > len(CORRECT_HEADER):
        for col_idx in range(len(CORRECT_HEADER) + 1, max_col + 1):
            for row in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=col_idx, max_col=col_idx):
                for cell in row:
                    cell.value = None

    wb.save(EXCEL_FILENAME)
    wb.close()
    print(f"Header row {HEADER_ROW} in '{SHEET_NAME}' repaired and cleaned in '{EXCEL_FILENAME}'.")

if __name__ == "__main__":
    repair_header()
