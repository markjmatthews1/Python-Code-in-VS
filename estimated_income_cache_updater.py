#!/usr/bin/env python3
"""
Estimated Income 2025 Cache-Based Updater

This script updates the Estimated Income 2025 sheet using cached portfolio data
instead of making direct API calls. It places data in account-specific rows
rather than appending to the bottom of the sheet.

Account Row Mapping:
- Row 4: E*TRADE IRA
- Row 5: E*TRADE Taxable  
- Row 6: Schwab IRA
- Row 7: Schwab Individual
"""

import os
import sys
import json
import openpyxl
from datetime import datetime, timedelta
import traceback
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.styles.numbers import FORMAT_CURRENCY_USD_SIMPLE

class EstimatedIncomeCacheUpdater:
    def __init__(self):
        """Initialize the cache-based updater"""
        self.main_dir = r"c:\Users\mjmat\Python Code in VS"
        self.app_dir = os.path.join(self.main_dir, "dividend_tracker", "DividendTrackerApp")
        self.cache_file = os.path.join(self.app_dir, "portfolio_data_cache.json")
        self.target_file = os.path.join(self.main_dir, "dividend_stocks.xlsx")
        self.ticker_yields_file = os.path.join(self.main_dir, "ticker_yields.json")
        
        # Account row mappings (1-based row numbers) - cache keys to display names and rows
        self.account_mapping = {
            'etrade_ira': {'display_name': 'E*TRADE IRA', 'row': 4},
            'etrade_taxable': {'display_name': 'E*TRADE Taxable', 'row': 5},
            'schwab_ira': {'display_name': 'Schwab IRA', 'row': 6},
            'schwab_individual': {'display_name': 'Schwab Individual', 'row': 7}
        }
        
        print(f"🎯 Estimated Income Cache Updater Initialized")
        print(f"📁 Main directory: {self.main_dir}")
        print(f"📊 Cache file: {self.cache_file}")
        print(f"📋 Target file: {self.target_file}")

    def load_cache(self):
        """Load portfolio data from cache"""
        try:
            if not os.path.exists(self.cache_file):
                print(f"❌ Cache file not found: {self.cache_file}")
                return None
                
            with open(self.cache_file, 'r') as f:
                cache_data = json.load(f)
                
            print(f"✅ Cache loaded successfully")
            print(f"📅 Cache timestamp: {cache_data.get('timestamp', 'Unknown')}")
            
            return cache_data
            
        except Exception as e:
            print(f"❌ Error loading cache: {e}")
            return None

    def is_cache_fresh(self, cache_data, max_age_hours=2):
        """Check if cache is fresh enough to use"""
        try:
            timestamp_str = cache_data.get('timestamp', '')
            if not timestamp_str:
                return False
                
            cache_time = datetime.fromisoformat(timestamp_str.replace('Z', '+00:00'))
            current_time = datetime.now()
            
            # Convert cache_time to local timezone for comparison
            if cache_time.tzinfo:
                import pytz
                local_tz = pytz.timezone('America/New_York')  # Adjust to your timezone
                cache_time = cache_time.astimezone(local_tz).replace(tzinfo=None)
                
            age_hours = (current_time - cache_time).total_seconds() / 3600
            
            print(f"⏰ Cache age: {age_hours:.1f} hours")
            
            return age_hours <= max_age_hours
            
        except Exception as e:
            print(f"⚠️ Error checking cache freshness: {e}")
            return False

    def load_ticker_yields(self):
        """Load ticker yield data"""
        try:
            if not os.path.exists(self.ticker_yields_file):
                print(f"❌ Ticker yields file not found: {self.ticker_yields_file}")
                return {}
                
            with open(self.ticker_yields_file, 'r') as f:
                yields_data = json.load(f)
                
            print(f"✅ Loaded {len(yields_data)} ticker yields")
            return yields_data
            
        except Exception as e:
            print(f"❌ Error loading ticker yields: {e}")
            return {}

    def calculate_account_dividend_income(self, positions, ticker_yields):
        """
        Calculate dividend income for an account's positions
        
        Args:
            positions (list): List of position dictionaries
            ticker_yields (dict): Ticker yield data
            
        Returns:
            dict: Account dividend summary
        """
        account_summary = {
            'tickers': {},
            'total_yearly_dividend': 0,
            'dividend_tickers_count': 0,
            'total_position_value': 0
        }
        
        for position in positions:
            ticker = position.get('symbol', '').upper()  # Fixed: use 'symbol' not 'ticker'
            quantity = float(position.get('quantity', 0))
            
            # Handle both 'price' and 'current_price' fields - some positions may have different structures
            current_price = float(position.get('price', 0))
            if current_price <= 0:
                # Try getting from market_value/quantity if price is 0
                market_value = float(position.get('market_value', 0))
                if market_value > 0 and quantity > 0:
                    current_price = market_value / quantity
            
            if not ticker or quantity <= 0 or current_price <= 0:
                continue
                
            position_value = quantity * current_price
            account_summary['total_position_value'] += position_value
            
            # Get yield data
            yield_data = ticker_yields.get(ticker, {})
            annual_yield_percent = yield_data.get('yield', 0)
            
            # Calculate dividend income
            has_dividend = annual_yield_percent > 0
            annual_dividend_per_share = (current_price * annual_yield_percent / 100) if has_dividend else 0
            total_annual_dividend = annual_dividend_per_share * quantity
            
            if has_dividend:
                account_summary['dividend_tickers_count'] += 1
                account_summary['total_yearly_dividend'] += total_annual_dividend
            
            # Store ticker details
            account_summary['tickers'][ticker] = {
                'quantity': quantity,
                'current_price': current_price,
                'position_value': position_value,
                'yield_percent': annual_yield_percent,
                'annual_dividend_per_share': annual_dividend_per_share,
                'total_annual_dividend': total_annual_dividend,
                'has_dividend': has_dividend,
                'monthly_dividend': total_annual_dividend / 12
            }
        
        return account_summary

    def update_estimated_income_sheet_with_cache(self):
        """
        Update the Estimated Income 2025 sheet using cached data and proper row placement
        """
        try:
            print(f"\n💰 UPDATING ESTIMATED INCOME 2025 SHEET FROM CACHE")
            print("=" * 70)
            
            # Load cache data
            cache_data = self.load_cache()
            if not cache_data:
                print("❌ Cannot proceed without cache data")
                return False
                
            # Check cache freshness
            if not self.is_cache_fresh(cache_data):
                print("⚠️ Cache is stale, but proceeding anyway")
            
            # Load ticker yields - use cache yields first, then fallback to file
            ticker_yields = cache_data.get('ticker_yields', {})
            if not ticker_yields:
                print("⚠️ No ticker yields in cache, trying file...")
                ticker_yields = self.load_ticker_yields()
            
            if not ticker_yields:
                print("⚠️ No ticker yield data available")
            else:
                print(f"📋 Available tickers with yields: {list(ticker_yields.keys())}")
                # Show some example yields
                sample_tickers = list(ticker_yields.keys())[:5]
                for ticker in sample_tickers:
                    yield_pct = ticker_yields[ticker].get('yield', 0)
                    print(f"   {ticker}: {yield_pct:.2f}%")
                
            # Load target workbook
            if not os.path.exists(self.target_file):
                print(f"❌ Target file not found: {self.target_file}")
                return False
                
            workbook = openpyxl.load_workbook(self.target_file)
            
            # Check if Estimated Income 2025 sheet exists
            if 'Estimated Income 2025' not in workbook.sheetnames:
                print(f"⚠️ Creating 'Estimated Income 2025' sheet...")
                self.create_estimated_income_sheet(workbook)
                
            sheet = workbook['Estimated Income 2025']
            print(f"✅ Found 'Estimated Income 2025' sheet")
            
            # Update timestamp
            current_date = datetime.now().strftime("%m/%d/%Y")
            current_time = datetime.now().strftime("%H:%M")
            sheet['B2'] = f"{current_date} {current_time}"
            
            # Process each account and place data in specific rows
            all_account_data = {}
            total_portfolio_yearly = 0
            
            for cache_key, account_info in self.account_mapping.items():
                display_name = account_info['display_name']
                target_row = account_info['row']
                
                print(f"\n📊 Processing {display_name} (Row {target_row})...")
                print("-" * 50)
                
                # Get account data from cache
                positions = cache_data.get('positions', {}).get(cache_key, [])
                
                if positions:
                    # Calculate dividend income
                    dividend_data = self.calculate_account_dividend_income(positions, ticker_yields)
                    all_account_data[display_name] = dividend_data
                    
                    yearly_income = dividend_data.get('total_yearly_dividend', 0)
                    monthly_income = yearly_income / 12
                    total_portfolio_yearly += yearly_income
                    
                    # Show position details for debugging
                    position_count = len(positions)
                    div_tickers = dividend_data.get('dividend_tickers_count', 0)
                    print(f"   📊 {position_count} positions, {div_tickers} with dividends")
                    
                    # Place data in specific row for this account
                    self.update_account_row(sheet, target_row, display_name, dividend_data)
                    
                    print(f"✅ {display_name}: ${yearly_income:,.2f} yearly (${monthly_income:,.2f} monthly)")
                    print(f"   📍 Data placed in row {target_row}")
                    
                else:
                    print(f"⚠️ No positions found for {display_name}")
                    # Clear the row for this account
                    self.clear_account_row(sheet, target_row, display_name)
            
            # Update total monthly income in row 9
            total_monthly = total_portfolio_yearly / 12
            sheet.cell(row=9, column=7).value = "TOTAL MONTHLY INCOME:"
            sheet.cell(row=9, column=7).font = Font(name='Arial', size=12, bold=True)
            sheet.cell(row=9, column=9).value = total_monthly
            sheet.cell(row=9, column=9).number_format = FORMAT_CURRENCY_USD_SIMPLE
            sheet.cell(row=9, column=9).font = Font(name='Arial', size=12, bold=True)
            
            # Save workbook
            workbook.save(self.target_file)
            workbook.close()
            
            print(f"\n🎉 ESTIMATED INCOME 2025 SHEET UPDATED SUCCESSFULLY!")
            print(f"   📈 Total Yearly Dividend Income: ${total_portfolio_yearly:,.2f}")
            print(f"   📅 Total Monthly Dividend Income: ${total_monthly:,.2f}")
            
            return True
            
        except Exception as e:
            print(f"❌ Error updating Estimated Income sheet: {e}")
            print(f"Traceback: {traceback.format_exc()}")
            return False

    def update_account_row(self, sheet, row, account_name, dividend_data):
        """
        Update a specific row with account dividend data
        
        Args:
            sheet: Excel worksheet
            row (int): Target row number
            account_name (str): Account name
            dividend_data (dict): Dividend calculation results
        """
        try:
            # Account name in column A
            sheet.cell(row=row, column=1).value = account_name
            sheet.cell(row=row, column=1).font = Font(name='Arial', size=11, bold=True)
            
            # Total position value in column B (if available)
            total_value = dividend_data.get('total_position_value', 0)
            if total_value > 0:
                sheet.cell(row=row, column=2).value = total_value
                sheet.cell(row=row, column=2).number_format = FORMAT_CURRENCY_USD_SIMPLE
            
            # Number of dividend-paying tickers in column C
            div_count = dividend_data.get('dividend_tickers_count', 0)
            sheet.cell(row=row, column=3).value = div_count
            
            # Yearly dividend income in column H
            yearly_dividend = dividend_data.get('total_yearly_dividend', 0)
            sheet.cell(row=row, column=8).value = yearly_dividend
            sheet.cell(row=row, column=8).number_format = FORMAT_CURRENCY_USD_SIMPLE
            if yearly_dividend > 0:
                sheet.cell(row=row, column=8).font = Font(name='Arial', size=10, bold=True)
            
            # Monthly dividend income in column I
            monthly_dividend = yearly_dividend / 12
            sheet.cell(row=row, column=9).value = monthly_dividend
            sheet.cell(row=row, column=9).number_format = FORMAT_CURRENCY_USD_SIMPLE
            if monthly_dividend > 0:
                sheet.cell(row=row, column=9).font = Font(name='Arial', size=10, bold=True)
                
        except Exception as e:
            print(f"⚠️ Error updating row {row} for {account_name}: {e}")

    def clear_account_row(self, sheet, row, account_name):
        """Clear a specific account row when no data is available"""
        try:
            # Keep account name but clear values
            sheet.cell(row=row, column=1).value = account_name
            sheet.cell(row=row, column=1).font = Font(name='Arial', size=11, bold=True)
            
            # Clear other columns
            for col in range(2, 10):
                sheet.cell(row=row, column=col).value = None
                
        except Exception as e:
            print(f"⚠️ Error clearing row {row} for {account_name}: {e}")

    def create_estimated_income_sheet(self, workbook):
        """Create the Estimated Income 2025 sheet with proper structure"""
        print("🏗️ Creating Estimated Income 2025 sheet...")
        
        sheet = workbook.create_sheet("Estimated Income 2025")
        
        # Set up headers
        sheet['A1'] = '📊 ESTIMATED DIVIDEND INCOME 2025'
        sheet['A1'].font = Font(name='Arial', size=14, bold=True)
        
        sheet['A2'] = 'Updated:'
        sheet['B2'] = datetime.now().strftime("%m/%d/%Y %H:%M")
        
        sheet['A3'] = ''  # Empty row
        
        # Column headers
        headers = [
            'Account', 'Total Value', 'Div Tickers', 'Avg Yield %', 'Notes',
            '', '', 'Annual Dividend', 'Monthly Dividend'
        ]
        
        for col, header in enumerate(headers, 1):
            if header:  # Skip empty headers
                cell = sheet.cell(row=3, column=col)
                cell.value = header
                cell.font = Font(name='Arial', size=11, bold=True, color='FFFFFF')
                cell.fill = PatternFill(start_color='4F81BD', end_color='4F81BD', fill_type='solid')
                cell.alignment = Alignment(horizontal='center')
        
        print("✅ Estimated Income 2025 sheet structure created")

def main():
    """Main execution function"""
    print("🚀 Starting Estimated Income Cache Update...")
    print("=" * 60)
    
    updater = EstimatedIncomeCacheUpdater()
    
    # Update the sheet
    success = updater.update_estimated_income_sheet_with_cache()
    
    if success:
        print("\n🎊 Update completed successfully!")
    else:
        print("\n💥 Update failed!")
        sys.exit(1)

if __name__ == "__main__":
    main()
