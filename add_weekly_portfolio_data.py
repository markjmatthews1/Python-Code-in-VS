#!/usr/bin/env python3
"""
Check Portfolio Values 2025 Sheet and Add This Week's Column
"""

import pandas as pd
import numpy as np
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
from datetime import datetime

def check_portfolio_sheet():
    """Check what's currently in Portfolio Values 2025"""
    
    excel_file = 'dividend_tracker/DividendTrackerApp/outputs/Dividends_2025.xlsx'
    
    try:
        # Load and examine the sheet
        wb = load_workbook(excel_file)
        ws = wb['Portfolio Values 2025']
        
        print("📊 Current Portfolio Values 2025 Sheet:")
        print(f"   Max column: {ws.max_column}")
        print(f"   Max row: {ws.max_row}")
        
        # Check the structure - look at rows 3-10 and the last few columns
        last_col = ws.max_column
        print(f"\n📋 Sheet Structure (Last 3 columns):")
        
        for col in range(max(1, last_col-2), last_col+1):
            col_letter = ws.cell(row=1, column=col).column_letter
            print(f"\n   Column {col_letter} (#{col}):")
            for row in [3, 4, 5, 6, 7, 8, 10]:
                cell_value = ws.cell(row=row, column=col).value
                if cell_value is not None:
                    if isinstance(cell_value, (int, float)):
                        print(f"     Row {row}: ${cell_value:,.2f}" if row != 3 else f"     Row {row}: {cell_value}")
                    else:
                        print(f"     Row {row}: {cell_value}")
                else:
                    print(f"     Row {row}: (empty)")
        
        wb.close()
        return last_col
        
    except Exception as e:
        print(f"❌ Error checking sheet: {e}")
        return 0

def add_this_weeks_data():
    """Add column for this week (08/23/2025) with current portfolio values"""
    
    excel_file = 'dividend_tracker/DividendTrackerApp/outputs/Dividends_2025.xlsx'
    
    try:
        wb = load_workbook(excel_file)
        ws = wb['Portfolio Values 2025']
        
        # Find the next available column
        next_col = ws.max_column + 1
        col_letter = ws.cell(row=1, column=next_col).column_letter
        
        print(f"\n💰 Adding this week's data to column {col_letter} (#{next_col}):")
        
        # This week's data (08/23/2025)
        today_date = "08/23/2025"
        
        # Current portfolio values using your Schwab data
        etrade_ira = 279339.15      # From API
        etrade_taxable = 62622.72   # From API  
        schwab_ira = 50558.40      # Your manual data
        schwab_individual = 2603.64 # Your manual data
        retirement_401k = 124315.15 # Manual entry
        
        # Calculate total
        total_portfolio = etrade_ira + etrade_taxable + schwab_ira + schwab_individual + retirement_401k
        
        print(f"   Row 3 (Date): {today_date}")
        print(f"   Row 4 (E*TRADE IRA): ${etrade_ira:,.2f}")
        print(f"   Row 5 (E*TRADE Taxable): ${etrade_taxable:,.2f}")
        print(f"   Row 6 (Schwab IRA): ${schwab_ira:,.2f}")
        print(f"   Row 7 (Schwab Individual): ${schwab_individual:,.2f}")
        print(f"   Row 8 (401k): ${retirement_401k:,.2f}")
        print(f"   Row 10 (Total): ${total_portfolio:,.2f}")
        
        # Add the data to the sheet
        ws.cell(row=3, column=next_col, value=today_date)        # Date
        ws.cell(row=4, column=next_col, value=etrade_ira)        # E*TRADE IRA  
        ws.cell(row=5, column=next_col, value=etrade_taxable)    # E*TRADE Taxable
        ws.cell(row=6, column=next_col, value=schwab_ira)        # Schwab IRA
        ws.cell(row=7, column=next_col, value=schwab_individual) # Schwab Individual
        ws.cell(row=8, column=next_col, value=retirement_401k)   # 401k
        ws.cell(row=10, column=next_col, value=total_portfolio)  # Total
        
        # Format as currency (except date row)
        for row in [4, 5, 6, 7, 8, 10]:
            ws.cell(row=row, column=next_col).number_format = '$#,##0.00'
        
        # Format date with bold font
        ws.cell(row=3, column=next_col).font = Font(name='Arial', size=12, bold=True)
        
        # Format value cells with Arial 12
        for row in [4, 5, 6, 7, 8, 10]:
            ws.cell(row=row, column=next_col).font = Font(name='Arial', size=12)
        
        # Save the workbook
        wb.save(excel_file)
        wb.close()
        
        print(f"\n✅ Successfully added column {col_letter} to Portfolio Values 2025")
        print(f"📈 Total Portfolio Value: ${total_portfolio:,.2f}")
        
        return total_portfolio, col_letter
        
    except Exception as e:
        print(f"❌ Error adding this week's data: {e}")
        import traceback
        traceback.print_exc()
        return 0, ""

def main():
    """Check current sheet and add this week's data"""
    
    print("🎯 Portfolio Values 2025 - Add This Week's Data")
    print("="*60)
    
    # First check what's in the sheet
    current_max_col = check_portfolio_sheet()
    
    if current_max_col > 0:
        # Add this week's data
        total, col = add_this_weeks_data()
        
        if total > 0:
            print(f"\n🎉 SUCCESS!")
            print(f"   • Added column {col} with 08/23/2025 data")
            print(f"   • Portfolio total: ${total:,.2f}")
            print(f"   • Ready to update Portfolio Summary sheet")
        else:
            print("\n❌ Failed to add this week's data")
    else:
        print("\n❌ Could not read Portfolio Values 2025 sheet")

if __name__ == "__main__":
    main()
