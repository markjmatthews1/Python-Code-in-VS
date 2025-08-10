#!/usr/bin/env python3
"""Show all SOFI trades with details"""

from openpyxl import load_workbook

def show_sofi_trades():
    excel_filename = "Bryan Perry Transactions.xlsx"
    sheet_name = "Open_Trades_2025"
    
    print("All SOFI trades in Excel:")
    print("=" * 80)
    
    wb = load_workbook(excel_filename)
    ws = wb[sheet_name]
    header = [cell.value for cell in ws[1]]
    
    # Column positions
    ticker_col = header.index("Stock ticker") + 1
    type_col = header.index("Stock Purchase Type") + 1
    strike_col = header.index("Strike price") + 1
    expiry_col = header.index("Expiration Date") + 1
    cost_col = header.index("Cost") + 1
    shares_col = header.index("Number of shares +/-") + 1
    current_price_col = header.index("Current Price") + 1
    current_pl_col = header.index("Current P/L") + 1
    
    sofi_count = 0
    for r in range(2, ws.max_row + 1):
        ticker = ws.cell(row=r, column=ticker_col).value
        if ticker == "SOFI":
            sofi_count += 1
            trade_type = ws.cell(row=r, column=type_col).value
            strike = ws.cell(row=r, column=strike_col).value
            expiry = ws.cell(row=r, column=expiry_col).value
            cost = ws.cell(row=r, column=cost_col).value
            shares = ws.cell(row=r, column=shares_col).value
            current_price = ws.cell(row=r, column=current_price_col).value
            current_pl = ws.cell(row=r, column=current_pl_col).value
            
            print(f"Row {r}: SOFI {trade_type}")
            print(f"  Strike: ${strike}, Expiry: {expiry}")
            print(f"  Cost: ${cost}, Shares: {shares}")
            print(f"  Current Price: ${current_price}")
            print(f"  Current P/L: ${current_pl}")
            print()
    
    print(f"Total SOFI trades found: {sofi_count}")
    wb.close()

if __name__ == "__main__":
    show_sofi_trades()
