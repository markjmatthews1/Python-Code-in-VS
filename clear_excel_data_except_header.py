"""
clear_excel_data_except_header.py

This script will backup the Excel file, then clear all rows except the header (row 2)
in the '2025 Results' sheet of 'Bryan Perry Transactions.xlsx'.
"""

import shutil
from openpyxl import load_workbook

EXCEL_FILENAME = "Bryan Perry Transactions.xlsx"
BACKUP_FILENAME = "Bryan Perry Transactions_BACKUP_BEFORE_CLEAR.xlsx"
SHEET_NAME = "2025 Results"
HEADER_ROW = 2

# Backup the file
shutil.copy(EXCEL_FILENAME, BACKUP_FILENAME)
print(f"Backup created: {BACKUP_FILENAME}")

# Load workbook and worksheet
wb = load_workbook(EXCEL_FILENAME)
ws = wb[SHEET_NAME]

# Delete all rows except header (row 2)
max_row = ws.max_row
if max_row > HEADER_ROW:
    ws.delete_rows(HEADER_ROW + 1, max_row - HEADER_ROW)
    print(f"Cleared all rows below header (row {HEADER_ROW}) in '{SHEET_NAME}'.")
else:
    print("No data rows to clear.")

wb.save(EXCEL_FILENAME)
wb.close()
print(f"Data cleared. Please try adding a new trade to a clean sheet.")
