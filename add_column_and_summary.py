#!/usr/bin/env python3
"""
Add Column AK to Portfolio Values 2025 and Create Portfolio Summary Sheet
"""

import pandas as pd
import numpy as np
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
from datetime import datetime

def add_portfolio_column_ak():
    """Add new column AK with 08/23/2025 data"""
    
    excel_file = 'dividend_tracker/DividendTrackerApp/outputs/Dividends_2025.xlsx'
    
    try:
        # Load the workbook
        wb = load_workbook(excel_file)
        ws = wb['Portfolio Values 2025']
        
        print("📊 Adding column AK to Portfolio Values 2025...")
        
        # Column AK is column 37
        ak_col = 37
        
        # Today's date and values
        today_date = "08/23/2025"
        
        # Current portfolio values
        etrade_ira = 279339.15
        etrade_taxable = 62622.72
        schwab_ira = 50558.40      # Your provided data
        schwab_individual = 2603.64 # Your provided data  
        retirement_401k = 124315.15
        
        # Calculate total
        total_portfolio = etrade_ira + etrade_taxable + schwab_ira + schwab_individual + retirement_401k
        
        # Add data to column AK following the row structure
        ws.cell(row=3, column=ak_col, value=today_date)  # Date header
        ws.cell(row=4, column=ak_col, value=etrade_ira)  # E*TRADE IRA
        ws.cell(row=5, column=ak_col, value=etrade_taxable)  # E*TRADE Taxable
        ws.cell(row=6, column=ak_col, value=schwab_ira)  # Schwab IRA
        ws.cell(row=7, column=ak_col, value=schwab_individual)  # Schwab Individual
        ws.cell(row=8, column=ak_col, value=retirement_401k)  # 401k
        ws.cell(row=10, column=ak_col, value=total_portfolio)  # Total
        
        # Format currency for value rows
        for row in [4, 5, 6, 7, 8, 10]:
            ws.cell(row=row, column=ak_col).number_format = '$#,##0.00'
        
        # Format date header
        ws.cell(row=3, column=ak_col).font = Font(bold=True)
        
        print(f"✅ Added column AK with data for {today_date}")
        print(f"   Total Portfolio: ${total_portfolio:,.2f}")
        
        return wb, total_portfolio
        
    except Exception as e:
        print(f"❌ Error adding column: {e}")
        return None, 0

def create_portfolio_summary_sheet(wb, current_total):
    """Create a comprehensive Portfolio Summary sheet"""
    
    try:
        # Create or get the Portfolio Summary sheet
        if 'Portfolio Summary' in wb.sheetnames:
            ws = wb['Portfolio Summary']
            ws.delete_rows(1, ws.max_row)  # Clear existing data
        else:
            ws = wb.create_sheet('Portfolio Summary')
        
        print("📋 Creating Portfolio Summary sheet...")
        
        # Title
        ws['A1'] = 'PORTFOLIO SUMMARY - 2025'
        ws['A1'].font = Font(size=16, bold=True)
        ws['A1'].fill = PatternFill(start_color='366092', end_color='366092', fill_type='solid')
        ws['A1'].font = Font(size=16, bold=True, color='FFFFFF')
        
        # Current Portfolio Value Section
        ws['A3'] = 'CURRENT PORTFOLIO VALUE (08/23/2025)'
        ws['A3'].font = Font(size=12, bold=True)
        ws['A3'].fill = PatternFill(start_color='D9E1F2', end_color='D9E1F2', fill_type='solid')
        
        # Account breakdown
        accounts = [
            ('E*TRADE IRA:', 279339.15),
            ('E*TRADE Taxable:', 62622.72),
            ('Schwab IRA:', 50558.40),
            ('Schwab Individual:', 2603.64),
            ('401k Retirement:', 124315.15)
        ]
        
        row = 4
        for account, value in accounts:
            ws[f'A{row}'] = account
            ws[f'B{row}'] = value
            ws[f'B{row}'].number_format = '$#,##0.00'
            row += 1
        
        # Total
        ws[f'A{row}'] = 'TOTAL PORTFOLIO:'
        ws[f'A{row}'].font = Font(bold=True)
        ws[f'B{row}'] = current_total
        ws[f'B{row}'].number_format = '$#,##0.00'
        ws[f'B{row}'].font = Font(bold=True)
        ws[f'B{row}'].fill = PatternFill(start_color='92D050', end_color='92D050', fill_type='solid')
        
        # Portfolio Allocation
        ws['A11'] = 'PORTFOLIO ALLOCATION'
        ws['A11'].font = Font(size=12, bold=True)
        ws['A11'].fill = PatternFill(start_color='D9E1F2', end_color='D9E1F2', fill_type='solid')
        
        row = 12
        for account, value in accounts:
            percentage = (value / current_total) * 100
            ws[f'A{row}'] = account
            ws[f'B{row}'] = f"{percentage:.1f}%"
            row += 1
        
        # Recent Performance (if we can calculate from last column)
        try:
            portfolio_ws = wb['Portfolio Values 2025']
            # Get last two total values to calculate change
            last_total = portfolio_ws.cell(row=10, column=36).value  # Column AJ
            if last_total and isinstance(last_total, (int, float)):
                change = current_total - last_total
                change_pct = (change / last_total) * 100 if last_total != 0 else 0
                
                ws['A18'] = 'RECENT PERFORMANCE'
                ws['A18'].font = Font(size=12, bold=True)
                ws['A18'].fill = PatternFill(start_color='D9E1F2', end_color='D9E1F2', fill_type='solid')
                
                ws['A19'] = 'Weekly Change:'
                ws['B19'] = change
                ws['B19'].number_format = '$#,##0.00'
                if change >= 0:
                    ws['B19'].font = Font(color='00B050')
                else:
                    ws['B19'].font = Font(color='FF0000')
                
                ws['A20'] = 'Weekly % Change:'
                ws['B20'] = f"{change_pct:.2f}%"
                if change_pct >= 0:
                    ws['B20'].font = Font(color='00B050')
                else:
                    ws['B20'].font = Font(color='FF0000')
        except:
            print("   Could not calculate recent performance")
        
        # Dividend Summary Section
        ws['A22'] = 'DIVIDEND SUMMARY'
        ws['A22'].font = Font(size=12, bold=True)
        ws['A22'].fill = PatternFill(start_color='D9E1F2', end_color='D9E1F2', fill_type='solid')
        
        ws['A23'] = 'Annual Dividend Goal:'
        ws['B23'] = 'Check Estimated Income 2025 sheet'
        
        ws['A24'] = 'YTD Dividends Received:'
        ws['B24'] = 'Check All account weekly dividends sheet'
        
        # Formatting
        for col in ['A', 'B']:
            ws.column_dimensions[col].width = 25 if col == 'A' else 15
        
        # Move Portfolio Summary to be near the beginning
        try:
            sheets = wb.sheetnames
            if 'Portfolio Summary' in sheets and len(sheets) > 1:
                wb._sheets.insert(1, wb._sheets.pop(wb._sheets.index(wb['Portfolio Summary'])))
        except:
            pass  # If moving fails, just leave it where it is
        
        print("✅ Portfolio Summary sheet created with:")
        print(f"   • Current portfolio breakdown")
        print(f"   • Account allocations")
        print(f"   • Recent performance tracking")
        print(f"   • Dividend summary links")
        
        return True
        
    except Exception as e:
        print(f"❌ Error creating summary sheet: {e}")
        import traceback
        traceback.print_exc()
        return False

def main():
    """Add column AK and create portfolio summary sheet"""
    
    print("🎯 Adding Portfolio Column AK & Creating Summary Sheet")
    print("="*60)
    
    # Add the new column
    wb, total = add_portfolio_column_ak()
    
    if wb and total > 0:
        # Create the summary sheet
        if create_portfolio_summary_sheet(wb, total):
            # Save the workbook
            try:
                wb.save('dividend_tracker/DividendTrackerApp/outputs/Dividends_2025.xlsx')
                wb.close()
                
                print("\n✅ SUCCESS!")
                print("   • Added column AK to Portfolio Values 2025")
                print("   • Created Portfolio Summary sheet")
                print(f"   • Total Portfolio Value: ${total:,.2f}")
                print("\n📊 The Portfolio Summary sheet now provides:")
                print("   • Current account balances and allocations")
                print("   • Recent performance tracking")
                print("   • Links to dividend data in other sheets")
                
            except Exception as e:
                print(f"❌ Error saving workbook: {e}")
        else:
            print("❌ Failed to create summary sheet")
    else:
        print("❌ Failed to add portfolio column")

if __name__ == "__main__":
    main()
