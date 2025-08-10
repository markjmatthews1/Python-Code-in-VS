#!/usr/bin/env python3
"""
Add Historical Yield Data to Ticker Analysis 2025 Sheet
Integrates historical yield progression directly into existing ticker analysis
"""

import pandas as pd
import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment
from openpyxl.utils.dataframe import dataframe_to_rows
import os
from datetime import datetime, timedelta
import yfinance as yf

def load_historical_data():
    """Load historical dividend data from CSV files"""
    historical_data = {}
    
    # Look for historical CSV files
    csv_files = [
        'fmp_daily_results_20250624.csv',
        'fmp_daily_results_20250625.csv', 
        'fmp_daily_results_20250626.csv',
        'fmp_daily_results_20250627.csv',
        'fmp_daily_results_20250701.csv',
        'fmp_daily_results_20250702.csv',
        'fmp_daily_results_20250703.csv',
        'fmp_daily_results_20250704.csv',
        'fmp_daily_results_20250705.csv'
    ]
    
    for csv_file in csv_files:
        if os.path.exists(csv_file):
            try:
                df = pd.read_csv(csv_file)
                # Extract date from filename (YYYYMMDD format)
                date_str = csv_file.split('_')[-1].replace('.csv', '')
                date = datetime.strptime(date_str, '%Y%m%d').strftime('%m/%d')
                
                print(f"✅ Loading historical data from {csv_file} for date {date}")
                
                # Store yield data by ticker
                for _, row in df.iterrows():
                    ticker = row.get('Ticker', row.get('Symbol', ''))
                    yield_val = row.get('Yield', row.get('Dividend Yield', 0))
                    
                    if ticker and yield_val:
                        if ticker not in historical_data:
                            historical_data[ticker] = {}
                        historical_data[ticker][date] = float(yield_val)
                        
            except Exception as e:
                print(f"⚠️ Could not load {csv_file}: {e}")
    
    return historical_data

def add_historical_yield_columns():
    """Add historical yield progression columns to Ticker Analysis 2025"""
    
    # Load the Excel file
    excel_file = r"c:\Users\mjmat\Python Code in VS\dividend_tracker\DividendTrackerApp\outputs\Dividends_2025.xlsx"
    
    if not os.path.exists(excel_file):
        print(f"❌ Excel file not found: {excel_file}")
        return
    
    # Load historical data
    historical_data = load_historical_data()
    print(f"📊 Loaded historical data for {len(historical_data)} tickers")
    
    # Load workbook
    wb = openpyxl.load_workbook(excel_file)
    
    if 'Ticker Analysis 2025' not in wb.sheetnames:
        print("❌ Ticker Analysis 2025 sheet not found")
        return
    
    ws = wb['Ticker Analysis 2025']
    print(f"✅ Found Ticker Analysis 2025 sheet with {ws.max_row} rows, {ws.max_column} columns")
    
    # Find the current structure
    headers = []
    for col in range(1, ws.max_column + 1):
        cell_value = ws.cell(row=1, column=col).value
        if cell_value:
            headers.append(str(cell_value))
        else:
            headers.append(f"Col_{col}")
    
    print(f"📋 Current headers: {headers}")
    
    # Add historical yield columns after the current yield column
    yield_col_index = None
    for i, header in enumerate(headers):
        if 'yield' in header.lower() or 'Yield' in header:
            yield_col_index = i + 1  # Convert to 1-based indexing
            break
    
    if yield_col_index is None:
        print("❌ Could not find yield column")
        return
    
    # Define historical date columns to add
    historical_dates = ['6/24', '6/25', '6/26', '6/27', '7/1', '7/2', '7/3', '7/4', '7/5']
    
    # Insert new columns for historical data
    insert_col = yield_col_index + 1
    
    for i, date in enumerate(historical_dates):
        ws.insert_cols(insert_col + i)
        # Add header
        cell = ws.cell(row=1, column=insert_col + i)
        cell.value = f"Yield {date}"
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        cell.alignment = Alignment(horizontal="center")
    
    # Fill in historical data
    for row in range(2, ws.max_row + 1):
        ticker_cell = ws.cell(row=row, column=1)  # Assuming ticker is in column A
        if ticker_cell.value:
            ticker = str(ticker_cell.value).strip()
            
            # Add historical yield data
            for i, date in enumerate(historical_dates):
                col = insert_col + i
                yield_value = historical_data.get(ticker, {}).get(date, '')
                
                cell = ws.cell(row=row, column=col)
                if yield_value:
                    cell.value = f"{yield_value:.2f}%"
                    # Color code based on yield level
                    if yield_value >= 15:
                        cell.fill = PatternFill(start_color="90EE90", end_color="90EE90", fill_type="solid")  # Light green
                    elif yield_value >= 10:
                        cell.fill = PatternFill(start_color="FFFF90", end_color="FFFF90", fill_type="solid")  # Light yellow
                    elif yield_value >= 5:
                        cell.fill = PatternFill(start_color="FFE4B5", end_color="FFE4B5", fill_type="solid")  # Light orange
                    else:
                        cell.fill = PatternFill(start_color="FFB6C1", end_color="FFB6C1", fill_type="solid")  # Light pink
                else:
                    cell.value = "N/A"
                    cell.fill = PatternFill(start_color="D3D3D3", end_color="D3D3D3", fill_type="solid")  # Light gray
    
    # Add a column for yield trend analysis
    trend_col = insert_col + len(historical_dates)
    ws.insert_cols(trend_col)
    
    trend_header = ws.cell(row=1, column=trend_col)
    trend_header.value = "Yield Trend"
    trend_header.font = Font(bold=True, color="FFFFFF")
    trend_header.fill = PatternFill(start_color="FF6B6B", end_color="FF6B6B", fill_type="solid")
    trend_header.alignment = Alignment(horizontal="center")
    
    # Calculate trend for each ticker
    for row in range(2, ws.max_row + 1):
        ticker_cell = ws.cell(row=row, column=1)
        if ticker_cell.value:
            ticker = str(ticker_cell.value).strip()
            
            # Get historical yields for this ticker
            ticker_yields = []
            for date in historical_dates:
                yield_val = historical_data.get(ticker, {}).get(date)
                if yield_val:
                    ticker_yields.append(yield_val)
            
            # Calculate trend
            trend_cell = ws.cell(row=row, column=trend_col)
            if len(ticker_yields) >= 2:
                # Simple trend: compare first and last available yields
                first_yield = ticker_yields[0]
                last_yield = ticker_yields[-1]
                change = last_yield - first_yield
                
                if change > 1:
                    trend_cell.value = "⬆️ Rising"
                    trend_cell.fill = PatternFill(start_color="90EE90", end_color="90EE90", fill_type="solid")
                elif change < -1:
                    trend_cell.value = "⬇️ Falling"
                    trend_cell.fill = PatternFill(start_color="FFB6C1", end_color="FFB6C1", fill_type="solid")
                else:
                    trend_cell.value = "➡️ Stable"
                    trend_cell.fill = PatternFill(start_color="FFFF90", end_color="FFFF90", fill_type="solid")
            else:
                trend_cell.value = "📊 Limited Data"
                trend_cell.fill = PatternFill(start_color="D3D3D3", end_color="D3D3D3", fill_type="solid")
    
    # Auto-adjust column widths
    for col in range(1, ws.max_column + 1):
        max_length = 0
        column_letter = openpyxl.utils.get_column_letter(col)
        
        for row in range(1, min(ws.max_row + 1, 20)):  # Check first 20 rows for width
            try:
                if ws[f"{column_letter}{row}"].value:
                    max_length = max(max_length, len(str(ws[f"{column_letter}{row}"].value)))
            except:
                pass
        
        # Set column width (with some padding)
        adjusted_width = min(max_length + 2, 20)  # Cap at 20 characters
        ws.column_dimensions[column_letter].width = adjusted_width
    
    # Save the updated workbook
    backup_name = f"Dividends_2025_backup_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    backup_path = backup_name
    
    # Create backup
    wb.save(backup_path)
    print(f"✅ Created backup: {backup_path}")
    
    # Save main file
    wb.save(excel_file)
    print(f"✅ Updated {excel_file} with historical yield data")
    
    # Print summary
    print(f"""
📊 HISTORICAL YIELD DATA INTEGRATION COMPLETE!

✅ Added {len(historical_dates)} historical yield columns
✅ Added yield trend analysis column
✅ Applied color coding for yield levels:
   🟢 Green: Yields ≥ 15%
   🟡 Yellow: Yields 10-15%
   🟠 Orange: Yields 5-10%
   🟣 Pink: Yields < 5%
   ⚫ Gray: No data available

✅ Trend Analysis:
   ⬆️ Rising: Yield increased > 1%
   ⬇️ Falling: Yield decreased > 1%
   ➡️ Stable: Yield changed < 1%
   📊 Limited Data: Insufficient historical data

The Ticker Analysis 2025 sheet now shows complete yield progression!
""")

if __name__ == "__main__":
    print("📈 Adding Historical Yield Data to Ticker Analysis 2025...")
    add_historical_yield_columns()
