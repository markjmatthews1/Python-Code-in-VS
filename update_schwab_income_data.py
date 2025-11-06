#!/usr/bin/env python3
"""
Update Schwab Data Using Existing Tokens
Uses the existing Schwab_auth.py token system to fetch fresh account data
for the Estimated Income 2025 sheet without requiring new authentication.
"""

import sys
import os
import json
from datetime import datetime
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill

# Import your existing Schwab authentication
try:
    from Schwab_auth import get_valid_access_token, ensure_fresh_token, fetch_quote
    import schwabdev
except ImportError as e:
    print(f"❌ Error importing Schwab modules: {e}")
    print("Make sure Schwab_auth.py and schwabdev are available")
    sys.exit(1)

def check_token_status():
    """Check current token status without triggering refresh"""
    try:
        with open("tokens.json", "r") as f:
            token_data = json.load(f)
        
        tokens = token_data.get("token_dictionary", {})
        if not tokens:
            return False, "No tokens found"
        
        current_time = datetime.now()
        expires_at = tokens.get("expires_at", 0)
        expires_datetime = datetime.fromtimestamp(expires_at)
        
        if current_time.timestamp() < expires_at:
            time_left = expires_datetime - current_time
            return True, f"Token valid for {time_left}"
        else:
            return False, f"Token expired at {expires_datetime}"
            
    except Exception as e:
        return False, f"Error checking tokens: {e}"

def get_schwab_account_data():
    """Get account data using existing Schwab tokens"""
    try:
        # Check if we have valid tokens
        token_valid, status_msg = check_token_status()
        print(f"📊 Token status: {status_msg}")
        
        if not token_valid:
            print("🔄 Attempting to refresh token...")
            ensure_fresh_token()
        
        # Get fresh access token
        access_token = get_valid_access_token()
        print("✅ Successfully obtained access token")
        
        # Initialize Schwab client with existing token
        client = schwabdev.Client(app_key="n3uMFJH8tsA9z2SB2ag0sqNUNm4uPjai", 
                                app_secret="h9YybKHnDVoDM1Jw",
                                tokens_file="tokens.json")
        
        # Get account information
        print("📈 Fetching Schwab account data...")
        accounts = client.account_numbers()
        
        if not accounts:
            print("❌ No accounts found")
            return None
        
        account_data = {}
        
        for account in accounts:
            account_number = account.get('accountNumber')
            if account_number:
                print(f"   Processing account: {account_number}")
                
                # Get account details
                account_info = client.account_details(account_number, fields='positions')
                
                if account_info:
                    # Extract relevant data
                    securities_account = account_info.get('securitiesAccount', {})
                    
                    # Calculate estimated dividend income
                    estimated_income = calculate_dividend_estimate(securities_account)
                    
                    account_type = securities_account.get('type', 'Unknown')
                    account_data[f"Schwab_{account_type}"] = estimated_income
                    
                    print(f"   Account {account_number} estimated annual income: ${estimated_income:,.2f}")
        
        return account_data
        
    except Exception as e:
        print(f"❌ Error getting Schwab data: {e}")
        import traceback
        traceback.print_exc()
        return None

def calculate_dividend_estimate(securities_account):
    """Calculate estimated dividend income from account positions"""
    try:
        positions = securities_account.get('positions', [])
        total_estimated_income = 0
        
        for position in positions:
            try:
                instrument = position.get('instrument', {})
                symbol = instrument.get('symbol', '')
                quantity = position.get('longQuantity', 0) or position.get('shortQuantity', 0) or 0
                
                if quantity > 0 and symbol:
                    # For now, use a simple estimate - you could enhance this with actual dividend data
                    # This is a placeholder calculation
                    market_value = position.get('marketValue', 0)
                    
                    # Estimate 3% yield on market value as a baseline
                    # You could replace this with actual dividend yield data
                    estimated_yield = 0.03
                    estimated_income = market_value * estimated_yield
                    
                    total_estimated_income += estimated_income
                    
            except Exception as e:
                print(f"   Warning: Could not process position {position}: {e}")
                continue
        
        return total_estimated_income
        
    except Exception as e:
        print(f"Error calculating dividend estimate: {e}")
        return 0

def update_estimated_income_with_schwab():
    """Update the Estimated Income sheet with fresh Schwab data"""
    
    excel_file = 'dividend_tracker/DividendTrackerApp/outputs/Dividends_2025.xlsx'
    
    try:
        print("🚀 Starting Schwab data update for Estimated Income sheet...")
        
        # Get fresh Schwab data
        schwab_data = get_schwab_account_data()
        
        if not schwab_data:
            print("❌ Could not get Schwab data")
            return False
        
        print(f"✅ Retrieved Schwab data: {schwab_data}")
        
        # Open the Excel file
        wb = load_workbook(excel_file)
        
        if 'Estimated Income 2025' not in wb.sheetnames:
            print("❌ Estimated Income 2025 sheet not found!")
            return False
        
        ws = wb['Estimated Income 2025']
        
        # Find the current date column (last column with data)
        current_col = None
        for col in range(2, ws.max_column + 2):
            date_cell = ws.cell(row=3, column=col)
            if not date_cell.value:
                current_col = col
                break
        
        if not current_col:
            current_col = ws.max_column + 1
        
        # Add current date
        current_date = datetime.now()
        date_str = f"{current_date.month}/{current_date.day}/{current_date.year}"
        
        date_cell = ws.cell(row=3, column=current_col)
        date_cell.value = date_str
        date_cell.font = Font(name='Arial', size=12, color='FFFFFF')
        date_cell.fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
        
        # Update Schwab account rows
        account_mapping = {
            "Schwab_IRA": 6,      # Row 6 for Schwab IRA
            "Schwab_Individual": 7 # Row 7 for Schwab Individual  
        }
        
        for account_name, estimated_income in schwab_data.items():
            if account_name in account_mapping:
                row = account_mapping[account_name]
                cell = ws.cell(row=row, column=current_col)
                cell.value = estimated_income
                cell.number_format = '$#,##0.00'
                cell.font = Font(name='Arial', size=12)
                
                # Apply color coding based on previous value
                if current_col > 2:
                    prev_cell = ws.cell(row=row, column=current_col-1)
                    prev_value = prev_cell.value or 0
                    
                    if estimated_income > prev_value:
                        cell.fill = PatternFill(start_color='00B050', end_color='00B050', fill_type='solid')
                    elif estimated_income < prev_value:
                        cell.fill = PatternFill(start_color='FF0000', end_color='FF0000', fill_type='solid')
                    else:
                        cell.fill = PatternFill(start_color='FFFF00', end_color='FFFF00', fill_type='solid')
                
                print(f"   Updated {account_name}: ${estimated_income:,.2f}")
        
        # Update monthly total (row 9)
        monthly_row = 9
        yearly_total = sum(schwab_data.values()) + get_other_account_totals(ws, current_col)
        monthly_total = yearly_total / 12
        
        monthly_cell = ws.cell(row=monthly_row, column=current_col)
        monthly_cell.value = monthly_total
        monthly_cell.number_format = '$#,##0.00'
        monthly_cell.font = Font(name='Arial', size=12, bold=True)
        
        print(f"   Updated Monthly Total: ${monthly_total:,.2f}")
        
        # Save the file
        wb.save(excel_file)
        wb.close()
        
        print("✅ Estimated Income sheet updated with fresh Schwab data!")
        return True
        
    except Exception as e:
        print(f"❌ Error updating Estimated Income sheet: {e}")
        import traceback
        traceback.print_exc()
        return False

def get_other_account_totals(ws, col):
    """Get totals from other account rows (E*TRADE, etc.)"""
    total = 0
    for row in [4, 5]:  # E*TRADE rows
        cell = ws.cell(row=row, column=col)
        if isinstance(cell.value, (int, float)):
            total += cell.value
    return total

if __name__ == "__main__":
    print("🔄 Schwab Data Update for Estimated Income")
    print("="*50)
    
    # Check token status first
    token_valid, status_msg = check_token_status()
    print(f"Token Status: {status_msg}")
    
    if update_estimated_income_with_schwab():
        print("\n✅ SUCCESS!")
        print("• Fresh Schwab data retrieved using existing tokens")
        print("• Estimated Income 2025 sheet updated")
        print("• Color coding applied for changes")
        print("• No new authentication required")
    else:
        print("\n❌ Failed to update with Schwab data")
        print("Consider running this again or checking token status")
