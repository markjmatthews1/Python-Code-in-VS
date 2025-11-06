#!/usr/bin/env python3
"""
Update Schwab Data for Estimated Income - Using Existing Token System
Fetches fresh Schwab account data and updates the Estimated Income 2025 sheet
"""

import sys
import json
import time
from datetime import datetime
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill

# Import your existing Schwab authentication
from Schwab_auth import get_valid_access_token, ensure_fresh_token
import schwabdev

def get_schwab_positions():
    """Get current positions from Schwab accounts using existing tokens"""
    try:
        print("📊 Fetching Schwab account data using existing tokens...")
        
        # Ensure we have a fresh token (refresh if needed)
        ensure_fresh_token(buffer_seconds=300)  # 5 minute buffer
        
        # Create Schwab client using existing tokens
        client = schwabdev.Client(
            app_key="n3uMFJH8tsA9z2SB2ag0sqNUNm4uPjai",
            app_secret="h9YybKHnDVoDM1Jw", 
            tokens_file="tokens.json"
        )
        
        print("✅ Schwab client initialized with existing tokens")
        
        # Get all account details at once
        accounts_response = client.account_details_all(fields='positions')
        
        # Extract accounts data from response
        if hasattr(accounts_response, 'json'):
            accounts_data = accounts_response.json()
        else:
            accounts_data = accounts_response
            
        if not accounts_data:
            print("❌ No account data received")
            return None
            
        # Ensure accounts_data is a list
        if not isinstance(accounts_data, list):
            accounts_data = [accounts_data]
            
        print(f"📋 Found {len(accounts_data)} Schwab accounts")
        
        account_data = {}
        
        for account in accounts_data:
            securities_account = account.get('securitiesAccount', {})
            account_number = securities_account.get('accountNumber', '')
            account_type = securities_account.get('type', 'Unknown')
            
            print(f"   Processing account: ...{account_number[-4:]} ({account_type})")
            
            try:
                # Get current value and calculate estimated dividend income
                current_value = securities_account.get('currentBalances', {}).get('liquidationValue', 0)
                
                # Calculate estimated annual dividend income
                # Using a conservative estimate based on account value and typical dividend yields
                if current_value > 0:
                    # Estimate 3.5% dividend yield for dividend-focused accounts
                    estimated_annual_dividend = current_value * 0.035
                    
                    account_key = f"Schwab_{account_type}"
                    account_data[account_key] = {
                        'current_value': current_value,
                        'estimated_annual_dividend': estimated_annual_dividend,
                        'account_number': account_number
                    }
                    
                    print(f"   Account Value: ${current_value:,.2f}")
                    print(f"   Estimated Annual Dividends: ${estimated_annual_dividend:,.2f}")
                
            except Exception as e:
                print(f"   ⚠️ Error processing account: {e}")
                continue
        
        return account_data
        
    except Exception as e:
        print(f"❌ Error fetching Schwab data: {e}")
        import traceback
        traceback.print_exc()
        return None

def update_estimated_income_sheet():
    """Update the Estimated Income 2025 sheet with fresh Schwab data"""
    
    excel_file = 'dividend_tracker/DividendTrackerApp/outputs/Dividends_2025.xlsx'
    
    try:
        print("📈 Updating Estimated Income 2025 sheet...")
        
        # Get fresh Schwab data
        schwab_data = get_schwab_positions()
        
        if not schwab_data:
            print("❌ No Schwab data available")
            return False
        
        # Load the Excel file
        wb = load_workbook(excel_file)
        
        if 'Estimated Income 2025' not in wb.sheetnames:
            print("❌ Estimated Income 2025 sheet not found!")
            return False
        
        ws = wb['Estimated Income 2025']
        print(f"📋 Current sheet size: {ws.max_row} rows x {ws.max_column} columns")
        
        # Find the next available column or use the current weekend column
        current_date = datetime.now()
        weekend_date = f"{current_date.month}/{current_date.day}/{current_date.year}"
        
        # Look for today's column or create new one
        target_col = None
        for col in range(2, ws.max_column + 2):
            date_cell = ws.cell(row=3, column=col)
            if date_cell.value == weekend_date:
                target_col = col
                break
            elif not date_cell.value:
                target_col = col
                break
        
        if not target_col:
            target_col = ws.max_column + 1
        
        # Set up the date header
        date_cell = ws.cell(row=3, column=target_col)
        date_cell.value = weekend_date
        date_cell.font = Font(name='Arial', size=12, color='FFFFFF')
        date_cell.fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
        
        print(f"📅 Updating column {target_col} for date: {weekend_date}")
        
        # Map Schwab accounts to rows (check what account types we got)
        schwab_rows = {
            'Schwab_MARGIN': 4,        # First account is likely IRA at row 4
            'Schwab_IRA': 4,           # Standard IRA row
            'Schwab_Individual': 5     # Individual account at row 5
        }
        
        print("🔍 Account type mapping:")
        for account_type, data in schwab_data.items():
            print(f"   {account_type}: ${data['estimated_annual_dividend']:,.2f}")
        
        # Handle multiple MARGIN accounts - assign them to different rows
        margin_accounts = [k for k in schwab_data.keys() if 'MARGIN' in k]
        
        # Update Schwab account data
        total_schwab_annual = 0
        row_assignments = {}
        
        # Assign rows for MARGIN accounts
        if len(margin_accounts) >= 1:
            row_assignments[margin_accounts[0]] = 4  # First MARGIN -> row 4 (IRA equivalent)
        if len(margin_accounts) >= 2:
            row_assignments[margin_accounts[1]] = 5  # Second MARGIN -> row 5 (Individual equivalent)
        
        # Also handle specific account types
        for account_type in schwab_data.keys():
            if 'IRA' in account_type:
                row_assignments[account_type] = 4
            elif 'Individual' in account_type:
                row_assignments[account_type] = 5
        
        for account_type, data in schwab_data.items():
            row = row_assignments.get(account_type)
            
            if row:
                estimated_income = data['estimated_annual_dividend']
                total_schwab_annual += estimated_income
                
                cell = ws.cell(row=row, column=target_col)
                cell.value = estimated_income
                cell.number_format = '$#,##0.00'
                cell.font = Font(name='Arial', size=12)
                
                # Apply color coding compared to previous column
                if target_col > 2:
                    prev_cell = ws.cell(row=row, column=target_col-1)
                    prev_value = prev_cell.value or 0
                    
                    if estimated_income > prev_value:
                        cell.fill = PatternFill(start_color='00B050', end_color='00B050', fill_type='solid')
                        change_indicator = "📈"
                    elif estimated_income < prev_value:
                        cell.fill = PatternFill(start_color='FF0000', end_color='FF0000', fill_type='solid') 
                        change_indicator = "📉"
                    else:
                        cell.fill = PatternFill(start_color='FFFF00', end_color='FFFF00', fill_type='solid')
                        change_indicator = "➡️"
                    
                    change_amount = estimated_income - prev_value
                    print(f"   {account_type}: ${estimated_income:,.2f} {change_indicator} (${change_amount:+,.2f})")
                else:
                    print(f"   {account_type}: ${estimated_income:,.2f}")
        
        # Update monthly total (row 9) - sum all accounts and divide by 12
        monthly_row = 9
        
        # Sum all account rows for this column (including our Schwab data)
        total_annual = total_schwab_annual  # Start with our Schwab total
        
        # Add other account rows (E*TRADE, etc.) if they have values
        for row in range(4, 8):  # Rows 4-7 for all accounts
            cell_value = ws.cell(row=row, column=target_col).value
            if isinstance(cell_value, (int, float)) and row not in row_assignments.values():
                # Only add non-Schwab accounts to avoid double counting
                total_annual += cell_value
        
        monthly_total = total_annual / 12
        monthly_cell = ws.cell(row=monthly_row, column=target_col)
        monthly_cell.value = monthly_total
        monthly_cell.number_format = '$#,##0.00'
        monthly_cell.font = Font(name='Arial', size=12, bold=True)
        
        # Apply color coding to monthly total
        if target_col > 2:
            prev_monthly = ws.cell(row=monthly_row, column=target_col-1).value or 0
            if monthly_total > prev_monthly:
                monthly_cell.fill = PatternFill(start_color='00B050', end_color='00B050', fill_type='solid')
            elif monthly_total < prev_monthly:
                monthly_cell.fill = PatternFill(start_color='FF0000', end_color='FF0000', fill_type='solid')
            else:
                monthly_cell.fill = PatternFill(start_color='FFFF00', end_color='FFFF00', fill_type='solid')
        
        print(f"   Monthly Total: ${monthly_total:,.2f} (Annual: ${total_annual:,.2f})")
        
        # Save the workbook
        wb.save(excel_file)
        wb.close()
        
        print("✅ Estimated Income 2025 sheet updated successfully!")
        return True
        
    except Exception as e:
        print(f"❌ Error updating sheet: {e}")
        import traceback
        traceback.print_exc()
        return False

if __name__ == "__main__":
    print("🚀 Schwab Data Update for Estimated Income")
    print("Using existing token system - no new authentication required")
    print("="*60)
    
    success = update_estimated_income_sheet()
    
    if success:
        print("\n✅ SUCCESS!")
        print("• Fresh Schwab data retrieved using existing tokens")
        print("• Estimated Income 2025 sheet updated with latest values")
        print("• Color coding applied for week-to-week changes")
        print("• Monthly totals recalculated")
        print("• No token refresh popup needed!")
    else:
        print("\n❌ Update failed - check errors above")
    
    print("\n" + "="*60)
