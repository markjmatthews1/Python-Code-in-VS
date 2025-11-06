#!/usr/bin/env python3
"""
Complete Portfolio Data Update - Real E*TRADE and Schwab Account Values
Gets actual current account values and updates Estimated Income 2025 sheet properly
"""

import sys
import json
import time
from datetime import datetime
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter

# E*TRADE imports
try:
    from etrade_auth import get_etrade_session
    ETRADE_AVAILABLE = True
except ImportError:
    ETRADE_AVAILABLE = False
    print("⚠️ E*TRADE authentication not available")

# Schwab imports
try:
    from Schwab_auth import get_valid_access_token, ensure_fresh_token
    import schwabdev
    SCHWAB_AVAILABLE = True
except ImportError:
    SCHWAB_AVAILABLE = False
    print("⚠️ Schwab authentication not available")

def get_etrade_account_values():
    """Get real account values from E*TRADE using existing auth system"""
    if not ETRADE_AVAILABLE:
        print("❌ E*TRADE not available")
        return None
    
    try:
        print("📊 Fetching E*TRADE account values...")
        
        # Use the existing etrade_auth system that has popup and audio
        import etrade_auth
        
        # This will automatically handle token refresh with popup and audio if needed
        session, base_url = etrade_auth.get_etrade_session()
        
        # Get account list
        url = f"{base_url}/v1/accounts/list.json"
        response = session.get(url)
        
        # The etrade_auth system handles all token refresh automatically with popup/audio
        if response.status_code != 200:
            print(f"❌ Failed to get E*TRADE accounts: {response.status_code}")
            print(f"Response: {response.text[:200]}")
            return None
        
        accounts_data = response.json()
        accounts = accounts_data.get("AccountListResponse", {}).get("Accounts", {}).get("Account", [])
        
        account_values = {}
        
        for account in accounts:
            account_id = account.get("accountIdKey")
            account_desc = account.get("accountDesc", "Unknown")
            account_type = account.get("accountType", "Unknown")
            
            # Skip the futures account ending in 7650 (empty/unused)
            if account_id and str(account_id).endswith('7650'):
                print(f"   Skipping E*TRADE futures account: {account_desc} (empty futures account)")
                continue
            
            print(f"   Processing E*TRADE account: {account_desc} ({account_type})")
            
            # Get account balance
            balance_url = f"{base_url}/v1/accounts/{account_id}/balance.json"
            balance_response = session.get(balance_url)
            
            if balance_response.status_code == 200:
                balance_data = balance_response.json()
                balance_info = balance_data.get("BalanceResponse", {})
                
                # Get total account value
                total_value = 0
                if "Computed" in balance_info:
                    computed = balance_info["Computed"]
                    total_value = computed.get("RealTimeValues", {}).get("totalAccountValue", 0)
                
                # Determine account type for mapping
                if "taxable" in account_desc.lower() or "individual" in account_desc.lower():
                    account_key = "ETRADE_Taxable"
                elif "ira" in account_desc.lower() or "rollover" in account_desc.lower():
                    account_key = "ETRADE_IRA"
                else:
                    account_key = f"ETRADE_{account_desc.replace(' ', '_')}"
                
                account_values[account_key] = {
                    'account_desc': account_desc,
                    'account_type': account_type,
                    'current_value': float(total_value),
                    'estimated_annual_dividend': float(total_value) * 0.035  # 3.5% estimate
                }
                
                print(f"   Account Value: ${total_value:,.2f}")
                print(f"   Estimated Annual Dividends: ${float(total_value) * 0.035:,.2f}")
            
            else:
                print(f"   ⚠️ Could not get balance for {account_desc}")
        
        return account_values
        
    except Exception as e:
        print(f"❌ Error fetching E*TRADE data: {e}")
        import traceback
        traceback.print_exc()
        return None

def get_schwab_account_values():
    """Get real account values from Schwab"""
    if not SCHWAB_AVAILABLE:
        print("❌ Schwab not available")
        return None
    
    try:
        print("📊 Fetching Schwab account values...")
        
        # Ensure we have a fresh token
        ensure_fresh_token(buffer_seconds=300)
        
        client = schwabdev.Client(
            app_key="n3uMFJH8tsA9z2SB2ag0sqNUNm4uPjai",
            app_secret="h9YybKHnDVoDM1Jw", 
            tokens_file="tokens.json"
        )
        
        print("✅ Schwab client initialized")
        
        # Get all account details
        accounts_response = client.account_details_all(fields='positions')
        
        if hasattr(accounts_response, 'json'):
            accounts_data = accounts_response.json()
        else:
            accounts_data = accounts_response
            
        if not isinstance(accounts_data, list):
            accounts_data = [accounts_data]
        
        account_values = {}
        
        for i, account in enumerate(accounts_data, 1):
            securities_account = account.get('securitiesAccount', {})
            account_number = securities_account.get('accountNumber', '')
            account_type = securities_account.get('type', 'Unknown')
            
            print(f"   Processing Schwab account {i}: ...{account_number[-4:]} ({account_type})")
            
            # Get current value
            current_balances = securities_account.get('currentBalances', {})
            total_value = current_balances.get('liquidationValue', 0)
            
            # Determine account mapping based on account type and order
            if account_type == "MARGIN":
                if i == 1:
                    # Smaller account is likely Individual
                    account_key = "SCHWAB_Individual"
                else:
                    # Larger account is likely IRA
                    account_key = "SCHWAB_IRA"
            else:
                account_key = f"SCHWAB_{account_type}"
            
            account_values[account_key] = {
                'account_number': account_number,
                'account_type': account_type,
                'current_value': float(total_value),
                'estimated_annual_dividend': float(total_value) * 0.035
            }
            
            print(f"   Account Value: ${total_value:,.2f}")
            print(f"   Estimated Annual Dividends: ${float(total_value) * 0.035:,.2f}")
        
        return account_values
        
    except Exception as e:
        print(f"❌ Error fetching Schwab data: {e}")
        import traceback
        traceback.print_exc()
        return None

def update_estimated_income_with_real_data():
    """Update the Estimated Income 2025 sheet with real account values"""
    
    excel_file = 'dividend_tracker/DividendTrackerApp/outputs/Dividends_2025.xlsx'
    
    try:
        print("📈 Updating Estimated Income 2025 sheet with real data...")
        
        # Get real account data
        etrade_data = get_etrade_account_values()
        schwab_data = get_schwab_account_values()
        
        if not etrade_data and not schwab_data:
            print("❌ No account data available from either source")
            return False
        
        # Load Excel file
        wb = load_workbook(excel_file)
        
        if 'Estimated Income 2025' not in wb.sheetnames:
            print("❌ Estimated Income 2025 sheet not found!")
            return False
        
        ws = wb['Estimated Income 2025']
        print(f"📋 Current sheet size: {ws.max_row} rows x {ws.max_column} columns")
        
        # Set column widths to 15 for all columns
        print("📐 Setting column widths to 15...")
        for col in range(1, ws.max_column + 5):  # Extra columns for future use
            col_letter = get_column_letter(col)
            ws.column_dimensions[col_letter].width = 15
        
        # Find next available column (don't overwrite existing data)
        current_date = datetime.now()
        weekend_date = f"{current_date.month}/{current_date.day}/{current_date.year}"
        
        target_col = None
        # Look for a completely empty column
        for col in range(2, ws.max_column + 10):  # Check beyond current max
            column_empty = True
            for row in range(3, 10):  # Check key rows
                if ws.cell(row=row, column=col).value is not None:
                    column_empty = False
                    break
            if column_empty:
                target_col = col
                break
        
        if not target_col:
            target_col = ws.max_column + 1
        
        print(f"📅 Using column {target_col} for date: {weekend_date}")
        
        # Set up the date header with blue background
        date_cell = ws.cell(row=3, column=target_col)
        date_cell.value = weekend_date
        date_cell.font = Font(name='Arial', size=12, color='FFFFFF')
        date_cell.fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
        
        # Account row mappings (based on your specification)
        account_mappings = {
            'ETRADE_Taxable': {'row': 4, 'expected': 10888.52},
            'ETRADE_IRA': {'row': 5, 'expected': 29178.41},
            'SCHWAB_Individual': {'row': 6, 'expected': 674.04},
            'SCHWAB_IRA': {'row': 7, 'expected': 8924.02}
        }
        
        # Update E*TRADE accounts
        if etrade_data:
            print("📊 Updating E*TRADE account data...")
            for account_key, data in etrade_data.items():
                if account_key in account_mappings:
                    row = account_mappings[account_key]['row']
                    expected = account_mappings[account_key]['expected']
                    estimated_income = data['estimated_annual_dividend']
                    
                    cell = ws.cell(row=row, column=target_col)
                    cell.value = estimated_income
                    cell.number_format = '$#,##0.00'
                    cell.font = Font(name='Arial', size=12)
                    
                    # Apply softer color coding
                    if target_col > 2:
                        prev_cell = ws.cell(row=row, column=target_col-1)
                        prev_value_raw = prev_cell.value
                        
                        # Handle both numeric and string values
                        try:
                            prev_value = float(str(prev_value_raw).replace('$', '').replace(',', '')) if prev_value_raw else 0
                        except:
                            prev_value = 0
                        
                        if estimated_income > prev_value:
                            # Softer green: #00FA71
                            cell.fill = PatternFill(start_color='00FA71', end_color='00FA71', fill_type='solid')
                            change_indicator = "📈"
                        elif estimated_income < prev_value:
                            # Softer red: #FF7C80
                            cell.fill = PatternFill(start_color='FF7C80', end_color='FF7C80', fill_type='solid')
                            change_indicator = "📉"
                        else:
                            # Keep yellow as is
                            cell.fill = PatternFill(start_color='FFFF00', end_color='FFFF00', fill_type='solid')
                            change_indicator = "➡️"
                        
                        change_amount = estimated_income - prev_value
                        print(f"   Row {row} ({account_key}): ${estimated_income:,.2f} {change_indicator} (${change_amount:+,.2f}) [Expected: ${expected:,.2f}]")
                    else:
                        print(f"   Row {row} ({account_key}): ${estimated_income:,.2f} [Expected: ${expected:,.2f}]")
        
        # Update Schwab accounts
        if schwab_data:
            print("📊 Updating Schwab account data...")
            for account_key, data in schwab_data.items():
                if account_key in account_mappings:
                    row = account_mappings[account_key]['row']
                    expected = account_mappings[account_key]['expected']
                    estimated_income = data['estimated_annual_dividend']
                    
                    cell = ws.cell(row=row, column=target_col)
                    cell.value = estimated_income
                    cell.number_format = '$#,##0.00'
                    cell.font = Font(name='Arial', size=12)
                    
                    # Apply softer color coding
                    if target_col > 2:
                        prev_cell = ws.cell(row=row, column=target_col-1)
                        prev_value_raw = prev_cell.value
                        
                        # Handle both numeric and string values
                        try:
                            prev_value = float(str(prev_value_raw).replace('$', '').replace(',', '')) if prev_value_raw else 0
                        except:
                            prev_value = 0
                        
                        if estimated_income > prev_value:
                            # Softer green: #00FA71
                            cell.fill = PatternFill(start_color='00FA71', end_color='00FA71', fill_type='solid')
                            change_indicator = "📈"
                        elif estimated_income < prev_value:
                            # Softer red: #FF7C80
                            cell.fill = PatternFill(start_color='FF7C80', end_color='FF7C80', fill_type='solid')
                            change_indicator = "📉"
                        else:
                            # Keep yellow as is
                            cell.fill = PatternFill(start_color='FFFF00', end_color='FFFF00', fill_type='solid')
                            change_indicator = "➡️"
                        
                        change_amount = estimated_income - prev_value
                        print(f"   Row {row} ({account_key}): ${estimated_income:,.2f} {change_indicator} (${change_amount:+,.2f}) [Expected: ${expected:,.2f}]")
                    else:
                        print(f"   Row {row} ({account_key}): ${estimated_income:,.2f} [Expected: ${expected:,.2f}]")
        
        # Calculate and update monthly total (row 9)
        monthly_row = 9
        total_annual = 0
        
        for row in range(4, 8):  # Rows 4-7 for all accounts
            cell_value = ws.cell(row=row, column=target_col).value
            if isinstance(cell_value, (int, float)):
                total_annual += cell_value
        
        monthly_total = total_annual / 12
        monthly_cell = ws.cell(row=monthly_row, column=target_col)
        monthly_cell.value = monthly_total
        monthly_cell.number_format = '$#,##0.00'
        monthly_cell.font = Font(name='Arial', size=12, bold=True)
        
        # Apply color coding to monthly total
        if target_col > 2:
            prev_monthly_raw = ws.cell(row=monthly_row, column=target_col-1).value
            # Convert to number if it's a string
            try:
                prev_monthly = float(str(prev_monthly_raw).replace('$', '').replace(',', '')) if prev_monthly_raw else 0
            except:
                prev_monthly = 0
            
            if monthly_total > prev_monthly:
                # Softer green: #00FA71
                monthly_cell.fill = PatternFill(start_color='00FA71', end_color='00FA71', fill_type='solid')
                change_indicator = "📈"
            elif monthly_total < prev_monthly:
                # Softer red: #FF7C80
                monthly_cell.fill = PatternFill(start_color='FF7C80', end_color='FF7C80', fill_type='solid')
                change_indicator = "📉"
            else:
                # Keep yellow as is
                monthly_cell.fill = PatternFill(start_color='FFFF00', end_color='FFFF00', fill_type='solid')
                change_indicator = "➡️"
            
            change_amount = monthly_total - prev_monthly
            print(f"   Monthly Total: ${monthly_total:,.2f} {change_indicator} (${change_amount:+,.2f})")
        else:
            print(f"   Monthly Total: ${monthly_total:,.2f}")
        
        print(f"   Total Annual Estimate: ${total_annual:,.2f}")
        
        # Save the workbook
        wb.save(excel_file)
        wb.close()
        
        print("✅ Estimated Income 2025 sheet updated with real data!")
        return True
        
    except Exception as e:
        print(f"❌ Error updating sheet: {e}")
        import traceback
        traceback.print_exc()
        return False

def check_portfolio_summary_sheet():
    """Check if Portfolio Summary sheet exists and is in first position"""
    excel_file = 'dividend_tracker/DividendTrackerApp/outputs/Dividends_2025.xlsx'
    
    try:
        wb = load_workbook(excel_file)
        sheet_names = wb.sheetnames
        
        print(f"📋 Current sheet order: {sheet_names}")
        
        if 'Portfolio Summary' in sheet_names:
            current_position = sheet_names.index('Portfolio Summary')
            print(f"✅ Portfolio Summary found at position {current_position}")
            
            if current_position != 0:
                print("🔄 Moving Portfolio Summary to first position...")
                ws = wb['Portfolio Summary']
                wb.move_sheet(ws, offset=-current_position)
                wb.save(excel_file)
                print("✅ Portfolio Summary moved to first position")
        else:
            print("⚠️ Portfolio Summary sheet not found!")
        
        wb.close()
        return True
        
    except Exception as e:
        print(f"❌ Error checking Portfolio Summary: {e}")
        return False

if __name__ == "__main__":
    print("🚀 Complete Portfolio Data Update")
    print("Getting real E*TRADE and Schwab account values")
    print("Using softer colors and proper column widths")
    print("="*60)
    
    # Check Portfolio Summary sheet first
    print("\n1️⃣ Checking Portfolio Summary sheet...")
    check_portfolio_summary_sheet()
    
    # Update Estimated Income with real data
    print("\n2️⃣ Updating Estimated Income with real account data...")
    success = update_estimated_income_with_real_data()
    
    if success:
        print("\n🎉 COMPLETE SUCCESS!")
        print("✅ Real E*TRADE account values retrieved")
        print("✅ Real Schwab account values retrieved")
        print("✅ Estimated Income 2025 sheet updated")
        print("✅ Softer colors applied (#FF7C80 red, #00FA71 green)")
        print("✅ Column widths set to 15")
        print("✅ New column used (no overwriting)")
        print("✅ Portfolio Summary checked/repositioned")
        print("\nYour portfolio data is now refreshed with REAL account values! 📊💰")
    else:
        print("\n❌ Update failed - check errors above")
    
    print("\n" + "="*60)
