#!/usr/bin/env python3
"""
Enhanced Portfolio Summary Sheet
- Year-to-date gain/loss calculation
- Bold green formatting for positive values
- Week-over-week performance tracking
- Professional styling with Arial 12 font
"""

import pandas as pd
import numpy as np
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
from datetime import datetime, date

def create_enhanced_portfolio_summary():
    """Create enhanced Portfolio Summary with YTD tracking and improved formatting"""
    
    excel_file = 'dividend_tracker/DividendTrackerApp/outputs/Dividends_2025.xlsx'
    
    try:
        # Load the workbook
        wb = load_workbook(excel_file)
        
        # Get current data from Portfolio Values 2025 sheet
        portfolio_ws = wb['Portfolio Values 2025']
        print("📊 Reading portfolio data from Portfolio Values 2025 sheet...")
        
        # Find the last column with actual data (not just empty cells)
        last_col = None
        for col in range(portfolio_ws.max_column, 0, -1):  # Search backwards
            # Check if this column has meaningful data in the total row (row 10)
            total_value = portfolio_ws.cell(row=10, column=col).value
            if total_value and isinstance(total_value, (int, float)) and total_value > 1000:  # Reasonable portfolio value
                last_col = col
                break
        
        if not last_col:
            # Fallback: find any column with data
            for col in range(portfolio_ws.max_column, 0, -1):
                for row in range(3, 11):
                    cell_val = portfolio_ws.cell(row=row, column=col).value
                    if cell_val and isinstance(cell_val, (int, float)) and cell_val > 0:
                        last_col = col
                        break
                if last_col:
                    break
        
        if not last_col:
            last_col = portfolio_ws.max_column
            
        print(f"   Last data column: {last_col}")
        
        # Also check the previous column for comparison
        prev_col = last_col - 1 if last_col > 1 else None
        
        # Get current values (last column)
        current_date_cell = portfolio_ws.cell(row=3, column=last_col).value
        current_date = current_date_cell if current_date_cell else "Latest"
        
        # Try to get values from different possible row locations
        etrade_ira_current = None
        etrade_taxable_current = None
        schwab_ira_current = None
        schwab_individual_current = None
        retirement_401k_current = None
        current_total = None
        
        # Check rows 4-10 for account values
        for row_num in range(4, 11):
            cell_val = portfolio_ws.cell(row=row_num, column=last_col).value
            if cell_val and isinstance(cell_val, (int, float)):
                if row_num == 4 and not etrade_ira_current:
                    etrade_ira_current = cell_val
                elif row_num == 5 and not etrade_taxable_current:
                    etrade_taxable_current = cell_val
                elif row_num == 6 and not schwab_ira_current:
                    schwab_ira_current = cell_val
                elif row_num == 7 and not schwab_individual_current:
                    schwab_individual_current = cell_val
                elif row_num == 8 and not retirement_401k_current:
                    retirement_401k_current = cell_val
                elif row_num == 10 and not current_total:
                    current_total = cell_val
        
        # Set defaults if not found
        etrade_ira_current = etrade_ira_current or 0
        etrade_taxable_current = etrade_taxable_current or 0
        schwab_ira_current = schwab_ira_current or 0
        schwab_individual_current = schwab_individual_current or 0
        retirement_401k_current = retirement_401k_current or 0
        current_total = current_total or (etrade_ira_current + etrade_taxable_current + 
                                       schwab_ira_current + schwab_individual_current + 
                                       retirement_401k_current)
        
        # Get previous week's total (if available)
        prev_total = None
        if prev_col:
            prev_total = portfolio_ws.cell(row=10, column=prev_col).value
        
        # Find year-end 2024 values (look for first column - should be year-end baseline)
        year_end_total = None
        year_end_col = None
        
        # Look through columns to find the earliest data (year-end baseline)
        for col in range(1, last_col + 1):
            cell_value = portfolio_ws.cell(row=10, column=col).value
            if cell_value and isinstance(cell_value, (int, float)) and cell_value > 0:
                year_end_total = cell_value
                year_end_col = col
                year_end_date = portfolio_ws.cell(row=3, column=col).value
                break
        
        print(f"   Current date: {current_date}")
        print(f"   Current total: ${current_total:,.2f}")
        if year_end_total:
            print(f"   Year-end baseline ({year_end_date}): ${year_end_total:,.2f}")
        
        # Calculate YTD performance
        ytd_gain_loss = None
        ytd_gain_loss_pct = None
        if year_end_total and isinstance(year_end_total, (int, float)) and year_end_total != 0:
            ytd_gain_loss = current_total - year_end_total
            ytd_gain_loss_pct = ytd_gain_loss / year_end_total
        
        # Clear and recreate Portfolio Summary sheet
        if 'Portfolio Summary' in wb.sheetnames:
            # Delete the existing sheet and create a new one
            del wb['Portfolio Summary']
        
        ws = wb.create_sheet('Portfolio Summary')
        
        print("📋 Creating enhanced Portfolio Summary sheet...")
        
        # Set default font (Arial 12)
        default_font = Font(name='Arial', size=12)
        bold_font = Font(name='Arial', size=12, bold=True)
        
        # Title
        ws['A1'] = 'PORTFOLIO SUMMARY - 2025'
        ws['A1'].font = Font(name='Arial', size=16, bold=True, color='FFFFFF')
        ws['A1'].fill = PatternFill(start_color='366092', end_color='366092', fill_type='solid')
        ws.merge_cells('A1:B1')
        
        # Current Portfolio Value Section
        ws['A3'] = f'CURRENT PORTFOLIO VALUE ({current_date})'
        ws['A3'].font = bold_font
        ws['A3'].fill = PatternFill(start_color='D9E1F2', end_color='D9E1F2', fill_type='solid')
        ws.merge_cells('A3:B3')
        
        # Account breakdown with current values
        accounts = [
            ('E*TRADE IRA:', etrade_ira_current),
            ('E*TRADE Taxable:', etrade_taxable_current),
            ('Schwab IRA:', schwab_ira_current),
            ('Schwab Individual:', schwab_individual_current),
            ('401k Retirement:', retirement_401k_current)
        ]
        
        row = 4
        for account, value in accounts:
            ws[f'A{row}'] = account
            ws[f'A{row}'].font = default_font
            ws[f'B{row}'] = value
            ws[f'B{row}'].number_format = '$#,##0.00'
            ws[f'B{row}'].font = default_font
            row += 1
        
        # Total
        ws[f'A{row}'] = 'TOTAL PORTFOLIO:'
        ws[f'A{row}'].font = bold_font
        ws[f'B{row}'] = current_total
        ws[f'B{row}'].number_format = '$#,##0.00'
        ws[f'B{row}'].font = bold_font
        ws[f'B{row}'].fill = PatternFill(start_color='92D050', end_color='92D050', fill_type='solid')
        
        # Portfolio Allocation
        ws['A11'] = 'PORTFOLIO ALLOCATION'
        ws['A11'].font = bold_font
        ws['A11'].fill = PatternFill(start_color='D9E1F2', end_color='D9E1F2', fill_type='solid')
        ws.merge_cells('A11:B11')
        
        row = 12
        for account, value in accounts:
            percentage_decimal = value / current_total if current_total != 0 else 0
            ws[f'A{row}'] = account
            ws[f'A{row}'].font = default_font
            ws[f'B{row}'] = percentage_decimal
            ws[f'B{row}'].number_format = '0.0%'
            ws[f'B{row}'].font = default_font
            row += 1
        
        # Recent Performance Section
        row = 18
        ws[f'A{row}'] = 'RECENT PERFORMANCE'
        ws[f'A{row}'].font = bold_font
        ws[f'A{row}'].fill = PatternFill(start_color='D9E1F2', end_color='D9E1F2', fill_type='solid')
        ws.merge_cells(f'A{row}:B{row}')
        row += 1
        
        # Weekly performance (with bold green for positive values)
        if prev_total and isinstance(prev_total, (int, float)) and prev_total != 0:
            weekly_change = current_total - prev_total
            weekly_change_pct = weekly_change / prev_total
            
            # Weekly dollar change
            ws[f'A{row}'] = 'Weekly Change:'
            ws[f'A{row}'].font = default_font
            ws[f'B{row}'] = weekly_change
            ws[f'B{row}'].number_format = '$#,##0.00'
            if weekly_change >= 0:
                ws[f'B{row}'].font = Font(name='Arial', size=12, color='00B050', bold=True)  # Bold green
            else:
                ws[f'B{row}'].font = Font(name='Arial', size=12, color='FF0000')
            row += 1
            
            # Weekly percentage change
            ws[f'A{row}'] = 'Weekly % Change:'
            ws[f'A{row}'].font = default_font
            ws[f'B{row}'] = weekly_change_pct
            ws[f'B{row}'].number_format = '0.00%'
            if weekly_change_pct >= 0:
                ws[f'B{row}'].font = Font(name='Arial', size=12, color='00B050', bold=True)  # Bold green
            else:
                ws[f'B{row}'].font = Font(name='Arial', size=12, color='FF0000')
            row += 1
            
            print(f"   Weekly change: ${weekly_change:,.2f} ({weekly_change_pct:.2%})")
        
        # Year-to-Date Performance Section
        row += 1
        ws[f'A{row}'] = 'YEAR-TO-DATE PERFORMANCE'
        ws[f'A{row}'].font = bold_font
        ws[f'A{row}'].fill = PatternFill(start_color='D9E1F2', end_color='D9E1F2', fill_type='solid')
        ws.merge_cells(f'A{row}:B{row}')
        row += 1
        
        if ytd_gain_loss is not None:
            # YTD dollar gain/loss
            ws[f'A{row}'] = 'YTD Gain/Loss:'
            ws[f'A{row}'].font = default_font
            ws[f'B{row}'] = ytd_gain_loss
            ws[f'B{row}'].number_format = '$#,##0.00'
            if ytd_gain_loss >= 0:
                ws[f'B{row}'].font = Font(name='Arial', size=12, color='00B050', bold=True)  # Bold green
            else:
                ws[f'B{row}'].font = Font(name='Arial', size=12, color='FF0000')
            row += 1
            
            # YTD percentage gain/loss
            ws[f'A{row}'] = 'YTD % Gain/Loss:'
            ws[f'A{row}'].font = default_font
            ws[f'B{row}'] = ytd_gain_loss_pct
            ws[f'B{row}'].number_format = '0.00%'
            if ytd_gain_loss_pct >= 0:
                ws[f'B{row}'].font = Font(name='Arial', size=12, color='00B050', bold=True)  # Bold green
            else:
                ws[f'B{row}'].font = Font(name='Arial', size=12, color='FF0000')
            row += 1
            
            print(f"   YTD performance: ${ytd_gain_loss:,.2f} ({ytd_gain_loss_pct:.2%})")
        else:
            ws[f'A{row}'] = 'YTD data not available'
            ws[f'A{row}'].font = Font(name='Arial', size=12, italic=True)
            row += 1
        
        # Baseline Information
        if year_end_total:
            row += 1
            ws[f'A{row}'] = f'Baseline Value ({year_end_date}):'
            ws[f'A{row}'].font = default_font
            ws[f'B{row}'] = year_end_total
            ws[f'B{row}'].number_format = '$#,##0.00'
            ws[f'B{row}'].font = default_font
            row += 1
        
        # Dividend Summary Section
        row += 1
        ws[f'A{row}'] = 'DIVIDEND SUMMARY'
        ws[f'A{row}'].font = bold_font
        ws[f'A{row}'].fill = PatternFill(start_color='D9E1F2', end_color='D9E1F2', fill_type='solid')
        ws.merge_cells(f'A{row}:B{row}')
        row += 1
        
        ws[f'A{row}'] = 'Annual Dividend Goal:'
        ws[f'A{row}'].font = default_font
        ws[f'B{row}'] = 'Check Estimated Income 2025 sheet'
        ws[f'B{row}'].font = default_font
        row += 1
        
        ws[f'A{row}'] = 'YTD Dividends Received:'
        ws[f'A{row}'].font = default_font
        ws[f'B{row}'] = 'Check All account weekly dividends sheet'
        ws[f'B{row}'].font = default_font
        
        # Column widths
        ws.column_dimensions['A'].width = 25
        ws.column_dimensions['B'].width = 20
        
        # Save the workbook
        wb.save(excel_file)
        wb.close()
        
        print("✅ Enhanced Portfolio Summary sheet created with:")
        print("   • Year-to-date gain/loss tracking")
        print("   • Bold green formatting for positive values")
        print("   • Week-over-week performance comparison")
        print("   • Professional Arial 12 font styling")
        print(f"   • Current portfolio data from {current_date}")
        print(f"   • Total Portfolio Value: ${current_total:,.2f}")
        if ytd_gain_loss is not None:
            print(f"   • YTD Performance: ${ytd_gain_loss:,.2f} ({ytd_gain_loss_pct:.2%})")
        
        return True
        
    except Exception as e:
        print(f"❌ Error creating enhanced Portfolio Summary: {e}")
        import traceback
        traceback.print_exc()
        return False

if __name__ == "__main__":
    print("🚀 Creating Enhanced Portfolio Summary Sheet")
    print("="*60)
    
    if create_enhanced_portfolio_summary():
        print("\n✅ SUCCESS!")
        print("Enhanced Portfolio Summary now includes:")
        print("• Year-to-date gain/loss calculation")
        print("• Bold green formatting for positive values")
        print("• Week-over-week performance tracking")
        print("• Professional styling throughout")
        print("• Comprehensive performance metrics")
    else:
        print("❌ Failed to create enhanced Portfolio Summary sheet")
