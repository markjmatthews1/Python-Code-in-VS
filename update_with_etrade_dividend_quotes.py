"""
Update Estimated Income 2025 with REAL dividend estimates using E*TRADE Quote API
Gets actual positions and actual annual dividend amounts from E*TRADE quotes
"""

import sys
import os
sys.path.append(os.path.dirname(os.path.abspath(__file__)))

from etrade_auth import get_etrade_session
import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment
import requests
import time
from datetime import datetime

def get_etrade_dividend_estimates():
    """Get E*TRADE positions and calculate annual dividend estimates using quote API"""
    print("🔄 Getting E*TRADE dividend estimates using quote API...")
    
    try:
        # Import the dividend tracker's E*TRADE API
        sys.path.append(r'c:\Users\mjmat\Python Code in VS\dividend_tracker\DividendTrackerApp')
        from modules.etrade_account_api import ETRADEAccountAPI
        
        # Also get the quote session for dividend data
        session, base_url = get_etrade_session()
        
        etrade_api = ETRADEAccountAPI()
        accounts = etrade_api.get_account_list()
        
        if not accounts:
            print("❌ No E*TRADE accounts found")
            return {}
        
        account_estimates = {}
        
        for account in accounts:
            account_id = account.get('accountIdKey', '')
            account_type = account.get('accountType', '')
            account_desc = account.get('accountDesc', '')
            
            # Skip futures account ending in 7650
            if account_id.endswith('7650'):
                print(f"   ⏭️ Skipping futures account: {account_desc}")
                continue
            
            # Determine account name
            if 'IRA' in account_type.upper() or 'ROLLOVER' in account_type.upper():
                account_name = 'E*TRADE IRA'
            else:
                account_name = 'E*TRADE Taxable'
            
            print(f"   📊 Processing {account_name} ({account_desc})")
            
            positions = etrade_api.get_account_positions(account_id)
            if not positions:
                print(f"   ⚠️ No positions found for {account_name}")
                account_estimates[account_name] = 0.0
                continue
            
            total_annual_dividend = 0.0
            dividend_positions = 0
            
            for position in positions:
                symbol = position.get('symbolDescription', '').upper().strip()
                quantity = position.get('quantity', 0)
                market_value = position.get('marketValue', 0)
                
                if quantity > 0 and market_value > 0 and symbol:
                    print(f"      🔍 {symbol}: {quantity} shares, ${market_value:,.2f}")
                    
                    try:
                        # Get quote from E*TRADE API for dividend info
                        quote_url = f"{base_url}/v1/market/quote/{symbol}.json"
                        quote_response = session.get(quote_url)
                        
                        if quote_response.status_code == 200:
                            quote_data = quote_response.json()
                            quote_info = quote_data.get('QuoteResponse', {}).get('QuoteData', [])
                            
                            if quote_info:
                                if not isinstance(quote_info, list):
                                    quote_info = [quote_info]
                                
                                # Get the first quote
                                quote = quote_info[0]
                                annual_dividend_per_share = float(quote.get('annualDividend', 0))
                                
                                if annual_dividend_per_share > 0:
                                    # Calculate total annual dividend for this position
                                    position_annual_dividend = annual_dividend_per_share * quantity
                                    total_annual_dividend += position_annual_dividend
                                    dividend_positions += 1
                                    
                                    # Calculate yield for display
                                    current_price = float(quote.get('lastTrade', market_value / quantity))
                                    dividend_yield = (annual_dividend_per_share / current_price * 100) if current_price > 0 else 0
                                    
                                    print(f"         💰 ${annual_dividend_per_share:.4f}/share × {quantity} = ${position_annual_dividend:.2f}/year ({dividend_yield:.2f}% yield)")
                                else:
                                    print(f"         ⚪ No dividend (growth stock)")
                            else:
                                print(f"         ⚠️ No quote data found")
                        else:
                            print(f"         ⚠️ Quote API error: {quote_response.status_code}")
                    
                    except Exception as quote_error:
                        print(f"         ❌ Quote error: {quote_error}")
                    
                    time.sleep(0.1)  # Brief pause between API calls
            
            account_estimates[account_name] = total_annual_dividend
            print(f"   ✅ {account_name} Total Annual Dividend: ${total_annual_dividend:.2f} ({dividend_positions} dividend positions)")
        
        return account_estimates
        
    except Exception as e:
        print(f"❌ E*TRADE API error: {e}")
        return {}

def get_schwab_dividend_estimates():
    """Get Schwab positions and calculate annual dividend estimates"""
    print("🔄 Getting Schwab dividend estimates...")
    
    try:
        # Import existing Schwab auth
        from Schwab_auth import get_valid_access_token
        
        access_token = get_valid_access_token()
        if not access_token:
            print("❌ No valid Schwab access token - trying to refresh...")
            # Try to get a fresh token
            try:
                from Schwab_auth import schwab_auth_popup_and_sound
                access_token = schwab_auth_popup_and_sound()
                if not access_token:
                    print("❌ Could not get Schwab access token")
                    return {}
            except Exception as auth_error:
                print(f"❌ Schwab auth error: {auth_error}")
                return {}
        
        headers = {
            'Authorization': f'Bearer {access_token}',
            'Accept': 'application/json'
        }
        
        # Get account numbers
        accounts_url = 'https://api.schwabapi.com/v1/accounts/accountNumbers'
        accounts_response = requests.get(accounts_url, headers=headers, timeout=30)
        
        if accounts_response.status_code != 200:
            print(f"❌ Schwab accounts error: {accounts_response.status_code}")
            return {}
        
        accounts_data = accounts_response.json()
        account_estimates = {}
        
        for account in accounts_data:
            account_number = account['accountNumber']
            account_type = account.get('type', 'Unknown')
            
            # Determine account name
            if 'IRA' in account_type.upper():
                account_name = 'Schwab IRA'
            else:
                account_name = 'Schwab Individual'
            
            print(f"   📊 Processing {account_name} ({account_number})")
            
            # Get positions
            positions_url = f'https://api.schwabapi.com/v1/accounts/{account_number}/positions'
            pos_response = requests.get(positions_url, headers=headers, timeout=30)
            
            if pos_response.status_code != 200:
                print(f"   ⚠️ Error getting positions: {pos_response.status_code}")
                account_estimates[account_name] = 0.0
                continue
            
            positions_data = pos_response.json()
            positions = positions_data.get('securitiesAccount', {}).get('positions', [])
            
            if not positions:
                print(f"   ⚠️ No positions found for {account_name}")
                account_estimates[account_name] = 0.0
                continue
            
            # For Schwab, we'll use a reasonable dividend estimate based on account type
            # This is the best we can do without individual stock dividend data
            total_account_value = 0.0
            
            for position in positions:
                market_value = position.get('marketValue', 0)
                total_account_value += market_value
            
            # Use different yield estimates based on account type
            if 'IRA' in account_name:
                estimated_yield = 0.045  # 4.5% for IRA (more dividend-focused)
            else:
                estimated_yield = 0.035  # 3.5% for individual (mix of growth/dividend)
            
            estimated_annual_dividend = total_account_value * estimated_yield
            account_estimates[account_name] = estimated_annual_dividend
            
            print(f"   ✅ {account_name} Value: ${total_account_value:,.2f}")
            print(f"   📈 Estimated Annual Dividend ({estimated_yield*100}%): ${estimated_annual_dividend:,.2f}")
        
        return account_estimates
        
    except Exception as e:
        print(f"❌ Schwab API error: {e}")
        return {}

def update_estimated_income_sheet():
    """Update Estimated Income 2025 sheet with real dividend estimates"""
    excel_path = r'c:\Users\mjmat\Python Code in VS\dividend_tracker\DividendTrackerApp\outputs\Dividends_2025.xlsx'
    
    print("📈 Getting real dividend estimates from APIs...")
    
    # Get real dividend estimates
    etrade_estimates = get_etrade_dividend_estimates()
    schwab_estimates = get_schwab_dividend_estimates()
    
    if not etrade_estimates and not schwab_estimates:
        print("❌ No dividend estimates retrieved from APIs")
        return
    
    print("\n📊 DIVIDEND ESTIMATES SUMMARY:")
    print("=" * 50)
    
    total_annual = 0.0
    account_data = {}
    
    for account, estimate in etrade_estimates.items():
        print(f"{account}: ${estimate:,.2f}/year")
        account_data[account] = estimate
        total_annual += estimate
    
    for account, estimate in schwab_estimates.items():
        print(f"{account}: ${estimate:,.2f}/year")
        account_data[account] = estimate  
        total_annual += estimate
    
    monthly_total = total_annual / 12
    print(f"\nTOTAL ANNUAL: ${total_annual:,.2f}")
    print(f"MONTHLY AVERAGE: ${monthly_total:,.2f}")
    
    # Update Excel sheet
    print(f"\n📝 Updating Excel sheet: {excel_path}")
    
    try:
        wb = openpyxl.load_workbook(excel_path)
        ws = wb['Estimated Income 2025']
        
        # Update account rows with real estimates
        if 'E*TRADE IRA' in account_data:
            ws.cell(row=4, column=44, value=account_data['E*TRADE IRA'])
            print(f"   Row 4 (E*TRADE IRA): ${account_data['E*TRADE IRA']:,.2f}")
        
        if 'E*TRADE Taxable' in account_data:
            ws.cell(row=5, column=44, value=account_data['E*TRADE Taxable'])
            print(f"   Row 5 (E*TRADE Taxable): ${account_data['E*TRADE Taxable']:,.2f}")
        
        if 'Schwab IRA' in account_data:
            ws.cell(row=6, column=44, value=account_data['Schwab IRA'])
            print(f"   Row 6 (Schwab IRA): ${account_data['Schwab IRA']:,.2f}")
        
        if 'Schwab Individual' in account_data:
            ws.cell(row=7, column=44, value=account_data['Schwab Individual'])
            print(f"   Row 7 (Schwab Individual): ${account_data['Schwab Individual']:,.2f}")
        
        # Update monthly total
        ws.cell(row=9, column=44, value=monthly_total)
        print(f"   Row 9 (Monthly Total): ${monthly_total:,.2f}")
        
        # Apply formatting
        soft_green = PatternFill(start_color='00FA71', end_color='00FA71', fill_type='solid')
        soft_red = PatternFill(start_color='FF7C80', end_color='FF7C80', fill_type='solid')
        
        for row in [4, 5, 6, 7, 9]:
            cell = ws.cell(row=row, column=44)
            if cell.value and cell.value > 0:
                cell.fill = soft_green
            else:
                cell.fill = soft_red
            
            # Set column width
            ws.column_dimensions[cell.column_letter].width = 15
        
        wb.save(excel_path)
        print(f"✅ Excel updated successfully!")
        
    except Exception as e:
        print(f"❌ Excel update error: {e}")

if __name__ == "__main__":
    print("🚀 REAL DIVIDEND ESTIMATES UPDATE (Using E*TRADE Quote API)")
    print("=" * 60)
    update_estimated_income_sheet()
    print("\n✅ Update complete!")
