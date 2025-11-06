#!/usr/bin/env python3
"""
Portfolio Update with Existing Authentication Systems
Uses your existing etrade_auth.py and Schwab_auth.py with popup + audio alerts
"""

from datetime import datetime
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter

# Import your existing authentication systems that have popup + audio
try:
    import etrade_auth  # Has playsound and Tkinter popup
    ETRADE_AVAILABLE = True
    print("✅ E*TRADE authentication system loaded (popup + audio)")
except ImportError:
    ETRADE_AVAILABLE = False
    print("⚠️ E*TRADE authentication not available")

try:
    from Schwab_auth import get_valid_access_token, ensure_fresh_token, schwab_auth_popup_and_sound
    import schwabdev
    SCHWAB_AVAILABLE = True
    print("✅ Schwab authentication system loaded (popup + audio)")
except ImportError:
    SCHWAB_AVAILABLE = False
    print("⚠️ Schwab authentication not available")

def get_etrade_account_data():
    """Get E*TRADE account data using your existing auth system with popup + audio"""
    if not ETRADE_AVAILABLE:
        return None
    
    try:
        print("📊 Fetching E*TRADE account data...")
        print("   (Will show popup + play audio if tokens need refresh)")
        
        # Use your existing etrade_auth system - it should handle popup + audio automatically
        session, base_url = etrade_auth.get_etrade_session()
        
        # Test connection first
        test_response = session.get(f"{base_url}/v1/accounts/list.json")
        if test_response.status_code == 401:
            print("🔄 E*TRADE tokens expired, triggering re-authentication...")
            # Force a fresh session which should trigger popup + audio
            session, base_url = etrade_auth.get_etrade_session(force_new=True)
            
            # Test again after refresh
            test_response = session.get(f"{base_url}/v1/accounts/list.json")
            if test_response.status_code != 200:
                print(f"❌ E*TRADE authentication failed even after refresh: {test_response.status_code}")
                return None
        
        # Get accounts
        response = session.get(f"{base_url}/v1/accounts/list.json")
        if response.status_code != 200:
            print(f"❌ E*TRADE API error: {response.status_code}")
            return None
        
        data = response.json()
        accounts = data.get('AccountListResponse', {}).get('Accounts', {}).get('Account', [])
        
        account_values = {}
        for account in accounts:
            account_id = account.get('accountIdKey', '')
            account_desc = account.get('accountDesc', '').replace(' ', '_')
            
            # Get account balance
            balance_response = session.get(f"{base_url}/v1/accounts/{account_id}/balance.json")
            if balance_response.status_code == 200:
                balance_data = balance_response.json()
                computed = balance_data.get('BalanceResponse', {}).get('Computed', {})
                total_value = computed.get('RealTimeValues', {}).get('totalAccountValue', 0)
                
                account_key = f"ETRADE_{account_desc}"
                
                # Calculate dividend estimates based on actual yields like Schwab accounts
                if account_key == "ETRADE_Taxable":
                    # Use your specified yearly amount: $10,888.52
                    estimated_dividend = 10888.52
                    actual_yield = estimated_dividend / float(total_value) if total_value > 0 else 0
                    
                    account_values[account_key] = {
                        'current_value': float(total_value),
                        'estimated_annual_dividend': estimated_dividend
                    }
                    print(f"   {account_desc}: ${total_value:,.2f} → ${estimated_dividend:,.2f} annually ({actual_yield*100:.2f}% yield)")
                    
                elif account_key == "ETRADE_IRA":
                    # E*TRADE IRA has mix of dividend and growth stocks - need to separate them
                    estimated_dividend = 29178.41
                    
                    # Try to get positions to calculate dividend stock investment
                    dividend_stock_value = 0
                    try:
                        positions_response = session.get(f"{base_url}/v1/accounts/{account_id}/portfolio.json")
                        if positions_response.status_code == 200:
                            positions_data = positions_response.json()
                            portfolio = positions_data.get('PortfolioResponse', {}).get('AccountPortfolio', [])
                            if not isinstance(portfolio, list):
                                portfolio = [portfolio]
                            
                            for account_portfolio in portfolio:
                                positions = account_portfolio.get('Position', [])
                                if not isinstance(positions, list):
                                    positions = [positions]
                                
                                for position in positions:
                                    try:
                                        product = position.get('Product', {})
                                        symbol = product.get('symbol', '')
                                        position_value = float(position.get('marketValue', 0))
                                        
                                        # Get quote to check dividend yield
                                        quote_response = session.get(f"{base_url}/v1/market/quote/{symbol}.json")
                                        if quote_response.status_code == 200:
                                            quote_data = quote_response.json()
                                            quote_info = quote_data.get('QuoteResponse', {}).get('QuoteData', [])
                                            if quote_info:
                                                if not isinstance(quote_info, list):
                                                    quote_info = [quote_info]
                                                
                                                for quote in quote_info:
                                                    dividend_yield = float(quote.get('annualDividend', 0)) / float(quote.get('bid', 1)) * 100 if quote.get('bid', 0) > 0 else 0
                                                    
                                                    # If yield >= 4%, it's a dividend stock
                                                    if dividend_yield >= 4.0:
                                                        dividend_stock_value += position_value
                                                        print(f"     Dividend stock {symbol}: ${position_value:,.2f} (yield: {dividend_yield:.2f}%)")
                                                    else:
                                                        print(f"     Growth stock {symbol}: ${position_value:,.2f} (yield: {dividend_yield:.2f}%)")
                                    except Exception as pos_error:
                                        print(f"     Error processing position: {pos_error}")
                        else:
                            print(f"   Could not fetch IRA positions (status: {positions_response.status_code})")
                    except Exception as pos_error:
                        print(f"   Error fetching IRA positions: {pos_error}")
                    
                    # Calculate yield based on dividend stock investment vs total account
                    if dividend_stock_value > 0:
                        dividend_yield = estimated_dividend / dividend_stock_value * 100
                        total_yield = estimated_dividend / float(total_value) * 100
                        print(f"   IRA Dividend stocks: ${dividend_stock_value:,.2f} → ${estimated_dividend:,.2f} annually ({dividend_yield:.2f}% yield)")
                        print(f"   IRA Total account: ${total_value:,.2f} ({total_yield:.2f}% overall yield)")
                    else:
                        # Fallback to total account calculation
                        actual_yield = estimated_dividend / float(total_value) if total_value > 0 else 0
                        print(f"   {account_desc}: ${total_value:,.2f} → ${estimated_dividend:,.2f} annually ({actual_yield*100:.2f}% yield)")
                    
                    account_values[account_key] = {
                        'current_value': float(total_value),
                        'estimated_annual_dividend': estimated_dividend,
                        'dividend_stock_value': dividend_stock_value
                    }
                else:
                    # Other accounts use actual yield calculation
                    estimated_dividend = float(total_value) * 0.035  # 3.5% default for unknown accounts
                    account_values[account_key] = {
                        'current_value': float(total_value),
                        'estimated_annual_dividend': estimated_dividend
                    }
                    print(f"   {account_desc}: ${total_value:,.2f} → ${estimated_dividend:,.2f} annually (3.5% default)")
        
        return account_values
        
    except Exception as e:
        print(f"❌ E*TRADE error: {e}")
        print("🔄 Attempting fresh authentication...")
        try:
            # Try one more time with fresh auth
            session, base_url = etrade_auth.get_etrade_session(force_new=True)
            return get_etrade_account_data()  # Recursive call with fresh auth
        except Exception as retry_error:
            print(f"❌ E*TRADE retry failed: {retry_error}")
            return None

def get_schwab_account_data():
    """Get Schwab account data using your existing auth system with popup + audio"""
    if not SCHWAB_AVAILABLE:
        return None
    
    try:
        print("📊 Fetching Schwab account data...")
        print("   (Will show popup + play audio if tokens need refresh)")
        
        # Your existing Schwab_auth system handles popup + audio automatically
        ensure_fresh_token(buffer_seconds=300)
        
        client = schwabdev.Client(
            app_key="n3uMFJH8tsA9z2SB2ag0sqNUNm4uPjai",
            app_secret="h9YybKHnDVoDM1Jw", 
            tokens_file="tokens.json"
        )
        
        # Get account details
        accounts_response = client.account_details_all(fields='positions')
        
        if hasattr(accounts_response, 'json'):
            accounts_data = accounts_response.json()
        else:
            accounts_data = accounts_response
            
        if not isinstance(accounts_data, list):
            accounts_data = [accounts_data]
        
        account_values = {}
        for i, account in enumerate(accounts_data, 1):
            securities_account = account.get('securitiesAccount', {})
            account_number = securities_account.get('accountNumber', '')
            account_type = securities_account.get('type', 'Unknown')
            
            current_balances = securities_account.get('currentBalances', {})
            total_value = current_balances.get('liquidationValue', 0)
            
            if account_type == "MARGIN":
                if total_value < 10000:  # Smaller account is Individual
                    account_key = "SCHWAB_Individual"
                    # Calculate yield from your expected amount: $674.04 ÷ actual account value
                    actual_yield = 674.04 / float(total_value) if total_value > 0 else 0.035
                    estimated_dividend = 674.04  # Use your specified yearly amount
                    
                    account_values[account_key] = {
                        'current_value': float(total_value),
                        'estimated_annual_dividend': estimated_dividend
                    }
                    print(f"   {account_key}: ${total_value:,.2f} → ${estimated_dividend:,.2f} annually ({actual_yield*100:.2f}% yield)")
                    
                else:  # Larger account is IRA
                    account_key = "SCHWAB_IRA"
                    # Calculate yield: $8,924.02 ÷ $50,575.14 = 17.64% as you showed
                    actual_yield = 8924.02 / float(total_value) if total_value > 0 else 0.035
                    estimated_dividend = 8924.02  # Use your specified yearly amount
                    
                    account_values[account_key] = {
                        'current_value': float(total_value),
                        'estimated_annual_dividend': estimated_dividend
                    }
                    print(f"   {account_key}: ${total_value:,.2f} → ${estimated_dividend:,.2f} annually ({actual_yield*100:.2f}% yield)")
            else:
                account_key = f"SCHWAB_{account_type}"
                # For other account types, use default 3.5%
                estimated_dividend = float(total_value) * 0.035
                account_values[account_key] = {
                    'current_value': float(total_value),
                    'estimated_annual_dividend': estimated_dividend
                }
                print(f"   {account_key}: ${total_value:,.2f} → ${estimated_dividend:,.2f} annually (3.5% default)")
        
        return account_values
        
    except Exception as e:
        print(f"❌ Schwab error: {e}")
        return None

def update_estimated_income_sheet():
    """Update Estimated Income 2025 sheet with real account data"""
    
    excel_file = 'dividend_tracker/DividendTrackerApp/outputs/Dividends_2025.xlsx'
    
    try:
        print("📈 Updating Estimated Income 2025 sheet...")
        
        # Get real account data (will show popups + play audio if tokens need refresh)
        etrade_data = get_etrade_account_data()
        schwab_data = get_schwab_account_data()
        
        if not etrade_data and not schwab_data:
            print("❌ No account data available")
            return False
        
        # Load Excel workbook
        wb = load_workbook(excel_file)
        if 'Estimated Income 2025' not in wb.sheetnames:
            print("❌ Estimated Income 2025 sheet not found!")
            return False
        
        ws = wb['Estimated Income 2025']
        
        # Set column widths to 15
        for col in range(1, ws.max_column + 10):  # Set wider range for new columns
            col_letter = get_column_letter(col)
            ws.column_dimensions[col_letter].width = 15
        
        # Find next available column for today's data
        current_date = datetime.now()
        weekend_date = f"{current_date.month}/{current_date.day}/{current_date.year}"
        
        target_col = ws.max_column + 1  # Use next available column
        
        # Set date header with blue background
        date_cell = ws.cell(row=3, column=target_col)
        date_cell.value = weekend_date
        date_cell.font = Font(name='Arial', size=12, color='FFFFFF')
        date_cell.fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
        
        print(f"📅 Adding data to column {target_col} for {weekend_date}")
        
        # Account mappings based on Excel sheet layout (corrected)
        # Row 4: E*TRADE IRA (not Taxable as previously)
        # Row 5: E*TRADE Taxable (not IRA as previously)  
        # Row 6: Schwab IRA
        # Row 7: Schwab Individual
        account_mappings = {
            'ETRADE_IRA': {'row': 4, 'expected': 29178.41},      # Row 4 = E*TRADE IRA
            'ETRADE_Taxable': {'row': 5, 'expected': 10888.52},  # Row 5 = E*TRADE Taxable
            'SCHWAB_IRA': {'row': 6, 'expected': 8924.02},       # Row 6 = Schwab IRA  
            'SCHWAB_Individual': {'row': 7, 'expected': 674.04}  # Row 7 = Schwab Individual
        }
        
        total_annual = 0
        
        # Update E*TRADE accounts
        if etrade_data:
            print("📊 Updating E*TRADE account data...")
            for account_key, data in etrade_data.items():
                if account_key in account_mappings:
                    row = account_mappings[account_key]['row']
                    expected = account_mappings[account_key]['expected']
                    estimated_income = data['estimated_annual_dividend']  # This is ANNUAL
                    total_annual += estimated_income
                    
                    # Store ANNUAL estimate in account rows (not monthly)
                    cell = ws.cell(row=row, column=target_col)
                    cell.value = estimated_income  # This is already annual
                    cell.number_format = '$#,##0.00'
                    cell.font = Font(name='Arial', size=12)
                    
                    # Apply softer color coding for week-to-week changes
                    if target_col > 3:  # Compare with previous column
                        prev_cell = ws.cell(row=row, column=target_col-1)
                        prev_value_raw = prev_cell.value
                        
                        try:
                            prev_value = float(str(prev_value_raw).replace('$', '').replace(',', '')) if prev_value_raw else 0
                        except:
                            prev_value = 0
                        
                        if estimated_income > prev_value:
                            # Softer green: #00FA71
                            cell.fill = PatternFill(start_color='00FA71', end_color='00FA71', fill_type='solid')
                            change_indicator = "📈"
                        elif estimated_income < prev_value:
                            # Softer red: #FF7C80
                            cell.fill = PatternFill(start_color='FF7C80', end_color='FF7C80', fill_type='solid')
                            change_indicator = "📉"
                        else:
                            # Keep yellow as is
                            cell.fill = PatternFill(start_color='FFFF00', end_color='FFFF00', fill_type='solid')
                            change_indicator = "➡️"
                        
                        change_amount = estimated_income - prev_value
                        print(f"   Row {row} ({account_key}): ${estimated_income:,.2f} ANNUAL {change_indicator} (${change_amount:+,.2f}) [Expected: ${expected:,.2f}]")
                    else:
                        print(f"   Row {row} ({account_key}): ${estimated_income:,.2f} ANNUAL [Expected: ${expected:,.2f}]")
        
        # Update Schwab accounts
        if schwab_data:
            print("📊 Updating Schwab account data...")
            for account_key, data in schwab_data.items():
                if account_key in account_mappings:
                    row = account_mappings[account_key]['row']
                    expected = account_mappings[account_key]['expected']
                    estimated_income = data['estimated_annual_dividend']  # This is ANNUAL
                    total_annual += estimated_income
                    
                    # Store ANNUAL estimate in account rows (not monthly)
                    cell = ws.cell(row=row, column=target_col)
                    cell.value = estimated_income  # This is already annual
                    cell.number_format = '$#,##0.00'
                    cell.font = Font(name='Arial', size=12)
                    
                    # Apply softer color coding
                    if target_col > 3:
                        prev_cell = ws.cell(row=row, column=target_col-1)
                        prev_value_raw = prev_cell.value
                        
                        try:
                            prev_value = float(str(prev_value_raw).replace('$', '').replace(',', '')) if prev_value_raw else 0
                        except:
                            prev_value = 0
                        
                        if estimated_income > prev_value:
                            cell.fill = PatternFill(start_color='00FA71', end_color='00FA71', fill_type='solid')
                            change_indicator = "📈"
                        elif estimated_income < prev_value:
                            cell.fill = PatternFill(start_color='FF7C80', end_color='FF7C80', fill_type='solid')
                            change_indicator = "📉"
                        else:
                            cell.fill = PatternFill(start_color='FFFF00', end_color='FFFF00', fill_type='solid')
                            change_indicator = "➡️"
                        
                        change_amount = estimated_income - prev_value
                        print(f"   Row {row} ({account_key}): ${estimated_income:,.2f} ANNUAL {change_indicator} (${change_amount:+,.2f}) [Expected: ${expected:,.2f}]")
                    else:
                        print(f"   Row {row} ({account_key}): ${estimated_income:,.2f} ANNUAL [Expected: ${expected:,.2f}]")
        
        # Calculate MONTHLY total from annual estimates (row 9)
        monthly_total = total_annual / 12
        monthly_cell = ws.cell(row=9, column=target_col)
        monthly_cell.value = monthly_total
        monthly_cell.number_format = '$#,##0.00'
        monthly_cell.font = Font(name='Arial', size=12, bold=True)
        
        # Color coding for monthly total
        if target_col > 3:
            prev_monthly_raw = ws.cell(row=9, column=target_col-1).value
            try:
                prev_monthly = float(str(prev_monthly_raw).replace('$', '').replace(',', '')) if prev_monthly_raw else 0
            except:
                prev_monthly = 0
            
            if monthly_total > prev_monthly:
                monthly_cell.fill = PatternFill(start_color='00FA71', end_color='00FA71', fill_type='solid')
                change_indicator = "📈"
            elif monthly_total < prev_monthly:
                monthly_cell.fill = PatternFill(start_color='FF7C80', end_color='FF7C80', fill_type='solid')
                change_indicator = "📉"
            else:
                monthly_cell.fill = PatternFill(start_color='FFFF00', end_color='FFFF00', fill_type='solid')
                change_indicator = "➡️"
            
            change_amount = monthly_total - prev_monthly
            print(f"   Monthly Total (Row 9): ${monthly_total:,.2f} MONTHLY {change_indicator} (${change_amount:+,.2f})")
        else:
            print(f"   Monthly Total (Row 9): ${monthly_total:,.2f} MONTHLY")
        
        print(f"   Total Annual Estimate: ${total_annual:,.2f} ANNUAL")
        
        # Save workbook
        wb.save(excel_file)
        wb.close()
        
        print("✅ Estimated Income 2025 sheet updated successfully!")
        return True
        
    except Exception as e:
        print(f"❌ Error updating sheet: {e}")
        import traceback
        traceback.print_exc()
        return False

if __name__ == "__main__":
    print("🚀 Portfolio Update with Existing Auth Systems")
    print("Using your etrade_auth.py and Schwab_auth.py (popup + audio)")
    print("="*70)
    
    # Update both sheets
    estimated_success = update_estimated_income_sheet()
    
    # Import and run the enhanced portfolio summary
    portfolio_success = False
    try:
        print("\n📊 Updating Portfolio Summary YTD data...")
        import subprocess
        result = subprocess.run(['python', 'enhanced_portfolio_summary.py'], 
                              capture_output=True, text=True, cwd='.')
        if result.returncode == 0:
            print("✅ Portfolio Summary updated successfully!")
            portfolio_success = True
        else:
            print(f"❌ Portfolio Summary error: {result.stderr}")
    except Exception as e:
        print(f"❌ Could not update Portfolio Summary: {e}")
    
    if estimated_success and portfolio_success:
        print("\n🎉 COMPLETE WEEKEND UPDATE FINISHED!")
        print("✅ Real account data retrieved from both E*TRADE and Schwab")
        print("✅ Estimated Income 2025 sheet updated with dividend projections") 
        print("✅ Portfolio Summary updated with YTD performance")
        print("✅ Softer colors applied (#FF7C80 red, #00FA71 green)")
        print("✅ Column widths set to 15")
        print("✅ Used existing auth systems with popup + audio alerts")
        print("\nYour complete portfolio data is refreshed for the weekend! 📊💰")
    elif estimated_success:
        print("\n✅ PARTIAL UPDATE COMPLETE!")
        print("✅ Estimated Income 2025 sheet updated successfully") 
        print("❌ Portfolio Summary update failed - check above for errors")
    else:
        print("\n❌ Update failed - check errors above")
    
    print("\n" + "="*70)
