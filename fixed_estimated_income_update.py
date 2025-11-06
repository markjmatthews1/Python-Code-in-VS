#!/usr/bin/env python3
"""
Fixed Estimated Income 2025 Update - Corrects row placement and annual/monthly calculations
- E*TRADE data goes in rows 4-5 (annual amounts)
- Schwab data goes in rows 6-7 (annual amounts) 
- Row 9 shows monthly total (annual amounts from rows 4-7 divided by 12)
- Skips E*TRADE futures account ending in 7650
"""

import sys
import json
import time
from datetime import datetime
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter

# E*TRADE imports
try:
    from etrade_auth import get_etrade_session
    ETRADE_AVAILABLE = True
except ImportError:
    ETRADE_AVAILABLE = False
    print("⚠️ E*TRADE authentication not available")

# Schwab imports
try:
    from Schwab_auth import get_valid_access_token, ensure_fresh_token
    import schwabdev
    SCHWAB_AVAILABLE = True
except ImportError:
    SCHWAB_AVAILABLE = False
    print("⚠️ Schwab authentication not available")

def get_etrade_account_values():
    """Get real account values from E*TRADE using existing auth system"""
    if not ETRADE_AVAILABLE:
        print("❌ E*TRADE not available")
        return None
    
    try:
        print("📊 Fetching E*TRADE account values...")
        
        # Use the existing etrade_auth system that has popup and audio
        import etrade_auth
        
        # This will automatically handle token refresh with popup and audio if needed
        session, base_url = etrade_auth.get_etrade_session()
        
        # Get account list
        url = f"{base_url}/v1/accounts/list.json"
        response = session.get(url)
        
        if response.status_code != 200:
            print(f"❌ Failed to get E*TRADE accounts: {response.status_code}")
            return None
        
        data = response.json()
        accounts = data.get("AccountListResponse", {}).get("Accounts", {}).get("Account", [])
        
        if not isinstance(accounts, list):
            accounts = [accounts]
        
        print(f"✅ Found {len(accounts)} E*TRADE accounts")
        
        account_values = {}
        
        for account in accounts:
            account_id = account.get("accountIdKey")
            account_desc = account.get("accountDesc", "Unknown")
            account_type = account.get("accountType", "Unknown")
            
            # Skip the futures account ending in 7650 (empty/unused)
            if account_id and str(account_id).endswith('7650'):
                print(f"   ⏭️ Skipping E*TRADE futures account: {account_desc} (empty futures account)")
                continue
            
            print(f"   Processing E*TRADE account: {account_desc} ({account_type})")
            
            # Get account balance
            balance_url = f"{base_url}/v1/accounts/{account_id}/balance.json"
            balance_response = session.get(balance_url)
            
            if balance_response.status_code == 200:
                balance_data = balance_response.json()
                balance_info = balance_data.get("BalanceResponse", {})
                
                # Get total account value
                total_value = 0
                if "Computed" in balance_info:
                    computed = balance_info["Computed"]
                    total_value = computed.get("RealTimeValues", {}).get("totalAccountValue", 0)
                
                # Determine account type for mapping
                if "taxable" in account_desc.lower() or "individual" in account_desc.lower():
                    account_key = "ETRADE_Taxable"
                elif "ira" in account_desc.lower() or "rollover" in account_desc.lower():
                    account_key = "ETRADE_IRA"
                else:
                    account_key = f"ETRADE_{account_desc.replace(' ', '_')}"
                
                # Calculate ANNUAL dividend estimate (3.5% for taxable, 4% for IRA)
                if "ira" in account_desc.lower():
                    annual_dividend_rate = 0.04  # 4% for IRA
                else:
                    annual_dividend_rate = 0.035  # 3.5% for taxable
                
                annual_dividend = float(total_value) * annual_dividend_rate
                
                account_values[account_key] = {
                    'account_desc': account_desc,
                    'account_type': account_type,
                    'current_value': float(total_value),
                    'estimated_annual_dividend': annual_dividend  # This is the ANNUAL amount
                }
                
                print(f"   💰 Account Value: ${total_value:,.2f}")
                print(f"   📈 Estimated ANNUAL Dividends ({annual_dividend_rate*100}%): ${annual_dividend:,.2f}")
            
            else:
                print(f"   ⚠️ Could not get balance for {account_desc}")
        
        return account_values
        
    except Exception as e:
        print(f"❌ Error fetching E*TRADE data: {e}")
        import traceback
        traceback.print_exc()
        return None

def get_schwab_account_values():
    """Get real account values from Schwab"""
    if not SCHWAB_AVAILABLE:
        print("❌ Schwab not available")
        return None
    
    try:
        print("📊 Fetching Schwab account values...")
        
        # Ensure we have a fresh token
        ensure_fresh_token(buffer_seconds=300)
        
        client = schwabdev.Client(
            app_key="n3uMFJH8tsA9z2SB2ag0sqNUNm4uPjai",
            app_secret="h9YybKHnDVoDM1Jw", 
            tokens_file="tokens.json"
        )
        
        print("✅ Schwab client initialized")
        
        # Get all account details
        accounts_response = client.account_details_all(fields='positions')
        
        if hasattr(accounts_response, 'json'):
            accounts_data = accounts_response.json()
        else:
            accounts_data = accounts_response
            
        if not isinstance(accounts_data, list):
            accounts_data = [accounts_data]
        
        account_values = {}
        
        for i, account in enumerate(accounts_data, 1):
            securities_account = account.get('securitiesAccount', {})
            account_number = securities_account.get('accountNumber', '')
            account_type = securities_account.get('type', 'Unknown')
            
            print(f"   Processing Schwab account {i}: ...{account_number[-4:]} ({account_type})")
            
            # Get current balance
            initial_balances = securities_account.get('initialBalances', {})
            current_balances = securities_account.get('currentBalances', {})
            
            # Try different balance fields
            total_value = (
                current_balances.get('totalValue') or
                current_balances.get('liquidationValue') or
                current_balances.get('longMarketValue') or
                initial_balances.get('totalValue') or
                0
            )
            
            # Account mapping
            if 'ira' in account_type.lower():
                account_key = 'SCHWAB_IRA'
                annual_dividend_rate = 0.045  # 4.5% for IRA
            else:
                account_key = 'SCHWAB_Individual'
                annual_dividend_rate = 0.035  # 3.5% for individual
            
            annual_dividend = float(total_value) * annual_dividend_rate
            
            account_values[account_key] = {
                'account_number': account_number,
                'account_type': account_type,
                'current_value': float(total_value),
                'estimated_annual_dividend': annual_dividend  # This is the ANNUAL amount
            }
            
            print(f"   💰 Account Value: ${total_value:,.2f}")
            print(f"   📈 Estimated ANNUAL Dividends ({annual_dividend_rate*100}%): ${annual_dividend:,.2f}")
        
        return account_values
        
    except Exception as e:
        print(f"❌ Error fetching Schwab data: {e}")
        import traceback
        traceback.print_exc()
        return None

def update_estimated_income_sheet(excel_file="Bryan Perry Transactions.xlsx"):
    """Update the Estimated Income 2025 sheet with real account data"""
    print("🔄 Starting Estimated Income 2025 update...")
    
    try:
        # Get data from both sources
        etrade_data = get_etrade_account_values()
        schwab_data = get_schwab_account_values()
        
        if not etrade_data and not schwab_data:
            print("❌ No account data available from either source")
            return False
        
        # Load Excel file
        wb = load_workbook(excel_file)
        
        if 'Estimated Income 2025' not in wb.sheetnames:
            print("❌ Estimated Income 2025 sheet not found!")
            return False
        
        ws = wb['Estimated Income 2025']
        print(f"📋 Current sheet size: {ws.max_row} rows x {ws.max_column} columns")
        
        # Set column widths to 15 for all columns
        print("📐 Setting column widths to 15...")
        for col in range(1, ws.max_column + 5):  # Extra columns for future use
            col_letter = get_column_letter(col)
            ws.column_dimensions[col_letter].width = 15
        
        # Find next available column (don't overwrite existing data)
        current_date = datetime.now()
        weekend_date = f"{current_date.month}/{current_date.day}/{current_date.year}"
        
        target_col = None
        # Look for a completely empty column
        for col in range(2, ws.max_column + 10):  # Check beyond current max
            column_empty = True
            for row in range(3, 10):  # Check key rows
                if ws.cell(row=row, column=col).value is not None:
                    column_empty = False
                    break
            if column_empty:
                target_col = col
                break
        
        if not target_col:
            target_col = ws.max_column + 1
        
        print(f"📅 Using column {target_col} for date: {weekend_date}")
        
        # Set up the date header with blue background
        date_cell = ws.cell(row=3, column=target_col)
        date_cell.value = weekend_date
        date_cell.font = Font(name='Arial', size=12, color='FFFFFF')
        date_cell.fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
        
        # CORRECTED Account row mappings (annual amounts in rows 4-7)
        account_mappings = {
            'ETRADE_Taxable': {'row': 4, 'expected': 10888.52, 'description': 'E*TRADE Taxable'},
            'ETRADE_IRA': {'row': 5, 'expected': 29178.41, 'description': 'E*TRADE IRA'},
            'SCHWAB_Individual': {'row': 6, 'expected': 674.04, 'description': 'Schwab Individual'},
            'SCHWAB_IRA': {'row': 7, 'expected': 8924.02, 'description': 'Schwab IRA'}
        }
        
        # Update E*TRADE accounts (rows 4-5 with ANNUAL amounts)
        if etrade_data:
            print("📊 Updating E*TRADE account data (ANNUAL amounts in rows 4-5)...")
            for account_key, data in etrade_data.items():
                if account_key in account_mappings:
                    row = account_mappings[account_key]['row']
                    expected = account_mappings[account_key]['expected']
                    description = account_mappings[account_key]['description']
                    
                    # Use ANNUAL dividend amount (not monthly)
                    annual_dividend = data['estimated_annual_dividend']
                    
                    cell = ws.cell(row=row, column=target_col)
                    cell.value = annual_dividend  # ANNUAL amount goes in the row
                    cell.number_format = '$#,##0.00'
                    cell.font = Font(name='Arial', size=12)
                    
                    # Apply softer color coding
                    if target_col > 3:
                        prev_cell = ws.cell(row=row, column=target_col-1)
                        prev_value_raw = prev_cell.value
                        
                        # Handle both numeric and string values
                        try:
                            prev_value = float(str(prev_value_raw).replace('$', '').replace(',', '')) if prev_value_raw else 0
                        except:
                            prev_value = 0
                        
                        if annual_dividend > prev_value:
                            # Softer green: #00FA71
                            cell.fill = PatternFill(start_color='00FA71', end_color='00FA71', fill_type='solid')
                            change_indicator = "📈"
                        elif annual_dividend < prev_value:
                            # Softer red: #FF7C80
                            cell.fill = PatternFill(start_color='FF7C80', end_color='FF7C80', fill_type='solid')
                            change_indicator = "📉"
                        else:
                            # Keep yellow as is
                            cell.fill = PatternFill(start_color='FFFF00', end_color='FFFF00', fill_type='solid')
                            change_indicator = "➡️"
                        
                        change_amount = annual_dividend - prev_value
                        print(f"   Row {row} ({description}): ${annual_dividend:,.2f} ANNUAL {change_indicator} (${change_amount:+,.2f})")
                    else:
                        print(f"   Row {row} ({description}): ${annual_dividend:,.2f} ANNUAL")
                else:
                    print(f"   ⚠️ Unknown E*TRADE account key: {account_key}")
        else:
            print("⚠️ No E*TRADE data available")
        
        # Update Schwab accounts (rows 6-7 with ANNUAL amounts)
        if schwab_data:
            print("📊 Updating Schwab account data (ANNUAL amounts in rows 6-7)...")
            for account_key, data in schwab_data.items():
                if account_key in account_mappings:
                    row = account_mappings[account_key]['row']
                    expected = account_mappings[account_key]['expected']
                    description = account_mappings[account_key]['description']
                    
                    # Use ANNUAL dividend amount (not monthly)
                    annual_dividend = data['estimated_annual_dividend']
                    
                    cell = ws.cell(row=row, column=target_col)
                    cell.value = annual_dividend  # ANNUAL amount goes in the row
                    cell.number_format = '$#,##0.00'
                    cell.font = Font(name='Arial', size=12)
                    
                    # Apply softer color coding
                    if target_col > 3:
                        prev_cell = ws.cell(row=row, column=target_col-1)
                        prev_value_raw = prev_cell.value
                        
                        # Handle both numeric and string values
                        try:
                            prev_value = float(str(prev_value_raw).replace('$', '').replace(',', '')) if prev_value_raw else 0
                        except:
                            prev_value = 0
                        
                        if annual_dividend > prev_value:
                            # Softer green: #00FA71
                            cell.fill = PatternFill(start_color='00FA71', end_color='00FA71', fill_type='solid')
                            change_indicator = "📈"
                        elif annual_dividend < prev_value:
                            # Softer red: #FF7C80
                            cell.fill = PatternFill(start_color='FF7C80', end_color='FF7C80', fill_type='solid')
                            change_indicator = "📉"
                        else:
                            # Keep yellow as is
                            cell.fill = PatternFill(start_color='FFFF00', end_color='FFFF00', fill_type='solid')
                            change_indicator = "➡️"
                        
                        change_amount = annual_dividend - prev_value
                        print(f"   Row {row} ({description}): ${annual_dividend:,.2f} ANNUAL {change_indicator} (${change_amount:+,.2f})")
                    else:
                        print(f"   Row {row} ({description}): ${annual_dividend:,.2f} ANNUAL")
                else:
                    print(f"   ⚠️ Unknown Schwab account key: {account_key}")
        else:
            print("⚠️ No Schwab data available")
        
        # Calculate and update monthly total (row 9) - sum ANNUAL amounts from rows 4-7 and divide by 12
        monthly_row = 9
        total_annual = 0
        
        print("🧮 Calculating monthly total from annual amounts...")
        for row in range(4, 8):  # Rows 4-7 contain ANNUAL amounts
            cell_value = ws.cell(row=row, column=target_col).value
            if isinstance(cell_value, (int, float)):
                total_annual += cell_value
                print(f"   Row {row}: ${cell_value:,.2f} (annual)")
            else:
                print(f"   Row {row}: No data")
        
        monthly_total = total_annual / 12  # Divide annual by 12 to get monthly
        monthly_cell = ws.cell(row=monthly_row, column=target_col)
        monthly_cell.value = monthly_total
        monthly_cell.number_format = '$#,##0.00'
        monthly_cell.font = Font(name='Arial', size=12, bold=True)
        monthly_cell.fill = PatternFill(start_color='D9E2F3', end_color='D9E2F3', fill_type='solid')
        
        print(f"📋 Row 9 (Monthly Total): ${monthly_total:,.2f} (${total_annual:,.2f} annual ÷ 12)")
        
        # Save the workbook
        wb.save(excel_file)
        wb.close()
        
        print(f"✅ Estimated Income 2025 sheet updated successfully!")
        print(f"💾 Data saved to: {excel_file}")
        
        return True
        
    except Exception as e:
        print(f"❌ Error updating Estimated Income 2025: {e}")
        import traceback
        traceback.print_exc()
        return False

def main():
    """Main function to update the Estimated Income 2025 sheet"""
    print("=" * 60)
    print("🔄 FIXED ESTIMATED INCOME 2025 UPDATE")
    print("=" * 60)
    print("📋 This script will:")
    print("   • Put E*TRADE data in rows 4-5 (ANNUAL amounts)")
    print("   • Put Schwab data in rows 6-7 (ANNUAL amounts)")
    print("   • Calculate monthly total in row 9 (annual ÷ 12)")
    print("   • Skip E*TRADE futures account ending in 7650")
    print("   • Use softer colors (#00FA71 green, #FF7C80 red)")
    print("=" * 60)
    
    success = update_estimated_income_sheet()
    
    if success:
        print("🎉 Update completed successfully!")
    else:
        print("💥 Update failed!")
    
    return success

if __name__ == "__main__":
    main()
