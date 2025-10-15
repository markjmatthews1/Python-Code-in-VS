#!/usr/bin/env python3
"""
Calculate yearly estimated dividend income for Schwab IRA account
using ticker yield data from IRA account and current positions.
"""

import json
import sys
import os
from datetime import datetime
import openpyxl
from openpyxl.utils import column_index_from_string
import requests

# Use the root directory Schwab authentication system
from Schwab_auth import get_valid_access_token, ensure_fresh_token

def format_dividend_cell(ws, row, col, value):
    """Apply standard formatting to dividend income cells."""
    from openpyxl.styles import PatternFill, Font
    
    # Set the value
    ws.cell(row=row, column=col, value=value)
    
    # Apply Arial 12 font
    arial_font = Font(name='Arial', size=12)
    ws.cell(row=row, column=col).font = arial_font
    
    # Apply currency formatting
    ws.cell(row=row, column=col).number_format = '$#,##0.00'
    
    # Apply light red background color (#FF7C80)
    fill = PatternFill(start_color='FF7C80', end_color='FF7C80', fill_type='solid')
    ws.cell(row=row, column=col).fill = fill
    
    return ws.cell(row=row, column=col)

def load_ticker_yields():
    """Load ticker yield data from the IRA dividend data file."""
    try:
        with open('actual_ira_dividend_data_20250825.json', 'r') as f:
            data = json.load(f)
        
        ticker_yields = {}
        for symbol, info in data.get('all_current_holdings', {}).items():
            if info.get('has_dividend', False):
                ticker_yields[symbol] = {
                    'yield': info.get('yield', 0.0),
                    'dividend_per_share': info.get('dividend_per_share', 0.0),
                    'annual_dividend': info.get('dividend_per_share', 0.0) * 4  # Assume quarterly
                }
        
        # Add manual entries for tickers not in IRA but known to pay dividends
        # (This will be updated as we find Schwab-specific tickers)
        
        print(f"✅ Loaded yield data for {len(ticker_yields)} dividend-paying tickers")
        return ticker_yields
    except Exception as e:
        print(f"❌ Error loading ticker yields: {e}")
        return {}

def get_schwab_ira_positions():
    """Get current positions from Schwab IRA account using existing auth system."""
    try:
        print("🔄 Getting Schwab IRA account positions...")
        
        # Ensure we have a valid token using existing auth system
        access_token = get_valid_access_token()
        if not access_token:
            print("❌ Could not get valid Schwab access token")
            ensure_fresh_token()  # Try to refresh
            access_token = get_valid_access_token()
            if not access_token:
                print("❌ Still no valid token after refresh attempt")
                return []
        
        print("✅ Using existing Schwab access token")
        
        # Set up headers with the access token
        headers = {
            'Authorization': f'Bearer {access_token}',
            'Accept': 'application/json'
        }
        
        # Get all accounts with positions included
        base_url = "https://api.schwabapi.com"
        accounts_url = f"{base_url}/trader/v1/accounts?fields=positions"
        
        response = requests.get(accounts_url, headers=headers)
        
        if response.status_code != 200:
            print(f"❌ Error getting accounts: {response.status_code} - {response.text}")
            return []
        
        accounts = response.json()
        print(f"📊 Found {len(accounts)} Schwab accounts")
        
        # Find the IRA account by account number (91562183) and get its positions
        ira_positions = None
        target_account_number = "91562183"
        
        for account in accounts:
            # Get account info from securitiesAccount
            securities_account = account.get('securitiesAccount', {})
            account_number = securities_account.get('accountNumber', '')
            
            if target_account_number == str(account_number):
                # Get positions from this account
                ira_positions = securities_account.get('positions', [])
                print(f"✅ Found Schwab IRA account: {account_number}")
                break
        
        if ira_positions is None:
            print(f"❌ Could not find Schwab IRA account {target_account_number}")
            # Show all accounts for debugging
            for account in accounts:
                securities_account = account.get('securitiesAccount', {})
                account_number = securities_account.get('accountNumber', 'Unknown')
                print(f"   Available: {account_number}")
            return []
        
        if not ira_positions:
            print("❌ No positions found in Schwab IRA account")
            return []
        
        print(f"📈 Found {len(ira_positions)} positions in Schwab IRA account")
        
        # Convert to our expected format
        formatted_positions = []
        for position in ira_positions:
            try:
                instrument = position.get('instrument', {})
                symbol = instrument.get('symbol', '').strip().upper()
                quantity = float(position.get('longQuantity', 0))
                market_value = float(position.get('marketValue', 0))
                
                if symbol and quantity > 0 and market_value > 0:
                    formatted_positions.append({
                        'symbol': symbol,
                        'quantity': quantity,
                        'market_value': market_value
                    })
                    print(f"   📊 {symbol}: {quantity:,.0f} shares = ${market_value:,.2f}")
                    
            except Exception as e:
                print(f"   ⚠️ Error processing position: {e}")
                continue
        
        return formatted_positions
        
    except Exception as e:
        print(f"❌ Error getting Schwab positions: {e}")
        return []

def calculate_dividend_income(positions, ticker_yields):
    """Calculate total annual dividend income from positions."""
    total_annual_income = 0.0
    dividend_breakdown = {}
    
    print("\n💰 CALCULATING ANNUAL DIVIDEND INCOME:")
    print("=" * 60)
    
    for position in positions:
        try:
            # This will need to be updated based on Schwab API response format
            symbol = position.get('symbol', 'UNKNOWN')
            current_value = float(position.get('market_value', 0))
            quantity = float(position.get('quantity', 0))
            
            # Skip if no value
            if current_value <= 0:
                continue
            
            # Check if we have dividend data for this ticker
            if symbol in ticker_yields:
                yield_data = ticker_yields[symbol]
                yield_percent = yield_data['yield'] / 100.0  # Convert percentage to decimal
                
                # Calculate annual dividend income using market value and yield
                annual_income = current_value * yield_percent
                total_annual_income += annual_income
                
                dividend_breakdown[symbol] = {
                    'quantity': quantity,
                    'current_value': current_value,
                    'yield_percent': yield_data['yield'],
                    'annual_income': annual_income
                }
                
                print(f"📊 {symbol}: ${current_value:,.2f} × {yield_data['yield']:.2f}% = ${annual_income:,.2f}")
            else:
                print(f"⚠️  {symbol}: ${current_value:,.2f} (no dividend data)")
                
        except Exception as e:
            print(f"❌ Error processing position {symbol}: {e}")
            continue
    
    print("=" * 60)
    print(f"🎯 TOTAL ANNUAL DIVIDEND INCOME: ${total_annual_income:,.2f}")
    
    return total_annual_income, dividend_breakdown

def update_excel_sheet(annual_income):
    """Update the Estimated Income 2025 sheet with the calculated amount."""
    try:
        excel_path = r'c:\Users\mjmat\Python Code in VS\dividend_tracker\DividendTrackerApp\outputs\Dividends_2025.xlsx'
        wb = openpyxl.load_workbook(excel_path)
        
        if 'Estimated Income 2025' not in wb.sheetnames:
            print("❌ 'Estimated Income 2025' sheet not found")
            return False
        
        ws = wb['Estimated Income 2025']
        
        # Use column AI (35) directly for 8/24/2025
        target_col = column_index_from_string('AI')  # Column 35
        
        # Verify the date is correct
        date_cell = ws.cell(row=3, column=target_col).value
        if str(date_cell).strip() != "8/24/2025":
            print(f"⚠️  Expected 8/24/2025 in column AI, found: {date_cell}")
        
        # Add the annual income to row 6 (Schwab IRA) with proper formatting
        format_dividend_cell(ws, 6, target_col, annual_income)
        
        wb.save(excel_path)
        print(f"✅ Updated Estimated Income 2025 sheet:")
        print(f"   Date: 8/24/2025 (Column AI)")
        print(f"   Row 6 (Schwab IRA): ${annual_income:,.2f}")
        print(f"   Formatting: Arial 12, Currency ($#,##0.00), Light Red Background (#FF7C80)")
        
        return True
            
    except Exception as e:
        print(f"❌ Error updating Excel sheet: {e}")
        return False

def main():
    """Main function to calculate and update Schwab IRA dividend income."""
    print("💼 CALCULATING SCHWAB IRA ANNUAL DIVIDEND INCOME")
    print("=" * 60)
    
    # Load ticker yields from IRA data
    ticker_yields = load_ticker_yields()
    if not ticker_yields:
        print("❌ No ticker yield data available")
        return
    
    # Get Schwab IRA positions
    positions = get_schwab_ira_positions()
    if not positions:
        print("❌ No positions found in Schwab IRA account")
        print("ℹ️  This may be due to Schwab API authentication issues")
        return
    
    # Calculate dividend income
    annual_income, breakdown = calculate_dividend_income(positions, ticker_yields)
    
    if annual_income > 0:
        # Update Excel sheet
        if update_excel_sheet(annual_income):
            print("\n🎉 SUCCESS! Schwab IRA dividend income calculated and updated.")
        else:
            print("\n⚠️  Calculation complete but Excel update failed.")
            print(f"   Manual entry needed: ${annual_income:,.2f} in row 6, column AI")
    else:
        print("\n⚠️  No dividend income calculated - check positions and ticker data")

if __name__ == "__main__":
    main()
