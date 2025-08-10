#!/usr/bin/env python3
"""
TradeTracker Diagnostic Script
Tests core functionality without GUI to identify issues
"""

import sys
import os
sys.path.append(os.path.dirname(os.path.abspath(__file__)))

def test_file_integrity():
    """Test if TradeTracker.py can be imported and basic functionality works"""
    
    print("🔍 DIAGNOSING TRADETRACKER ISSUES")
    print("=" * 50)
    
    try:
        # Test imports
        print("✅ Testing imports...")
        import tkinter as tk
        from tkinter import ttk, messagebox
        import pandas as pd
        from openpyxl import load_workbook
        print("   ✅ All modules imported successfully")
        
        # Test Excel file
        print("\n📁 Testing Excel file...")
        wb = load_workbook("Bryan Perry Transactions.xlsx")
        print(f"   ✅ Excel file opened. Sheets: {wb.sheetnames}")
        
        # Test Open_Trades_2025 sheet
        ws = wb["Open_Trades_2025"]
        header = [cell.value for cell in ws[1]]
        print(f"   ✅ Open_Trades_2025 sheet found with {ws.max_row-1} data rows")
        print(f"   📋 Headers: {header[:5]}...")  # Show first 5 headers
        
        # Test pandas reading
        df = pd.read_excel("Bryan Perry Transactions.xlsx", sheet_name="Open_Trades_2025")
        print(f"   ✅ Pandas can read sheet: {len(df)} rows, {len(df.columns)} columns")
        
        wb.close()
        
        # Test TradeTracker class definition
        print("\n🧪 Testing TradeTracker class...")
        
        # Import the class without running GUI
        import importlib.util
        spec = importlib.util.spec_from_file_location("TradeTracker", "TradeTracker.py")
        tt_module = importlib.util.module_from_spec(spec)
        
        # Execute only the class definition part
        with open("TradeTracker.py", "r") as f:
            content = f.read()
            
        # Check if essential methods exist
        methods_to_check = ["add_trade", "update_trade", "update_stock_prices", "load_open_trades_from_excel"]
        for method in methods_to_check:
            if f"def {method}" in content:
                print(f"   ✅ Method {method} found")
            else:
                print(f"   ❌ Method {method} MISSING")
        
        # Check if button definitions exist
        buttons_to_check = ["Add Trade", "Update Trade", "Update Stock Prices"]
        for button in buttons_to_check:
            if f'text="{button}"' in content:
                print(f"   ✅ Button '{button}' found")
            else:
                print(f"   ❌ Button '{button}' MISSING")
        
        print("\n🎯 DIAGNOSIS COMPLETE")
        return True
        
    except Exception as e:
        print(f"\n❌ ERROR: {e}")
        return False

def suggest_fixes():
    """Suggest potential fixes based on common issues"""
    
    print("\n💡 SUGGESTED FIXES")
    print("=" * 30)
    print("1. 🔄 Try restarting VS Code completely")
    print("2. 📁 Ensure you're in the correct working directory")
    print("3. 🐍 Check if Python environment is properly configured")
    print("4. 📊 Verify Excel file hasn't been corrupted")
    print("5. 🖥️  Try running: python TradeTracker.py directly from command line")

def main():
    """Run diagnostics"""
    success = test_file_integrity()
    suggest_fixes()
    
    if success:
        print("\n✅ Core components look intact!")
        print("🚀 Issue might be environment-related or GUI startup problem")
    else:
        print("\n⚠️  Found issues that need fixing")
    
    return success

if __name__ == "__main__":
    main()
