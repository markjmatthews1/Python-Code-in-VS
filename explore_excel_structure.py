#!/usr/bin/env python3
"""
Explore Excel File Structure
Find where your historical yield data is actually stored
"""

import pandas as pd
import openpyxl
import os

def explore_excel_structure():
    """Explore the Excel file to understand the data structure"""
    print("🔍 Exploring your Excel file structure...")
    
    excel_file = r"c:\Users\mjmat\Python Code in VS\dividend_tracker\DividendTrackerApp\outputs\Dividends_2025.xlsx"
    
    if not os.path.exists(excel_file):
        print(f"❌ Excel file not found: {excel_file}")
        return
    
    wb = openpyxl.load_workbook(excel_file, data_only=True)
    
    print(f"📋 Found {len(wb.sheetnames)} sheets:")
    for i, sheet_name in enumerate(wb.sheetnames, 1):
        print(f"   {i}. {sheet_name}")
    
    # Check each sheet for data structure
    for sheet_name in wb.sheetnames:
        print(f"\n🔍 Analyzing sheet: {sheet_name}")
        ws = wb[sheet_name]
        
        # Get headers (first row)
        headers = []
        for col in range(1, min(ws.max_column + 1, 50)):  # Check first 50 columns
            cell_value = ws.cell(row=1, column=col).value
            if cell_value:
                headers.append(str(cell_value))
            else:
                break
        
        print(f"   📊 Dimensions: {ws.max_row} rows × {ws.max_column} columns")
        print(f"   📋 Headers ({len(headers)}): {headers[:10]}...")  # Show first 10 headers
        
        # Look for date columns
        date_columns = []
        for header in headers:
            if '-2025' in str(header) or '2025' in str(header):
                date_columns.append(header)
        
        if date_columns:
            print(f"   📅 Date columns found ({len(date_columns)}): {date_columns[:5]}...")
        
        # Check for tickers
        has_tickers = False
        ticker_column = None
        for i, header in enumerate(headers):
            if 'ticker' in header.lower() or 'symbol' in header.lower():
                has_tickers = True
                ticker_column = i + 1
                break
        
        if has_tickers:
            print(f"   🎯 Has ticker data in column {ticker_column}")
            
            # Show sample ticker data
            sample_tickers = []
            for row in range(2, min(ws.max_row + 1, 7)):  # Check first 5 data rows
                ticker_cell = ws.cell(row=row, column=ticker_column)
                if ticker_cell.value:
                    sample_tickers.append(str(ticker_cell.value))
            
            if sample_tickers:
                print(f"   📈 Sample tickers: {sample_tickers}")
                
                # If this sheet has both tickers and date columns, show sample data
                if date_columns:
                    print(f"   ✅ This sheet appears to have historical yield data!")
                    
                    # Show sample yield data for first ticker
                    first_ticker = sample_tickers[0]
                    sample_yields = {}
                    
                    for date_col in date_columns[:5]:  # Check first 5 date columns
                        try:
                            col_index = headers.index(date_col) + 1
                            yield_value = ws.cell(row=2, column=col_index).value
                            if yield_value is not None:
                                sample_yields[date_col] = yield_value
                        except:
                            pass
                    
                    if sample_yields:
                        print(f"   📊 Sample yields for {first_ticker}: {sample_yields}")
        
        print("   " + "-" * 60)
    
    wb.close()

if __name__ == "__main__":
    explore_excel_structure()
