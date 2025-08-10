#!/usr/bin/env python3
"""
Check for 0.00 cosmetic issues
"""

import openpyxl

def check_zeros():
    print("=== CHECKING FOR COSMETIC 0.00 VALUES ===\n")
    
    wb = openpyxl.load_workbook('Bryan Perry Transactions.xlsx')
    ws = wb['Open_Trades_2025']
    
    print("Sample data showing potential 0.00 issues:")
    print()
    
    for row in range(2, min(ws.max_row + 1, 8)):
        ticker = ws.cell(row=row, column=4).value  # Stock ticker column
        if ticker:
            print(f'Row {row} ({ticker}):')
            zero_fields = []
            for col in range(1, 21):
                header = ws.cell(row=1, column=col).value
                value = ws.cell(row=row, column=col).value
                if value == 0 or (isinstance(value, (int, float)) and value == 0.0):
                    zero_fields.append(f'  {header}: {value}')
            
            if zero_fields:
                for field in zero_fields:
                    print(field)
            else:
                print("  No zero values found")
            print()
    
    wb.close()

if __name__ == "__main__":
    check_zeros()
