import tkinter as tk
from tkinter import ttk, messagebox
import os
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Alignment, Font

# --- Explicit mapping from Excel columns to app field names ---
EXCEL_TO_APP_COLS = {
    "Stock Purchase Type": "Type",
    "Trade Type": "Type",  # Alternative header name
    "Expiration Date": "Expiration",
    "Strike price": "Strike",
    "Stock ticker": "Ticker",
    "Ticker": "Ticker",  # Alternative header name
    "Date of transaction": "Date",
    "Trade Date": "Date",  # Alternative header name
    "Number of shares +/-": "Shares",
    "Shares": "Shares",  # Alternative header name
    "Quantity": "Shares",  # Alternative header name
    "Cost": "Cost",
    "Total investment": "Investment",
    "Investment": "Investment",  # Alternative header name
    "Assigned/exercised Y/N": "Assigned/Exercised",
    "Assigned/Exercised": "Assigned/Exercised",  # Alternative header name
    "Call Value": "Call Value",
    "Put Value": "Put Value",
    "Put cash requirement": "Put Cash Req",
    "Cash Requirement": "Put Cash Req",  # Alternative header name
    "Notes": "Notes",
    "Close Date": "Close Date",
    "Close Price": "Close Price",
    "Realized profit loss": "Realized P/L",
    "Realized P/L": "Realized P/L",  # Alternative header name
    "P&L %": "P&L %",
    "Notes2": "Notes2",
    "Current Price": "Current Price",
    "Current P/L": "Current P/L"
}

# Columns used in the app and Excel
COLUMNS = [
    "Type", "Expiration", "Strike", "Ticker", "Date", "Shares", "Cost", "Investment",
    "Assigned/Exercised", "Call Value", "Put Value", "Put Cash Req", "Notes", "Close Date",
    "Close Price", "Realized P/L", "P&L %", "Notes2", "Current Price", "Current P/L"
]
DATE_COLUMNS = ["Expiration", "Date", "Close Date"]


class TradeTrackerApp:
    def find_header_column(self, header, possible_names):
        """Return 1-based column index for the first matching name in header, or None if not found."""
        for idx, col_name in enumerate(header, start=1):
            for name in possible_names:
                if col_name and str(col_name).strip().replace(" ", "").lower() == str(name).strip().replace(" ", "").lower():
                    return idx
        return None
    def clean_blank(self, value):
        """Return an empty string if value is None or blank, else stripped value."""
        if value is None:
            return ""
        if isinstance(value, str):
            return value.strip()
        return value
    # --- Helper functions and constants available to all methods ---
    @staticmethod
    def format_currency(val):
        try:
            num = float(str(val).replace("$","").replace(",","").strip())
            if num == 0.0:
                return ""
            return "$" + format(num, ",.2f")
        except Exception:
            return ""

    @staticmethod
    def parse_num(val):
        try:
            return float(str(val).replace("$","").replace(",","").strip())
        except Exception:
            return 0.0


    def __init__(self, root):
        # Columns that represent monetary values
        self.MONEY_COLS = [
            "Cost", "Investment", "Close Price", "Realized P/L", "Current Price", "Current P/L",
            "Call Value", "Put Value", "Put Cash Req", "Strike"
        ]
        self.root = root
        self.root.title("TradeTracker - Open Trades")
        
        # Set window size and make it resizable
        self.root.geometry("1400x800")  # Width x Height in pixels
        self.root.minsize(800, 600)     # Minimum size
        self.root.state('zoomed')       # Start maximized on Windows
        
        self.trades = []

        ENTRY_COLS = [
            "Type", "Expiration", "Strike", "Ticker", "Date", "Shares", "Cost", "Investment",
            "Assigned/Exercised", "Notes", "Close Date", "Close Price", "Notes2"
        ]
        self.vars = {col: tk.StringVar() for col in ENTRY_COLS}
        
        # Create main container for entry fields with horizontal scrolling
        entry_container = tk.Frame(self.root)
        entry_container.pack(fill="x", padx=10, pady=5)
        
        # Canvas and scrollbar for horizontal scrolling of entry fields
        entry_canvas = tk.Canvas(entry_container, height=80)
        entry_h_scrollbar = tk.Scrollbar(entry_container, orient="horizontal", command=entry_canvas.xview)
        entry_canvas.configure(xscrollcommand=entry_h_scrollbar.set)
        
        entry_frame = tk.Frame(entry_canvas)
        entry_canvas.create_window((0, 0), window=entry_frame, anchor="nw")
        
        entry_canvas.pack(side="top", fill="x", expand=True)
        entry_h_scrollbar.pack(side="bottom", fill="x")

        # Entry field widths based on content
        field_widths = {
            "Type": 15,
            "Expiration": 12,
            "Strike": 8,
            "Ticker": 8,
            "Date": 12,
            "Shares": 8,
            "Cost": 10,
            "Investment": 12,
            "Assigned/Exercised": 16,
            "Notes": 20,
            "Close Date": 12,
            "Close Price": 10,
            "Notes2": 20
        }

        for idx, col in enumerate(ENTRY_COLS):
            width = field_widths.get(col, 12)
            lbl = tk.Label(entry_frame, text=col, font=("Arial", 12))
            lbl.grid(row=0, column=idx, sticky="ew", padx=2)
            ent = tk.Entry(entry_frame, textvariable=self.vars[col], width=width, font=("Arial", 12))
            ent.grid(row=1, column=idx, sticky="ew", padx=2)

        # Update canvas scroll region
        entry_frame.update_idletasks()
        entry_canvas.configure(scrollregion=entry_canvas.bbox("all"))
        
        # Add mouse wheel scrolling for entry canvas
        def on_entry_mousewheel(event):
            if event.state & 0x1:  # Shift key pressed
                entry_canvas.xview_scroll(int(-1*(event.delta/120)), "units")
        
        entry_canvas.bind("<MouseWheel>", on_entry_mousewheel)
        entry_frame.bind("<MouseWheel>", on_entry_mousewheel)

        # Buttons in a separate frame below the entry fields
        button_frame = tk.Frame(self.root)
        button_frame.pack(fill="x", padx=10, pady=5)
        
        add_btn = tk.Button(button_frame, text="Add Trade", command=self.add_trade, bg="#4CAF50", fg="white", font=("Arial", 12, "bold"))
        add_btn.pack(side="left", padx=5, pady=2)
        update_btn = tk.Button(button_frame, text="Update Trade", command=self.update_trade, bg="#2196F3", fg="white", font=("Arial", 12, "bold"))
        update_btn.pack(side="left", padx=5, pady=2)
        update_prices_btn = tk.Button(button_frame, text="Update Stock Prices", command=self.update_stock_prices, bg="#FFEB3B", fg="black", font=("Arial", 12, "bold"))
        update_prices_btn.pack(side="left", padx=5, pady=2)
        fix_pl_btn = tk.Button(button_frame, text="Fix P/L Calculations", command=self.fix_missing_current_pl, bg="#FF9800", fg="white", font=("Arial", 12, "bold"))
        fix_pl_btn.pack(side="left", padx=5, pady=2)
        results_btn = tk.Button(button_frame, text="Generate Results", command=self.generate_results, bg="#9C27B0", fg="white", font=("Arial", 12, "bold"))
        results_btn.pack(side="left", padx=5, pady=2)

        # Treeview for displaying trades
        tree_frame = tk.Frame(self.root)
        tree_frame.pack(fill="both", expand=True, padx=10, pady=10)
        
        # Configure treeview style for larger font
        style = ttk.Style()
        style.configure("Treeview", font=("Arial", 12))
        style.configure("Treeview.Heading", font=("Arial", 12, "bold"))
        
        self.tree = ttk.Treeview(tree_frame, columns=COLUMNS, show="headings", height=15)
        
        # Set column widths appropriate for larger font
        column_widths = {
            "Type": 120,
            "Expiration": 100,
            "Strike": 80,
            "Ticker": 80,
            "Date": 100,
            "Shares": 80,
            "Cost": 100,
            "Investment": 120,
            "Assigned/Exercised": 140,
            "Call Value": 120,
            "Put Value": 120,
            "Put Cash Req": 130,
            "Notes": 150,
            "Close Date": 100,
            "Close Price": 100,
            "Realized P/L": 120,
            "P&L %": 80,
            "Notes2": 150,
            "Current Price": 120,
            "Current P/L": 120
        }
        
        for col in COLUMNS:
            self.tree.heading(col, text=col)
            width = column_widths.get(col, 100)
            self.tree.column(col, width=width, anchor="center", minwidth=50)
        vsb = ttk.Scrollbar(tree_frame, orient="vertical", command=self.tree.yview)
        hsb = ttk.Scrollbar(tree_frame, orient="horizontal", command=self.tree.xview)
        self.tree.configure(yscrollcommand=vsb.set, xscrollcommand=hsb.set)
        self.tree.grid(row=0, column=0, sticky="nsew")
        vsb.grid(row=0, column=1, sticky="ns")
        hsb.grid(row=1, column=0, sticky="ew")
        tree_frame.grid_rowconfigure(0, weight=1)
        tree_frame.grid_columnconfigure(0, weight=1)
        
        # Bind treeview selection event to populate entry fields
        self.tree.bind("<<TreeviewSelect>>", self.on_tree_select)

        # Load trades on startup
        self.load_open_trades_from_excel()
    
    def on_tree_select(self, event):
        """Handle treeview row selection to populate entry fields for editing"""
        selected = self.tree.selection()
        if not selected:
            return
        
        # Get the selected item
        item = selected[0]
        values = self.tree.item(item, "values")
        
        # Populate entry fields with selected row data
        for idx, col in enumerate(COLUMNS):
            if idx < len(values) and col in self.vars:
                self.vars[col].set(values[idx] if values[idx] else "")
                
        # Debug: Show what was loaded into Cost field
        if "Cost" in self.vars:
            print(f"DEBUG: Cost field populated with: '{self.vars['Cost'].get()}'")

    def force_refresh_cost(self):
        """Force refresh cost field to ensure changes are captured"""
        if "Cost" in self.vars:
            cost_widget = None
            # Find the Cost entry widget
            for widget in self.root.winfo_children():
                if hasattr(widget, 'winfo_children'):
                    for child in widget.winfo_children():
                        if hasattr(child, 'winfo_children'):
                            for grandchild in child.winfo_children():
                                if hasattr(grandchild, 'get') and hasattr(grandchild, 'textvariable'):
                                    if grandchild.textvariable == self.vars['Cost']:
                                        cost_widget = grandchild
                                        break
            if cost_widget:
                # Force the widget to update its textvariable
                cost_widget.update()
                print(f"DEBUG: Force refreshed Cost field: '{self.vars['Cost'].get()}'")
                
    def add_trade(self):
        # Build trade entry with calculated fields
        from datetime import datetime
        entry = {}
        for col in COLUMNS:
            entry[col] = self.vars[col].get().strip() if col in self.vars else ""

        # Auto-convert ticker to uppercase
        if entry["Ticker"]:
            entry["Ticker"] = entry["Ticker"].upper()



        # Format and calculate fields
        # Dates: m/d/yyyy (no leading zeros, no time)
        def format_date(val):
            if val:
                try:
                    val_str = str(val).split()[0]  # Remove time part first
                    dt = None
                    for fmt in ("%m/%d/%Y", "%Y-%m-%d", "%m-%d-%Y", "%Y/%m/%d"):
                        try:
                            dt = datetime.strptime(val_str, fmt)
                            break
                        except Exception:
                            continue
                    if dt:
                        # Format without leading zeros: m/d/yyyy
                        month = dt.month
                        day = dt.day  
                        year = dt.year
                        return f"{month}/{day}/{year}"
                except Exception:
                    pass
            return val

        # Currency formatting for GUI
        def format_currency(val):
            try:
                num = float(str(val).replace("$","").replace(",","").strip())
                # Return blank for zero values in most cases (cosmetic cleanup)
                if num == 0.0:
                    return ""
                return "$" + format(num, ",.2f")
            except Exception:
                return ""

        # List of monetary columns (including Strike)
        MONEY_COLS = ["Cost", "Investment", "Close Price", "Realized P/L", "Current Price", "Current P/L", "Call Value", "Put Value", "Put Cash Req", "Strike"]

        # Dates
        for dcol in DATE_COLUMNS:
            val = entry[dcol]
            # If Bought Stock, Expiration is always blank
            if dcol == "Expiration" and entry["Type"].strip().lower() == "bought stock":
                entry[dcol] = ""
            elif not val or val.lower() in ("nat", "none", "nan"):
                entry[dcol] = ""
            else:
                entry[dcol] = format_date(val.split()[0])  # Remove time if present

        # Numbers and calculated fields
        def parse_num(val):
            try:
                return float(str(val).replace("$","").replace(",","").strip())
            except Exception:
                return 0.0

        shares = parse_num(entry["Shares"])
        cost = parse_num(entry["Cost"])
        strike = parse_num(entry["Strike"])
        trade_type = entry["Type"].lower().strip()

        # Trade type specific calculations with assignment handling
        assigned_exercised = entry.get("Assigned/Exercised", "").strip().upper()
        
        if "put" in trade_type and "sold" in trade_type:
            # Sold Put
            if assigned_exercised == "Y":
                # Assigned put: calculate investment as (strike - cost) * shares
                if strike and cost and shares:
                    assignment_investment = (strike - cost) * abs(shares)
                    entry["Investment"] = self.format_currency(assignment_investment) if assignment_investment > 0 else ""
                else:
                    entry["Investment"] = ""
                # Clear option-specific fields for assignments
                entry["Put Value"] = ""
                entry["Put Cash Req"] = ""
                entry["Call Value"] = ""  # Not applicable for put trades
            else:
                # Regular sold put
                entry["Investment"] = ""  # No investment for sold options
                entry["Put Value"] = self.format_currency(cost * abs(shares)) if cost and shares and cost > 0 else ""
                entry["Put Cash Req"] = self.format_currency(strike * abs(shares)) if strike and shares and strike > 0 else ""
                entry["Call Value"] = ""  # Not applicable for put trades
        elif "call" in trade_type and "sold" in trade_type:
            # Sold Call
            if assigned_exercised == "Y":
                # Exercised call: calculate based on assignment
                entry["Investment"] = ""  # No additional investment needed
                entry["Call Value"] = ""
                entry["Put Value"] = ""  # Not applicable for call trades
                entry["Put Cash Req"] = ""  # Not applicable for calls
            else:
                # Regular sold call
                entry["Investment"] = ""  # No investment for sold options
                entry["Call Value"] = self.format_currency(cost * abs(shares)) if cost and shares and cost > 0 else ""
                entry["Put Value"] = ""  # Not applicable for call trades
                entry["Put Cash Req"] = ""  # Not applicable for calls
        elif "put" in trade_type and "bought" in trade_type:
            # Bought Put
            if assigned_exercised == "Y":
                # Exercised bought put: clear option values
                entry["Investment"] = ""
                entry["Put Value"] = ""
                entry["Put Cash Req"] = ""
                entry["Call Value"] = ""  # Not applicable for put trades
            else:
                # Regular bought put
                entry["Investment"] = self.format_currency(abs(shares) * cost) if shares and cost and cost > 0 else ""
                entry["Put Value"] = self.format_currency(cost * abs(shares)) if cost and shares and cost > 0 else ""
                entry["Put Cash Req"] = ""  # Not applicable for bought puts
                entry["Call Value"] = ""  # Not applicable for put trades
        elif "call" in trade_type and "bought" in trade_type:
            # Bought Call
            if assigned_exercised == "Y":
                # Exercised bought call: calculate investment as strike * shares
                if strike and shares:
                    exercise_investment = strike * abs(shares)
                    entry["Investment"] = self.format_currency(exercise_investment) if exercise_investment > 0 else ""
                else:
                    entry["Investment"] = ""
                entry["Call Value"] = ""
                entry["Put Value"] = ""  # Not applicable for call trades
                entry["Put Cash Req"] = ""  # Not applicable for calls
            else:
                # Regular bought call
                entry["Investment"] = self.format_currency(abs(shares) * cost) if shares and cost and cost > 0 else ""
                entry["Call Value"] = self.format_currency(cost * abs(shares)) if cost and shares and cost > 0 else ""
                entry["Put Value"] = ""  # Not applicable for call trades
                entry["Put Cash Req"] = ""  # Not applicable for calls
        else:
            # Bought Stock or other
            entry["Investment"] = self.format_currency(shares * cost) if shares and cost and cost > 0 else ""
            entry["Put Value"] = ""  # Not applicable for stock trades
            entry["Put Cash Req"] = ""  # Not applicable for stock trades
            entry["Call Value"] = ""  # Not applicable for stock trades

        # Current price handling (use cost as placeholder for now)
        current_price = cost if cost and cost > 0 else 0
        entry["Current Price"] = format_currency(current_price) if current_price > 0 else ""

        # Realized P/L and P&L % (if closed) - these can show zero if meaningful
        close_price = parse_num(entry["Close Price"])
        if close_price and shares and cost:
            realized_pl = (close_price - cost) * shares
            entry["Realized P/L"] = f"${realized_pl:,.2f}" if realized_pl != 0 else "$0.00"  # Show zero for closed trades
            entry["P&L %"] = f"{((close_price-cost)/cost*100):.2f}%" if cost else ""
        else:
            entry["Realized P/L"] = ""
            entry["P&L %"] = ""

        # Current P/L (open) - can show zero if meaningful
        if current_price and shares and cost:
            current_pl = (current_price - cost) * shares
            entry["Current P/L"] = f"${current_pl:,.2f}" if current_pl != 0 else "$0.00"  # Show zero for active trades
        else:
            entry["Current P/L"] = ""

        # Format all monetary columns for GUI (except Shares) - BUT PRESERVE CALCULATED VALUES
        for mcol in MONEY_COLS:
            if mcol in entry and entry[mcol] != "" and mcol != "Shares":
                # Don't reformat if it's already been calculated as blank for sold options
                if mcol == "Investment" and entry[mcol] == "":
                    continue  # Keep it blank for sold options
                try:
                    entry[mcol] = format_currency(float(str(entry[mcol]).replace("$","").replace(",","").strip()))
                except Exception:
                    entry[mcol] = format_currency(entry[mcol])




        # Build trade row in COLUMNS order for GUI
        trade = [entry[col] for col in COLUMNS]
        self.trades.append(trade)
        self.tree.insert("", "end", values=trade)

        # Write to Excel (store numbers as numbers, currency as string only for display columns)
        excel_filename = "Put Transactions.xlsx"
        sheet_name = "Open_Trades_2025"
        try:
            wb = load_workbook(excel_filename)
            ws = wb[sheet_name]
            header = [cell.value for cell in ws[1]]
            # Find the first empty row after the header (row 1)
            insert_row = None
            for row in range(2, ws.max_row + 2):
                if all((ws.cell(row=row, column=col).value in (None, "") for col in range(1, len(header) + 1))):
                    insert_row = row
                    break
            if insert_row is None:
                insert_row = ws.max_row + 1
            ws.insert_rows(insert_row)
            for col_idx, col_name in enumerate(header, start=1):
                app_col = EXCEL_TO_APP_COLS.get(col_name, col_name)
                value = entry[app_col] if app_col in entry else ""
                # Check for assignment/exercise status
                assigned_exercised = entry.get("Assigned/Exercised", "").strip().upper()
                # Calculate and store as number for numeric columns with assignment logic
                if app_col in ["Strike", "Shares", "Cost", "Investment", "Close Price", "Realized P/L", "Current Price", "Current P/L", "Call Value", "Put Value", "Put Cash Req"]:
                    if app_col == "Investment":
                        if assigned_exercised == "Y":
                            if "put" in trade_type and "sold" in trade_type:
                                value = (strike - cost) * abs(shares) if strike and cost and shares else None
                            elif "call" in trade_type and "bought" in trade_type:
                                value = strike * abs(shares) if strike and shares else None
                            else:
                                value = None
                        else:
                            if ("put" in trade_type and "sold" in trade_type) or ("call" in trade_type and "sold" in trade_type):
                                value = None
                            elif "put" in trade_type and "bought" in trade_type:
                                value = abs(shares) * cost if shares and cost else None
                            elif "call" in trade_type and "bought" in trade_type:
                                value = abs(shares) * cost if shares and cost else None
                            else:
                                value = shares * cost if shares and cost else None
                    elif app_col == "Put Value":
                        if assigned_exercised == "Y":
                            value = None
                        elif "put" in trade_type and cost and shares:
                            value = cost * abs(shares)
                        else:
                            value = None
                    elif app_col == "Put Cash Req":
                        if assigned_exercised == "Y":
                            value = None
                        elif "put" in trade_type and "sold" in trade_type and strike and shares:
                            value = strike * abs(shares)
                        else:
                            value = None
                    elif app_col == "Call Value":
                        if assigned_exercised == "Y":
                            value = None
                        elif "call" in trade_type and cost and shares:
                            value = cost * abs(shares)
                        else:
                            value = None
                    elif app_col == "Realized P/L":
                        value = (close_price - cost) * shares if close_price and cost and shares else None
                    elif app_col == "Current P/L":
                        value = (current_price - cost) * shares if current_price and cost and shares else None
                    elif app_col == "Current Price":
                        value = current_price if current_price else None
                    elif app_col == "Close Price":
                        value = close_price if close_price else None
                    elif app_col == "Strike":
                        value = strike if strike else None
                    elif app_col == "Cost":
                        value = cost if cost else None
                    elif app_col == "Shares":
                        value = shares if shares else None
                    else:
                        try:
                            value = float(str(value).replace("$","").replace(",","").strip()) if value != "" else None
                        except Exception:
                            value = None
                else:
                    value = entry[app_col] if app_col in entry else ""
                cell = ws.cell(row=insert_row, column=col_idx, value=value)
                if app_col in DATE_COLUMNS:
                    cell.alignment = Alignment(horizontal="right")
                elif app_col == "Assigned/Exercised":
                    cell.alignment = Alignment(horizontal="center")
                    cell.font = Font(name="Arial", size=12, bold=True)
                elif app_col in ["Strike", "Cost", "Investment", "Close Price", "Realized P/L", "Current Price", "Current P/L", "Call Value", "Put Value", "Put Cash Req"] and value is not None:
                    cell.number_format = "$#,##0.00"
            wb.save(excel_filename)
            wb.close()
            # Sort Excel trades by ticker (existing logic)
            self.sort_excel_trades()
            self.load_open_trades_from_excel()
            self.sort_treeview_by_ticker_and_date()
        except Exception as e:
            messagebox.showerror("Excel Error", f"Could not add trade to Excel.\n{e}")
        for var in self.vars.values():
            var.set("")

    def update_trade(self):
        print("DEBUG: update_trade() called")
        
        # CRITICAL: Capture ALL entered values FIRST before any other processing
        entered_close_date = ""
        entered_close_price = ""
        entered_notes = ""
        entered_assigned = ""
        if hasattr(self, 'vars') and "Close Date" in self.vars:
            entered_close_date = self.vars["Close Date"].get().strip()
        if hasattr(self, 'vars') and "Close Price" in self.vars:
            entered_close_price = self.vars["Close Price"].get().strip()
        if hasattr(self, 'vars') and "Notes" in self.vars:
            entered_notes = self.vars["Notes"].get().strip()
        if hasattr(self, 'vars') and "Assigned/Exercised" in self.vars:
            entered_assigned = self.vars["Assigned/Exercised"].get().strip()
        print(f"DEBUG: Captured entered values - Close Date: '{entered_close_date}', Close Price: '{entered_close_price}', Notes: '{entered_notes}', Assigned: '{entered_assigned}'")
        
        selected = self.tree.selection()
        if not selected:
            messagebox.showwarning("No Selection", "Select a trade to update.")
            return
        item = selected[0]
        values = self.tree.item(item, "values")
        
        # Only populate entry fields that exist in the GUI
        ENTRY_COLS = [
            "Type", "Expiration", "Strike", "Ticker", "Date", "Shares", "Cost", "Investment",
            "Assigned/Exercised", "Notes", "Close Date", "Close Price", "Notes2"
        ]
        
        for idx, col in enumerate(COLUMNS):
            if col in ENTRY_COLS and col in self.vars:
                value = values[idx] if idx < len(values) else ""
                # Don't overwrite entered values
                if col == "Close Date" and entered_close_date:
                    value = entered_close_date
                elif col == "Close Price" and entered_close_price:
                    value = entered_close_price
                elif col == "Notes" and entered_notes:
                    value = entered_notes
                elif col == "Assigned/Exercised" and entered_assigned:
                    value = entered_assigned
                self.vars[col].set(value)
                print(f"DEBUG: Setting {col} = '{value}'")
        
        def save_update():
            print("DEBUG: save_update() called")
            # Always define Excel workbook and worksheet at the start
            excel_filename = "Put Transactions.xlsx"
            sheet_name = "Open_Trades_2025"
            wb = load_workbook(excel_filename)
            ws = wb[sheet_name]
            # Build updated entry - preserve original values, only update assignment fields
            from datetime import datetime
            entry = {}
            
            # First, get current values from treeview as baseline (preserve original data)
            current_values = self.tree.item(item, "values")
            for idx, col in enumerate(COLUMNS):
                if idx < len(current_values):
                    entry[col] = current_values[idx] if current_values[idx] else ""
                else:
                    entry[col] = ""

            # Store original identifying values before GUI update for row matching
            original_values = {}
            for col in ["Ticker", "Date", "Shares", "Strike", "Cost", "Notes"]:
                original_values[col] = entry.get(col, "")

            # Always update all entry fields from the GUI (including Cost)
            for col in ENTRY_COLS:
                if col in self.vars:
                    gui_value = self.vars[col].get().strip()
                    entry[col] = gui_value
                    print(f"DEBUG: Updated from GUI - {col} = '{gui_value}'")

            # Always use the Cost from the GUI for all calculations and updates
            cost_from_gui = entry["Cost"]

            print(f"DEBUG: Final Close Date = '{entry.get('Close Date', '')}', Close Price = '{entry.get('Close Price', '')}'")

            # Auto-convert ticker to uppercase
            if entry["Ticker"]:
                entry["Ticker"] = entry["Ticker"].upper()

            # Recalculate Investment for assigned puts before writing to Excel
            trade_type = entry["Type"].lower().strip()
            assigned_exercised = entry.get("Assigned/Exercised", "").strip().upper()
            try:
                shares = float(str(entry["Shares"]).replace("$","").replace(",","").strip())
            except Exception:
                shares = 0.0
            try:
                cost = float(str(entry["Cost"]).replace("$","").replace(",","").strip())
            except Exception:
                cost = 0.0
            try:
                strike = float(str(entry["Strike"]).replace("$","").replace(",","").strip())
            except Exception:
                strike = 0.0

            # Always get header and column indices before any for-loop that uses them
            header = [cell.value for cell in ws[1]]
            ticker_col = self.find_header_column(header, ["Stock ticker", "Ticker"])
            date_col = self.find_header_column(header, ["Date of transaction", "Trade Date", "Date"])
            shares_col = self.find_header_column(header, ["Number of shares +/-", "Shares", "Quantity"])
            strike_col = self.find_header_column(header, ["Strike price", "Strike"])
            cost_col = self.find_header_column(header, ["Cost"])
            notes_col = self.find_header_column(header, ["Notes"])

            investment_numeric = None
            if "put" in trade_type and "sold" in trade_type and assigned_exercised == "Y":
                # Always calculate and set Investment for assigned Sold Put
                if strike and cost and shares:
                    assignment_investment = (strike - cost) * abs(shares)
                    investment_numeric = assignment_investment
                    entry["Investment"] = self.format_currency(assignment_investment) if assignment_investment != 0 else ""
                    # Update the GUI field immediately
                    if "Investment" in self.vars:
                        self.vars["Investment"].set(entry["Investment"])
                    print(f"DEBUG: Assigned Put - Investment calculated for GUI/Excel: (${strike} - ${cost}) * {abs(shares)} = ${assignment_investment}")
                else:
                    entry["Investment"] = ""
                # Find row in Excel matching Ticker, Date, Shares
                match_row = None
                print(f"DEBUG: Looking for ticker='{original_values['Ticker']}', date='{original_values['Date']}', shares='{original_values['Shares']}', strike='{original_values['Strike']}', cost='{original_values['Cost']}', notes='{original_values['Notes']}'")
                print(f"DEBUG: Excel columns - ticker_col={ticker_col}, date_col={date_col}, shares_col={shares_col}, strike_col={strike_col}, cost_col={cost_col}, notes_col={notes_col}")
                potential_matches = []  # Store all potential matches to handle duplicates
                for r in range(2, ws.max_row+1):
                    ticker = ws.cell(row=r, column=ticker_col).value
                    date = ws.cell(row=r, column=date_col).value
                    shares = ws.cell(row=r, column=shares_col).value
                    strike = ws.cell(row=r, column=strike_col).value if strike_col else None
                    cost = ws.cell(row=r, column=cost_col).value if cost_col else None
                    notes = ws.cell(row=r, column=notes_col).value if notes_col else None

                    # Use original values for matching
                    match_score = 0
                    if str(ticker).strip().upper() == str(original_values["Ticker"]).strip().upper():
                        match_score += 2
                    if str(date).strip() == str(original_values["Date"]).strip():
                        match_score += 2
                    if str(shares).strip() == str(original_values["Shares"]).strip():
                        match_score += 2
                    if strike is not None and str(strike).strip() == str(original_values["Strike"]).strip():
                        match_score += 1
                    if cost is not None and str(cost).strip() == str(original_values["Cost"]).strip():
                        match_score += 1
                    if notes is not None and str(notes).strip() == str(original_values["Notes"]).strip():
                        match_score += 1
                    if match_score >= 7:
                        potential_matches.append((r, match_score, notes))

            # Now, after the for-loop, handle assignment/exercise and trade-type logic
            assigned_exercised = entry.get("Assigned/Exercised", "").strip().upper()
            print(f"DEBUG: Assignment status captured: '{assigned_exercised}' (raw: '{entry.get('Assigned/Exercised', 'NOT_FOUND')}')")
            
            if "put" in trade_type and "sold" in trade_type:
                # Sold Put
                if assigned_exercised == "Y":
                    # Assigned put: calculate investment as (strike - cost) * shares
                    if strike and cost and shares:
                        assignment_investment = (strike - cost) * abs(shares)
                        entry["Investment"] = self.format_currency(assignment_investment) if assignment_investment > 0 else ""
                        print(f"DEBUG: Assigned Put - Investment calculated: (${strike} - ${cost}) * {abs(shares)} = ${assignment_investment}")
                    else:
                        entry["Investment"] = ""
                    # Clear option-specific fields for assignments
                    entry["Put Value"] = ""
                    entry["Put Cash Req"] = ""
                    entry["Call Value"] = ""  # Not applicable for put trades
                else:
                    # Regular sold put
                    entry["Investment"] = ""  # No investment for sold options
                    entry["Put Value"] = self.format_currency(cost * abs(shares)) if cost and shares and cost > 0 else ""
                    entry["Put Cash Req"] = self.format_currency(strike * abs(shares)) if strike and shares and strike > 0 else ""
                    entry["Call Value"] = ""  # Not applicable for put trades
                print(f"DEBUG UPDATE: Sold Put (Assigned: {assigned_exercised}) - Investment: '{entry['Investment']}', Put Value: '{entry['Put Value']}', Put Cash Req: '{entry['Put Cash Req']}'")
            elif "call" in trade_type and "sold" in trade_type:
                # Sold Call
                if assigned_exercised == "Y":
                    # Exercised call: calculate based on assignment
                    entry["Investment"] = ""  # No additional investment needed
                    entry["Call Value"] = ""
                    entry["Put Value"] = ""  # Not applicable for call trades
                    entry["Put Cash Req"] = ""  # Not applicable for calls
                else:
                    # Regular sold call
                    entry["Investment"] = ""  # No investment for sold options
                    entry["Call Value"] = self.format_currency(cost * abs(shares)) if cost and shares and cost > 0 else ""
                    entry["Put Value"] = ""  # Not applicable for call trades
                    entry["Put Cash Req"] = ""  # Not applicable for calls
            elif "put" in trade_type and "bought" in trade_type:
                # Bought Put
                if assigned_exercised == "Y":
                    # Exercised bought put: clear option values
                    entry["Investment"] = ""
                    entry["Put Value"] = ""
                    entry["Put Cash Req"] = ""
                    entry["Call Value"] = ""  # Not applicable for put trades
                else:
                    # Regular bought put
                    entry["Investment"] = self.format_currency(abs(shares) * cost) if shares and cost and cost > 0 else ""
                    entry["Put Value"] = self.format_currency(cost * abs(shares)) if cost and shares and cost > 0 else ""
                    entry["Put Cash Req"] = ""  # Not applicable for bought puts
                    entry["Call Value"] = ""  # Not applicable for put trades
            elif "call" in trade_type and "bought" in trade_type:
                # Bought Call
                if assigned_exercised == "Y":
                    # Exercised bought call: calculate investment as strike * shares
                    if strike and shares:
                        exercise_investment = strike * abs(shares)
                        entry["Investment"] = self.format_currency(exercise_investment) if exercise_investment > 0 else ""
                    else:
                        entry["Investment"] = ""
                    entry["Call Value"] = ""
                    entry["Put Value"] = ""  # Not applicable for call trades
                    entry["Put Cash Req"] = ""  # Not applicable for calls
                else:
                    # Regular bought call
                    entry["Investment"] = self.format_currency(abs(shares) * cost) if shares and cost and cost > 0 else ""
                    entry["Call Value"] = self.format_currency(cost * abs(shares)) if cost and shares and cost > 0 else ""
                    entry["Put Value"] = ""  # Not applicable for call trades
                    entry["Put Cash Req"] = ""  # Not applicable for calls
            else:
                # Bought Stock or other
                entry["Investment"] = self.format_currency(shares * cost) if shares and cost and cost > 0 else ""
                entry["Put Value"] = ""  # Not applicable for stock trades
                entry["Put Cash Req"] = ""  # Not applicable for stock trades
                entry["Call Value"] = ""  # Not applicable for stock trades

            # Current price handling (use cost as placeholder for now)
            current_price = cost if cost and cost > 0 else 0
            entry["Current Price"] = self.format_currency(current_price) if current_price > 0 else ""

            close_price = self.parse_num(entry["Close Price"])
            if close_price and shares and cost:
                realized_pl = (close_price - cost) * shares
                entry["Realized P/L"] = f"${realized_pl:,.2f}" if realized_pl != 0 else "$0.00"  # Show zero for closed trades
                entry["P&L %"] = f"{((close_price-cost)/cost*100):.2f}%" if cost else ""
            else:
                entry["Realized P/L"] = ""
                entry["P&L %"] = ""

            if current_price and shares and cost:
                current_pl = (current_price - cost) * shares
                entry["Current P/L"] = f"${current_pl:,.2f}" if current_pl != 0 else "$0.00"  # Show zero for active trades
            else:
                entry["Current P/L"] = ""


            # Format all monetary columns for GUI - BUT PRESERVE CALCULATED VALUES
            for mcol in self.MONEY_COLS:
                if mcol in entry and entry[mcol] != "" and mcol != "Shares":
                    # Don't reformat if it's already been calculated as blank for sold options
                    if mcol == "Investment" and entry[mcol] == "":
                        continue  # Keep it blank for sold options
                    if mcol == "Cost":
                        # Always use the Cost from the GUI, and format for display
                        entry[mcol] = self.format_currency(float(str(cost_from_gui).replace("$","").replace(",","").strip())) if str(cost_from_gui).strip() != "" else ""
                        continue
                    try:
                        entry[mcol] = self.format_currency(float(str(entry[mcol]).replace("$","").replace(",","").strip()))
                    except Exception:
                        entry[mcol] = self.format_currency(entry[mcol])

            # Build trade row in COLUMNS order for GUI
            trade = [entry[col] for col in COLUMNS]
            # Ensure Investment value is up to date in the GUI row
            if "Investment" in entry:
                trade[COLUMNS.index("Investment")] = entry["Investment"]
            gui_row_index = self.tree.index(item)
            print(f"DEBUG: Updating GUI row {gui_row_index} for {entry.get('Ticker', 'UNKNOWN')} trade")
            self.tree.item(item, values=trade)
            self.trades[gui_row_index] = trade
            # Update Excel
            excel_filename = "Put Transactions.xlsx"
            sheet_name = "Open_Trades_2025"
            try:
                wb = load_workbook(excel_filename)
                ws = wb[sheet_name]
                header = [cell.value for cell in ws[1]]
                # Find row in Excel matching Ticker, Date, Shares
                match_row = None
                
                ticker_col = self.find_header_column(header, ["Stock ticker", "Ticker"])
                date_col = self.find_header_column(header, ["Date of transaction", "Trade Date", "Date"])
                shares_col = self.find_header_column(header, ["Number of shares +/-", "Shares", "Quantity"])
                strike_col = self.find_header_column(header, ["Strike price", "Strike"])
                cost_col = self.find_header_column(header, ["Cost"])
                notes_col = self.find_header_column(header, ["Notes"])
                
                print(f"DEBUG: Looking for ticker='{original_values['Ticker']}', date='{original_values['Date']}', shares='{original_values['Shares']}', strike='{original_values['Strike']}', cost='{original_values['Cost']}', notes='{original_values['Notes']}'")
                print(f"DEBUG: Excel columns - ticker_col={ticker_col}, date_col={date_col}, shares_col={shares_col}, strike_col={strike_col}, cost_col={cost_col}, notes_col={notes_col}")
                
                potential_matches = []  # Store all potential matches to handle duplicates
                
                for r in range(2, ws.max_row+1):
                    ticker = ws.cell(row=r, column=ticker_col).value
                    date = ws.cell(row=r, column=date_col).value
                    shares = ws.cell(row=r, column=shares_col).value
                    strike = ws.cell(row=r, column=strike_col).value if strike_col else None
                    cost = ws.cell(row=r, column=cost_col).value if cost_col else None
                    notes = ws.cell(row=r, column=notes_col).value if notes_col else None
                    
                    # Convert dates to comparable format
                    trade_date_str = str(trade[COLUMNS.index("Date")])
                    
                    # Handle different date formats in Excel
                    excel_date_str = str(date)
                    if date and hasattr(date, 'strftime'):
                        # Excel date is datetime object - convert to m/d/yyyy format
                        excel_date_str = date.strftime("%m/%d/%Y").lstrip('0').replace('/0', '/')
                    elif excel_date_str and excel_date_str != 'None':
                        # Excel date is string, try to normalize it
                        try:
                            from datetime import datetime
                            for fmt in ["%Y-%m-%d %H:%M:%S", "%Y-%m-%d", "%m/%d/%Y"]:
                                try:
                                    dt = datetime.strptime(excel_date_str.split()[0], fmt)
                                    excel_date_str = f"{dt.month}/{dt.day}/{dt.year}"
                                    break
                                except:
                                    continue
                        except:
                            pass
                    
                    # Normalize strike and cost values for comparison
                    def normalize_value(val):
                        if val is None:
                            return ""
                        try:
                            return str(float(str(val).replace("$","").replace(",","").strip()))
                        except:
                            return str(val).strip()
                    
                    excel_strike = normalize_value(strike)
                    excel_cost = normalize_value(cost)
                    excel_notes = str(notes).strip() if notes else ""
                    trade_strike = normalize_value(trade[COLUMNS.index("Strike")])
                    trade_cost = normalize_value(trade[COLUMNS.index("Cost")])
                    trade_notes = str(trade[COLUMNS.index("Notes")]).strip()
                    
                    print(f"DEBUG: Row {r}: ticker='{ticker}', date='{excel_date_str}', shares='{shares}', strike='{excel_strike}', cost='{excel_cost}', notes='{excel_notes}'")
                    print(f"DEBUG: Comparing with: ticker='{str(trade[COLUMNS.index('Ticker')])}', date='{trade_date_str}', shares='{str(trade[COLUMNS.index('Shares')])}', strike='{trade_strike}', cost='{trade_cost}', notes='{trade_notes}'")
                    
                    # Enhanced matching: ticker, date, shares must match
                    basic_match = (str(ticker) == str(trade[COLUMNS.index("Ticker")]) and 
                                   excel_date_str == trade_date_str and 
                                   str(shares) == str(trade[COLUMNS.index("Shares")]))
                    
                    if basic_match:
                        # For identical trades, try to distinguish by notes first, then by strike/cost
                        score = 0
                        if excel_notes == trade_notes and trade_notes != "":
                            score += 10  # Highest priority: exact notes match
                        if excel_strike == trade_strike:
                            score += 5
                        if excel_cost == trade_cost:
                            score += 3
                        
                        potential_matches.append((r, score, excel_notes, excel_strike, excel_cost))
                        print(f"DEBUG: Potential match at row {r} with score {score}")
                
                # Select the best match from potential matches
                if potential_matches:
                    # Sort by score (highest first), then by row number (lowest first for consistency)
                    potential_matches.sort(key=lambda x: (-x[1], x[0]))
                    best_match = potential_matches[0]
                    match_row = best_match[0]
                    print(f"DEBUG: Selected best match at row {match_row} with score {best_match[1]}")
                    
                    # If we have multiple matches with the same top score, and we're updating with new notes,
                    # prefer the first one that doesn't already have those notes
                    if len(potential_matches) > 1 and potential_matches[0][1] == potential_matches[1][1]:
                        trade_notes = str(trade[COLUMNS.index("Notes")]).strip()
                        if trade_notes:
                            for match in potential_matches:
                                if match[2] != trade_notes:  # This row doesn't have the new notes yet
                                    match_row = match[0]
                                    print(f"DEBUG: Preferring row {match_row} for new notes update")
                                    break
                if match_row:
                    try:
                        for col_idx, col_name in enumerate(header, start=1):
                            app_col = EXCEL_TO_APP_COLS.get(col_name, col_name)
                            value = entry[app_col] if app_col in entry else ""
                            
                            # Check for assignment/exercise status
                            assigned_exercised = entry.get("Assigned/Exercised", "").strip().upper()
                            
                            # Calculate and store as number for numeric columns with assignment logic
                            if app_col in ["Strike", "Shares", "Cost", "Investment", "Close Price", "Realized P/L", "Current Price", "Current P/L", "Call Value", "Put Value", "Put Cash Req"]:
                                if app_col == "Investment":
                                    # Always write as a number for Excel, but keep currency formatting
                                    # For assigned Sold Put, always use investment_numeric if set
                                    if investment_numeric is not None:
                                        value = investment_numeric
                                    else:
                                        val_str = str(entry[app_col])
                                        if val_str.strip() != "":
                                            # Remove $ and commas for numeric value
                                            try:
                                                value = float(val_str.replace("$", "").replace(",", "").strip())
                                            except Exception:
                                                value = None
                                        else:
                                            value = None
                                elif app_col == "Put Value":
                                    if assigned_exercised == "Y":
                                        value = None  # ALWAYS clear for assignments
                                    elif "put" in trade_type and cost and shares:
                                        value = cost * abs(shares)
                                    else:
                                        value = None
                                elif app_col == "Put Cash Req":
                                    if assigned_exercised == "Y":
                                        value = None  # ALWAYS clear for assignments
                                    elif "put" in trade_type and "sold" in trade_type and strike and shares:
                                        value = strike * abs(shares)
                                    else:
                                        value = None
                                elif app_col == "Call Value":
                                    if assigned_exercised == "Y":
                                        value = None  # ALWAYS clear for assignments
                                    elif "call" in trade_type and cost and shares:
                                        value = cost * abs(shares)
                                    else:
                                        value = None
                                elif app_col == "Realized P/L":
                                    value = (close_price - cost) * shares if close_price and cost and shares else None
                                elif app_col == "Current P/L":
                                    value = (current_price - cost) * shares if current_price and cost and shares else None
                                elif app_col == "Current Price":
                                    value = current_price if current_price else None
                                elif app_col == "Close Price":
                                    value = close_price if close_price else None
                                elif app_col == "Strike":
                                    # Use entry value instead of parsed variable
                                    entry_value = entry.get(app_col, "")
                                    value = self.parse_num(entry_value) if entry_value else None
                                elif app_col == "Cost":
                                    # Use entry value instead of parsed variable
                                    entry_value = entry.get(app_col, "")
                                    value = self.parse_num(entry_value) if entry_value else None
                                elif app_col == "Shares":
                                    # Use entry value instead of parsed variable
                                    entry_value = entry.get(app_col, "")
                                    value = self.parse_num(entry_value) if entry_value else None
                                else:
                                    try:
                                        value = float(str(value).replace("$","").replace(",","").strip()) if value != "" else None
                                    except Exception:
                                        value = None
                            else:
                                # Handle non-numeric columns
                                if app_col == "Assigned/Exercised":
                                    # Always ensure consistent formatting for assignment status
                                    value = assigned_exercised if assigned_exercised in ["Y", "N"] else entry.get(app_col, "")
                                else:
                                    value = entry[app_col] if app_col in entry else ""
                            
                            print(f"DEBUG: Writing to Excel row {match_row}, col {col_idx} ({app_col}): {value}")
                            cell = ws.cell(row=match_row, column=col_idx, value=value)
                            # Excel formatting: right align for dates, currency for money fields, center for assignments
                            if app_col in DATE_COLUMNS:
                                cell.alignment = Alignment(horizontal="right")
                            elif app_col == "Assigned/Exercised":
                                cell.alignment = Alignment(horizontal="center")
                                if assigned_exercised == "Y":
                                    cell.font = Font(name="Arial", size=12, bold=True)
                                else:
                                    cell.font = Font(name="Arial", size=12, bold=False)

                            elif app_col in ["Strike", "Cost", "Investment", "Close Price", "Realized P/L", "Current Price", "Current P/L", "Call Value", "Put Value", "Put Cash Req"]:
                                # For assigned puts, ensure Put Value and Put Cash Req are written as blank
                                if app_col in ["Put Value", "Put Cash Req"] and "put" in trade_type and "sold" in trade_type and assigned_exercised == "Y":
                                    cell.value = ""
                                if value is not None:
                                    cell.number_format = "$#,##0.00"
                    
                    except Exception as excel_update_error:
                        print(f"DEBUG: Error during Excel update: {excel_update_error}")
                        messagebox.showerror("Excel Update Error", f"Could not update Excel file. Make sure the file is not open in another application.\n\nError: {excel_update_error}")
                        return
                    
                    # Check if this trade should be closed (has both Close Date and Close Price)
                    close_date = entry.get("Close Date", "").strip()
                    close_price_val = entry.get("Close Price", "").strip()
                    
                    print(f"DEBUG: Checking close conditions - close_date='{close_date}', close_price_val='{close_price_val}'")
                    
                    if close_date and close_price_val:
                        print("DEBUG: Close conditions met, entering close trade logic")
                        try:
                            print("DEBUG: Trade is being closed!")
                            # Trade is being closed - calculate final P/L and clear open position fields
                            close_price = self.parse_num(close_price_val)
                            print(f"DEBUG: Parsed close price = {close_price}")
                        
                            # Parse shares and cost for calculations
                            shares_num = self.parse_num(entry["Shares"])
                            cost_num = self.parse_num(entry["Cost"])
                            print(f"DEBUG: Parsed shares = {shares_num}, cost = {cost_num}")
                        
                            # Calculate realized P/L and P&L %
                            if close_price and shares_num and cost_num:
                                realized_pl = (close_price - cost_num) * shares_num
                                entry["Realized P/L"] = f"${realized_pl:,.2f}"
                                
                                # P&L % calculation depends on trade type - store as integer rounded number
                                trade_type = entry["Type"].lower().strip()
                                if "sold" in trade_type:
                                    # For sold options: (premium_received - buy_back_cost) / premium_received * 100
                                    # Since shares are negative for sold options, and cost is the premium received
                                    pl_percent = ((cost_num - close_price)/cost_num*100) if cost_num else 0
                                    entry["P&L %"] = str(round(pl_percent)) if cost_num else ""
                                else:
                                    # For bought options/stocks: standard calculation
                                    pl_percent = ((close_price-cost_num)/cost_num*100) if cost_num else 0
                                    entry["P&L %"] = str(round(pl_percent)) if cost_num else ""
                            else:
                                entry["Realized P/L"] = ""
                                entry["P&L %"] = ""
                            
                            # Clear fields that are only relevant for open trades
                            entry["Investment"] = ""
                            entry["Call Value"] = ""
                            entry["Put Value"] = ""
                            entry["Put Cash Req"] = ""
                            entry["Current Price"] = ""
                            entry["Current P/L"] = ""
                            print(f"DEBUG: Cleared open trade fields for closed trade")
                            
                            # Update Excel with final calculations for closed trade
                            for col_idx, col_name in enumerate(header, start=1):
                                app_col = EXCEL_TO_APP_COLS.get(col_name, col_name)
                                value = entry.get(app_col, "")  # Use get() to handle all columns
                                
                                # Store calculated values as numbers in Excel
                                if app_col == "Realized P/L" and value:
                                    try:
                                        value = float(str(value).replace("$","").replace(",","").strip())
                                    except:
                                        value = None
                                elif app_col == "P&L %" and value:
                                    try:
                                        # Store P&L % as decimal number for Excel (e.g., 15 becomes 0.15)
                                        percent_num = float(str(value).replace("%","").strip())
                                        value = percent_num / 100.0  # Convert to decimal for percentage format
                                    except:
                                        value = None
                                elif app_col in ["Investment", "Call Value", "Put Value", "Put Cash Req", "Current Price", "Current P/L"]:
                                    print(f"DEBUG: Explicitly clearing field {app_col} (column {col_idx}) for closed trade - was: {value}")
                                    value = None  # Force clear these fields for closed trades
                                elif app_col == "Close Price":
                                    value = close_price if close_price else None
                                elif app_col in ["Strike", "Shares", "Cost"]:
                                    # Keep these as numbers for proper Excel formatting
                                    try:
                                        value = float(str(value).replace("$","").replace(",","").strip()) if value else None
                                    except:
                                        value = None
                                
                                cell = ws.cell(row=match_row, column=col_idx, value=value)
                                # Improved Excel formatting
                                if app_col in DATE_COLUMNS:
                                    cell.alignment = Alignment(horizontal="right")
                                    cell.font = Font(name="Arial", size=12)
                                elif app_col in ["Strike", "Cost", "Close Price", "Realized P/L", "Shares"]:
                                    if app_col == "Shares":
                                        cell.alignment = Alignment(horizontal="right")
                                        cell.font = Font(name="Arial", size=12)
                                    elif value is not None:
                                        cell.number_format = "$#,##0.00"
                                        cell.alignment = Alignment(horizontal="right")
                                        cell.font = Font(name="Arial", size=12)
                                elif app_col == "P&L %":
                                    if value is not None:
                                        cell.number_format = "0%"  # Integer percentage format for P&L %
                                    cell.alignment = Alignment(horizontal="right")
                                    cell.font = Font(name="Arial", size=12)
                                else:
                                    cell.font = Font(name="Arial", size=12)
                            
                            wb.save(excel_filename)
                            wb.close()
                            
                            # Update the treeview with final values - ensure cleared fields show empty
                            trade = []
                            for col in COLUMNS:
                                if col in ["Investment", "Call Value", "Put Value", "Put Cash Req", "Current Price", "Current P/L"]:
                                    trade.append("")  # Show these as empty for closed trades
                                else:
                                    trade.append(entry[col])
                            self.tree.item(item, values=trade)
                            
                            # Move to closed trades sheet and remove from GUI - pass the cleaned entry data
                            self.move_trade_to_closed(entry, match_row, excel_filename)
                            # Remove from GUI
                            self.tree.delete(item)
                            # Clear entry fields
                            for var in self.vars.values():
                                var.set("")
                            messagebox.showinfo("Trade Closed", f"Trade for {entry['Ticker']} has been closed and moved to Closed Trades.\nRealized P/L: {entry.get('Realized P/L', 'N/A')}")
                            return
                            
                        except Exception as close_error:
                            print(f"DEBUG: Error in close trade logic: {close_error}")
                            messagebox.showerror("Close Trade Error", f"Error closing trade: {close_error}")
                            # Continue with regular update if close fails
                            wb.save(excel_filename)
                            wb.close()
                            messagebox.showinfo("Trade Updated", "Trade updated in Excel and GUI (close failed).")
                    else:
                        print("DEBUG: No close conditions, proceeding with regular Excel update")
                        wb.save(excel_filename)
                        wb.close()
                        print("DEBUG: Excel file saved successfully")
                        messagebox.showinfo("Trade Updated", "Trade updated in Excel and GUI.")
                else:
                    print("DEBUG: No matching row found in Excel!")
                    messagebox.showwarning("Excel Update", "Could not find matching trade in Excel.")
            except Exception as e:
                messagebox.showerror("Excel Error", f"Could not update trade in Excel.\n{e}")
            
            # Clear entry fields after successful update
            for var in self.vars.values():
                var.set("")
        
        # Execute the save_update function directly (no popup needed)
        save_update()

    def move_trade_to_closed(self, entry, match_row, excel_filename):
        """Move a trade from Open_Trades_2025 to Closed_Trades_2025 sheet"""
        try:
            print(f"DEBUG: Starting move_trade_to_closed for {entry['Ticker']}")
            # Wait a moment and reopen to ensure we get the latest saved data
            import time
            time.sleep(0.1)  # Brief pause to ensure file save is complete
            
            wb = load_workbook(excel_filename)
            open_sheet = wb["Open_Trades_2025"]
            print(f"DEBUG: Loaded workbook, open sheet has {open_sheet.max_row} rows")
            
            # Create Closed_Trades_2025 sheet if it doesn't exist (matching the naming pattern)
            closed_sheet_name = "Closed_Trades_2025"
            if closed_sheet_name not in wb.sheetnames:
                print(f"DEBUG: Creating new {closed_sheet_name} sheet")
                closed_sheet = wb.create_sheet(closed_sheet_name)
                # Copy headers from open trades sheet
                open_headers = [cell.value for cell in open_sheet[1]]
                for col_idx, header in enumerate(open_headers, start=1):
                    closed_sheet.cell(row=1, column=col_idx, value=header)
                    # Format headers
                    closed_sheet.cell(row=1, column=col_idx).font = Font(bold=True)
                print(f"DEBUG: Created closed sheet with {len(open_headers)} headers")
            else:
                print(f"DEBUG: Using existing {closed_sheet_name} sheet")
                closed_sheet = wb[closed_sheet_name]
            
            print(f"DEBUG: Closed sheet currently has {closed_sheet.max_row} rows")
            
            # Instead of copying from Excel row (which may have old values), 
            # build the row data from the cleaned entry data
            headers = [cell.value for cell in open_sheet[1]]
            cleaned_row_data = []
            
            for col_idx, header_name in enumerate(headers):
                app_col = EXCEL_TO_APP_COLS.get(header_name, header_name)
                
                # Use the cleaned entry data with explicit clearing for closed trades
                if app_col in ["Investment", "Call Value", "Put Value", "Put Cash Req", "Current Price", "Current P/L"]:
                    value = None  # Force clear these fields for closed trades
                    print(f"DEBUG: Force clearing {app_col} (column {col_idx+1}) - setting to None")
                elif app_col in entry:
                    value = entry[app_col]
                    # Convert formatted values back to raw numbers for Excel storage
                    if app_col == "Realized P/L" and value:
                        try:
                            value = float(str(value).replace("$","").replace(",","").strip())
                        except:
                            value = None
                    elif app_col == "P&L %" and value:
                        try:
                            # Convert percentage string to decimal for Excel
                            percent_num = float(str(value))
                            value = percent_num / 100.0
                        except:
                            value = None
                    elif app_col == "Close Price" and value:
                        try:
                            value = float(str(value).replace("$","").replace(",","").strip()) if value else None
                        except:
                            value = None
                    elif app_col in ["Strike", "Shares", "Cost"] and value:
                        try:
                            value = float(str(value).replace("$","").replace(",","").strip()) if value else None
                        except:
                            value = None
                else:
                    # Get from the current Excel row if not in entry
                    try:
                        value = open_sheet.cell(row=match_row, column=col_idx+1).value
                    except:
                        value = None
                
                cleaned_row_data.append(value)
            
            next_closed_row = closed_sheet.max_row + 1
            print(f"DEBUG: Adding cleaned row to closed row {next_closed_row}")
            print(f"DEBUG: Cleaned row data has {len(cleaned_row_data)} columns")
            
            for col_idx, value in enumerate(cleaned_row_data, start=1):
                cell = closed_sheet.cell(row=next_closed_row, column=col_idx, value=value)
                print(f"DEBUG: Set column {col_idx}: {value}")
                
                # Get the header to determine column type
                header_cell = closed_sheet.cell(row=1, column=col_idx)
                header_name = header_cell.value if header_cell.value else ""
                
                # Apply proper formatting based on column type
                cell.font = Font(name="Arial", size=12)
                
                if "Date" in header_name or "Expiration" in header_name:
                    # Date columns: right aligned
                    cell.alignment = Alignment(horizontal="right")
                elif any(term in header_name for term in ["Strike", "Cost", "investment", "Price", "profit", "P/L", "Value", "requirement"]):
                    # Monetary columns: right aligned with currency format if numeric
                    cell.alignment = Alignment(horizontal="right")
                    if isinstance(value, (int, float)) and value is not None:
                        if "%" not in header_name:
                            cell.number_format = "$#,##0.00"
                elif "shares" in header_name.lower():
                    # Shares column: right aligned, no currency format
                    cell.alignment = Alignment(horizontal="right")
                
                # Copy number format from source if applicable
                source_cell = open_sheet.cell(row=match_row, column=col_idx)
                try:
                    if source_cell.number_format and source_cell.number_format != "General":
                        cell.number_format = source_cell.number_format
                except Exception:
                    pass  # Skip formatting if there's any issue
            
            # Delete the row from open trades sheet
            print(f"DEBUG: Deleting row {match_row} from open trades")
            open_sheet.delete_rows(match_row)
            
            print(f"DEBUG: Saving workbook to {excel_filename}")
            wb.save(excel_filename)
            wb.close()
            print(f"DEBUG: Workbook saved and closed successfully")
            
            print(f"Successfully moved closed trade {entry['Ticker']} to {closed_sheet_name} sheet")
            
        except Exception as e:
            print(f"Error moving trade to closed: {e}")
            messagebox.showerror("Close Trade Error", f"Could not move trade to closed trades.\n{e}")
            
            # Reopen the file to reload fresh data if there was an error
            try:
                wb = load_workbook(excel_filename)
                wb.close()
                self.load_open_trades_from_excel()
            except:
                pass

    def get_stock_price_from_etrade(self, ticker):
        """Fetch current stock price from E*TRADE"""
        try:
            from etrade_auth import get_etrade_session
            session, base_url = get_etrade_session()
            url = f"{base_url}/v1/market/quote/{ticker}.json"
            response = session.get(url)
            
            if response.status_code == 200:
                data = response.json()
                quote_data = data.get("QuoteResponse", {}).get("QuoteData", [])
                if quote_data and len(quote_data) > 0:
                    all_data = quote_data[0].get("All", {})
                    last_trade = all_data.get("lastTrade")
                    if last_trade:
                        return float(last_trade)
        except Exception as e:
            print(f"Error fetching stock price for {ticker}: {e}")
        return None

    def get_option_price_from_etrade(self, ticker, strike, option_type, expiry_date):
        """Fetch current option price from E*TRADE"""
        try:
            from etrade_auth import get_etrade_session
            from datetime import datetime
            
            session, base_url = get_etrade_session()
            
            # Parse expiry date if it's a string
            target_date = None
            if isinstance(expiry_date, str):
                try:
                    target_date = datetime.strptime(expiry_date, "%m/%d/%Y").date()
                except:
                    try:
                        target_date = datetime.strptime(expiry_date, "%Y-%m-%d").date()
                    except:
                        print(f"Could not parse expiry date: {expiry_date}")
                        return None
            elif hasattr(expiry_date, 'date'):
                target_date = expiry_date.date()
            else:
                target_date = expiry_date
                
            if not target_date:
                print(f"Invalid expiry date format: {expiry_date}")
                return None
            
            # Get available expiration dates to find the closest match
            exp_url = f"{base_url}/v1/market/optionexpiredate.json?symbol={ticker}"
            exp_response = session.get(exp_url)
            
            if exp_response.status_code != 200:
                print(f"Could not get expiration dates for {ticker}")
                return None
                
            exp_data = exp_response.json()
            dates = exp_data.get("OptionExpireDateResponse", {}).get("ExpirationDate", [])
            
            # Find the best matching expiration date
            best_match = None
            min_diff = float('inf')
            
            for date_info in dates:
                exp_datetime = datetime(date_info['year'], date_info['month'], date_info['day']).date()
                diff = abs((exp_datetime - target_date).days)
                if diff < min_diff:
                    min_diff = diff
                    best_match = date_info
            
            if not best_match:
                print(f"No expiration dates found for {ticker}")
                return None
                
            # Format URL for option chain with the best matching date
            url = f"{base_url}/v1/market/optionchains.json?symbol={ticker}&expiryDay={best_match['day']:02d}&expiryMonth={best_match['month']:02d}&expiryYear={best_match['year']}&chainType={option_type.upper()}"
            response = session.get(url)
            
            if response.status_code == 200:
                data = response.json()
                chain = data.get("OptionChainResponse", {})
                option_pairs = chain.get("OptionPair", [])
                
                # Find option with closest strike price
                best_option = None
                min_strike_diff = float('inf')
                
                for pair in option_pairs:
                    if option_type.lower() == "call":
                        option_data = pair.get("Call", {})
                    else:
                        option_data = pair.get("Put", {})
                    
                    if option_data:
                        option_strike = float(option_data.get("strikePrice", 0))
                        strike_diff = abs(option_strike - float(strike))
                        if strike_diff < min_strike_diff:
                            min_strike_diff = strike_diff
                            best_option = option_data
                
                if best_option:
                    last_price = best_option.get("lastPrice")
                    if last_price:
                        print(f"Found {option_type} option for {ticker}: Strike ${best_option.get('strikePrice')}, Price ${last_price}")
                        return float(last_price)
                        
        except Exception as e:
            print(f"Error fetching option price for {ticker} {strike} {option_type}: {e}")
        return None

    def update_stock_prices(self):
        updated_count = 0
        total_trades = len(self.trades)
        excel_filename = "Put Transactions.xlsx"
        sheet_name = "Open_Trades_2025"
        
        if total_trades == 0:
            messagebox.showinfo("No Trades", "No trades to update prices for.")
            return
        
        # Ask user to confirm
        if not messagebox.askyesno("Update Prices", f"Fetch current prices from E*TRADE for {total_trades} trades?\n\nThis may take a few moments."):
            return
        
        try:
            wb = load_workbook(excel_filename)
            ws = wb[sheet_name]
            header = [cell.value for cell in ws[1]]
            
            print(f"Updating prices for {total_trades} trades...")
            
            for idx, trade in enumerate(self.trades):
                ticker = trade[COLUMNS.index("Ticker")]
                trade_type = trade[COLUMNS.index("Type")].lower().strip()
                strike_str = trade[COLUMNS.index("Strike")]
                expiry_str = trade[COLUMNS.index("Expiration")]
                
                print(f"Processing {idx+1}/{total_trades}: {ticker} ({trade_type})")
                print(f"  Strike: {strike_str}, Expiry: {expiry_str}")
                current_price = None
                
                # Determine if this is a stock or option trade
                if "put" in trade_type or "call" in trade_type:
                    # This is an option trade
                    try:
                        strike = float(str(strike_str).replace("$","").replace(",","").strip()) if strike_str else None
                        if strike and expiry_str:
                            option_type = "put" if "put" in trade_type else "call"
                            print(f"  Fetching {option_type} option price for {ticker}, strike ${strike}, expiry {expiry_str}")
                            current_price = self.get_option_price_from_etrade(ticker, strike, option_type, expiry_str)
                        else:
                            print(f"  Missing strike ({strike_str}) or expiry ({expiry_str}) for option {ticker}")
                    except Exception as e:
                        print(f"  Error processing option {ticker}: {e}")
                else:
                    # This is a stock trade
                    print(f"  Fetching stock price for {ticker}")
                    current_price = self.get_stock_price_from_etrade(ticker)
                
                # Update the price if we got one
                if current_price:
                    formatted_price = f"${current_price:,.2f}"
                    trade[COLUMNS.index("Current Price")] = formatted_price
                    
                    # Calculate Current P/L based on new price
                    try:
                        cost_str = trade[COLUMNS.index("Cost")]
                        shares_str = trade[COLUMNS.index("Shares")]
                        cost = float(str(cost_str).replace("$","").replace(",","").strip()) if cost_str else 0.0
                        shares = float(str(shares_str).replace("$","").replace(",","").strip()) if shares_str else 0.0
                        
                        if cost and shares:
                            current_pl = (current_price - cost) * shares
                            formatted_pl = f"${current_pl:,.2f}"
                            trade[COLUMNS.index("Current P/L")] = formatted_pl
                            print(f"  Calculated P/L: {formatted_pl} (Price: {current_price}, Cost: {cost}, Shares: {shares})")
                        else:
                            print(f"  Could not calculate P/L - Cost: {cost}, Shares: {shares}")
                    except Exception as e:
                        print(f"  Error calculating P/L for {ticker}: {e}")
                    
                    # Update Excel row
                    match_row = None
                    for r in range(2, ws.max_row+1):
                        ticker_col = self.find_header_column(header, ["Stock ticker", "Ticker"])
                        date_col = self.find_header_column(header, ["Date of transaction", "Trade Date", "Date"])
                        shares_col = self.find_header_column(header, ["Number of shares +/-", "Shares", "Quantity"])
                        
                        excel_ticker = ws.cell(row=r, column=ticker_col).value
                        excel_date = ws.cell(row=r, column=date_col).value
                        excel_shares = ws.cell(row=r, column=shares_col).value
                        if str(excel_ticker) == str(trade[COLUMNS.index("Ticker")]) and str(excel_date) == str(trade[COLUMNS.index("Date")]) and str(excel_shares) == str(trade[COLUMNS.index("Shares")]):
                            match_row = r
                            break
                    
                    if match_row:
                        # Update Current Price
                        current_price_col = self.find_header_column(header, ["Current Price"])
                        ws.cell(row=match_row, column=current_price_col, value=current_price)
                        ws.cell(row=match_row, column=current_price_col).number_format = '$#,##0.00'
                        
                        # Update Current P/L if we calculated it
                        if cost and shares:
                            current_pl = (current_price - cost) * shares
                            current_pl_col = self.find_header_column(header, ["Current P/L"])
                            ws.cell(row=match_row, column=current_pl_col, value=current_pl)
                            ws.cell(row=match_row, column=current_pl_col).number_format = '$#,##0.00'
                    
                    updated_count += 1
                    print(f"  ✓ Updated {ticker}: {formatted_price}")
                else:
                    print(f"  ✗ Could not get price for {ticker}")
            
            wb.save(excel_filename)
            wb.close()
            
            # Refresh the GUI display
            self.tree.delete(*self.tree.get_children())
            for trade in self.trades:
                self.tree.insert("", "end", values=trade)
            
            # Fix any existing trades that have Current Price but missing Current P/L
            self.fix_missing_current_pl()
            
            messagebox.showinfo("Price Update Complete", f"Successfully updated prices for {updated_count} out of {total_trades} trades.")
            
        except Exception as e:
            print(f"Excel Error: {e}")
            messagebox.showerror("Excel Error", f"Could not update stock prices in Excel.\n{e}")

    def fix_missing_current_pl(self):
        """Fix trades that have Current Price but missing Current P/L"""
        excel_filename = "Put Transactions.xlsx"
        sheet_name = "Open_Trades_2025"
        
        print(f"Starting P/L fix process...")
        
        try:
            wb = load_workbook(excel_filename)
            ws = wb[sheet_name]
            header = [cell.value for cell in ws[1]]
            
            current_price_col = self.find_header_column(header, ["Current Price"])
            current_pl_col = self.find_header_column(header, ["Current P/L"])
            cost_col = self.find_header_column(header, ["Cost"])
            shares_col = self.find_header_column(header, ["Number of shares +/-", "Shares", "Quantity"])
            ticker_col = self.find_header_column(header, ["Stock ticker", "Ticker"])
            
            print(f"Total rows to check: {ws.max_row - 1}")
            fixed_count = 0
            
            for r in range(2, ws.max_row + 1):
                current_price = ws.cell(row=r, column=current_price_col).value
                current_pl = ws.cell(row=r, column=current_pl_col).value
                cost = ws.cell(row=r, column=cost_col).value
                shares = ws.cell(row=r, column=shares_col).value
                ticker = ws.cell(row=r, column=ticker_col).value
                
                print(f"Row {r}: {ticker} - Price: {current_price}, P/L: {current_pl}, Cost: {cost}, Shares: {shares}")
                
                # If we have a current price and cost and shares, but missing or zero P/L
                if current_price and cost and shares and (not current_pl or current_pl == 0):
                    try:
                        # Parse the current price (remove $ and commas if it's a string)
                        if isinstance(current_price, str):
                            price_val = float(current_price.replace("$", "").replace(",", ""))
                        else:
                            price_val = float(current_price)
                        
                        cost_val = float(cost)
                        shares_val = float(shares)
                        
                        # Calculate P/L
                        pl_val = (price_val - cost_val) * shares_val
                        
                        # Update the cell
                        ws.cell(row=r, column=current_pl_col, value=pl_val)
                        ws.cell(row=r, column=current_pl_col).number_format = '$#,##0.00'
                        
                        print(f"  ✓ Fixed P/L for {ticker}: ${pl_val:,.2f} = ({price_val} - {cost_val}) × {shares_val}")
                        fixed_count += 1
                        
                    except Exception as e:
                        print(f"  ✗ Error fixing P/L for {ticker}: {e}")
                else:
                    if not current_price:
                        print(f"  Skipping {ticker}: No current price")
                    elif not cost:
                        print(f"  Skipping {ticker}: No cost")
                    elif not shares:
                        print(f"  Skipping {ticker}: No shares")
                    elif current_pl and current_pl != 0:
                        print(f"  Skipping {ticker}: P/L already calculated ({current_pl})")
            
            if fixed_count > 0:
                wb.save(excel_filename)
                print(f"✓ Fixed Current P/L for {fixed_count} trades")
                
                # Reload the data to refresh the GUI
                self.load_open_trades_from_excel()
            else:
                print("No trades needed P/L fixes")
            
            wb.close()
            
            # Show message to user
            if fixed_count > 0:
                messagebox.showinfo("P/L Fix Complete", f"Fixed Current P/L calculations for {fixed_count} trades.")
            else:
                messagebox.showinfo("P/L Fix Complete", "No trades needed P/L fixes. All Current P/L values are already calculated.")
            
        except Exception as e:
            print(f"Error fixing Current P/L: {e}")
            messagebox.showerror("P/L Fix Error", f"Could not fix P/L calculations.\n{e}")
    def sort_excel_trades(self, sheet_name="Open_Trades_2025"):
        """Sort Excel trades by ticker and date without data loss"""
        excel_filename = "Put Transactions.xlsx"
        try:
            from datetime import datetime
            wb = load_workbook(excel_filename)
            ws = wb[sheet_name]
            
            # Get header row
            header = [cell.value for cell in ws[1]]
            ticker_col = self.find_header_column(header, ["Stock ticker", "Ticker"])
            date_col = self.find_header_column(header, ["Date of transaction", "Trade Date", "Date"])
            
            # Get all data rows with their row numbers
            data_with_rows = []
            for row_num in range(2, ws.max_row + 1):
                ticker = ws.cell(row=row_num, column=ticker_col).value or ""
                date_val = ws.cell(row=row_num, column=date_col).value
                
                # Parse date for sorting
                sort_date = datetime.min
                try:
                    if isinstance(date_val, datetime):
                        sort_date = date_val
                    elif date_val:
                        date_str = str(date_val).split()[0]
                        for fmt in ("%m/%d/%Y", "%Y-%m-%d"):
                            try:
                                sort_date = datetime.strptime(date_str, fmt)
                                break
                            except:
                                continue
                except:
                    pass
                
                data_with_rows.append((str(ticker), sort_date, row_num))
            
            # Sort by ticker and date
            data_with_rows.sort(key=lambda x: (x[0], x[1]))
            
            # Get the sorted row numbers
            sorted_row_nums = [row_data[2] for row_data in data_with_rows]
            
            # Only proceed if we actually need to sort
            if sorted_row_nums != list(range(2, ws.max_row + 1)):
                # Read all row data first
                all_rows_data = []
                for row_num in sorted_row_nums:
                    row_data = []
                    for col in range(1, len(header) + 1):
                        cell = ws.cell(row=row_num, column=col)
                        row_data.append({
                            'value': cell.value,
                            'number_format': cell.number_format
                        })
                    all_rows_data.append(row_data)
                
                # Clear data rows and rewrite in sorted order
                for row_num in range(ws.max_row, 1, -1):
                    ws.delete_rows(row_num)
                
                # Write sorted data back
                for new_row, row_data in enumerate(all_rows_data, start=2):
                    for col, cell_data in enumerate(row_data, start=1):
                        cell = ws.cell(row=new_row, column=col)
                        cell.value = cell_data['value']
                        cell.number_format = cell_data['number_format']
                        
                        # Apply standard formatting
                        app_col = EXCEL_TO_APP_COLS.get(header[col-1], header[col-1])
                        if app_col in DATE_COLUMNS:
                            cell.alignment = Alignment(horizontal="right")
            
            wb.save(excel_filename)
            wb.close()
            print("Excel sorting completed successfully")
            
        except Exception as e:
            print(f"Excel sort error: {e}")

    def sort_treeview_by_ticker_and_date(self):
        # Sort the treeview rows by ticker and date
        try:
            from datetime import datetime
            def normalize_date(date_val):
                if not date_val or str(date_val).lower() in ("nat", "none", "nan"):
                    return ""
                if isinstance(date_val, datetime):
                    # Format without leading zeros: m/d/yyyy
                    month = date_val.month
                    day = date_val.day  
                    year = date_val.year
                    return f"{month}/{day}/{year}"
                try:
                    # Remove time component first
                    date_str = str(date_val).split()[0]
                    dt = datetime.strptime(date_str, "%Y-%m-%d")
                    month = dt.month
                    day = dt.day  
                    year = dt.year
                    return f"{month}/{day}/{year}"
                except Exception:
                    try:
                        date_str = str(date_val).split()[0]
                        dt = datetime.strptime(date_str, "%m/%d/%Y")
                        month = dt.month
                        day = dt.day  
                        year = dt.year
                        return f"{month}/{day}/{year}"
                    except Exception:
                        return str(date_val)
            items = [(self.tree.set(i, "Ticker"), normalize_date(self.tree.set(i, "Date")), i) for i in self.tree.get_children()]
            def date_key(date_str):
                try:
                    return datetime.strptime(date_str, "%m/%d/%Y")
                except Exception:
                    return datetime.min
            items.sort(key=lambda x: (x[0], date_key(x[1]) if x[1] else datetime.min))
            for idx, (_, _, item_id) in enumerate(items):
                self.tree.move(item_id, '', idx)
        except Exception as e:
            print(f"Sort Treeview Error: {e}")
            messagebox.showerror("Sort Error", f"Could not sort trades by ticker/date.\n{e}")

    def sort_open_trades_gui_and_excel(self):
        # Placeholder for future sorting logic if needed
        # All sorting is now handled in load_open_trades_from_excel
        pass

    def load_open_trades_from_excel(self):
        excel_filename = "Put Transactions.xlsx"
        if not os.path.exists(excel_filename):
            return  # File not found, skip loading

        # Read from the correct sheet and header row
        try:
            df = pd.read_excel(excel_filename, header=0, sheet_name="Open_Trades_2025")
        except ValueError as e:
            print(f"Excel loading error: {e}")
            messagebox.showerror("Excel Error", f"Could not find sheet 'Open_Trades_2025'.\n{e}")
            return
        # Replace all NaN with empty string for clean display
        df = df.fillna("")
        print("Excel columns:", df.columns.tolist())  # Debug print
        print("First 5 rows of Excel data:")
        print(df.head())
        if df.empty:
            print("WARNING: DataFrame is empty. No trades to display.")
        # Clear current trades in GUI and memory
        self.tree.delete(*self.tree.get_children())
        self.trades.clear()

        print("load_open_trades_from_excel called")
        print(f"DataFrame has {len(df)} rows to process")
        
        # Build each trade row in COLUMNS order, mapping from Excel columns
        def format_date(val):
            if val:
                try:
                    from datetime import datetime
                    dt = None
                    if isinstance(val, datetime):
                        dt = val
                    else:
                        val_str = str(val).split()[0]  # Remove time part first
                        for fmt in ("%m/%d/%Y", "%Y-%m-%d", "%m-%d-%Y", "%Y/%m/%d"):
                            try:
                                dt = datetime.strptime(val_str, fmt)
                                break
                            except Exception:
                                continue
                    if dt:
                        # Format without leading zeros: m/d/yyyy
                        month = dt.month
                        day = dt.day  
                        year = dt.year
                        return f"{month}/{day}/{year}"
                except Exception:
                    pass
            return val

        def format_currency(val):
            try:
                num = float(str(val).replace("$","").replace(",","").strip())
                return "$" + format(num, ",.2f")
            except Exception:
                return ""

        MONEY_COLS = ["Cost", "Investment", "Close Price", "Realized P/L", "Current Price", "Current P/L", "Call Value", "Put Value", "Put Cash Req", "Strike"]

        processed_count = 0
        for row_idx, row in df.iterrows():
            print(f"Processing row {row_idx + 1}/{len(df)}: {row.get('Stock ticker', 'Unknown')}")
            trade = []
            type_val = self.clean_blank(row.get("Stock Purchase Type", ""))
            
            for col in COLUMNS:
                excel_col = next((k for k, v in EXCEL_TO_APP_COLS.items() if v == col), None)
                val = row.get(excel_col, "") if excel_col else ""
                val = self.clean_blank(val)
                # If Bought Stock, Expiration is always blank
                if col == "Expiration" and type_val.strip().lower() == "bought stock":
                    val = ""
                elif col in DATE_COLUMNS:
                    val = format_date(val)
                elif col in MONEY_COLS and val != "" and col != "Shares":
                    val = format_currency(val)
                trade.append(val)
                
            if any(str(v).strip() != "" for v in trade):
                self.trades.append(trade)
                self.tree.insert("", "end", values=trade)
                processed_count += 1
                
        print(f"Processed {processed_count} trades successfully")
        print("Starting GUI sort...")

        # Sort the GUI after loading all trades
        try:
            self.sort_treeview_by_ticker_and_date()
            print("GUI sort completed")
        except Exception as e:
            print(f"Error during GUI sort: {e}")
            import traceback
            traceback.print_exc()

    def generate_results(self):
        """Generate comprehensive results analysis in Excel with real-time data"""
        try:
            from datetime import datetime
            import pandas as pd
            from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
            
            excel_filename = "Put Transactions.xlsx"
            
            # First update current prices and P&L for open trades
            print("Updating current prices for open trades...")
            self.update_current_prices_for_all_trades()
            
            wb = load_workbook(excel_filename)
            
            # Create or clear Results sheet
            if "Results" in wb.sheetnames:
                wb.remove(wb["Results"])
            results_sheet = wb.create_sheet("Results")
            
            print("Generating comprehensive trading results...")
            
            # --- SECTION 1: TITLE ---
            row = 1
            
            # Title
            title_cell = results_sheet.cell(row=row, column=1, value="🏆 TRADING RESULTS DASHBOARD")
            title_cell.font = Font(name="Arial", size=20, bold=True, color="FFFFFF")
            title_cell.fill = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
            title_cell.alignment = Alignment(horizontal="center")
            results_sheet.merge_cells(f"A{row}:G{row}")
            row += 1
            
            # Timestamp
            timestamp_cell = results_sheet.cell(row=row, column=1, value=f"Generated: {datetime.now().strftime('%m/%d/%Y %I:%M %p')}")
            timestamp_cell.font = Font(name="Arial", size=10, italic=True, color="666666")
            timestamp_cell.alignment = Alignment(horizontal="center")
            results_sheet.merge_cells(f"A{row}:G{row}")
            row += 3
            
            # --- SECTION 2: CLOSED TRADES ANALYSIS ---
            
            # Get closed trades data
            closed_data = self.get_closed_trades_data()
            
            # Closed Trades Summary Header
            header_cell = results_sheet.cell(row=row, column=1, value="📈 CLOSED TRADES PERFORMANCE")
            header_cell.font = Font(name="Arial", size=16, bold=True, color="FFFFFF")
            header_cell.fill = PatternFill(start_color="4CAF50", end_color="4CAF50", fill_type="solid")
            header_cell.alignment = Alignment(horizontal="center")
            results_sheet.merge_cells(f"A{row}:E{row}")
            row += 2
            
            # Yearly P&L Analysis
            yearly_stats = self.calculate_yearly_stats(closed_data)
            
            # Headers for yearly table
            headers = ["Year", "Total P&L", "Average P&L %", "Win Rate", "Number of Trades"]
            for col, header in enumerate(headers, 1):
                cell = results_sheet.cell(row=row, column=col, value=header)
                cell.font = Font(name="Arial", size=12, bold=True, color="FFFFFF")
                cell.fill = PatternFill(start_color="2196F3", end_color="2196F3", fill_type="solid")
                cell.alignment = Alignment(horizontal="center")
            row += 1
            
            # Yearly data
            total_pl = 0
            total_trades = 0
            total_avg_percent = 0
            
            for year, stats in yearly_stats.items():
                results_sheet.cell(row=row, column=1, value=year).font = Font(name="Arial", size=11, bold=True)
                
                # P&L with color coding
                pl_cell = results_sheet.cell(row=row, column=2, value=stats['total_pl'])
                pl_cell.number_format = "$#,##0.00"
                pl_cell.font = Font(name="Arial", size=11, bold=True)
                if stats['total_pl'] > 0:
                    pl_cell.font = Font(name="Arial", size=11, bold=True, color="006400")  # Green
                    pl_cell.fill = PatternFill(start_color="E8F5E8", end_color="E8F5E8", fill_type="solid")
                elif stats['total_pl'] < 0:
                    pl_cell.font = Font(name="Arial", size=11, bold=True, color="CC0000")  # Red
                    pl_cell.fill = PatternFill(start_color="FFF2F2", end_color="FFF2F2", fill_type="solid")
                
                # Average P&L %
                avg_cell = results_sheet.cell(row=row, column=3, value=stats['avg_percent'])
                avg_cell.number_format = "0.0%"
                avg_cell.font = Font(name="Arial", size=11)
                
                # Win Rate
                win_rate_cell = results_sheet.cell(row=row, column=4, value=stats['win_rate'])
                win_rate_cell.number_format = "0.0%"
                win_rate_cell.font = Font(name="Arial", size=11)
                if stats['win_rate'] >= 0.6:
                    win_rate_cell.fill = PatternFill(start_color="E8F5E8", end_color="E8F5E8", fill_type="solid")
                elif stats['win_rate'] < 0.4:
                    win_rate_cell.fill = PatternFill(start_color="FFF2F2", end_color="FFF2F2", fill_type="solid")
                
                # Number of trades
                results_sheet.cell(row=row, column=5, value=stats['trade_count']).font = Font(name="Arial", size=11)
                
                total_pl += stats['total_pl']
                total_trades += stats['trade_count']
                total_avg_percent += stats['avg_percent']
                row += 1
            
            # Totals row
            results_sheet.cell(row=row, column=1, value="TOTAL").font = Font(name="Arial", size=12, bold=True)
            total_pl_cell = results_sheet.cell(row=row, column=2, value=total_pl)
            total_pl_cell.number_format = "$#,##0.00"
            total_pl_cell.font = Font(name="Arial", size=12, bold=True, color="FFFFFF")
            if total_pl > 0:
                total_pl_cell.fill = PatternFill(start_color="4CAF50", end_color="4CAF50", fill_type="solid")
            else:
                total_pl_cell.fill = PatternFill(start_color="F44336", end_color="F44336", fill_type="solid")
            
            overall_avg = total_avg_percent / len(yearly_stats) if yearly_stats else 0
            avg_total_cell = results_sheet.cell(row=row, column=3, value=overall_avg)
            avg_total_cell.number_format = "0.0%"
            avg_total_cell.font = Font(name="Arial", size=12, bold=True, color="FFFFFF")
            avg_total_cell.fill = PatternFill(start_color="FF9800", end_color="FF9800", fill_type="solid")
            
            # Overall win rate
            overall_win_rate = sum(stats['win_rate'] for stats in yearly_stats.values()) / len(yearly_stats) if yearly_stats else 0
            win_total_cell = results_sheet.cell(row=row, column=4, value=overall_win_rate)
            win_total_cell.number_format = "0.0%"
            win_total_cell.font = Font(name="Arial", size=12, bold=True, color="FFFFFF")
            win_total_cell.fill = PatternFill(start_color="FF9800", end_color="FF9800", fill_type="solid")
            
            total_trades_cell = results_sheet.cell(row=row, column=5, value=total_trades)
            total_trades_cell.font = Font(name="Arial", size=12, bold=True, color="FFFFFF")
            total_trades_cell.fill = PatternFill(start_color="FF9800", end_color="FF9800", fill_type="solid")
            
            row += 4
            
            # --- SECTION 3: OPEN TRADES ANALYSIS (REAL-TIME) ---
            
            # Open Trades Header
            open_header_cell = results_sheet.cell(row=row, column=1, value="📊 OPEN TRADES SUMMARY (REAL-TIME)")
            open_header_cell.font = Font(name="Arial", size=16, bold=True, color="FFFFFF")
            open_header_cell.fill = PatternFill(start_color="E91E63", end_color="E91E63", fill_type="solid")
            open_header_cell.alignment = Alignment(horizontal="center")
            results_sheet.merge_cells(f"A{row}:F{row}")
            row += 2
            
            # Get updated open trades data
            open_stats = self.calculate_enhanced_open_trades_stats()
            
            # Create attractive metrics display
            metrics_data = [
                ("📊 Portfolio Overview", "", "header"),
                ("Number of Open Trades", open_stats['count'], "count"),
                ("Total Cash Investment", open_stats['cash_investment'], "currency"),
                ("", "", "spacer"),
                ("💰 Option Values", "", "header"),
                ("Total Call Value", open_stats['call_value'], "currency"),
                ("Total Put Value", open_stats['put_value'], "currency"),
                ("Cash at Risk (Put Req.)", open_stats['put_cash_req'], "currency"),
                ("", "", "spacer"),
                ("📈 Current Performance", "", "header"),
                ("Current Unrealized P&L", open_stats['current_pl'], "pl"),
                ("Current P&L %", open_stats['current_pl_percent'], "percent"),
                ("Portfolio Value", open_stats['portfolio_value'], "currency"),
                ("", "", "spacer"),
                ("🎯 Risk Metrics", "", "header"),
                ("Total Exposure", open_stats['total_exposure'], "currency"),
                ("Options Premium Value", open_stats['options_premium'], "currency"),
                ("Net Liquidity", open_stats['net_liquidity'], "currency")
            ]
            
            for metric_name, value, metric_type in metrics_data:
                if metric_type == "header":
                    # Section header
                    header_cell = results_sheet.cell(row=row, column=1, value=metric_name)
                    header_cell.font = Font(name="Arial", size=13, bold=True, color="FFFFFF")
                    header_cell.fill = PatternFill(start_color="607D8B", end_color="607D8B", fill_type="solid")
                    header_cell.alignment = Alignment(horizontal="center")
                    results_sheet.merge_cells(f"A{row}:C{row}")
                    row += 1
                elif metric_type == "spacer":
                    row += 1
                else:
                    # Metric name
                    name_cell = results_sheet.cell(row=row, column=1, value=metric_name)
                    name_cell.font = Font(name="Arial", size=11, bold=True)
                    name_cell.fill = PatternFill(start_color="F8F9FA", end_color="F8F9FA", fill_type="solid")
                    
                    # Value with appropriate formatting and colors
                    value_cell = results_sheet.cell(row=row, column=2, value=value)
                    value_cell.font = Font(name="Arial", size=11, bold=True)
                    
                    if metric_type == "currency":
                        value_cell.number_format = "$#,##0.00"
                        if value > 0:
                            value_cell.fill = PatternFill(start_color="E8F5E8", end_color="E8F5E8", fill_type="solid")
                    elif metric_type == "percent":
                        value_cell.number_format = "0.0%"
                        if value > 0:
                            value_cell.font = Font(name="Arial", size=11, bold=True, color="006400")
                            value_cell.fill = PatternFill(start_color="E8F5E8", end_color="E8F5E8", fill_type="solid")
                        elif value < 0:
                            value_cell.font = Font(name="Arial", size=11, bold=True, color="CC0000")
                            value_cell.fill = PatternFill(start_color="FFF2F2", end_color="FFF2F2", fill_type="solid")
                    elif metric_type == "pl":
                        value_cell.number_format = "$#,##0.00"
                        if value > 0:
                            value_cell.font = Font(name="Arial", size=11, bold=True, color="006400")
                            value_cell.fill = PatternFill(start_color="E8F5E8", end_color="E8F5E8", fill_type="solid")
                        elif value < 0:
                            value_cell.font = Font(name="Arial", size=11, bold=True, color="CC0000")
                            value_cell.fill = PatternFill(start_color="FFF2F2", end_color="FFF2F2", fill_type="solid")
                    elif metric_type == "count":
                        value_cell.font = Font(name="Arial", size=11, bold=True, color="1F4E79")
                        value_cell.fill = PatternFill(start_color="E3F2FD", end_color="E3F2FD", fill_type="solid")
                    
                    row += 1
            
            row += 2
            
            # --- SECTION 3: POSITION BREAKDOWN ---
            breakdown_header = results_sheet.cell(row=row, column=1, value="📋 POSITION TYPE BREAKDOWN")
            breakdown_header.font = Font(name="Arial", size=14, bold=True, color="FFFFFF")
            breakdown_header.fill = PatternFill(start_color="9C27B0", end_color="9C27B0", fill_type="solid")
            breakdown_header.alignment = Alignment(horizontal="center")
            results_sheet.merge_cells(f"A{row}:D{row}")
            row += 2
            
            # Position breakdown
            position_breakdown = self.get_position_breakdown()
            
            # Headers
            breakdown_headers = ["Position Type", "Count", "Investment", "Current Value"]
            for col, header in enumerate(breakdown_headers, 1):
                cell = results_sheet.cell(row=row, column=col, value=header)
                cell.font = Font(name="Arial", size=11, bold=True, color="FFFFFF")
                cell.fill = PatternFill(start_color="607D8B", end_color="607D8B", fill_type="solid")
                cell.alignment = Alignment(horizontal="center")
            row += 1
            
            # Position data
            for pos_type, data in position_breakdown.items():
                results_sheet.cell(row=row, column=1, value=pos_type).font = Font(name="Arial", size=10)
                results_sheet.cell(row=row, column=2, value=data['count']).font = Font(name="Arial", size=10)
                
                inv_cell = results_sheet.cell(row=row, column=3, value=data['investment'])
                inv_cell.number_format = "$#,##0.00"
                inv_cell.font = Font(name="Arial", size=10)
                
                val_cell = results_sheet.cell(row=row, column=4, value=data['current_value'])
                val_cell.number_format = "$#,##0.00"
                val_cell.font = Font(name="Arial", size=10)
                
                row += 1
            
            # Auto-adjust column widths
            for col in range(1, 7):
                max_length = 0
                column = results_sheet.column_dimensions[chr(64 + col)]
                for row_cells in results_sheet.iter_rows():
                    try:
                        if len(str(row_cells[col-1].value)) > max_length:
                            max_length = len(str(row_cells[col-1].value))
                    except:
                        pass
                adjusted_width = min(max_length + 2, 25)
                column.width = adjusted_width
            
            # Add borders to all cells with content
            thin_border = Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )
            
            for row_cells in results_sheet.iter_rows():
                for cell in row_cells:
                    if cell.value is not None:
                        cell.border = thin_border
            
            wb.save(excel_filename)
            wb.close()
            
            messagebox.showinfo("Results Generated", "Comprehensive trading results with real-time data have been generated in the 'Results' sheet!")
            
        except Exception as e:
            messagebox.showerror("Results Error", f"Could not generate results.\n{e}")
            print(f"Results generation error: {e}")
            import traceback
            traceback.print_exc()

    def update_current_prices_for_all_trades(self):
        """Update current prices for all open trades in Excel"""
        try:
            excel_filename = "Put Transactions.xlsx"
            wb = load_workbook(excel_filename)
            
            if "Open_Trades_2025" not in wb.sheetnames:
                wb.close()
                return
            
            ws = wb["Open_Trades_2025"]
            headers = [cell.value for cell in ws[1]]
            
            # Find relevant columns
            ticker_col = self.find_header_column(headers, ["Stock ticker", "Ticker"])
            current_price_col = self.find_header_column(headers, ["Current Price"])
            current_pl_col = self.find_header_column(headers, ["Current P/L"])
            cost_col = self.find_header_column(headers, ["Cost"])
            shares_col = self.find_header_column(headers, ["Number of shares +/-", "Shares"])
            
            if not all([ticker_col, current_price_col, cost_col, shares_col]):
                print("Could not find required columns for price update")
                wb.close()
                return
            
            updated_count = 0
            for r in range(2, ws.max_row+1):
                try:
                    ticker = ws.cell(row=r, column=ticker_col).value
                    cost = ws.cell(row=r, column=cost_col).value if cost_col else 0
                    shares = ws.cell(row=r, column=shares_col).value if shares_col else 0
                    current_price = ws.cell(row=r, column=current_price_col).value if current_price_col else 0
                    # Calculate current P/L
                    try:
                        cost_val = float(str(cost).replace("$", "").replace(",", "")) if cost else 0
                        shares_val = float(str(shares).replace(",", "")) if shares else 0
                        price_val = float(str(current_price).replace("$", "").replace(",", "")) if current_price else 0
                        if cost_val > 0 and shares_val != 0:
                            current_pl = (price_val - cost_val) * shares_val
                            ws.cell(row=r, column=current_pl_col, value=current_pl)
                            ws.cell(row=r, column=current_pl_col).number_format = "$#,##0.00"
                            # Color code the P&L
                            pl_cell = ws.cell(row=r, column=current_pl_col)
                            if current_pl > 0:
                                pl_cell.font = Font(name="Arial", size=12, color="006400")
                            elif current_pl < 0:
                                pl_cell.font = Font(name="Arial", size=12, color="CC0000")
                    except Exception as calc_error:
                        print(f"Error calculating P&L for {ticker}: {calc_error}")
                    updated_count += 1
                    print(f"Updated {ticker}: ${current_price}")
                except Exception as price_error:
                    print(f"Could not get price for {ticker}: {price_error}")
                    continue
            
            wb.save(excel_filename)
            wb.close()
            
            # Reload the GUI data to reflect updates
            self.load_open_trades_from_excel()
            
            print(f"Successfully updated prices for {updated_count} trades")
            
        except Exception as e:
            print(f"Error updating current prices: {e}")

    def calculate_enhanced_open_trades_stats(self):
        """Calculate comprehensive statistics for open trades with real-time data"""
        stats = {
            'count': 0,
            'cash_investment': 0,
            'call_value': 0,
            'put_value': 0,
            'put_cash_req': 0,
            'current_pl': 0,
            'current_pl_percent': 0,
            'portfolio_value': 0,
            'total_exposure': 0,
            'options_premium': 0,
            'net_liquidity': 0
        }
        
        try:
            total_investment = 0
            
            for trade in self.trades:
                if len(trade) >= len(COLUMNS):
                    stats['count'] += 1
                    
                    # Cash investment (Total investment + Put cash requirement)
                    investment = self.parse_currency(trade[COLUMNS.index('Investment')])
                    put_cash_req = self.parse_currency(trade[COLUMNS.index('Put Cash Req')])
                    cash_inv = investment + put_cash_req
                    stats['cash_investment'] += cash_inv
                    total_investment += cash_inv
                    
                    # Call and Put values
                    call_val = self.parse_currency(trade[COLUMNS.index('Call Value')])
                    put_val = self.parse_currency(trade[COLUMNS.index('Put Value')])
                    stats['call_value'] += call_val
                    stats['put_value'] += put_val
                    stats['put_cash_req'] += put_cash_req
                    
                    # Current P&L
                    current_pl = self.parse_currency(trade[COLUMNS.index('Current P/L')])
                    stats['current_pl'] += current_pl
                    
                    # Portfolio calculations
                    stats['options_premium'] += call_val + put_val
                    stats['total_exposure'] += cash_inv + put_cash_req
            
            # Calculate derived metrics
            if total_investment > 0:
                stats['current_pl_percent'] = stats['current_pl'] / total_investment
            
            stats['portfolio_value'] = stats['cash_investment'] + stats['current_pl']
            stats['net_liquidity'] = stats['call_value'] + stats['put_value'] - stats['put_cash_req']
        
        except Exception as e:
            print(f"Error calculating enhanced open trades stats: {e}")
        
        return stats

    def get_closed_trades_data(self):
        """Get all closed trades data from Excel"""
        try:
            excel_filename = "Put Transactions.xlsx"
            wb = load_workbook(excel_filename)
            
            if "Closed_Trades_2025" not in wb.sheetnames:
                wb.close()
                return []
            
            closed_sheet = wb["Closed_Trades_2025"]
            headers = [cell.value for cell in closed_sheet[1]]
            
            closed_trades = []
            for row in range(2, closed_sheet.max_row + 1):
                trade = {}
                for col, header in enumerate(headers, 1):
                    value = closed_sheet.cell(row=row, column=col).value
                    app_col = EXCEL_TO_APP_COLS.get(header, header)
                    trade[app_col] = value
                closed_trades.append(trade)
            
            wb.close()
            return closed_trades
            
        except Exception as e:
            print(f"Error getting closed trades data: {e}")
            return []

    def calculate_yearly_stats(self, closed_data):
        """Calculate yearly statistics from closed trades including win rate"""
        from datetime import datetime
        yearly_stats = {}
        
        for trade in closed_data:
            try:
                # Get close date
                close_date = trade.get('Close Date', '')
                if not close_date:
                    continue
                
                # Parse date to get year
                if hasattr(close_date, 'year'):
                    year = close_date.year
                else:
                    # Parse string date
                    date_str = str(close_date).split()[0]
                    for fmt in ["%m/%d/%Y", "%Y-%m-%d", "%m-%d-%Y"]:
                        try:
                            dt = datetime.strptime(date_str, fmt)
                            year = dt.year
                            break
                        except:
                            continue
                    else:
                        continue
                
                if year not in yearly_stats:
                    yearly_stats[year] = {
                        'total_pl': 0, 
                        'total_percent': 0, 
                        'trade_count': 0,
                        'winning_trades': 0,
                        'losing_trades': 0
                    }
                
                # Add P&L
                pl = trade.get('Realized P/L', 0) or 0
                if isinstance(pl, str):
                    pl = float(pl.replace('$', '').replace(',', ''))
                yearly_stats[year]['total_pl'] += pl
                
                # Count wins/losses
                if pl > 0:
                    yearly_stats[year]['winning_trades'] += 1
                elif pl < 0:
                    yearly_stats[year]['losing_trades'] += 1
                
                # Add P&L percentage
                pl_percent = trade.get('P&L %', 0) or 0
                if isinstance(pl_percent, str):
                    pl_percent = float(pl_percent.replace('%', ''))
                elif isinstance(pl_percent, float) and pl_percent < 1:
                    # Convert decimal to percentage
                    pl_percent = pl_percent * 100
                yearly_stats[year]['total_percent'] += pl_percent
                
                yearly_stats[year]['trade_count'] += 1
                
            except Exception as e:
                print(f"Error processing trade for yearly stats: {e}")
                continue
        
        # Calculate averages and win rates
        for year, stats in yearly_stats.items():
            if stats['trade_count'] > 0:
                stats['avg_percent'] = stats['total_percent'] / stats['trade_count'] / 100
                stats['win_rate'] = stats['winning_trades'] / stats['trade_count']
            else:
                stats['avg_percent'] = 0
                stats['win_rate'] = 0
        
        return yearly_stats

    def calculate_open_trades_stats(self):
        """Calculate statistics for open trades"""
        stats = {
            'count': 0,
            'cash_investment': 0,
            'call_value': 0,
            'put_value': 0,
            'put_cash_req': 0,
            'current_pl': 0
        }
        
        try:
            for trade in self.trades:
                if len(trade) >= len(COLUMNS):
                    stats['count'] += 1
                    
                    # Cash investment (Total investment + Put cash requirement)
                    investment = self.parse_currency(trade[COLUMNS.index('Investment')])
                    put_cash_req = self.parse_currency(trade[COLUMNS.index('Put Cash Req')])
                    stats['cash_investment'] += investment + put_cash_req
                    
                    # Call and Put values
                    stats['call_value'] += self.parse_currency(trade[COLUMNS.index('Call Value')])
                    stats['put_value'] += self.parse_currency(trade[COLUMNS.index('Put Value')])
                    stats['put_cash_req'] += put_cash_req
                    
                    # Current P&L
                    stats['current_pl'] += self.parse_currency(trade[COLUMNS.index('Current P/L')])
        
        except Exception as e:
            print(f"Error calculating open trades stats: {e}")
        
        return stats

    def get_position_breakdown(self):
        """Get breakdown of positions by type"""
        breakdown = {}
        
        try:
            for trade in self.trades:
                if len(trade) >= len(COLUMNS):
                    position_type = trade[COLUMNS.index('Type')]
                    
                    if position_type not in breakdown:
                        breakdown[position_type] = {
                            'count': 0,
                            'investment': 0,
                            'current_value': 0
                        }
                    
                    breakdown[position_type]['count'] += 1
                    breakdown[position_type]['investment'] += self.parse_currency(trade[COLUMNS.index('Investment')])
                    
                    # Current value based on type
                    if 'call' in position_type.lower():
                        breakdown[position_type]['current_value'] += self.parse_currency(trade[COLUMNS.index('Call Value')])
                    elif 'put' in position_type.lower():
                        breakdown[position_type]['current_value'] += self.parse_currency(trade[COLUMNS.index('Put Value')])
        
        except Exception as e:
            print(f"Error calculating position breakdown: {e}")
        
        return breakdown

    def parse_currency(self, value):
        """Parse currency string to float"""
        if not value or value == "":
            return 0.0
        try:
            return float(str(value).replace('$', '').replace(',', ''))
        except:
            return 0.0


        
if __name__ == "__main__":
    root = tk.Tk()
    app = TradeTrackerApp(root)
    root.mainloop()
