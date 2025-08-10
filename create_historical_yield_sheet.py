#!/usr/bin/env python3
"""
Create Historical Yield Data Sheet
Extracts your actual weekly historical yield data and creates a new sheet ready for integration
"""

import pandas as pd
import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils.dataframe import dataframe_to_rows
import os
from datetime import datetime

def load_real_historical_data():
    """Load your actual historical yield data from the main Excel file"""
    print("📊 Loading REAL historical yield data from your Excel file...")
    
    excel_file = r"c:\Users\mjmat\Python Code in VS\dividend_tracker\DividendTrackerApp\outputs\Dividends_2025.xlsx"
    
    if not os.path.exists(excel_file):
        print(f"❌ Excel file not found: {excel_file}")
        return {}, []
    
    # Try to find the sheet with historical data
    wb = openpyxl.load_workbook(excel_file)
    
    # Look for sheets that contain historical yield data
    potential_sheets = []
    for sheet_name in wb.sheetnames:
        if any(keyword in sheet_name.lower() for keyword in ['historic', 'yield', 'ticker', 'analysis']):
            potential_sheets.append(sheet_name)
    
    print(f"📋 Found potential data sheets: {potential_sheets}")
    
    historical_data = {}
    date_columns = []
    
    # Try to read from the most likely sheets (prioritize historic yield sheets)
    sorted_sheets = sorted(potential_sheets, key=lambda x: 'historic' in x.lower(), reverse=True)
    
    for sheet_name in sorted_sheets:
        try:
            print(f"🔍 Checking sheet: {sheet_name}")
            df = pd.read_excel(excel_file, sheet_name=sheet_name)
            
            # Look for date columns in various formats
            found_date_columns = []
            for col in df.columns:
                col_str = str(col)
                # Look for MM-DD-YYYY or YYYY-MM-DD patterns
                if ('-2025' in col_str or '2025-' in col_str) and any(char.isdigit() for char in col_str):
                    found_date_columns.append(col_str)
            
            if found_date_columns and 'Ticker' in df.columns:
                print(f"   ✅ Found {len(found_date_columns)} date columns in {sheet_name}")
                print(f"   📅 Sample dates: {found_date_columns[:5]}")
                
                # Extract historical yield data
                ticker_count = 0
                for _, row in df.iterrows():
                    if pd.notna(row.get('Ticker')):
                        ticker = str(row['Ticker']).strip().upper()
                        historical_data[ticker] = {}
                        has_yield_data = False
                        
                        for date_col in found_date_columns:
                            yield_value = row.get(date_col)
                            if pd.notna(yield_value) and isinstance(yield_value, (int, float)) and yield_value > 0:
                                historical_data[ticker][date_col] = float(yield_value)
                                has_yield_data = True
                        
                        if has_yield_data:
                            ticker_count += 1
                        else:
                            # Remove ticker if no yield data
                            del historical_data[ticker]
                
                print(f"   📊 Successfully loaded {ticker_count} tickers with yield data")
                
                # Use the longest list of date columns found
                if len(found_date_columns) > len(date_columns):
                    date_columns = found_date_columns
                    print(f"   📈 Using {len(date_columns)} historical weeks from {sheet_name}")
                
                if historical_data:  # If we found good data, use it
                    break
                
        except Exception as e:
            print(f"   ⚠️ Could not read {sheet_name}: {e}")
    
    if not historical_data:
        print("❌ No historical yield data found in Excel file")
        return {}, []
    
    print(f"✅ Loaded historical data for {len(historical_data)} tickers across {len(date_columns)} weeks")
    return historical_data, date_columns

def create_historical_yield_sheet():
    """Create a new sheet with historical yield data ready for integration"""
    print("🚀 Creating Historical Yield Data Sheet...")
    
    # Load real historical data
    historical_data, date_columns = load_real_historical_data()
    
    if not historical_data:
        print("❌ No historical data to work with")
        return
    
    # Sort date columns chronologically (newest first)
    try:
        sorted_dates = sorted(date_columns, 
                            key=lambda x: datetime.strptime(x, '%m-%d-%Y'), 
                            reverse=True)
    except:
        # Fallback sorting
        sorted_dates = sorted(date_columns, reverse=True)
    
    print(f"📅 Date range: {sorted_dates[-1]} to {sorted_dates[0]} ({len(sorted_dates)} weeks)")
    
    # Create new Excel file for the historical data sheet
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Historical Yield Data"
    
    # Set up headers
    headers = ['Ticker', 'Account', 'Current Yield', 'Yield Trend'] + sorted_dates
    
    # Style the headers
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col)
        cell.value = header
        cell.font = Font(bold=True, color="FFFFFF", size=12)
        
        if col <= 4:
            # Static columns
            cell.fill = PatternFill(start_color="2F4F4F", end_color="2F4F4F", fill_type="solid")
        else:
            # Date columns
            cell.fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
    
    # Add data for each ticker
    row = 2
    for ticker in sorted(historical_data.keys()):
        ticker_history = historical_data[ticker]
        
        # Calculate current yield (most recent date)
        current_yield = 0
        for date in sorted_dates:
            if date in ticker_history:
                current_yield = ticker_history[date]
                break
        
        # Calculate yield trend
        yield_trend = "No Data"
        trend_color = "D3D3D3"  # Gray
        
        if len(sorted_dates) >= 2:
            recent_yields = []
            for date in sorted_dates[:4]:  # Look at last 4 weeks
                if date in ticker_history:
                    recent_yields.append(ticker_history[date])
            
            if len(recent_yields) >= 2:
                trend_change = recent_yields[0] - recent_yields[-1]
                if trend_change > 0.5:
                    yield_trend = "⬆️ Rising"
                    trend_color = "90EE90"  # Light green
                elif trend_change < -0.5:
                    yield_trend = "⬇️ Falling"
                    trend_color = "FFB6C1"  # Light pink
                else:
                    yield_trend = "➡️ Stable"
                    trend_color = "FFFF90"  # Light yellow
        
        # Ticker name
        ws.cell(row=row, column=1).value = ticker
        ws.cell(row=row, column=1).font = Font(bold=True, size=11)
        
        # Account (placeholder - you can fill this in)
        ws.cell(row=row, column=2).value = "TBD"
        ws.cell(row=row, column=2).fill = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")
        
        # Current yield
        current_cell = ws.cell(row=row, column=3)
        current_cell.value = current_yield
        current_cell.number_format = '0.00'
        if current_yield >= 15:
            current_cell.fill = PatternFill(start_color="90EE90", end_color="90EE90", fill_type="solid")
        elif current_yield >= 10:
            current_cell.fill = PatternFill(start_color="FFFF90", end_color="FFFF90", fill_type="solid")
        elif current_yield >= 4:
            current_cell.fill = PatternFill(start_color="FFE4B5", end_color="FFE4B5", fill_type="solid")
        
        # Yield trend
        trend_cell = ws.cell(row=row, column=4)
        trend_cell.value = yield_trend
        trend_cell.fill = PatternFill(start_color=trend_color, end_color=trend_color, fill_type="solid")
        trend_cell.font = Font(size=10)
        
        # Historical yield data
        for date_idx, date in enumerate(sorted_dates):
            col = 5 + date_idx
            yield_value = ticker_history.get(date, '')
            
            cell = ws.cell(row=row, column=col)
            if yield_value:
                cell.value = yield_value
                cell.number_format = '0.00'
                
                # Color code based on yield level
                if yield_value >= 15:
                    cell.fill = PatternFill(start_color="90EE90", end_color="90EE90", fill_type="solid")  # Green
                elif yield_value >= 10:
                    cell.fill = PatternFill(start_color="FFFF90", end_color="FFFF90", fill_type="solid")  # Yellow
                elif yield_value >= 4:
                    cell.fill = PatternFill(start_color="FFE4B5", end_color="FFE4B5", fill_type="solid")  # Orange
                else:
                    cell.fill = PatternFill(start_color="FFB6C1", end_color="FFB6C1", fill_type="solid")  # Pink
            else:
                cell.value = "N/A"
                cell.fill = PatternFill(start_color="D3D3D3", end_color="D3D3D3", fill_type="solid")  # Gray
                cell.font = Font(size=9, italic=True)
        
        row += 1
    
    # Auto-adjust column widths
    for col in range(1, len(headers) + 1):
        max_length = 0
        column_letter = openpyxl.utils.get_column_letter(col)
        
        for row_num in range(1, min(ws.max_row + 1, 20)):
            try:
                if ws[f"{column_letter}{row_num}"].value:
                    max_length = max(max_length, len(str(ws[f"{column_letter}{row_num}"].value)))
            except:
                pass
        
        adjusted_width = min(max_length + 2, 15)
        ws.column_dimensions[column_letter].width = adjusted_width
    
    # Add summary information
    summary_row = row + 2
    ws.cell(row=summary_row, column=1).value = "📊 HISTORICAL YIELD DATA SUMMARY"
    ws.cell(row=summary_row, column=1).font = Font(bold=True, size=14, color="2F4F4F")
    
    ws.cell(row=summary_row + 1, column=1).value = f"✅ Total Tickers: {len(historical_data)}"
    ws.cell(row=summary_row + 2, column=1).value = f"📅 Weeks of Data: {len(sorted_dates)}"
    ws.cell(row=summary_row + 3, column=1).value = f"🗓️ Date Range: {sorted_dates[-1]} to {sorted_dates[0]}"
    ws.cell(row=summary_row + 4, column=1).value = f"⏰ Generated: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}"
    
    # Color coding legend
    legend_row = summary_row + 6
    ws.cell(row=legend_row, column=1).value = "🎨 COLOR CODING LEGEND:"
    ws.cell(row=legend_row, column=1).font = Font(bold=True, size=12)
    
    legend_items = [
        ("🟢 Green: Yields ≥ 15%", "90EE90"),
        ("🟡 Yellow: Yields 10-15%", "FFFF90"),
        ("🟠 Orange: Yields 4-10%", "FFE4B5"),
        ("🟣 Pink: Yields < 4%", "FFB6C1"),
        ("⚫ Gray: No data available", "D3D3D3")
    ]
    
    for i, (label, color) in enumerate(legend_items):
        cell = ws.cell(row=legend_row + 1 + i, column=1)
        cell.value = label
        cell.fill = PatternFill(start_color=color, end_color=color, fill_type="solid")
        cell.font = Font(size=10)
    
    # Save the file
    output_file = "Historical_Yield_Data_Ready_for_Integration.xlsx"
    wb.save(output_file)
    
    print(f"""
🎉 HISTORICAL YIELD DATA SHEET CREATED!

✅ File saved: {output_file}
✅ Tickers included: {len(historical_data)}
✅ Historical weeks: {len(sorted_dates)}
✅ Date range: {sorted_dates[-1]} to {sorted_dates[0]}

📋 WHAT'S INCLUDED:
• Ticker column
• Account column (marked TBD for you to fill in)
• Current yield with color coding
• Yield trend analysis (Rising/Falling/Stable)
• {len(sorted_dates)} weeks of historical yield data
• Professional color coding by yield level
• Summary statistics and legend

🔧 NEXT STEPS:
1. Fill in the Account column with proper account names
2. Copy/move the historical date columns to your Ticker Analysis 2025 sheet
3. The data is already color-coded and formatted for easy integration

This sheet contains your REAL historical yield data in a format ready for integration!
""")

if __name__ == "__main__":
    print("📈 Creating Historical Yield Data Sheet from your actual data...")
    create_historical_yield_sheet()
