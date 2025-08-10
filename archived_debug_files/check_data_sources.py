#!/usr/bin/env python3
"""
Check Excel Structure and Available Data Sources
Examine the current Excel structure and find the correct source for historical weekly data
"""

import pandas as pd
import openpyxl
import os
from datetime import datetime

def examine_excel_structure():
    """Examine the Excel file structure and available data"""
    
    excel_file = r"c:\Users\mjmat\Python Code in VS\dividend_tracker\DividendTrackerApp\outputs\Dividends_2025.xlsx"
    
    if not os.path.exists(excel_file):
        print(f"❌ Excel file not found: {excel_file}")
        return
    
    # Load workbook
    wb = openpyxl.load_workbook(excel_file)
    print(f"📊 Excel file has {len(wb.sheetnames)} sheets:")
    
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        print(f"   📋 {sheet_name}: {ws.max_row} rows x {ws.max_column} columns")
        
        # Show first few headers
        headers = []
        for col in range(1, min(ws.max_column + 1, 10)):  # First 10 columns max
            cell_value = ws.cell(row=1, column=col).value
            if cell_value:
                headers.append(str(cell_value))
        print(f"      Headers: {headers}")
        print()
    
    # Check the "All account weekly dividends" sheet specifically for historical data
    if "All account weekly dividends" in wb.sheetnames:
        print("🔍 Examining 'All account weekly dividends' sheet for historical data...")
        ws = wb["All account weekly dividends"]
        
        # Show all headers
        headers = []
        for col in range(1, ws.max_column + 1):
            cell_value = ws.cell(row=1, column=col).value
            if cell_value:
                headers.append(str(cell_value))
        
        print(f"   📋 Full headers: {headers}")
        
        # Show a few sample rows
        print("   📊 Sample data:")
        for row in range(2, min(ws.max_row + 1, 5)):  # First few data rows
            row_data = []
            for col in range(1, min(ws.max_column + 1, 8)):  # First 8 columns
                cell_value = ws.cell(row=row, column=col).value
                row_data.append(str(cell_value) if cell_value else "")
            print(f"      Row {row}: {row_data}")
    
    wb.close()

def examine_csv_files():
    """Check what CSV files might have ticker-level historical data"""
    
    print("🔍 Checking for CSV files with ticker-level data...")
    
    # Check scored holdings
    holdings_file = "scored_my_holdings.csv"
    if os.path.exists(holdings_file):
        print(f"✅ Found {holdings_file}")
        df = pd.read_csv(holdings_file)
        print(f"   📊 {len(df)} rows x {len(df.columns)} columns")
        print(f"   📋 Columns: {list(df.columns)}")
        if 'Ticker' in df.columns and 'Yield Est (%)' in df.columns:
            print(f"   🎯 Contains ticker yield data!")
            print(f"   📈 Sample yields: {df[['Ticker', 'Yield Est (%)']].head()}")
        print()
    
    # Check for backup ticker files (these might have historical weekly data)
    backup_files = [f for f in os.listdir('.') if f.startswith('backup_ticker_list_')]
    print(f"📁 Found {len(backup_files)} backup ticker files:")
    for file in sorted(backup_files)[:5]:  # Show first 5
        if file.endswith('.json'):
            print(f"   📄 {file}")
    
    if backup_files:
        # Check one backup file structure
        import json
        latest_backup = sorted(backup_files)[-1]
        if latest_backup.endswith('.json'):
            try:
                with open(latest_backup, 'r') as f:
                    data = json.load(f)
                print(f"\n🔍 Examining {latest_backup}:")
                if isinstance(data, dict):
                    print(f"   📊 Contains {len(data)} entries")
                    # Show sample entry
                    sample_key = list(data.keys())[0] if data else None
                    if sample_key:
                        print(f"   📋 Sample entry '{sample_key}': {data[sample_key]}")
                elif isinstance(data, list):
                    print(f"   📊 Contains {len(data)} items")
                    if data:
                        print(f"   📋 Sample item: {data[0]}")
            except Exception as e:
                print(f"   ❌ Could not read {latest_backup}: {e}")

if __name__ == "__main__":
    print("🔍 EXAMINING DATA SOURCES FOR HISTORICAL YIELDS")
    print("=" * 60)
    
    examine_excel_structure()
    print()
    examine_csv_files()
