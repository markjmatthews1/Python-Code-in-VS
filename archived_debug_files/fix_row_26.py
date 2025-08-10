#!/usr/bin/env python3
"""Fix row 26 specifically"""

import os
from openpyxl import load_workbook
from datetime import datetime

def fix_row_26():
    """Fix the SOFI Sold Put in row 26"""
    excel_filename = "Bryan Perry Transactions.xlsx"
    sheet_name = "Open_Trades_2025"
    
    print("Checking and fixing row 26 (SOFI Sold Put)...")
    
    try:
        from etrade_auth import get_etrade_session
        session, base_url = get_etrade_session()
        
        wb = load_workbook(excel_filename)
        ws = wb[sheet_name]
        header = [cell.value for cell in ws[1]]
        
        # Column positions
        ticker_col = header.index("Stock ticker") + 1
        type_col = header.index("Stock Purchase Type") + 1
        strike_col = header.index("Strike price") + 1
        expiry_col = header.index("Expiration Date") + 1
        cost_col = header.index("Cost") + 1
        shares_col = header.index("Number of shares +/-") + 1
        current_price_col = header.index("Current Price") + 1
        current_pl_col = header.index("Current P/L") + 1
        
        r = 26  # Row 26
        
        ticker = ws.cell(row=r, column=ticker_col).value
        trade_type = str(ws.cell(row=r, column=type_col).value).lower()
        strike = ws.cell(row=r, column=strike_col).value
        expiry = ws.cell(row=r, column=expiry_col).value
        cost = ws.cell(row=r, column=cost_col).value
        shares = ws.cell(row=r, column=shares_col).value
        current_price = ws.cell(row=r, column=current_price_col).value
        current_pl = ws.cell(row=r, column=current_pl_col).value
        
        print(f"Row 26 current data:")
        print(f"  Ticker: {ticker}")
        print(f"  Type: {trade_type}")
        print(f"  Strike: ${strike}")
        print(f"  Expiry: {expiry}")
        print(f"  Cost: ${cost}")
        print(f"  Shares: {shares}")
        print(f"  Current Price: {current_price}")
        print(f"  Current P/L: {current_pl}")
        
        # Fetch fresh option price for SOFI put
        if "put" in trade_type:
            print(f"\nFetching fresh SOFI put option price...")
            
            try:
                # Parse expiry date
                if isinstance(expiry, datetime):
                    exp_date = expiry.date()
                else:
                    exp_date = datetime.strptime(str(expiry)[:10], "%Y-%m-%d").date()
                
                # Get available expiration dates
                exp_url = f"{base_url}/v1/market/optionexpiredate.json?symbol={ticker}"
                exp_response = session.get(exp_url)
                
                if exp_response.status_code == 200:
                    exp_data = exp_response.json()
                    dates = exp_data.get("OptionExpireDateResponse", {}).get("ExpirationDate", [])
                    
                    # Find closest expiration date
                    best_match = None
                    min_diff = float('inf')
                    
                    for date_info in dates:
                        exp_datetime = datetime(date_info['year'], date_info['month'], date_info['day']).date()
                        diff = abs((exp_datetime - exp_date).days)
                        if diff < min_diff:
                            min_diff = diff
                            best_match = date_info
                    
                    if best_match:
                        print(f"Using expiration date: {best_match['month']}/{best_match['day']}/{best_match['year']}")
                        
                        # Get option chain
                        chain_url = f"{base_url}/v1/market/optionchains.json?symbol={ticker}&expiryDay={best_match['day']:02d}&expiryMonth={best_match['month']:02d}&expiryYear={best_match['year']}&chainType=PUT"
                        chain_response = session.get(chain_url)
                        
                        if chain_response.status_code == 200:
                            chain_data = chain_response.json()
                            option_pairs = chain_data.get("OptionChainResponse", {}).get("OptionPair", [])
                            
                            # Find closest strike
                            target_strike = float(strike)
                            best_option = None
                            min_strike_diff = float('inf')
                            
                            print(f"Looking for strike ${target_strike} in {len(option_pairs)} option pairs...")
                            
                            for pair in option_pairs:
                                put_data = pair.get("Put", {})
                                if put_data:
                                    option_strike = float(put_data.get("strikePrice", 0))
                                    strike_diff = abs(option_strike - target_strike)
                                    print(f"  Found strike ${option_strike}, diff: {strike_diff}")
                                    if strike_diff < min_strike_diff:
                                        min_strike_diff = strike_diff
                                        best_option = put_data
                            
                            if best_option:
                                last_price = best_option.get("lastPrice")
                                if last_price:
                                    new_price = float(last_price)
                                    print(f"✓ Found fresh option price: ${new_price}")
                                    
                                    # Force update the Excel cells
                                    ws.cell(row=r, column=current_price_col, value=new_price)
                                    ws.cell(row=r, column=current_price_col).number_format = '$#,##0.00'
                                    
                                    # Calculate and update Current P/L
                                    if cost and shares:
                                        cost_val = float(cost)
                                        shares_val = float(shares)
                                        pl_val = (new_price - cost_val) * shares_val
                                        
                                        ws.cell(row=r, column=current_pl_col, value=pl_val)
                                        ws.cell(row=r, column=current_pl_col).number_format = '$#,##0.00'
                                        
                                        print(f"✓ Updated: Price=${new_price:,.2f}, P/L=${pl_val:,.2f}")
                                        print(f"  Calculation: ({new_price} - {cost_val}) × {shares_val} = {pl_val}")
                                        
                                        wb.save(excel_filename)
                                        print(f"✅ Row 26 updated and saved to Excel!")
                                    else:
                                        print(f"✗ Cannot calculate P/L: missing cost or shares")
                                else:
                                    print(f"✗ No lastPrice in option data")
                            else:
                                print(f"✗ No matching option found for strike ${target_strike}")
                        else:
                            print(f"✗ Option chain request failed: {chain_response.status_code}")
                    else:
                        print(f"✗ No matching expiration date found")
                else:
                    print(f"✗ Expiration dates request failed: {exp_response.status_code}")
                    
            except Exception as e:
                print(f"✗ Error fetching option price: {e}")
                import traceback
                traceback.print_exc()
        
        wb.close()
        
    except Exception as e:
        print(f"Error: {e}")
        import traceback
        traceback.print_exc()

if __name__ == "__main__":
    fix_row_26()
