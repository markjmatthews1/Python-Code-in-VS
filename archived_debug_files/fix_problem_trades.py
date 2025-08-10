#!/usr/bin/env python3
"""Direct price update for the problematic trades"""

import os
from openpyxl import load_workbook
from datetime import datetime

def update_problem_trades():
    """Update prices for the specific problematic trades"""
    excel_filename = "Bryan Perry Transactions.xlsx"
    sheet_name = "Open_Trades_2025"
    
    print("Updating prices for problematic trades (rows 23-25)...")
    
    try:
        from etrade_auth import get_etrade_session
        session, base_url = get_etrade_session()
        
        wb = load_workbook(excel_filename)
        ws = wb[sheet_name]
        header = [cell.value for cell in ws[1]]
        
        # Column positions
        ticker_col = header.index("Stock ticker") + 1
        type_col = header.index("Stock Purchase Type") + 1
        strike_col = header.index("Strike price") + 1
        expiry_col = header.index("Expiration Date") + 1
        cost_col = header.index("Cost") + 1
        shares_col = header.index("Number of shares +/-") + 1
        current_price_col = header.index("Current Price") + 1
        current_pl_col = header.index("Current P/L") + 1
        
        problem_rows = [23, 24, 25]
        updated_count = 0
        
        for r in problem_rows:
            ticker = ws.cell(row=r, column=ticker_col).value
            trade_type = str(ws.cell(row=r, column=type_col).value).lower()
            strike = ws.cell(row=r, column=strike_col).value
            expiry = ws.cell(row=r, column=expiry_col).value
            cost = ws.cell(row=r, column=cost_col).value
            shares = ws.cell(row=r, column=shares_col).value
            
            print(f"\n--- Updating Row {r}: {ticker} ({trade_type}) ---")
            
            current_price = None
            
            if "put" in trade_type or "call" in trade_type:
                # Option trade
                print(f"  Fetching option price: {ticker} ${strike} {'call' if 'call' in trade_type else 'put'} {expiry}")
                
                try:
                    # Parse expiry date
                    if isinstance(expiry, datetime):
                        exp_date = expiry.date()
                    else:
                        exp_date = datetime.strptime(str(expiry)[:10], "%Y-%m-%d").date()
                    
                    option_type = "call" if "call" in trade_type else "put"
                    
                    # Get available expiration dates
                    exp_url = f"{base_url}/v1/market/optionexpiredate.json?symbol={ticker}"
                    exp_response = session.get(exp_url)
                    
                    if exp_response.status_code == 200:
                        exp_data = exp_response.json()
                        dates = exp_data.get("OptionExpireDateResponse", {}).get("ExpirationDate", [])
                        
                        # Find closest expiration date
                        best_match = None
                        min_diff = float('inf')
                        
                        for date_info in dates:
                            exp_datetime = datetime(date_info['year'], date_info['month'], date_info['day']).date()
                            diff = abs((exp_datetime - exp_date).days)
                            if diff < min_diff:
                                min_diff = diff
                                best_match = date_info
                        
                        if best_match:
                            # Get option chain
                            chain_url = f"{base_url}/v1/market/optionchains.json?symbol={ticker}&expiryDay={best_match['day']:02d}&expiryMonth={best_match['month']:02d}&expiryYear={best_match['year']}&chainType={option_type.upper()}"
                            chain_response = session.get(chain_url)
                            
                            if chain_response.status_code == 200:
                                chain_data = chain_response.json()
                                option_pairs = chain_data.get("OptionChainResponse", {}).get("OptionPair", [])
                                
                                # Find closest strike
                                target_strike = float(strike)
                                best_option = None
                                min_strike_diff = float('inf')
                                
                                for pair in option_pairs:
                                    option_data = pair.get("Call" if option_type == "call" else "Put", {})
                                    if option_data:
                                        option_strike = float(option_data.get("strikePrice", 0))
                                        strike_diff = abs(option_strike - target_strike)
                                        if strike_diff < min_strike_diff:
                                            min_strike_diff = strike_diff
                                            best_option = option_data
                                
                                if best_option:
                                    last_price = best_option.get("lastPrice")
                                    if last_price:
                                        current_price = float(last_price)
                                        print(f"  ✓ Found option price: ${current_price}")
                                
                except Exception as e:
                    print(f"  ✗ Error fetching option price: {e}")
            
            else:
                # Stock trade
                print(f"  Fetching stock price for {ticker}")
                try:
                    url = f"{base_url}/v1/market/quote/{ticker}.json"
                    response = session.get(url)
                    
                    if response.status_code == 200:
                        data = response.json()
                        quote_data = data.get("QuoteResponse", {}).get("QuoteData", [])
                        if quote_data and len(quote_data) > 0:
                            all_data = quote_data[0].get("All", {})
                            last_trade = all_data.get("lastTrade")
                            if last_trade:
                                current_price = float(last_trade)
                                print(f"  ✓ Found stock price: ${current_price}")
                
                except Exception as e:
                    print(f"  ✗ Error fetching stock price: {e}")
            
            # Update Excel if we got a price
            if current_price:
                # Update Current Price
                ws.cell(row=r, column=current_price_col, value=current_price)
                ws.cell(row=r, column=current_price_col).number_format = '$#,##0.00'
                
                # Calculate and update Current P/L
                if cost and shares:
                    cost_val = float(cost)
                    shares_val = float(shares)
                    pl_val = (current_price - cost_val) * shares_val
                    
                    ws.cell(row=r, column=current_pl_col, value=pl_val)
                    ws.cell(row=r, column=current_pl_col).number_format = '$#,##0.00'
                    
                    print(f"  ✓ Updated: Price=${current_price:,.2f}, P/L=${pl_val:,.2f}")
                    updated_count += 1
                else:
                    print(f"  ✗ Cannot calculate P/L: missing cost or shares")
            else:
                print(f"  ✗ Could not get price for {ticker}")
        
        if updated_count > 0:
            wb.save(excel_filename)
            print(f"\n✅ Successfully updated {updated_count} trades and saved to Excel!")
        else:
            print(f"\n❌ No trades were updated")
        
        wb.close()
        
    except Exception as e:
        print(f"Error: {e}")
        import traceback
        traceback.print_exc()

if __name__ == "__main__":
    update_problem_trades()
