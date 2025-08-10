#!/usr/bin/env python3
"""
Check Current Excel Headers and Data Structure
Analyze what headers are actually being used in the Excel file
"""

import pandas as pd
import openpyxl
import os

def check_current_headers():
    """Check the current headers in both main Excel files"""
    print("🔍 Checking current Excel file headers and structure...")
    
    # Check the main dividend tracking file
    dividend_file = r"c:\Users\mjmat\Python Code in VS\dividend_tracker\DividendTrackerApp\outputs\Dividends_2025.xlsx"
    
    if os.path.exists(dividend_file):
        print(f"\n📊 DIVIDEND TRACKING FILE: {dividend_file}")
        wb = openpyxl.load_workbook(dividend_file, data_only=True)
        
        for sheet_name in wb.sheetnames:
            print(f"\n📋 Sheet: {sheet_name}")
            ws = wb[sheet_name]
            
            # Get headers
            headers = []
            for col in range(1, min(ws.max_column + 1, 30)):  # Check first 30 columns
                cell_value = ws.cell(row=1, column=col).value
                if cell_value:
                    headers.append(str(cell_value))
                else:
                    break
            
            print(f"   📊 Dimensions: {ws.max_row} rows × {ws.max_column} columns")
            print(f"   📋 Headers: {headers}")
            
            # If this is Ticker Analysis 2025, show some sample data
            if 'ticker analysis' in sheet_name.lower():
                print(f"   📈 Sample data from first few rows:")
                for row in range(2, min(ws.max_row + 1, 5)):
                    ticker = ws.cell(row=row, column=1).value
                    if ticker:
                        sample_data = []
                        for col in range(1, min(len(headers) + 1, 8)):  # First 8 columns
                            cell_value = ws.cell(row=row, column=col).value
                            sample_data.append(str(cell_value) if cell_value is not None else "")
                        print(f"      Row {row}: {sample_data}")
        
        wb.close()
    
    # Check the trade tracking file  
    trade_file = "Bryan Perry Transactions.xlsx"
    
    if os.path.exists(trade_file):
        print(f"\n📊 TRADE TRACKING FILE: {trade_file}")
        try:
            wb = openpyxl.load_workbook(trade_file, data_only=True)
            
            for sheet_name in wb.sheetnames:
                print(f"\n📋 Sheet: {sheet_name}")
                ws = wb[sheet_name]
                
                # Get headers
                headers = []
                for col in range(1, min(ws.max_column + 1, 25)):  # Check first 25 columns
                    cell_value = ws.cell(row=1, column=col).value
                    if cell_value:
                        headers.append(str(cell_value))
                    else:
                        break
                
                print(f"   📊 Dimensions: {ws.max_row} rows × {ws.max_column} columns")
                print(f"   📋 Headers: {headers}")
                
                # Show sample data for Open_Trades_2025
                if 'open_trades' in sheet_name.lower():
                    print(f"   📈 Sample data from first few rows:")
                    for row in range(2, min(ws.max_row + 1, 4)):
                        ticker = None
                        sample_data = []
                        for col in range(1, min(len(headers) + 1, 8)):  # First 8 columns
                            cell_value = ws.cell(row=row, column=col).value
                            if col == 1:  # Assuming first column might be ticker-related
                                ticker = cell_value
                            sample_data.append(str(cell_value) if cell_value is not None else "")
                        if ticker:
                            print(f"      Row {row}: {sample_data}")
            
            wb.close()
            
        except Exception as e:
            print(f"   ❌ Error reading trade file: {e}")
    
    # Check the historical yield data file
    hist_file = "Historical_Yield_Data_Ready_for_Integration.xlsx"
    
    if os.path.exists(hist_file):
        print(f"\n📊 HISTORICAL YIELD FILE: {hist_file}")
        try:
            wb = openpyxl.load_workbook(hist_file, data_only=True)
            ws = wb.active
            
            # Get headers
            headers = []
            for col in range(1, min(ws.max_column + 1, 15)):  # Check first 15 columns
                cell_value = ws.cell(row=1, column=col).value
                if cell_value:
                    headers.append(str(cell_value))
                else:
                    break
            
            print(f"   📊 Dimensions: {ws.max_row} rows × {ws.max_column} columns")
            print(f"   📋 Headers: {headers}")
            
            # Show sample tickers
            sample_tickers = []
            for row in range(2, min(ws.max_row + 1, 6)):
                ticker = ws.cell(row=row, column=1).value
                if ticker:
                    sample_tickers.append(str(ticker))
            
            if sample_tickers:
                print(f"   📈 Sample tickers: {sample_tickers}")
            
            wb.close()
            
        except Exception as e:
            print(f"   ❌ Error reading historical file: {e}")

if __name__ == "__main__":
    check_current_headers()
