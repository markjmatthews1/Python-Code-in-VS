#!/usr/bin/env python3
"""Fix row 26 specifically - simplified version"""

import os
from openpyxl import load_workbook
from datetime import datetime

def fix_row_26():
    """Fix the SOFI Sold Put in row 26"""
    excel_filename = "Bryan Perry Transactions.xlsx"
    sheet_name = "Open_Trades_2025"
    
    print("Checking and fixing row 26 (SOFI Sold Put)...")
    
    try:
        # Import E*TRADE auth
        import etrade_auth
        auth = etrade_auth.EtradeAuth()
        print("✅ Using stored OAuth token. In Etrade auth.")
        
        wb = load_workbook(excel_filename)
        ws = wb[sheet_name]
        header = [cell.value for cell in ws[1]]
        
        # Column positions
        ticker_col = header.index("Stock ticker") + 1
        type_col = header.index("Stock Purchase Type") + 1
        strike_col = header.index("Strike price") + 1
        expiry_col = header.index("Expiration Date") + 1
        cost_col = header.index("Cost") + 1
        shares_col = header.index("Number of shares +/-") + 1
        current_price_col = header.index("Current Price") + 1
        current_pl_col = header.index("Current P/L") + 1
        
        r = 26  # Row 26
        
        ticker = ws.cell(row=r, column=ticker_col).value
        trade_type = str(ws.cell(row=r, column=type_col).value).lower()
        strike = ws.cell(row=r, column=strike_col).value
        expiry = ws.cell(row=r, column=expiry_col).value
        cost = ws.cell(row=r, column=cost_col).value
        shares = ws.cell(row=r, column=shares_col).value
        current_price = ws.cell(row=r, column=current_price_col).value
        current_pl = ws.cell(row=r, column=current_pl_col).value
        
        print(f"Row 26 current data:")
        print(f"  Ticker: {ticker}")
        print(f"  Type: {trade_type}")
        print(f"  Strike: ${strike}")
        print(f"  Expiry: {expiry}")
        print(f"  Cost: ${cost}")
        print(f"  Shares: {shares}")
        print(f"  Current Price: {current_price}")
        print(f"  Current P/L: {current_pl}")
        
        # Handle date conversion
        if isinstance(expiry, datetime):
            expiry_str = expiry.strftime('%Y-%m-%d')
        else:
            # If it's a string like "9/19/2025", convert it
            try:
                expiry_dt = datetime.strptime(str(expiry), '%m/%d/%Y')
                expiry_str = expiry_dt.strftime('%Y-%m-%d')
            except:
                # Try other formats
                try:
                    expiry_dt = datetime.strptime(str(expiry)[:10], '%Y-%m-%d')
                    expiry_str = expiry_dt.strftime('%Y-%m-%d')
                except:
                    print(f"✗ Could not parse expiry date: {expiry}")
                    return
        
        print(f"  Converted expiry to: {expiry_str}")
        
        # Fetch fresh option price for SOFI put
        if "put" in trade_type:
            print(f"\nFetching fresh SOFI put option price...")
            
            try:
                option_price = auth.get_option_price_from_etrade(ticker, "put", strike, expiry_str)
                print(f"✓ Fresh option price: ${option_price}")
                
                # Force update the Excel cells
                ws.cell(row=r, column=current_price_col, value=option_price)
                ws.cell(row=r, column=current_price_col).number_format = '$#,##0.00'
                
                # Calculate and update Current P/L
                if cost and shares:
                    cost_val = float(cost)
                    shares_val = float(shares)
                    pl_val = (option_price - cost_val) * shares_val
                    
                    ws.cell(row=r, column=current_pl_col, value=pl_val)
                    ws.cell(row=r, column=current_pl_col).number_format = '$#,##0.00'
                    
                    print(f"✓ Updated: Price=${option_price:,.2f}, P/L=${pl_val:,.2f}")
                    print(f"  Calculation: ({option_price} - {cost_val}) × {shares_val} = {pl_val}")
                    
                    wb.save(excel_filename)
                    print(f"✅ Row 26 updated and saved to Excel!")
                else:
                    print(f"✗ Cannot calculate P/L: missing cost or shares")
                    
            except Exception as e:
                print(f"✗ Error fetching option price: {e}")
                import traceback
                traceback.print_exc()
        
        wb.close()
        
    except Exception as e:
        print(f"Error: {e}")
        import traceback
        traceback.print_exc()

if __name__ == "__main__":
    fix_row_26()
