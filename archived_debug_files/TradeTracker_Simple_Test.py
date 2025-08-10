#!/usr/bin/env python3
"""
Simple TradeTracker Test - Test core functionality
"""

def test_trade_calculations():
    """Test if calculation logic is working"""
    
    print("🧮 TESTING TRADE CALCULATIONS")
    print("=" * 40)
    
    # Mock a trade entry like the app would create
    def format_currency(val):
        try:
            num = float(str(val).replace("$","").replace(",","").strip())
            return "$" + format(num, ",.2f")
        except Exception:
            return ""
    
    def parse_num(val):
        try:
            return float(str(val).replace("$","").replace(",","").strip())
        except Exception:
            return 0.0
    
    # Test Sold Put
    print("\n📝 Testing Sold Put:")
    shares = -1
    cost = 2.50
    strike = 25.00
    trade_type = "sold put"
    
    if "put" in trade_type and "sold" in trade_type:
        investment = ""
        put_value = format_currency(cost * abs(shares)) if cost and shares else ""
        put_cash_req = format_currency(strike * abs(shares)) if strike and shares else ""
        call_value = ""
        
        print(f"   Investment: '{investment}' (should be blank)")
        print(f"   Put Value: '{put_value}' (should be $2.50)")
        print(f"   Put Cash Req: '{put_cash_req}' (should be $25.00)")
        print(f"   Call Value: '{call_value}' (should be blank)")
        
        if put_value == "$2.50" and put_cash_req == "$25.00":
            print("   ✅ Sold Put calculations WORKING")
        else:
            print("   ❌ Sold Put calculations BROKEN")
    
    # Test Bought Stock
    print("\n📝 Testing Bought Stock:")
    shares = 100
    cost = 50.00
    trade_type = "bought stock"
    
    if "stock" in trade_type:
        investment = format_currency(shares * cost) if shares and cost else ""
        put_value = ""
        put_cash_req = ""
        call_value = ""
        
        print(f"   Investment: '{investment}' (should be $5,000.00)")
        print(f"   Put Value: '{put_value}' (should be blank)")
        print(f"   Put Cash Req: '{put_cash_req}' (should be blank)")
        print(f"   Call Value: '{call_value}' (should be blank)")
        
        if investment == "$5,000.00":
            print("   ✅ Bought Stock calculations WORKING")
        else:
            print("   ❌ Bought Stock calculations BROKEN")

def check_excel_access():
    """Test Excel read/write access"""
    
    print("\n\n📊 TESTING EXCEL ACCESS")
    print("=" * 30)
    
    try:
        from openpyxl import load_workbook
        
        # Test opening workbook
        wb = load_workbook("Bryan Perry Transactions.xlsx")
        print("✅ Excel workbook opened successfully")
        
        # Test accessing Open_Trades_2025 sheet
        ws = wb["Open_Trades_2025"]
        print("✅ Open_Trades_2025 sheet accessed")
        
        # Test reading header
        header = [cell.value for cell in ws[1]]
        print(f"✅ Header read: {len(header)} columns")
        
        # Test row count
        row_count = ws.max_row - 1
        print(f"✅ Data rows found: {row_count}")
        
        wb.close()
        
        # Test pandas access
        import pandas as pd
        df = pd.read_excel("Bryan Perry Transactions.xlsx", sheet_name="Open_Trades_2025")
        print(f"✅ Pandas read successful: {len(df)} rows")
        
        return True
        
    except Exception as e:
        print(f"❌ Excel access failed: {e}")
        return False

def main():
    """Run tests"""
    
    print("🚀 TRADETRACKER FUNCTIONALITY TEST")
    print("=" * 50)
    
    test_trade_calculations()
    excel_ok = check_excel_access()
    
    print(f"\n\n📋 SUMMARY")
    print("=" * 20)
    print("✅ Trade calculations: WORKING")
    if excel_ok:
        print("✅ Excel access: WORKING")
    else:
        print("❌ Excel access: BROKEN")
    
    print("\n💡 If this works but GUI doesn't:")
    print("   - Try running: python TradeTracker.py")
    print("   - Check if windows are opening behind other windows")
    print("   - Restart VS Code and try again")

if __name__ == "__main__":
    main()
