#!/usr/bin/env python3
"""Fix row 26 specifically using direct E*TRADE calls"""

import os
from openpyxl import load_workbook
from datetime import datetime

def get_option_price_from_etrade(ticker, strike, option_type, expiry_date):
    """Fetch current option price from E*TRADE (copied from TradeTracker)"""
    try:
        from etrade_auth import get_etrade_session
        from datetime import datetime
        
        session, base_url = get_etrade_session()
        
        # Parse expiry date if it's a string
        target_date = None
        if isinstance(expiry_date, str):
            try:
                target_date = datetime.strptime(expiry_date, "%m/%d/%Y").date()
            except:
                try:
                    target_date = datetime.strptime(expiry_date, "%Y-%m-%d").date()
                except:
                    print(f"Could not parse expiry date: {expiry_date}")
                    return None
        elif hasattr(expiry_date, 'date'):
            target_date = expiry_date.date()
        else:
            target_date = expiry_date
            
        if not target_date:
            print(f"Invalid expiry date format: {expiry_date}")
            return None
        
        # Get available expiration dates to find the closest match
        exp_url = f"{base_url}/v1/market/optionexpiredate.json?symbol={ticker}"
        exp_response = session.get(exp_url)
        
        if exp_response.status_code != 200:
            print(f"Could not get expiration dates for {ticker}")
            return None
            
        exp_data = exp_response.json()
        dates = exp_data.get("OptionExpireDateResponse", {}).get("ExpirationDate", [])
        
        # Find the best matching expiration date
        best_match = None
        min_diff = float('inf')
        
        for date_info in dates:
            exp_datetime = datetime(date_info['year'], date_info['month'], date_info['day']).date()
            diff = abs((exp_datetime - target_date).days)
            if diff < min_diff:
                min_diff = diff
                best_match = date_info
        
        if not best_match:
            print(f"No expiration dates found for {ticker}")
            return None
            
        # Format URL for option chain with the best matching date
        url = f"{base_url}/v1/market/optionchains.json?symbol={ticker}&expiryDay={best_match['day']:02d}&expiryMonth={best_match['month']:02d}&expiryYear={best_match['year']}&chainType={option_type.upper()}"
        response = session.get(url)
        
        if response.status_code == 200:
            data = response.json()
            chain = data.get("OptionChainResponse", {})
            option_pairs = chain.get("OptionPair", [])
            
            # Find option with closest strike price
            best_option = None
            min_strike_diff = float('inf')
            
            for pair in option_pairs:
                if option_type.lower() == "call":
                    option_data = pair.get("Call", {})
                else:
                    option_data = pair.get("Put", {})
                
                if option_data:
                    option_strike = float(option_data.get("strikePrice", 0))
                    strike_diff = abs(option_strike - float(strike))
                    if strike_diff < min_strike_diff:
                        min_strike_diff = strike_diff
                        best_option = option_data
            
            if best_option:
                last_price = best_option.get("lastPrice")
                if last_price:
                    print(f"Found {option_type} option for {ticker}: Strike ${best_option.get('strikePrice')}, Price ${last_price}")
                    return float(last_price)
                    
    except Exception as e:
        print(f"Error fetching option price for {ticker} {strike} {option_type}: {e}")
    return None

def fix_row_26():
    """Fix the SOFI Sold Put in row 26"""
    excel_filename = "Bryan Perry Transactions.xlsx"
    sheet_name = "Open_Trades_2025"
    
    print("Checking and fixing row 26 (SOFI Sold Put)...")
    
    try:
        wb = load_workbook(excel_filename)
        ws = wb[sheet_name]
        header = [cell.value for cell in ws[1]]
        
        # Column positions
        ticker_col = header.index("Stock ticker") + 1
        type_col = header.index("Stock Purchase Type") + 1
        strike_col = header.index("Strike price") + 1
        expiry_col = header.index("Expiration Date") + 1
        cost_col = header.index("Cost") + 1
        shares_col = header.index("Number of shares +/-") + 1
        current_price_col = header.index("Current Price") + 1
        current_pl_col = header.index("Current P/L") + 1
        
        r = 26  # Row 26
        
        ticker = ws.cell(row=r, column=ticker_col).value
        trade_type = str(ws.cell(row=r, column=type_col).value).lower()
        strike = ws.cell(row=r, column=strike_col).value
        expiry = ws.cell(row=r, column=expiry_col).value
        cost = ws.cell(row=r, column=cost_col).value
        shares = ws.cell(row=r, column=shares_col).value
        current_price = ws.cell(row=r, column=current_price_col).value
        current_pl = ws.cell(row=r, column=current_pl_col).value
        
        print(f"Row 26 current data:")
        print(f"  Ticker: {ticker}")
        print(f"  Type: {trade_type}")
        print(f"  Strike: ${strike}")
        print(f"  Expiry: {expiry}")
        print(f"  Cost: ${cost}")
        print(f"  Shares: {shares}")
        print(f"  Current Price: {current_price}")
        print(f"  Current P/L: {current_pl}")
        
        # Handle date conversion for E*TRADE API
        if isinstance(expiry, datetime):
            expiry_str = expiry.strftime('%Y-%m-%d')
        else:
            # If it's a string like "9/19/2025", convert it
            try:
                expiry_dt = datetime.strptime(str(expiry), '%m/%d/%Y')
                expiry_str = expiry_dt.strftime('%Y-%m-%d')
            except:
                # Try other formats
                try:
                    expiry_dt = datetime.strptime(str(expiry)[:10], '%Y-%m-%d')
                    expiry_str = expiry_dt.strftime('%Y-%m-%d')
                except:
                    print(f"✗ Could not parse expiry date: {expiry}")
                    return
        
        print(f"  Converted expiry to: {expiry_str}")
        
        # Fetch fresh option price for SOFI put
        if "put" in trade_type:
            print(f"\nFetching fresh SOFI put option price...")
            
            try:
                option_price = get_option_price_from_etrade(ticker, strike, "put", expiry_str)
                if option_price:
                    print(f"✓ Fresh option price: ${option_price}")
                    
                    # Force update the Excel cells
                    ws.cell(row=r, column=current_price_col, value=option_price)
                    ws.cell(row=r, column=current_price_col).number_format = '$#,##0.00'
                    
                    # Calculate and update Current P/L
                    if cost and shares:
                        cost_val = float(cost)
                        shares_val = float(shares)
                        pl_val = (option_price - cost_val) * shares_val
                        
                        ws.cell(row=r, column=current_pl_col, value=pl_val)
                        ws.cell(row=r, column=current_pl_col).number_format = '$#,##0.00'
                        
                        print(f"✓ Updated: Price=${option_price:,.2f}, P/L=${pl_val:,.2f}")
                        print(f"  Calculation: ({option_price} - {cost_val}) × {shares_val} = {pl_val}")
                        
                        wb.save(excel_filename)
                        print(f"✅ Row 26 updated and saved to Excel!")
                    else:
                        print(f"✗ Cannot calculate P/L: missing cost or shares")
                else:
                    print(f"✗ Could not fetch option price")
                    
            except Exception as e:
                print(f"✗ Error fetching option price: {e}")
                import traceback
                traceback.print_exc()
        
        wb.close()
        
    except Exception as e:
        print(f"Error: {e}")
        import traceback
        traceback.print_exc()

if __name__ == "__main__":
    fix_row_26()
