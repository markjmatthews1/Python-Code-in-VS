#!/usr/bin/env python3
"""Direct Excel P/L fix without GUI"""

import os
from openpyxl import load_workbook

def direct_pl_fix():
    """Fix P/L calculations directly in Excel"""
    excel_filename = "Bryan Perry Transactions.xlsx"
    sheet_name = "Open_Trades_2025"
    
    print(f"Opening {excel_filename}...")
    
    if not os.path.exists(excel_filename):
        print(f"Error: {excel_filename} not found")
        return
        
    try:
        wb = load_workbook(excel_filename)
        ws = wb[sheet_name]
        header = [cell.value for cell in ws[1]]
        
        print(f"Excel headers: {header}")
        
        current_price_col = header.index("Current Price") + 1
        current_pl_col = header.index("Current P/L") + 1
        cost_col = header.index("Cost") + 1
        shares_col = header.index("Number of shares +/-") + 1
        ticker_col = header.index("Stock ticker") + 1
        
        print(f"Column positions: Price={current_price_col}, P/L={current_pl_col}, Cost={cost_col}, Shares={shares_col}, Ticker={ticker_col}")
        print(f"Total rows: {ws.max_row}")
        
        fixed_count = 0
        
        for r in range(2, ws.max_row + 1):
            current_price = ws.cell(row=r, column=current_price_col).value
            current_pl = ws.cell(row=r, column=current_pl_col).value
            cost = ws.cell(row=r, column=cost_col).value
            shares = ws.cell(row=r, column=shares_col).value
            ticker = ws.cell(row=r, column=ticker_col).value
            
            print(f"Row {r}: {ticker} - Price: {current_price}, P/L: {current_pl}, Cost: {cost}, Shares: {shares}")
            
            # Check specifically for rows 23-25 (the problematic ones)
            if r in [23, 24, 25]:
                print(f"  *** PROBLEM ROW {r} ***")
            
            # If we have a current price and cost and shares, but missing or zero P/L
            if current_price and cost and shares and (not current_pl or current_pl == 0):
                try:
                    # Parse the current price (remove $ and commas if it's a string)
                    if isinstance(current_price, str):
                        price_val = float(current_price.replace("$", "").replace(",", ""))
                    else:
                        price_val = float(current_price)
                    
                    cost_val = float(cost)
                    shares_val = float(shares)
                    
                    # Calculate P/L
                    pl_val = (price_val - cost_val) * shares_val
                    
                    # Update the cell
                    ws.cell(row=r, column=current_pl_col, value=pl_val)
                    ws.cell(row=r, column=current_pl_col).number_format = '$#,##0.00'
                    
                    print(f"  ✓ Fixed P/L for {ticker}: ${pl_val:,.2f} = ({price_val} - {cost_val}) × {shares_val}")
                    fixed_count += 1
                    
                except Exception as e:
                    print(f"  ✗ Error fixing P/L for {ticker}: {e}")
            else:
                if not current_price:
                    print(f"  Skipping {ticker}: No current price")
                elif not cost:
                    print(f"  Skipping {ticker}: No cost")
                elif not shares:
                    print(f"  Skipping {ticker}: No shares")
                elif current_pl and current_pl != 0:
                    print(f"  Skipping {ticker}: P/L already calculated ({current_pl})")
        
        if fixed_count > 0:
            wb.save(excel_filename)
            print(f"✓ Fixed Current P/L for {fixed_count} trades and saved to Excel")
        else:
            print("No trades needed P/L fixes")
        
        wb.close()
        
    except Exception as e:
        print(f"Error: {e}")
        import traceback
        traceback.print_exc()

if __name__ == "__main__":
    direct_pl_fix()
