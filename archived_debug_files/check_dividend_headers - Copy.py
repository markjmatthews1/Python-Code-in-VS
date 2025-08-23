#!/usr/bin/env python3
"""
Simple Dividend Headers Check
Check current headers in dividend tracker Excel
"""

import openpyxl
import os

def check_dividend_headers():
    div_file = r'dividend_tracker\DividendTrackerApp\outputs\Dividends_2025.xlsx'
    
    if os.path.exists(div_file):
        print(f'✅ FOUND: {div_file}')
        wb = openpyxl.load_workbook(div_file)
        print(f'📊 Sheets: {wb.sheetnames}')
        
        for sheet_name in wb.sheetnames:
            print(f'\n📋 Sheet: {sheet_name}')
            ws = wb[sheet_name]
            
            headers = []
            for col in range(1, 21):  # Check first 20 columns
                header = ws.cell(row=1, column=col).value
                if header:
                    headers.append(str(header))
                else:
                    break
            
            print(f'   Headers ({len(headers)}): {headers}')
            
            # Show dimensions
            print(f'   Dimensions: {ws.max_row} rows × {ws.max_column} columns')
            
            # If this looks like ticker analysis, show sample data
            if any(keyword in sheet_name.lower() for keyword in ['ticker', 'analysis', '2025']):
                print(f'   📈 Sample data:')
                for row in range(2, min(ws.max_row + 1, 4)):
                    sample = []
                    for col in range(1, min(len(headers) + 1, 6)):
                        cell_value = ws.cell(row=row, column=col).value
                        sample.append(str(cell_value) if cell_value is not None else "")
                    if any(sample):
                        print(f'      Row {row}: {sample}')
        
        wb.close()
    else:
        print(f'❌ NOT FOUND: {div_file}')

if __name__ == "__main__":
    check_dividend_headers()
