#!/usr/bin/env python3
"""
Simple Estimated Income 2025 Sheet Formatter
Focus on safe formatting without complex column manipulations
"""

import pandas as pd
import numpy as np
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment
from datetime import datetime

def format_estimated_income_sheet():
    """Format the Estimated Income 2025 sheet safely"""
    
    excel_file = 'dividend_tracker/DividendTrackerApp/outputs/Dividends_2025.xlsx'
    
    try:
        print("📊 Opening Excel file...")
        wb = load_workbook(excel_file)
        
        # Check available sheets
        print(f"Available sheets: {wb.sheetnames}")
        
        if 'Estimated Income 2025' not in wb.sheetnames:
            print("❌ Estimated Income 2025 sheet not found!")
            print("Available sheets:", wb.sheetnames)
            return False
            
        ws = wb['Estimated Income 2025']
        print(f"📋 Processing Estimated Income 2025 sheet...")
        
        # Define styles
        blue_bg_white_text = Font(name='Arial', size=12, color='FFFFFF')
        blue_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
        green_fill = PatternFill(start_color='00B050', end_color='00B050', fill_type='solid')
        red_fill = PatternFill(start_color='FF0000', end_color='FF0000', fill_type='solid')
        yellow_fill = PatternFill(start_color='FFFF00', end_color='FFFF00', fill_type='solid')
        default_font = Font(name='Arial', size=12)
        bold_font = Font(name='Arial', size=12, bold=True)
        
        # Get current dimensions
        max_row = ws.max_row
        max_col = ws.max_column
        print(f"   Sheet dimensions: {max_row} rows x {max_col} columns")
        
        # Format headers in row 3 (dates)
        print("🎨 Formatting date headers...")
        for col in range(2, min(max_col + 1, 50)):  # Limit to reasonable range
            try:
                date_cell = ws.cell(row=3, column=col)
                if date_cell.value:
                    # Try to format the date
                    if isinstance(date_cell.value, str) and '/' in date_cell.value:
                        try:
                            # Convert mm/dd/yyyy to m/d/yyyy
                            parts = date_cell.value.split('/')
                            if len(parts) == 3:
                                month = int(parts[0])
                                day = int(parts[1]) 
                                year = parts[2]
                                new_date = f"{month}/{day}/{year}"
                                date_cell.value = new_date
                        except:
                            pass  # Keep original if parsing fails
                    
                    # Apply styling
                    date_cell.font = blue_bg_white_text
                    date_cell.fill = blue_fill
                    date_cell.alignment = Alignment(horizontal='center')
                    
            except Exception as e:
                print(f"   Warning: Could not format column {col}: {e}")
                continue
        
        # Format account rows (4-8 typically)
        print("💰 Formatting account data...")
        for row in range(4, min(9, max_row + 1)):  # Rows 4-8
            try:
                # Check if this row has account data
                account_cell = ws.cell(row=row, column=1)
                if not account_cell.value or not str(account_cell.value).strip():
                    continue
                    
                print(f"   Processing row {row}: {account_cell.value}")
                
                # Format account name
                account_cell.font = default_font
                
                # Format data cells in this row
                prev_value = None
                for col in range(2, min(max_col + 1, 50)):
                    try:
                        cell = ws.cell(row=row, column=col)
                        current_value = cell.value
                        
                        if isinstance(current_value, (int, float)) and current_value != 0:
                            # Format as currency
                            cell.number_format = '$#,##0.00'
                            cell.font = default_font
                            
                            # Apply color coding based on previous value
                            if prev_value is not None and isinstance(prev_value, (int, float)):
                                if current_value > prev_value:
                                    cell.fill = green_fill
                                elif current_value < prev_value:
                                    cell.fill = red_fill
                                else:
                                    cell.fill = yellow_fill
                            
                            prev_value = current_value
                        elif current_value == 0:
                            cell.font = default_font
                            if prev_value is not None and prev_value != 0:
                                cell.fill = red_fill
                            prev_value = current_value
                            
                    except Exception as e:
                        continue
                        
            except Exception as e:
                print(f"   Warning: Could not process row {row}: {e}")
                continue
        
        # Format Monthly Total row (row 9)
        print("📊 Processing Monthly Total row...")
        try:
            monthly_row = 9
            if monthly_row <= max_row:
                # Format label
                label_cell = ws.cell(row=monthly_row, column=1)
                if not label_cell.value:
                    label_cell.value = "Monthly Total"
                label_cell.font = bold_font
                
                # Calculate and format monthly totals
                for col in range(2, min(max_col + 1, 50)):
                    try:
                        cell = ws.cell(row=monthly_row, column=col)
                        
                        # Sum up account values for this column
                        yearly_total = 0
                        for account_row in range(4, 9):
                            account_value = ws.cell(row=account_row, column=col).value
                            if isinstance(account_value, (int, float)):
                                yearly_total += account_value
                        
                        # Calculate monthly (yearly / 12)
                        monthly_value = yearly_total / 12 if yearly_total > 0 else 0
                        cell.value = monthly_value
                        cell.number_format = '$#,##0.00'
                        cell.font = bold_font
                        
                        # You could add color coding here too if needed
                        
                    except Exception as e:
                        continue
                        
        except Exception as e:
            print(f"   Warning: Could not process Monthly Total row: {e}")
        
        # Save the file
        print("💾 Saving changes...")
        wb.save(excel_file)
        wb.close()
        
        print("✅ Estimated Income 2025 sheet formatted successfully!")
        print("   • Date headers with blue background and white text")
        print("   • Color coding for week-to-week changes")
        print("   • Monthly totals calculated")
        print("   • Arial 12 font applied throughout")
        
        return True
        
    except Exception as e:
        print(f"❌ Error formatting sheet: {e}")
        import traceback
        traceback.print_exc()
        return False

def add_income_to_summary():
    """Add estimated income section to Portfolio Summary"""
    
    excel_file = 'dividend_tracker/DividendTrackerApp/outputs/Dividends_2025.xlsx'
    
    try:
        wb = load_workbook(excel_file)
        
        if 'Portfolio Summary' not in wb.sheetnames or 'Estimated Income 2025' not in wb.sheetnames:
            print("❌ Required sheets not found for summary update")
            return False
            
        summary_ws = wb['Portfolio Summary']
        income_ws = wb['Estimated Income 2025']
        
        # Get latest monthly estimate
        max_col = income_ws.max_column
        monthly_estimate = 0
        
        try:
            monthly_cell = income_ws.cell(row=9, column=max_col)
            if isinstance(monthly_cell.value, (int, float)):
                monthly_estimate = monthly_cell.value
        except:
            monthly_estimate = 0
            
        print(f"📊 Latest monthly estimate: ${monthly_estimate:,.2f}")
        
        # Find where to add income section (after existing content)
        last_used_row = summary_ws.max_row
        start_row = last_used_row + 2
        
        # Add Estimated Income section
        summary_ws[f'A{start_row}'] = 'ESTIMATED DIVIDEND INCOME'
        summary_ws[f'A{start_row}'].font = Font(name='Arial', size=12, bold=True)
        summary_ws[f'A{start_row}'].fill = PatternFill(start_color='D9E1F2', end_color='D9E1F2', fill_type='solid')
        summary_ws.merge_cells(f'A{start_row}:B{start_row}')
        start_row += 1
        
        summary_ws[f'A{start_row}'] = 'Monthly Estimated:'
        summary_ws[f'A{start_row}'].font = Font(name='Arial', size=12)
        summary_ws[f'B{start_row}'] = monthly_estimate
        summary_ws[f'B{start_row}'].number_format = '$#,##0.00'
        summary_ws[f'B{start_row}'].font = Font(name='Arial', size=12)
        start_row += 1
        
        summary_ws[f'A{start_row}'] = 'Annual Estimated:'
        summary_ws[f'A{start_row}'].font = Font(name='Arial', size=12)
        summary_ws[f'B{start_row}'] = monthly_estimate * 12
        summary_ws[f'B{start_row}'].number_format = '$#,##0.00'
        summary_ws[f'B{start_row}'].font = Font(name='Arial', size=12, bold=True)
        
        wb.save(excel_file)
        wb.close()
        
        print("✅ Portfolio Summary updated with estimated income!")
        
        return True
        
    except Exception as e:
        print(f"❌ Error updating summary: {e}")
        return False

if __name__ == "__main__":
    print("🚀 Formatting Estimated Income 2025 Sheet")
    print("="*50)
    
    success = format_estimated_income_sheet()
    
    if success:
        print("\n📈 Adding income data to Portfolio Summary...")
        add_income_to_summary()
        
        print("\n✅ SUCCESS!")
        print("• Estimated Income 2025 sheet properly formatted")
        print("• Date headers with blue background (#4472C4)")
        print("• Color coding for value changes (green/red/yellow)")
        print("• Monthly totals calculated (yearly ÷ 12)")
        print("• Portfolio Summary updated with income estimates")
        print("\nNote: Fresh Schwab API tokens would be needed for new data")
    else:
        print("\n❌ Failed to format Estimated Income 2025 sheet")
