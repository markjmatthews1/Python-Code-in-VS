#!/usr/bin/env python3
"""
Calculate yearly estimated dividend income for Schwab Individual account
using ticker yield data from IRA account and current positions.
"""

import json
import sys
import os
from datetime import datetime
import openpyxl
from openpyxl.utils import column_index_from_string
import requests

def format_dividend_cell(ws, row, col, value):
    """Apply standard formatting to dividend income cells."""
    from openpyxl.styles import PatternFill, Font
    
    # Set the value
    ws.cell(row=row, column=col, value=value)
    
    # Apply Arial 12 font
    arial_font = Font(name='Arial', size=12)
    ws.cell(row=row, column=col).font = arial_font
    
    # Apply currency formatting
    ws.cell(row=row, column=col).number_format = '$#,##0.00'
    
    # Apply light red background fill
    light_red_fill = PatternFill(start_color="FF7C80", end_color="FF7C80", fill_type="solid")
    ws.cell(row=row, column=col).fill = light_red_fill

def load_ticker_yields():
    """Load ticker yield data from the IRA dividend data file."""
    try:
        with open('actual_ira_dividend_data_20250825.json', 'r') as f:
            data = json.load(f)
        
        ticker_yields = {}
        for symbol, info in data.get('all_current_holdings', {}).items():
            if info.get('has_dividend', False):
                ticker_yields[symbol] = {
                    'yield': info.get('yield', 0.0),
                    'dividend_per_share': info.get('dividend_per_share', 0.0),
                    'annual_dividend': info.get('dividend_per_share', 0.0) * 4  # Assume quarterly
                }
        
        # Add manual entries for tickers not in IRA but known to pay dividends
        # (This will be updated as we find Schwab-specific tickers)
        
        print(f"✅ Loaded yield data for {len(ticker_yields)} dividend-paying tickers")
        return ticker_yields
    except Exception as e:
        print(f"❌ Error loading ticker yields: {e}")
        return {}

def get_schwab_individual_positions():
    """Get positions from Schwab Individual account using the API."""
    try:
        # Load token from tokens.json
        with open("tokens.json", "r") as f:
            token_data = json.load(f)
        
        tokens = token_data.get("token_dictionary", {})
        access_token = tokens.get("access_token")
        
        if not access_token:
            print("❌ No access token found")
            return []
        
        print("✅ Using existing Schwab access token")
        
        headers = {
            'Authorization': f'Bearer {access_token}',
            'Accept': 'application/json'
        }
        
        # Get all accounts with positions included
        base_url = "https://api.schwabapi.com"
        accounts_url = f"{base_url}/trader/v1/accounts?fields=positions"
        
        response = requests.get(accounts_url, headers=headers)
        
        if response.status_code != 200:
            print(f"❌ Error getting accounts: {response.status_code} - {response.text}")
            return []
        
        accounts = response.json()
        print(f"📊 Found {len(accounts)} Schwab accounts")
        
        # Find the Individual account by account number (74501314) and get its positions
        individual_positions = None
        target_account_number = "74501314"
        
        for account in accounts:
            # Get account info from securitiesAccount
            securities_account = account.get('securitiesAccount', {})
            account_number = securities_account.get('accountNumber', '')
            
            if target_account_number == str(account_number):
                # Get positions from this account
                individual_positions = securities_account.get('positions', [])
                print(f"✅ Found Schwab Individual account: {account_number}")
                break
        
        if individual_positions is None:
            print(f"❌ Could not find Schwab Individual account {target_account_number}")
            # Show all accounts for debugging
            for account in accounts:
                securities_account = account.get('securitiesAccount', {})
                account_number = securities_account.get('accountNumber', 'Unknown')
                print(f"   Available: {account_number}")
            return []
        
        if not individual_positions:
            print("❌ No positions found in Schwab Individual account")
            return []
        
        print(f"📈 Found {len(individual_positions)} positions in Schwab Individual account")
        
        # Convert to our expected format
        formatted_positions = []
        for position in individual_positions:
            try:
                instrument = position.get('instrument', {})
                symbol = instrument.get('symbol', '').strip().upper()
                quantity = float(position.get('longQuantity', 0))
                market_value = float(position.get('marketValue', 0))
                
                if quantity > 0 and symbol:  # Only include actual holdings
                    formatted_positions.append({
                        'symbol': symbol,
                        'quantity': quantity,
                        'market_value': market_value
                    })
                    print(f"   📊 {symbol}: {quantity} shares = ${market_value:,.2f}")
            
            except (ValueError, KeyError) as e:
                print(f"⚠️  Error processing position: {e}")
                continue
        
        return formatted_positions
        
    except Exception as e:
        print(f"❌ Error getting Schwab positions: {e}")
        return []

def calculate_dividend_income(positions, ticker_yields):
    """Calculate annual dividend income from positions and yields."""
    total_income = 0.0
    breakdown = []
    
    print("\n💰 CALCULATING ANNUAL DIVIDEND INCOME:")
    print("=" * 60)
    
    for position in positions:
        symbol = position['symbol']
        quantity = position['quantity']
        market_value = position['market_value']
        
        if symbol in ticker_yields:
            yield_pct = ticker_yields[symbol]['yield']
            annual_income = market_value * (yield_pct / 100)
            total_income += annual_income
            
            print(f"📊 {symbol}: ${market_value:,.2f} × {yield_pct:.2f}% = ${annual_income:,.2f}")
            
            breakdown.append({
                'symbol': symbol,
                'market_value': market_value,
                'yield': yield_pct,
                'annual_income': annual_income
            })
        else:
            print(f"⚠️  {symbol}: No yield data available (${market_value:,.2f} market value)")
    
    print("=" * 60)
    print(f"🎯 TOTAL ANNUAL DIVIDEND INCOME: ${total_income:,.2f}")
    
    return total_income, breakdown

def update_excel_sheet(annual_income):
    """Update the Excel sheet with the calculated dividend income."""
    try:
        # Use the same Excel file as the IRA script
        excel_file = r'c:\Users\mjmat\Python Code in VS\dividend_tracker\DividendTrackerApp\outputs\Dividends_2025.xlsx'
        wb = openpyxl.load_workbook(excel_file)
        
        # Work with the "Estimated Income 2025" sheet
        sheet_name = "Estimated Income 2025"
        if sheet_name not in wb.sheetnames:
            print(f"❌ Sheet '{sheet_name}' not found in {excel_file}")
            return False
        
        ws = wb[sheet_name]
        
        # Find column AI (column 35) for 8/24/2025
        target_column = column_index_from_string('AI')  # Column 35
        target_row = 7      # Row 7 for Schwab Individual
        
        # Verify the date is correct
        date_cell = ws.cell(row=3, column=target_column).value
        if str(date_cell).strip() != "8/24/2025":
            print(f"⚠️  Expected 8/24/2025 in column AI, found: {date_cell}")
        
        # Update the cell with formatted value
        format_dividend_cell(ws, target_row, target_column, annual_income)
        
        # Save the workbook
        wb.save(excel_file)
        
        print(f"✅ Updated Estimated Income 2025 sheet:")
        print(f"   Date: 8/24/2025 (Column AI)")
        print(f"   Row 7 (Schwab Individual): ${annual_income:,.2f}")
        print(f"   Formatting: Arial 12, Currency ($#,##0.00), Light Red Background (#FF7C80)")
        
        return True
        
    except Exception as e:
        print(f"❌ Error updating Excel sheet: {e}")
        return False

def main():
    """Main function to calculate and update Schwab Individual dividend income."""
    print("💼 CALCULATING SCHWAB INDIVIDUAL ANNUAL DIVIDEND INCOME")
    print("=" * 60)
    
    # Load ticker yield data
    ticker_yields = load_ticker_yields()
    if not ticker_yields:
        print("❌ Could not load ticker yield data")
        return
    
    # Get Schwab Individual positions
    positions = get_schwab_individual_positions()
    if not positions:
        print("❌ No positions found in Schwab Individual account")
        print("ℹ️  This may be due to Schwab API authentication issues")
        return
    
    # Calculate dividend income
    annual_income, breakdown = calculate_dividend_income(positions, ticker_yields)
    
    if annual_income > 0:
        # Update Excel sheet
        if update_excel_sheet(annual_income):
            print("\n🎉 SUCCESS! Schwab Individual dividend income calculated and updated.")
        else:
            print("\n⚠️  Calculation complete but Excel update failed.")
            print(f"   Manual entry needed: ${annual_income:,.2f} in row 7, column AI")
    else:
        print("\n⚠️  No dividend income calculated - check positions and ticker data")

if __name__ == "__main__":
    main()
