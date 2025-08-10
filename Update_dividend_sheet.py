                                                       # ***** imports for Update Dividend Sheet *****

from etrade_auth import get_etrade_session
import pyetrade
import openpyxl
from openpyxl.styles import PatternFill
from datetime import date, datetime
import time
import sys
import tkinter as tk
from tkinter import messagebox, filedialog

                                                        # ***** Endimports for Update Dividend Sheet *****

# Call Etrade_auth.py to get the access token

session, base_url = get_etrade_session()

print("✅ Got E*TRADE session, starting update...")
# Now use `session` for authenticated requests, and `base_url` for endpoints.

# Function to update dividend sheet (modified to use existing authorization)
def update_dividend_sheet():
    """
    Updates the dividend Excel sheet with latest market data.
    Prompts user to select account and file path.
    Handles E*TRADE 401 errors by refreshing tokens and retrying once.
    """
        # Helper for account and file selection
    def select_account_and_file(account_options, default_path):
        selected = {"account": None, "file": default_path}

        def on_account_select():
            try:
                idx = account_listbox.curselection()[0]
                selected["account"] = list(account_options.values())[idx]
                for i in range(account_listbox.size()):
                    account_listbox.itemconfig(i, bg="#222244", fg="white")
                account_listbox.itemconfig(idx, bg="#4CAF50", fg="white")
            except IndexError:
                pass

        def browse_file():
            file = filedialog.askopenfilename(
                title="Select Excel File",
                filetypes=[("Excel files", "*.xlsx *.xls")]
            )
            if file:
                file_entry.delete(0, tk.END)
                file_entry.insert(0, file)
                selected["file"] = file

        def submit():
            if selected["account"] is None:
                messagebox.showerror("Error", "Please select an account.")
                return
            selected["file"] = file_entry.get()
            root.destroy()

        root = tk.Tk()
        root.title("Select Account and Excel File")
        root.geometry("500x400")
        root.configure(bg="#222244")

        tk.Label(root, text="Select E*TRADE Account:", font=("Arial", 14, "bold"), bg="#222244", fg="white").pack(pady=10)
        account_listbox = tk.Listbox(root, font=("Arial", 12), bg="#222244", fg="white", selectbackground="#4CAF50", selectforeground="white", height=6)
        for desc in account_options.keys():
            account_listbox.insert(tk.END, desc)
        account_listbox.pack(pady=10, fill="x", padx=40)
        account_listbox.bind("<<ListboxSelect>>", lambda e: on_account_select())

        tk.Label(root, text="Excel File Path:", font=("Arial", 14, "bold"), bg="#222244", fg="white").pack(pady=10)
        file_entry = tk.Entry(root, font=("Arial", 12), width=40, bg="#e0e0e0")
        file_entry.insert(0, default_path)
        file_entry.pack(pady=5)
        tk.Button(root, text="Browse...", font=("Arial", 12), command=browse_file, bg="#2196F3", fg="white").pack(pady=5)

        tk.Button(root, text="Submit", font=("Arial", 14, "bold"), bg="#4CAF50", fg="white", command=submit).pack(pady=20)

        root.mainloop()
        return selected["account"], selected["file"]

    # --- Robust E*TRADE accounts fetch with 401 handling ---
    def fetch_accounts_with_retry():
        session, base_url = get_etrade_session()
        url = f"{base_url}/v1/accounts/list.json"
        response = session.get(url)
        if response.status_code == 401:
            print("⚠️ 401 Unauthorized. Refreshing E*TRADE session and retrying...")
            session, base_url = get_etrade_session(force_new=True) if 'force_new' in get_etrade_session.__code__.co_varnames else get_etrade_session()
            response = session.get(url)
        accounts_data = response.json()
        print("E*TRADE accounts API response:", accounts_data)
        if 'AccountListResponse' in accounts_data and \
           'Accounts' in accounts_data['AccountListResponse'] and \
           'Account' in accounts_data['AccountListResponse']['Accounts']:
            account_options = {}
            for account in accounts_data['AccountListResponse']['Accounts']['Account']:
                account_options[account['accountDesc']] = account['accountIdKey']
            return account_options, session, base_url
        else:
            print("❌ Unexpected response format from E*TRADE. Check your OAuth token and API permissions.")
            return None, session, base_url

    # Get E*TRADE accounts robustly
    account_options, session, base_url = fetch_accounts_with_retry()
    if not account_options:
        print("Error retrieving accounts. Exiting.")
        return

    # Set the default Excel file path
    default_path = "dividend_stocks.xlsx"

    # Now call the GUI selection
    accountID, file_path = select_account_and_file(account_options, default_path)

    try:
        workbook = openpyxl.load_workbook(file_path)
        sheet = workbook.active

        # Insert a new column with today's date as the header
        today = date.today().strftime("%m-%d-%Y")
        sheet.insert_cols(17)
        sheet.cell(row=1, column=17).value = today

        # Iterate through rows and update with market quote data
        for index, row in enumerate(sheet.iter_rows(min_row=2, values_only=False), start=2):
            symbol = row[0].value
            if symbol is None or not isinstance(symbol, str) or symbol.strip() == "":
                break

            try:
                quote_url = f"{base_url}/v1/market/quote/{symbol}.json"
                quote_response = session.get(quote_url)
                quote_json = quote_response.json()
                if 'QuoteResponse' in quote_json and 'QuoteData' in quote_json['QuoteResponse'] and isinstance(quote_json['QuoteResponse']['QuoteData'], list):
                    quote_data = quote_json['QuoteResponse']['QuoteData'][0]
                    dividend_payable_date_str = quote_data['All'].get('dividendPayableDate', '')
                    if dividend_payable_date_str:
                        try:
                            dividend_payable_date_secs = datetime.fromtimestamp(int(dividend_payable_date_str))
                            dividend_payable_date = dividend_payable_date_secs
                            sheet.cell(row=index, column=11).value = dividend_payable_date
                            sheet.cell(row=index, column=11).number_format = 'mm-dd'
                        except ValueError as e:
                            print(f"   Error converting to date: {e}")
                            sheet.cell(row=index, column=11).value = ''
                    else:
                        sheet.cell(row=index, column=11).value = ''

                    last_trade = quote_data['All'].get('lastTrade', '')
                    if last_trade:
                        sheet.cell(row=index, column=4).value = round(float(last_trade), 2)
                    else:
                        sheet.cell(row=index, column=4).value = last_trade

                    current_yield = quote_data['All'].get('yield', '')
                    if current_yield:
                        sheet.cell(row=index, column=17).value = round(float(current_yield), 2)
                    else:
                        sheet.cell(row=index, column=17).value = current_yield

                    col_q_value = sheet.cell(row=index, column=17).value
                    col_p_value = sheet.cell(row=index, column=16).value
                    try:
                        col_q_value = float(col_q_value)
                        col_p_value = float(col_p_value)
                        if col_q_value < col_p_value:
                            fill_color = "FF0000"
                        elif col_q_value > col_p_value:
                            fill_color = "00FF00"
                        else:
                            fill_color = "FFFF00"
                        fill = PatternFill(start_color=fill_color, end_color=fill_color, fill_type="solid")
                        sheet.cell(row=index, column=17).fill = fill
                    except (TypeError, ValueError):
                        print(f"Skipping row {index} due to invalid data types for comparison.")
                        time.sleep(5)
                        break

            except FileNotFoundError:
                print(f"File not found at: {file_path}")

        workbook.save(file_path)
        print(f"File processed and saved at: {file_path}")
    except Exception as e:
        print(f"An error occurred: {e}")
        try:
            workbook.save(file_path)
            print(f"File processed and saved at: {file_path}")
        except Exception:
            pass

# At the very end of your script
if __name__ == "__main__":
    update_dividend_sheet()
    print("✅ Update complete. Exiting.")
    
