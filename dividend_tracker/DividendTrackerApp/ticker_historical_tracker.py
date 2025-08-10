#!/usr/bin/env python3
"""
Enhanced Ticker Tracker with Historical Yield Tracking
Maintains historical yield data by date columns (like your current sheet)
"""

import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment
from openpyxl.styles.numbers import FORMAT_CURRENCY_USD, FORMAT_PERCENTAGE_00
import pandas as pd
from datetime import datetime, date, timedelta
import os
import sys

def create_ticker_tracker_with_history():
    """
    Create ticker tracker that maintains historical yield data
    """
    
    print("🚀 Creating Ticker Tracker with Historical Yield Tracking...")
    print("=" * 60)
    
    # File paths
    data_file = os.path.join("data", "dividend_stocks.xlsx")
    output_file = os.path.join("outputs", "Dividends_2025.xlsx")
    
    # Load existing data and structure
    ticker_data, historical_yields = load_existing_ticker_data(data_file)
    
    # Get current live data
    current_data = get_current_ticker_data()
    
    # Create enhanced sheet with historical yield tracking
    create_historical_yield_sheet(ticker_data, historical_yields, current_data, output_file)
    
    print("✅ Historical yield tracking created!")

def load_existing_ticker_data(data_file):
    """
    Load existing ticker data including historical yield columns
    """
    print("📊 Loading existing ticker data with historical yields...")
    
    ticker_data = {}
    historical_yields = {}
    
    if os.path.exists(data_file):
        try:
            # Read the Excel file to get both ticker data and historical yields
            df = pd.read_excel(data_file)
            
            # Process each row
            for idx, row in df.iterrows():
                if pd.notna(row.get('Ticker')):
                    ticker = str(row['Ticker']).strip()
                    
                    # Basic ticker information
                    ticker_data[ticker] = {
                        'quantity': float(row.get('Qty #', 0)),
                        'price_paid': float(row.get('Price Paid $', 0)),
                        'payment_cycle': str(row.get('Payment cycle', 'Monthly')),
                        'rate_per_share': float(row.get('Rate per share', 0)),
                        'account': 'E*TRADE' if 'Etrade' in str(row.get('Account', '')) else 'Schwab'
                    }
                    
                    # Extract historical yield data from date columns
                    historical_yields[ticker] = {}
                    
                    # Look for date columns (your format: 07-19-2025, 07-12-2025, etc.)
                    for col in df.columns:
                        col_str = str(col)
                        if '-2025' in col_str and col_str != 'Ticker':
                            try:
                                yield_value = row.get(col)
                                if pd.notna(yield_value) and isinstance(yield_value, (int, float)):
                                    historical_yields[ticker][col_str] = float(yield_value)
                            except:
                                pass
            
            print(f"   📈 Loaded {len(ticker_data)} tickers")
            print(f"   📅 Found historical data for {len(historical_yields)} tickers")
            
            # Show sample of historical dates found
            if historical_yields:
                sample_ticker = list(historical_yields.keys())[0]
                sample_dates = list(historical_yields[sample_ticker].keys())
                print(f"   📊 Sample dates: {sample_dates[:5]}...")
            
        except Exception as e:
            print(f"   ❌ Error loading data: {e}")
            
    else:
        print(f"   ⚠️ {data_file} not found - creating sample structure")
        # Create sample data structure
        ticker_data = create_sample_ticker_data()
        historical_yields = create_sample_historical_data()
    
    return ticker_data, historical_yields

def create_sample_ticker_data():
    """Create sample ticker data structure"""
    return {
        'ABR': {'quantity': 1050, 'price_paid': 13.88, 'payment_cycle': 'Quarterly', 'rate_per_share': 0.10, 'account': 'E*TRADE'},
        'ACP': {'quantity': 1625, 'price_paid': 6.64, 'payment_cycle': 'Monthly', 'rate_per_share': 0.08, 'account': 'E*TRADE'},
        'AGNC': {'quantity': 736, 'price_paid': 10.37, 'payment_cycle': 'Monthly', 'rate_per_share': 0.12, 'account': 'E*TRADE'},
        'DX': {'quantity': 125, 'price_paid': 11.91, 'payment_cycle': 'Monthly', 'rate_per_share': 0.17, 'account': 'Schwab'},
        'ECC': {'quantity': 1330, 'price_paid': 7.43, 'payment_cycle': 'Monthly', 'rate_per_share': 0.14, 'account': 'Schwab'}
    }

def create_sample_historical_data():
    """Create sample historical yield data"""
    # Generate weekly dates going back
    dates = []
    current_date = datetime.now()
    for i in range(20):  # 20 weeks of history
        week_date = current_date - timedelta(weeks=i)
        dates.append(week_date.strftime("%m-%d-%Y"))
    
    # Sample historical yields for each ticker
    historical_data = {}
    base_yields = {'ABR': 14.20, 'ACP': 15.53, 'AGNC': 13.81, 'DX': 14.62, 'ECC': 19.35}
    
    for ticker in base_yields:
        historical_data[ticker] = {}
        base_yield = base_yields[ticker]
        
        for i, date_str in enumerate(dates):
            # Simulate yield variations over time
            variation = (i * 0.1) + ((i % 3) * 0.05)  # Some fluctuation
            yield_value = base_yield + variation
            historical_data[ticker][date_str] = round(yield_value, 2)
    
    return historical_data

def get_current_ticker_data():
    """Get current market data for tickers"""
    print("🔌 Getting current ticker market data...")
    
    # Sample current data - this would be replaced with actual API calls
    current_data = {
        'ABR': {'current_price': 11.21, 'change': -0.12, 'change_pct': -0.87, 'current_yield': 14.20},
        'ACP': {'current_price': 5.94, 'change': -0.02, 'change_pct': -0.33, 'current_yield': 15.53},
        'AGNC': {'current_price': 9.25, 'change': -0.07, 'change_pct': -0.69, 'current_yield': 13.81},
        'DX': {'current_price': 12.53, 'change': -0.07, 'change_pct': -0.52, 'current_yield': 14.62},
        'ECC': {'current_price': 7.45, 'change': 0.03, 'change_pct': 0.34, 'current_yield': 19.35}
    }
    
    print(f"   📈 Retrieved current data for {len(current_data)} tickers")
    return current_data

def create_historical_yield_sheet(ticker_data, historical_yields, current_data, output_file):
    """
    Create sheet with historical yield tracking
    """
    print("📝 Creating Historical Yield Tracking sheet...")
    
    # Load workbook
    wb = openpyxl.load_workbook(output_file) if os.path.exists(output_file) else openpyxl.Workbook()
    
    # Create sheet
    sheet_name = "Ticker Historical Analysis"
    if sheet_name in wb.sheetnames:
        del wb[sheet_name]
    
    ws = wb.create_sheet(sheet_name)
    
    # Get all unique historical dates and sort them
    all_dates = set()
    for ticker_history in historical_yields.values():
        all_dates.update(ticker_history.keys())
    
    sorted_dates = sorted(list(all_dates), key=lambda x: datetime.strptime(x, "%m-%d-%Y"), reverse=True)
    
    # Add today's date as the first column
    today = datetime.now().strftime("%m-%d-%Y")
    if today not in sorted_dates:
        sorted_dates.insert(0, today)
    
    print(f"   📅 Tracking yields across {len(sorted_dates)} dates")
    
    # Define headers
    static_headers = [
        'Ticker', 'Account', 'Qty #', 'Price Paid $', 'Current Price $', 'Current Value $',
        'Original Value $', 'Total Gain $', 'Total Gain %', 'Payment Cycle', 
        'Rate/Share', 'Monthly Dividend', 'Annual Dividend'
    ]
    
    # Add date headers for yield tracking
    all_headers = static_headers + sorted_dates
    
    # Create headers
    for col, header in enumerate(all_headers, 1):
        cell = ws.cell(row=1, column=col)
        cell.value = header
        cell.font = Font(bold=True, color="FFFFFF")
        
        # Color code different sections
        if col <= len(static_headers):
            cell.fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")  # Blue for static data
        else:
            cell.fill = PatternFill(start_color="70AD47", end_color="70AD47", fill_type="solid")  # Green for dates
        
        cell.alignment = Alignment(horizontal="center")
    
    # Add data rows
    row = 2
    totals = {'current_value': 0, 'original_value': 0, 'monthly_dividend': 0, 'annual_dividend': 0}
    
    for ticker, data in ticker_data.items():
        # Get current market data
        current = current_data.get(ticker, {})
        current_price = current.get('current_price', data['price_paid'])
        current_yield = current.get('current_yield', 0)
        
        # Calculate values
        quantity = data['quantity']
        price_paid = data['price_paid']
        current_value = quantity * current_price
        original_value = quantity * price_paid
        total_gain = current_value - original_value
        total_gain_pct = (total_gain / original_value * 100) if original_value > 0 else 0
        
        # Calculate dividends
        rate_per_share = data['rate_per_share']
        payment_cycle = data['payment_cycle']
        
        if payment_cycle == 'Monthly':
            monthly_dividend = rate_per_share * quantity
            annual_dividend = monthly_dividend * 12
        elif payment_cycle == 'Quarterly':
            quarterly_dividend = rate_per_share * quantity
            monthly_dividend = quarterly_dividend / 3
            annual_dividend = quarterly_dividend * 4
        elif payment_cycle == 'Weekly':
            weekly_dividend = rate_per_share * quantity
            monthly_dividend = weekly_dividend * 4.33
            annual_dividend = weekly_dividend * 52
        else:
            monthly_dividend = rate_per_share * quantity / 12
            annual_dividend = rate_per_share * quantity
        
        # Static data
        static_data = [
            ticker,                    # Ticker
            data['account'],           # Account
            quantity,                  # Qty #
            price_paid,               # Price Paid $
            current_price,            # Current Price $
            current_value,            # Current Value $
            original_value,           # Original Value $
            total_gain,               # Total Gain $
            total_gain_pct,           # Total Gain %
            payment_cycle,            # Payment Cycle
            rate_per_share,           # Rate/Share
            monthly_dividend,         # Monthly Dividend
            annual_dividend           # Annual Dividend
        ]
        
        # Add static data
        for col, value in enumerate(static_data, 1):
            cell = ws.cell(row=row, column=col)
            cell.value = value
            
            # Apply formatting
            if col in [4, 5, 6, 7, 8, 11, 12, 13]:  # Currency columns
                if isinstance(value, (int, float)):
                    cell.number_format = '"$"#,##0.00'
            elif col in [9]:  # Percentage columns
                if isinstance(value, (int, float)):
                    cell.number_format = '0.00%'
        
        # Add historical yield data
        ticker_history = historical_yields.get(ticker, {})
        
        for date_idx, date_str in enumerate(sorted_dates):
            col = len(static_headers) + date_idx + 1
            
            if date_str == today:
                # Use current yield for today
                yield_value = current_yield
            else:
                # Use historical data
                yield_value = ticker_history.get(date_str, 0)
            
            cell = ws.cell(row=row, column=col)
            cell.value = yield_value
            cell.number_format = '0.00'
            
            # Color code yield changes
            if date_idx > 0:  # Compare with previous week
                prev_date = sorted_dates[date_idx - 1]
                if prev_date == today:
                    prev_yield = current_yield
                else:
                    prev_yield = ticker_history.get(prev_date, yield_value)
                
                if yield_value > prev_yield + 0.1:
                    cell.fill = PatternFill(start_color="00FF00", end_color="00FF00", fill_type="solid")  # Green - increase
                elif yield_value < prev_yield - 0.1:
                    cell.fill = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")  # Red - decrease
                elif abs(yield_value - prev_yield) <= 0.1:
                    cell.fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")  # Yellow - stable
        
        # Update totals
        totals['current_value'] += current_value
        totals['original_value'] += original_value
        totals['monthly_dividend'] += monthly_dividend
        totals['annual_dividend'] += annual_dividend
        
        row += 1
    
    # Add totals row
    totals_row = row + 1
    ws.cell(row=totals_row, column=1, value="TOTALS").font = Font(bold=True)
    
    # Add total values
    ws.cell(row=totals_row, column=6, value=totals['current_value']).font = Font(bold=True)
    ws.cell(row=totals_row, column=6).number_format = '"$"#,##0.00'
    
    ws.cell(row=totals_row, column=7, value=totals['original_value']).font = Font(bold=True)
    ws.cell(row=totals_row, column=7).number_format = '"$"#,##0.00'
    
    ws.cell(row=totals_row, column=8, value=totals['current_value'] - totals['original_value']).font = Font(bold=True)
    ws.cell(row=totals_row, column=8).number_format = '"$"#,##0.00'
    
    ws.cell(row=totals_row, column=12, value=totals['monthly_dividend']).font = Font(bold=True)
    ws.cell(row=totals_row, column=12).number_format = '"$"#,##0.00'
    
    ws.cell(row=totals_row, column=13, value=totals['annual_dividend']).font = Font(bold=True)
    ws.cell(row=totals_row, column=13).number_format = '"$"#,##0.00'
    
    # Auto-adjust column widths
    for col in range(1, len(all_headers) + 1):
        if col <= len(static_headers):
            ws.column_dimensions[openpyxl.utils.get_column_letter(col)].width = 15
        else:
            ws.column_dimensions[openpyxl.utils.get_column_letter(col)].width = 12
    
    # Add legend
    legend_row = row + 3
    ws.cell(row=legend_row, column=1, value="Yield Change Legend:").font = Font(bold=True)
    
    # Green cell
    green_cell = ws.cell(row=legend_row + 1, column=1, value="Yield Increased")
    green_cell.fill = PatternFill(start_color="00FF00", end_color="00FF00", fill_type="solid")
    
    # Red cell  
    red_cell = ws.cell(row=legend_row + 2, column=1, value="Yield Decreased")
    red_cell.fill = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")
    
    # Yellow cell
    yellow_cell = ws.cell(row=legend_row + 3, column=1, value="Yield Stable")
    yellow_cell.fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
    
    # Save workbook
    wb.save(output_file)
    
    print(f"✅ Historical yield tracking sheet created")
    print(f"   📊 {len(ticker_data)} tickers tracked")
    print(f"   📅 {len(sorted_dates)} weeks of yield history")
    print(f"   💰 Total Monthly Dividend: ${totals['monthly_dividend']:,.2f}")
    print(f"   💰 Total Annual Dividend: ${totals['annual_dividend']:,.2f}")

if __name__ == "__main__":
    create_ticker_tracker_with_history()
