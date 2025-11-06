r"""
Module: estimated_income_tracker.py
Author: Mark  
Created: July 25, 2025
Updated: September 1, 2025 - Integrated with comprehensive dividend income system
Purpose: Track comprehensive dividend income using ticker_yields.json and real API data
Location: C:\Python_Projects\DividendTrackerApp\modules\estimated_income_tracker.py
"""

import os
import sys
import traceback
from datetime import datetime

def main():
    """
    Main function to update estimated income using comprehensive dividend calculation
    
    This integrates with the enhanced portfolio updater to use:
    - ticker_yields.json for actual dividend yields
    - Real-time position data from E*TRADE and Schwab APIs  
    - Comprehensive dividend calculations for all accounts
    """
    try:
        print("📊 Running estimated income tracker...")
        print("💡 Portfolio values already updated by enhanced updater")
        print(">> Using comprehensive dividend income calculation with ticker_yields.json")
        
        # Add the parent directory to sys.path to import enhanced_portfolio_updater_with_schwab
        parent_dir = os.path.dirname(os.path.dirname(__file__))
        sys.path.insert(0, parent_dir)
        
        try:
            # Import the enhanced portfolio updater with comprehensive dividend calculations
            from enhanced_portfolio_updater_with_schwab import EnhancedPortfolioUpdater
            
            # Create updater instance
            updater = EnhancedPortfolioUpdater()
            
            # Run comprehensive dividend income calculation
            print("🔄 Running comprehensive estimated income calculation...")
            success = updater.update_estimated_income_2025_comprehensive()
            
            if success:
                print("✅ Comprehensive estimated income calculation completed successfully!")
                return os.path.join(parent_dir, "outputs", "Dividends_2025.xlsx")
            else:
                print("❌ Comprehensive estimated income calculation failed")
                # Fall back to basic update
                return fallback_estimated_income_update()
                
        except ImportError as e:
            print(f"⚠️ Import error with enhanced updater: {e}")
            print("📄 Falling back to basic estimated income update...")
            return fallback_estimated_income_update()
            
    except Exception as e:
        print(f"❌ Error in estimated income tracker: {e}")
        print(f"Traceback: {traceback.format_exc()}")
        return fallback_estimated_income_update()

def fallback_estimated_income_update():
    """
    Fallback function for basic estimated income update
    """
    try:
        print("🔄 Running fallback estimated income update...")
        
        # Basic implementation - just ensure the sheet exists
        target_file = os.path.join(os.path.dirname(__file__), "..", "outputs", "Dividends_2025.xlsx")
        
        if os.path.exists(target_file):
            import openpyxl
            
            workbook = openpyxl.load_workbook(target_file)
            
            # Ensure Estimated Income 2025 sheet exists
            if 'Estimated Income 2025' not in workbook.sheetnames:
                print("🏗️ Creating basic Estimated Income 2025 sheet...")
                sheet = workbook.create_sheet("Estimated Income 2025")
                
                # Basic headers
                sheet['A1'] = '📊 ESTIMATED DIVIDEND INCOME 2025'
                sheet['A2'] = 'Updated:'
                sheet['B2'] = datetime.now().strftime("%m/%d/%Y %H:%M")
                sheet['A4'] = 'Ticker'
                sheet['B4'] = 'Account'
                sheet['C4'] = 'Estimated Monthly Dividend'
                
                # Save the workbook
                workbook.save(target_file)
                print("✅ Basic Estimated Income 2025 sheet created")
            else:
                # Update timestamp
                sheet = workbook['Estimated Income 2025']
                sheet['B2'] = datetime.now().strftime("%m/%d/%Y %H:%M")
                workbook.save(target_file)
                print("✅ Estimated Income 2025 sheet timestamp updated")
            
            workbook.close()
            return target_file
        else:
            print(f"❌ Target file not found: {target_file}")
            return None
            
    except Exception as e:
        print(f"❌ Error in fallback update: {e}")
        return None

if __name__ == "__main__":
    result = main()
    if result:
        print(f"✅ Estimated income tracker updated: {result}")
    else:
        print(f"❌ Estimated income tracker failed")
