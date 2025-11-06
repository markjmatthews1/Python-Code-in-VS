#!/usr/bin/env python3
"""
Working Estimated Income Update - Using Real Position & Yield Data
==================================================================

This script updates the Estimated Income 2025 sheet with real data from:
- E*TRADE positions (43 positions with real quantities)
- E*TRADE dividend yields (25 dividend-paying stocks with real yields)
- Calculated estimated income: $34,245.19 annually ($2,853.77 monthly)

Author: AI Assistant
Date: September 6, 2025
"""
import sys
import os
sys.path.append(os.path.join(os.path.dirname(__file__), 'modules'))

import openpyxl
from datetime import datetime
import pandas as pd

def update_estimated_income_with_real_data():
    """Update Estimated Income 2025 sheet with real position and yield data"""
    print("🚀 UPDATING ESTIMATED INCOME 2025 WITH REAL DATA")
    print("=" * 60)
    
    try:
        # 1. Collect real position and yield data
        from portfolio_data_collector import PortfolioDataCollector
        collector = PortfolioDataCollector()
        
        print("📊 Step 1: Collecting real ticker yields...")
        ticker_yields = collector.collect_fresh_ticker_yields_from_etrade_ira()
        if not ticker_yields:
            print("❌ Failed to collect ticker yields")
            return False
        
        dividend_count = len([t for t in ticker_yields.values() if t.get('has_dividend', False)])
        print(f"✅ Found {len(ticker_yields)} tickers, {dividend_count} with dividends")
        
        print("🏦 Step 2: Collecting real position data...")
        positions = collector.get_etrade_positions_by_account()
        
        ira_positions = positions.get('etrade_ira', [])
        taxable_positions = positions.get('etrade_taxable', [])
        
        print(f"✅ E*TRADE IRA: {len(ira_positions)} positions")
        print(f"✅ E*TRADE Taxable: {len(taxable_positions)} positions")
        
        # 2. Calculate estimated income for each position
        print("💰 Step 3: Calculating estimated income...")
        
        estimated_income_data = []
        total_annual_income = 0.0
        all_positions = ira_positions + taxable_positions
        
        for position in all_positions:
            symbol = position.get('symbol', '')
            quantity = position.get('quantity', 0)
            market_value = position.get('market_value', 0)
            account_name = position.get('account_name', 'Unknown')
            
            # Determine account type for sheet
            if 'IRA' in account_name.upper():
                account_type = 'E*TRADE IRA'
            else:
                account_type = 'E*TRADE Taxable'
            
            # Get dividend data
            annual_dividend_income = 0.0
            dividend_yield = 0.0
            payment_frequency = 'quarterly'
            
            if symbol in ticker_yields:
                yield_data = ticker_yields[symbol]
                annual_dividend = yield_data.get('annual_dividend', 0)
                dividend_yield = yield_data.get('yield', 0)
                
                if annual_dividend > 0:
                    # Use actual annual dividend per share
                    annual_dividend_income = annual_dividend * quantity
                elif dividend_yield > 0:
                    # Use yield percentage * market value
                    annual_dividend_income = (dividend_yield / 100) * market_value
                    
                payment_frequency = yield_data.get('payment_frequency', 'quarterly')
            
            if annual_dividend_income > 0:
                total_annual_income += annual_dividend_income
                
                # Calculate monthly income
                monthly_dividend_income = annual_dividend_income / 12
                
                estimated_income_data.append({
                    'Symbol': symbol,
                    'Account': account_type,
                    'Quantity #': quantity,
                    'Dividend Yield %': dividend_yield,
                    'Annual Income $': annual_dividend_income,
                    'Monthly Income $': monthly_dividend_income,
                    'Payment Frequency': payment_frequency,
                    'Market Value $': market_value,
                    'Status': 'REAL_API_DATA',
                    'Last Updated': datetime.now().strftime('%Y-%m-%d %H:%M:%S')
                })
                
                print(f"   💰 {symbol}: {quantity} shares × {dividend_yield:.2f}% = ${annual_dividend_income:,.2f}/year")
        
        monthly_total = total_annual_income / 12
        print(f"\n🎯 TOTALS:")
        print(f"📈 Annual estimated income: ${total_annual_income:,.2f}")
        print(f"📅 Monthly estimated income: ${monthly_total:,.2f}")
        print(f"🎊 Dividend positions: {len(estimated_income_data)}")
        
        # 3. Update Excel sheet
        print("📝 Step 4: Updating Excel sheet...")
        
        excel_path = os.path.join(os.path.dirname(__file__), 'outputs', 'Dividends_2025.xlsx')
        if not os.path.exists(excel_path):
            print(f"❌ Excel file not found: {excel_path}")
            return False
        
        # Load workbook
        workbook = openpyxl.load_workbook(excel_path)
        
        if 'Estimated Income 2025' not in workbook.sheetnames:
            print("❌ 'Estimated Income 2025' sheet not found")
            return False
        
        sheet = workbook['Estimated Income 2025']
        
        # Clear existing data (keep headers)
        max_row = sheet.max_row
        if max_row > 1:
            sheet.delete_rows(2, max_row - 1)
        
        # Add new data
        for row_idx, data in enumerate(estimated_income_data, start=2):
            sheet[f'A{row_idx}'] = data['Symbol']
            sheet[f'B{row_idx}'] = data['Account'] 
            sheet[f'C{row_idx}'] = data['Quantity #']
            sheet[f'D{row_idx}'] = data['Dividend Yield %']
            sheet[f'E{row_idx}'] = data['Annual Income $']
            sheet[f'F{row_idx}'] = data['Monthly Income $']
            sheet[f'G{row_idx}'] = data['Payment Frequency']
            sheet[f'H{row_idx}'] = data['Market Value $']
            sheet[f'I{row_idx}'] = data['Status']
            sheet[f'J{row_idx}'] = data['Last Updated']
        
        # Add summary row
        summary_row = len(estimated_income_data) + 3
        sheet[f'A{summary_row}'] = 'TOTAL'
        sheet[f'E{summary_row}'] = total_annual_income
        sheet[f'F{summary_row}'] = monthly_total
        
        # Format numbers
        for row in range(2, len(estimated_income_data) + 2):
            sheet[f'C{row}'].number_format = '#,##0'  # Quantity
            sheet[f'D{row}'].number_format = '0.00%'  # Yield percentage
            sheet[f'E{row}'].number_format = '$#,##0.00'  # Annual income
            sheet[f'F{row}'].number_format = '$#,##0.00'  # Monthly income
            sheet[f'H{row}'].number_format = '$#,##0.00'  # Market value
        
        # Format totals
        sheet[f'E{summary_row}'].number_format = '$#,##0.00'
        sheet[f'F{summary_row}'].number_format = '$#,##0.00'
        
        # Add timestamp
        sheet[f'A1'] = f"Estimated Income 2025 - Updated {datetime.now().strftime('%Y-%m-%d %H:%M:%S')} with REAL API DATA"
        
        # Save workbook
        workbook.save(excel_path)
        workbook.close()
        
        print(f"✅ Excel sheet updated successfully!")
        print(f"📁 File: {excel_path}")
        print(f"📊 Added {len(estimated_income_data)} dividend positions")
        print(f"💰 Total estimated annual income: ${total_annual_income:,.2f}")
        
        return True
        
    except Exception as e:
        print(f"❌ Error updating estimated income: {e}")
        import traceback
        traceback.print_exc()
        return False

if __name__ == "__main__":
    success = update_estimated_income_with_real_data()
    if success:
        print("\n🎉 ESTIMATED INCOME UPDATE COMPLETED SUCCESSFULLY!")
    else:
        print("\n❌ ESTIMATED INCOME UPDATE FAILED!")
    
    input("Press Enter to close...")
