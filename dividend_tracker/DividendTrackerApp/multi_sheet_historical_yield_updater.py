#!/usr/bin/env python3
"""
Multi-Sheet Historical Yield Updater
====================================
Updated for separate account sheets structure:
- Etrade_IRA
- Etrade_Individual  
- Schwab_IRA
- Schwab_Individual

This replaces the old single-sheet updater now that we have
separate sheets per account.

Author: Claude & Mark
Date: October 18, 2025
"""

import openpyxl
from openpyxl.styles import Font, PatternFill
import json
import os
from datetime import datetime

class MultiSheetHistoricalYieldUpdater:
    def __init__(self):
        self.excel_path = r"C:\Users\mjmat\Python Code in VS\dividend_tracker\DividendTrackerApp\outputs\Dividends_2025.xlsx"
        self.cache_file = r"C:\Users\mjmat\Python Code in VS\dividend_tracker\DividendTrackerApp\portfolio_data_cache.json"
        
        # Map cache account names to new sheet names
        self.account_sheet_mapping = {
            "etrade_ira": "Etrade_IRA",
            "etrade_taxable": "Etrade_Individual",
            "schwab_ira": "Schwab_IRA",
            "schwab_individual": "Schwab_Individual"
        }
        
        # Minimum yield threshold (only include tickers with yield > 4%)
        self.min_yield_threshold = 4.0
    
    def load_cache_data(self):
        """Load data from portfolio cache"""
        try:
            if not os.path.exists(self.cache_file):
                print(f"ERROR: Cache file not found: {self.cache_file}")
                return None
            
            with open(self.cache_file, 'r') as f:
                cache_data = json.load(f)
            
            return cache_data
        except Exception as e:
            print(f"ERROR: Error loading cache: {e}")
            return None
    
    def update_account_sheet(self, wb, sheet_name, account_key, cache_data):
        """
        Update a single account sheet with fresh dividend data
        
        Args:
            wb: Workbook object
            sheet_name: Name of the sheet (e.g., "Etrade_IRA")
            account_key: Cache key for this account (e.g., "etrade_ira")
            cache_data: Full cache data dictionary
        """
        print(f"\n   Updating {sheet_name}...")
        
        if sheet_name not in wb.sheetnames:
            print(f"      WARNING: Sheet '{sheet_name}' not found - skipping")
            return False
        
        ws = wb[sheet_name]
        
        # Get positions and yields for this account
        positions = cache_data.get("positions", {}).get(account_key, [])
        yields_data = cache_data.get("ticker_yields", {})
        
        if not positions:
            print(f"      INFO: No positions found for {account_key}")
            return True
        
        print(f"      - Found {len(positions)} positions in cache")
        
        # Filter for high-yield dividend stocks (>4%)
        high_yield_positions = []
        for pos in positions:
            # Try 'ticker' first, then 'symbol' (cache uses 'symbol')
            ticker = pos.get("ticker") or pos.get("symbol", "")
            # Yield data is nested: ticker_yields[ticker]['yield']
            ticker_yield_data = yields_data.get(ticker, {})
            yield_pct = ticker_yield_data.get('yield', 0.0) if isinstance(ticker_yield_data, dict) else 0.0
            
            if yield_pct >= self.min_yield_threshold:
                # Calculate last_price from market_value if not available
                quantity = pos.get("quantity", 0)
                last_price = pos.get("last_price", 0.0)
                
                # If last_price is 0 but we have market_value and quantity, calculate it
                if last_price == 0 and quantity > 0:
                    market_value = pos.get("market_value", 0.0)
                    if market_value > 0:
                        last_price = market_value / quantity
                
                high_yield_positions.append({
                    "ticker": ticker,
                    "quantity": quantity,
                    "price_paid": pos.get("price_paid", last_price),  # Use last_price as fallback
                    "last_price": last_price,
                    "yield": yield_pct
                })
        
        print(f"      - {len(high_yield_positions)} tickers meet {self.min_yield_threshold}% yield threshold")
        
        if not high_yield_positions:
            print(f"      INFO: No high-yield positions to update")
            return True
        
        # Find header row (row with "Ticker")
        header_row = None
        for row in range(1, min(ws.max_row + 1, 20)):
            cell_value = ws.cell(row=row, column=1).value
            if cell_value and str(cell_value).strip().lower() == "ticker":
                header_row = row
                break
        
        if not header_row:
            print(f"      ERROR: Could not find header row with 'Ticker'")
            return False
        
        print(f"      - Header row found at row {header_row}")
        
        # Get existing tickers in the sheet (starting from header_row + 1)
        existing_tickers = {}
        data_start_row = header_row + 1
        
        for row in range(data_start_row, ws.max_row + 1):
            ticker_cell = ws.cell(row=row, column=1).value
            if ticker_cell and isinstance(ticker_cell, str) and ticker_cell.strip():
                ticker = ticker_cell.strip().upper()
                existing_tickers[ticker] = row
        
        print(f"      - Found {len(existing_tickers)} existing tickers in sheet")
        
        # UPDATE Column D (Last Price) for ALL existing tickers FIRST
        print(f"      - Updating Last Price (Column D) for all existing tickers...")
        for pos in high_yield_positions:
            ticker = pos["ticker"].upper()
            if ticker in existing_tickers:
                row_num = existing_tickers[ticker]
                # Round to 2 decimals and format with 2 decimal places
                cell = ws.cell(row=row_num, column=4)
                cell.value = round(pos["last_price"], 2)
                cell.number_format = '0.00'  # Force 2 decimal places
        
        # INSERT new column P to push historical data right
        print(f"      - Inserting new column P (current yield)...")
        ws.insert_cols(16)  # Insert at column P (16)
        
        # Update column P header with today's date
        ws.cell(row=header_row, column=16).value = datetime.now().strftime('%m/%d/%Y')
        
        # Update existing tickers and track which ones need to be added
        updated_count = 0
        missing_tickers = []
        
        # Get yield data from cache
        yields_data = cache_data.get("ticker_yields", {})
        
        for pos in high_yield_positions:
            ticker = pos["ticker"].upper()
            
            if ticker in existing_tickers:
                # Update existing row
                row_num = existing_tickers[ticker]
                
                # Column B: Quantity
                ws.cell(row=row_num, column=2).value = pos["quantity"]
                
                # Column C: Price Paid - NEVER overwrite existing, only for new tickers
                # (Skip - will be handled when adding new tickers)
                
                # Column D: Last Price - Already updated above before column insertion
                
                # Column E: Change $ - Skip (calculated field)
                
                # Column P: Current Yield with color coding
                current_yield = pos["yield"]
                col_p_cell = ws.cell(row=row_num, column=16)  # Column P (newly inserted)
                # Round to 2 decimals
                col_p_cell.value = round(current_yield, 2)
                col_p_cell.number_format = '0.00'  # Force 2 decimal places
                
                # Get previous yield from Column O (Beginning Dividend Yield)
                previous_yield = ws.cell(row=row_num, column=15).value  # Column O
                if previous_yield:
                    # Color code based on comparison
                    if current_yield > previous_yield:
                        col_p_cell.fill = PatternFill(start_color="00FF00", end_color="00FF00", fill_type="solid")  # Green
                    elif current_yield < previous_yield:
                        col_p_cell.fill = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")  # Red
                    else:
                        col_p_cell.fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")  # Yellow
                
                updated_count += 1
            else:
                # Track missing ticker
                missing_tickers.append(pos)
        
        print(f"      - Updated {updated_count} existing tickers")
        
        # Add missing tickers
        if missing_tickers:
            print(f"      - Adding {len(missing_tickers)} new tickers...")
            
            # Find next empty row after the last ticker (not max_row which includes calculations)
            # Start from header_row and find the last row with a ticker
            last_ticker_row = header_row
            for row in range(header_row + 1, ws.max_row + 1):
                ticker_val = ws.cell(row=row, column=1).value
                if ticker_val and isinstance(ticker_val, str) and ticker_val.strip():
                    last_ticker_row = row
                else:
                    break  # Stop at first empty ticker cell
            
            next_row = last_ticker_row + 1
            
            # Define formatting standards
            ticker_font = Font(name='Arial', size=12, bold=True, color='3072C2')
            qty_font = Font(name='Arial', size=12, bold=True, color='3072C2')
            data_font = Font(name='Calibri', size=12)
            
            # Get yield data from cache
            yields_data = cache_data.get("ticker_yields", {})
            
            for pos in missing_tickers:
                ticker = pos["ticker"]
                
                # Column A: Ticker (Arial 12 Bold #3072C2)
                ticker_cell = ws.cell(row=next_row, column=1)
                ticker_cell.value = ticker
                ticker_cell.font = ticker_font
                
                # Column B: Quantity (Arial 12 Bold #3072C2)
                qty_cell = ws.cell(row=next_row, column=2)
                qty_cell.value = pos["quantity"]
                qty_cell.font = qty_font
                
                # Column C: Price Paid (from API if available, rounded to 2 decimals)
                price_paid_cell = ws.cell(row=next_row, column=3)
                price_paid = pos.get("price_paid", pos["last_price"])
                price_paid_cell.value = round(price_paid, 2)
                price_paid_cell.font = data_font
                price_paid_cell.number_format = '0.00'  # Force 2 decimal places
                
                # Column D: Last Price (current price, rounded to 2 decimals)
                last_price_cell = ws.cell(row=next_row, column=4)
                last_price_cell.value = round(pos["last_price"], 2)
                last_price_cell.font = data_font
                last_price_cell.number_format = '0.00'  # Force 2 decimal places
                
                # Column E: Change $ - Formula (=D-C)
                change_cell = ws.cell(row=next_row, column=5)
                change_cell.value = f"=D{next_row}-C{next_row}"
                change_cell.font = data_font
                
                # Column O: Beginning Dividend Yield (set to current yield for new tickers, 2 decimals)
                beginning_yield_cell = ws.cell(row=next_row, column=15)
                beginning_yield_cell.value = round(pos["yield"], 2)
                beginning_yield_cell.font = data_font
                beginning_yield_cell.number_format = '0.00'  # Force 2 decimal places
                
                # Column P: Current Yield (same as O for new tickers, yellow color, 2 decimals)
                current_yield_cell = ws.cell(row=next_row, column=16)
                current_yield_cell.value = round(pos["yield"], 2)
                current_yield_cell.font = data_font
                current_yield_cell.number_format = '0.00'  # Force 2 decimal places
                current_yield_cell.fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")  # Yellow (same as beginning)
                
                print(f"         ADDED: {ticker} (Qty: {pos['quantity']}, Price: ${pos['last_price']:.2f}, Yield: {pos['yield']:.2f}%)")
                next_row += 1
        
        # Remove obsolete tickers (those in sheet but not in cache or below yield threshold)
        cache_tickers = {pos["ticker"].upper() for pos in high_yield_positions}
        removed_count = 0
        
        for ticker, row_num in list(existing_tickers.items()):
            if ticker not in cache_tickers:
                # Clear the row (set all cells to empty)
                for col in range(1, ws.max_column + 1):
                    ws.cell(row=row_num, column=col).value = None
                
                print(f"         REMOVED: {ticker} (no longer in account or below yield threshold)")
                removed_count += 1
        
        if removed_count > 0:
            print(f"      - Removed {removed_count} obsolete tickers")
        
        print(f"      SUCCESS: {sheet_name} updated successfully")
        return True
    
    def update_dividend_statistics(self, wb, sheet_name, account_key, cache_data):
        """
        Update dividend statistics section at the top of each account sheet
        
        Statistics calculated:
        - Total Annual Dividends
        - Weekly Average
        - Monthly Average  
        - Current Yield %
        - Number of Dividend Positions
        - Last Updated timestamp
        """
        print(f"\n   Updating dividend statistics for {sheet_name}...")
        
        if sheet_name not in wb.sheetnames:
            return False
        
        ws = wb[sheet_name]
        
        # Get pre-calculated dividend estimates from cache
        dividend_estimates = cache_data.get('dividend_estimates', {})
        positions = cache_data.get("positions", {}).get(account_key, [])
        yields_data = cache_data.get("ticker_yields", {})
        portfolio_values = cache_data.get('portfolio_values', {})
        
        # Map account keys to display names
        account_display_names = {
            'etrade_ira': 'E*TRADE IRA',
            'etrade_taxable': 'E*TRADE Taxable',
            'schwab_ira': 'Schwab IRA',
            'schwab_individual': 'Schwab Individual'
        }
        
        account_display_name = account_display_names.get(account_key, account_key.title())
        
        # Get annual dividend for this account from pre-calculated estimates
        annual_dividend = dividend_estimates.get(account_display_name, 0)
        
        # Calculate weekly and monthly averages
        weekly_dividend = annual_dividend / 52
        monthly_dividend = annual_dividend / 12
        
        # Get portfolio value for this account
        account_value = portfolio_values.get(account_display_name, 0)
        
        # Calculate current yield %
        current_yield_pct = (annual_dividend / account_value * 100) if account_value > 0 else 0
        
        # Count dividend-paying positions (yield >0)
        dividend_position_count = 0
        for pos in positions:
            ticker = pos.get("ticker") or pos.get("symbol", "")
            ticker_yield_data = yields_data.get(ticker, {})
            yield_pct = ticker_yield_data.get('yield', 0.0) if isinstance(ticker_yield_data, dict) else 0.0
            if yield_pct > 0:
                dividend_position_count += 1
        
        # Update statistics section (assuming it's in cells B1:C6 or similar)
        # This is a typical layout - adjust row numbers if your sheets differ
        try:
            # Row 1: Account name header
            ws.cell(row=1, column=1).value = f"{account_display_name} - Dividend Statistics"
            
            # Row 2: Annual Dividends
            ws.cell(row=2, column=1).value = "Annual Dividends:"
            ws.cell(row=2, column=2).value = annual_dividend
            ws.cell(row=2, column=2).number_format = '$#,##0.00'
            
            # Row 3: Monthly Average
            ws.cell(row=3, column=1).value = "Monthly Average:"
            ws.cell(row=3, column=2).value = monthly_dividend
            ws.cell(row=3, column=2).number_format = '$#,##0.00'
            
            # Row 4: Weekly Average
            ws.cell(row=4, column=1).value = "Weekly Average:"
            ws.cell(row=4, column=2).value = weekly_dividend
            ws.cell(row=4, column=2).number_format = '$#,##0.00'
            
            # Row 5: Current Yield %
            ws.cell(row=5, column=1).value = "Current Yield:"
            ws.cell(row=5, column=2).value = current_yield_pct / 100  # Store as decimal for percentage format
            ws.cell(row=5, column=2).number_format = '0.00%'
            
            # Row 6: Number of Positions
            ws.cell(row=6, column=1).value = "Dividend Positions:"
            ws.cell(row=6, column=2).value = dividend_position_count
            
            # Row 7: Last Updated
            ws.cell(row=7, column=1).value = "Last Updated:"
            ws.cell(row=7, column=2).value = datetime.now().strftime('%m/%d/%Y %I:%M %p')
            
            print(f"      - Annual: ${annual_dividend:,.2f}")
            print(f"      - Monthly: ${monthly_dividend:,.2f}")
            print(f"      - Weekly: ${weekly_dividend:,.2f}")
            print(f"      - Yield: {current_yield_pct:.2f}%")
            print(f"      - Positions: {dividend_position_count}")
            
            return True
            
        except Exception as e:
            print(f"      ERROR: Failed to update statistics: {e}")
            import traceback
            traceback.print_exc()
            return False
    
    def run_update(self):
        """Main update process"""
        print("=" * 70)
        print("MULTI-SHEET HISTORICAL YIELD UPDATER")
        print("=" * 70)
        print(f"Update Date: {datetime.now().strftime('%m/%d/%Y %I:%M %p')}")
        print(f"Excel File: {os.path.basename(self.excel_path)}")
        print(f"Minimum Yield Threshold: {self.min_yield_threshold}%")
        print()
        
        # Step 1: Load cache
        print("[1] Loading cache data...")
        cache_data = self.load_cache_data()
        if not cache_data:
            print("ERROR: Failed to load cache data")
            return False
        
        print(f"   OK: Cache loaded: {cache_data.get('timestamp', 'Unknown time')}")
        print()
        
        # Step 2: Load workbook
        print("[2] Loading Excel workbook...")
        if not os.path.exists(self.excel_path):
            print(f"   ERROR: Excel file not found: {self.excel_path}")
            return False
        
        try:
            wb = openpyxl.load_workbook(self.excel_path)
            print(f"   OK: Workbook loaded")
            print(f"   Available sheets: {wb.sheetnames}")
        except Exception as e:
            print(f"   ERROR: Error loading workbook: {e}")
            return False
        
        print()
        
        # Step 3: Update each account sheet
        print("[3] Updating account sheets...")
        
        success_count = 0
        for account_key, sheet_name in self.account_sheet_mapping.items():
            try:
                success = self.update_account_sheet(wb, sheet_name, account_key, cache_data)
                if success:
                    success_count += 1
            except Exception as e:
                print(f"      ERROR: Error updating {sheet_name}: {e}")
                import traceback
                traceback.print_exc()
        
        print()
        print(f"   OK: Updated {success_count}/{len(self.account_sheet_mapping)} account sheets")
        
        # Step 4: Save workbook
        print()
        print("[4] Saving workbook...")
        try:
            wb.save(self.excel_path)
            wb.close()
            print(f"   OK: Workbook saved successfully")
        except Exception as e:
            print(f"   ERROR: Error saving workbook: {e}")
            wb.close()
            return False
        
        print()
        print("=" * 70)
        print("SUCCESS: UPDATE COMPLETE!")
        print("=" * 70)
        print()
        print(f"Summary:")
        print(f"   - Account sheets updated: {success_count}/{len(self.account_sheet_mapping)}")
        print(f"   - Minimum yield threshold: {self.min_yield_threshold}%")
        print(f"   - File: {os.path.basename(self.excel_path)}")
        
        return True

if __name__ == "__main__":
    updater = MultiSheetHistoricalYieldUpdater()
    success = updater.run_update()
    
    if success:
        print("\nSUCCESS: All account sheets updated successfully!")
    else:
        print("\nERROR: Update failed - check errors above")
