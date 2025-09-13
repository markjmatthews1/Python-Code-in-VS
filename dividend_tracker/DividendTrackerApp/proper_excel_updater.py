#!/usr/bin/env python3
"""
ENHANCED Excel Append-Only Updater with Color Coding
===================================================

ENHANCED FEATURES:
1. Correct color codes: Green #90EE90, Red #FF7C80, Yellow #FFFF00
2. Color coding for Portfolio Values rows 4-9 (including 401K)
3. Color coding for Estimated Income rows 4-9 (including Monthly Average)
4. Proper 401K integration in Portfolio Values
5. Row 9 formula correction: =SUM(rows 4:7)/12
6. Time-series column structure preservation
"""

import os
import sys
import openpyxl
import openpyxl.utils
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.styles.numbers import FORMAT_CURRENCY_USD_SIMPLE
import traceback
from datetime import datetime
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.styles.numbers import FORMAT_CURRENCY_USD_SIMPLE, FORMAT_PERCENTAGE
from datetime import datetime
import traceback

# Add modules to path
sys.path.append(os.path.join(os.path.dirname(__file__), 'modules'))

class ProperExcelUpdater:
    """Enhanced append-only Excel updater with color coding"""
    
    def __init__(self):
        self.script_dir = os.path.dirname(os.path.abspath(__file__))
        self.outputs_dir = os.path.join(self.script_dir, "outputs")
        self.excel_file = os.path.join(self.outputs_dir, "Dividends_2025.xlsx")
        self.cache_file = os.path.join(self.script_dir, "portfolio_data_cache.json")
        self.today = datetime.now()
        self.today_str = self.today.strftime("%Y-%m-%d")
    
    def get_previous_column_values(self, ws, row_range):
        """Get values from the previous column for color comparison"""
        previous_values = {}
        
        # Find the last column with data in row 3 (dates row)
        last_col = 1
        for col in range(1, ws.max_column + 1):
            if ws.cell(row=3, column=col).value:
                last_col = col
        
        # Get values from the previous column
        if last_col > 1:
            prev_col = last_col
            for row in row_range:
                account_name = ws.cell(row=row, column=1).value
                if account_name:
                    account_key = str(account_name).strip()
                    prev_value = ws.cell(row=row, column=prev_col).value
                    if prev_value and isinstance(prev_value, (int, float)):
                        previous_values[account_key] = prev_value
        
        return previous_values
    
    def apply_color_coding(self, cell, new_value, old_value):
        """Apply color coding based on value comparison"""
        if old_value is None or old_value == 0:
            # No previous value, use default formatting
            cell.fill = PatternFill(fill_type=None)
        elif new_value > old_value:
            # Increase - Green #90EE90
            cell.fill = PatternFill(start_color='90EE90', end_color='90EE90', fill_type='solid')
        elif new_value < old_value:
            # Decrease - Red #FF7C80
            cell.fill = PatternFill(start_color='FF7C80', end_color='FF7C80', fill_type='solid')
        else:
            # Same value - Yellow #FFFF00
            cell.fill = PatternFill(start_color='FFFF00', end_color='FFFF00', fill_type='solid')
        
        print("🔧 PROPER EXCEL UPDATER - APPEND-ONLY")
        print("=" * 45)
        print(f"📅 Today: {self.today_str}")
        print(f"📁 Excel File: {os.path.basename(self.excel_file)}")
        
    def run_update(self):
        """Run the proper append-only update"""
        
        print("\n📋 This will ADD new data columns while preserving ALL historical data")
        print("-" * 60)
        
        # Step 1: Get 401K value (working popup)
        print("\n💰 STEP 1: Getting 401K value...")
        k401_value = self.get_401k_value()
        if not k401_value:
            return False
            
        # Step 2: Get fresh API data  
        print("\n📊 STEP 2: Getting fresh API data...")
        fresh_data = self.get_fresh_api_data(k401_value)
        if not fresh_data:
            return False
            
        # Step 3: Create backup
        print("\n💾 STEP 3: Creating backup...")
        self.create_backup()
        
        # Step 4: Update sheets properly
        print("\n📝 STEP 4: Adding new columns to existing sheets...")
        
        success_count = 0
        
        # Update Portfolio Values 2025 (time-series)
        if self.update_portfolio_values_timeseries(fresh_data, k401_value):
            print("   ✅ Portfolio Values 2025: New column added")
            success_count += 1
        else:
            print("   ❌ Portfolio Values 2025: Failed")
            
        # Update Estimated Income 2025 (time-series)
        if self.update_estimated_income_timeseries(fresh_data):
            print("   ✅ Estimated Income 2025: New column added")
            success_count += 1
        else:
            print("   ❌ Estimated Income 2025: Failed")
            
        # Update Accounts Div historical yield sheet (NEW!)
        if self.update_historical_yield_sheet():
            print("   ✅ Accounts Div historical yield: New column added with color coding")
            success_count += 1
        else:
            print("   ❌ Accounts Div historical yield: Failed")
            
        # Update other sheets safely
        if self.update_other_sheets_safely(fresh_data):
            print("   ✅ Other sheets: Updated safely")
            success_count += 1
        else:
            print("   ❌ Other sheets: Some issues")
        
        print(f"\n🎯 RESULTS: {success_count}/4 components updated")
        
        if success_count >= 3:
            print("✅ SUCCESS: Key sheets updated including historical yields!")
            return True
        else:
            print("❌ PARTIAL: Some updates may need attention")
            return False
            
    def get_401k_value(self):
        """Get 401K value using the working popup"""
        try:
            from gui_prompts import get_k401_value
            print("   📱 Opening 401K dialog (working version)...")
            k401_value = get_k401_value()
            
            if k401_value and k401_value > 0:
                print(f"   ✅ 401K Value: ${k401_value:,.2f}")
                return k401_value
            else:
                print("   ❌ Invalid 401K value")
                return None
                
        except Exception as e:
            print(f"   ❌ Error with 401K popup: {e}")
            # Console fallback
            try:
                value_str = input("   💰 Enter 401K value: $")
                k401_value = float(value_str.replace(',', '').replace('$', ''))
                if k401_value > 0:
                    return k401_value
            except:
                pass
            return None
            
    def get_fresh_api_data(self, k401_value):
        """Get fresh data using the working portfolio collector"""
        try:
            from portfolio_data_collector import PortfolioDataCollector
            
            collector = PortfolioDataCollector()
            
            print("   🔄 Collecting from working APIs...")
            
            # Use the working collect method with fallback
            fresh_data = collector.collect_all_data_with_fallback(k401_value)
            
            if fresh_data and fresh_data.get('totals', {}).get('total_portfolio', 0) > 0:
                totals = fresh_data['totals']
                print(f"   ✅ Portfolio: ${totals['total_portfolio']:,.2f}")
                print(f"   ✅ Annual Dividends: ${totals['total_yearly_dividends']:,.2f}")
                return fresh_data
            else:
                print("   ❌ No valid API data collected")
                return None
                
        except Exception as e:
            print(f"   ❌ API collection error: {e}")
            return None
            
    def create_backup(self):
        """Create backup before making changes"""
        try:
            if os.path.exists(self.excel_file):
                timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
                backup_file = self.excel_file.replace('.xlsx', f'_backup_enhanced_{timestamp}.xlsx')
                
                import shutil
                shutil.copy2(self.excel_file, backup_file)
                print(f"   💾 Backup: {os.path.basename(backup_file)}")
                return True
        except Exception as e:
            print(f"   ⚠️ Backup error: {e}")
        return False
        
    def update_portfolio_values_timeseries(self, fresh_data, k401_value):
        """Update Portfolio Values 2025 by adding new date column"""
        try:
            wb = openpyxl.load_workbook(self.excel_file)
            
            if "Portfolio Values 2025" not in wb.sheetnames:
                print("      ❌ Portfolio Values 2025 sheet not found")
                return False
                
            ws = wb["Portfolio Values 2025"]
            
            # Find the structure: dates should be in row 3, accounts in column A starting row 4
            print("      📊 Analyzing existing Portfolio Values structure...")
            
            # Get previous values for color comparison (rows 4-9)
            previous_values = self.get_previous_column_values(ws, range(4, 10))
            
            # Find last column with data in row 3 (dates row)
            last_col = 1
            for col in range(1, ws.max_column + 1):
                if ws.cell(row=3, column=col).value:
                    last_col = col
            
            # Add new date column
            new_col = last_col + 1
            
            # Add today's date to the header row (row 3) with proper formatting
            date_cell = ws.cell(row=3, column=new_col, value=self.today_str)
            date_cell.font = Font(name='Arial', size=12, bold=True, color='FFFFFF')
            date_cell.fill = PatternFill(start_color='4F81BD', end_color='4F81BD', fill_type='solid')
            date_cell.alignment = Alignment(horizontal='right')
            date_cell.number_format = 'm/d/yyyy'
            
            # Add account values to the new column
            portfolio_values = fresh_data.get('portfolio_values', {})
            
            # Map the account names to what's in the Excel (including 401K)
            account_mapping = {
                'E*TRADE IRA': portfolio_values.get('E*TRADE IRA', 0),
                'E*TRADE Taxable': portfolio_values.get('E*TRADE Taxable', 0),
                'Schwab IRA': portfolio_values.get('Schwab IRA', 0),
                'Schwab Individual': portfolio_values.get('Schwab Individual', 0),
                '401k Retirement (Manual)': k401_value  # Match exact Excel label
            }
            
            # Find and update account rows with color coding (rows 4-9)
            for row in range(4, 10):  # Rows 4-9 (accounts + total)
                account_name = ws.cell(row=row, column=1).value
                if account_name:
                    account_key = str(account_name).strip()
                    
                    # Handle exact matches and variations
                    value = None
                    if account_key in account_mapping:
                        value = account_mapping[account_key]
                    elif 'total' in account_key.lower():
                        # Calculate total for Total row - 401K is already included in portfolio_values
                        value = sum(portfolio_values.values())
                    elif any(partial in account_key for partial in ['E*TRADE', 'Etrade', 'Schwab']):
                        # Handle variations in account names
                        for map_key, map_value in account_mapping.items():
                            if map_key.replace('*', '') in account_key or map_key in account_key:
                                value = map_value
                                break
                    elif '401k' in account_key.lower() or '401' in account_key:
                        # Handle 401k variations
                        value = k401_value
                    
                    if value is not None:
                        cell = ws.cell(row=row, column=new_col, value=value)
                        cell.number_format = FORMAT_CURRENCY_USD_SIMPLE
                        cell.font = Font(name='Arial', size=12, bold=('total' in account_key.lower()))
                        
                        # Apply color coding
                        old_value = previous_values.get(account_key, None)
                        self.apply_color_coding(cell, value, old_value)
                        
                        print(f"      💰 {account_key}: ${value:,.2f} (Color coded)")
            
            # Legacy total calculation for older rows (kept for compatibility)
            total_portfolio = sum(portfolio_values.values())  # 401K already included in portfolio_values
            for row in range(10, ws.max_row + 1):
                account_name = ws.cell(row=row, column=1).value
                if account_name and 'total' in str(account_name).lower():
                    cell = ws.cell(row=row, column=new_col, value=total_portfolio)
                    cell.number_format = FORMAT_CURRENCY_USD_SIMPLE
                    cell.font = Font(name='Arial', size=12, bold=True)
                    
                    # Apply color coding
                    account_key = str(account_name).strip()
                    old_value = previous_values.get(account_key, None)
                    self.apply_color_coding(cell, total_portfolio, old_value)
                    
                    print(f"      💰 {account_key}: ${total_portfolio:,.2f} (Color coded)")
                    break
            
            # Set column width for proper display
            ws.column_dimensions[openpyxl.utils.get_column_letter(new_col)].width = 15
            
            wb.save(self.excel_file)
            wb.close()
            return True
            
        except Exception as e:
            print(f"      ❌ Portfolio Values update error: {e}")
            traceback.print_exc()
            return False
            
    def update_estimated_income_timeseries(self, fresh_data):
        """Update Estimated Income 2025 by adding new date column"""
        try:
            wb = openpyxl.load_workbook(self.excel_file)
            
            if "Estimated Income 2025" not in wb.sheetnames:
                print("      ❌ Estimated Income 2025 sheet not found")
                return False
                
            ws = wb["Estimated Income 2025"]
            
            # Find the structure: dates should be in row 3, accounts in column A starting row 4
            print("      📊 Analyzing existing Estimated Income structure...")
            
            # Get previous values for color comparison (rows 4-9)
            previous_values = self.get_previous_column_values(ws, range(4, 10))
            
            # Find last column with data in row 3 (dates row) 
            last_col = 1
            for col in range(1, ws.max_column + 1):
                if ws.cell(row=3, column=col).value:
                    last_col = col
            
            # Add new date column
            new_col = last_col + 1
            
            # Add today's date to the header row (row 3) with proper formatting
            date_cell = ws.cell(row=3, column=new_col, value=self.today_str)
            date_cell.font = Font(name='Arial', size=12, bold=True, color='FFFFFF')
            date_cell.fill = PatternFill(start_color='4F81BD', end_color='4F81BD', fill_type='solid')
            date_cell.alignment = Alignment(horizontal='right')
            date_cell.number_format = 'm/d/yyyy'
            
            # Add dividend estimates to the new column
            dividend_estimates = fresh_data.get('dividend_estimates', {})
            
            # Map the account names
            account_mapping = {
                'E*TRADE IRA': dividend_estimates.get('E*TRADE IRA', 0),
                'E*TRADE Taxable': dividend_estimates.get('E*TRADE Taxable', 0), 
                'Schwab IRA': dividend_estimates.get('Schwab IRA', 0),
                'Schwab Individual': dividend_estimates.get('Schwab Individual', 0)
            }
            
            # Find and update account rows with color coding (rows 4-7)
            for row in range(4, 8):  # Dividend account rows
                account_name = ws.cell(row=row, column=1).value
                if account_name:
                    account_str = str(account_name).strip()
                    
                    # Try exact match first
                    value = None
                    if account_str in account_mapping:
                        value = account_mapping[account_str]
                    else:
                        # Try partial matches for account names
                        for map_key, map_value in account_mapping.items():
                            if map_key.replace('*', '') in account_str or map_key in account_str:
                                value = map_value
                                break
                    
                    if value is not None:
                        cell = ws.cell(row=row, column=new_col, value=value)
                        cell.number_format = FORMAT_CURRENCY_USD_SIMPLE
                        cell.font = Font(name='Arial', size=12)
                        
                        # Apply color coding
                        old_value = previous_values.get(account_str, None)
                        self.apply_color_coding(cell, value, old_value)
                        
                        print(f"      📈 {account_str}: ${value:,.2f} (Color coded)")
            
            # CRITICAL: Add Row 9 Monthly Average Calculation with color coding
            # Row 9 should calculate: =SUM(rows 4:7)/12 (ALL dividend accounts monthly average)
            monthly_row = None
            for row in range(8, 12):  # Look for Monthly Average row
                cell_value = ws.cell(row=row, column=1).value
                if cell_value and 'monthly' in str(cell_value).lower():
                    monthly_row = row
                    break
            
            if monthly_row:
                # Add the formula: =SUM(E4:E7)/12 format (rows 4-7, not 5-7)
                col_letter = openpyxl.utils.get_column_letter(new_col)
                formula = f"=SUM({col_letter}4:{col_letter}7)/12"
                
                cell = ws.cell(row=monthly_row, column=new_col)
                cell.value = formula
                cell.number_format = FORMAT_CURRENCY_USD_SIMPLE
                cell.font = Font(name='Arial', size=12, bold=True)  # Make Monthly Average bold
                
                # For formulas, save first to calculate the value, then apply color coding
                wb.save(self.excel_file)
                wb.close()
                wb = openpyxl.load_workbook(self.excel_file)
                ws = wb["Estimated Income 2025"]
                
                # Manual calculation for reliable color coding comparison
                # Sum dividend values in rows 4-7 of the new column, then divide by 12
                current_sum = 0
                for row in range(4, 8):  # Rows 4-7 (dividend accounts)
                    cell_value = ws.cell(row=row, column=new_col).value
                    if isinstance(cell_value, (int, float)):
                        current_sum += cell_value
                
                calculated_value = current_sum / 12 if current_sum > 0 else 0
                
                if calculated_value > 0:
                    account_name = ws.cell(row=monthly_row, column=1).value
                    account_key = str(account_name).strip() if account_name else "Monthly"
                    
                    # Get previous monthly average for comparison
                    old_value = None
                    if new_col > 1:
                        prev_sum = 0
                        for row in range(4, 8):
                            prev_cell_value = ws.cell(row=row, column=new_col - 1).value
                            if isinstance(prev_cell_value, (int, float)):
                                prev_sum += prev_cell_value
                        old_value = prev_sum / 12 if prev_sum > 0 else None
                    
                    # Reapply formatting with color coding and bold
                    cell = ws.cell(row=monthly_row, column=new_col)
                    cell.value = formula
                    cell.number_format = FORMAT_CURRENCY_USD_SIMPLE
                    cell.font = Font(name='Arial', size=12, bold=True)  # Keep bold after color coding
                    self.apply_color_coding(cell, calculated_value, old_value)
                    
                    print(f"      🧮 Monthly Average: {formula} = ${calculated_value:,.2f} (Bold & Color coded)")
                else:
                    print(f"      🧮 Monthly Average: {formula} (Bold formatting)")
            
            # Add total dividends with color coding
            total_dividends = fresh_data.get('totals', {}).get('total_yearly_dividends', 0)
            for row in range(4, ws.max_row + 1):
                account_name = ws.cell(row=row, column=1).value
                if account_name and 'total' in str(account_name).lower():
                    cell = ws.cell(row=row, column=new_col, value=total_dividends)
                    cell.number_format = FORMAT_CURRENCY_USD_SIMPLE
                    cell.font = Font(name='Arial', size=12, bold=True)
                    
                    # Apply color coding
                    account_key = str(account_name).strip()
                    old_value = previous_values.get(account_key, None)
                    self.apply_color_coding(cell, total_dividends, old_value)
                    
                    print(f"      📈 Total Dividends: ${total_dividends:,.2f} (Color coded)")
                    cell = ws.cell(row=row, column=new_col, value=total_dividends)
                    cell.number_format = FORMAT_CURRENCY_USD_SIMPLE
                    cell.font = Font(name='Arial', size=12, bold=True)
                    print(f"      📈 Total Dividends: ${total_dividends:,.2f}")
                    break
            
            # Set column width for proper display
            ws.column_dimensions[openpyxl.utils.get_column_letter(new_col)].width = 15
            
            wb.save(self.excel_file)
            wb.close()
            return True
            
        except Exception as e:
            print(f"      ❌ Estimated Income update error: {e}")
            traceback.print_exc()
            return False
            
    def update_historical_yield_sheet(self):
        """Update Accounts Div historical yield sheet using cache data"""
        try:
            print("   📊 Updating historical yield sheet from cache...")
            
            # Import and run the cache-based historical yield updater
            import subprocess
            import sys
            
            # Run the dedicated historical yield updater
            updater_script = os.path.join(self.script_dir, "cache_historical_yield_updater.py")
            
            if not os.path.exists(updater_script):
                print(f"      ❌ Historical yield updater not found: {updater_script}")
                return False
            
            # Run the updater as a subprocess to avoid import conflicts
            result = subprocess.run([sys.executable, updater_script], 
                                  capture_output=True, text=True, cwd=self.script_dir)
            
            if result.returncode == 0:
                print("   ✅ Historical yield sheet updated successfully")
                return True
            else:
                print(f"   ❌ Historical yield updater failed: {result.stderr}")
                return False
                
        except Exception as e:
            print(f"   ❌ Historical yield update error: {e}")
            traceback.print_exc()
            return False
            
    def update_other_sheets_safely(self, fresh_data):
        """Update other sheets without destroying existing data"""
        try:
            # Just update the timestamp on Portfolio Summary to show it was touched
            wb = openpyxl.load_workbook(self.excel_file)
            
            if "Portfolio Summary" in wb.sheetnames:
                ws = wb["Portfolio Summary"]
                
                # Find a safe place to put an update timestamp
                # Look for an empty area or existing timestamp
                for row in range(1, min(10, ws.max_row + 1)):
                    for col in range(1, min(5, ws.max_column + 1)):
                        cell_value = ws.cell(row=row, column=col).value
                        if cell_value and "updated" in str(cell_value).lower():
                            # Update existing timestamp
                            ws.cell(row=row, column=col, value=f"Last Updated: {self.today_str}")
                            break
                else:
                    # Add timestamp in a safe location (bottom of sheet)
                    safe_row = ws.max_row + 2
                    ws.cell(row=safe_row, column=1, value=f"API Data Updated: {self.today_str}")
                
            wb.save(self.excel_file)
            wb.close()
            return True
            
        except Exception as e:
            print(f"      ⚠️ Other sheets update: {e}")
            return False


def main():
    """Main execution"""
    updater = ProperExcelUpdater()
    
    try:
        success = updater.run_update()
        
        if success:
            print("\n🎉 SUCCESS! Time-series data updated properly")
            print("📊 Your historical data has been preserved")
            print("📈 New columns added with today's fresh API data")
        else:
            print("\n⚠️ Some issues occurred, but historical data preserved")
            
        print(f"\nPress Enter to close...")
        input()
        
    except KeyboardInterrupt:
        print("\n❌ Update cancelled by user")
    except Exception as e:
        print(f"\n❌ Critical error: {e}")
        traceback.print_exc()


if __name__ == "__main__":
    main()
