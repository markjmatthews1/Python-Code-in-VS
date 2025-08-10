#!/usr/bin/env python3
"""
Enhanced Automated Ticker Dividend Tracker
- Automatically updates share counts from E*TRADE and Schwab APIs
- Auto-calculates all portfolio metrics
- Maintains yield tracking with color coding
- Full automation of dividend_stocks.xlsx equivalent
"""

import openpyxl
from openpyxl.styles import PatternFill, Font
from openpyxl.styles.numbers import FORMAT_CURRENCY_USD, FORMAT_PERCENTAGE_00
import pandas as pd
from datetime import datetime, date
import requests
import os
import sys

# Add modules to path
sys.path.append(os.path.join(os.path.dirname(__file__), 'modules'))

from etrade_auth import get_etrade_session
from etrade_account_api import get_account_info, get_positions

def create_enhanced_ticker_tracker():
    """
    Create a fully automated ticker tracking system
    """
    
    print("🚀 Creating Enhanced Automated Ticker Tracker...")
    
    # File paths
    data_file = os.path.join("data", "dividend_stocks.xlsx")
    output_file = os.path.join("outputs", "Dividends_2025.xlsx")
    
    # Check if dividend_stocks.xlsx exists in data folder
    if not os.path.exists(data_file):
        print(f"❌ Please copy dividend_stocks.xlsx to {data_file}")
        return
    
    # Load existing ticker data for reference
    print("📊 Loading existing ticker reference data...")
    try:
        df_reference = pd.read_excel(data_file)
        reference_tickers = df_reference['Ticker'].tolist() if 'Ticker' in df_reference.columns else []
        print(f"   Found {len(reference_tickers)} reference tickers")
    except Exception as e:
        print(f"   ⚠️ Could not load reference data: {e}")
        reference_tickers = []
    
    # Get live data from APIs
    print("🔌 Connecting to APIs...")
    
    # E*TRADE data
    etrade_positions = get_live_etrade_positions()
    schwab_positions = get_live_schwab_positions()  # We'll need to implement this
    
    # Combine all positions
    all_positions = combine_account_positions(etrade_positions, schwab_positions, reference_tickers)
    
    # Create or update the ticker analysis sheet
    create_ticker_analysis_sheet(all_positions, output_file)
    
    print("✅ Enhanced ticker tracker created!")

def get_live_etrade_positions():
    """
    Get current positions from E*TRADE API
    """
    print("   📈 Fetching E*TRADE positions...")
    
    try:
        session, base_url = get_etrade_session()
        
        # Get accounts
        accounts_url = f"{base_url}/v1/accounts/list.json"
        accounts_response = session.get(accounts_url)
        accounts_data = accounts_response.json()
        
        all_positions = {}
        
        if 'AccountListResponse' in accounts_data:
            for account in accounts_data['AccountListResponse']['Accounts']['Account']:
                account_id = account['accountIdKey']
                account_desc = account['accountDesc']
                
                # Get positions for this account
                positions_url = f"{base_url}/v1/accounts/{account_id}/portfolio.json"
                positions_response = session.get(positions_url)
                
                if positions_response.status_code == 200:
                    positions_data = positions_response.json()
                    
                    if 'PortfolioResponse' in positions_data and 'AccountPortfolio' in positions_data['PortfolioResponse']:
                        for portfolio in positions_data['PortfolioResponse']['AccountPortfolio']:
                            if 'Position' in portfolio:
                                for position in portfolio['Position']:
                                    product = position['Product']
                                    symbol = product['symbol']
                                    
                                    # Only track dividend-paying stocks (exclude options, etc.)
                                    if product['securityType'] == 'EQ':
                                        quantity = float(position['quantity'])
                                        price_paid = float(position.get('pricePaid', 0))
                                        market_value = float(position.get('marketValue', 0))
                                        current_price = market_value / quantity if quantity > 0 else 0
                                        
                                        # Get quote data for dividend info
                                        quote_data = get_etrade_quote(session, base_url, symbol)
                                        
                                        all_positions[symbol] = {
                                            'account': f'E*TRADE ({account_desc})',
                                            'quantity': quantity,
                                            'price_paid': price_paid,
                                            'current_price': current_price,
                                            'market_value': market_value,
                                            'original_value': quantity * price_paid,
                                            'quote_data': quote_data
                                        }
                                        
                                        print(f"      {symbol}: {quantity} shares @ ${current_price:.2f}")
        
        return all_positions
        
    except Exception as e:
        print(f"   ❌ Error getting E*TRADE positions: {e}")
        return {}

def get_etrade_quote(session, base_url, symbol):
    """
    Get quote data including dividend information
    """
    try:
        quote_url = f"{base_url}/v1/market/quote/{symbol}.json"
        quote_response = session.get(quote_url)
        
        if quote_response.status_code == 200:
            quote_json = quote_response.json()
            if 'QuoteResponse' in quote_json and 'QuoteData' in quote_json['QuoteResponse']:
                quote_data = quote_json['QuoteResponse']['QuoteData'][0]['All']
                
                return {
                    'last_trade': float(quote_data.get('lastTrade', 0)),
                    'change': float(quote_data.get('netChange', 0)),
                    'change_pct': float(quote_data.get('netChangePct', 0)),
                    'dividend_yield': float(quote_data.get('yield', 0)),
                    'dividend_amount': float(quote_data.get('dividendAmount', 0)),
                    'dividend_date': quote_data.get('dividendPayableDate', ''),
                    'ex_dividend_date': quote_data.get('exDividendDate', '')
                }
    except Exception as e:
        print(f"      ⚠️ Could not get quote for {symbol}: {e}")
    
    return {}

def get_live_schwab_positions():
    """
    Get current positions from Schwab (placeholder for now)
    Note: Schwab API integration would need to be implemented
    """
    print("   📊 Schwab positions (manual entry for now)...")
    
    # For now, return known Schwab positions from the spreadsheet data
    # In the future, this would connect to Schwab API
    schwab_positions = {
        'DX': {
            'account': 'Schwab Individual',
            'quantity': 125,
            'price_paid': 11.91,
            'current_price': 12.53,  # Will be updated by quote
            'market_value': 125 * 12.53,
            'original_value': 125 * 11.91,
            'quote_data': {}
        },
        'AGNC': {
            'account': 'Schwab Individual', 
            'quantity': 1060,
            'price_paid': 9.42,
            'current_price': 9.25,
            'market_value': 1060 * 9.25,
            'original_value': 1060 * 9.42,
            'quote_data': {}
        },
        'QDTE': {
            'account': 'Schwab Individual',
            'quantity': 285,
            'price_paid': 35.30,
            'current_price': 35.36,
            'market_value': 285 * 35.36,
            'original_value': 285 * 35.30,
            'quote_data': {}
        },
        'ECC': {
            'account': 'Schwab Individual',
            'quantity': 1330,
            'price_paid': 7.43,
            'current_price': 7.45,
            'market_value': 1330 * 7.45,
            'original_value': 1330 * 7.43,
            'quote_data': {}
        }
    }
    
    print(f"      Found {len(schwab_positions)} Schwab positions")
    return schwab_positions

def combine_account_positions(etrade_positions, schwab_positions, reference_tickers):
    """
    Combine positions from all accounts and supplement with reference data
    """
    print("🔗 Combining account positions...")
    
    combined_positions = {}
    
    # Add E*TRADE positions
    for symbol, data in etrade_positions.items():
        combined_positions[symbol] = data
        combined_positions[symbol]['source'] = 'E*TRADE API'
    
    # Add Schwab positions  
    for symbol, data in schwab_positions.items():
        if symbol in combined_positions:
            # If we have the same ticker in multiple accounts, we need to handle this
            # For now, just add a suffix
            symbol_key = f"{symbol}_Schwab"
        else:
            symbol_key = symbol
            
        combined_positions[symbol_key] = data
        combined_positions[symbol_key]['source'] = 'Schwab'
    
    print(f"   📊 Total combined positions: {len(combined_positions)}")
    
    return combined_positions

def create_ticker_analysis_sheet(positions_data, output_file):
    """
    Create a new 'Ticker Analysis 2025' sheet with automated calculations
    """
    print("📝 Creating Ticker Analysis 2025 sheet...")
    
    # Load the workbook
    if os.path.exists(output_file):
        wb = openpyxl.load_workbook(output_file)
    else:
        wb = openpyxl.Workbook()
        # Remove default sheet
        if 'Sheet' in wb.sheetnames:
            wb.remove(wb['Sheet'])
    
    # Create or get the ticker analysis sheet
    sheet_name = "Ticker Analysis 2025"
    if sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        ws.delete_rows(1, ws.max_row)  # Clear existing data
    else:
        ws = wb.create_sheet(sheet_name)
    
    # Create headers
    headers = [
        'Ticker', 'Account', 'Qty #', 'Price Paid $', 'Last Price $', 'Day\'s Gain $', 
        'Change $', 'Change %', 'Current Value $', 'Original Value $', 'Total Gain $',
        'Total Gain %', 'Pay Date', 'Payment Cycle', 'Rate per Share', 
        'Monthly Dividend', 'Annual Dividend', 'Current Yield %', 'Previous Yield %',
        'Yield Change', 'Today\'s Date'
    ]
    
    # Add headers
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col)
        cell.value = header
        cell.font = Font(bold=True)
        cell.fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    
    # Add data rows
    today = date.today().strftime("%m/%d/%Y")
    row = 2
    
    total_current_value = 0
    total_original_value = 0
    total_monthly_dividend = 0
    
    for symbol, data in positions_data.items():
        # Clean symbol name (remove account suffix if any)
        clean_symbol = symbol.replace('_Schwab', '')
        
        # Calculate values
        quantity = data['quantity']
        price_paid = data['price_paid']
        current_price = data['current_price']
        current_value = quantity * current_price
        original_value = quantity * price_paid
        total_gain = current_value - original_value
        total_gain_pct = (total_gain / original_value * 100) if original_value > 0 else 0
        
        # Get quote data
        quote_data = data.get('quote_data', {})
        day_change = quote_data.get('change', 0) * quantity
        price_change = quote_data.get('change', 0)
        change_pct = quote_data.get('change_pct', 0)
        dividend_yield = quote_data.get('dividend_yield', 0)
        dividend_amount = quote_data.get('dividend_amount', 0)
        
        # Estimate monthly dividend (this would need refinement based on payment frequency)
        monthly_dividend = dividend_amount * quantity / 12  # Rough estimate
        
        # Add row data
        row_data = [
            clean_symbol,                    # Ticker
            data['account'],                 # Account
            quantity,                        # Qty #
            price_paid,                      # Price Paid $
            current_price,                   # Last Price $
            day_change,                      # Day's Gain $
            price_change,                    # Change $
            change_pct,                      # Change %
            current_value,                   # Current Value $
            original_value,                  # Original Value $
            total_gain,                      # Total Gain $
            total_gain_pct,                  # Total Gain %
            '',                              # Pay Date (would need dividend calendar)
            '',                              # Payment Cycle
            dividend_amount,                 # Rate per Share
            monthly_dividend,                # Monthly Dividend
            dividend_amount * quantity,      # Annual Dividend
            dividend_yield,                  # Current Yield %
            0,                               # Previous Yield % (would need historical data)
            0,                               # Yield Change
            today                            # Today's Date
        ]
        
        for col, value in enumerate(row_data, 1):
            cell = ws.cell(row=row, column=col)
            cell.value = value
            
            # Apply formatting
            if col in [4, 5, 6, 7, 9, 10, 11, 15, 16, 17]:  # Currency columns
                cell.number_format = FORMAT_CURRENCY_USD
            elif col in [8, 12, 18, 19, 20]:  # Percentage columns
                cell.number_format = FORMAT_PERCENTAGE_00
        
        # Add totals
        total_current_value += current_value
        total_original_value += original_value
        total_monthly_dividend += monthly_dividend
        
        row += 1
    
    # Add totals row
    totals_row = row + 1
    ws.cell(row=totals_row, column=1).value = "TOTALS"
    ws.cell(row=totals_row, column=1).font = Font(bold=True)
    
    ws.cell(row=totals_row, column=9).value = total_current_value  # Current Value
    ws.cell(row=totals_row, column=9).number_format = FORMAT_CURRENCY_USD
    ws.cell(row=totals_row, column=9).font = Font(bold=True)
    
    ws.cell(row=totals_row, column=10).value = total_original_value  # Original Value
    ws.cell(row=totals_row, column=10).number_format = FORMAT_CURRENCY_USD
    ws.cell(row=totals_row, column=10).font = Font(bold=True)
    
    ws.cell(row=totals_row, column=16).value = total_monthly_dividend  # Monthly Dividend
    ws.cell(row=totals_row, column=16).number_format = FORMAT_CURRENCY_USD
    ws.cell(row=totals_row, column=16).font = Font(bold=True)
    
    # Auto-size columns
    for column in ws.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 20)
        ws.column_dimensions[column_letter].width = adjusted_width
    
    # Save the workbook
    wb.save(output_file)
    print(f"✅ Ticker Analysis sheet created in {output_file}")
    
    # Print summary
    print(f"\n📊 Summary:")
    print(f"   Total Positions: {len(positions_data)}")
    print(f"   Total Current Value: ${total_current_value:,.2f}")
    print(f"   Total Original Value: ${total_original_value:,.2f}")
    print(f"   Total Gain: ${total_current_value - total_original_value:,.2f}")
    print(f"   Total Monthly Dividend: ${total_monthly_dividend:,.2f}")

if __name__ == "__main__":
    create_enhanced_ticker_tracker()
