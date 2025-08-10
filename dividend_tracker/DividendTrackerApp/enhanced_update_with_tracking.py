#!/usr/bin/env python3
"""
Enhanced Update Dividend Sheet - Integrated Ticker Tracking
Modifies the original Update_dividend_sheet.py to automatically update
our comprehensive ticker tracking system with yield history.
"""

import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment
from openpyxl.styles.numbers import FORMAT_CURRENCY_USD, FORMAT_PERCENTAGE_00
from datetime import datetime, date
import os
import sys

# Add modules path
sys.path.append(os.path.join(os.path.dirname(__file__), 'modules'))

def enhanced_update_dividend_sheet_with_tracking():
    """
    Enhanced version that updates both dividend_stocks.xlsx AND our tracking system
    """
    
    print("🚀 ENHANCED DIVIDEND SHEET UPDATE WITH TRACKING")
    print("=" * 60)
    print(f"Started: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}")
    
    # Step 1: Run the original dividend sheet update logic
    print("\n📊 STEP 1: Updating Original Dividend Sheet...")
    print("-" * 50)
    
    try:
        # Import and run the original update logic
        from modules.Update_dividend_sheet import update_dividend_sheet
        print("   🔌 Running original E*TRADE dividend sheet update...")
        
        # This will update your dividend_stocks.xlsx with current prices and yields
        update_dividend_sheet()
        print("   ✅ Original dividend sheet updated successfully")
        
    except Exception as e:
        print(f"   ⚠️ Original update had issues: {e}")
        print("   🔄 Continuing with ticker tracking anyway...")
    
    # Step 2: Extract data from updated dividend_stocks.xlsx
    print("\n📈 STEP 2: Extracting Updated Data...")
    print("-" * 50)
    
    dividend_file = "dividend_stocks.xlsx"  # Your original file
    extracted_data = extract_dividend_sheet_data(dividend_file)
    
    # Step 3: Update our comprehensive tracking system
    print("\n🎯 STEP 3: Updating Comprehensive Tracking...")
    print("-" * 50)
    
    update_comprehensive_tracking(extracted_data)
    
    # Step 4: Create integrated analysis in Dividends_2025.xlsx
    print("\n📊 STEP 4: Creating Integrated Analysis...")
    print("-" * 50)
    
    create_integrated_analysis(extracted_data)
    
    # Step 5: Show console completion display
    print("\n🎯 STEP 5: Showing Completion Display...")
    print("-" * 50)
    
    try:
        # Use console-based display (doesn't interfere with other apps)
        from console_completion_display import show_console_completion_display
        print("🎯 Processing complete - showing results!")
        show_console_completion_display()
    except Exception as e:
        print(f"⚠️ Console display failed: {e}")
        print("📊 Check outputs/Dividends_2025.xlsx for results")
    
    print(f"\n✅ ENHANCED UPDATE COMPLETED")
    print(f"Finished: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}")
    print("📊 Your dividend_stocks.xlsx AND comprehensive tracking are now updated!")
    print("=" * 60)

def extract_dividend_sheet_data(dividend_file):
    """
    Extract all data from the updated dividend_stocks.xlsx
    """
    
    extracted_data = {
        'tickers': {},
        'yield_history': {},
        'current_date': datetime.now().strftime("%m-%d-%Y")
    }
    
    if not os.path.exists(dividend_file):
        print(f"   ⚠️ {dividend_file} not found")
        return extracted_data
    
    try:
        print(f"   📂 Reading updated data from {dividend_file}...")
        
        # Load the Excel file
        wb = openpyxl.load_workbook(dividend_file, data_only=True)
        ws = wb.active
        
        # Get header row to find columns by name
        headers = {}
        for col in range(1, ws.max_column + 1):
            header_value = ws.cell(row=1, column=col).value
            if header_value:
                headers[str(header_value)] = col
        
        # Helper function to get column index by header name
        def get_col(header_name, fallback_col=None):
            return headers.get(header_name, fallback_col)
        
        # Find the most recent date column (look for today's date)
        today_date = None
        today_column = None
        for header_name, col_num in headers.items():
            if '-2025' in str(header_name) or '/2025' in str(header_name):
                # This is a date column - use the first one found (should be most recent)
                if today_column is None:
                    today_column = col_num
                    today_date = header_name
                    break
        
        if today_column is None:
            # Fallback to looking for any yield column
            today_column = get_col('08-02-2025', 18)  # Default to most recent known date
            today_date = '08-02-2025'
        
        print(f"   📅 Found new data for date: {today_date}")
        
        # Extract ticker data
        for row in range(2, ws.max_row + 1):
            ticker = ws.cell(row=row, column=get_col('Ticker', 1)).value  # Column A: Ticker
            
            if ticker and isinstance(ticker, str) and ticker.strip():
                ticker = ticker.strip()
                
                # Extract all ticker information using header-based column lookup
                extracted_data['tickers'][ticker] = {
                    'quantity': ws.cell(row=row, column=get_col('Qty #', 3)).value or 0,
                    'price_paid': ws.cell(row=row, column=get_col('Price Paid $', 4)).value or 0,
                    'current_price': ws.cell(row=row, column=get_col('Current Price $', 5)).value or 0,
                    'day_change': ws.cell(row=row, column=get_col('Day Change $', 6)).value or 0,
                    'price_change': ws.cell(row=row, column=get_col('Price Change $', 7)).value or 0,
                    'change_percent': ws.cell(row=row, column=get_col('Change %', 8)).value or 0,
                    'current_value': ws.cell(row=row, column=get_col('Current Value $', 9)).value or 0,
                    'original_value': ws.cell(row=row, column=get_col('Original Value $', 10)).value or 0,
                    'total_gain': ws.cell(row=row, column=get_col('Total Gain $', 11)).value or 0,
                    'total_gain_percent': ws.cell(row=row, column=get_col('Total Gain %', 12)).value or 0,
                    'payment_cycle': ws.cell(row=row, column=get_col('Payment Cycle', 13)).value or 'Monthly',
                    'rate_per_share': ws.cell(row=row, column=get_col('Rate Per Share', 14)).value or 0,
                    'original_payment': ws.cell(row=row, column=get_col('Original Payment', 15)).value or 0,
                    'new_payment': ws.cell(row=row, column=get_col('New Payment', 16)).value or 0,
                    'original_yield': ws.cell(row=row, column=get_col('Original Yield %', 17)).value or 0,
                    'current_yield': ws.cell(row=row, column=today_column).value or 0,  # Today's yield
                    'account': ws.cell(row=row, column=get_col('Account', 2)).value or 'Unknown'
                }
                
                # Initialize yield history for this ticker
                extracted_data['yield_history'][ticker] = {}
                
                # Extract historical yield data from all date columns (17 onwards)
                for col in range(17, ws.max_column + 1):
                    header = ws.cell(row=1, column=col).value
                    yield_value = ws.cell(row=row, column=col).value
                    
                    if header and yield_value is not None:
                        # Convert header to standard date format if needed
                        date_str = standardize_date_format(str(header))
                        extracted_data['yield_history'][ticker][date_str] = float(yield_value)
        
        wb.close()
        print(f"   ✅ Extracted data for {len(extracted_data['tickers'])} tickers")
        
    except Exception as e:
        print(f"   ❌ Error extracting data: {e}")
    
    return extracted_data

def standardize_date_format(date_str):
    """Convert date string to MM-DD-YYYY format"""
    try:
        # Handle various date formats
        if '/' in date_str:
            parts = date_str.split('/')
            if len(parts) == 3:
                return f"{parts[0].zfill(2)}-{parts[1].zfill(2)}-{parts[2]}"
        elif '-' in date_str and len(date_str.split('-')) == 3:
            parts = date_str.split('-')
            if len(parts[0]) == 4:  # YYYY-MM-DD format
                return f"{parts[1].zfill(2)}-{parts[2].zfill(2)}-{parts[0]}"
            else:  # MM-DD-YYYY format
                return date_str
    except:
        pass
    return date_str

def update_comprehensive_tracking(extracted_data):
    """
    Update our comprehensive tracking system with the extracted data
    """
    
    # This integrates the data into our Dividends_2025.xlsx system
    output_file = os.path.join("outputs", "Dividends_2025.xlsx")
    
    # Load or create workbook
    wb = openpyxl.load_workbook(output_file) if os.path.exists(output_file) else openpyxl.Workbook()
    
    # Update/create the integrated ticker sheet
    sheet_name = "Ticker Analysis 2025"
    if sheet_name in wb.sheetnames:
        del wb[sheet_name]
    
    ws = wb.create_sheet(sheet_name)
    
    print("   📊 Creating comprehensive ticker analysis...")
    
    # Get all historical dates and sort them
    all_dates = set()
    for ticker_history in extracted_data['yield_history'].values():
        all_dates.update(ticker_history.keys())
    
    sorted_dates = sorted(list(all_dates), 
                         key=lambda x: datetime.strptime(x, "%m-%d-%Y"), reverse=True)
    
    # Define headers to match actual Excel structure
    static_headers = [
        'Ticker', 'Account', 'Qty #', 'Price Paid $', 'Current Price $', 'Day Change $',
        'Price Change $', 'Change %', 'Current Value $', 'Original Value $', 
        'Total Gain $', 'Total Gain %', 'Payment Cycle', 'Rate Per Share',
        'Original Payment', 'New Payment', 'Original Yield %'
    ]
    
    all_headers = static_headers + sorted_dates
    
    # Create headers
    for col, header in enumerate(all_headers, 1):
        cell = ws.cell(row=1, column=col)
        cell.value = header
        cell.font = Font(bold=True, color="FFFFFF")
        cell.alignment = Alignment(horizontal="center")
        
        if col <= len(static_headers):
            cell.fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        else:
            cell.fill = PatternFill(start_color="70AD47", end_color="70AD47", fill_type="solid")
    
    # Add data rows
    row = 2
    totals = {'current_value': 0, 'original_value': 0, 'new_payment': 0}
    
    for ticker, data in extracted_data['tickers'].items():
        # Static data
        static_data = [
            ticker,
            data['quantity'],
            data['price_paid'],
            data['last_price'],
            data['day_gain'],
            data['change_amount'],
            data['change_percent'],
            data['current_value'],
            data['original_value'],
            data['total_gain_percent'],
            data['pay_date'],
            data['payment_cycle'],
            data['rate_per_share'],
            data['original_payment'],
            data['new_payment'],
            data['beginning_yield']
        ]
        
        # Add static data
        for col, value in enumerate(static_data, 1):
            cell = ws.cell(row=row, column=col)
            cell.value = value
            
            # Format currency and percentage columns
            if col in [3, 4, 5, 6, 8, 9, 13, 14, 15]:  # Currency columns
                if isinstance(value, (int, float)):
                    cell.number_format = '"$"#,##0.00'
            elif col in [7, 10, 16]:  # Percentage columns
                if isinstance(value, (int, float)):
                    cell.number_format = '0.00%'
        
        # Add historical yield data with color coding
        ticker_history = extracted_data['yield_history'].get(ticker, {})
        
        for date_idx, date_str in enumerate(sorted_dates):
            col = len(static_headers) + date_idx + 1
            yield_value = ticker_history.get(date_str, 0)
            
            cell = ws.cell(row=row, column=col)
            cell.value = yield_value
            cell.number_format = '0.00'
            
            # Color code based on change from previous period
            if date_idx > 0:
                prev_date = sorted_dates[date_idx - 1]
                prev_yield = ticker_history.get(prev_date, yield_value)
                
                yield_diff = yield_value - prev_yield
                if yield_diff > 0.1:
                    cell.fill = PatternFill(start_color="00FF00", end_color="00FF00", fill_type="solid")  # Green
                elif yield_diff < -0.1:
                    cell.fill = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")  # Red
                else:
                    cell.fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")  # Yellow
        
        # Update totals
        totals['current_value'] += data['current_value']
        totals['original_value'] += data['original_value'] 
        totals['new_payment'] += data['new_payment']
        
        row += 1
    
    # Add totals row
    totals_row = row + 1
    ws.cell(row=totals_row, column=1, value="TOTALS").font = Font(bold=True)
    
    for col, value in [(8, totals['current_value']), (9, totals['original_value']), (15, totals['new_payment'])]:
        cell = ws.cell(row=totals_row, column=col)
        cell.value = value
        cell.font = Font(bold=True)
        cell.number_format = '"$"#,##0.00'
    
    # Auto-adjust columns
    for col in range(1, len(all_headers) + 1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(col)].width = 12
    
    # Save workbook
    wb.save(output_file)
    
    print(f"   ✅ Updated comprehensive tracking with {len(extracted_data['tickers'])} tickers")
    print(f"   💰 Total Current Value: ${totals['current_value']:,.2f}")
    print(f"   💰 Total Monthly Payment: ${totals['new_payment']:,.2f}")

def create_integrated_analysis(extracted_data):
    """
    Create integrated analysis showing the connection between systems
    """
    
    print("   🔗 Creating integrated analysis report...")
    
    # Calculate summary statistics
    ticker_count = len(extracted_data['tickers'])
    total_current_value = sum(data['current_value'] for data in extracted_data['tickers'].values())
    total_dividends = sum(data['new_payment'] for data in extracted_data['tickers'].values())
    
    print(f"   📊 Analysis Complete:")
    print(f"      • Tickers Tracked: {ticker_count}")
    print(f"      • Total Value: ${total_current_value:,.2f}")
    print(f"      • Monthly Dividends: ${total_dividends:,.2f}")
    print(f"      • Yield History: {len(extracted_data['yield_history'])} tickers with historical data")

if __name__ == "__main__":
    enhanced_update_dividend_sheet_with_tracking()
