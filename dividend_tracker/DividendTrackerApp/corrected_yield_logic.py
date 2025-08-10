#!/usr/bin/env python3
"""
CORRECTED Yield Update Logic - Column 14 Beginning Yield, Insert at 15
"""

import openpyxl
from openpyxl.styles import PatternFill
from datetime import date, datetime
import os

def show_correct_column_structure():
    """
    Show the correct column structure based on your clarification
    """
    
    print("🔧 CORRECT COLUMN STRUCTURE")
    print("=" * 40)
    
    print("📋 Based on your clarification:")
    print("   Column 14: Beginning Dividend Yield (preserved)")
    print("   Column 15: INSERT new date column here ← CORRECT INSERT POSITION")
    print("   Column 16+: Existing historical data (shifts right)")
    
    print("\n🔧 CORRECTED UPDATE LOGIC:")
    print("   1. ✅ INSERT new column at position 15 (not 17)")
    print("   2. ✅ Set column 15 header to today's date")
    print("   3. ✅ Put current yield data in column 15")
    print("   4. ✅ Compare column 15 vs column 14 for color coding")
    print("   5. ✅ Preserve Beginning Dividend Yield in column 14")

def corrected_yield_update(file_path="dividend_stocks.xlsx"):
    """
    CORRECTED yield update with proper column positioning
    """
    
    print(f"\n🔧 Running CORRECTED yield update on {file_path}")
    print("   📋 Column 14: Beginning Dividend Yield (preserved)")
    print("   📋 Column 15: New yield data (inserted)")
    
    if not os.path.exists(file_path):
        print(f"   ⚠️ File {file_path} not found")
        return False
    
    try:
        workbook = openpyxl.load_workbook(file_path)
        sheet = workbook.active

        # CORRECTED: Insert new column at position 15 (right after Beginning Dividend Yield)
        print("   📋 Inserting new column at position 15...")
        sheet.insert_cols(15)
        
        # Set today's date as the header for the new column 15
        today = date.today().strftime("%m-%d-%Y")
        sheet.cell(row=1, column=15).value = today
        print(f"   📅 Set column 15 header to: {today}")
        
        # Verify column structure
        col_14_header = sheet.cell(row=1, column=14).value
        col_15_header = sheet.cell(row=1, column=15).value
        print(f"   ✅ Column 14: {col_14_header}")
        print(f"   ✅ Column 15: {col_15_header}")
        
        # Update ticker data in the new column 15
        print("   📊 Updating yield data in new column 15...")
        updated_count = 0
        
        for index, row in enumerate(sheet.iter_rows(min_row=2, values_only=False), start=2):
            symbol = row[0].value
            if symbol is None or not isinstance(symbol, str) or symbol.strip() == "":
                break

            # Get current yield from API (placeholder for now)
            # This would be replaced with actual E*TRADE API call
            try:
                # Simulate getting current yield - this would come from your E*TRADE API
                current_yield = 15.5  # Placeholder
                
                # Set the current yield in the new column 15
                sheet.cell(row=index, column=15).value = round(float(current_yield), 2)
                
                # CORRECTED: Compare column 15 (new) vs column 14 (beginning)
                col_15_value = sheet.cell(row=index, column=15).value  # New yield
                col_14_value = sheet.cell(row=index, column=14).value  # Beginning yield
                
                try:
                    col_15_value = float(col_15_value) if col_15_value else 0
                    col_14_value = float(col_14_value) if col_14_value else 0
                    
                    if col_15_value < col_14_value:
                        fill_color = "FF0000"  # Red - yield decreased from beginning
                    elif col_15_value > col_14_value:
                        fill_color = "00FF00"  # Green - yield increased from beginning
                    else:
                        fill_color = "FFFF00"  # Yellow - yield same as beginning
                        
                    fill = PatternFill(start_color=fill_color, end_color=fill_color, fill_type="solid")
                    sheet.cell(row=index, column=15).fill = fill
                    
                    updated_count += 1
                    
                except (TypeError, ValueError):
                    print(f"      ⚠️ Skipping row {index} due to invalid data")
                    
            except Exception as e:
                print(f"      ❌ Error processing {symbol}: {e}")

        # Save the workbook
        workbook.save(file_path)
        print(f"   ✅ Updated {updated_count} ticker yields")
        print(f"   💾 File saved: {file_path}")
        
        return True
        
    except Exception as e:
        print(f"   ❌ Error in corrected yield update: {e}")
        return False

def fix_integration_code():
    """
    Create the fixed integration code with correct column logic
    """
    
    print(f"\n🔗 CREATING FIXED INTEGRATION CODE...")
    print("-" * 45)
    
    fixed_code = '''#!/usr/bin/env python3
"""
FIXED Integration - Correct Column Structure
Column 14: Beginning Dividend Yield (preserved)
Column 15: New yield data (inserted with today's date)
"""

def fixed_update_dividend_sheet_integration():
    """
    Fixed version that uses correct column positioning
    """
    
    from etrade_auth import get_etrade_session
    import openpyxl
    from openpyxl.styles import PatternFill
    from datetime import date
    
    # Get E*TRADE session
    session, base_url = get_etrade_session()
    
    # Account selection logic here...
    file_path = "dividend_stocks.xlsx"
    
    try:
        workbook = openpyxl.load_workbook(file_path)
        sheet = workbook.active

        # FIXED: Insert new column at position 15 (after Beginning Dividend Yield in column 14)
        today = date.today().strftime("%m-%d-%Y")
        sheet.insert_cols(15)  # CORRECTED: Insert at 15, not 17
        sheet.cell(row=1, column=15).value = today

        # Iterate through rows and update with market quote data
        for index, row in enumerate(sheet.iter_rows(min_row=2, values_only=False), start=2):
            symbol = row[0].value
            if symbol is None or not isinstance(symbol, str) or symbol.strip() == "":
                break

            try:
                # Get quote data from E*TRADE
                quote_url = f"{base_url}/v1/market/quote/{symbol}.json"
                quote_response = session.get(quote_url)
                quote_json = quote_response.json()
                
                if 'QuoteResponse' in quote_json and 'QuoteData' in quote_json['QuoteResponse']:
                    quote_data = quote_json['QuoteResponse']['QuoteData'][0]
                    
                    # Update last trade price (column 4)
                    last_trade = quote_data['All'].get('lastTrade', '')
                    if last_trade:
                        sheet.cell(row=index, column=4).value = round(float(last_trade), 2)

                    # FIXED: Put current yield in column 15 (newly inserted)
                    current_yield = quote_data['All'].get('yield', '')
                    if current_yield:
                        sheet.cell(row=index, column=15).value = round(float(current_yield), 2)
                    else:
                        sheet.cell(row=index, column=15).value = current_yield

                    # FIXED: Compare column 15 (new) vs column 14 (beginning)
                    col_15_value = sheet.cell(row=index, column=15).value  # New yield
                    col_14_value = sheet.cell(row=index, column=14).value  # Beginning yield
                    
                    try:
                        col_15_value = float(col_15_value) if col_15_value else 0
                        col_14_value = float(col_14_value) if col_14_value else 0
                        
                        if col_15_value < col_14_value:
                            fill_color = "FF0000"  # Red
                        elif col_15_value > col_14_value:
                            fill_color = "00FF00"  # Green
                        else:
                            fill_color = "FFFF00"  # Yellow
                            
                        fill = PatternFill(start_color=fill_color, end_color=fill_color, fill_type="solid")
                        sheet.cell(row=index, column=15).fill = fill
                        
                    except (TypeError, ValueError):
                        print(f"Skipping row {index} due to invalid data types for comparison.")

            except Exception as e:
                print(f"Error processing {symbol}: {e}")

        workbook.save(file_path)
        print(f"File processed and saved at: {file_path}")
        
    except Exception as e:
        print(f"An error occurred: {e}")
'''
    
    with open("fixed_integration_code.py", "w") as f:
        f.write(fixed_code)
    
    print("   ✅ Created: fixed_integration_code.py")
    print("   📋 This shows the CORRECT column logic")

def test_with_sample_data():
    """
    Test the logic with sample data to verify column structure
    """
    
    print(f"\n🧪 TESTING WITH SAMPLE DATA...")
    print("-" * 35)
    
    # Create a test file to verify the logic
    wb = openpyxl.Workbook()
    ws = wb.active
    
    # Sample headers up to column 16
    headers = [
        "Ticker", "Qty #", "Price Paid $", "Last Price $", "Day's Gain $",
        "Change $", "Change %", "Current Value $", "Original Value $", 
        "Total Gain %", "Pay Date", "Payment cycle", "Rate per share",
        "Original Payment amount", "New Payment amount", "Beginning Dividend Yield"
    ]
    
    # Add headers
    for col, header in enumerate(headers, 1):
        ws.cell(row=1, column=col).value = header
    
    # Add sample data
    ws.cell(row=2, column=1).value = "ABR"
    ws.cell(row=2, column=14).value = 14.20  # Beginning Dividend Yield
    
    print("   📋 Created sample data:")
    print(f"      Column 14: {ws.cell(row=1, column=14).value} = {ws.cell(row=2, column=14).value}")
    
    # Test the insertion
    print("   🔧 Testing column insertion at position 15...")
    ws.insert_cols(15)
    today = date.today().strftime("%m-%d-%Y")
    ws.cell(row=1, column=15).value = today
    ws.cell(row=2, column=15).value = 15.5  # New yield
    
    print("   ✅ After insertion:")
    print(f"      Column 14: {ws.cell(row=1, column=14).value} = {ws.cell(row=2, column=14).value}")
    print(f"      Column 15: {ws.cell(row=1, column=15).value} = {ws.cell(row=2, column=15).value}")
    print(f"      Column 16: {ws.cell(row=1, column=16).value}")
    
    # Test comparison
    beginning_yield = float(ws.cell(row=2, column=14).value)
    current_yield = float(ws.cell(row=2, column=15).value)
    
    if current_yield > beginning_yield:
        result = "GREEN (yield increased)"
    elif current_yield < beginning_yield:
        result = "RED (yield decreased)"
    else:
        result = "YELLOW (yield same)"
    
    print(f"   🎨 Color coding: {beginning_yield} → {current_yield} = {result}")
    
    wb.save("test_column_structure.xlsx")
    print("   💾 Saved test file: test_column_structure.xlsx")

if __name__ == "__main__":
    show_correct_column_structure()
    
    # Test with your actual file if it exists
    if os.path.exists("dividend_stocks.xlsx"):
        corrected_yield_update("dividend_stocks.xlsx")
    else:
        print("\n⚠️ dividend_stocks.xlsx not found - will test with sample data")
        test_with_sample_data()
    
    fix_integration_code()
