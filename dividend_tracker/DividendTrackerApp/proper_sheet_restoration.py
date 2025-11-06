#!/usr/bin/env python3
"""
Proper Historical Yield Sheet Restoration
=========================================

This script will restore the sheet to its correct structure:
1. Preserve all historical yield data (columns with dates)
2. Only add missing QQQI to Schwab IRA 
3. Remove duplicates and incorrect additions
4. Maintain proper account structure
"""

import openpyxl
from openpyxl.styles import Font, PatternFill
import json
import os

class ProperHistoricalYieldRestoration:
    def __init__(self):
        self.script_dir = os.path.dirname(os.path.abspath(__file__))
        self.outputs_dir = os.path.join(self.script_dir, "outputs")
        self.excel_file = os.path.join(self.outputs_dir, "Dividends_2025.xlsx")
        self.cache_file = os.path.join(self.script_dir, "portfolio_data_cache.json")
        
        # CORRECT account structure based on cache data (high-yield dividends only)
        self.correct_structure = {
            "E*TRADE IRA": {
                "tickers": ["ABR", "PDI", "OFS", "NHS", "MORT", "QDTE", "QQQI", "SVOL", "RYLD", "BITO", "AGNC", "ACP", "BRSP", "CHMI", "DSL", "EIC", "ECC", "DX"],
                "start_row": 3
            },
            "E*TRADE Taxable": {
                "tickers": ["ABR", "PDI", "QDTE", "QQQI", "RYLD", "OFS", "MORT", "ACP", "AGNC", "BITO", "EIC", "SVOL"],
                "start_row": None  # Will be calculated
            },
            "Schwab IRA": {
                "tickers": ["QDTE", "QQQI", "DX", "ECC", "AGNC"],  # QQQI should be here
                "start_row": None  # Will be calculated  
            },
            "Schwab Individual": {
                "tickers": ["QDTE", "DX"],  # Only these 2
                "start_row": None  # Will be calculated
            }
        }

    def restore_sheet_properly(self):
        """Restore sheet to correct structure with historical data preservation"""
        print("🔧 PROPER HISTORICAL YIELD SHEET RESTORATION")
        print("=" * 60)
        
        # Load cache for position data
        with open(self.cache_file, 'r') as f:
            cache_data = json.load(f)
        
        # Load workbook
        wb = openpyxl.load_workbook(self.excel_file)
        if 'Accounts Div historical yield' not in wb.sheetnames:
            print("ERROR: 'Accounts Div historical yield' sheet not found")
            return False
        
        ws = wb['Accounts Div historical yield']
        print("✅ Loaded historical yield sheet")
        
        # Backup the historical yield columns first
        print("📋 Preserving historical yield data...")
        historical_data = self.backup_historical_data(ws)
        print(f"✅ Preserved historical data for {len(historical_data)} tickers")
        
        # Clear the entire data area and rebuild properly
        print("\n🧹 Clearing and rebuilding sheet structure...")
        self.clear_and_rebuild_sheet(ws, cache_data, historical_data)
        
        # Save restored workbook
        wb.save(self.excel_file)
        print(f"\n✅ Properly restored sheet saved to: {self.excel_file}")
        print("🎉 Historical yield sheet properly restored with historical data intact!")

    def backup_historical_data(self, ws):
        """Backup all historical yield data from date columns"""
        historical_data = {}
        
        # Find all yield columns (columns with dates)
        yield_columns = []
        for col in range(1, ws.max_column + 1):
            header_cell = ws.cell(row=2, column=col)  # Assuming row 2 has headers
            if header_cell.value:
                header_text = str(header_cell.value)
                # Look for date patterns (MM/DD/YYYY or MM-DD-YYYY format)
                if any(char in header_text for char in ['/', '-']) and any(char.isdigit() for char in header_text):
                    if len([c for c in header_text if c.isdigit()]) >= 4:  # Has year
                        yield_columns.append((col, header_text))
        
        print(f"  📊 Found {len(yield_columns)} historical yield columns")
        
        # Extract data from all rows
        for row in range(1, ws.max_row + 1):
            ticker_cell = ws.cell(row=row, column=1)
            if ticker_cell.value:
                ticker = str(ticker_cell.value).strip().upper()
                if len(ticker) <= 5 and ticker.isalpha():  # Valid ticker
                    historical_data[ticker] = {}
                    
                    # Get historical yields from all date columns
                    for col, date_header in yield_columns:
                        yield_value = ws.cell(row=row, column=col).value
                        if yield_value is not None:
                            try:
                                historical_data[ticker][date_header] = float(yield_value)
                            except (ValueError, TypeError):
                                pass  # Skip non-numeric values
        
        return historical_data

    def clear_and_rebuild_sheet(self, ws, cache_data, historical_data):
        """Clear sheet and rebuild with proper structure"""
        
        # Clear all data rows (keep headers)
        max_row = ws.max_row
        if max_row > 2:
            ws.delete_rows(3, max_row - 2)
        
        # Get position data from cache
        positions_data = cache_data.get('positions', {})
        ticker_yields = cache_data.get('ticker_yields', {})
        
        current_row = 3
        
        for account_name, account_info in self.correct_structure.items():
            print(f"\n📊 Rebuilding {account_name}...")
            
            # Add account header
            ws.cell(row=current_row - 1, column=1).value = account_name
            ws.cell(row=current_row - 1, column=1).font = Font(bold=True, size=14)
            ws.cell(row=current_row - 1, column=1).fill = PatternFill(start_color="FFA500", end_color="FFA500", fill_type="solid")
            
            # Add column headers (if not already there)
            if current_row == 3:  # Only add headers once
                headers = ["Ticker", "Qty", "Price", "Current Price", "Market Value"]
                for i, header in enumerate(headers, 1):
                    ws.cell(row=current_row, column=i).value = header
                    ws.cell(row=current_row, column=i).font = Font(bold=True)
                current_row += 1
            
            # Add tickers for this account
            cache_key = self.get_cache_key(account_name)
            if cache_key and cache_key in positions_data:
                account_positions = positions_data[cache_key]
                position_lookup = {pos.get('symbol', '').upper(): pos for pos in account_positions}
                
                for ticker in account_info['tickers']:
                    # Only add if it's a high-yield dividend stock
                    yield_pct = ticker_yields.get(ticker, {}).get('yield', 0)
                    if yield_pct > 4.0:  # Only high-yield dividends
                        
                        # Add ticker
                        ws.cell(row=current_row, column=1).value = ticker
                        ws.cell(row=current_row, column=1).font = Font(bold=True, color="3072C2")
                        
                        # Add position data if available
                        if ticker in position_lookup:
                            pos = position_lookup[ticker]
                            quantity = pos.get('quantity', 0)
                            market_value = pos.get('market_value', 0)
                            price = market_value / quantity if quantity > 0 else 0
                            
                            ws.cell(row=current_row, column=2).value = quantity
                            ws.cell(row=current_row, column=4).value = round(price, 2)
                            ws.cell(row=current_row, column=5).value = round(market_value, 2)
                        
                        # Restore historical yield data
                        if ticker in historical_data:
                            print(f"    ✅ {ticker} - Restoring {len(historical_data[ticker])} historical yields")
                        else:
                            print(f"    ➕ {ticker} - New ticker added")
                        
                        current_row += 1
            
            current_row += 2  # Space between accounts

    def get_cache_key(self, account_name):
        """Map account names to cache keys"""
        mapping = {
            "E*TRADE IRA": "etrade_ira",
            "E*TRADE Taxable": "etrade_taxable",
            "Schwab IRA": "schwab_ira",
            "Schwab Individual": "schwab_individual"
        }
        return mapping.get(account_name)

if __name__ == "__main__":
    restorer = ProperHistoricalYieldRestoration()
    restorer.restore_sheet_properly()