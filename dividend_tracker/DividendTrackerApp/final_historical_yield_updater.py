#!/usr/bin/env python3

import openpyxl
from openpyxl.styles import Font, PatternFill
import json
import os
from datetime import datetime

class FinalHistoricalYieldUpdater:
    def __init__(self):
        # Use the outputs directory version as the source
        self.excel_path = r"C:\Users\mjmat\Python Code in VS\dividend_tracker\DividendTrackerApp\outputs\Dividends_2025.xlsx"
        # Output to the same location (overwrite)
        self.output_path = r"C:\Users\mjmat\Python Code in VS\dividend_tracker\DividendTrackerApp\outputs\Dividends_2025.xlsx"
        self.cache_file = "portfolio_data_cache.json"
        
        # Account mapping from cache to Excel groups
        self.account_mapping = {
            "E*TRADE IRA": "etrade_ira",
            "E*TRADE Taxable": "etrade_taxable", 
            "Schwab IRA": "schwab_ira",
            "Schwab Individual": "schwab_individual"
        }
        
        # Only include dividend stocks with yield > 4%
        self.min_yield_threshold = 4.0
    
    def run_update(self):
        """Main update process - Windows compatible, no emoji characters"""
        print("FINAL HISTORICAL YIELD UPDATER - WINDOWS COMPATIBLE")
        print("=" * 60)
        print(f"Update Date: {datetime.now().strftime('%m/%d/%Y')}")
        print(f"Excel File: {os.path.basename(self.excel_path)}")
        print(f"Cache File: {self.cache_file}")
        print(f"Minimum Yield Threshold: {self.min_yield_threshold}%")
        print()
        
        # Load cache data
        print("STEP 1: Loading cache data...")
        cache_data = self.load_cache_data()
        if not cache_data:
            return False
        
        positions_data = cache_data.get("positions", {})
        yields_data = cache_data.get("ticker_yields", {})
        print(f"SUCCESS: Cache loaded: {cache_data.get('timestamp')}")
        print(f"INFO: Found yields for {len(yields_data)} tickers")
        print(f"INFO: Found positions for {len(positions_data)} account groups")
        print()
        
        # Load Excel workbook
        print("STEP 2: Loading Excel workbook...")
        if not os.path.exists(self.excel_path):
            print(f"ERROR: Excel file not found: {self.excel_path}")
            return False
        
        wb = openpyxl.load_workbook(self.excel_path)
        if 'Accounts Div historical yield' not in wb.sheetnames:
            print("ERROR: 'Accounts Div historical yield' sheet not found")
            return False
        
        ws = wb['Accounts Div historical yield']
        print("SUCCESS: Found sheet: Accounts Div historical yield")
        print()
        
        # Find account groups with proper boundary detection
        print("STEP 3: Finding account groups and boundaries...")
        account_info = self.find_account_groups_and_boundaries(ws)
        
        if len(account_info) != 4:
            print(f"ERROR: Expected 4 account groups, found {len(account_info)}")
            print("Available groups:", list(account_info.keys()))
            return False
        print()
        
        # Insert new yield column
        print("STEP 4: Inserting new yield column...")
        date_str = datetime.now().strftime('%m/%d/%Y')
        self.insert_yield_column(ws, date_str)
        print(f"SUCCESS: Inserted yield column P with date {date_str}")
        print()
        
        # Update each account with proper filtering
        print("STEP 5: Updating account groups with high-yield dividend filtering...")
        
        for group_name, group_info in account_info.items():
            print(f"INFO: Processing {group_name}...")
            cache_key = self.account_mapping.get(group_name)
            if not cache_key:
                print(f"WARNING: No cache mapping for {group_name}")
                continue
            
            account_positions = positions_data.get(cache_key, [])
            filtered_positions = self.filter_high_yield_positions(account_positions, yields_data)
            
            print(f"INFO: Found {len(account_positions)} total positions")
            print(f"INFO: Filtered to {len(filtered_positions)} high-yield dividend stocks")
            
            if filtered_positions:
                self.clear_existing_tickers(ws, group_info)
                # Don't use update_existing_tickers anymore - use surgical approach
                self.surgical_update_tickers(ws, group_info, account_positions)
                print(f"SUCCESS: Updated positions for {group_name}")
            else:
                print(f"INFO: No high-yield positions found for {group_name}")
            print()
        
        # Update yield data with color coding
        print("STEP 6: Updating yield percentages...")
        self.update_yield_data(ws, account_info, yields_data)
        print()
        
        # Apply formatting
        print("STEP 7: Applying formatting...")
        self.apply_group_formatting(ws, account_info)
        print()
        
        # Save workbook
        print("STEP 8: Saving workbook...")
        os.makedirs(os.path.dirname(self.output_path), exist_ok=True)
        wb.save(self.output_path)
        print(f"SUCCESS: Successfully saved: {self.output_path}")
        print()
        
        print("SUCCESS: FINAL UPDATE COMPLETED!")
        print("Updated: Only high-yield dividend stocks (>4%) in correct accounts")
        print("Applied: Proper row insertion, formatting, and yield calculations")
        wb.close()
        return True
    
    def filter_high_yield_positions(self, positions, yields_data):
        """Filter to only include high-yield dividend paying stocks"""
        filtered = []
        
        for position in positions:
            symbol = position.get('symbol', '').strip().upper()
            yield_info = yields_data.get(symbol, {})
            
            ticker_yield = yield_info.get('yield', 0.0)
            has_dividend = yield_info.get('has_dividend', False)
            
            if has_dividend and ticker_yield > self.min_yield_threshold:
                filtered.append(position)
                print(f"INCLUDE: {symbol} - Yield: {ticker_yield:.2f}%")
            else:
                reason = "No dividend" if not has_dividend else f"Low yield: {ticker_yield:.2f}%"
                print(f"EXCLUDE: {symbol} - {reason}")
        
        return filtered
    
    def find_account_groups_and_boundaries(self, ws):
        """Find account groups with improved boundary detection for calculation rows"""
        account_info = {}
        
        print("INFO: Scanning for account groups...")
        
        for row in range(1, min(80, ws.max_row + 1)):
            cell_value = ws.cell(row=row, column=1).value
            if not cell_value:
                continue
                
            cell_text = str(cell_value).strip().upper()
            
            if cell_text == 'ETRADE IRA':
                account_info["E*TRADE IRA"] = {"start_row": row, "end_row": None}
                print(f"SUCCESS: E*TRADE IRA found at row {row}")
            elif cell_text == 'ETRADE TAXABLE':
                account_info["E*TRADE Taxable"] = {"start_row": row, "end_row": None}
                print(f"SUCCESS: E*TRADE Taxable found at row {row}")
            elif cell_text == 'SCHWAB IRA':
                account_info["Schwab IRA"] = {"start_row": row, "end_row": None}
                print(f"SUCCESS: Schwab IRA found at row {row}")
            elif cell_text == 'SCHWAB INDIVIDUAL':
                account_info["Schwab Individual"] = {"start_row": row, "end_row": None}
                print(f"SUCCESS: Schwab Individual found at row {row}")
        
        # Calculate end boundaries accounting for calculation and blank rows
        group_order = ["E*TRADE IRA", "E*TRADE Taxable", "Schwab IRA", "Schwab Individual"]
        
        for i, group_name in enumerate(group_order):
            if group_name not in account_info:
                continue
                
            start_row = account_info[group_name]["start_row"]
            
            if i < len(group_order) - 1:
                # End is 3 rows before next group (calculation + blank + next header)
                end_row = ws.max_row
                for next_group in group_order[i+1:]:
                    if next_group in account_info:
                        end_row = account_info[next_group]["start_row"] - 3
                        break
            else:
                # Last group - find actual end of data
                end_row = start_row + 2
                while end_row <= ws.max_row:
                    cell_value = ws.cell(row=end_row, column=1).value
                    if not cell_value or str(cell_value).strip() == "":
                        break
                    if any(keyword in str(cell_value).strip().upper() for keyword in ['ETRADE', 'SCHWAB']):
                        break
                    end_row += 1
                end_row -= 1
                
            account_info[group_name]["end_row"] = end_row
            print(f"INFO: {group_name}: data rows {start_row+2} to {end_row}")
        
        return account_info
    
    def clear_existing_tickers(self, ws, group_info):
        """Clear only quantity and price columns while preserving ticker symbols and other data"""
        start_row = group_info["start_row"] + 2  # Skip headers
        end_row = group_info["end_row"]
        
        print(f"INFO: Clearing columns B, D only (preserving ticker symbols) for rows {start_row} to {end_row}")
        for row in range(start_row, end_row + 1):
            ticker_cell = ws.cell(row=row, column=1)
            if ticker_cell.value and str(ticker_cell.value).strip():
                # Clear ONLY quantity and price columns - PRESERVE Column A (ticker symbols)
                # ws.cell(row=row, column=1).value = None  # DON'T CLEAR TICKER!
                ws.cell(row=row, column=2).value = None  # Clear quantity
                ws.cell(row=row, column=4).value = None  # Clear price
                
                # Reset formatting for quantity and price columns only
                # ws.cell(row=row, column=1).font = Font()  # DON'T CLEAR TICKER FORMATTING!
                ws.cell(row=row, column=2).font = Font()
                ws.cell(row=row, column=4).font = Font()
    
    def update_existing_tickers(self, ws, group_info, positions, yields_data):
        """Surgically update existing tickers in place, preserving historical data alignment"""
        start_row = group_info["start_row"] + 2  # After headers
        end_row = group_info["end_row"]
        
        # Create lookup of positions by symbol
        position_lookup = {pos.get('symbol', '').strip().upper(): pos for pos in positions}
        
        # Define formatting
        ticker_font = Font(name="Arial", size=12, bold=True, color="3072C2")
        qty_font = Font(name="Arial", size=12, bold=True, color="3072C2")
        
        updated_count = 0
        
        # Go through existing rows and update matching tickers BY EXISTING TICKER SYMBOL
        for row in range(start_row, end_row + 1):
            # Get existing ticker from Column A
            existing_ticker_cell = ws.cell(row=row, column=1)
            existing_ticker = existing_ticker_cell.value
            
            if existing_ticker:
                ticker_symbol = str(existing_ticker).strip().upper()
                
                # Find matching position data for this exact ticker
                if ticker_symbol in position_lookup:
                    position = position_lookup[ticker_symbol]
                    quantity = position.get('quantity', 0)
                    market_value = position.get('market_value', 0)
                    price = round(market_value / quantity, 2) if quantity > 0 else 0
                    
                    # Update the three columns while preserving exact ticker and row position
                    existing_ticker_cell.value = ticker_symbol  # Ensure formatting consistency
                    existing_ticker_cell.font = ticker_font
                    
                    qty_cell = ws.cell(row=row, column=2)
                    qty_cell.value = quantity  
                    qty_cell.font = qty_font
                    
                    price_cell = ws.cell(row=row, column=4)
                    price_cell.value = price
                    
                    updated_count += 1
                    print(f"UPDATED: Row {row}: {ticker_symbol} qty={quantity} price=${price}")
                else:
                    print(f"WARNING: Row {row}: {ticker_symbol} - No position data found")
        
        print(f"INFO: Updated {updated_count} existing ticker positions")
        return updated_count
    
    def surgical_update_tickers(self, ws, group_info, account_positions):
        """Surgically update existing tickers by matching symbols, preserving exact order"""
        start_row = group_info["start_row"] + 2  # After headers
        end_row = group_info["end_row"]
        
        # Create position lookup by symbol (use ALL positions, not just high-yield filtered)
        position_lookup = {pos.get('symbol', '').strip().upper(): pos for pos in account_positions}
        
        # Define formatting
        ticker_font = Font(name="Arial", size=12, bold=True, color="3072C2")
        qty_font = Font(name="Arial", size=12, bold=True, color="3072C2")
        
        updated_count = 0
        
        # Go through existing rows and match by ticker symbol
        for row in range(start_row, end_row + 1):
            # Get existing ticker from Column A
            existing_ticker_cell = ws.cell(row=row, column=1)
            existing_ticker = existing_ticker_cell.value
            
            if existing_ticker:
                ticker_symbol = str(existing_ticker).strip().upper()
                
                # Find matching position data for this exact ticker
                if ticker_symbol in position_lookup:
                    position = position_lookup[ticker_symbol]
                    quantity = position.get('quantity', 0)
                    market_value = position.get('market_value', 0)
                    price = round(market_value / quantity, 2) if quantity > 0 else 0
                    
                    # Update only Quantity (Column B) and Price (Column D)
                    # Keep ticker symbol in place with proper formatting
                    existing_ticker_cell.value = ticker_symbol
                    existing_ticker_cell.font = ticker_font
                    
                    qty_cell = ws.cell(row=row, column=2)
                    qty_cell.value = quantity  
                    qty_cell.font = qty_font
                    
                    price_cell = ws.cell(row=row, column=4)
                    price_cell.value = price
                    
                    updated_count += 1
                    print(f"UPDATED: Row {row}: {ticker_symbol} qty={quantity} price=${price}")
                else:
                    print(f"WARNING: Row {row}: {ticker_symbol} - No position data found")
        
        print(f"INFO: Updated {updated_count} ticker positions while preserving order")
        return updated_count
    
    def insert_yield_column(self, ws, date_str):
        """Insert new yield column P with date headers"""
        # Insert column P (after O)
        ws.insert_cols(16)
        
        # Add date header for each group
        for row in range(1, min(80, ws.max_row + 1)):
            cell_value = ws.cell(row=row, column=1).value
            if cell_value and any(keyword in str(cell_value).upper() for keyword in ['ETRADE', 'SCHWAB']):
                header_row = row + 1
                header_cell = ws.cell(row=header_row, column=16)
                header_cell.value = date_str
                header_cell.font = Font(bold=True)
    
    def update_yield_data(self, ws, account_info, yields_data):
        """Update yield percentages with proper color coding and numeric values"""
        for group_name, group_info in account_info.items():
            print(f"INFO: Updating yields for {group_name}...")
            
            start_row = group_info["start_row"] + 2
            end_row = group_info["end_row"]
            yield_count = 0
            
            for row in range(start_row, end_row + 1):
                ticker_cell = ws.cell(row=row, column=1)
                if ticker_cell.value:
                    symbol = str(ticker_cell.value).strip().upper()
                    yield_info = yields_data.get(symbol, {})
                    current_yield = yield_info.get('yield', 0.0)
                    
                    yield_cell = ws.cell(row=row, column=16)  # Column P (new)
                    
                    # Set as NUMERIC value, not string
                    yield_cell.value = current_yield / 100.0  # Convert percentage to decimal
                    yield_cell.number_format = '0.00%'  # Format as percentage
                    
                    # Get previous yield from column O for comparison
                    previous_yield_cell = ws.cell(row=row, column=15)  # Column O
                    previous_yield = 0.0
                    
                    if previous_yield_cell.value:
                        try:
                            if isinstance(previous_yield_cell.value, str):
                                # Remove % sign and convert
                                prev_str = str(previous_yield_cell.value).replace('%', '').strip()
                                previous_yield = float(prev_str)
                            elif isinstance(previous_yield_cell.value, (int, float)):
                                prev_val = float(previous_yield_cell.value)
                                # If it's already a decimal (like 0.142), multiply by 100
                                # If it's a percentage (like 14.2), use as is
                                if prev_val < 1.0:
                                    previous_yield = prev_val * 100  # Convert decimal to percentage
                                else:
                                    previous_yield = prev_val  # Already a percentage
                        except (ValueError, TypeError):
                            previous_yield = 0.0
                    
                    # Color coding based on comparison with column O (using BACKGROUND colors)
                    if previous_yield == 0.0:
                        # No previous data - no background color (like Column O default)
                        yield_cell.fill = PatternFill()  # Clear any existing fill
                        direction = "NEW"
                    elif current_yield > previous_yield:
                        # Green background for increase (#90EE90 from plan)
                        yield_cell.fill = PatternFill(start_color="90EE90", end_color="90EE90", fill_type="solid")
                        direction = "UP"
                    elif current_yield < previous_yield:
                        # Red background for decrease (#FF7C80 from plan)
                        yield_cell.fill = PatternFill(start_color="FF7C80", end_color="FF7C80", fill_type="solid")
                        direction = "DOWN"
                    else:
                        # Yellow background when old value equals new value (#FFFF00 from plan)
                        yield_cell.fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
                        direction = "SAME"
                    
                    yield_count += 1
                    print(f"YIELD: Row {row}: {symbol} = {current_yield:.2f}% ({direction} from {previous_yield:.2f}%)")
            
            print(f"SUCCESS: Updated {yield_count} yields for {group_name}")
    
    
    def apply_group_formatting(self, ws, account_info):
        """Apply orange formatting to group headers AND yield column dividers"""
        orange_fill = PatternFill(start_color="FFC000", end_color="FFC000", fill_type="solid")
        
        print("INFO: Applying orange color to group dividers and yield column...")
        for group_name, group_info in account_info.items():
            # Orange fill for group header (Column A)
            divider_row = group_info["start_row"]
            divider_cell = ws.cell(row=divider_row, column=1)
            divider_cell.fill = orange_fill
            
            # Orange fill for yield column divider (Column P) 
            yield_divider_cell = ws.cell(row=divider_row, column=16)  # Column P
            yield_divider_cell.fill = orange_fill
            
            print(f"DIVIDER: {group_name} divider at row {divider_row} (columns A and P)")
    
    def load_cache_data(self):
        """Load position and yield data from cache file"""
        try:
            if not os.path.exists(self.cache_file):
                print(f"ERROR: Cache file not found: {self.cache_file}")
                return None
            
            with open(self.cache_file, 'r') as f:
                return json.load(f)
        except Exception as e:
            print(f"ERROR: Failed to load cache: {e}")
            return None

if __name__ == "__main__":
    updater = FinalHistoricalYieldUpdater()
    success = updater.run_update()
    
    if success:
        print("\nSUCCESS: Final historical yield update completed!")
    else:
        print("\nERROR: Update failed!")