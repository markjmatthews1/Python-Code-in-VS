#!/usr/bin/env python3
"""
Rename Sheet to "Accounts Div historical yield"
==============================================
"""

import openpyxl
import os

def rename_sheet_to_shorter():
    """Rename the sheet to fit Excel's 31-character limit"""
    
    excel_path = r"c:\Users\mjmat\Python Code in VS\dividend_tracker\DividendTrackerApp\outputs\Dividends_2025.xlsx"
    
    try:
        workbook = openpyxl.load_workbook(excel_path)
        
        # Find the current sheet (might be truncated)
        target_sheet = None
        for sheet_name in workbook.sheetnames:
            if 'dividend' in sheet_name.lower() and 'historical' in sheet_name.lower():
                target_sheet = workbook[sheet_name]
                old_name = sheet_name
                break
        
        if target_sheet:
            print(f"📋 Current sheet name: '{old_name}' ({len(old_name)} characters)")
            
            # Set the new name
            new_name = "Accounts Div historical yield"
            target_sheet.title = new_name
            
            print(f"✅ Renamed to: '{new_name}' ({len(new_name)} characters)")
            
            # Save the changes
            workbook.save(excel_path)
            workbook.close()
            
            print(f"💾 File saved with new sheet name")
            
        else:
            print(f"❌ Could not find dividend historical sheet")
            print(f"📋 Available sheets: {workbook.sheetnames}")
            
    except Exception as e:
        print(f"❌ Error: {e}")

if __name__ == "__main__":
    rename_sheet_to_shorter()
