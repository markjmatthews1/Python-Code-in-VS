r"""
Module: comprehensive_retirement_tracker.py
Author: Mark
Created: July 26, 2025
Purpose: Complete retirement tracking system matching user's Excel format
Location: C:\Python_Projects\DividendTrackerApp\modules\comprehensive_retirement_tracker.py
"""

import os
import pandas as pd
import openpyxl
from openpyxl.styles import Font, Border, Side, PatternFill
from openpyxl.styles.numbers import FORMAT_CURRENCY_USD_SIMPLE, FORMAT_PERCENTAGE
from datetime import datetime, timedelta
from modules.portfolio_value_tracker import PortfolioValueTracker
from modules.hybrid_dividend_tracker import get_hybrid_dividend_data
from modules.portfolio_config import PortfolioConfig

# Styling constants
HEADER_FONT = Font(bold=True, name="Arial", size=12)
NORMAL_FONT = Font(name="Arial", size=12)
SMALL_FONT = Font(name="Arial", size=10)
BOX_BORDER = Border(
    left=Side(style='thin'), right=Side(style='thin'),
    top=Side(style='thin'), bottom=Side(style='thin')
)
GREEN_FILL = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
RED_FILL = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")

class ComprehensiveRetirementTracker:
    """Complete retirement financial tracking system"""
    
    def __init__(self):
        self.portfolio_config = PortfolioConfig()
        self.retirement_date = datetime(2024, 5, 16)  # 5/16/24
        self.retirement_value = 452975.00  # Starting retirement value
        
        # Static values from your data
        self.marks_ss_gross = 2590.00
        self.marks_medicare = 201.30
        self.marks_ss_net = 2388.70
        
        self.carolyns_ss_gross = 1295.00
        self.carolyns_medicare = 185.00
        self.carolyns_ss_net = 1110.00
        
        self.total_ss_net = 3498.70
        
    def create_comprehensive_tracking_sheet(self, wb):
        """Create the complete retirement tracking sheet matching your format"""
        
        # Remove existing sheet if it exists
        sheet_name = "Retirement Dashboard 2025"
        if sheet_name in wb.sheetnames:
            del wb[sheet_name]
            print(f"🗑️ Removed existing {sheet_name} sheet")
        
        # Create new sheet
        ws = wb.create_sheet(title=sheet_name, index=0)
        
        # Get current data using the new portfolio config with full API integration
        portfolio_breakdown = self.portfolio_config.get_account_breakdown(use_api=True)
        dividend_data = get_hybrid_dividend_data()
        
        # Calculate monthly dividend income
        monthly_dividend_income = self.calculate_monthly_dividend_income(dividend_data)
        
        # Get all historical dates
        historical_portfolio = self.portfolio_tracker.load_historical_portfolio_values()
        if not historical_portfolio.empty:
            all_dates = sorted(historical_portfolio['Date'].unique())
        else:
            all_dates = [datetime.now().strftime('%Y-%m-%d')]
        
        current_row = 1
        
        # === HEADER SECTION ===
        self.create_header_section(ws, current_row)
        current_row += 4
        
        # === PROFIT/LOSS TRACKING SECTION ===
        current_row = self.create_profit_loss_section(ws, current_row, all_dates)
        current_row += 2
        
        # === DATE HEADERS ===
        current_row = self.create_date_headers(ws, current_row, all_dates)
        current_row += 1
        
        # === PORTFOLIO VALUES SECTION ===
        current_row = self.create_portfolio_section(ws, current_row, all_dates, portfolio_breakdown, historical_portfolio)
        current_row += 2
        
        # === DIVIDEND INCOME SECTION ===
        current_row = self.create_dividend_section(ws, current_row, all_dates, monthly_dividend_income)
        
        # Auto-size columns
        for col in range(1, len(all_dates) + 2):
            ws.column_dimensions[ws.cell(row=1, column=col).column_letter].width = 12
        
        print("📊 Comprehensive retirement tracking sheet created")
        return ws
    
    def create_header_section(self, ws, start_row):
        """Create the header section with retirement info and SS details"""
        
        # Retirement summary
        ws.cell(row=start_row, column=1, value="Retired 5/16/24 total retirement value on 5/16/24")
        ws.cell(row=start_row, column=7, value="$452,975.00")
        ws.cell(row=start_row, column=7).number_format = FORMAT_CURRENCY_USD_SIMPLE
        
        # Mark's SS info
        ws.cell(row=start_row, column=8, value="Marks 2025 SS payments will be")
        ws.cell(row=start_row, column=11, value=self.marks_ss_gross)
        ws.cell(row=start_row, column=11).number_format = FORMAT_CURRENCY_USD_SIMPLE
        
        ws.cell(row=start_row, column=13, value="Medicare payment with prescription drug will be")
        ws.cell(row=start_row, column=17, value=self.marks_medicare)
        ws.cell(row=start_row, column=17).number_format = FORMAT_CURRENCY_USD_SIMPLE
        
        ws.cell(row=start_row, column=18, value="Take home SS")
        ws.cell(row=start_row, column=20, value=self.marks_ss_net)
        ws.cell(row=start_row, column=20).number_format = FORMAT_CURRENCY_USD_SIMPLE
        
        # Carolyn's SS info (next row)
        ws.cell(row=start_row+1, column=8, value="Carolyns 2025 SS payments will be")
        ws.cell(row=start_row+1, column=11, value=self.carolyns_ss_gross)
        ws.cell(row=start_row+1, column=11).number_format = FORMAT_CURRENCY_USD_SIMPLE
        
        ws.cell(row=start_row+1, column=13, value="Medicare payment will be")
        ws.cell(row=start_row+1, column=16, value=self.carolyns_medicare)
        ws.cell(row=start_row+1, column=16).number_format = FORMAT_CURRENCY_USD_SIMPLE
        
        ws.cell(row=start_row+1, column=18, value="Take home SS")
        ws.cell(row=start_row+1, column=20, value=self.carolyns_ss_net)
        ws.cell(row=start_row+1, column=20).number_format = FORMAT_CURRENCY_USD_SIMPLE
        
        # Headers for weekly tracking
        headers = [
            "Total SS take home", "Amount taken out of IRA before tax", 
            "Amount taken out of IRA after tax", "Amount taken out of taxed account",
            "Total after tax withdraws", "Total monthly income", "Yearly income", "Yearly taxes taken out"
        ]
        
        for i, header in enumerate(headers, start=21):
            ws.cell(row=start_row, column=i, value=header)
            ws.cell(row=start_row, column=i).font = SMALL_FONT
        
        # Total SS take home value
        ws.cell(row=start_row+1, column=21, value=self.total_ss_net)
        ws.cell(row=start_row+1, column=21).number_format = FORMAT_CURRENCY_USD_SIMPLE
        
        # Manual entry placeholders
        manual_values = [1200.00, 957.00, 393.00, 1350.00, 4848.70, 58184.40, 2916.00]
        for i, value in enumerate(manual_values, start=22):
            ws.cell(row=start_row+1, column=i, value=value)
            ws.cell(row=start_row+1, column=i).number_format = FORMAT_CURRENCY_USD_SIMPLE
    
    def create_profit_loss_section(self, ws, start_row, all_dates):
        """Create profit/loss tracking rows"""
        
        # Profit/loss since retirement
        ws.cell(row=start_row, column=1, value="Retired 5/16/24 total retirement value on 5/16/24 - $452,975.00, P/L since retirement")
        
        # Profit/loss since beginning of year
        ws.cell(row=start_row+1, column=1, value="Profit/loss since beginning of the year")
        
        # Calculate and populate P/L values for each date
        for col_idx, date in enumerate(all_dates, start=2):
            # Placeholder calculations - you'd replace with actual historical data
            retirement_pl = 29552.00  # This would be calculated from actual portfolio history
            ytd_pl = 8401.00  # This would be calculated from year start
            
            ws.cell(row=start_row, column=col_idx, value=retirement_pl)
            ws.cell(row=start_row, column=col_idx).number_format = FORMAT_CURRENCY_USD_SIMPLE
            
            ws.cell(row=start_row+1, column=col_idx, value=ytd_pl)
            ws.cell(row=start_row+1, column=col_idx).number_format = FORMAT_CURRENCY_USD_SIMPLE
            
            # Color coding
            if retirement_pl > 0:
                ws.cell(row=start_row, column=col_idx).fill = GREEN_FILL
            else:
                ws.cell(row=start_row, column=col_idx).fill = RED_FILL
                
            if ytd_pl > 0:
                ws.cell(row=start_row+1, column=col_idx).fill = GREEN_FILL
            else:
                ws.cell(row=start_row+1, column=col_idx).fill = RED_FILL
        
        return start_row + 2
    
    def create_date_headers(self, ws, start_row, all_dates):
        """Create date header row"""
        for col_idx, date in enumerate(all_dates, start=2):
            try:
                parsed_date = datetime.strptime(date, '%Y-%m-%d')
                display_date = f"{parsed_date.month}/{parsed_date.day}/{parsed_date.year}"
            except:
                display_date = date
            
            ws.cell(row=start_row, column=col_idx, value=display_date)
            ws.cell(row=start_row, column=col_idx).font = HEADER_FONT
            ws.cell(row=start_row, column=col_idx).border = BOX_BORDER
        
        return start_row
    
    def create_portfolio_section(self, ws, start_row, all_dates, portfolio_breakdown, historical_data):
        """Create portfolio values section using live API data"""
        
        accounts = [
            "E*TRADE IRA", "E*TRADE Taxable", "Schwab IRA", 
            "Schwab Individual", "401K", "Total portfolio value", 
            "80% of 401k & IRA taxed full value", "Weekly change"
        ]
        
        for row_idx, account in enumerate(accounts):
            ws.cell(row=start_row + row_idx, column=1, value=account)
            
            if account == "Total portfolio value":
                ws.cell(row=start_row + row_idx, column=1).font = Font(bold=True)
            elif account == "Weekly change":
                ws.cell(row=start_row + row_idx, column=1).font = Font(italic=True)
            
            previous_value = None
            
            for col_idx, date in enumerate(all_dates, start=2):
                if account in ["E*TRADE IRA", "E*TRADE Taxable", "Schwab IRA", "Schwab Individual", "401K"]:
                    # Use live API values from portfolio breakdown
                    value = portfolio_breakdown.get(account, 0)
                elif account == "Total portfolio value":
                    # Get total from portfolio breakdown
                    value = portfolio_breakdown.get("Total portfolio value", 0)
                elif account == "80% of 401k & IRA taxed full value":
                    # Get 80% calculation from portfolio breakdown
                    value = portfolio_breakdown.get("80% of 401k & IRA taxed full value", 0)
                elif account == "Weekly change":
                    # Calculate change from previous week
                    if previous_value is not None:
                        current_total = portfolio_breakdown.get("Total portfolio value", 0)
                        value = current_total - previous_value
                    else:
                        value = 0
                else:
                    value = 0
                
                cell = ws.cell(row=start_row + row_idx, column=col_idx, value=value)
                cell.number_format = FORMAT_CURRENCY_USD_SIMPLE
                cell.border = BOX_BORDER
                
                # Color coding for changes
                if account == "Weekly change" and value != 0:
                    if value > 0:
                        cell.fill = GREEN_FILL
                    else:
                        cell.fill = RED_FILL
                
                if account == "Total portfolio value":
                    previous_value = value
        
        return start_row + len(accounts)
    
    def create_dividend_section(self, ws, start_row, all_dates, monthly_income):
        """Create dividend income tracking section"""
        
        # Headers
        for col_idx in range(2, len(all_dates) + 2):
            ws.cell(row=start_row, column=col_idx, value="Estimated Income Monthly")
            ws.cell(row=start_row, column=col_idx).font = SMALL_FONT
            ws.cell(row=start_row, column=col_idx).border = BOX_BORDER
        
        # Account rows
        dividend_accounts = ["Taxed account", "IRA taxed on withdrawls only", "Schwab Individual", "Schwab IRA"]
        
        for row_idx, account in enumerate(dividend_accounts, start=1):
            ws.cell(row=start_row + row_idx, column=1, value=account)
            
            for col_idx, date in enumerate(all_dates, start=2):
                # Get dividend income for this account and date
                if account == "Taxed account":
                    income = monthly_income.get('ETRADE_Taxable', 0)
                elif account == "IRA taxed on withdrawls only":
                    income = monthly_income.get('ETRADE_IRA', 0)
                else:
                    # Manual entry for Schwab
                    income = 106.25 if "Individual" in account else 646.04
                
                cell = ws.cell(row=start_row + row_idx, column=col_idx, value=income)
                cell.number_format = FORMAT_CURRENCY_USD_SIMPLE
                cell.border = BOX_BORDER
        
        # Monthly dividend total
        total_row = start_row + len(dividend_accounts) + 1
        ws.cell(row=total_row, column=1, value="Monthly dividend")
        ws.cell(row=total_row, column=1).font = Font(bold=True)
        
        for col_idx, date in enumerate(all_dates, start=2):
            total_dividend = (
                monthly_income.get('ETRADE_Taxable', 0) + 
                monthly_income.get('ETRADE_IRA', 0) + 
                106.25 + 646.04  # Schwab placeholders
            )
            
            cell = ws.cell(row=total_row, column=col_idx, value=total_dividend)
            cell.number_format = FORMAT_CURRENCY_USD_SIMPLE
            cell.font = Font(bold=True)
            cell.border = BOX_BORDER
            
            # Color coding based on historical comparison
            if total_dividend > 3000:
                cell.fill = GREEN_FILL
            elif total_dividend < 2800:
                cell.fill = RED_FILL
    
    def calculate_monthly_dividend_income(self, dividend_data):
        """Calculate monthly dividend income from dividend data"""
        monthly_income = {}
        
        for account_type, df in dividend_data.items():
            if not df.empty and 'Est. Income $' in df.columns:
                # Sum all dividend estimates for this account
                monthly_total = df['Est. Income $'].sum()
                monthly_income[account_type] = monthly_total
            else:
                monthly_income[account_type] = 0
        
        return monthly_income

def create_comprehensive_retirement_dashboard(wb_path):
    """Create the complete retirement dashboard"""
    
    if not os.path.exists(wb_path):
        print(f"⚠️ Workbook not found: {wb_path}")
        return
    
    try:
        wb = openpyxl.load_workbook(wb_path)
        tracker = ComprehensiveRetirementTracker()
        tracker.create_comprehensive_tracking_sheet(wb)
        wb.save(wb_path)
        print(f"✅ Comprehensive retirement dashboard created: {wb_path}")
        
    except Exception as e:
        print(f"❌ Error creating retirement dashboard: {e}")

if __name__ == "__main__":
    # Test path
    wb_path = os.path.join(os.path.dirname(__file__), "..", "outputs", "Dividends_2025.xlsx")
    create_comprehensive_retirement_dashboard(wb_path)
