"""
Enhanced Dividend Tracker with Alpha Vantage Integration
Combines your existing historical data with new Alpha Vantage API for current estimates
Last Updated: July 27, 2025
"""

import sys
import pandas as pd
from datetime import datetime
import os
import openpyxl
from openpyxl.styles import Font, Border, Side, PatternFill
from openpyxl.styles.numbers import FORMAT_CURRENCY_USD_SIMPLE

# Add modules to path
sys.path.append(os.path.dirname(__file__))

# Individual imports with error handling
PORTFOLIO_MANAGER_AVAILABLE = False
SCHWAB_API_AVAILABLE = False
ETRADE_API_AVAILABLE = False
DIVIDEND_PROVIDER_AVAILABLE = False

try:
    from portfolio_history_manager import PortfolioHistoryManager
    PORTFOLIO_MANAGER_AVAILABLE = True
    print("✅ Portfolio manager imported")
except ImportError as e:
    print(f"⚠️ Portfolio manager import error: {e}")

try:
    from dividend_data_provider import get_dividend_estimates_for_positions, get_dividend_summary
    from estimated_income_tracker import load_all_historical_data, save_current_estimates
    DIVIDEND_PROVIDER_AVAILABLE = True
    print("✅ Dividend provider imported")
except ImportError as e:
    print(f"⚠️ Dividend provider import error: {e}")

try:
    from schwab_api_integrated import SchwabAPI
    SCHWAB_API_AVAILABLE = True
    print("✅ Schwab API imported")
except ImportError as e:
    print(f"⚠️ Schwab API import error: {e}")

try:
    from etrade_account_api import ETRADEAccountAPI
    ETRADE_API_AVAILABLE = True
    print("✅ E*TRADE API imported")
except ImportError as e:
    print(f"⚠️ E*TRADE API import error: {e}")

# Configuration
OUTPUT_DIR = os.path.join(os.path.dirname(__file__), "..", "outputs")
TARGET_FILE = os.path.join(OUTPUT_DIR, "Dividends_2025.xlsx")
HISTORY_DIR = os.path.join(os.path.dirname(__file__), "..", "data", "history")

# Styling
HEADER_FONT = Font(bold=True, name="Arial", size=12)
NORMAL_FONT = Font(name="Arial", size=12)
BOX_BORDER = Border(
    left=Side(style='thin'), right=Side(style='thin'),
    top=Side(style='thin'), bottom=Side(style='thin')
)
GREEN_FILL = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
RED_FILL = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")

class EnhancedDividendTracker:
    def __init__(self):
        """Initialize the enhanced dividend tracker with portfolio history support"""
        self.current_positions = []
        self.historical_data = pd.DataFrame()
        
        # Initialize portfolio manager if available
        if PORTFOLIO_MANAGER_AVAILABLE:
            self.portfolio_manager = PortfolioHistoryManager()
        else:
            self.portfolio_manager = None
            print("⚠️ Portfolio history manager not available")
        
    def get_all_current_positions(self):
        """Get current positions from both E*TRADE and Schwab APIs"""
        all_positions = []
        
        # Get E*TRADE positions
        try:
            if ETRADE_API_AVAILABLE:
                print("📊 Getting E*TRADE positions...")
                etrade_api = ETRADEAccountAPI()
                # Use the existing dividend estimates method since it has position data
                etrade_accounts = etrade_api.get_dividend_estimates()
                
                for account_type, df in etrade_accounts.items():
                    if not df.empty and 'Symbol' in df.columns:
                        for _, row in df.iterrows():
                            quantity = row.get('Quantity', 0)
                            if quantity > 0:
                                all_positions.append({
                                    'symbol': row.get('Symbol', ''),
                                    'quantity': quantity,
                                    'account_type': f"E*TRADE_{account_type}",
                                    'source': 'E*TRADE'
                                })
            else:
                print("⚠️ E*TRADE API not available")
        except Exception as e:
            print(f"⚠️ Could not get E*TRADE positions: {e}")
        
        # Get Schwab positions
        try:
            if SCHWAB_API_AVAILABLE:
                print("📊 Getting Schwab positions...")
                schwab_api = SchwabAPI()
                portfolio_values = schwab_api.get_portfolio_values()
                
                for account_data in portfolio_values:
                    account_type = f"Schwab_{account_data.get('account_type', 'Unknown')}"
                    positions = account_data.get('positions', [])
                    
                    for pos in positions:
                        if pos.get('quantity', 0) > 0:
                            all_positions.append({
                                'symbol': pos.get('symbol', ''),
                                'quantity': pos.get('quantity', 0),
                                'account_type': account_type,
                                'source': 'Schwab'
                            })
            else:
                print("⚠️ Schwab API not available")
        except Exception as e:
            print(f"⚠️ Could not get Schwab positions: {e}")
        
        print(f"📊 Total positions found: {len(all_positions)}")
        self.current_positions = all_positions
        return all_positions
    
    def get_enhanced_dividend_estimates(self):
        """Get dividend estimates using Alpha Vantage for current positions"""
        if not self.current_positions:
            print("⚠️ No current positions available")
            return pd.DataFrame()
        
        print("💰 Getting dividend estimates with Alpha Vantage...")
        dividend_df = get_dividend_estimates_for_positions(self.current_positions)
        
        if not dividend_df.empty:
            # Add source information
            dividend_df['Data_Source'] = 'Alpha_Vantage_API'
            dividend_df['Generated_Date'] = datetime.now().strftime('%m/%d/%Y')
            
            # Save current estimates to history
            self.save_weekly_estimates(dividend_df)
            
        return dividend_df
    
    def save_weekly_estimates(self, dividend_df):
        """Save current week's estimates to historical data"""
        if dividend_df.empty:
            return
            
        # Create history directory if it doesn't exist
        if not os.path.exists(HISTORY_DIR):
            os.makedirs(HISTORY_DIR)
        
        # Create weekly summary for history
        week_date = datetime.now().strftime('%Y-%m-%d')
        history_file = os.path.join(HISTORY_DIR, f"estimates_{week_date}.csv")
        
        # Prepare data for CSV format
        weekly_data = []
        for _, row in dividend_df.iterrows():
            weekly_data.append({
                'symbol': row['Symbol'],
                'quantity': row['Quantity'],
                'account_type': row['Account_Type'],
                'dividend_per_share': row['Dividend_Per_Share'],
                'annual_dividend': row['Annual_Dividend'],
                'monthly_estimate': row['Monthly_Estimate'],
                'dividend_yield': row['Dividend_Yield'],
                'frequency': row['Frequency'],
                'source': 'Alpha_Vantage',
                'week_captured': week_date
            })
        
        # Save to CSV
        weekly_df = pd.DataFrame(weekly_data)
        weekly_df.to_csv(history_file, index=False)
        print(f"💾 Saved weekly estimates to: {history_file}")
    
    def load_historical_estimates(self):
        """Load all historical dividend estimates"""
        try:
            self.historical_data = load_all_historical_data()
            print(f"📊 Loaded historical data: {len(self.historical_data)} records")
            return self.historical_data
        except Exception as e:
            print(f"⚠️ Could not load historical data: {e}")
            return pd.DataFrame()
    
    def create_comprehensive_dividend_sheet(self, wb):
        """Create enhanced dividend tracking sheet with historical comparison"""
        
        # Remove existing sheet if it exists
        sheet_name = "Enhanced Dividend Tracker"
        if sheet_name in wb.sheetnames:
            del wb[sheet_name]
            print(f"🗑️ Removed existing {sheet_name} sheet")
        
        # Create new sheet
        ws = wb.create_sheet(title=sheet_name, index=0)
        
        # Get current dividend estimates
        current_estimates = self.get_enhanced_dividend_estimates()
        
        if current_estimates.empty:
            ws['A1'] = "No dividend estimates available"
            return ws
        
        # Load historical data
        historical_data = self.load_historical_estimates()
        
        current_row = 1
        
        # === HEADER SECTION ===
        ws.cell(row=current_row, column=1, value="Enhanced Dividend Tracker with Alpha Vantage")
        ws.cell(row=current_row, column=1).font = HEADER_FONT
        current_row += 2
        
        ws.cell(row=current_row, column=1, value=f"Generated: {datetime.now().strftime('%m/%d/%Y %H:%M')}")
        ws.cell(row=current_row, column=1).font = NORMAL_FONT
        current_row += 1
        
        # Get summary stats
        summary = get_dividend_summary(self.current_positions)
        ws.cell(row=current_row, column=1, value=f"Total Monthly Dividend: ${summary['total_monthly']:.2f}")
        ws.cell(row=current_row, column=1).font = NORMAL_FONT
        current_row += 1
        
        ws.cell(row=current_row, column=1, value=f"Total Annual Dividend: ${summary['total_annual']:.2f}")
        ws.cell(row=current_row, column=1).font = NORMAL_FONT
        current_row += 2
        
        # === CURRENT ESTIMATES TABLE ===
        ws.cell(row=current_row, column=1, value="Current Dividend Estimates (Alpha Vantage)")
        ws.cell(row=current_row, column=1).font = HEADER_FONT
        current_row += 1
        
        # Headers
        headers = ['Symbol', 'Quantity', 'Account', 'Div/Share', 'Annual Div', 'Monthly Est', 'Yield', 'Frequency', 'Last Updated']
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=current_row, column=col, value=header)
            cell.font = HEADER_FONT
            cell.border = BOX_BORDER
        current_row += 1
        
        # Data rows
        for _, row in current_estimates.iterrows():
            ws.cell(row=current_row, column=1, value=row['Symbol']).border = BOX_BORDER
            ws.cell(row=current_row, column=2, value=row['Quantity']).border = BOX_BORDER
            ws.cell(row=current_row, column=3, value=row['Account_Type']).border = BOX_BORDER
            
            # Currency formatting for dividend amounts
            div_cell = ws.cell(row=current_row, column=4, value=row['Dividend_Per_Share'])
            div_cell.border = BOX_BORDER
            div_cell.number_format = FORMAT_CURRENCY_USD_SIMPLE
            
            annual_cell = ws.cell(row=current_row, column=5, value=row['Annual_Dividend'])
            annual_cell.border = BOX_BORDER
            annual_cell.number_format = FORMAT_CURRENCY_USD_SIMPLE
            
            monthly_cell = ws.cell(row=current_row, column=6, value=row['Monthly_Estimate'])
            monthly_cell.border = BOX_BORDER
            monthly_cell.number_format = FORMAT_CURRENCY_USD_SIMPLE
            
            ws.cell(row=current_row, column=7, value=row['Dividend_Yield']).border = BOX_BORDER
            ws.cell(row=current_row, column=8, value=row['Frequency']).border = BOX_BORDER
            ws.cell(row=current_row, column=9, value=row['Last_Updated']).border = BOX_BORDER
            
            current_row += 1
        
        current_row += 2
        
        # === HISTORICAL COMPARISON ===
        if not historical_data.empty:
            ws.cell(row=current_row, column=1, value="Historical Weekly Progression")
            ws.cell(row=current_row, column=1).font = HEADER_FONT
            current_row += 2
            
            # Get unique weeks from historical data
            unique_weeks = sorted(historical_data['week_captured'].unique()) if 'week_captured' in historical_data.columns else []
            
            if unique_weeks:
                # Create weekly comparison table
                ws.cell(row=current_row, column=1, value="Week")
                ws.cell(row=current_row, column=1).font = HEADER_FONT
                ws.cell(row=current_row, column=1).border = BOX_BORDER
                
                ws.cell(row=current_row, column=2, value="Total Monthly Est")
                ws.cell(row=current_row, column=2).font = HEADER_FONT
                ws.cell(row=current_row, column=2).border = BOX_BORDER
                
                ws.cell(row=current_row, column=3, value="Change")
                ws.cell(row=current_row, column=3).font = HEADER_FONT
                ws.cell(row=current_row, column=3).border = BOX_BORDER
                
                current_row += 1
                
                previous_total = None
                for week in unique_weeks[-10:]:  # Show last 10 weeks
                    week_data = historical_data[historical_data['week_captured'] == week]
                    week_total = week_data['monthly_estimate'].sum() if 'monthly_estimate' in week_data.columns else 0
                    
                    ws.cell(row=current_row, column=1, value=week).border = BOX_BORDER
                    
                    total_cell = ws.cell(row=current_row, column=2, value=week_total)
                    total_cell.border = BOX_BORDER
                    total_cell.number_format = FORMAT_CURRENCY_USD_SIMPLE
                    
                    # Calculate change
                    if previous_total is not None:
                        change = week_total - previous_total
                        change_cell = ws.cell(row=current_row, column=3, value=change)
                        change_cell.border = BOX_BORDER
                        change_cell.number_format = FORMAT_CURRENCY_USD_SIMPLE
                        
                        # Color coding
                        if change > 0:
                            change_cell.fill = GREEN_FILL
                        elif change < 0:
                            change_cell.fill = RED_FILL
                    
                    previous_total = week_total
                    current_row += 1
        
        # Auto-size columns
        for col in range(1, 10):
            ws.column_dimensions[ws.cell(row=1, column=col).column_letter].width = 15
        
        print(f"✅ Enhanced dividend tracker sheet created")
        return ws
    
    def update_dividend_tracking_workbook(self):
        """Update the main dividend tracking workbook with enhanced data"""
        
        # Create output directory if it doesn't exist
        if not os.path.exists(OUTPUT_DIR):
            os.makedirs(OUTPUT_DIR)
        
        # Load or create workbook
        if os.path.exists(TARGET_FILE):
            wb = openpyxl.load_workbook(TARGET_FILE)
        else:
            wb = openpyxl.Workbook()
            # Remove default sheet
            if 'Sheet' in wb.sheetnames:
                del wb['Sheet']
        
        # Create the enhanced dividend tracking sheet
        self.create_comprehensive_dividend_sheet(wb)
        
        # Create portfolio tracking sheet with 401K manual entry
        if self.portfolio_manager:
            self.portfolio_manager.create_portfolio_tracking_sheet(wb)
        else:
            print("⚠️ Portfolio tracking sheet not created - portfolio manager unavailable")
        
        # Save workbook
        wb.save(TARGET_FILE)
        print(f"💾 Enhanced dividend tracker with portfolio tracking saved to: {TARGET_FILE}")
        
        return TARGET_FILE
    
    def import_portfolio_history(self):
        """Import historical portfolio data from existing Excel structure"""
        if not self.portfolio_manager:
            print("⚠️ Portfolio manager not available")
            return False
        print("📊 Importing portfolio history from existing Excel file...")
        return self.portfolio_manager.import_portfolio_history_from_excel()
    
    def update_401k_value(self, new_value):
        """Update the 401K value for current week"""
        if not self.portfolio_manager:
            print("⚠️ Portfolio manager not available")
            return
        self.portfolio_manager.update_401k_value(new_value)
    
    def get_portfolio_summary(self, date_str=None):
        """Get portfolio summary for a specific week"""
        if not self.portfolio_manager:
            print("⚠️ Portfolio manager not available")
            return {}
        return self.portfolio_manager.get_portfolio_summary_for_week(date_str)

def main():
    """Main function to run the enhanced dividend tracker"""
    print("=== Enhanced Dividend Tracker with Alpha Vantage ===")
    print("1. Portfolio & Dividend Tracking")
    print("2. Import Historical Portfolio Data")
    print("3. Update 401K Value")
    print("4. Show Portfolio Summary")
    
    choice = input("Select option (1-4): ").strip()
    
    tracker = EnhancedDividendTracker()
    
    if choice == "1":
        # Get current positions from all sources
        positions = tracker.get_all_current_positions()
        
        if not positions:
            print("⚠️ No positions found from APIs")
            return
        
        # Update the tracking workbook
        file_path = tracker.update_dividend_tracking_workbook()
        
        # Show summary
        summary = get_dividend_summary(positions)
        print(f"\n=== Dividend Summary ===")
        print(f"📊 Total Positions: {len(positions)}")
        print(f"💰 Monthly Dividend Income: ${summary['total_monthly']:.2f}")
        print(f"💰 Annual Dividend Income: ${summary['total_annual']:.2f}")
        print(f"📈 Average Dividend Yield: {summary['avg_yield']:.1f}%")
        print(f"📁 Workbook saved: {file_path}")
        
    elif choice == "2":
        print("📊 Importing historical portfolio data...")
        success = tracker.import_portfolio_history()
        if success:
            print("✅ Portfolio history imported successfully!")
        else:
            print("⚠️ Portfolio history import failed")
            
    elif choice == "3":
        if not tracker.portfolio_manager:
            print("⚠️ Portfolio manager not available")
            return
            
        current_value = tracker.portfolio_manager.get_current_401k_value()
        if current_value:
            print(f"Current 401K value: ${current_value:,.2f}")
        else:
            print("No current 401K value found")
        
        try:
            new_value_str = input("Enter new 401K value (numbers only): $")
            new_value = float(new_value_str.replace(',', '').replace('$', ''))
            tracker.update_401k_value(new_value)
            print("✅ 401K value updated!")
        except ValueError:
            print("⚠️ Invalid value entered")
            
    elif choice == "4":
        summary = tracker.get_portfolio_summary()
        if summary:
            print(f"\n📊 Current Portfolio Summary:")
            for account, value in summary.items():
                print(f"  {account}: ${value:,.2f}")
        else:
            print("No portfolio data found for current week")
    
    else:
        print("Invalid option")

if __name__ == "__main__":
    main()
