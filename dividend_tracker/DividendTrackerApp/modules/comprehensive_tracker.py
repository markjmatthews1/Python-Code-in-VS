#!/usr/bin/env python3
"""
Comprehensive Tracking Sheet Builder
Builds enhanced sheets with complete historical data and color coding
"""

import os
import pandas as pd
import openpyxl
from openpyxl.styles import Font, Border, Side, PatternFill
from datetime import datetime

TARGET_FILE = os.path.join(os.path.dirname(__file__), "..", "outputs", "Dividends_2025.xlsx")

class ComprehensiveTracker:
    """Build comprehensive tracking sheets with full historical data and color coding"""
    
    def __init__(self):
        self.wb = None
        self.dividend_data = {}
        self.portfolio_data = {}
        
        # Color definitions for progression tracking
        self.colors = {
            'increase': PatternFill(start_color="90EE90", end_color="90EE90", fill_type="solid"),  # Light green
            'decrease': PatternFill(start_color="FFB6C1", end_color="FFB6C1", fill_type="solid"),  # Light red  
            'same': PatternFill(start_color="FFFFE0", end_color="FFFFE0", fill_type="solid"),      # Light yellow
            'header': PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")    # Blue header
        }
        
        # Font definitions
        self.header_font = Font(bold=True, name="Arial", size=12, color="FFFFFF")
        self.normal_font = Font(name="Arial", size=12)
        
        # Border definition
        self.border = Border(
            left=Side(style='thin'), right=Side(style='thin'),
            top=Side(style='thin'), bottom=Side(style='thin')
        )
    
    def import_dividend_historical_data(self):
        """Import dividend historical data from 'All account weekly dividends' sheet"""
        print("📊 Importing dividend historical data from 'All account weekly dividends'...")
        
        try:
            wb = openpyxl.load_workbook(TARGET_FILE)
            if "All account weekly dividends" not in wb.sheetnames:
                print("⚠️ 'All account weekly dividends' sheet not found")
                return False
            
            ws = wb["All account weekly dividends"]
            dates = []
            
            # Get dates from row 1 starting column B
            for col in range(2, ws.max_column + 1):
                date_cell = ws.cell(row=1, column=col).value
                if date_cell:
                    if isinstance(date_cell, datetime):
                        dates.append(date_cell.strftime('%Y-%m-%d'))
                    else:
                        # Calculate date based on position
                        base_date = ws.cell(row=1, column=2).value
                        if isinstance(base_date, datetime):
                            week_offset = col - 2
                            calc_date = base_date + pd.Timedelta(days=7 * week_offset)
                            dates.append(calc_date.strftime('%Y-%m-%d'))
            
            # Account mapping
            account_mapping = {
                'Etrade IRA': 'E*TRADE IRA',
                'Etrade Taxable': 'E*TRADE Taxable', 
                'Schwab IRA': 'Schwab IRA',
                'Schwab Individual': 'Schwab Individual',
                'Monthly totals': 'Monthly Totals'
            }
            
            # Process each account row
            for row in range(2, ws.max_row + 1):
                account_name = ws.cell(row=row, column=1).value
                if account_name and account_name in account_mapping:
                    mapped_name = account_mapping[account_name]
                    self.dividend_data[mapped_name] = {}
                    
                    for col_idx, date in enumerate(dates):
                        col = col_idx + 2
                        value = ws.cell(row=row, column=col).value
                        if value is not None:
                            self.dividend_data[mapped_name][date] = float(value)
            
            print(f"✅ Imported dividend data for {len(self.dividend_data)} accounts across {len(dates)} weeks")
            return True
            
        except Exception as e:
            print(f"❌ Error importing dividend data: {e}")
            return False
    
    def import_portfolio_historical_data(self):
        """Import portfolio historical data from 'All account weekly totals' sheet"""
        print("📊 Importing portfolio historical data from 'All account weekly totals'...")
        
        try:
            wb = openpyxl.load_workbook(TARGET_FILE)
            if "All account weekly totals" not in wb.sheetnames:
                print("⚠️ 'All account weekly totals' sheet not found")
                return False
            
            ws = wb["All account weekly totals"]
            dates = []
            
            # Get dates from row 1 starting column B
            for col in range(2, ws.max_column + 1):
                date_cell = ws.cell(row=1, column=col).value
                if date_cell:
                    if isinstance(date_cell, datetime):
                        dates.append(date_cell.strftime('%Y-%m-%d'))
                    else:
                        # Calculate date based on position
                        base_date = ws.cell(row=1, column=2).value
                        if isinstance(base_date, datetime):
                            week_offset = col - 2
                            calc_date = base_date + pd.Timedelta(days=7 * week_offset)
                            dates.append(calc_date.strftime('%Y-%m-%d'))
            
            # Account mapping
            account_mapping = {
                '401K': '401k Retirement (Manual)',
                'Etrade IRA': 'E*TRADE IRA',
                'Etrade Taxable': 'E*TRADE Taxable',
                'Schwab IRA': 'Schwab IRA', 
                'Schwab Individual': 'Schwab Individual'
            }
            
            # Process each account row
            for row in range(2, ws.max_row + 1):
                account_name = ws.cell(row=row, column=1).value
                if account_name and account_name in account_mapping:
                    mapped_name = account_mapping[account_name]
                    self.portfolio_data[mapped_name] = {}
                    
                    for col_idx, date in enumerate(dates):
                        col = col_idx + 2
                        value = ws.cell(row=row, column=col).value
                        if value is not None:
                            self.portfolio_data[mapped_name][date] = float(value)
            
            print(f"✅ Imported portfolio data for {len(self.portfolio_data)} accounts across {len(dates)} weeks")
            return True
            
        except Exception as e:
            print(f"❌ Error importing portfolio data: {e}")
            return False
    
    def create_enhanced_estimated_income_sheet(self):
        """Create enhanced Estimated Income 2025 sheet with complete historical data and color coding"""
        print("🏗️ Creating enhanced Estimated Income 2025 sheet...")
        
        # Remove existing sheet if it exists
        if "Estimated Income 2025" in self.wb.sheetnames:
            del self.wb["Estimated Income 2025"]
        
        ws = self.wb.create_sheet("Estimated Income 2025", 0)
        
        # Get all dates and sort them
        all_dates = set()
        for account_data in self.dividend_data.values():
            all_dates.update(account_data.keys())
        sorted_dates = sorted(list(all_dates))
        
        # Title
        ws.cell(row=1, column=1, value="Dividend Income Tracking & Progression Analysis")
        ws.cell(row=1, column=1).font = Font(bold=True, size=14, name="Arial")
        ws.merge_cells(f"A1:{chr(66 + len(sorted_dates))}1")
        
        # Headers
        row = 3
        ws.cell(row=row, column=1, value="Account")
        ws.cell(row=row, column=1).font = self.header_font
        ws.cell(row=row, column=1).fill = self.colors['header']
        ws.cell(row=row, column=1).border = self.border
        
        for col_idx, date in enumerate(sorted_dates):
            col = col_idx + 2
            # Format date for display
            display_date = datetime.strptime(date, '%Y-%m-%d').strftime('%m/%d/%Y')
            ws.cell(row=row, column=col, value=display_date)
            ws.cell(row=row, column=col).font = self.header_font
            ws.cell(row=row, column=col).fill = self.colors['header']
            ws.cell(row=row, column=col).border = self.border
        
        # Account data with color coding
        current_row = row + 1
        for account_name in ['E*TRADE IRA', 'E*TRADE Taxable', 'Schwab IRA', 'Schwab Individual', 'Monthly Totals']:
            if account_name in self.dividend_data:
                ws.cell(row=current_row, column=1, value=account_name)
                ws.cell(row=current_row, column=1).font = self.normal_font
                ws.cell(row=current_row, column=1).border = self.border
                
                prev_value = None
                for col_idx, date in enumerate(sorted_dates):
                    col = col_idx + 2
                    current_value = self.dividend_data[account_name].get(date, 0)
                    
                    cell = ws.cell(row=current_row, column=col, value=current_value)
                    cell.font = self.normal_font
                    cell.border = self.border
                    cell.number_format = '"$"#,##0.00_-'
                    
                    # Apply color coding based on previous week
                    if prev_value is not None and current_value != 0:
                        if current_value > prev_value:
                            cell.fill = self.colors['increase']
                        elif current_value < prev_value:
                            cell.fill = self.colors['decrease']
                        elif current_value == prev_value:
                            cell.fill = self.colors['same']
                    
                    prev_value = current_value if current_value != 0 else prev_value
                
                current_row += 1
        
        # Set column widths
        ws.column_dimensions['A'].width = 25
        for col_idx in range(len(sorted_dates)):
            col_letter = chr(66 + col_idx)  # B, C, D, etc.
            ws.column_dimensions[col_letter].width = 12
        
        print("✅ Enhanced Estimated Income 2025 sheet created with color coding")
    
    def create_enhanced_portfolio_values_sheet(self, current_values=None, k401_value=None):
        """Create enhanced Portfolio Values 2025 sheet with complete historical data and color coding"""
        print("🏗️ Creating enhanced Portfolio Values 2025 sheet...")
        
        # Remove existing sheet if it exists
        if "Portfolio Values 2025" in self.wb.sheetnames:
            del self.wb["Portfolio Values 2025"]
        
        ws = self.wb.create_sheet("Portfolio Values 2025", 1)
        
        # Add current week's values if provided
        if current_values and k401_value is not None:
            current_date = datetime.now().strftime('%Y-%m-%d')
            if '401k Retirement (Manual)' not in self.portfolio_data:
                self.portfolio_data['401k Retirement (Manual)'] = {}
            self.portfolio_data['401k Retirement (Manual)'][current_date] = k401_value
            
            for account, value in current_values.items():
                if account not in self.portfolio_data:
                    self.portfolio_data[account] = {}
                self.portfolio_data[account][current_date] = value
        
        # Get all dates and sort them
        all_dates = set()
        for account_data in self.portfolio_data.values():
            all_dates.update(account_data.keys())
        sorted_dates = sorted(list(all_dates))
        
        # Title
        ws.cell(row=1, column=1, value="Portfolio Values & Growth Tracking")
        ws.cell(row=1, column=1).font = Font(bold=True, size=14, name="Arial")
        ws.merge_cells(f"A1:{chr(66 + len(sorted_dates))}1")
        
        # Headers
        row = 3
        ws.cell(row=row, column=1, value="Account")
        ws.cell(row=row, column=1).font = self.header_font
        ws.cell(row=row, column=1).fill = self.colors['header']
        ws.cell(row=row, column=1).border = self.border
        
        for col_idx, date in enumerate(sorted_dates):
            col = col_idx + 2
            # Format date for display
            display_date = datetime.strptime(date, '%Y-%m-%d').strftime('%m/%d/%Y')
            ws.cell(row=row, column=col, value=display_date)
            ws.cell(row=row, column=col).font = self.header_font
            ws.cell(row=row, column=col).fill = self.colors['header']
            ws.cell(row=row, column=col).border = self.border
        
        # Account data with color coding
        current_row = row + 1
        total_row_data = {}
        
        for account_name in ['E*TRADE IRA', 'E*TRADE Taxable', 'Schwab IRA', 'Schwab Individual', '401k Retirement (Manual)']:
            if account_name in self.portfolio_data:
                ws.cell(row=current_row, column=1, value=account_name)
                ws.cell(row=current_row, column=1).font = self.normal_font
                ws.cell(row=current_row, column=1).border = self.border
                
                prev_value = None
                for col_idx, date in enumerate(sorted_dates):
                    col = col_idx + 2
                    current_value = self.portfolio_data[account_name].get(date, 0)
                    
                    if current_value != 0:
                        cell = ws.cell(row=current_row, column=col, value=current_value)
                        cell.font = self.normal_font
                        cell.border = self.border
                        cell.number_format = '"$"#,##0.00_-'
                        
                        # Apply color coding based on previous week
                        if prev_value is not None:
                            if current_value > prev_value:
                                cell.fill = self.colors['increase']
                            elif current_value < prev_value:
                                cell.fill = self.colors['decrease']
                            elif current_value == prev_value:
                                cell.fill = self.colors['same']
                        
                        # Add to total
                        if date not in total_row_data:
                            total_row_data[date] = 0
                        total_row_data[date] += current_value
                        
                        prev_value = current_value
                
                current_row += 1
        
        # Add total row
        ws.cell(row=current_row, column=1, value="TOTAL PORTFOLIO")
        ws.cell(row=current_row, column=1).font = Font(bold=True, name="Arial", size=12)
        ws.cell(row=current_row, column=1).border = self.border
        
        prev_total = None
        for col_idx, date in enumerate(sorted_dates):
            col = col_idx + 2
            total_value = total_row_data.get(date, 0)
            
            if total_value != 0:
                cell = ws.cell(row=current_row, column=col, value=total_value)
                cell.font = Font(bold=True, name="Arial", size=12)
                cell.border = self.border
                cell.number_format = '"$"#,##0.00_-'
                
                # Apply color coding for totals
                if prev_total is not None:
                    if total_value > prev_total:
                        cell.fill = self.colors['increase']
                    elif total_value < prev_total:
                        cell.fill = self.colors['decrease']
                    elif total_value == prev_total:
                        cell.fill = self.colors['same']
                
                prev_total = total_value
        
        # Set column widths
        ws.column_dimensions['A'].width = 25
        for col_idx in range(len(sorted_dates)):
            col_letter = chr(66 + col_idx)  # B, C, D, etc.
            ws.column_dimensions[col_letter].width = 15
        
        print("✅ Enhanced Portfolio Values 2025 sheet created with color coding")
    
    def build_comprehensive_sheets(self, current_values=None, k401_value=None):
        """Build all comprehensive sheets with historical data and color coding"""
        print("🏗️ Building comprehensive tracking sheets...")
        
        # Load workbook
        try:
            self.wb = openpyxl.load_workbook(TARGET_FILE)
        except:
            self.wb = openpyxl.Workbook()
        
        # Import historical data
        self.import_dividend_historical_data()
        self.import_portfolio_historical_data()
        
        # Create enhanced sheets
        self.create_enhanced_estimated_income_sheet()
        self.create_enhanced_portfolio_values_sheet(current_values, k401_value)
        
        # Save workbook
        self.wb.save(TARGET_FILE)
        print(f"✅ Comprehensive tracking sheets saved to {TARGET_FILE}")

# Test function
if __name__ == "__main__":
    tracker = ComprehensiveTracker()
    tracker.build_comprehensive_sheets()
