"""
Module: excel_generator.py
Author: Mark
Created: [Insert Date]
Purpose: Format and export dividend data into Excel workbook with clear styling
Location: C:\Python_Projects\DividendTrackerApp\modules\excel_generator.py
"""

import pandas as pd
import openpyxl
from openpyxl.styles import Font, Border, Side
from openpyxl.utils.dataframe import dataframe_to_rows
import os

# ------------------------- Configuration Section -------------------------
OUTPUT_PATH = os.path.join(os.path.dirname(__file__), "..", "outputs")
OUTPUT_PATH = os.path.abspath(OUTPUT_PATH)

# Auto-create folder if missing
if not os.path.exists(OUTPUT_PATH):
    os.makedirs(OUTPUT_PATH)
DEFAULT_FONT_NAME = "Arial"
DEFAULT_FONT_SIZE = 12
HEADER_FONT = Font(bold=True, name=DEFAULT_FONT_NAME, size=DEFAULT_FONT_SIZE)
NORMAL_FONT = Font(name=DEFAULT_FONT_NAME, size=DEFAULT_FONT_SIZE)
BOX_BORDER = Border(
    left=Side(style='thin'),
    right=Side(style='thin'),
    top=Side(style='thin'),
    bottom=Side(style='thin')
)

# ------------------------- Function: Export DataFrame -------------------------
def export_to_excel(df, filename="dividends_export.xlsx", sheet_name="Dividends"):
    """
    Exports a cleaned DataFrame to an Excel file with formatting:
    - Bold headers with box borders
    - Regular cells: Arial 12, blank if value is 0.0 or NaN
    """

    # Replace NaN and 0.0 with blank before exporting
    df = df.replace({pd.NA: '', 'NaN': '', 'nan': '', 0.0: '', 0: '', None: ''})

    # Define full output file path
    file_path = os.path.join(OUTPUT_PATH, filename)

    # Create workbook and worksheet
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = sheet_name

    # Write DataFrame rows to worksheet with formatting
    for row_index, row in enumerate(dataframe_to_rows(df, index=False, header=True), start=1):
        for col_index, value in enumerate(row, start=1):
            cell = ws.cell(row=row_index, column=col_index, value=value)

            # Apply header style to first row
            if row_index == 1:
                cell.font = HEADER_FONT
                cell.border = BOX_BORDER
            else:
                cell.font = NORMAL_FONT
                cell.border = BOX_BORDER

    # Auto-adjust column widths
    for column_cells in ws.columns:
        length = max(len(str(cell.value)) if cell.value is not None else 0 for cell in column_cells)
        ws.column_dimensions[column_cells[0].column_letter].width = length + 2

    # Save the workbook
    print(f"Saving to 71: {file_path}")
    try:
        wb.save(file_path)
        print(f"✅ Excel export complete: {file_path}")
    except Exception as e:
        print(f"❌ Excel export failed: {e}")

def write_to_workbook(wb, sheet_name, df):
    """
    Appends a formatted DataFrame to an existing workbook object under a new sheet name.
    - Uses Arial 12 font for all cells
    - Bold headers with box borders
    - Auto-adjusts column widths
    """
    # Replace NaN and 0.0 with blank before exporting
    df = df.replace({pd.NA: '', 'NaN': '', 'nan': '', 0.0: '', 0: '', None: ''})

    # Create a new sheet
    ws = wb.create_sheet(title=sheet_name)

    # Write DataFrame to worksheet with formatting
    for row_index, row in enumerate(dataframe_to_rows(df, index=False, header=True), start=1):
        for col_index, value in enumerate(row, start=1):
            cell = ws.cell(row=row_index, column=col_index, value=value)

            # Apply styles
            cell.font = HEADER_FONT if row_index == 1 else NORMAL_FONT
            cell.border = BOX_BORDER

    # Auto-adjust column widths
    for column_cells in ws.columns:
        length = max(len(str(cell.value)) if cell.value is not None else 0 for cell in column_cells)
        ws.column_dimensions[column_cells[0].column_letter].width = length + 2

# ------------------------- Debug Mode -------------------------
if __name__ == "__main__":
    import sys
    sys.path.append(os.path.join(os.path.dirname(__file__)))  # Include current module folder

    from dividend_loader import load_dividend_file, list_available_files

    files = list_available_files()
    if files:
        sample_df = load_dividend_file(files[0])
        export_to_excel(sample_df, filename="dividends_test_export.xlsx")
    else:
        print("⚠️ No files found to export.")