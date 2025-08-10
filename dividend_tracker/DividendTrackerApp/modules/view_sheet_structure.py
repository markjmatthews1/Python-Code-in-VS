"""
Quick test to view the Estimated Income 2025 sheet structure
"""

import os
import pandas as pd
from openpyxl import load_workbook

def show_sheet_structure():
    """Show the structure of the Estimated Income 2025 sheet"""
    
    outputs_dir = os.path.join(os.path.dirname(__file__), "..", "outputs")
    file_path = os.path.join(outputs_dir, "Dividends_2025.xlsx")
    
    if not os.path.exists(file_path):
        print("❌ Dividends_2025.xlsx not found")
        return
    
    try:
        # Load the workbook
        wb = load_workbook(file_path)
        
        print("📊 Available sheets:")
        for sheet_name in wb.sheetnames:
            print(f"  - {sheet_name}")
        
        # Check if Estimated Income 2025 sheet exists
        if "Estimated Income 2025" in wb.sheetnames:
            ws = wb["Estimated Income 2025"]
            
            print(f"\n📋 Estimated Income 2025 sheet structure:")
            print(f"   Dimensions: {ws.max_row} rows x {ws.max_column} columns")
            
            # Show the first few rows to see the structure
            print(f"\n📅 Header row:")
            for col in range(1, min(6, ws.max_column + 1)):
                cell_value = ws.cell(row=1, column=col).value
                print(f"   Column {col}: {cell_value}")
            
            print(f"\n🏦 Account types found:")
            for row in range(2, min(10, ws.max_row + 1)):
                account_name = ws.cell(row=row, column=1).value
                if account_name and account_name not in ["TOTAL", "PORTFOLIO VALUE TRACKING", ""]:
                    value_col2 = ws.cell(row=row, column=2).value
                    print(f"   Row {row}: {account_name} = {value_col2}")
            
            # Look for portfolio section
            portfolio_section_found = False
            for row in range(1, ws.max_row + 1):
                cell_value = ws.cell(row=row, column=1).value
                if cell_value == "PORTFOLIO VALUE TRACKING":
                    print(f"\n💰 Portfolio section found at row {row}")
                    portfolio_section_found = True
                    
                    # Show portfolio accounts
                    for portfolio_row in range(row + 2, min(row + 10, ws.max_row + 1)):
                        portfolio_account = ws.cell(row=portfolio_row, column=1).value
                        if portfolio_account and portfolio_account != "TOTAL PORTFOLIO":
                            portfolio_value = ws.cell(row=portfolio_row, column=2).value
                            print(f"     {portfolio_account}: {portfolio_value}")
                        elif portfolio_account == "TOTAL PORTFOLIO":
                            total_value = ws.cell(row=portfolio_row, column=2).value
                            print(f"     *** {portfolio_account}: {total_value} ***")
                            break
                    break
            
            if not portfolio_section_found:
                print(f"\n⚠️ Portfolio section not found in sheet")
        
        else:
            print(f"\n❌ 'Estimated Income 2025' sheet not found")
        
        wb.close()
        
    except Exception as e:
        print(f"❌ Error reading Excel file: {e}")

if __name__ == "__main__":
    show_sheet_structure()
