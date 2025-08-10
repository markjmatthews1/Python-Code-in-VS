"""
Portfolio History Integration Module
Integrates historical portfolio data into the Estimated Income 2025 sheet
Replaces static values with actual weekly historical data
Last Updated: July 27, 2025
"""

import os
import pandas as pd
import openpyxl
from datetime import datetime, timedelta
from openpyxl.styles import Font, Border, Side, PatternFill
from openpyxl.styles.numbers import FORMAT_CURRENCY_USD_SIMPLE

# Configuration
OUTPUT_DIR = os.path.join(os.path.dirname(__file__), "..", "outputs")
TARGET_FILE = os.path.join(OUTPUT_DIR, "Dividends_2025.xlsx")

class PortfolioHistoryIntegrator:
    def __init__(self):
        """Initialize the portfolio history integrator"""
        
        # Your actual historical data with multiple date format keys
        base_data = {
            (2024, 12, 29): {'401K': 165551.00, 'E*TRADE IRA': 254846.00, 'E*TRADE TAXABLE': 62130.00, 'TOTAL': 482527.00},
            (2025, 1, 5): {'401K': 165819.00, 'E*TRADE IRA': 262064.00, 'E*TRADE TAXABLE': 63045.00, 'TOTAL': 490928.00},
            (2025, 1, 12): {'401K': 162065.00, 'E*TRADE IRA': 256772.00, 'E*TRADE TAXABLE': 63204.00, 'TOTAL': 482041.00},
            (2025, 1, 19): {'401K': 166210.00, 'E*TRADE IRA': 267398.00, 'E*TRADE TAXABLE': 63872.00, 'TOTAL': 497480.00},
            (2025, 1, 26): {'401K': 169990.00, 'E*TRADE IRA': 270572.00, 'E*TRADE TAXABLE': 63600.00, 'TOTAL': 504162.00},
            (2025, 2, 2): {'401K': 166949.00, 'E*TRADE IRA': 269937.00, 'E*TRADE TAXABLE': 63995.00, 'TOTAL': 500881.00},
            (2025, 2, 9): {'401K': 166709.00, 'E*TRADE IRA': 277913.00, 'E*TRADE TAXABLE': 64369.00, 'TOTAL': 508991.00},
            (2025, 2, 16): {'401K': 169029.00, 'E*TRADE IRA': 279767.00, 'E*TRADE TAXABLE': 65010.00, 'TOTAL': 513806.00},
            (2025, 2, 23): {'401K': 164087.00, 'E*TRADE IRA': 272834.00, 'E*TRADE TAXABLE': 63649.00, 'TOTAL': 500570.00},
            (2025, 3, 2): {'401K': 160462.00, 'E*TRADE IRA': 263612.00, 'E*TRADE TAXABLE': 62948.00, 'TOTAL': 487022.00},
            (2025, 3, 9): {'401K': 153493.00, 'E*TRADE IRA': 255640.00, 'E*TRADE TAXABLE': 63742.00, 'TOTAL': 472875.00},
            (2025, 3, 16): {'401K': 150494.00, 'E*TRADE IRA': 252656.00, 'E*TRADE TAXABLE': 61477.00, 'TOTAL': 464627.00},
            (2025, 3, 23): {'401K': 151362.00, 'E*TRADE IRA': 252762.00, 'E*TRADE TAXABLE': 61627.00, 'TOTAL': 465751.00},
            (2025, 3, 30): {'401K': 147403.00, 'E*TRADE IRA': 243756.00, 'E*TRADE TAXABLE': 60483.00, 'TOTAL': 451642.00},
            (2025, 4, 6): {'401K': 132519.00, 'E*TRADE IRA': 219603.00, 'E*TRADE TAXABLE': 56557.00, 'TOTAL': 408679.00},
            (2025, 4, 13): {'401K': 140458.00, 'E*TRADE IRA': 224151.00, 'E*TRADE TAXABLE': 57385.00, 'TOTAL': 421994.00},
            (2025, 4, 20): {'401K': 138329.00, 'E*TRADE IRA': 223601.00, 'E*TRADE TAXABLE': 56721.00, 'TOTAL': 418651.00},
            (2025, 4, 27): {'401K': 146480.00, 'E*TRADE IRA': 236658.00, 'E*TRADE TAXABLE': 58556.00, 'TOTAL': 441694.00},
            (2025, 5, 4): {'401K': 151256.00, 'E*TRADE IRA': 243672.00, 'E*TRADE TAXABLE': 59427.00, 'TOTAL': 454355.00},
            (2025, 5, 11): {'401K': 150857.00, 'E*TRADE IRA': 244978.00, 'E*TRADE TAXABLE': 60844.00, 'TOTAL': 456679.00},
            (2025, 5, 18): {'401K': 160921.00, 'E*TRADE IRA': 258118.00, 'E*TRADE TAXABLE': 60318.00, 'TOTAL': 479357.00},
            (2025, 5, 25): {'401K': 156962.00, 'E*TRADE IRA': 249420.00, 'E*TRADE TAXABLE': 58830.00, 'TOTAL': 465212.00},
            (2025, 6, 1): {'401K': 159656.00, 'E*TRADE IRA': 253461.00, 'E*TRADE TAXABLE': 59465.00, 'TOTAL': 472582.00},
            (2025, 6, 8): {'401K': 163658.00, 'E*TRADE IRA': 264617.00, 'E*TRADE TAXABLE': 60286.00, 'TOTAL': 488561.00},
            (2025, 6, 15): {'401K': 162113.00, 'E*TRADE IRA': 262824.00, 'E*TRADE TAXABLE': 61572.00, 'Schwab IRA': 50145.00, 'Schwab Individual': 1500.00, 'TOTAL': 486509.00},
            (2025, 6, 18): {'401K': 162957.00, 'E*TRADE IRA': 264981.00, 'E*TRADE TAXABLE': 60281.00, 'Schwab IRA': 50145.00, 'Schwab Individual': 1526.00, 'TOTAL': 488219.00},
            (2025, 6, 22): {'401K': 112508.00, 'E*TRADE IRA': 263152.00, 'E*TRADE TAXABLE': 60015.00, 'Schwab IRA': 50145.00, 'Schwab Individual': 1613.00, 'TOTAL': 487320.00},
            (2025, 6, 29): {'401K': 117082.00, 'E*TRADE IRA': 272767.00, 'E*TRADE TAXABLE': 60668.00, 'Schwab IRA': 50124.00, 'Schwab Individual': 1605.00, 'TOTAL': 502188.00},
            (2025, 7, 6): {'401K': 119228.00, 'E*TRADE IRA': 278134.00, 'E*TRADE TAXABLE': 61658.00, 'Schwab IRA': 50077.00, 'Schwab Individual': 1566.00, 'TOTAL': 510778.00},
            (2025, 7, 13): {'401K': 119082.00, 'E*TRADE IRA': 278002.00, 'E*TRADE TAXABLE': 63658.00, 'Schwab IRA': 49951.00, 'Schwab Individual': 1605.60, 'TOTAL': 512471.00},
            (2025, 7, 20): {'401K': 120374.00, 'E*TRADE IRA': 277430.00, 'E*TRADE TAXABLE': 61708.00, 'Schwab IRA': 49951.00, 'Schwab Individual': 1605.60, 'TOTAL': 511155.00},
            (2025, 7, 27): {'401K': 125000.00, 'E*TRADE IRA': 278418.00, 'E*TRADE TAXABLE': 62110.00, 'Schwab IRA': 49951.53, 'Schwab Individual': 1605.60, 'TOTAL': 517085.13}
        }
        
        # Create historical_data with date lookup flexibility
        self.historical_data = {}
        for date_tuple, values in base_data.items():
            year, month, day = date_tuple
            date_obj = datetime(year, month, day)
            
            # Create multiple format keys for flexibility
            self.historical_data[date_obj] = values
        
        # Account mapping for sheet integration
        self.account_mapping = {
            '401K': '401k Retirement',
            'E*TRADE IRA': 'E*TRADE IRA', 
            'E*TRADE TAXABLE': 'E*TRADE Taxable',
            'Schwab IRA': 'Schwab IRA',
            'Schwab Individual': 'Schwab Individual',
            'TOTAL': 'TOTAL PORTFOLIO'
        }
    
    def update_estimated_income_sheet(self):
        """Update the Estimated Income 2025 sheet with historical portfolio data"""
        
        if not os.path.exists(TARGET_FILE):
            print(f"❌ Target file not found: {TARGET_FILE}")
            return False
            
        try:
            print("📊 Opening Dividends_2025.xlsx for portfolio data integration...")
            wb = openpyxl.load_workbook(TARGET_FILE)
            
            # Find the Estimated Income sheet
            sheet_name = "Estimated Income 2025"
            if sheet_name not in wb.sheetnames:
                print(f"❌ Sheet '{sheet_name}' not found")
                return False
                
            ws = wb[sheet_name]
            print(f"✅ Found sheet: {sheet_name}")
            
            # Find date row and account column structure
            date_row = 1
            account_col = 1
            
            # Parse existing dates and find the data range
            existing_dates = []
            start_col = 2  # Assuming dates start in column B
            
            for col in range(start_col, ws.max_column + 1):
                cell_value = ws.cell(row=date_row, column=col).value
                if cell_value:
                    if isinstance(cell_value, datetime):
                        date_str = cell_value.strftime('%m/%d/%Y')
                    else:
                        # Try to parse as string
                        try:
                            if '/' in str(cell_value):
                                date_obj = datetime.strptime(str(cell_value), '%m/%d/%Y')
                                date_str = date_obj.strftime('%m/%d/%Y')
                            else:
                                continue
                        except:
                            continue
                    existing_dates.append((col, date_str))
                    
            print(f"📅 Found {len(existing_dates)} date columns")
            
            # Update historical data for each account
            accounts_updated = 0
            for row in range(2, ws.max_row + 1):
                account_name = ws.cell(row=row, column=account_col).value
                if not account_name:
                    continue
                    
                account_str = str(account_name).strip()
                
                # Find matching account in our historical data
                historical_account = None
                for hist_acc, sheet_acc in self.account_mapping.items():
                    if sheet_acc.lower() in account_str.lower() or account_str.lower() in sheet_acc.lower():
                        historical_account = hist_acc
                        break
                
                if not historical_account:
                    continue
                    
                print(f"📊 Updating {account_str} ({historical_account})...")
                values_updated = 0
                
                # Update each date column with historical data
                for col, date_str in existing_dates:
                    # Parse the Excel date and find matching historical data
                    try:
                        # Try different date parsing approaches
                        date_obj = None
                        if isinstance(date_str, datetime):
                            date_obj = date_str
                        else:
                            # Try parsing as string with various formats
                            for fmt in ['%m/%d/%Y', '%m/%d/%y', '%Y-%m-%d']:
                                try:
                                    date_obj = datetime.strptime(str(date_str), fmt)
                                    break
                                except:
                                    continue
                        
                        if date_obj is None:
                            continue
                            
                        # Look for matching historical data by date object
                        historical_value = None
                        for hist_date, hist_data in self.historical_data.items():
                            if (hist_date.year == date_obj.year and 
                                hist_date.month == date_obj.month and 
                                hist_date.day == date_obj.day):
                                if historical_account in hist_data:
                                    historical_value = hist_data[historical_account]
                                    break
                        
                        if historical_value is not None:
                            # Update the cell
                            cell = ws.cell(row=row, column=col)
                            cell.value = historical_value
                            cell.number_format = FORMAT_CURRENCY_USD_SIMPLE
                            values_updated += 1
                            
                    except Exception as e:
                        continue
                
                if values_updated > 0:
                    print(f"  ✅ Updated {values_updated} values for {account_str}")
                    accounts_updated += 1
                else:
                    print(f"  ⚠️ No values updated for {account_str}")
            
            # Save the updated workbook
            wb.save(TARGET_FILE)
            print(f"💾 Portfolio data integration complete!")
            print(f"📊 Updated {accounts_updated} accounts with historical data")
            print(f"📁 File saved: {TARGET_FILE}")
            
            return True
            
        except Exception as e:
            print(f"❌ Error during portfolio integration: {e}")
            import traceback
            traceback.print_exc()
            return False
    
    def verify_current_values(self):
        """Verify that current values are correct"""
        
        current_date_obj = datetime(2025, 7, 27)
        current_values = self.historical_data.get(current_date_obj, {})
        
        if current_values:
            print(f"\n📊 Expected values for 7/27/2025:")
            for account, value in current_values.items():
                print(f"  {account}: ${value:,.2f}")
        
        return current_values

def main():
    """Main function to integrate portfolio history"""
    print("=== Portfolio History Integration ===")
    
    integrator = PortfolioHistoryIntegrator()
    
    print("Expected current values:")
    current_values = integrator.verify_current_values()
    
    choice = input("\nProceed with integration? (y/n): ").lower()
    
    if choice.startswith('y'):
        success = integrator.update_estimated_income_sheet()
        if success:
            print("\n✅ Integration completed successfully!")
            print("The Estimated Income 2025 sheet now contains your actual historical portfolio data.")
            print(f"Schwab Individual should now show ${current_values.get('Schwab Individual', 0):,.2f}")
        else:
            print("\n❌ Integration failed. Check errors above.")
    else:
        print("Integration cancelled.")

if __name__ == "__main__":
    main()
