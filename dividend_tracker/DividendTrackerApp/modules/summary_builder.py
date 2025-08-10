r"""
Module: summary_builder.py
Author: Mark
Created: [Insert Date]
Purpose: Add unified Totals sheet to dividend workbook
Location: C:\Python_Projects\DividendTrackerApp\modules\summary_builder.py
"""

import os
import pandas as pd
import openpyxl
from openpyxl.styles import Font, Border, Side
from openpyxl.utils.dataframe import dataframe_to_rows

# ------------------------- Configuration -------------------------
OUTPUT_DIR = os.path.join(os.path.dirname(__file__), "..", "outputs")
TARGET_FILE = os.path.join(OUTPUT_DIR, "Dividends_2025.xlsx")

HEADER_FONT = Font(bold=True, name="Arial", size=12)
NORMAL_FONT = Font(name="Arial", size=12)
BOX_BORDER = Border(
    left=Side(style='thin'),
    right=Side(style='thin'),
    top=Side(style='thin'),
    bottom=Side(style='thin')
)

MONTH_ORDER = [
    "January", "February", "March", "April", "May", "June",
    "July", "August", "September", "October", "November", "December"
]

# ------------------------- Helper: Write DataFrame to Sheet -------------------------
def write_block(ws, df, start_row):
    for row_index, row in enumerate(dataframe_to_rows(df, index=False, header=True), start=start_row):
        for col_index, value in enumerate(row, start=1):
            cell = ws.cell(row=row_index, column=col_index, value=value)
            cell.font = HEADER_FONT if row_index == start_row else NORMAL_FONT
            cell.border = BOX_BORDER
    return row_index + 2  # Leave a blank row after block

# ------------------------- Main Function -------------------------
def build_totals_sheet():
    if not os.path.exists(TARGET_FILE):
        print(f"⚠️ Workbook not found: {TARGET_FILE}")
        return

    try:
        wb = openpyxl.load_workbook(TARGET_FILE)
        if "Totals 2025" in wb.sheetnames:
            del wb["Totals 2025"]
        ws = wb.create_sheet(title="Totals 2025")

        all_data = []

        for sheet_name in wb.sheetnames:
            if sheet_name.lower() == "totals 2025":
                continue

            df = pd.DataFrame(wb[sheet_name].values)
            df.columns = df.iloc[0]
            df = df.drop(index=0)
            df.columns = [str(col).strip().lower() for col in df.columns]

            if "amount" not in df.columns or "symbol" not in df.columns:
                continue

            df["amount"] = pd.to_numeric(df["amount"], errors="coerce").fillna(0)
            df["symbol"] = df["symbol"].astype(str).str.strip().str.upper()
            df["account type"] = df.get("account type", "Unknown")

            if "transactiondate" in df.columns:
                df["transactiondate"] = pd.to_datetime(df["transactiondate"], errors="coerce")
                df["month"] = df["transactiondate"].dt.strftime('%B')
                df["week"] = df["transactiondate"].dt.to_period("W").apply(lambda r: r.start_time.strftime('%Y-%m-%d'))
            else:
                df["month"] = ""
                df["week"] = ""

            df["month"] = pd.Categorical(df["month"], categories=MONTH_ORDER, ordered=True)
            all_data.append(df)

        if not all_data:
            print("⚠️ No valid data found.")
            return

        combined_df = pd.concat(all_data)

        # 1. Weekly Totals
        weekly = combined_df.groupby(["account type", "month", "week"], as_index=False)["amount"].sum()
        weekly = weekly.sort_values(by=["month", "week", "account type"])
        weekly["change"] = weekly.groupby(["account type", "month"])["amount"].diff().fillna(0)
        start_row = write_block(ws, weekly, 1)

        # 2. Monthly Totals by Account
        monthly_by_account = combined_df.groupby(["account type", "month"], as_index=False)["amount"].sum()
        monthly_by_account = monthly_by_account.sort_values(by=["month", "account type"])
        start_row = write_block(ws, monthly_by_account, start_row)

        # 3. Combined Monthly Totals
        monthly_combined = combined_df.groupby("month", as_index=False)["amount"].sum()
        monthly_combined = monthly_combined.sort_values(by="month")
        monthly_combined["change"] = monthly_combined["amount"].diff().fillna(0)
        start_row = write_block(ws, monthly_combined, start_row)

        # 4. Ticker Totals by Month and Account
        ticker_monthly = combined_df.groupby(["account type", "month", "symbol"], as_index=False)["amount"].sum()
        ticker_monthly = ticker_monthly.sort_values(by=["month", "account type", "symbol"])
        write_block(ws, ticker_monthly, start_row)

        wb.save(TARGET_FILE)
        print("✅ Unified Totals 2025 sheet created successfully.")

    except Exception as e:
        print(f"❌ Error building summary: {e}")

# ------------------------- Debug Mode -------------------------
if __name__ == "__main__":
    build_totals_sheet()