r"""
Module: estimated_income_tracker.py
Author: Mark
Created: July 25, 2025
Purpose: Track weekly changes in estimated monthly dividend income from E*TRADE
Location: C:\Python_Projects\DividendTrackerApp\modules\estimated_income_tracker.py
"""

import os
import pandas as pd
import openpyxl
from openpyxl.styles import Font, Border, Side, PatternFill
from openpyxl.styles.numbers import FORMAT_CURRENCY_USD_SIMPLE
from datetime import datetime
import calendar

# ------------------------- Configuration -------------------------
OUTPUT_DIR = os.path.join(os.path.dirname(__file__), "..", "outputs")
TARGET_FILE = os.path.join(OUTPUT_DIR, "Dividends_2025.xlsx")
HISTORY_DIR = os.path.join(os.path.dirname(__file__), "..", "data", "history")

HEADER_FONT = Font(bold=True, name="Arial", size=12)
NORMAL_FONT = Font(name="Arial", size=12)
BOX_BORDER = Border(
    left=Side(style='thin'), right=Side(style='thin'),
    top=Side(style='thin'), bottom=Side(style='thin')
)

# Month order for proper sorting
MONTH_ORDER = [
    "January", "February", "March", "April", "May", "June",
    "July", "August", "September", "October", "November", "December"
]

# Color fills for income changes
GREEN_FILL = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")  # Light green
RED_FILL = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")    # Light red

def save_current_estimates(monthly_summaries):
    """Save current week's estimates for future comparison"""
    if not os.path.exists(HISTORY_DIR):
        os.makedirs(HISTORY_DIR)
    
    current_week = datetime.now().strftime('%Y-%m-%d')
    history_file = os.path.join(HISTORY_DIR, f"estimates_{current_week}.csv")
    
    # Combine all account data for storage
    all_data = []
    for account_type, df in monthly_summaries.items():
        df_copy = df.copy()
        df_copy['account_type'] = account_type
        all_data.append(df_copy)
    
    if all_data:
        combined_history = pd.concat(all_data)
        combined_history.to_csv(history_file, index=False)
        print(f"📄 Saved current estimates to: {history_file}")

def load_previous_week_estimates():
    """Load previous week's estimates for comparison"""
    if not os.path.exists(HISTORY_DIR):
        return None
    
    # Find the most recent history file (excluding today)
    today = datetime.now().strftime('%Y-%m-%d')
    history_files = [f for f in os.listdir(HISTORY_DIR) if f.startswith('estimates_') and f.endswith('.csv')]
    history_files = [f for f in history_files if not f.endswith(f"{today}.csv")]
    
    if not history_files:
        return None
    
    # Get the most recent file
    history_files.sort(reverse=True)
    recent_file = os.path.join(HISTORY_DIR, history_files[0])
    
    try:
        previous_data = pd.read_csv(recent_file)
        print(f"📄 Loaded previous estimates from: {recent_file}")
        return previous_data
    except Exception as e:
        print(f"⚠️ Could not load previous estimates: {e}")
        return None

def load_historical_portfolio_data():
    """Load historical portfolio data from CSV files"""
    portfolio_history_dir = os.path.join(os.path.dirname(__file__), "..", "data", "portfolio_history")
    
    if not os.path.exists(portfolio_history_dir):
        print("⚠️ No portfolio history directory found")
        return {}
    
    historical_portfolio = {}
    
    # Get all portfolio history files
    for filename in os.listdir(portfolio_history_dir):
        if filename.startswith('portfolio_') and filename.endswith('.csv'):
            date_str = filename.replace('portfolio_', '').replace('.csv', '')
            
            try:
                file_path = os.path.join(portfolio_history_dir, filename)
                df = pd.read_csv(file_path)
                
                # Convert to dictionary format
                week_data = {}
                for _, row in df.iterrows():
                    account = row['Account']
                    value = float(row['Value']) if pd.notna(row['Value']) else 0
                    week_data[account] = value
                
                historical_portfolio[date_str] = week_data
                
            except Exception as e:
                print(f"⚠️ Error loading {filename}: {e}")
                continue
    
    print(f"📊 Loaded historical portfolio data for {len(historical_portfolio)} weeks")
    return historical_portfolio


def load_estimate_files(use_api=False, use_hybrid=False):
    """Load both E*TRADE and Schwab estimate files and return processed data
    If use_api=True, fetches data directly from both E*TRADE and Schwab APIs instead of Excel files
    If use_hybrid=True, uses API for validation + Excel for dividend estimates
    """
    
    if use_hybrid:
        print("🔄 Loading estimate data using hybrid approach (API + Excel)...")
        try:
            from hybrid_dividend_tracker import get_hybrid_dividend_data
            account_data = get_hybrid_dividend_data()
            return account_data
        except Exception as e:
            print(f"❌ Error loading hybrid data: {e}")
            print("📄 Falling back to Excel files...")
    
    if use_api:
        print("🔄 Loading estimate data from E*TRADE and Schwab APIs...")
        account_data = {}
        
        # Get E*TRADE dividend estimates
        try:
            from modules.etrade_account_api import ETRADEAccountAPI
            etrade_api = ETRADEAccountAPI()
            etrade_data = etrade_api.get_dividend_estimates()
            account_data.update(etrade_data)
            print("✅ Successfully loaded E*TRADE dividend estimates")
        except Exception as e:
            print(f"❌ Error loading from E*TRADE API: {e}")
        
        # Get Schwab dividend estimates
        try:
            from modules.schwab_api_integrated import SchwabAPI
            schwab_api = SchwabAPI()
            schwab_estimates = schwab_api.get_dividend_estimates()
            
            # Convert Schwab data to match our format
            for account_key, positions in schwab_estimates.items():
                if positions:
                    # Convert to DataFrame format matching E*TRADE structure
                    df_data = []
                    for pos in positions:
                        df_data.append({
                            'Symbol': pos['Symbol'],
                            'Quantity #': pos['Quantity'],
                            'Payable Date': None,  # Schwab API doesn't provide dividend dates
                            'Est. Income $': 0,    # Would need additional dividend research
                            'Status': 'SCHWAB_API_POSITION',
                            'Frequency': 'Unknown',
                            'Income Type': 'Dividend',
                            'Rate': 0,
                            'account_type': account_key,
                            'Position_Type': 'EQ'
                        })
                    
                    if df_data:
                        import pandas as pd
                        schwab_df = pd.DataFrame(df_data)
                        account_data[account_key] = schwab_df
                        print(f"✅ Successfully loaded {len(df_data)} Schwab positions from {account_key}")
        except Exception as e:
            print(f"❌ Error loading from Schwab API: {e}")
        
        if account_data:
            return account_data
        else:
            print("📄 No API data available, falling back to Excel files...")
    
    # Original Excel file loading logic
    data_dir = os.path.join(os.path.dirname(__file__), "..", "data")
    estimate_files = {
        'ETRADE_IRA': 'etrade_ira_estimates_2025.xlsx',
        'ETRADE_Taxable': 'etrade_taxable_estimates_2025.xlsx'
    }
    
    account_data = {}
    
    for account_type, filename in estimate_files.items():
        filepath = os.path.join(data_dir, filename)
        if not os.path.exists(filepath):
            print(f"⚠️ Estimate file not found: {filename}")
            continue
            
        try:
            # Read the entire Excel file first to find the data section
            all_data = pd.read_excel(filepath, header=None)
            
            # Find the row with "ESTIMATED INCOME DETAILS"
            details_row = None
            for idx, row in all_data.iterrows():
                if row.astype(str).str.contains('ESTIMATED INCOME DETAILS', na=False).any():
                    details_row = idx
                    break
            
            if details_row is None:
                print(f"⚠️ Could not find 'ESTIMATED INCOME DETAILS' section in {filename}")
                continue
            
            # Read from the row after "ESTIMATED INCOME DETAILS" (which contains headers)
            df = pd.read_excel(filepath, skiprows=details_row + 1)
            
            # Print original columns for debugging
            print(f"📋 Original columns in {filename}: {list(df.columns)}")
            
            # Clean column names (keep original structure but clean spaces)
            df.columns = [str(col).strip() for col in df.columns]
            
            # Filter out empty rows and TOTAL row
            df = df[df['Symbol'].notna() & (df['Symbol'] != 'TOTAL')]
            
            # Remove any rows that might be footers or empty
            df = df[df['Symbol'].astype(str).str.strip() != '']
            
            # Add account identifier
            df['account_type'] = account_type
            
            # Convert date and income columns using exact column names
            df['Payable Date'] = pd.to_datetime(df['Payable Date'], errors='coerce')
            df['Est. Income $'] = pd.to_numeric(df['Est. Income $'], errors='coerce')
            
            print(f"✅ Loaded {len(df)} estimates from {account_type}")
            account_data[account_type] = df
            
        except Exception as e:
            print(f"❌ Error loading {filename}: {e}")
            continue
    
    return account_data

def calculate_monthly_totals(account_data):
    """Calculate estimated monthly income totals by account"""
    monthly_summaries = {}
    
    for account_type, df in account_data.items():
        if df.empty:
            continue
        
        # Check if we have 'Payable Date' column and if it contains valid dates
        if 'Payable Date' in df.columns and not df['Payable Date'].isna().all():
            # Convert date column for Excel data
            df['Payable Date'] = pd.to_datetime(df['Payable Date'], errors='coerce')
            # Extract month from payable date using correct column name
            df['month'] = df['Payable Date'].dt.strftime('%B')
            df['year'] = df['Payable Date'].dt.year
            
            # Group by month and sum estimated income using correct column name
            monthly_total = df.groupby(['year', 'month'])['Est. Income $'].sum().reset_index()
        else:
            # For API data without specific payable dates, create a summary row
            total_income = df['Est. Income $'].sum() if 'Est. Income $' in df.columns else 0
            monthly_total = pd.DataFrame({
                'year': [datetime.now().year],
                'month': ['Total'],
                'Est. Income $': [total_income]
            })
        
        monthly_total['account_type'] = account_type
        
        # Sort by proper month order (skip if using 'Total')
        if 'Total' not in monthly_total['month'].values:
            monthly_total['month'] = pd.Categorical(monthly_total['month'], categories=MONTH_ORDER, ordered=True)
            monthly_total = monthly_total.sort_values('month')
        
        # Add current week capture date
        monthly_total['week_captured'] = datetime.now().strftime('%Y-%m-%d')
        
        monthly_summaries[account_type] = monthly_total
    
    # Save current estimates for future weekly comparison
    save_current_estimates(monthly_summaries)
        
    return monthly_summaries

def import_historical_portfolio_data_from_excel():
    """Import historical portfolio tracking data from 'All account weekly totals' sheet and create CSV files"""
    
    print("🔄 Looking for 'All account weekly totals' sheet for portfolio data...")
    
    try:
        wb = openpyxl.load_workbook(TARGET_FILE)
        
        if "All account weekly totals" not in wb.sheetnames:
            print("⚠️ No 'All account weekly totals' sheet found for portfolio data.")
            return False
        
        ws = wb["All account weekly totals"]
        
        print("🔄 Processing historical portfolio data from 'All account weekly totals'...")
        
        # Portfolio history directory
        portfolio_history_dir = os.path.join(os.path.dirname(__file__), "..", "data", "portfolio_history")
        if not os.path.exists(portfolio_history_dir):
            os.makedirs(portfolio_history_dir)
        
        # Get date headers from first row (starting from column B)
        date_headers = []
        base_date = None
        
        for col in range(2, ws.max_column + 1):  # Start from column B (index 2)
            header_cell = ws.cell(row=1, column=col)
            header_value = header_cell.value
            
            if header_value:
                if isinstance(header_value, datetime):
                    # This is the first evaluated date
                    base_date = header_value
                    date_headers.append(header_value.strftime('%Y-%m-%d'))
                    print(f"📅 Column {col}: Base date {header_value.strftime('%Y-%m-%d')}")
                elif str(header_value).startswith('=') and base_date:
                    # This is a formula - calculate the date manually (=B1+7, =C1+7, etc.)
                    week_offset = col - 2  # How many weeks from the base date
                    calculated_date = base_date + pd.Timedelta(days=7 * week_offset)
                    date_headers.append(calculated_date.strftime('%Y-%m-%d'))
                    print(f"📅 Column {col}: Calculated date {calculated_date.strftime('%Y-%m-%d')}")
                elif isinstance(header_value, (int, float)) and base_date:
                    # Sometimes Excel evaluates the formula to a number, treat as days since base
                    try:
                        days_offset = int(header_value - base_date.toordinal()) if hasattr(base_date, 'toordinal') else (col - 2) * 7
                        calculated_date = base_date + pd.Timedelta(days=days_offset)
                        date_headers.append(calculated_date.strftime('%Y-%m-%d'))
                        print(f"📅 Column {col}: Date from number {calculated_date.strftime('%Y-%m-%d')}")
                    except:
                        # Fallback to weekly calculation
                        week_offset = col - 2
                        calculated_date = base_date + pd.Timedelta(days=7 * week_offset)
                        date_headers.append(calculated_date.strftime('%Y-%m-%d'))
                        print(f"📅 Column {col}: Fallback calculated date {calculated_date.strftime('%Y-%m-%d')}")
        
        # Process portfolio account data - map from your sheet names to our internal names
        account_mapping = {
            '401K': '401k Retirement (Manual)',
            'Etrade IRA': 'E*TRADE IRA', 
            'Etrade Taxable': 'E*TRADE Taxable',
            'Schwab IRA': 'Schwab IRA',
            'Schwab Individual': 'Schwab Individual'
        }
        
        for row in range(2, ws.max_row + 1):
            account_name = ws.cell(row=row, column=1).value
            
            if account_name and account_name in account_mapping:
                mapped_account_name = account_mapping[account_name]
                print(f"🔄 Processing portfolio history for {account_name} -> {mapped_account_name}...")
                
                # Process each week's data for this account
                for col_idx, date_str in enumerate(date_headers):
                    col = col_idx + 2  # Excel columns start at B (index 2)
                    portfolio_value = ws.cell(row=row, column=col).value
                    
                    if portfolio_value is not None and portfolio_value != 0:
                        # Save to portfolio history CSV for this week
                        history_file = os.path.join(portfolio_history_dir, f"portfolio_{date_str}.csv")
                        
                        # Read existing data or create new
                        if os.path.exists(history_file):
                            existing_df = pd.read_csv(history_file)
                        else:
                            existing_df = pd.DataFrame(columns=['Account', 'Value', 'Date'])
                        
                        # Update or add this account's data using mapped name
                        account_exists = existing_df['Account'] == mapped_account_name
                        if account_exists.any():
                            existing_df.loc[account_exists, 'Value'] = portfolio_value
                        else:
                            new_row = pd.DataFrame({
                                'Account': [mapped_account_name], 
                                'Value': [portfolio_value], 
                                'Date': [date_str]
                            })
                            existing_df = pd.concat([existing_df, new_row], ignore_index=True)
                        
                        # Save updated data
                        existing_df.to_csv(history_file, index=False)
                
                print(f"✅ Imported portfolio history for {account_name} -> {mapped_account_name}")
        
        print("✅ Portfolio historical data import completed from 'All account weekly totals'!")
        return True
        
    except Exception as e:
        print(f"❌ Error importing portfolio historical data: {e}")
        import traceback
        traceback.print_exc()
        return False


def import_historical_data_from_excel():
    """Import historical tracking data from Excel sheet and create CSV files"""
    
    print("🔄 Looking for Historical_Import sheet...")
    
    try:
        wb = openpyxl.load_workbook(TARGET_FILE)
        
        # Debug: Print all sheet names to see what's actually there
        print(f"🔍 Available sheets: {wb.sheetnames}")
        
        if "Historical_Import" not in wb.sheetnames:
            print("⚠️ No 'Historical_Import' sheet found. Please create one with your historical data.")
            print("🔍 Make sure the sheet is named exactly 'Historical_Import' (case sensitive)")
            return False
        
        ws = wb["Historical_Import"]
        
        # Debug: Check the actual content of the first few rows and columns
        print("🔍 Debugging sheet content:")
        print(f"Sheet dimensions: {ws.max_row} rows x {ws.max_column} columns")
        for row in range(1, min(6, ws.max_row + 1)):  # Check first 5 rows or max rows
            row_values = []
            for col in range(1, min(11, ws.max_column + 1)):  # Check first 10 columns or max columns
                cell_value = ws.cell(row=row, column=col).value
                row_values.append(str(cell_value) if cell_value is not None else "None")
            print(f"  Row {row}: {row_values[:5]}...")  # Show first 5 values
        
        # Read the data from the sheet
        data = []
        
        # Based on the debugging output, the structure is:
        # Row 1: [None, date1, date2, date3, ...]
        # Row 2+: [account_name, value1, value2, value3, ...]
        
        # Get date headers from first row (skip column 1 which is None)
        # Excel formulas like =B1+7 will be evaluated only for the first cell,
        # then we need to calculate the progression manually
        date_headers = []
        base_date = None
        
        for col in range(2, ws.max_column + 1):
            header_cell = ws.cell(row=1, column=col)
            header_value = header_cell.value
            
            if header_value:
                if isinstance(header_value, datetime):
                    # This is the first evaluated date
                    base_date = header_value
                    date_headers.append(header_value.strftime('%m/%d/%Y'))
                    print(f"📅 Column {col}: Base date {header_value.strftime('%m/%d/%Y')}")
                elif str(header_value).startswith('=') and base_date:
                    # This is a formula - calculate the date manually
                    # Each formula adds 7 days to the previous date
                    week_offset = col - 2  # How many weeks from the base date
                    calculated_date = base_date + pd.Timedelta(days=7 * week_offset)
                    date_headers.append(calculated_date.strftime('%m/%d/%Y'))
                    print(f"📅 Column {col}: Calculated date {calculated_date.strftime('%m/%d/%Y')} (formula: {header_value})")
                else:
                    # Handle other formats
                    date_str = str(header_value).strip()
                    try:
                        parsed_date = datetime.strptime(date_str, '%Y-%m-%d %H:%M:%S')
                        if not base_date:
                            base_date = parsed_date
                        date_headers.append(parsed_date.strftime('%m/%d/%Y'))
                        print(f"📅 Column {col}: Parsed datetime string to {parsed_date.strftime('%m/%d/%Y')}")
                    except:
                        print(f"⚠️ Column {col}: Could not parse '{date_str}' as date")
                        break
            else:
                break
        
        print(f"📋 Generated {len(date_headers)} date columns: {date_headers[:3]}..." if len(date_headers) > 3 else f"📋 Generated date headers: {date_headers}")
        
        # Read account data rows
        for row in range(2, ws.max_row + 1):
            account_name = ws.cell(row=row, column=1).value
            if not account_name or account_name == 'TOTAL':
                continue
                
            row_data = {'Account Type': str(account_name).strip()}
            has_data = False
            
            # Read values for each date column
            for col_idx, date_header in enumerate(date_headers, start=2):
                cell = ws.cell(row=row, column=col_idx)
                cell_value = cell.value
                
                if cell_value is not None and cell_value != '':
                    try:
                        # Handle currency formatting - the value should already be numeric
                        # even if displayed as currency in Excel
                        if isinstance(cell_value, (int, float)):
                            row_data[date_header] = float(cell_value)
                            has_data = True
                        else:
                            # Try to convert string values (in case of currency symbols)
                            value_str = str(cell_value).replace('$', '').replace(',', '').strip()
                            if value_str:
                                row_data[date_header] = float(value_str)
                                has_data = True
                    except (ValueError, TypeError) as e:
                        # Skip non-numeric values but log them
                        print(f"⚠️ Skipping non-numeric value in {account_name}, column {date_header}: {cell_value}")
                        pass
            
            if has_data:
                data.append(row_data)
        
        print(f"📄 Found {len(data)} account rows with data")
        
        # Process the imported data and create historical CSV files
        if not os.path.exists(HISTORY_DIR):
            os.makedirs(HISTORY_DIR)
        
        # Group data by week dates (assuming the date headers are dates)
        date_columns = date_headers  # Use the date headers we extracted
        
        for date_col in date_columns:
            try:
                # Parse the date from column header - try multiple formats
                parsed_date = None
                
                # Try MM/DD/YYYY format first
                if '/' in date_col:
                    try:
                        parsed_date = datetime.strptime(date_col, '%m/%d/%Y')
                    except:
                        try:
                            parsed_date = datetime.strptime(date_col, '%m/%d/%y')
                        except:
                            pass
                
                # Try other formats like 12/29/2024
                if not parsed_date and '/' in date_col:
                    try:
                        parsed_date = datetime.strptime(date_col, '%m/%d/%Y')
                    except:
                        pass
                
                # Skip if we can't parse the date
                if not parsed_date:
                    print(f"⚠️ Skipping column '{date_col}' - cannot parse as date")
                    continue
                
                week_date = parsed_date.strftime('%Y-%m-%d')
                history_file = os.path.join(HISTORY_DIR, f"estimates_{week_date}.csv")
                
                print(f"🔄 Processing date column '{date_col}' -> {week_date}")
                
                # Create data for this week
                week_data = []
                
                for row in data:
                    account_name = row.get('Account Type', '')
                    if not account_name or account_name == 'TOTAL':
                        continue
                    
                    # Map account names back to internal format
                    account_mapping = {
                        'E*TRADE IRA': 'ETRADE_IRA',
                        'E*TRADE Taxable': 'ETRADE_Taxable',
                        'Schwab Individual': 'Schwab_Individual',
                        'Schwab IRA': 'Schwab_IRA'
                    }
                    
                    internal_account = account_mapping.get(account_name, account_name)
                    total_value = row.get(date_col, 0)
                    
                    if total_value and total_value != 0:
                        # The total_value from Excel is the yearly total, so convert to monthly
                        monthly_total = float(total_value) / 12
                        
                        # Create a single entry representing the monthly total for this account/week
                        week_data.append({
                            'year': parsed_date.year,
                            'month': 'Total',  # Use a single entry for the total
                            'Est. Income $': monthly_total,
                            'account_type': internal_account,
                            'week_captured': week_date
                        })
                
                # Save the CSV file
                if week_data:
                    df = pd.DataFrame(week_data)
                    df.to_csv(history_file, index=False)
                    print(f"✅ Created: {history_file}")
                
            except Exception as e:
                print(f"⚠️ Could not process date column '{date_col}': {e}")
        
        print("🎉 Historical data import completed!")
        return True
        
    except Exception as e:
        print(f"❌ Error importing historical data: {e}")
        return False

def load_all_historical_data():
    """Load all historical weekly data and combine with current week"""
    if not os.path.exists(HISTORY_DIR):
        return pd.DataFrame()
    
    # Get all history files
    history_files = [f for f in os.listdir(HISTORY_DIR) if f.startswith('estimates_') and f.endswith('.csv')]
    
    if not history_files:
        return pd.DataFrame()
    
    all_weeks_data = []
    
    for file in sorted(history_files):
        file_path = os.path.join(HISTORY_DIR, file)
        try:
            # Extract date from filename (estimates_2025-07-26.csv)
            week_date = file.replace('estimates_', '').replace('.csv', '')
            
            df = pd.read_csv(file_path)
            df['week_date'] = week_date
            all_weeks_data.append(df)
        except Exception as e:
            print(f"⚠️ Could not load {file}: {e}")
    
    if all_weeks_data:
        return pd.concat(all_weeks_data, ignore_index=True)
    else:
        return pd.DataFrame()

def create_estimated_income_sheet(wb, use_api=False, use_hybrid=False):
    """Create weekly progression tracking sheet matching manual format"""
    
    # Remove existing sheet if it exists
    if "Estimated Income 2025" in wb.sheetnames:
        del wb["Estimated Income 2025"]
        print("🗑️ Removed existing Estimated Income 2025 sheet")
    
    # Create new sheet
    ws = wb.create_sheet(title="Estimated Income 2025", index=0)
    
    # Load current week's estimate data
    account_data = load_estimate_files(use_api=use_api, use_hybrid=use_hybrid)
    
    if not account_data:
        ws['A1'] = "No estimate files found"
        return ws
    
    # Calculate current week's monthly totals
    current_monthly_summaries = calculate_monthly_totals(account_data)
    
    # Load all historical data
    all_historical_data = load_all_historical_data()
    
    # Create the weekly progression data structure
    weekly_progression = {}
    
    # Get all unique week dates (sorted)
    if not all_historical_data.empty:
        all_weeks = sorted(all_historical_data['week_date'].unique())
    else:
        all_weeks = [datetime.now().strftime('%Y-%m-%d')]
    
    # Account mapping to match your manual sheet
    account_mapping = {
        'ETRADE_IRA': 'E*TRADE IRA',
        'ETRADE_Taxable': 'E*TRADE Taxable'
        # Add more mappings as needed for Schwab accounts
    }
    
    # Build progression data
    for account_type in current_monthly_summaries.keys():
        account_display_name = account_mapping.get(account_type, account_type)
        weekly_progression[account_display_name] = {}
        
        for week_date in all_weeks:
            # Get data for this week and account
            if not all_historical_data.empty:
                week_data = all_historical_data[
                    (all_historical_data['week_date'] == week_date) & 
                    (all_historical_data['account_type'] == account_type)
                ]
                if not week_data.empty:
                    # Sum all monthly estimates for this account/week (should be just one entry now)
                    total_estimate = week_data['Est. Income $'].sum()
                    weekly_progression[account_display_name][week_date] = total_estimate
                else:
                    weekly_progression[account_display_name][week_date] = 0
            else:
                # First time running - use current data
                if account_type in current_monthly_summaries:
                    # Sum the monthly estimates to get the total monthly dividend income
                    total_estimate = current_monthly_summaries[account_type]['Est. Income $'].sum()
                    weekly_progression[account_display_name][week_date] = total_estimate
                else:
                    weekly_progression[account_display_name][week_date] = 0
    
    # Create the sheet layout matching your manual format
    current_row = 1
    
    # Create date headers across the top
    ws.cell(row=current_row, column=1, value="Account Type")
    ws.cell(row=current_row, column=1).font = HEADER_FONT
    
    for col_idx, week_date in enumerate(all_weeks, start=2):
        # Format date for display as MM/DD/YYYY (US format)
        try:
            parsed_date = datetime.strptime(week_date, '%Y-%m-%d')
            display_date = f"{parsed_date.month:02d}/{parsed_date.day:02d}/{parsed_date.year}"
        except:
            display_date = week_date
        ws.cell(row=current_row, column=col_idx, value=display_date)
        ws.cell(row=current_row, column=col_idx).font = HEADER_FONT
        ws.cell(row=current_row, column=col_idx).border = BOX_BORDER
    
    current_row += 1
    
    # Add account rows
    total_by_week = {}  # Track totals for bottom row
    
    for account_name, week_data in weekly_progression.items():
        ws.cell(row=current_row, column=1, value=account_name)
        ws.cell(row=current_row, column=1).font = NORMAL_FONT
        
        previous_value = None
        
        for col_idx, week_date in enumerate(all_weeks, start=2):
            value = week_data.get(week_date, 0)
            
            # Add to weekly totals
            if week_date not in total_by_week:
                total_by_week[week_date] = 0
            total_by_week[week_date] += value
            
            cell = ws.cell(row=current_row, column=col_idx, value=value)
            cell.number_format = FORMAT_CURRENCY_USD_SIMPLE
            cell.font = NORMAL_FONT
            cell.border = BOX_BORDER
            
            # Color coding based on change from previous week
            if previous_value is not None and value != 0:
                if value > previous_value:
                    cell.fill = GREEN_FILL
                elif value < previous_value:
                    cell.fill = RED_FILL
            
            previous_value = value
        
        current_row += 1
    
    # Add total row (matching your red formatting)
    ws.cell(row=current_row, column=1, value="TOTAL")
    ws.cell(row=current_row, column=1).font = Font(bold=True, name="Arial", size=12)
    
    previous_total = None
    
    for col_idx, week_date in enumerate(all_weeks, start=2):
        total_value = total_by_week.get(week_date, 0)
        
        total_cell = ws.cell(row=current_row, column=col_idx, value=total_value)
        total_cell.number_format = FORMAT_CURRENCY_USD_SIMPLE
        total_cell.font = Font(bold=True, name="Arial", size=12)
        total_cell.border = BOX_BORDER
        
        # Color coding for total row changes
        if previous_total is not None and total_value != 0:
            if total_value > previous_total:
                total_cell.fill = GREEN_FILL
            elif total_value < previous_total:
                total_cell.fill = RED_FILL
        
        previous_total = total_value
    
    # Auto-size columns
    for col in range(1, len(all_weeks) + 2):
        ws.column_dimensions[ws.cell(row=1, column=col).column_letter].width = 15

    # Add Portfolio Value Tracking Section
    current_row += 3  # Leave some space between sections
    
    # Portfolio tracking header
    ws.cell(row=current_row, column=1, value="PORTFOLIO VALUE TRACKING")
    ws.cell(row=current_row, column=1).font = Font(bold=True, name="Arial", size=14)
    current_row += 2
    
    # Portfolio section headers
    ws.cell(row=current_row, column=1, value="Account Type")
    ws.cell(row=current_row, column=1).font = HEADER_FONT
    
    for col_idx, week_date in enumerate(all_weeks, start=2):
        try:
            parsed_date = datetime.strptime(week_date, '%Y-%m-%d')
            display_date = f"{parsed_date.month:02d}/{parsed_date.day:02d}/{parsed_date.year}"
        except:
            display_date = week_date
        ws.cell(row=current_row, column=col_idx, value=display_date)
        ws.cell(row=current_row, column=col_idx).font = HEADER_FONT
        ws.cell(row=current_row, column=col_idx).border = BOX_BORDER
    
    current_row += 1
    
    # Load portfolio value tracking data
    try:
        from portfolio_config import PortfolioConfig
        portfolio_config = PortfolioConfig()
        current_portfolio_values = portfolio_config.get_current_portfolio_values(use_api=True)
        
        # Load historical portfolio data
        historical_portfolio = load_historical_portfolio_data()
        
        # Define portfolio accounts to track
        portfolio_accounts = ['E*TRADE IRA', 'E*TRADE Taxable', 'Schwab Individual', 'Schwab IRA', '401k Retirement']
        
        # Portfolio tracking for each account
        portfolio_total_by_week = {}
        
        for account in portfolio_accounts:
            ws.cell(row=current_row, column=1, value=account)
            ws.cell(row=current_row, column=1).font = NORMAL_FONT
            
            previous_portfolio_value = None
            
            for col_idx, week_date in enumerate(all_weeks, start=2):
                # First try to get historical data for this week
                value = 0
                if week_date in historical_portfolio and account in historical_portfolio[week_date]:
                    value = historical_portfolio[week_date][account]
                    print(f"📊 Using historical data for {account} on {week_date}: ${value:,.2f}")
                else:
                    # Fall back to current API values for recent weeks
                    if account == 'E*TRADE IRA':
                        value = current_portfolio_values.get('etrade_ira', 0)
                    elif account == 'E*TRADE Taxable':
                        value = current_portfolio_values.get('etrade_taxable', 0)
                    elif account == 'Schwab IRA':
                        value = current_portfolio_values.get('schwab_ira', 0)
                    elif account == 'Schwab Individual':
                        value = current_portfolio_values.get('schwab_individual', 0)
                    elif account == '401k Retirement':
                        value = current_portfolio_values.get('k401_value', 0)
                    
                    # Add placeholder logic for accounts we don't have API data for yet
                    if account in ['Schwab Individual', 'Schwab IRA', '401k Retirement'] and value == 0:
                        if account == 'Schwab IRA':
                            value = 49951.00
                        elif account == 'Schwab Individual':
                            value = 1605.60
                        elif account == '401k Retirement':
                            value = 122122.00
                
                # Track totals
                if week_date not in portfolio_total_by_week:
                    portfolio_total_by_week[week_date] = 0
                portfolio_total_by_week[week_date] += value
                
                cell = ws.cell(row=current_row, column=col_idx, value=value)
                cell.number_format = FORMAT_CURRENCY_USD_SIMPLE
                cell.font = NORMAL_FONT
                cell.border = BOX_BORDER
                
                # Color coding for portfolio value changes
                if previous_portfolio_value is not None and value != 0:
                    if value > previous_portfolio_value:
                        cell.fill = GREEN_FILL
                    elif value < previous_portfolio_value:
                        cell.fill = RED_FILL
                
                previous_portfolio_value = value
            
            current_row += 1
        
        # Portfolio total row
        ws.cell(row=current_row, column=1, value="TOTAL PORTFOLIO")
        ws.cell(row=current_row, column=1).font = Font(bold=True, name="Arial", size=12)
        
        previous_portfolio_total = None
        
        for col_idx, week_date in enumerate(all_weeks, start=2):
            total_portfolio_value = portfolio_total_by_week.get(week_date, 0)
            
            total_cell = ws.cell(row=current_row, column=col_idx, value=total_portfolio_value)
            total_cell.number_format = FORMAT_CURRENCY_USD_SIMPLE
            total_cell.font = Font(bold=True, name="Arial", size=12)
            total_cell.border = BOX_BORDER
            
            # Color coding for portfolio total changes
            if previous_portfolio_total is not None and total_portfolio_value != 0:
                if total_portfolio_value > previous_portfolio_total:
                    total_cell.fill = GREEN_FILL
                elif total_portfolio_value < previous_portfolio_total:
                    total_cell.fill = RED_FILL
            
            previous_portfolio_total = total_portfolio_value
    
    except Exception as e:
        print(f"⚠️ Could not load portfolio data: {e}")
        # Add placeholder row indicating portfolio data unavailable
        ws.cell(row=current_row, column=1, value="Portfolio data unavailable")
        ws.cell(row=current_row, column=1).font = NORMAL_FONT

    print("📊 Weekly progression tracking sheet created with portfolio values")
    return ws

def build_estimated_income_tracker(import_historical=False, use_api=False, use_hybrid=False, include_portfolio=False, create_comprehensive=False, k401_value=None, use_enhanced_sheets=True):
    """
    Main function to build estimated income tracker with optional enhancements
    
    Args:
        use_enhanced_sheets: If True, creates comprehensive sheets with full historical data and color coding
    """
    """Main function to update estimated income tracking
    
    Args:
        import_historical (bool): Import historical data from Excel
        use_api (bool): Use E*TRADE API instead of Excel files
        use_hybrid (bool): Use hybrid approach (API + Excel)
        include_portfolio (bool): Include portfolio value tracking
        create_comprehensive (bool): Create comprehensive retirement dashboard
    """
    
    if not os.path.exists(TARGET_FILE):
        print(f"⚠️ Workbook not found: {TARGET_FILE}")
        return
    
    # Import historical data if requested
    if import_historical:
        print("📥 Importing historical data from Excel...")
        import_historical_data_from_excel()
        print("📥 Importing historical portfolio data from Excel...")
        import_historical_portfolio_data_from_excel()
    
    try:
        wb = openpyxl.load_workbook(TARGET_FILE)
        
        # Create dividend tracking sheet
        create_estimated_income_sheet(wb, use_api=use_api, use_hybrid=use_hybrid)
        
        # Create comprehensive enhanced sheets if requested
        if use_enhanced_sheets:
            print("🌟 Creating comprehensive enhanced sheets with full historical data and color coding...")
            try:
                from modules.comprehensive_tracker import ComprehensiveTracker
                
                # Get current portfolio values for integration
                current_portfolio_values = {}
                if include_portfolio:
                    from modules.portfolio_value_tracker import PortfolioValueTracker
                    tracker = PortfolioValueTracker()
                    
                    # Handle 401k value prompting if not provided
                    if k401_value is None:
                        from modules.gui_prompts import get_k401_value
                        print("🎨 Opening colorful popup for 401k value...")
                        k401_value = get_k401_value()
                        
                        if k401_value is None:
                            print("❌ 401k value input cancelled. Using default value.")
                            k401_value = 125000.00  # Default fallback
                        else:
                            print(f"✅ 401k value received from GUI: ${k401_value:,.2f}")
                    
                    # Get current API values
                    etrade_values = tracker.get_etrade_portfolio_values()
                    if etrade_values:
                        current_portfolio_values.update(etrade_values)
                    
                    try:
                        from modules.schwab_api_integrated import SchwabAPI
                        schwab_api = SchwabAPI()
                        schwab_values = schwab_api.get_account_values()
                        if schwab_values:
                            current_portfolio_values.update(schwab_values)
                    except Exception as e:
                        print(f"⚠️ Could not get Schwab values: {e}")
                
                # Build comprehensive sheets
                comp_tracker = ComprehensiveTracker()
                comp_tracker.wb = wb  # Use existing workbook
                comp_tracker.import_dividend_historical_data()
                comp_tracker.import_portfolio_historical_data()
                comp_tracker.create_enhanced_estimated_income_sheet()
                if include_portfolio:
                    comp_tracker.create_enhanced_portfolio_values_sheet(current_portfolio_values, k401_value)
                
                wb.save(TARGET_FILE)
                print("✅ Enhanced comprehensive tracking sheets created successfully!")
                return
                
            except Exception as e:
                print(f"⚠️ Error creating comprehensive sheets: {e}")
                print("🔄 Falling back to standard sheet creation...")
        
        # Standard sheet creation (fallback)
        # Create comprehensive retirement dashboard if requested
        if create_comprehensive:
            from modules.comprehensive_retirement_tracker import ComprehensiveRetirementTracker
            tracker = ComprehensiveRetirementTracker()
            tracker.create_comprehensive_tracking_sheet(wb)
        # Otherwise create portfolio value tracking sheet if requested
        elif include_portfolio:
            from modules.portfolio_value_tracker import PortfolioValueTracker
            tracker = PortfolioValueTracker()
            
            # Handle 401k value prompting if not provided
            if k401_value is None:
                from modules.gui_prompts import get_k401_value
                print("🎨 Opening colorful popup for 401k value...")
                k401_value = get_k401_value()
                
                if k401_value is None:
                    print("❌ 401k value input cancelled. Using default value.")
                    k401_value = 122122.00  # Default fallback
                else:
                    print(f"✅ 401k value received from GUI: ${k401_value:,.2f}")
            
            tracker.create_portfolio_value_sheet(wb, k401_value=k401_value)
        
        wb.save(TARGET_FILE)
        
        if create_comprehensive:
            print(f"🏠 Comprehensive retirement dashboard created: {TARGET_FILE}")
        else:
            print(f"✅ Estimated income tracker updated: {TARGET_FILE}")
        
    except Exception as e:
        print(f"❌ Error building estimated income tracker: {e}")

if __name__ == "__main__":
    # Check command line arguments
    import sys
    
    import_historical = False
    use_api = False
    use_hybrid = False
    include_portfolio = False
    create_comprehensive = False
    
    if len(sys.argv) > 1:
        for arg in sys.argv[1:]:
            if arg == "--import":
                import_historical = True
            elif arg == "--api":
                use_api = True
            elif arg == "--hybrid":
                use_hybrid = True
            elif arg == "--portfolio":
                include_portfolio = True
            elif arg == "--comprehensive":
                create_comprehensive = True
                
    if use_hybrid:
        print("🔗 Using hybrid approach (API validation + Excel dividend data)")
    elif use_api:
        print("🔗 Using E*TRADE API for data collection")
    else:
        print("📄 Using Excel files for data collection")
        
    if include_portfolio:
        print("📊 Including portfolio value tracking")
        
    if create_comprehensive:
        print("🏠 Creating comprehensive retirement dashboard")
        
    build_estimated_income_tracker(import_historical=import_historical, use_api=use_api, use_hybrid=use_hybrid, include_portfolio=include_portfolio, create_comprehensive=create_comprehensive)