"""
Portfolio History Importer and Tracker
Imports historical portfolio data from Dividends_2025.xlsx and tracks 401K manually
Last Updated: July 27, 2025
"""

import os
import pandas as pd
import openpyxl
from datetime import datetime, timedelta
from openpyxl.styles import Font, Border, Side, PatternFill
from openpyxl.styles.numbers import FORMAT_CURRENCY_USD_SIMPLE

# Configuration
OUTPUT_DIR = os.path.join(os.path.dirname(__file__), "..", "outputs")
TARGET_FILE = os.path.join(OUTPUT_DIR, "Dividends_2025.xlsx")
PORTFOLIO_HISTORY_DIR = os.path.join(os.path.dirname(__file__), "..", "data", "portfolio_history")

# Styling
HEADER_FONT = Font(bold=True, name="Arial", size=12)
NORMAL_FONT = Font(name="Arial", size=12)
BOX_BORDER = Border(
    left=Side(style='thin'), right=Side(style='thin'),
    top=Side(style='thin'), bottom=Side(style='thin')
)

class PortfolioHistoryManager:
    def __init__(self):
        """Initialize the portfolio history manager"""
        self.portfolio_accounts = [
            'E*TRADE IRA',
            'E*TRADE Taxable', 
            'Schwab IRA',
            'Schwab Taxable',
            '401K',  # Special manually-entered account
            'TOTAL'  # Total portfolio value row
        ]
        self.manual_401k_values = {}  # Store weekly 401K values
        
    def import_portfolio_history_from_excel(self):
        """Import portfolio history from existing Dividends_2025.xlsx structure"""
        
        if not os.path.exists(TARGET_FILE):
            print(f"⚠️ Target file not found: {TARGET_FILE}")
            return False
            
        try:
            wb = openpyxl.load_workbook(TARGET_FILE)
            
            # Look for a sheet with portfolio data (could be multiple sheet names)
            portfolio_sheet = None
            for sheet_name in wb.sheetnames:
                if any(keyword in sheet_name.lower() for keyword in ['portfolio', 'values', 'estimated', 'income']):
                    portfolio_sheet = wb[sheet_name]
                    break
            
            if not portfolio_sheet:
                print("⚠️ No portfolio data sheet found")
                return False
                
            print(f"📊 Reading portfolio data from sheet: {portfolio_sheet.title}")
            
            # Create portfolio history directory
            if not os.path.exists(PORTFOLIO_HISTORY_DIR):
                os.makedirs(PORTFOLIO_HISTORY_DIR)
            
            # Find the data section by looking for date patterns in row 1
            date_start_col = None
            for col in range(1, portfolio_sheet.max_column + 1):
                cell_value = portfolio_sheet.cell(row=1, column=col).value
                if cell_value and isinstance(cell_value, datetime):
                    date_start_col = col
                    break
                elif cell_value and str(cell_value).replace('-', '').replace('/', '').isdigit():
                    date_start_col = col
                    break
                    
            if not date_start_col:
                print("⚠️ Could not find date columns")
                return False
                
            print(f"📅 Found date columns starting at column {date_start_col}")
            
            # Extract dates from first row
            dates = []
            base_date = portfolio_sheet.cell(row=1, column=date_start_col).value
            
            # Handle different date formats
            if isinstance(base_date, datetime):
                current_date = base_date
            else:
                # Try to parse string date
                try:
                    current_date = datetime.strptime(str(base_date), '%Y-%m-%d')
                except:
                    try:
                        current_date = datetime.strptime(str(base_date), '%m/%d/%Y')
                    except:
                        print(f"⚠️ Could not parse base date: {base_date}")
                        return False
            
            # Generate weekly dates (B1 + 7, B1 + 14, etc.)
            for col in range(date_start_col, portfolio_sheet.max_column + 1):
                week_offset = col - date_start_col
                week_date = current_date + timedelta(weeks=week_offset)
                dates.append(week_date.strftime('%Y-%m-%d'))
                
            print(f"📅 Processing {len(dates)} weeks of data")
            
            # Process each account row
            for row in range(2, portfolio_sheet.max_row + 1):
                account_name = portfolio_sheet.cell(row=row, column=1).value
                
                if not account_name:
                    continue
                    
                # Check for various account name formats
                account_match = None
                account_str = str(account_name).strip()
                
                # Direct match first
                if account_str in self.portfolio_accounts:
                    account_match = account_str
                # Check for partial matches (case insensitive)
                else:
                    for known_account in self.portfolio_accounts:
                        if known_account.lower() in account_str.lower() or account_str.lower() in known_account.lower():
                            account_match = known_account
                            break
                    
                    # Special handling for total row
                    if 'total' in account_str.lower():
                        account_match = 'TOTAL'
                
                if not account_match:
                    continue
                    
                print(f"📊 Processing {account_match} (found as: {account_str})...")
                
                # Extract portfolio values for each week
                for col_idx, date_str in enumerate(dates):
                    col = date_start_col + col_idx
                    portfolio_value = portfolio_sheet.cell(row=row, column=col).value
                    
                    if portfolio_value is not None and portfolio_value != 0:
                        # Convert to float if it's a string with currency formatting
                        if isinstance(portfolio_value, str):
                            try:
                                portfolio_value = float(portfolio_value.replace('$', '').replace(',', ''))
                            except:
                                continue
                        
                        # Save to weekly portfolio CSV
                        self.save_weekly_portfolio_value(date_str, account_match, portfolio_value)
                        
            print("✅ Portfolio history import completed!")
            return True
            
        except Exception as e:
            print(f"❌ Error importing portfolio history: {e}")
            return False
    
    def save_weekly_portfolio_value(self, date_str, account_name, value):
        """Save a single portfolio value to the weekly CSV file"""
        
        history_file = os.path.join(PORTFOLIO_HISTORY_DIR, f"portfolio_{date_str}.csv")
        
        # Read existing data or create new
        if os.path.exists(history_file):
            try:
                existing_df = pd.read_csv(history_file)
            except:
                existing_df = pd.DataFrame(columns=['Account', 'Value', 'Date'])
        else:
            existing_df = pd.DataFrame(columns=['Account', 'Value', 'Date'])
        
        # Update or add this account's data
        account_exists = existing_df['Account'] == account_name
        if account_exists.any():
            existing_df.loc[account_exists, 'Value'] = value
        else:
            new_row = pd.DataFrame({
                'Account': [account_name],
                'Value': [value],
                'Date': [date_str]
            })
            existing_df = pd.concat([existing_df, new_row], ignore_index=True)
        
        # Save updated data
        existing_df.to_csv(history_file, index=False)
    
    def get_current_401k_value(self):
        """Get the current 401K value for manual entry"""
        
        current_week = datetime.now().strftime('%Y-%m-%d')
        
        # Check if we already have a value for this week
        if current_week in self.manual_401k_values:
            return self.manual_401k_values[current_week]
        
        # Try to get from latest portfolio file
        history_file = os.path.join(PORTFOLIO_HISTORY_DIR, f"portfolio_{current_week}.csv")
        if os.path.exists(history_file):
            try:
                df = pd.read_csv(history_file)
                k401_row = df[df['Account'] == '401K']
                if not k401_row.empty:
                    return k401_row.iloc[0]['Value']
            except:
                pass
        
        return None
    
    def update_401k_value(self, new_value):
        """Manually update the 401K value for current week"""
        
        current_week = datetime.now().strftime('%Y-%m-%d')
        
        print(f"💰 Updating 401K value for week {current_week}: ${new_value:,.2f}")
        
        # Save to memory
        self.manual_401k_values[current_week] = new_value
        
        # Save to CSV file
        self.save_weekly_portfolio_value(current_week, '401K', new_value)
        
        print("✅ 401K value updated successfully!")
    
    def get_portfolio_summary_for_week(self, date_str=None):
        """Get portfolio summary for a specific week"""
        
        if not date_str:
            date_str = datetime.now().strftime('%Y-%m-%d')
        
        history_file = os.path.join(PORTFOLIO_HISTORY_DIR, f"portfolio_{date_str}.csv")
        
        if not os.path.exists(history_file):
            return {}
        
        try:
            df = pd.read_csv(history_file)
            summary = {}
            
            for _, row in df.iterrows():
                account = row['Account']
                value = float(row['Value']) if pd.notna(row['Value']) else 0
                summary[account] = value
            
            # The TOTAL row is now included in the data, no need to calculate separately
            return summary
            
        except Exception as e:
            print(f"⚠️ Error getting portfolio summary: {e}")
            return {}
    
    def create_portfolio_tracking_sheet(self, wb):
        """Create a portfolio tracking sheet with 401K manual entry"""
        
        # Remove existing sheet if it exists
        sheet_name = "Portfolio Tracker"
        if sheet_name in wb.sheetnames:
            del wb[sheet_name]
            print(f"🗑️ Removed existing {sheet_name} sheet")
        
        # Create new sheet
        ws = wb.create_sheet(title=sheet_name, index=1)
        
        current_row = 1
        
        # === HEADER ===
        ws.cell(row=current_row, column=1, value="Portfolio Value Tracker")
        ws.cell(row=current_row, column=1).font = HEADER_FONT
        current_row += 2
        
        # === CURRENT WEEK SUMMARY ===
        current_week = datetime.now().strftime('%Y-%m-%d')
        ws.cell(row=current_row, column=1, value=f"Week of {current_week}")
        ws.cell(row=current_row, column=1).font = HEADER_FONT
        current_row += 1
        
        # Get current portfolio values
        current_summary = self.get_portfolio_summary_for_week()
        
        if current_summary:
            # Headers
            ws.cell(row=current_row, column=1, value="Account").font = HEADER_FONT
            ws.cell(row=current_row, column=1).border = BOX_BORDER
            ws.cell(row=current_row, column=2, value="Value").font = HEADER_FONT
            ws.cell(row=current_row, column=2).border = BOX_BORDER
            current_row += 1
            
            # Account values
            for account in self.portfolio_accounts:
                value = current_summary.get(account, 0)
                
                ws.cell(row=current_row, column=1, value=account).border = BOX_BORDER
                
                value_cell = ws.cell(row=current_row, column=2, value=value)
                value_cell.border = BOX_BORDER
                value_cell.number_format = FORMAT_CURRENCY_USD_SIMPLE
                
                # Highlight 401K for manual entry
                if account == '401K':
                    value_cell.fill = PatternFill(start_color="FFFF99", end_color="FFFF99", fill_type="solid")
                
                # Highlight TOTAL row
                if account == 'TOTAL':
                    ws.cell(row=current_row, column=1).font = HEADER_FONT
                    value_cell.font = HEADER_FONT
                    value_cell.fill = PatternFill(start_color="D3D3D3", end_color="D3D3D3", fill_type="solid")
                
                current_row += 1
            
            # Note: TOTAL row is now included in the account data
            
        current_row += 3
        
        # === MANUAL 401K UPDATE SECTION ===
        ws.cell(row=current_row, column=1, value="401K Manual Update")
        ws.cell(row=current_row, column=1).font = HEADER_FONT
        current_row += 1
        
        ws.cell(row=current_row, column=1, value="Current 401K Value:")
        current_401k = self.get_current_401k_value()
        
        k401_cell = ws.cell(row=current_row, column=2, value=current_401k if current_401k else "Enter manually")
        k401_cell.number_format = FORMAT_CURRENCY_USD_SIMPLE
        k401_cell.fill = PatternFill(start_color="FFFF99", end_color="FFFF99", fill_type="solid")
        current_row += 2
        
        ws.cell(row=current_row, column=1, value="Instructions:")
        current_row += 1
        ws.cell(row=current_row, column=1, value="1. Update 401K value manually each week")
        current_row += 1
        ws.cell(row=current_row, column=1, value="2. Use: portfolio.update_401k_value(new_amount)")
        current_row += 1
        ws.cell(row=current_row, column=1, value="3. Or edit the highlighted cell above")
        
        # Auto-size columns
        ws.column_dimensions['A'].width = 20
        ws.column_dimensions['B'].width = 15
        
        print("📊 Portfolio tracking sheet created with 401K manual entry")
        return ws

def update_401k_interactive():
    """Interactive function to update 401K value"""
    
    portfolio = PortfolioHistoryManager()
    
    current_value = portfolio.get_current_401k_value()
    
    print("=== 401K Value Update ===")
    if current_value:
        print(f"Current 401K value: ${current_value:,.2f}")
    else:
        print("No current 401K value found")
    
    try:
        new_value_str = input("Enter new 401K value (numbers only): $")
        new_value = float(new_value_str.replace(',', '').replace('$', ''))
        
        portfolio.update_401k_value(new_value)
        
        # Show updated summary
        summary = portfolio.get_portfolio_summary_for_week()
        if summary:
            print(f"\n📊 Updated Portfolio Summary:")
            for account, value in summary.items():
                print(f"  {account}: ${value:,.2f}")
                
    except ValueError:
        print("⚠️ Invalid value entered")
    except KeyboardInterrupt:
        print("\n⚠️ Update cancelled")

if __name__ == "__main__":
    portfolio = PortfolioHistoryManager()
    
    print("=== Portfolio History Manager ===")
    print("1. Import historical data from Excel")
    print("2. Update 401K value")
    print("3. Show current portfolio summary")
    
    choice = input("Select option (1-3): ")
    
    if choice == "1":
        portfolio.import_portfolio_history_from_excel()
    elif choice == "2":
        update_401k_interactive()
    elif choice == "3":
        summary = portfolio.get_portfolio_summary_for_week()
        if summary:
            print(f"\n📊 Current Portfolio Summary:")
            for account, value in summary.items():
                print(f"  {account}: ${value:,.2f}")
        else:
            print("No portfolio data found for current week")
    else:
        print("Invalid option")
