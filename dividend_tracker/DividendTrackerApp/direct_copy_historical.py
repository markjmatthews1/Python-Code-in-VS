#!/usr/bin/env python3
"""
Direct copy from Historical_Import and Portfolio_History_import to Estimated Income 2025
"""

import openpyxl
from openpyxl.styles import Font, Border, Side, PatternFill
from openpyxl.styles.numbers import FORMAT_CURRENCY_USD_SIMPLE
import os

OUTPUT_FILE = os.path.join(os.path.dirname(__file__), "outputs", "Dividends_2025.xlsx")

# Styling
HEADER_FONT = Font(bold=True, name="Arial", size=12)
SECTION_HEADER_FONT = Font(bold=True, name="Arial", size=14)
NORMAL_FONT = Font(name="Arial", size=12)
BOX_BORDER = Border(
    left=Side(style='thin'), right=Side(style='thin'),
    top=Side(style='thin'), bottom=Side(style='thin')
)
GREEN_FILL = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
RED_FILL = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")

def copy_historical_data_directly():
    """Direct copy approach - copy data exactly as it appears in source sheets"""
    
    if not os.path.exists(OUTPUT_FILE):
        print(f"❌ File not found: {OUTPUT_FILE}")
        return
    
    wb = openpyxl.load_workbook(OUTPUT_FILE, data_only=False)  # Keep formulas
    
    print("📊 Available sheets:", wb.sheetnames)
    
    # Check source sheets exist
    if "Historical_Import" not in wb.sheetnames:
        print("❌ Historical_Import sheet not found")
        return
    
    if "Portfolio_History_import" not in wb.sheetnames:
        print("❌ Portfolio_History_import sheet not found")  
        return
    
    hist_ws = wb["Historical_Import"]
    port_ws = wb["Portfolio_History_import"]
    
    print(f"\n📋 Historical_Import: {hist_ws.max_row} rows x {hist_ws.max_column} columns")
    print(f"📋 Portfolio_History_import: {port_ws.max_row} rows x {port_ws.max_column} columns")
    
    # Remove existing Estimated Income sheet
    if "Estimated Income 2025" in wb.sheetnames:
        del wb["Estimated Income 2025"]
        print("🗑️ Removed existing Estimated Income 2025 sheet")
    
    # Create new sheet
    ws = wb.create_sheet(title="Estimated Income 2025", index=0)
    
    current_row = 1
    
    # === SECTION 1: DIVIDEND ESTIMATES (Copy from Historical_Import) ===
    ws.cell(row=current_row, column=1, value="ESTIMATED MONTHLY DIVIDEND INCOME")
    ws.cell(row=current_row, column=1).font = SECTION_HEADER_FONT
    current_row += 2
    
    # Copy dividend data directly from Historical_Import
    print("\n📊 Copying dividend data from Historical_Import...")
    dividend_start_row = current_row
    
    for row_idx in range(1, hist_ws.max_row + 1):
        for col_idx in range(1, hist_ws.max_column + 1):
            source_cell = hist_ws.cell(row=row_idx, column=col_idx)
            target_cell = ws.cell(row=current_row, column=col_idx)
            
            # Copy value (could be formula or data)
            if source_cell.data_type == 'f':  # Formula
                target_cell.value = source_cell.value
            else:
                target_cell.value = source_cell.value
            
            # Copy formatting
            target_cell.font = HEADER_FONT if row_idx == 1 else NORMAL_FONT
            target_cell.border = BOX_BORDER
            
            # Apply currency formatting for numeric values
            if isinstance(source_cell.value, (int, float)) and source_cell.value != 0:
                target_cell.number_format = FORMAT_CURRENCY_USD_SIMPLE
        
        current_row += 1
    
    dividend_end_row = current_row - 1
    current_row += 3  # Space between sections
    
    # === SECTION 2: PORTFOLIO VALUES (Copy from Portfolio_History_import) ===
    ws.cell(row=current_row, column=1, value="PORTFOLIO VALUE TRACKING")
    ws.cell(row=current_row, column=1).font = SECTION_HEADER_FONT
    current_row += 2
    
    # Copy portfolio data directly from Portfolio_History_import
    print("📊 Copying portfolio data from Portfolio_History_import...")
    portfolio_start_row = current_row
    
    for row_idx in range(1, port_ws.max_row + 1):
        for col_idx in range(1, port_ws.max_column + 1):
            source_cell = port_ws.cell(row=row_idx, column=col_idx)
            target_cell = ws.cell(row=current_row, column=col_idx)
            
            # Copy value (could be formula or data)
            if source_cell.data_type == 'f':  # Formula
                target_cell.value = source_cell.value
            else:
                target_cell.value = source_cell.value
            
            # Copy formatting
            target_cell.font = HEADER_FONT if row_idx == 1 else NORMAL_FONT
            target_cell.border = BOX_BORDER
            
            # Apply currency formatting for numeric values
            if isinstance(source_cell.value, (int, float)) and source_cell.value != 0:
                target_cell.number_format = FORMAT_CURRENCY_USD_SIMPLE
        
        current_row += 1
    
    # Add color coding for changes week-over-week
    print("🎨 Adding color coding for week-over-week changes...")
    add_color_coding_to_sheet(ws, dividend_start_row, dividend_end_row, portfolio_start_row, current_row - 1)
    
    # Auto-size columns
    max_column = max(hist_ws.max_column, port_ws.max_column)
    for col in range(1, max_column + 1):
        ws.column_dimensions[ws.cell(row=1, column=col).column_letter].width = 15
    
    # Save workbook
    wb.save(OUTPUT_FILE)
    print(f"\n✅ Successfully rebuilt Estimated Income 2025 sheet!")
    print(f"📊 Copied {hist_ws.max_row} rows of dividend data")
    print(f"📊 Copied {port_ws.max_row} rows of portfolio data")
    print(f"📊 Sheet now includes all historical data through current week")

def add_color_coding_to_sheet(ws, div_start, div_end, port_start, port_end):
    """Add color coding for week-over-week changes"""
    
    # Find the data columns (skip first column which is account names)
    max_col = ws.max_column
    
    # Color code dividend section
    for row in range(div_start + 1, div_end + 1):  # Skip header row
        for col in range(3, max_col + 1):  # Start from column 3 (second data column)
            current_cell = ws.cell(row=row, column=col)
            previous_cell = ws.cell(row=row, column=col-1)
            
            if (isinstance(current_cell.value, (int, float)) and 
                isinstance(previous_cell.value, (int, float)) and
                current_cell.value != 0 and previous_cell.value != 0):
                
                if current_cell.value > previous_cell.value:
                    current_cell.fill = GREEN_FILL
                elif current_cell.value < previous_cell.value:
                    current_cell.fill = RED_FILL
    
    # Color code portfolio section
    for row in range(port_start + 1, port_end + 1):  # Skip header row
        for col in range(3, max_col + 1):  # Start from column 3 (second data column)
            current_cell = ws.cell(row=row, column=col)
            previous_cell = ws.cell(row=row, column=col-1)
            
            if (isinstance(current_cell.value, (int, float)) and 
                isinstance(previous_cell.value, (int, float)) and
                current_cell.value != 0 and previous_cell.value != 0):
                
                if current_cell.value > previous_cell.value:
                    current_cell.fill = GREEN_FILL
                elif current_cell.value < previous_cell.value:
                    current_cell.fill = RED_FILL

if __name__ == "__main__":
    copy_historical_data_directly()
