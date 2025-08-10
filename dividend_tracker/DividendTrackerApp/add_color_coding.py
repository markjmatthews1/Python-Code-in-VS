#!/usr/bin/env python3
"""
Add color coding to total rows by evaluating formulas
"""

import openpyxl
from openpyxl.styles import PatternFill
import os

OUTPUT_FILE = os.path.join(os.path.dirname(__file__), "outputs", "Dividends_2025.xlsx")

GREEN_FILL = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
RED_FILL = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")

def add_total_row_color_coding():
    """Add color coding to total rows based on week-over-week changes"""
    
    wb = openpyxl.load_workbook(OUTPUT_FILE, data_only=True)  # Load with calculated values
    
    if "Estimated Income 2025" not in wb.sheetnames:
        print("❌ Estimated Income 2025 sheet not found")
        return
    
    ws = wb["Estimated Income 2025"]
    
    print("🎨 Adding color coding to total rows...")
    
    # Color code dividend total row (row 8)
    dividend_total_row = 8
    print(f"🎨 Coloring dividend totals (row {dividend_total_row})")
    
    for col in range(3, ws.max_column + 1):  # Start from column 3 (second data column)
        current_cell = ws.cell(row=dividend_total_row, column=col)
        previous_cell = ws.cell(row=dividend_total_row, column=col-1)
        
        current_value = current_cell.value
        previous_value = previous_cell.value
        
        if (isinstance(current_value, (int, float)) and 
            isinstance(previous_value, (int, float)) and
            current_value != 0 and previous_value != 0):
            
            if current_value > previous_value:
                current_cell.fill = GREEN_FILL
                print(f"   Column {col}: Green (${current_value:.2f} > ${previous_value:.2f})")
            elif current_value < previous_value:
                current_cell.fill = RED_FILL
                print(f"   Column {col}: Red (${current_value:.2f} < ${previous_value:.2f})")
    
    # Color code portfolio total row (row 19)
    portfolio_total_row = 19
    print(f"🎨 Coloring portfolio totals (row {portfolio_total_row})")
    
    for col in range(3, ws.max_column + 1):  # Start from column 3 (second data column)
        current_cell = ws.cell(row=portfolio_total_row, column=col)
        previous_cell = ws.cell(row=portfolio_total_row, column=col-1)
        
        current_value = current_cell.value
        previous_value = previous_cell.value
        
        if (isinstance(current_value, (int, float)) and 
            isinstance(previous_value, (int, float)) and
            current_value != 0 and previous_value != 0):
            
            if current_value > previous_value:
                current_cell.fill = GREEN_FILL
                print(f"   Column {col}: Green (${current_value:,.2f} > ${previous_value:,.2f})")
            elif current_value < previous_value:
                current_cell.fill = RED_FILL
                print(f"   Column {col}: Red (${current_value:,.2f} < ${previous_value:,.2f})")
    
    # Also color code individual account rows for better visualization
    print("🎨 Adding color coding to individual account rows...")
    
    # Dividend account rows (4-7)
    for row in range(4, 8):
        account_name = ws.cell(row=row, column=1).value
        if account_name:
            print(f"   Coloring {account_name} (row {row})")
            for col in range(3, ws.max_column + 1):
                current_cell = ws.cell(row=row, column=col)
                previous_cell = ws.cell(row=row, column=col-1)
                
                current_value = current_cell.value
                previous_value = previous_cell.value
                
                if (isinstance(current_value, (int, float)) and 
                    isinstance(previous_value, (int, float)) and
                    current_value != 0 and previous_value != 0):
                    
                    if current_value > previous_value:
                        current_cell.fill = GREEN_FILL
                    elif current_value < previous_value:
                        current_cell.fill = RED_FILL
    
    # Portfolio account rows (14-18)
    for row in range(14, 19):
        account_name = ws.cell(row=row, column=1).value
        if account_name:
            print(f"   Coloring {account_name} (row {row})")
            for col in range(3, ws.max_column + 1):
                current_cell = ws.cell(row=row, column=col)
                previous_cell = ws.cell(row=row, column=col-1)
                
                current_value = current_cell.value
                previous_value = previous_cell.value
                
                if (isinstance(current_value, (int, float)) and 
                    isinstance(previous_value, (int, float)) and
                    current_value != 0 and previous_value != 0):
                    
                    if current_value > previous_value:
                        current_cell.fill = GREEN_FILL
                    elif current_value < previous_value:
                        current_cell.fill = RED_FILL
    
    # Save workbook
    wb.save(OUTPUT_FILE)
    print(f"\n✅ Successfully added color coding to all rows!")
    print("🟢 Green = Increase from previous week")
    print("🔴 Red = Decrease from previous week")

if __name__ == "__main__":
    add_total_row_color_coding()
