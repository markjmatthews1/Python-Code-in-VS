#!/usr/bin/env python3
"""
Check current workbook sheets and structure
"""

import openpyxl
import os

def check_workbook_structure():
    """Check what sheets currently exist in the workbook"""
    
    workbook_path = os.path.join("outputs", "Dividends_2025.xlsx")
    
    try:
        wb = openpyxl.load_workbook(workbook_path)
        print("📊 Current workbook sheets:")
        for i, sheet_name in enumerate(wb.sheetnames, 1):
            print(f"  {i}. {sheet_name}")
        
        # Check if our target sheets exist
        target_sheets = ["All account weekly totals", "All account weekly dividends"]
        for sheet in target_sheets:
            if sheet in wb.sheetnames:
                print(f"✅ Found data sheet: {sheet}")
                ws = wb[sheet]
                print(f"   - Rows: {ws.max_row}, Columns: {ws.max_column}")
            else:
                print(f"❌ Missing data sheet: {sheet}")
        
        return wb.sheetnames
        
    except Exception as e:
        print(f"❌ Error checking workbook: {e}")
        return []

if __name__ == "__main__":
    check_workbook_structure()
