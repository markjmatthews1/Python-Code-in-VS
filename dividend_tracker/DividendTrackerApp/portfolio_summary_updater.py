import openpyxl
import json
import os
from datetime import datetime

class PortfolioSummaryUpdater:
    """Updates Portfolio Summary sheet values while preserving all formatting"""
    
    def __init__(self):
        self.excel_file = "outputs/Dividends_2025.xlsx"
        self.cache_file = "portfolio_data_cache.json"
    
    def run_update(self):
        """Main update method for Portfolio Summary sheet"""
        print("PORTFOLIO SUMMARY UPDATER")
        print("=" * 35)
        
        # Load cache data
        cache_data = self.load_cache_data()
        if not cache_data:
            return False
        
        # Open workbook
        wb = openpyxl.load_workbook(self.excel_file)
        ws = wb["Portfolio Summary"]
        
        print("STEP 1: Loading current portfolio values...")
        portfolio_values = self.get_portfolio_values(cache_data)
        
        print("STEP 2: Loading dividend estimates...")
        dividend_estimates = self.get_dividend_estimates(cache_data)
        
        print("STEP 3: Calculating performance metrics...")
        performance_data = self.calculate_performance_metrics(portfolio_values)
        
        print("STEP 4: Updating Portfolio Summary values...")
        self.update_current_values(ws, portfolio_values)
        self.update_dividend_estimates(ws, dividend_estimates)
        self.update_account_dividend_breakdown(ws, dividend_estimates)
        self.update_next_update_status(ws, dividend_estimates)
        self.update_account_breakdown(ws, portfolio_values)
        self.update_withdrawal_data(ws)
        self.update_performance_tracking(ws, performance_data)
        self.update_dividend_metrics(ws, dividend_estimates, portfolio_values)
        self.update_dividend_summary_section(ws, dividend_estimates)
        self.update_bito_percentages(ws, cache_data)
        section_positions = self.update_dividend_cuts_section(ws, cache_data)
        self.update_ytd_gain_loss(ws, performance_data)  # New YTD feature
        self.update_dividend_performance_metric(ws, section_positions)  # Dynamic positioning
        self.update_last_updated_date(ws)
        
        # Save workbook
        wb.save(self.excel_file)
        wb.close()
        
        print("\nSUCCESS: Portfolio Summary updated with preserved formatting!")
        return True
    
    def get_portfolio_values(self, cache_data):
        """Get current portfolio values from cache data (use pre-calculated values)"""
        # Use pre-calculated portfolio values from cache instead of calculating from positions
        portfolio_values_data = cache_data.get('portfolio_values', {})
        
        values = {}
        
        # Map cache keys to our internal keys
        values['etrade_ira'] = portfolio_values_data.get('E*TRADE IRA', 0)
        values['etrade_taxable'] = portfolio_values_data.get('E*TRADE Taxable', 0)
        values['schwab_ira'] = portfolio_values_data.get('Schwab IRA', 0)
        values['schwab_individual'] = portfolio_values_data.get('Schwab Individual', 0)
        values['k401_retirement'] = portfolio_values_data.get('401K', 128693.17)
        
        # Calculate total
        values['total_portfolio'] = (values['etrade_ira'] + values['etrade_taxable'] + 
                                   values['schwab_ira'] + values['schwab_individual'] + 
                                   values['k401_retirement'])
        
        print(f"  E*TRADE IRA: ${values['etrade_ira']:,.2f}")
        print(f"  E*TRADE Taxable: ${values['etrade_taxable']:,.2f}") 
        print(f"  Schwab IRA: ${values['schwab_ira']:,.2f}")
        print(f"  Schwab Individual: ${values['schwab_individual']:,.2f}")
        print(f"  401K Retirement: ${values['k401_retirement']:,.2f}")
        print(f"  Total Portfolio: ${values['total_portfolio']:,.2f}")
        
        return values
    
    def get_dividend_estimates(self, cache_data):
        """Get dividend estimates from pre-calculated data in cache"""
        # Use pre-calculated dividend estimates from cache (calculated by portfolio_data_collector.py)
        dividend_estimates_data = cache_data.get('dividend_estimates', {})
        
        if not dividend_estimates_data:
            print("    Warning: No dividend_estimates in cache, calculating from positions...")
            return self.calculate_dividend_estimates_from_positions(cache_data)
        
        # Extract account breakdown from the pre-calculated data
        etrade_ira_annual = dividend_estimates_data.get('E*TRADE IRA', 0)
        etrade_taxable_annual = dividend_estimates_data.get('E*TRADE Taxable', 0)
        schwab_ira_annual = dividend_estimates_data.get('Schwab IRA', 0)
        schwab_individual_annual = dividend_estimates_data.get('Schwab Individual', 0)
        
        total_annual = etrade_ira_annual + etrade_taxable_annual + schwab_ira_annual + schwab_individual_annual
        total_weekly = total_annual / 52
        total_monthly = total_annual / 12
        
        account_breakdown = {
            'etrade_ira': etrade_ira_annual,
            'etrade_taxable': etrade_taxable_annual,
            'schwab_ira': schwab_ira_annual,
            'schwab_individual': schwab_individual_annual
        }
        
        estimates = {
            'weekly': total_weekly,
            'monthly': total_monthly,
            'annual': total_annual,
            'account_breakdown': account_breakdown
        }
        
        print(f"  Weekly Estimate: ${total_weekly:.2f}")
        print(f"  Monthly Estimate: ${total_monthly:.2f}")
        print(f"  Annual Estimate: ${total_annual:.2f}")
        print(f"    E*TRADE IRA: ${etrade_ira_annual:,.2f}")
        print(f"    E*TRADE Taxable: ${etrade_taxable_annual:,.2f}")
        print(f"    Schwab IRA: ${schwab_ira_annual:,.2f}")
        print(f"    Schwab Individual: ${schwab_individual_annual:,.2f}")
        
        return estimates
    
    def calculate_dividend_estimates_from_positions(self, cache_data):
        """Fallback: Calculate dividend estimates from positions if pre-calculated data not available"""
        positions_data = cache_data.get('positions', {})
        yields_data = cache_data.get('ticker_yields', {})
        
        print(f"    Found {len(yields_data)} tickers with yield data in cache")
        
        if not yields_data:
            print("    ERROR: No ticker_yields data found in cache!")
            return {'weekly': 0, 'monthly': 0, 'annual': 0, 'account_breakdown': {}}
        
        total_annual = 0
        account_breakdown = {}
        
        for account, positions in positions_data.items():
            account_annual = 0
            
            for position in positions:
                symbol = position.get('symbol', '').strip().upper()
                quantity = position.get('quantity', 0)
                yield_info = yields_data.get(symbol, {})
                
                # Use the CORRECT calculation method: shares × annual_dividend_per_share
                annual_dividend_per_share = yield_info.get('annual_dividend', 0.0)
                
                if annual_dividend_per_share > 0 and quantity > 0:
                    annual_dividend = quantity * annual_dividend_per_share
                    account_annual += annual_dividend
                    print(f"    {symbol}: {quantity} shares × ${annual_dividend_per_share:.2f} = ${annual_dividend:.2f}/year")
            
            # Map account names
            if account == 'etrade_ira':
                account_breakdown['etrade_ira'] = account_annual
            elif account == 'etrade_taxable':
                account_breakdown['etrade_taxable'] = account_annual
            elif account == 'schwab_ira':
                account_breakdown['schwab_ira'] = account_annual
            elif account == 'schwab_individual':
                account_breakdown['schwab_individual'] = account_annual
            
            total_annual += account_annual
        
        total_weekly = total_annual / 52
        total_monthly = total_annual / 12
        
        estimates = {
            'weekly': total_weekly,
            'monthly': total_monthly,
            'annual': total_annual,
            'account_breakdown': account_breakdown
        }
        
        print(f"  Calculated - Weekly: ${total_weekly:.2f}, Monthly: ${total_monthly:.2f}, Annual: ${total_annual:.2f}")
        
        return estimates
    
    def calculate_performance_metrics(self, portfolio_values):
        """Calculate performance tracking metrics from Portfolio Values 2025 sheet"""
        print("  Calculating performance metrics from Portfolio Values 2025...")
        
        try:
            # Open the Excel file to get historical data
            # Use data_only=True to get calculated formula values
            wb = openpyxl.load_workbook(self.excel_file, data_only=True)
            
            if 'Portfolio Values 2025' not in wb.sheetnames:
                print("    Warning: Portfolio Values 2025 sheet not found, using default values")
                return {'weekly_change_amount': 0, 'weekly_change_percent': 0, 'ytd_amount': 0, 'ytd_percent': 0}
            
            portfolio_ws = wb['Portfolio Values 2025']
            max_col = portfolio_ws.max_column
            
            # Year-start value is in column B (column 2), row 10
            # B10 = SUM(B4:B8) which is the 12/29/2024 beginning balance
            # With data_only=True, this will be the calculated value
            year_start_total = portfolio_ws.cell(row=10, column=2).value
            
            # If data_only didn't work, manually sum B4:B8
            if not isinstance(year_start_total, (int, float)) or year_start_total == 0:
                print("    B10 formula not calculated, manually summing B4:B8...")
                year_start_total = 0
                for row in range(4, 9):  # Rows 4-8
                    val = portfolio_ws.cell(row=row, column=2).value
                    if isinstance(val, (int, float)):
                        year_start_total += val
            
            print(f"    Year Start (B10 = SUM(B4:B8)): ${year_start_total:,.2f}")
            
            # Current week total is in the LAST data column (max_column), row 10
            current_total = portfolio_ws.cell(row=10, column=max_col).value
            
            # Handle formulas and ensure we get a numeric value
            if isinstance(current_total, str) or current_total is None:
                # If max_col is a formula or empty, go backwards to find last numeric value
                print(f"    Column {max_col} is not numeric, searching backwards...")
                for col_offset in range(0, min(5, max_col)):
                    test_col = max_col - col_offset
                    test_val = portfolio_ws.cell(row=10, column=test_col).value
                    if isinstance(test_val, (int, float)) and test_val > 1000:
                        current_total = test_val
                        max_col = test_col
                        print(f"    Using column {max_col} as current week: ${current_total:,.2f}")
                        break
                
                if not isinstance(current_total, (int, float)):
                    print(f"    ERROR: Could not find valid current total")
                    current_total = 0
            else:
                print(f"    Current Week (col {max_col}): ${current_total:,.2f}")
            
            # Previous week total is in the column BEFORE the current one (max_col - 1)
            prev_total = 0
            if max_col > 2:  # Make sure there's a previous column
                prev_total = portfolio_ws.cell(row=10, column=max_col - 1).value
                if not isinstance(prev_total, (int, float)):
                    # If previous column is not numeric, search backwards
                    print(f"    Column {max_col - 1} is not numeric, searching backwards...")
                    for col_offset in range(2, min(10, max_col)):
                        prev_col = max_col - col_offset
                        if prev_col < 3:  # Don't go before column 3
                            break
                        prev_value = portfolio_ws.cell(row=10, column=prev_col).value
                        if isinstance(prev_value, (int, float)) and prev_value > 1000:
                            prev_total = prev_value
                            print(f"    Using column {prev_col} as previous week: ${prev_total:,.2f}")
                            break
                else:
                    print(f"    Previous Week (col {max_col - 1}): ${prev_total:,.2f}")
            
            if prev_total == 0 or prev_total < 1000:
                prev_total = current_total  # Fallback to prevent division by zero
                print(f"    Warning: Using current total as previous (no valid previous week found)")
            
            # Calculate weekly change
            weekly_change_amount = current_total - prev_total
            weekly_change_percent = (weekly_change_amount / prev_total * 100) if prev_total > 0 else 0
            
            # Calculate YTD change from year start (B10)
            ytd_amount = current_total - year_start_total
            ytd_percent = (ytd_amount / year_start_total * 100) if year_start_total > 0 else 0
            
            print(f"    ✅ Weekly Change: ${weekly_change_amount:,.2f} ({weekly_change_percent:+.2f}%)")
            print(f"    ✅ YTD Gain: ${ytd_amount:,.2f} ({ytd_percent:+.1f}%)")
            
            wb.close()
            
            performance = {
                'weekly_change_amount': weekly_change_amount,
                'weekly_change_percent': weekly_change_percent,
                'ytd_amount': ytd_amount,
                'ytd_percent': ytd_percent
            }
            
            return performance
            
        except Exception as e:
            print(f"    Error calculating performance metrics: {e}")
            import traceback
            traceback.print_exc()
            # Return safe default values
            return {
                'weekly_change_amount': 0,
                'weekly_change_percent': 0,
                'ytd_amount': 0,
                'ytd_percent': 0
            }
    
    def update_current_values(self, ws, portfolio_values):
        """Update current portfolio values (Column B, rows 4-9)"""
        print("  Updating current values...")
        
        # Update individual account values
        ws.cell(row=4, column=2).value = portfolio_values.get('etrade_ira', 0)  # E*TRADE IRA
        ws.cell(row=5, column=2).value = portfolio_values.get('etrade_taxable', 0)  # E*TRADE Taxable
        ws.cell(row=6, column=2).value = portfolio_values.get('schwab_ira', 0)  # Schwab IRA  
        ws.cell(row=7, column=2).value = portfolio_values.get('schwab_individual', 0)  # Schwab Individual
        ws.cell(row=8, column=2).value = portfolio_values.get('k401_retirement', 0)  # 401k Retirement
        ws.cell(row=9, column=2).value = portfolio_values.get('total_portfolio', 0)  # Total Portfolio
        
        print(f"    Updated portfolio total: ${portfolio_values['total_portfolio']:,.2f}")
    
    def update_dividend_estimates(self, ws, dividend_estimates):
        """Update dividend estimates (Column E, rows 4-6)"""
        print("  Updating dividend estimates...")
        
        ws.cell(row=4, column=5).value = dividend_estimates['weekly']  # Weekly Estimate
        ws.cell(row=5, column=5).value = dividend_estimates['monthly']  # Monthly Estimate  
        ws.cell(row=6, column=5).value = dividend_estimates['annual']  # Annual Estimate
        
        print(f"    Updated annual estimate: ${dividend_estimates['annual']:,.2f}")
    
    def update_account_dividend_breakdown(self, ws, dividend_estimates):
        """Update account breakdown annual dividends (Column D-E, rows 9-12)"""
        print("  Updating account dividend breakdown...")
        
        account_breakdown = dividend_estimates.get('account_breakdown', {})
        
        # Update annual dividend amounts for each account (this is the "ACCOUNT BREAKDOWN (Annual)" section)
        ws.cell(row=9, column=4).value = "E*TRADE IRA:"  # Label
        ws.cell(row=9, column=5).value = account_breakdown.get('etrade_ira', 0)  # Annual amount
        
        ws.cell(row=10, column=4).value = "E*TRADE Taxable:"  # Label  
        ws.cell(row=10, column=5).value = account_breakdown.get('etrade_taxable', 0)  # Annual amount
        
        ws.cell(row=11, column=4).value = "Schwab IRA:"  # Label
        ws.cell(row=11, column=5).value = account_breakdown.get('schwab_ira', 0)  # Annual amount
        
        ws.cell(row=12, column=4).value = "Schwab Individual:"  # Label
        ws.cell(row=12, column=5).value = account_breakdown.get('schwab_individual', 0)  # Annual amount
        
        print(f"    Updated account breakdown: E*TRADE IRA ${account_breakdown.get('etrade_ira', 0):,.0f}, E*TRADE Taxable ${account_breakdown.get('etrade_taxable', 0):,.0f}")
    
    def update_next_update_status(self, ws, dividend_estimates):
        """Update dividend status/next update (Row 29, Column E)"""
        print("  Updating dividend status...")
        
        # Update the "Next Update" field which shows dividend status
        ws.cell(row=29, column=4).value = "Next Update:"  # Label
        ws.cell(row=29, column=5).value = "Weekly (Automated)"  # Status
        
        print("    Updated dividend status: Next Update - Weekly (Automated)")
    
    
    def update_account_breakdown(self, ws, portfolio_values):
        """Update account breakdown percentages (Column B, rows 14-18)"""
        print("  Updating account breakdown...")
        
        total = portfolio_values['total_portfolio']
        if total > 0:
            ws.cell(row=14, column=2).value = portfolio_values.get('etrade_ira', 0) / total  # E*TRADE IRA %
            ws.cell(row=15, column=2).value = portfolio_values.get('etrade_taxable', 0) / total  # E*TRADE Taxable %
            ws.cell(row=16, column=2).value = portfolio_values.get('schwab_ira', 0) / total  # Schwab IRA %
            ws.cell(row=17, column=2).value = portfolio_values.get('schwab_individual', 0) / total  # Schwab Individual %
            ws.cell(row=18, column=2).value = portfolio_values.get('k401_retirement', 0) / total  # 401k %
    
    def update_withdrawal_data(self, ws):
        """Update withdrawal data - keep existing values for now"""
        print("  Keeping existing withdrawal data...")
        # These values are likely manually set, so preserve them
    
    def update_performance_tracking(self, ws, performance_data):
        """Update performance tracking data"""
        print("  Updating performance tracking...")
        
        # Weekly change (row 10, column B) - Handle positive and negative values  
        weekly_amount = performance_data['weekly_change_amount']
        weekly_percent = performance_data['weekly_change_percent']
        
        if weekly_amount >= 0:
            weekly_change = f"+${weekly_amount:,.2f} (+{weekly_percent:.2f}%)"
        else:
            weekly_change = f"-${abs(weekly_amount):,.2f} ({weekly_percent:.2f}%)"
        
        ws.cell(row=10, column=2).value = weekly_change
        
        # YTD Performance (row 23, column B) - Handle positive and negative values
        ytd_amount = performance_data['ytd_amount']
        ytd_percent = performance_data['ytd_percent']
        
        if ytd_amount >= 0:
            ytd_performance = f"+${ytd_amount:,.0f} (+{ytd_percent:.1f}%)"
        else:
            ytd_performance = f"-${abs(ytd_amount):,.0f} ({ytd_percent:.1f}%)"
            
        ws.cell(row=23, column=2).value = ytd_performance
    
    def update_dividend_metrics(self, ws, dividend_estimates, portfolio_values):
        """Update dividend metrics (Column E, rows 21-28)"""
        print("  Updating dividend metrics...")
        
        # Net Monthly Income (row 21)
        monthly_income = dividend_estimates['monthly'] - 1718.33  # Subtract withdrawals
        ws.cell(row=21, column=5).value = monthly_income
        
        # Net Annual Income (row 22)
        annual_income = dividend_estimates['annual'] - (1718.33 * 12)  # Subtract annual withdrawals
        ws.cell(row=22, column=5).value = annual_income
        
        # Current Yield (row 26)
        current_yield = dividend_estimates['annual'] / portfolio_values['total_portfolio']
        ws.cell(row=26, column=5).value = current_yield
        
        # Monthly Dividend Coverage (row 27)
        monthly_coverage = dividend_estimates['monthly'] / 1718.33 if 1718.33 > 0 else 0
        coverage_text = f"{monthly_coverage:.1f}x"
        ws.cell(row=27, column=5).value = coverage_text
    
    def update_dividend_summary_section(self, ws, dividend_estimates):
        """Update dividend summary section (row 35 - Current Annual Estimate)"""
        print("  Updating dividend summary section...")
        
        # Row 35: Current Annual Estimate should match row 6 
        ws.cell(row=35, column=5).value = dividend_estimates['annual']
        
        print(f"    Updated row 35 annual estimate: ${dividend_estimates['annual']:,.2f}")
    
    def update_bito_percentages(self, ws, cache_data):
        """Update BITO percentage values in column H"""
        print("  Updating BITO percentages...")
        
        # Get BITO yield from cache data
        ticker_yields = cache_data.get('ticker_yields', {})
        bito_data = ticker_yields.get('BITO', {})
        bito_yield = bito_data.get('yield', 52.64)  # Default to current value if not found
        
        # Update row 24 (H24) 
        ws.cell(row=24, column=8).value = f"BITO: {bito_yield:.2f}% (18 total)"
        
        # Update row 25 (H25)
        ws.cell(row=25, column=8).value = f"BITO: {bito_yield:.2f}% (12 total)"
        
        print(f"    Updated BITO percentage to {bito_yield:.2f}%")
    
    def update_last_updated_date(self, ws):
        """Update the last updated date at K1 (just date/time, no label)"""
        print("  Updating last updated date...")
        
        from datetime import datetime
        from openpyxl.styles import Font, Alignment
        
        current_date = datetime.now().strftime("%m/%d %H:%M")
        
        # Row 1, Column K - Just the timestamp, no "Last Updated:" label
        cell = ws.cell(row=1, column=11)
        cell.value = current_date
        cell.font = Font(name='Arial', size=12)
        cell.alignment = Alignment(horizontal='right')
        
        print(f"    Updated timestamp at K1: {current_date}")
    
    def update_dividend_cuts_section(self, ws, cache_data):
        """Update dividend cuts (G&H) and increases (J&K) sections with FULLY DYNAMIC data"""
        print("  Updating dividend cuts and increases sections (fully dynamic)...")
        
        try:
            from openpyxl.styles import Font, PatternFill, Alignment
            
            # Define formatting styles
            header_font = Font(name='Arial', size=12, bold=True)
            header_fill = PatternFill(start_color='C5D9F1', end_color='C5D9F1', fill_type='solid')
            ticker_font = Font(name='Arial', size=12)
            red_font = Font(name='Arial', size=12, color='FF0000')  # Red for reductions
            green_font = Font(name='Arial', size=12, color='228B22')  # Green for increases
            
            # Styles for summary totals at the top
            blue_fill = PatternFill(start_color='4F81BD', end_color='4F81BD', fill_type='solid')
            white_fill = PatternFill(start_color='FFFFFF', end_color='FFFFFF', fill_type='solid')
            white_font_bold = Font(name='Arial', size=12, color='FFFFFF', bold=True)
            right_align = Alignment(horizontal='right')
            
            # Helper function to safely set cell value and format
            def safe_set_cell(row, col, value, font=None, fill=None, alignment=None):
                try:
                    cell = ws.cell(row=row, column=col)
                    cell.value = value
                    if font:
                        cell.font = font
                    if fill:
                        cell.fill = fill
                    if alignment:
                        cell.alignment = alignment
                    return True
                except Exception as e:
                    print(f"    Error setting cell at row {row}, col {col}: {e}")
                    return False
            
            # STEP 1: CLEAR ENTIRE DYNAMIC SECTION (G7:K35) - Delete all previous data
            print("    Clearing previous dividend change data (G7:K35)...")
            for row in range(7, 36):
                safe_set_cell(row, 7, None)   # Column G
                safe_set_cell(row, 8, None)   # Column H  
                safe_set_cell(row, 10, None)  # Column J
                safe_set_cell(row, 11, None)  # Column K
            
            # STEP 2: Get yield changes from individual account sheets
            dividend_cuts, dividend_increases = self.get_dynamic_yield_changes()
            
            # STEP 3: PLACE SUMMARY TOTALS AT THE TOP (Row 5) - Fixed position
            if dividend_cuts or dividend_increases:
                # Calculate averages for summary
                if dividend_cuts:
                    avg_cuts = abs(sum(item['change_percent'] for item in dividend_cuts) / len(dividend_cuts))
                else:
                    avg_cuts = 0
                    
                if dividend_increases:
                    avg_increases = sum(item['change_percent'] for item in dividend_increases) / len(dividend_increases)
                else:
                    avg_increases = 0
                
                # Calculate net performance (increases - cuts)
                net_dividend_performance = avg_increases - avg_cuts
                
                # Row 5, Column G: Dividend Performance Summary with blue background
                safe_set_cell(5, 7, "Dividend Performance:", font=white_font_bold, fill=blue_fill)
                
                # Row 5, Column H: Net performance with conditional formatting
                if net_dividend_performance >= 0:
                    perf_font = Font(name='Arial', size=12, bold=True, color='228B22')  # Green
                    performance_text = f"+{net_dividend_performance:.1f}% Net Gain"
                else:
                    perf_font = Font(name='Arial', size=12, bold=True, color='FF0000')  # Red
                    performance_text = f"{net_dividend_performance:.1f}% Net Loss"
                
                safe_set_cell(5, 8, performance_text, font=perf_font, fill=white_fill, alignment=right_align)
                
                # Row 5, Columns J&K: Summary counts
                safe_set_cell(5, 10, f"Total Changes:", font=white_font_bold, fill=blue_fill)
                summary_text = f"{len(dividend_cuts)} cuts, {len(dividend_increases)} increases"
                safe_set_cell(5, 11, summary_text, font=ticker_font, fill=white_fill, alignment=right_align)
                
                print(f"    Added dividend summary at row 5: {performance_text}")
                print(f"    (Avg Increases: {avg_increases:.1f}% - Avg Cuts: {avg_cuts:.1f}%)")
            
            # STEP 4: Start individual items below the summary (Row 7+)
            cuts_row = 7    # Start cuts at row 7 (below summary)
            increases_row = 7  # Start increases at row 7 (below summary)
            
            # Group dividend changes by account (ALL 4 ACCOUNTS - show even if no changes)
            account_sheet_names = [
                ('Etrade_IRA', 'E*TRADE IRA'),
                ('Etrade_Individual', 'E*TRADE Taxable'), 
                ('Schwab_IRA', 'Schwab IRA'),
                ('Schwab_Individual', 'Schwab Individual')
            ]
            
            # STEP 5: Process dividend CUTS (Columns G & H) - Show ALL accounts dynamically
            for sheet_name, display_name in account_sheet_names:
                # Find cuts for this account (filter list by account)
                account_cuts = [item for item in dividend_cuts if item['account'] == sheet_name]
                
                # Always show account header (even if no cuts)
                safe_set_cell(cuts_row, 7, display_name, font=header_font, fill=header_fill)
                if account_cuts:
                    safe_set_cell(cuts_row, 8, f"{len(account_cuts)} dividend cuts", font=header_font, fill=header_fill)
                else:
                    safe_set_cell(cuts_row, 8, "No dividend cuts", font=header_font, fill=header_fill)
                cuts_row += 1
                
                # Show each cut with RED font for values
                if account_cuts:
                    # Sort by ticker name
                    for item in sorted(account_cuts, key=lambda x: x['ticker']):
                        ticker = item['ticker']
                        old_yield = item['old_yield']
                        new_yield = item['new_yield']
                        change_pct = item['change_percent']
                        
                        safe_set_cell(cuts_row, 7, f"  {ticker} ↓", font=ticker_font)
                        cut_text = f"{change_pct:.1f}% ({old_yield:.2f}% → {new_yield:.2f}%)"
                        safe_set_cell(cuts_row, 8, cut_text, font=red_font)  # RED for reductions
                        cuts_row += 1
                
                # No extra space - keep tight formatting
            
            # STEP 6: Process dividend INCREASES (Columns J & K) - Show ALL accounts dynamically
            for sheet_name, display_name in account_sheet_names:
                # Find increases for this account (filter list by account)
                account_increases = [item for item in dividend_increases if item['account'] == sheet_name]
                
                # Always show account header (even if no increases)
                safe_set_cell(increases_row, 10, display_name, font=header_font, fill=header_fill)
                if account_increases:
                    safe_set_cell(increases_row, 11, f"{len(account_increases)} dividend increases", font=header_font, fill=header_fill)
                else:
                    safe_set_cell(increases_row, 11, "No dividend increases", font=header_font, fill=header_fill)
                increases_row += 1
                
                # Show each increase with GREEN font for values
                if account_increases:
                    # Sort by ticker name
                    for item in sorted(account_increases, key=lambda x: x['ticker']):
                        ticker = item['ticker']
                        old_yield = item['old_yield']
                        new_yield = item['new_yield']
                        change_pct = item['change_percent']
                        
                        safe_set_cell(increases_row, 10, f"  {ticker} ↑", font=ticker_font)
                        increase_text = f"+{change_pct:.1f}% ({old_yield:.2f}% → {new_yield:.2f}%)"
                        safe_set_cell(increases_row, 11, increase_text, font=green_font)  # GREEN for increases
                        increases_row += 1
                
                # No extra space - keep tight formatting
            
            print(f"    ✅ Updated dividend sections: {len(dividend_cuts)} cuts, {len(dividend_increases)} increases across all accounts")
            
            # Return the final row positions (no longer needed for dynamic summary placement)
            return {'cuts_final_row': cuts_row, 'increases_final_row': increases_row}
            
        except Exception as e:
            print(f"    Error updating dividend sections: {e}")
            import traceback
            traceback.print_exc()
            return {'cuts_final_row': 7, 'increases_final_row': 7}  # Safe fallback

    def get_dynamic_yield_changes(self):
        """Extract dividend cuts and increases by comparing individual account sheets Column O vs P"""
        try:
            # Load workbook to read individual account sheets
            import openpyxl
            wb = openpyxl.load_workbook(self.excel_file)
            
            # Store as list of (account, ticker, data) instead of dict to allow duplicates
            dividend_cuts_list = []
            dividend_increases_list = []
            
            # List of individual account sheets to scan
            account_sheets = ['Etrade_IRA', 'Etrade_Individual', 'Schwab_IRA', 'Schwab_Individual']
            
            for sheet_name in account_sheets:
                try:
                    ws = wb[sheet_name]
                    
                    # Scan each account sheet for yield changes (starting from row 3, typical data range)
                    for row in range(3, 100):  # Extended range to catch all tickers
                        ticker_cell = ws.cell(row=row, column=1)  # Column A - Ticker
                        prev_yield_cell = ws.cell(row=row, column=15)  # Column O - Beginning Dividend Yield
                        curr_yield_cell = ws.cell(row=row, column=16)  # Column P - Current Yield
                        
                        if ticker_cell.value and prev_yield_cell.value is not None and curr_yield_cell.value is not None:
                            ticker = str(ticker_cell.value).strip().upper()
                            
                            try:
                                # Handle percentage values
                                prev_yield = float(prev_yield_cell.value)
                                curr_yield = float(curr_yield_cell.value)
                                
                                # Convert from decimal to percentage if needed
                                if prev_yield < 1.0:
                                    prev_yield *= 100
                                if curr_yield < 1.0:
                                    curr_yield *= 100
                                
                                # Calculate change
                                if prev_yield > 0:  # Avoid division by zero
                                    change_percent = ((curr_yield - prev_yield) / prev_yield) * 100
                                    
                                    # Significant changes (more than 1% change)
                                    if abs(change_percent) > 1.0:
                                        change_data = {
                                            'ticker': ticker,
                                            'account': sheet_name,
                                            'old_yield': prev_yield,
                                            'new_yield': curr_yield,
                                            'change_percent': change_percent
                                        }
                                        
                                        if change_percent < 0:  # Dividend cut
                                            dividend_cuts_list.append(change_data)
                                        else:  # Dividend increase
                                            dividend_increases_list.append(change_data)
                            except (ValueError, TypeError):
                                continue  # Skip invalid yield data
                                
                except KeyError:
                    print(f"    Warning: Account sheet '{sheet_name}' not found, skipping...")
                    continue
            
            print(f"    Dynamic analysis found: {len(dividend_cuts_list)} cuts, {len(dividend_increases_list)} increases across all accounts")
            return dividend_cuts_list, dividend_increases_list
            
        except Exception as e:
            print(f"    Warning: Could not read account sheets for dynamic analysis: {e}")
            # Fall back to empty data if sheets unavailable
            return [], []

    def update_ytd_gain_loss(self, ws, performance_data):
        """Update YTD gain/loss at row 11 columns A&B with conditional formatting"""
        print("  Updating YTD gain/loss...")
        
        try:
            from openpyxl.styles import Font, PatternFill, Alignment
            
            # Get YTD data from performance_data using correct keys
            ytd_change = performance_data.get('ytd_amount', 0)
            ytd_percent = performance_data.get('ytd_percent', 0)
            
            # Define styling
            blue_fill = PatternFill(start_color='4F81BD', end_color='4F81BD', fill_type='solid')
            white_fill = PatternFill(start_color='FFFFFF', end_color='FFFFFF', fill_type='solid')
            white_font_bold = Font(name='Arial', size=12, color='FFFFFF', bold=True)  # White bold text
            right_align = Alignment(horizontal='right')  # Right alignment
            
            # Conditional formatting for positive/negative values on white background
            if ytd_change >= 0:
                value_font = Font(name='Arial', size=12, bold=True, color='228B22')  # Green for positive
            else:
                value_font = Font(name='Arial', size=12, bold=True, color='FF0000')  # Red for negative
            
            # Row 11 Column A: Description with white bold text on blue background
            ws.cell(row=11, column=1).value = "YTD Gain/Loss:"
            ws.cell(row=11, column=1).font = white_font_bold
            ws.cell(row=11, column=1).fill = blue_fill
            
            # Row 11 Column B: Value with conditional formatting, white background, right aligned
            if ytd_change >= 0:
                ytd_text = f"+${ytd_change:,.2f} (+{ytd_percent:.1f}%)"
            else:
                ytd_text = f"-${abs(ytd_change):,.2f} ({ytd_percent:.1f}%)"
            
            ws.cell(row=11, column=2).value = ytd_text
            ws.cell(row=11, column=2).font = value_font
            ws.cell(row=11, column=2).fill = white_fill  # White background for better readability
            ws.cell(row=11, column=2).alignment = right_align  # Right justify
            
            # Also update row 10 column B to be right aligned (Weekly Change)
            ws.cell(row=10, column=2).alignment = right_align
            
            print(f"    Updated YTD: {ytd_text}")
            
        except Exception as e:
            print(f"    Error updating YTD gain/loss: {e}")
            import traceback
            traceback.print_exc()

    def update_dividend_performance_metric(self, ws, section_positions=None):
        """Clean up any old dividend performance data since summary is now at the top (row 5)"""
        print("  Cleaning up old dividend performance data (summary now at top)...")
        
        try:
            # Clear any existing dividend performance data from the old dynamic position area
            for clear_row in range(25, 40):
                cell_j = ws.cell(row=clear_row, column=10)
                cell_k = ws.cell(row=clear_row, column=11)
                if cell_j.value and "Dividend Performance" in str(cell_j.value):
                    cell_j.value = None
                    cell_k.value = None
                    print(f"    Cleared old dividend performance data from row {clear_row}")
            
            print("    Dividend performance summary is now integrated at row 5 (top of cuts/increases section)")
            
        except Exception as e:
            print(f"    Error cleaning up old dividend performance data: {e}")
            import traceback
            traceback.print_exc()
    
    def load_cache_data(self):
        """Load cache data"""
        try:
            with open(self.cache_file, 'r') as f:
                return json.load(f)
        except Exception as e:
            print(f"ERROR loading cache: {e}")
            return None

if __name__ == "__main__":
    updater = PortfolioSummaryUpdater()
    success = updater.run_update()
    
    if success:
        print("\nSUCCESS: Portfolio Summary sheet updated!")
    else:
        print("\nERROR: Portfolio Summary update failed!")