import openpyxl

# Load the restored workbook
wb = openpyxl.load_workbook('outputs/Dividends_2025.xlsx')

print("RESTORED FILE VERIFICATION")
print("=" * 30)

print(f"\nSheets in workbook: {wb.sheetnames}")

# Check each sheet
for sheet_name in wb.sheetnames:
    ws = wb[sheet_name]
    print(f"\n{sheet_name}:")
    print(f"  Dimensions: {ws.max_row} rows, {ws.max_column} columns")
    
    # Sample first few cells to verify data
    if ws.max_row > 0:
        print(f"  First row data: {[ws.cell(1, j).value for j in range(1, min(6, ws.max_column+1))]}")

wb.close()
print("\n✅ File restoration verification complete")