#!/usr/bin/env python3
"""
Diagnostic tool to examine current sheet structure after previous updates
"""

import openpyxl
import os

def diagnose_current_sheet_structure():
    """Examine the current structure of the historical yield sheet"""
    
    excel_file = os.path.join(os.path.dirname(__file__), "outputs", "Dividends_2025.xlsx")
    
    if not os.path.exists(excel_file):
        print(f"ERROR: Excel file not found: {excel_file}")
        return
    
    try:
        workbook = openpyxl.load_workbook(excel_file)
        
        # Find the historical yield sheet
        historical_sheet = None
        for sheet_name in workbook.sheetnames:
            if "historical yield" in sheet_name.lower():
                historical_sheet = workbook[sheet_name]
                print(f"SUCCESS: Found sheet: {sheet_name}")
                break
        
        if not historical_sheet:
            print("ERROR: Could not find historical yield sheet")
            return
        
        print("\nCURRENT SHEET STRUCTURE ANALYSIS")
        print("="*50)
        
        # Check all rows with data in column A
        print("\nALL ROWS WITH DATA IN COLUMN A:")
        for row in range(1, min(80, historical_sheet.max_row + 1)):
            cell_value = historical_sheet.cell(row=row, column=1).value
            if cell_value:
                cell_text = str(cell_value).strip()
                
                # Check if this looks like a group header
                is_group_header = any(keyword in cell_text.upper() for keyword in ['ETRADE', 'SCHWAB', 'TAXABLE', 'INDIVIDUAL', 'IRA'])
                
                if is_group_header:
                    print(f"  Row {row:2d}: {cell_text} <-- POTENTIAL GROUP HEADER")
                else:
                    print(f"  Row {row:2d}: {cell_text}")
                    
                # Stop if we've gone too far
                if row > 75:
                    print("  ... (stopping at row 75)")
                    break
        
        print(f"\nSHEET INFO:")
        print(f"  - Total rows: {historical_sheet.max_row}")
        print(f"  - Total columns: {historical_sheet.max_column}")
        
        workbook.close()
        
    except Exception as e:
        print(f"ERROR: Error examining sheet: {e}")
        import traceback
        traceback.print_exc()

if __name__ == "__main__":
    diagnose_current_sheet_structure()