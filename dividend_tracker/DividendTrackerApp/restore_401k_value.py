#!/usr/bin/env python3
"""
Restore 401K Account Value in Row 8 Column B
Row 8 should show the actual dollar value ($124,315.15), not the percentage
Row 18 should show the percentage (23.9%)
"""
import openpyxl
from openpyxl.styles import Font, Alignment

def restore_401k_value():
    """Restore the actual 401K dollar value in row 8"""
    
    excel_file = 'outputs/Dividends_2025.xlsx'
    
    try:
        print("🔧 Restoring 401K account value in row 8...")
        
        wb = openpyxl.load_workbook(excel_file)
        portfolio_ws = wb['Portfolio Summary']
        
        # 401K account details
        k401_value = 124315.15  # Actual dollar value
        k401_pct = k401_value / 519439.06  # Percentage for row 18
        
        # Define formatting
        normal_font = Font(name='Arial', size=12)
        right_aligned = Alignment(horizontal='right')
        currency_format = '$#,##0.00'
        percentage_format = '0.0%'
        
        print(f"📊 Checking current values in rows 8 and 18...")
        
        # Check and fix row 8 (should have dollar value)
        row8_a = portfolio_ws.cell(row=8, column=1).value or ""
        row8_b = portfolio_ws.cell(row=8, column=2).value or ""
        print(f"   Row 8: '{row8_a}' | '{row8_b}'")
        
        if "401k" in str(row8_a).lower() and "retirement" in str(row8_a).lower():
            # This should be the dollar value row
            value_cell = portfolio_ws.cell(row=8, column=2)
            value_cell.value = k401_value
            value_cell.number_format = currency_format
            value_cell.font = normal_font
            value_cell.alignment = right_aligned
            print(f"   ✅ Restored 401K account value in row 8: ${k401_value:,.2f}")
        
        # Check row 18 (should have percentage)
        row18_a = portfolio_ws.cell(row=18, column=1).value or ""
        row18_b = portfolio_ws.cell(row=18, column=2).value or ""
        print(f"   Row 18: '{row18_a}' | '{row18_b}'")
        
        # Verify row 18 still has the percentage (should be good from previous fix)
        if isinstance(row18_b, (int, float)) and row18_b < 1:
            print(f"   ✅ Row 18 percentage confirmed: {row18_b:.1%}")
        else:
            print(f"   ⚠️ Row 18 may need percentage fix: '{row18_b}'")
        
        # Save the workbook
        wb.save(excel_file)
        wb.close()
        
        return True
        
    except Exception as e:
        print(f"❌ Error restoring 401K value: {e}")
        return False

def verify_401k_values():
    """Verify both 401K rows are correct"""
    excel_file = 'outputs/Dividends_2025.xlsx'
    
    try:
        wb = openpyxl.load_workbook(excel_file, data_only=True)
        portfolio_ws = wb['Portfolio Summary']
        
        print("\n📊 VERIFICATION - 401K Values:")
        print("=" * 60)
        
        # Check the Current Values section (around row 8)
        print("CURRENT VALUES Section:")
        for row in range(2, 12):
            col_a = portfolio_ws.cell(row=row, column=1).value or ""
            col_b = portfolio_ws.cell(row=row, column=2).value or ""
            
            if col_a and "401k" in str(col_a).lower():
                if isinstance(col_b, (int, float)) and col_b > 1000:
                    print(f"   Row {row}: {col_a:<25} ${col_b:,.2f} ✅")
                else:
                    print(f"   Row {row}: {col_a:<25} {col_b} ❌")
        
        print("\nACCOUNT BREAKDOWN Section:")
        # Check the Account Breakdown section (around row 18)
        for row in range(13, 20):
            col_a = portfolio_ws.cell(row=row, column=1).value or ""
            col_b = portfolio_ws.cell(row=row, column=2).value or ""
            
            if col_a and "401k" in str(col_a).lower():
                if isinstance(col_b, (int, float)) and col_b < 1 and col_b > 0:
                    print(f"   Row {row}: {col_a:<25} {col_b:.1%} ✅")
                else:
                    print(f"   Row {row}: {col_a:<25} {col_b} ❌")
        
        wb.close()
        
        print("\n🎯 Expected Values:")
        print("   Row 8 (Current Values): 401k Retirement: $124,315.15")
        print("   Row 18 (Account Breakdown): 401k Retirement: 23.9%")
        
    except Exception as e:
        print(f"Error: {e}")

if __name__ == "__main__":
    print("🎯 RESTORING 401K ACCOUNT VALUE")
    print("=" * 50)
    print("Issue: Row 8 shows percentage instead of dollar value")
    print("Fix: Restore $124,315.15 in row 8, keep 23.9% in row 18")
    print("=" * 50)
    
    success = restore_401k_value()
    
    if success:
        verify_401k_values()
        print("\n✅ 401K account value restoration completed!")
        print("   Row 8: Shows actual dollar amount")
        print("   Row 18: Shows percentage of portfolio")
    else:
        print("\n❌ Could not restore 401K account value")
