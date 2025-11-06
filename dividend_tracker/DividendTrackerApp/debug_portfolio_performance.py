#!/usr/bin/env python3
"""
Debug Portfolio Performance Calculation
"""

import openpyxl
import os

def debug_portfolio_performance():
    """Debug the portfolio performance calculation"""
    
    excel_file = "outputs/Dividends_2025.xlsx"
    
    try:
        wb = openpyxl.load_workbook(excel_file)
        
        if 'Portfolio Values 2025' not in wb.sheetnames:
            print("❌ Portfolio Values 2025 sheet not found")
            return
        
        portfolio_ws = wb['Portfolio Values 2025']
        max_col = portfolio_ws.max_column
        
        print(f"🔍 Debugging Portfolio Performance Calculation")
        print(f"Max column: {max_col}")
        print("=" * 50)
        
        # Check the last few columns and row 10 (totals)
        for col_offset in [2, 1, 0]:  # Look at last 3 columns
            col = max_col - col_offset
            if col > 0:
                total_cell = portfolio_ws.cell(row=10, column=col)
                total_value = total_cell.value
                
                print(f"Column {col} (row 10):")
                print(f"  Value: {total_value}")
                print(f"  Type: {type(total_value)}")
                print(f"  Is Number: {isinstance(total_value, (int, float))}")
                
                if isinstance(total_value, str):
                    print(f"  String content: '{total_value}'")
                    # Try to convert
                    try:
                        numeric_val = float(total_value.replace('$', '').replace(',', ''))
                        print(f"  Converted: {numeric_val}")
                    except:
                        print(f"  Cannot convert to number")
                print()
        
        # Check a few other key cells
        print("\nChecking other cells in row 10:")
        for col in range(max(1, max_col-5), max_col+1):
            cell = portfolio_ws.cell(row=10, column=col)
            if cell.value is not None:
                print(f"  Column {col}: {cell.value} (Type: {type(cell.value)})")
        
        wb.close()
        
    except Exception as e:
        print(f"Error: {e}")
        import traceback
        traceback.print_exc()

if __name__ == "__main__":
    debug_portfolio_performance()