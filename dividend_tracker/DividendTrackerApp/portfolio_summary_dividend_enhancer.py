"""
Portfolio Summary Dividend Status Enhancement
============================================

Adds dividend analysis columns to Portfolio Summary sheet starting at G1:
1. Identifies dividends that were reduced for each account
2. Provides top dividend payers % for each account
3. Pulls data from "Accounts Div historical yield" sheet
"""

import pandas as pd
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
import os
from datetime import datetime

class PortfolioSummaryDividendEnhancer:
    """Enhances Portfolio Summary with dividend analysis from historical yield data"""
    
    def __init__(self):
        self.workbook_path = os.path.join(os.path.dirname(__file__), "outputs", "Dividends_2025.xlsx")
        self.portfolio_sheet_name = "Portfolio Summary"
        self.dividend_sheet_name = "Accounts Div historical yield"
        
    def analyze_dividend_changes(self, workbook):
        """Analyze dividend changes from the historical yield sheet"""
        try:
            dividend_sheet = workbook[self.dividend_sheet_name]
            
            # Get the latest two columns of yield data for comparison
            max_col = dividend_sheet.max_column
            if max_col < 17:  # Need at least Beginning + 1 historical column
                return {"reduced": [], "top_payers": []}
            
            latest_col = max_col
            beginning_col = 15  # "Beginning Dividend Yield" column O
            
            dividend_analysis = {
                "reduced": [],
                "top_payers": []
            }
            
            # Account sections row ranges (using current shifted layout)
            account_sections = {
                "Etrade IRA": {"start": 3, "end": 20, "avg_row": 21},
                "Etrade Taxable": {"start": 25, "end": 36, "avg_row": 37}, 
                "Schwab IRA": {"start": 41, "end": 44, "avg_row": 45},
                "Schwab Individual": {"start": 49, "end": 50, "avg_row": 51}
            }
            
            for account, ranges in account_sections.items():
                print(f"Analyzing {account} dividend changes...")
                
                account_reduced = []
                account_yields = []
                
                # Check each ticker in this account
                for row in range(ranges["start"], ranges["end"] + 1):
                    ticker_cell = dividend_sheet.cell(row=row, column=1)
                    if not ticker_cell.value or str(ticker_cell.value).strip() == "":
                        continue
                        
                    ticker = str(ticker_cell.value).strip()
                    
                    # Get current and beginning yield values
                    current_yield = dividend_sheet.cell(row=row, column=latest_col).value
                    beginning_yield = dividend_sheet.cell(row=row, column=beginning_col).value
                    
                    try:
                        current_val = float(current_yield) if current_yield else 0
                        beginning_val = float(beginning_yield) if beginning_yield else 0
                        
                        # Current yield is in decimal format (0.1015 = 10.15%), convert to percentage
                        if current_val < 1.0 and current_val > 0:
                            current_val *= 100
                        
                        # Beginning yield is already in percentage format (14.2 = 14.2%)
                        # No conversion needed
                        
                        # Check for significant dividend reduction (>10% reduction threshold)
                        if beginning_val > 0 and current_val < (beginning_val * 0.90):  # 10% reduction threshold
                            reduction_pct = ((beginning_val - current_val) / beginning_val) * 100
                            account_reduced.append({
                                "ticker": ticker,
                                "beginning": beginning_val,
                                "current": current_val,
                                "reduction_pct": reduction_pct
                            })
                        
                        # Track yields for top payers analysis (use current percentage value)
                        if current_val > 0:
                            account_yields.append({
                                "ticker": ticker,
                                "yield": current_val
                            })
                            
                    except (ValueError, TypeError):
                        continue
                
                # Sort and get top dividend payers for this account
                account_yields.sort(key=lambda x: x["yield"], reverse=True)
                top_3_payers = account_yields[:3]
                
                dividend_analysis["reduced"].append({
                    "account": account,
                    "reductions": account_reduced
                })
                
                dividend_analysis["top_payers"].append({
                    "account": account,
                    "top_yields": top_3_payers
                })
                
                print(f"  {account}: {len(account_reduced)} dividend reductions found")
                print(f"  {account}: Top yield - {top_3_payers[0]['ticker']} at {top_3_payers[0]['yield']:.2f}%" if top_3_payers else "  No yield data")
                
            return dividend_analysis
            
        except Exception as e:
            print(f"❌ Error analyzing dividend changes: {e}")
            return {"reduced": [], "top_payers": []}
    
    def add_dividend_status_columns(self):
        """Add Dividend Status analysis columns starting at G1"""
        try:
            print("Adding Dividend Status columns to Portfolio Summary...")
            
            # Load workbook
            workbook = openpyxl.load_workbook(self.workbook_path)
            
            if self.portfolio_sheet_name not in workbook.sheetnames:
                print(f"❌ Sheet '{self.portfolio_sheet_name}' not found")
                return False
            
            portfolio_sheet = workbook[self.portfolio_sheet_name]
            
            # Analyze dividend changes
            dividend_analysis = self.analyze_dividend_changes(workbook)
            
            # Header styling
            header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
            header_font = Font(bold=True, color="FFFFFF", size=12)
            section_fill = PatternFill(start_color="B7DEE8", end_color="B7DEE8", fill_type="solid")
            section_font = Font(bold=True, size=11)
            
            # Add main header at G1
            portfolio_sheet.cell(row=1, column=7, value="DIVIDEND STATUS").fill = header_fill
            portfolio_sheet.cell(row=1, column=7).font = header_font
            portfolio_sheet.cell(row=1, column=7).alignment = Alignment(horizontal='center')
            
            # Merge header across columns G-H
            portfolio_sheet.merge_cells('G1:H1')
            
            current_row = 3
            
            # === DIVIDEND REDUCTIONS SECTION ===
            portfolio_sheet.cell(row=current_row, column=7, value="DIVIDEND REDUCTIONS").fill = section_fill
            portfolio_sheet.cell(row=current_row, column=7).font = section_font
            portfolio_sheet.merge_cells(f'G{current_row}:H{current_row}')
            current_row += 1
            
            # Add column headers for reductions
            portfolio_sheet.cell(row=current_row, column=7, value="Account").font = Font(bold=True)
            portfolio_sheet.cell(row=current_row, column=8, value="Reduced Tickers").font = Font(bold=True)
            current_row += 1
            
            for account_data in dividend_analysis["reduced"]:
                account = account_data["account"]
                reductions = account_data["reductions"]
                
                if reductions:
                    # Sort by highest reduction percentage
                    reductions.sort(key=lambda x: x["reduction_pct"], reverse=True)
                    
                    # Show account and worst reduction
                    worst_reduction = reductions[0]
                    reduction_text = f"{worst_reduction['ticker']}: -{worst_reduction['reduction_pct']:.1f}%"
                    
                    if len(reductions) > 1:
                        reduction_text += f" (+{len(reductions)-1} more)"
                    
                    portfolio_sheet.cell(row=current_row, column=7, value=account)
                    portfolio_sheet.cell(row=current_row, column=8, value=reduction_text)
                    
                    # Red text for reductions
                    portfolio_sheet.cell(row=current_row, column=8).font = Font(color="DC143C")
                else:
                    portfolio_sheet.cell(row=current_row, column=7, value=account)
                    portfolio_sheet.cell(row=current_row, column=8, value="No reductions - Good!")
                    portfolio_sheet.cell(row=current_row, column=8).font = Font(color="228B22")
                
                current_row += 1
            
            current_row += 1  # Space between sections
            
            # === TOP DIVIDEND PAYERS SECTION ===
            portfolio_sheet.cell(row=current_row, column=7, value="TOP DIVIDEND PAYERS").fill = section_fill
            portfolio_sheet.cell(row=current_row, column=7).font = section_font
            portfolio_sheet.merge_cells(f'G{current_row}:H{current_row}')
            current_row += 1
            
            # Add column headers for top payers
            portfolio_sheet.cell(row=current_row, column=7, value="Account").font = Font(bold=True)
            portfolio_sheet.cell(row=current_row, column=8, value="Top Yield").font = Font(bold=True)
            current_row += 1
            
            for account_data in dividend_analysis["top_payers"]:
                account = account_data["account"]
                top_yields = account_data["top_yields"]
                
                if top_yields:
                    # Show top yielder
                    top_yielder = top_yields[0]
                    yield_text = f"{top_yielder['ticker']}: {top_yielder['yield']:.2f}%"
                    
                    if len(top_yields) > 1:
                        yield_text += f" (#{len(top_yields)} total)"
                    
                    portfolio_sheet.cell(row=current_row, column=7, value=account)
                    portfolio_sheet.cell(row=current_row, column=8, value=yield_text)
                    
                    # Green text for high yields
                    portfolio_sheet.cell(row=current_row, column=8).font = Font(color="228B22", bold=True)
                else:
                    portfolio_sheet.cell(row=current_row, column=7, value=account)
                    portfolio_sheet.cell(row=current_row, column=8, value="No yield data")
                    portfolio_sheet.cell(row=current_row, column=8).font = Font(color="808080")
                
                current_row += 1
            
            current_row += 1  # Space between sections
            
            # === DIVIDEND SUMMARY METRICS ===
            portfolio_sheet.cell(row=current_row, column=7, value="DIVIDEND METRICS").fill = section_fill
            portfolio_sheet.cell(row=current_row, column=7).font = section_font
            portfolio_sheet.merge_cells(f'G{current_row}:H{current_row}')
            current_row += 1
            
            # Calculate overall metrics
            total_reductions = sum(len(acc["reductions"]) for acc in dividend_analysis["reduced"])
            total_positions = sum(len(acc["top_yields"]) for acc in dividend_analysis["top_payers"])
            
            # Add metrics
            metrics = [
                ("Total Positions:", total_positions),
                ("Dividend Cuts:", total_reductions),
                ("Cut Rate:", f"{(total_reductions/total_positions*100):.1f}%" if total_positions > 0 else "0%"),
                ("Last Updated:", datetime.now().strftime("%m/%d %H:%M"))
            ]
            
            for metric_name, metric_value in metrics:
                portfolio_sheet.cell(row=current_row, column=7, value=metric_name).font = Font(bold=True)
                portfolio_sheet.cell(row=current_row, column=8, value=str(metric_value))
                current_row += 1
            
            # Auto-size columns
            portfolio_sheet.column_dimensions['G'].width = 20
            portfolio_sheet.column_dimensions['H'].width = 25
            
            # Save workbook
            workbook.save(self.workbook_path)
            workbook.close()
            
            print("Dividend Status columns added successfully!")
            print(f"   Found {total_reductions} dividend reductions across all accounts")
            print(f"   Analyzed {total_positions} dividend-paying positions")
            print(f"   Overall dividend cut rate: {(total_reductions/total_positions*100):.1f}%" if total_positions > 0 else "   No positions analyzed")
            
            return True
            
        except Exception as e:
            print(f"❌ Error adding dividend status columns: {e}")
            import traceback
            traceback.print_exc()
            return False

def main():
    """Test the dividend status enhancement"""
    enhancer = PortfolioSummaryDividendEnhancer()
    
    print("Starting Portfolio Summary Dividend Status Enhancement...")
    
    if not os.path.exists(enhancer.workbook_path):
        print(f"❌ Workbook not found: {enhancer.workbook_path}")
        return
    
    success = enhancer.add_dividend_status_columns()
    
    if success:
        print("Portfolio Summary enhancement completed successfully!")
        print("Check columns G-H in the Portfolio Summary sheet for dividend analysis")
    else:
        print("❌ Enhancement failed - check error messages above")

if __name__ == "__main__":
    main()
