#!/usr/bin/env python3
"""
Move Dividend Statistics to Top - Simplified Approach
Move dividend stats to start at row 1, push existing content down
"""
import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment

def move_dividend_stats_to_top_simple():
    """Move dividend statistics to start at row 1 using insert approach"""
    
    excel_file = 'outputs/Dividends_2025.xlsx'
    
    try:
        print("🔄 Moving dividend statistics to top of Portfolio Summary...")
        
        wb = openpyxl.load_workbook(excel_file)
        portfolio_ws = wb['Portfolio Summary']
        
        # Get latest dividend data from Estimated Income sheet
        income_ws = wb['Estimated Income 2025']
        max_col = income_ws.max_column
        latest_monthly_estimate = income_ws.cell(row=9, column=max_col).value or 0
        
        # Calculate statistics
        latest_annual_estimate = latest_monthly_estimate * 12
        weekly_estimate = latest_monthly_estimate / 4.33
        
        etrade_ira_annual = income_ws.cell(row=4, column=max_col).value or 0
        etrade_taxable_annual = income_ws.cell(row=5, column=max_col).value or 0
        schwab_ira_annual = income_ws.cell(row=6, column=max_col).value or 0
        schwab_individual_annual = income_ws.cell(row=7, column=max_col).value or 0
        
        # Withdrawal data
        monthly_etrade_ira = 1200
        quarterly_etrade_ira = 370
        monthly_etrade_taxable = 395
        total_monthly_withdrawals = monthly_etrade_ira + monthly_etrade_taxable + (quarterly_etrade_ira / 3)
        net_monthly = latest_monthly_estimate - total_monthly_withdrawals
        net_annual = net_monthly * 12
        
        # Define styles
        header_font = Font(name='Arial', size=12, bold=True, color='FFFFFF')
        header_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
        section_font = Font(name='Arial', size=11, bold=True)
        section_fill = PatternFill(start_color='D9E2F3', end_color='D9E2F3', fill_type='solid')
        normal_font = Font(name='Arial', size=12)
        currency_format = '$#,##0.00'
        
        # Calculate how many rows we need for dividend stats
        stats_rows_needed = 35  # Approximately 35 rows for all dividend stats
        
        # Insert rows at the top to make space
        print(f"📊 Inserting {stats_rows_needed} rows at the top...")
        portfolio_ws.insert_rows(1, stats_rows_needed)
        
        # Now create dividend statistics starting at row 1
        print("📈 Creating dividend statistics at row 1...")
        
        current_row = 1
        
        # === DIVIDEND STATISTICS HEADER ===
        cell = portfolio_ws.cell(row=current_row, column=1)
        cell.value = "WEEKLY DIVIDEND STATISTICS"
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center')
        portfolio_ws.merge_cells(f'A{current_row}:B{current_row}')
        
        current_row += 2
        
        # Add all the dividend sections
        dividend_data = [
            ("CURRENT ESTIMATES", [
                ("Weekly Estimate:", weekly_estimate),
                ("Monthly Estimate:", latest_monthly_estimate),
                ("Annual Estimate:", latest_annual_estimate),
            ]),
            ("ACCOUNT BREAKDOWN (Annual)", [
                ("E*TRADE IRA:", etrade_ira_annual),
                ("E*TRADE Taxable:", etrade_taxable_annual),
                ("Schwab IRA:", schwab_ira_annual),
                ("Schwab Individual:", schwab_individual_annual),
            ]),
            ("WITHDRAWAL DATA", [
                ("Monthly E*TRADE IRA:", monthly_etrade_ira),
                ("Quarterly E*TRADE IRA:", quarterly_etrade_ira),
                ("Monthly E*TRADE Taxable:", monthly_etrade_taxable),
                ("Total Monthly Withdrawals:", total_monthly_withdrawals),
            ]),
            ("NET INCOME (After Withdrawals)", [
                ("Net Monthly Income:", net_monthly),
                ("Net Annual Income:", net_annual),
                ("Reinvestment Rate:", f"{(net_annual / latest_annual_estimate * 100):.1f}%" if latest_annual_estimate > 0 else "0%"),
            ]),
            ("DIVIDEND METRICS", [
                ("Current Yield:", f"{(latest_annual_estimate / 519000 * 100):.2f}%"),
                ("Monthly Dividend Coverage:", f"{latest_monthly_estimate / total_monthly_withdrawals:.1f}x" if total_monthly_withdrawals > 0 else "N/A"),
                ("Annual Div Growth Target:", "5-7%"),
                ("Next Update:", "Weekly (Automated)"),
            ])
        ]
        
        for section_title, items in dividend_data:
            # Section header
            cell = portfolio_ws.cell(row=current_row, column=1)
            cell.value = section_title
            cell.font = section_font
            cell.fill = section_fill
            current_row += 1
            
            # Section items
            for label, value in items:
                # Label
                label_cell = portfolio_ws.cell(row=current_row, column=1)
                label_cell.value = label
                label_cell.font = normal_font
                
                # Value
                value_cell = portfolio_ws.cell(row=current_row, column=2)
                
                if isinstance(value, (int, float)):
                    value_cell.value = value
                    value_cell.number_format = currency_format
                else:
                    value_cell.value = value
                
                value_cell.font = normal_font
                current_row += 1
            
            current_row += 1  # Extra spacing between sections
        
        # Add separator before original portfolio data
        separator_row = current_row + 1
        cell = portfolio_ws.cell(row=separator_row, column=1)
        cell.value = "═══════════════════════ PORTFOLIO SUMMARY ═══════════════════════"
        cell.font = Font(name='Arial', size=11, bold=True)
        cell.fill = PatternFill(start_color='F2F2F2', end_color='F2F2F2', fill_type='solid')
        portfolio_ws.merge_cells(f'A{separator_row}:B{separator_row}')
        
        print(f"✅ Dividend statistics successfully moved to top!")
        print(f"   Dividend stats: rows 1-{current_row}")
        print(f"   Original portfolio data: rows {stats_rows_needed + 1}+")
        
        # Save the workbook
        wb.save(excel_file)
        wb.close()
        
        return True
        
    except Exception as e:
        print(f"❌ Error moving dividend statistics: {e}")
        return False

def verify_new_layout():
    """Verify the new layout"""
    excel_file = 'outputs/Dividends_2025.xlsx'
    
    try:
        wb = openpyxl.load_workbook(excel_file, data_only=True)
        portfolio_ws = wb['Portfolio Summary']
        
        print("\n📊 NEW LAYOUT VERIFICATION")
        print("=" * 50)
        
        # Show the new structure
        for row in range(1, min(35, portfolio_ws.max_row + 1)):
            col_a_value = portfolio_ws.cell(row=row, column=1).value
            col_b_value = portfolio_ws.cell(row=row, column=2).value
            
            if col_a_value and str(col_a_value).strip():
                if col_b_value:
                    print(f"Row {row:2d}: {col_a_value} | {col_b_value}")
                else:
                    print(f"Row {row:2d}: {col_a_value}")
        
        wb.close()
        
    except Exception as e:
        print(f"Error: {e}")

if __name__ == "__main__":
    print("🚀 MOVING DIVIDEND STATISTICS TO TOP (SIMPLIFIED)")
    print("=" * 60)
    
    success = move_dividend_stats_to_top_simple()
    
    if success:
        verify_new_layout()
        
        print("\n🎯 Summary:")
        print("✅ Dividend statistics moved to start at row 1")
        print("✅ Original portfolio data pushed down below")
        print("✅ Professional formatting applied")
        print("✅ Clean separator between sections")
    else:
        print("\n❌ Failed to move dividend statistics")
