#!/usr/bin/env python3
"""
Check what data was created in the Ticker Analysis 2025 sheet
"""

import openpyxl
import pandas as pd

# Check Excel file
print("=== Checking Ticker Analysis 2025 Sheet ===")
wb = openpyxl.load_workbook('outputs/Dividends_2025.xlsx')
ws = wb['Ticker Analysis 2025']

print(f"Rows in sheet: {ws.max_row}")
print(f"Columns in sheet: {ws.max_column}")

print("\nFirst 20 rows:")
for row in range(1, min(21, ws.max_row + 1)):
    ticker = ws.cell(row=row, column=1).value
    account = ws.cell(row=row, column=2).value
    qty = ws.cell(row=row, column=3).value
    print(f"Row {row}: {ticker} | {account} | {qty}")

print("\n=== Checking Original Data File ===")
df = pd.read_excel('data/dividend_stocks.xlsx')
print("Columns:", df.columns.tolist())
print("Accounts:", df['Account'].value_counts().to_dict())
print(f"Total rows: {len(df)}")

print("\nSample data:")
print(df[['Ticker', 'Account', 'Qty #']].head(10))
