#!/usr/bin/env python3
"""
Portfolio Summary Options
Create either an enhanced Excel summary sheet or a simple GUI
"""

import pandas as pd
import numpy as np
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment, NamedStyle
from datetime import datetime, timedelta
import tkinter as tk
from tkinter import ttk, messagebox
import os

def create_excel_summary_sheet():
    """Create a comprehensive summary sheet in Excel"""
    
    excel_file = 'outputs/Dividends_2025.xlsx'
    
    try:
        # Load the workbook
        wb = load_workbook(excel_file)
        
        # Create or get the Portfolio Summary sheet
        if 'Portfolio Summary' in wb.sheetnames:
            ws = wb['Portfolio Summary']
            wb.remove(ws)  # Remove existing to recreate
        
        ws = wb.create_sheet('Portfolio Summary', 0)  # Insert as first sheet
        
        print("📊 Creating Portfolio Summary sheet...")
        
        # Get data from Portfolio Values 2025 sheet
        portfolio_ws = wb['Portfolio Values 2025']
        
        # Get the latest column (should be the one we just added)
        max_col = portfolio_ws.max_column
        latest_date = portfolio_ws.cell(row=3, column=max_col).value
        
        # Get current values
        current_values = {
            'ETrade_IRA': portfolio_ws.cell(row=4, column=max_col).value,
            'ETrade_Taxable': portfolio_ws.cell(row=5, column=max_col).value,
            'Schwab_IRA': portfolio_ws.cell(row=6, column=max_col).value,
            'Schwab_Individual': portfolio_ws.cell(row=7, column=max_col).value,
            'Retirement_401k': portfolio_ws.cell(row=8, column=max_col).value,
            'Total': portfolio_ws.cell(row=10, column=max_col).value
        }
        
        # Get previous week values for comparison
        prev_col = max_col - 1 if max_col > 1 else max_col
        prev_date = portfolio_ws.cell(row=3, column=prev_col).value
        prev_total = portfolio_ws.cell(row=10, column=prev_col).value
        
        # Calculate weekly change
        weekly_change = current_values['Total'] - prev_total if prev_total else 0
        weekly_change_pct = (weekly_change / prev_total * 100) if prev_total else 0
        
        # HEADER SECTION
        ws['A1'] = "DIVIDEND PORTFOLIO SUMMARY"
        ws['A2'] = f"Generated: {datetime.now().strftime('%B %d, %Y at %I:%M %p')}"
        ws['A3'] = f"Latest Data: {latest_date}"
        
        # Format header
        ws['A1'].font = Font(name='Arial', size=16, bold=True, color='FFFFFF')
        ws['A1'].fill = PatternFill(start_color='2E75B6', end_color='2E75B6', fill_type='solid')
        ws['A2'].font = Font(name='Arial', size=10, italic=True)
        ws['A3'].font = Font(name='Arial', size=10, italic=True)
        
        # CURRENT PORTFOLIO VALUES
        row = 5
        ws[f'A{row}'] = "CURRENT PORTFOLIO VALUES"
        ws[f'A{row}'].font = Font(name='Arial', size=14, bold=True, color='FFFFFF')
        ws[f'A{row}'].fill = PatternFill(start_color='70AD47', end_color='70AD47', fill_type='solid')
        
        row += 2
        accounts = [
            ('E*TRADE IRA', current_values['ETrade_IRA']),
            ('E*TRADE Taxable', current_values['ETrade_Taxable']),
            ('Schwab IRA', current_values['Schwab_IRA']),
            ('Schwab Individual', current_values['Schwab_Individual']),
            ('401(k) Retirement', current_values['Retirement_401k'])
        ]
        
        for account, value in accounts:
            ws[f'A{row}'] = account
            ws[f'B{row}'] = value
            ws[f'B{row}'].number_format = '$#,##0.00'
            ws[f'A{row}'].font = Font(name='Arial', size=11)
            ws[f'B{row}'].font = Font(name='Arial', size=11, bold=True)
            row += 1
        
        # Total row
        row += 1
        ws[f'A{row}'] = "TOTAL PORTFOLIO VALUE"
        ws[f'B{row}'] = current_values['Total']
        ws[f'A{row}'].font = Font(name='Arial', size=12, bold=True, color='FFFFFF')
        ws[f'B{row}'].font = Font(name='Arial', size=12, bold=True, color='FFFFFF')
        ws[f'A{row}'].fill = PatternFill(start_color='2E75B6', end_color='2E75B6', fill_type='solid')
        ws[f'B{row}'].fill = PatternFill(start_color='2E75B6', end_color='2E75B6', fill_type='solid')
        ws[f'B{row}'].number_format = '$#,##0.00'
        
        # WEEKLY PERFORMANCE
        row += 3
        ws[f'A{row}'] = "WEEKLY PERFORMANCE"
        ws[f'A{row}'].font = Font(name='Arial', size=14, bold=True, color='FFFFFF')
        ws[f'A{row}'].fill = PatternFill(start_color='C55A11', end_color='C55A11', fill_type='solid')
        
        row += 2
        ws[f'A{row}'] = "Previous Week Total"
        ws[f'B{row}'] = prev_total
        ws[f'B{row}'].number_format = '$#,##0.00'
        
        row += 1
        ws[f'A{row}'] = "Current Week Total"
        ws[f'B{row}'] = current_values['Total']
        ws[f'B{row}'].number_format = '$#,##0.00'
        
        row += 1
        ws[f'A{row}'] = "Weekly Change ($)"
        ws[f'B{row}'] = weekly_change
        ws[f'B{row}'].number_format = '$#,##0.00'
        change_color = '00B050' if weekly_change >= 0 else 'C5504B'
        ws[f'B{row}'].font = Font(color=change_color, bold=True)
        
        row += 1
        ws[f'A{row}'] = "Weekly Change (%)"
        ws[f'B{row}'] = weekly_change_pct / 100
        ws[f'B{row}'].number_format = '0.00%'
        ws[f'B{row}'].font = Font(color=change_color, bold=True)
        
        # ACCOUNT ALLOCATION
        row += 3
        ws[f'A{row}'] = "ACCOUNT ALLOCATION"
        ws[f'A{row}'].font = Font(name='Arial', size=14, bold=True, color='FFFFFF')
        ws[f'A{row}'].fill = PatternFill(start_color='7030A0', end_color='7030A0', fill_type='solid')
        
        row += 2
        total_value = current_values['Total']
        for account, value in accounts:
            percentage = (value / total_value * 100) if total_value > 0 else 0
            ws[f'A{row}'] = account
            ws[f'B{row}'] = value
            ws[f'C{row}'] = percentage / 100
            ws[f'B{row}'].number_format = '$#,##0.00'
            ws[f'C{row}'].number_format = '0.0%'
            row += 1
        
        # DIVIDEND SUMMARY (if dividend data exists)
        try:
            dividend_df = pd.read_excel(excel_file, sheet_name='All account weekly dividends')
            if not dividend_df.empty:
                row += 2
                ws[f'A{row}'] = "LATEST DIVIDEND ACTIVITY"
                ws[f'A{row}'].font = Font(name='Arial', size=14, bold=True, color='FFFFFF')
                ws[f'A{row}'].fill = PatternFill(start_color='D45087', end_color='D45087', fill_type='solid')
                
                # Get recent dividend info (this would need to be enhanced based on your dividend data structure)
                row += 2
                ws[f'A{row}'] = "Monthly Est. Dividends"
                ws[f'B{row}'] = "Data Available in Dividend Sheets"
                ws[f'A{row}'].font = Font(italic=True)
        except:
            pass
        
        # Format column widths
        ws.column_dimensions['A'].width = 25
        ws.column_dimensions['B'].width = 18
        ws.column_dimensions['C'].width = 12
        
        # Add borders to all used cells
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'), 
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        for row_num in range(1, row + 1):
            for col_num in range(1, 4):
                ws.cell(row=row_num, column=col_num).border = thin_border
        
        # Save the workbook
        wb.save(excel_file)
        wb.close()
        
        print("✅ Portfolio Summary sheet created successfully!")
        print(f"📊 Total Portfolio: ${current_values['Total']:,.2f}")
        print(f"📈 Weekly Change: ${weekly_change:,.2f} ({weekly_change_pct:+.2f}%)")
        
        return True
        
    except Exception as e:
        print(f"❌ Error creating summary sheet: {e}")
        import traceback
        traceback.print_exc()
        return False

def create_simple_gui():
    """Create a simple GUI for portfolio summary"""
    
    class PortfolioSummaryGUI:
        def __init__(self, root):
            self.root = root
            self.root.title("Dividend Portfolio Summary")
            self.root.geometry("600x500")
            self.root.configure(bg='#f0f0f0')
            
            # Load data
            self.load_portfolio_data()
            self.create_widgets()
            
        def load_portfolio_data(self):
            """Load portfolio data from Excel"""
            try:
                excel_file = 'outputs/Dividends_2025.xlsx'
                wb = load_workbook(excel_file)
                portfolio_ws = wb['Portfolio Values 2025']
                
                # Get latest data
                max_col = portfolio_ws.max_column
                self.latest_date = portfolio_ws.cell(row=3, column=max_col).value
                
                self.portfolio_data = {
                    'ETrade IRA': portfolio_ws.cell(row=4, column=max_col).value,
                    'ETrade Taxable': portfolio_ws.cell(row=5, column=max_col).value,
                    'Schwab IRA': portfolio_ws.cell(row=6, column=max_col).value,
                    'Schwab Individual': portfolio_ws.cell(row=7, column=max_col).value,
                    '401(k) Retirement': portfolio_ws.cell(row=8, column=max_col).value,
                    'Total': portfolio_ws.cell(row=10, column=max_col).value
                }
                
                # Calculate weekly change
                prev_col = max_col - 1 if max_col > 1 else max_col
                prev_total = portfolio_ws.cell(row=10, column=prev_col).value
                self.weekly_change = self.portfolio_data['Total'] - prev_total if prev_total else 0
                self.weekly_change_pct = (self.weekly_change / prev_total * 100) if prev_total else 0
                
                wb.close()
                
            except Exception as e:
                messagebox.showerror("Error", f"Failed to load portfolio data: {e}")
                self.portfolio_data = {}
        
        def create_widgets(self):
            """Create GUI widgets"""
            
            # Title
            title_frame = tk.Frame(self.root, bg='#2E75B6', height=60)
            title_frame.pack(fill='x')
            title_frame.pack_propagate(False)
            
            title_label = tk.Label(title_frame, text="DIVIDEND PORTFOLIO SUMMARY", 
                                 font=('Arial', 16, 'bold'), fg='white', bg='#2E75B6')
            title_label.pack(expand=True)
            
            date_label = tk.Label(title_frame, text=f"Data as of: {self.latest_date}", 
                                font=('Arial', 10), fg='white', bg='#2E75B6')
            date_label.pack()
            
            # Main content frame
            main_frame = tk.Frame(self.root, bg='#f0f0f0')
            main_frame.pack(fill='both', expand=True, padx=20, pady=20)
            
            # Portfolio values frame
            values_frame = tk.LabelFrame(main_frame, text="Current Portfolio Values", 
                                       font=('Arial', 12, 'bold'), bg='#f0f0f0')
            values_frame.pack(fill='x', pady=10)
            
            for i, (account, value) in enumerate(self.portfolio_data.items()):
                if account != 'Total':
                    frame = tk.Frame(values_frame, bg='#f0f0f0')
                    frame.pack(fill='x', padx=10, pady=2)
                    
                    tk.Label(frame, text=f"{account}:", font=('Arial', 11), 
                           bg='#f0f0f0', width=20, anchor='w').pack(side='left')
                    tk.Label(frame, text=f"${value:,.2f}", font=('Arial', 11, 'bold'), 
                           bg='#f0f0f0', anchor='e').pack(side='right')
            
            # Total frame
            total_frame = tk.Frame(values_frame, bg='#2E75B6', height=40)
            total_frame.pack(fill='x', padx=10, pady=5)
            total_frame.pack_propagate(False)
            
            tk.Label(total_frame, text="TOTAL PORTFOLIO:", font=('Arial', 12, 'bold'), 
                   fg='white', bg='#2E75B6').pack(side='left', padx=10, pady=5)
            tk.Label(total_frame, text=f"${self.portfolio_data['Total']:,.2f}", 
                   font=('Arial', 12, 'bold'), fg='white', bg='#2E75B6').pack(side='right', padx=10, pady=5)
            
            # Performance frame
            perf_frame = tk.LabelFrame(main_frame, text="Weekly Performance", 
                                     font=('Arial', 12, 'bold'), bg='#f0f0f0')
            perf_frame.pack(fill='x', pady=10)
            
            change_color = 'green' if self.weekly_change >= 0 else 'red'
            change_symbol = '+' if self.weekly_change >= 0 else ''
            
            perf_text = tk.Frame(perf_frame, bg='#f0f0f0')
            perf_text.pack(padx=10, pady=10)
            
            tk.Label(perf_text, text=f"Weekly Change: {change_symbol}${self.weekly_change:,.2f}", 
                   font=('Arial', 12, 'bold'), fg=change_color, bg='#f0f0f0').pack()
            tk.Label(perf_text, text=f"({change_symbol}{self.weekly_change_pct:.2f}%)", 
                   font=('Arial', 12, 'bold'), fg=change_color, bg='#f0f0f0').pack()
            
            # Buttons frame
            button_frame = tk.Frame(main_frame, bg='#f0f0f0')
            button_frame.pack(fill='x', pady=20)
            
            tk.Button(button_frame, text="Refresh Data", command=self.refresh_data,
                     font=('Arial', 10), bg='#70AD47', fg='white', width=15).pack(side='left', padx=5)
            
            tk.Button(button_frame, text="View Excel Summary", command=self.open_excel,
                     font=('Arial', 10), bg='#C55A11', fg='white', width=15).pack(side='left', padx=5)
            
            tk.Button(button_frame, text="Close", command=self.root.quit,
                     font=('Arial', 10), bg='#C5504B', fg='white', width=15).pack(side='right', padx=5)
        
        def refresh_data(self):
            """Refresh portfolio data"""
            self.load_portfolio_data()
            # Recreate widgets with new data
            for widget in self.root.winfo_children():
                widget.destroy()
            self.create_widgets()
        
        def open_excel(self):
            """Open Excel file"""
            try:
                os.startfile('outputs/Dividends_2025.xlsx')
            except Exception as e:
                messagebox.showerror("Error", f"Failed to open Excel file: {e}")
    
    # Create and run GUI
    root = tk.Tk()
    app = PortfolioSummaryGUI(root)
    root.mainloop()

if __name__ == "__main__":
    print("📊 Portfolio Summary Options")
    print("=" * 40)
    print("1. Create Enhanced Excel Summary Sheet")
    print("2. Launch Simple Portfolio GUI")
    print("3. Create Both")
    
    choice = input("\nEnter your choice (1, 2, or 3): ").strip()
    
    if choice in ['1', '3']:
        print("\n🏗️  Creating Excel Summary Sheet...")
        create_excel_summary_sheet()
    
    if choice in ['2', '3']:
        print("\n🖥️  Launching Portfolio GUI...")
        create_simple_gui()
    
    if choice not in ['1', '2', '3']:
        print("❌ Invalid choice. Please run again and choose 1, 2, or 3.")
