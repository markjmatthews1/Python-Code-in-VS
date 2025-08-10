#!/usr/bin/env python3
"""
Enhanced Ticker Analysis Creator with Dividend Filter
Creates ticker analysis using both historical data and live API data
Filters for dividend stocks (4%+ yield) only
"""

import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.styles.numbers import FORMAT_CURRENCY_USD_SIMPLE
import pandas as pd
from datetime import datetime
import os
import sys

# Add modules directory to path  
sys.path.append(os.path.join(os.path.dirname(__file__), 'modules'))

# Color definitions
GREEN_FILL = PatternFill(start_color="90EE90", end_color="90EE90", fill_type="solid")
RED_FILL = PatternFill(start_color="FFB6C1", end_color="FFB6C1", fill_type="solid")
YELLOW_FILL = PatternFill(start_color="FFFFE0", end_color="FFFFE0", fill_type="solid")
BLUE_FILL = PatternFill(start_color="ADD8E6", end_color="ADD8E6", fill_type="solid")

# Style definitions
HEADER_FONT = Font(bold=True, name="Arial", size=11, color="000000")
NORMAL_FONT = Font(name="Arial", size=10)
THIN_BORDER = Border(
    left=Side(style='thin'), right=Side(style='thin'),
    top=Side(style='thin'), bottom=Side(style='thin')
)

def load_historical_data():
    """Load historical ticker data from dividend_stocks.xlsx"""
    print("📊 Loading historical dividend data...")
    
    data_file = os.path.join(os.path.dirname(__file__), "data", "dividend_stocks.xlsx")
    if not os.path.exists(data_file):
        print(f"❌ Historical data file not found: {data_file}")
        return {}
    
    try:
        df = pd.read_excel(data_file)
        print(f"✅ Loaded historical data: {df.shape[0]} tickers, {df.shape[1]} columns")
        
        ticker_data = {}
        
        for idx, row in df.iterrows():
            if pd.notna(row.get('Ticker')):
                ticker = str(row['Ticker']).strip()
                
                # Get account data - check Account column first
                account = "Unknown"
                if 'Account' in df.columns and pd.notna(row.get('Account')):
                    account = str(row.get('Account')).strip()
                    print(f"  📊 {ticker}: {account}")
                    
                    # Only include dividend stocks (4%+ yield)
                    beginning_yield = float(row.get('Beginning Dividend Yield', 0))
                    if beginning_yield >= 4.0:
                        # Store ticker data
                        yield_history = {}
                        for col in df.columns:
                            if '2025' in str(col) and col not in ['Ticker', 'Account']:
                                yield_value = row.get(col)
                                if pd.notna(yield_value) and isinstance(yield_value, (int, float)):
                                    yield_history[col] = float(yield_value)
                        
                        ticker_data[ticker] = {
                            'account': account,
                            'quantity': float(row.get('Qty #', 0)),
                            'price_paid': float(row.get('Price Paid $', 0)),
                            'rate_per_share': float(row.get('Rate per share', 0)),
                            'payment_cycle': str(row.get('Payment cycle', 'Monthly')),
                            'beginning_yield': beginning_yield,
                            'yield_history': yield_history
                        }
                        print(f"    ✅ Added dividend stock: {ticker} - {beginning_yield}% yield")
                    else:
                        print(f"    ⏭️ Skipped {ticker} - {beginning_yield}% yield (below 4% threshold)")
        
        print(f"✅ Loaded {len(ticker_data)} dividend stocks from historical data")
        return ticker_data
        
    except Exception as e:
        print(f"❌ Error loading historical data: {e}")
        return {}

def get_api_positions():
    """Get current positions from E*TRADE and Schwab APIs with dividend filter"""
    print("📡 Getting current API positions for dividend stocks...")
    
    api_positions = {}
    
    # Get E*TRADE positions with dividend filter
    try:
        from etrade_account_api import ETRADEAccountAPI
        etrade_api = ETRADEAccountAPI()
        etrade_estimates = etrade_api.get_dividend_estimates()
        
        for account_type, df in etrade_estimates.items():
            if not df.empty and 'Symbol' in df.columns:
                for _, row in df.iterrows():
                    symbol = row.get('Symbol', '').strip()
                    if symbol:
                        # Get market quote to check dividend yield
                        try:
                            quote_url = f"{etrade_api.base_url}/v1/market/quote/{symbol}.json"
                            quote_response = etrade_api.session.get(quote_url)
                            
                            if quote_response.status_code == 200:
                                quote_data = quote_response.json()
                                if ('QuoteResponse' in quote_data and 
                                    'QuoteData' in quote_data['QuoteResponse'] and 
                                    isinstance(quote_data['QuoteResponse']['QuoteData'], list)):
                                    
                                    quote_info = quote_data['QuoteResponse']['QuoteData'][0]
                                    current_yield = quote_info['All'].get('yield', 0)
                                    current_price = quote_info['All'].get('lastTrade', 0)
                                    
                                    # Only include if dividend yield >= 4%
                                    if current_yield and float(current_yield) >= 4.0:
                                        api_positions[symbol] = {
                                            'current_quantity': row.get('Quantity', 0),
                                            'current_price': float(current_price) if current_price else 0,
                                            'current_yield': float(current_yield),
                                            'account': f"E*TRADE {account_type}",
                                            'source': 'E*TRADE'
                                        }
                                        print(f"  📊 Added dividend stock: {symbol} - {current_yield}% yield ({account_type})")
                                    else:
                                        print(f"  ⏭️ Skipped {symbol} - {current_yield}% yield (below 4% threshold)")
                        except Exception as e:
                            print(f"  ⚠️ Error getting quote for {symbol}: {e}")
        
        dividend_positions = len([p for p in api_positions.values() if p['source'] == 'E*TRADE'])
        print(f"✅ Found {dividend_positions} E*TRADE dividend positions (4%+ yield)")
        
    except Exception as e:
        print(f"⚠️ Could not get E*TRADE positions: {e}")
    
    # Get Schwab positions with dividend filter
    try:
        from schwab_api_integrated import SchwabAPI
        schwab_api = SchwabAPI()
        schwab_estimates = schwab_api.get_dividend_estimates()
        
        for account_type, positions in schwab_estimates.items():
            for position in positions:
                symbol = position.get('Symbol', '').strip()
                if symbol:
                    # For Schwab, check if we have dividend data or if it's in historical data
                    dividend_yield = position.get('Yield', 0)
                    if dividend_yield and float(dividend_yield) >= 4.0:
                        api_positions[symbol] = {
                            'current_quantity': position.get('Quantity', 0),
                            'current_price': position.get('Price', 0),
                            'current_yield': float(dividend_yield),
                            'account': f"Schwab {account_type}",
                            'source': 'Schwab'
                        }
                        print(f"  📊 Added dividend stock: {symbol} - {dividend_yield}% yield ({account_type})")
        
        schwab_dividend_positions = len([p for p in api_positions.values() if p['source'] == 'Schwab'])
        print(f"✅ Found {schwab_dividend_positions} Schwab dividend positions (4%+ yield)")
        
    except Exception as e:
        print(f"⚠️ Could not get Schwab positions: {e}")
    
    return api_positions

def create_ticker_analysis_sheet():
    """Create the complete ticker analysis sheet"""
    print("🚀 Creating Dividend Tracker Analysis...")
    print("=" * 60)
    
    # Load historical data
    historical_data = load_historical_data()
    
    # Get current API positions
    api_positions = get_api_positions()
    
    # Combine all tickers (historical + new from APIs)
    all_tickers = set(historical_data.keys()) | set(api_positions.keys())
    
    print(f"📊 Total dividend stocks to track: {len(all_tickers)}")
    print(f"   📚 Historical: {len(historical_data)}")
    print(f"   📡 API: {len(api_positions)}")
    
    # Create the workbook
    output_file = os.path.join(os.path.dirname(__file__), "outputs", "Dividends_2025.xlsx")
    
    try:
        wb = openpyxl.load_workbook(output_file)
    except FileNotFoundError:
        print("📁 Creating new workbook...")
        wb = openpyxl.Workbook()
        wb.remove(wb.active)  # Remove default sheet
    
    # Remove existing sheet if it exists
    if "Ticker Analysis 2025" in wb.sheetnames:
        wb.remove(wb["Ticker Analysis 2025"])
        print("🗑️ Removed existing sheet")
    
    # Create new sheet
    ws = wb.create_sheet(title="Ticker Analysis 2025")
    
    # Set up headers
    ws.cell(row=1, column=1, value="Dividend Tracker Analysis 2025")
    ws.cell(row=1, column=1).font = Font(bold=True, name="Arial", size=14)
    ws.cell(row=2, column=1, value="Historical Data + Live API Integration (4%+ Dividend Yield Filter)")
    
    # Static headers (removed Current Yield % and Yield Change)
    static_headers = [
        'Account', 'Ticker', 'Qty # (Historical)', 'Current Qty (API)', 'Price Paid $', 
        'Current Price $', 'Current Value $', 'Original Value $', 'Total Gain $', 'Total Gain %',
        'Payment Cycle', 'Rate/Share', 'Monthly Dividend', 'Annual Dividend',
        'Beginning Yield %'
    ]
    
    # Get date columns from historical data
    date_columns = set()
    for ticker_info in historical_data.values():
        date_columns.update(ticker_info.get('yield_history', {}).keys())
    
    # Sort dates (newest first)
    def sort_date_key(date_str):
        try:
            if '/' in date_str:
                return datetime.strptime(date_str, '%m/%d/%Y')
            elif '-' in date_str:
                if date_str.count('-') == 2 and len(date_str.split('-')[0]) == 4:
                    return datetime.strptime(date_str, '%Y-%m-%d')
                else:
                    return datetime.strptime(date_str, '%m-%d-%Y')
            return datetime.min
        except:
            return datetime.min
    
    sorted_dates = sorted(date_columns, key=sort_date_key, reverse=True)
    
    # Add today's column first
    today_header = datetime.now().strftime('%m/%d/%Y')
    weekly_headers = [today_header] + sorted_dates
    
    all_headers = static_headers + weekly_headers
    
    # Create headers in row 3
    header_row = 3
    for col, header in enumerate(all_headers, 1):
        cell = ws.cell(row=header_row, column=col)
        cell.value = header
        cell.font = HEADER_FONT
        cell.border = THIN_BORDER
        cell.fill = BLUE_FILL if col <= len(static_headers) else GREEN_FILL
        cell.alignment = Alignment(horizontal="center")
    
    # Sort tickers by account and ticker name
    def account_sort_key(ticker):
        data = historical_data.get(ticker, {})
        api_data = api_positions.get(ticker, {})
        account = data.get('account') or api_data.get('account', 'Unknown')
        
        if 'E*TRADE IRA' in account:
            return (1, ticker)
        elif 'E*TRADE Taxable' in account:
            return (2, ticker)
        elif 'Schwab Individual' in account:
            return (3, ticker)
        elif 'Schwab IRA' in account:
            return (4, ticker)
        else:
            return (5, ticker)
    
    sorted_tickers = sorted(all_tickers, key=account_sort_key)
    
    # Add data rows
    row = 4
    totals = {'current_value': 0, 'original_value': 0, 'monthly_dividend': 0, 'annual_dividend': 0}
    
    for ticker in sorted_tickers:
        hist_data = historical_data.get(ticker, {})
        api_data = api_positions.get(ticker, {})
        
        # Get account (prefer API data if available)
        account = api_data.get('account') or hist_data.get('account', 'Unknown')
        
        # Get quantities
        hist_quantity = hist_data.get('quantity', 0)
        current_quantity = api_data.get('current_quantity', hist_quantity)
        
        # Get prices
        price_paid = hist_data.get('price_paid', 0)
        current_price = api_data.get('current_price', price_paid)
        
        # Calculate values
        current_value = current_quantity * current_price
        original_value = hist_quantity * price_paid
        total_gain = current_value - original_value
        total_gain_pct = (total_gain / original_value * 100) if original_value > 0 else 0
        
        # Get dividend info
        payment_cycle = hist_data.get('payment_cycle', 'Monthly')
        rate_per_share = hist_data.get('rate_per_share', 0)
        beginning_yield = hist_data.get('beginning_yield', 0)
        
        # Calculate dividends
        monthly_dividend = current_quantity * rate_per_share
        if payment_cycle.lower() == 'quarterly':
            monthly_dividend = monthly_dividend / 3
        annual_dividend = monthly_dividend * 12
        
        # Static data
        static_data = [
            account, ticker, hist_quantity, current_quantity, price_paid,
            current_price, current_value, original_value, total_gain, total_gain_pct,
            payment_cycle, rate_per_share, monthly_dividend, annual_dividend,
            beginning_yield
        ]
        
        # Add static data
        for col, value in enumerate(static_data, 1):
            cell = ws.cell(row=row, column=col)
            cell.value = value
            cell.border = THIN_BORDER
            if col in [5, 6, 7, 8, 9, 12, 13, 14]:  # Currency columns
                cell.number_format = FORMAT_CURRENCY_USD_SIMPLE
            elif col == 10:  # Percentage
                cell.number_format = '0.00%'
        
        # Add weekly yield data
        yield_history = hist_data.get('yield_history', {})
        current_yield = api_data.get('current_yield', beginning_yield)
        
        # Today's yield (first weekly column)
        today_col = len(static_headers) + 1
        today_cell = ws.cell(row=row, column=today_col)
        today_cell.value = current_yield
        today_cell.number_format = '0.00'
        today_cell.border = THIN_BORDER
        
        # Color code today's yield
        if len(sorted_dates) > 0:
            last_week_yield = yield_history.get(sorted_dates[0], 0)
            if last_week_yield > 0 and current_yield > 0:
                if current_yield > last_week_yield + 0.1:
                    today_cell.fill = GREEN_FILL
                elif current_yield < last_week_yield - 0.1:
                    today_cell.fill = RED_FILL
                else:
                    today_cell.fill = YELLOW_FILL
        
        # Historical yields
        prev_yield = current_yield
        for date_idx, date_str in enumerate(sorted_dates):
            col = len(static_headers) + 2 + date_idx
            yield_value = yield_history.get(date_str, 0)
            
            cell = ws.cell(row=row, column=col)
            cell.value = yield_value
            cell.number_format = '0.00'
            cell.border = THIN_BORDER
            
            # Color code yield changes
            if prev_yield is not None and yield_value > 0:
                if yield_value > prev_yield + 0.1:
                    cell.fill = GREEN_FILL
                elif yield_value < prev_yield - 0.1:
                    cell.fill = RED_FILL
                elif abs(yield_value - prev_yield) <= 0.1:
                    cell.fill = YELLOW_FILL
            
            if yield_value > 0:
                prev_yield = yield_value
        
        # Update totals
        totals['current_value'] += current_value
        totals['original_value'] += original_value
        totals['monthly_dividend'] += monthly_dividend
        totals['annual_dividend'] += annual_dividend
        
        row += 1
    
    # Add totals row
    totals_row = row
    ws.cell(row=totals_row, column=1, value="TOTALS").font = Font(bold=True)
    
    total_columns = [7, 8, 9, 13, 14]  # Current Value, Original Value, Total Gain, Monthly Dividend, Annual Dividend
    total_values = [totals['current_value'], totals['original_value'], 
                   totals['current_value'] - totals['original_value'],
                   totals['monthly_dividend'], totals['annual_dividend']]
    
    for col, value in zip(total_columns, total_values):
        cell = ws.cell(row=totals_row, column=col)
        cell.value = value
        cell.font = Font(bold=True)
        cell.number_format = FORMAT_CURRENCY_USD_SIMPLE
        cell.border = THIN_BORDER
    
    # Auto-adjust column widths
    for col in range(1, len(all_headers) + 1):
        if col <= len(static_headers):
            ws.column_dimensions[openpyxl.utils.get_column_letter(col)].width = 15
        else:
            ws.column_dimensions[openpyxl.utils.get_column_letter(col)].width = 12
    
    # Add legend
    legend_row = totals_row + 3
    ws.cell(row=legend_row, column=1, value="Weekly Yield Tracking Legend:").font = Font(bold=True)
    ws.cell(row=legend_row + 1, column=1, value="🟢 Yield Increased").fill = GREEN_FILL
    ws.cell(row=legend_row + 2, column=1, value="🔴 Yield Decreased").fill = RED_FILL
    ws.cell(row=legend_row + 3, column=1, value="🟡 Yield Stable").fill = YELLOW_FILL
    
    # Save workbook
    try:
        wb.save(output_file)
        print("✅ Saved enhanced ticker analysis sheet")
    except PermissionError:
        print("❌ Error saving workbook: File may be open in Excel")
    
    print(f"\n📊 Summary:")
    print(f"  📈 Dividend stocks tracked: {len(sorted_tickers)}")
    print(f"  📅 Historical dates: {len(sorted_dates)}")
    print(f"  💰 Total Monthly Dividend: ${totals['monthly_dividend']:.2f}")
    print(f"  💰 Total Annual Dividend: ${totals['annual_dividend']:.2f}")
    print(f"  💰 Total Portfolio Value: ${totals['current_value']:.2f}")
    
    print("✅ SUCCESS: Dividend-focused ticker analysis created!")
    print("📊 Only stocks with 4%+ dividend yield included")

if __name__ == "__main__":
    create_ticker_analysis_sheet()
