#!/usr/bin/env python3
"""
Surgical QQQI Addition to Schwab IRA
====================================

This script will ONLY add QQQI to the Schwab IRA section if it's missing,
without disturbing any other part of the sheet or historical data.
"""

import openpyxl
from openpyxl.styles import Font
import json
import os

class SurgicalQQQIAddition:
    def __init__(self):
        self.script_dir = os.path.dirname(os.path.abspath(__file__))
        self.outputs_dir = os.path.join(self.script_dir, "outputs")
        self.excel_file = os.path.join(self.outputs_dir, "Dividends_2025.xlsx")
        self.cache_file = os.path.join(self.script_dir, "portfolio_data_cache.json")

    def add_qqqi_to_schwab_ira(self):
        """Add QQQI to Schwab IRA section only if missing"""
        print("🎯 SURGICAL QQQI ADDITION TO SCHWAB IRA")
        print("=" * 50)
        
        # Load cache for QQQI position data
        with open(self.cache_file, 'r') as f:
            cache_data = json.load(f)
        
        schwab_ira_positions = cache_data.get('positions', {}).get('schwab_ira', [])
        qqqi_position = None
        
        for pos in schwab_ira_positions:
            if pos.get('symbol', '').upper() == 'QQQI':
                qqqi_position = pos
                break
        
        if not qqqi_position:
            print("❌ QQQI position not found in Schwab IRA cache")
            return False
        
        print(f"✅ Found QQQI in cache: {qqqi_position['quantity']} shares")
        
        # Load workbook
        wb = openpyxl.load_workbook(self.excel_file)
        ws = wb['Accounts Div historical yield']
        
        # Find Schwab IRA section
        schwab_ira_start = None
        schwab_ira_end = None
        
        for row in range(1, ws.max_row + 1):
            cell_value = ws.cell(row=row, column=1).value
            if cell_value:
                cell_text = str(cell_value).strip().upper()
                if 'SCHWAB' in cell_text and 'IRA' in cell_text:
                    schwab_ira_start = row
                    print(f"✅ Found Schwab IRA header at row {row}")
                    break
        
        if not schwab_ira_start:
            print("❌ Schwab IRA section not found")
            return False
        
        # Find the data rows and check if QQQI exists
        data_start = schwab_ira_start + 2  # Skip header and column names
        qqqi_exists = False
        schwab_ira_end = data_start
        
        # Scan the Schwab IRA section
        for row in range(data_start, ws.max_row + 1):
            ticker_cell = ws.cell(row=row, column=1)
            if ticker_cell.value:
                ticker = str(ticker_cell.value).strip().upper()
                if ticker == 'QQQI':
                    qqqi_exists = True
                    print(f"✅ QQQI already exists at row {row}")
                    break
                # Check if we've left the Schwab IRA section
                elif any(keyword in ticker for keyword in ['SCHWAB', 'ETRADE']) and row != schwab_ira_start:
                    schwab_ira_end = row - 1
                    break
                else:
                    schwab_ira_end = row
            else:
                # Empty row might indicate end of section
                if row > data_start + 10:  # Give some buffer
                    schwab_ira_end = row - 1
                    break
        
        if qqqi_exists:
            print("ℹ️  QQQI already exists in Schwab IRA - no action needed")
            return True
        
        # Add QQQI at the end of Schwab IRA section
        insert_row = schwab_ira_end + 1
        print(f"➕ Adding QQQI at row {insert_row}")
        
        # Insert a new row
        ws.insert_rows(insert_row)
        
        # Add QQQI data
        ws.cell(row=insert_row, column=1).value = "QQQI"
        ws.cell(row=insert_row, column=1).font = Font(name="Arial", size=12, bold=True, color="3072C2")
        
        # Add position data
        quantity = qqqi_position.get('quantity', 0)
        market_value = qqqi_position.get('market_value', 0)
        price = market_value / quantity if quantity > 0 else 0
        
        ws.cell(row=insert_row, column=2).value = quantity
        ws.cell(row=insert_row, column=4).value = round(price, 2)
        ws.cell(row=insert_row, column=5).value = round(market_value, 2)
        
        print(f"✅ Added QQQI: {quantity} shares @ ${price:.2f} = ${market_value:,.2f}")
        
        # Save
        wb.save(self.excel_file)
        print(f"✅ Saved to: {self.excel_file}")
        print("🎉 QQQI successfully added to Schwab IRA!")
        
        return True

if __name__ == "__main__":
    adder = SurgicalQQQIAddition()
    adder.add_qqqi_to_schwab_ira()