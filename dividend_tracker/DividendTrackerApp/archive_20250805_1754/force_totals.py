#!/usr/bin/env python3
"""
Force calculation of totals and ensure they work properly
"""

import openpyxl
from openpyxl.styles import Font
from openpyxl.styles.numbers import FORMAT_CURRENCY_USD_SIMPLE
import os

OUTPUT_FILE = os.path.join(os.path.dirname(__file__), "outputs", "Dividends_2025.xlsx")

TOTAL_FONT = Font(bold=True, name="Arial", size=12)

def force_total_calculations():
    """Force the total calculations to work properly"""
    
    wb = openpyxl.load_workbook(OUTPUT_FILE, data_only=True)  # Get calculated values
    ws = wb["Estimated Income 2025"]
    
    print("🔧 Force calculating totals...")
    
    # Load original workbook for editing
    wb_edit = openpyxl.load_workbook(OUTPUT_FILE)
    ws_edit = wb_edit["Estimated Income 2025"]
    
    # === FORCE DIVIDEND TOTALS ===
    print("💰 Force calculating dividend totals (Row 8)...")
    
    for col in range(2, ws.max_column + 1):
        # Calculate manually from the actual values
        dividend_sum = 0
        for row in range(4, 8):  # Rows 4-7
            value = ws.cell(row=row, column=col).value
            if isinstance(value, (int, float)):
                dividend_sum += value
        
        # Set the calculated value
        total_cell = ws_edit.cell(row=8, column=col, value=dividend_sum)
        total_cell.number_format = FORMAT_CURRENCY_USD_SIMPLE
        total_cell.font = TOTAL_FONT
        
        if dividend_sum > 0:
            col_letter = ws.cell(row=1, column=col).column_letter
            print(f"   Column {col_letter}: ${dividend_sum:,.2f}")
    
    # === FORCE PORTFOLIO TOTALS ===
    print("💰 Force calculating portfolio totals (Row 19)...")
    
    for col in range(2, ws.max_column + 1):
        # Calculate manually from the actual values
        portfolio_sum = 0
        for row in range(14, 19):  # Rows 14-18
            value = ws.cell(row=row, column=col).value
            if isinstance(value, (int, float)):
                portfolio_sum += value
        
        # Set the calculated value
        total_cell = ws_edit.cell(row=19, column=col, value=portfolio_sum)
        total_cell.number_format = FORMAT_CURRENCY_USD_SIMPLE
        total_cell.font = TOTAL_FONT
        
        if portfolio_sum > 0:
            col_letter = ws.cell(row=1, column=col).column_letter
            print(f"   Column {col_letter}: ${portfolio_sum:,.2f}")
    
    # Save the workbook
    wb_edit.save(OUTPUT_FILE)
    wb.close()
    
    print("✅ Forced calculations complete!")
    
    # === NOW ADD COLOR CODING ===
    print("🎨 Adding color coding to calculated totals...")
    
    # Reload with new calculated values
    wb_calc = openpyxl.load_workbook(OUTPUT_FILE, data_only=True)
    ws_calc = wb_calc["Estimated Income 2025"]
    
    # Reload for editing
    wb_edit = openpyxl.load_workbook(OUTPUT_FILE)
    ws_edit = wb_edit["Estimated Income 2025"]
    
    from openpyxl.styles import PatternFill
    GREEN_FILL = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
    RED_FILL = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
    
    # Color code dividend totals
    for col in range(3, ws_calc.max_column + 1):
        current_value = ws_calc.cell(row=8, column=col).value
        previous_value = ws_calc.cell(row=8, column=col-1).value
        
        if (isinstance(current_value, (int, float)) and 
            isinstance(previous_value, (int, float)) and
            current_value > 0 and previous_value > 0):
            
            cell = ws_edit.cell(row=8, column=col)
            if current_value > previous_value:
                cell.fill = GREEN_FILL
            elif current_value < previous_value:
                cell.fill = RED_FILL
    
    # Color code portfolio totals
    for col in range(3, ws_calc.max_column + 1):
        current_value = ws_calc.cell(row=19, column=col).value
        previous_value = ws_calc.cell(row=19, column=col-1).value
        
        if (isinstance(current_value, (int, float)) and 
            isinstance(previous_value, (int, float)) and
            current_value > 0 and previous_value > 0):
            
            cell = ws_edit.cell(row=19, column=col)
            if current_value > previous_value:
                cell.fill = GREEN_FILL
            elif current_value < previous_value:
                cell.fill = RED_FILL
    
    # Final save
    wb_edit.save(OUTPUT_FILE)
    wb_calc.close()
    
    print("🎨 Color coding complete!")
    print("✅ All totals now working with proper color coding!")

if __name__ == "__main__":
    force_total_calculations()
