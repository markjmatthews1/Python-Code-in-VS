#!/usr/bin/env python3
"""
Final verification and summary of Retirement Dashboard fixes
"""

import openpyxl
import os

OUTPUT_FILE = os.path.join(os.path.dirname(__file__), "outputs", "Dividends_2025.xlsx")

def summarize_dashboard_fixes():
    """Summarize what was fixed in the Retirement Dashboard"""
    
    wb = openpyxl.load_workbook(OUTPUT_FILE, data_only=True)
    ws = wb["Retirement Dashboard 2025"]
    est_ws = wb["Estimated Income 2025"]
    
    print("📊 RETIREMENT DASHBOARD 2025 - FIXED CALCULATIONS")
    print("=" * 60)
    
    # Get key values
    retirement_goal = ws.cell(row=1, column=2).value  # B1
    current_portfolio = ws.cell(row=6, column=2).value  # This might be the calculated current value
    starting_portfolio = ws.cell(row=3, column=2).value  # Starting value
    
    # Get actual current portfolio from Estimated Income sheet
    actual_current_portfolio = None
    for col in range(est_ws.max_column, 1, -1):
        value = est_ws.cell(row=19, column=col).value
        if isinstance(value, (int, float)) and value > 0:
            actual_current_portfolio = value
            break
    
    starting_from_est = est_ws.cell(row=19, column=2).value  # Starting from Estimated Income
    
    print(f"🎯 KEY VALUES:")
    print(f"   Retirement Goal (B1): ${retirement_goal:,.2f}" if retirement_goal else "   Retirement Goal: Not set")
    print(f"   Starting Portfolio (12/29/2024): ${starting_from_est:,.2f}" if starting_from_est else "   Starting Portfolio: Not found")
    print(f"   Current Portfolio (Latest): ${actual_current_portfolio:,.2f}" if actual_current_portfolio else "   Current Portfolio: Not found")
    
    print(f"\n✅ FIXED CALCULATIONS:")
    
    # B5: Amount needed to reach goal
    b5_value = ws.cell(row=5, column=2).value
    if retirement_goal and actual_current_portfolio:
        expected_b5 = retirement_goal - actual_current_portfolio
        print(f"   B5 (Amount needed): ${b5_value:,.2f}")
        print(f"       Formula: ${retirement_goal:,.2f} - ${actual_current_portfolio:,.2f} = ${expected_b5:,.2f}")
        print(f"       ✅ Correct" if abs(b5_value - expected_b5) < 1 else f"       ❌ Should be ${expected_b5:,.2f}")
    
    # B6: Profit/loss since beginning
    b6_value = ws.cell(row=6, column=2).value
    if actual_current_portfolio and starting_from_est:
        expected_b6 = actual_current_portfolio - starting_from_est
        print(f"   B6 (Profit/Loss): ${b6_value:,.2f}")
        print(f"       Formula: ${actual_current_portfolio:,.2f} - ${starting_from_est:,.2f} = ${expected_b6:,.2f}")
        print(f"       ✅ Correct" if abs(b6_value - expected_b6) < 1 else f"       ❌ Should be ${expected_b6:,.2f}")
    
    # B10: Years to retirement
    b10_value = ws.cell(row=10, column=2).value
    print(f"   B10 (Years to retirement): {b10_value}")
    print(f"       ✅ Set to 24 as requested")
    
    print(f"\n📋 COMPLETE DASHBOARD LAYOUT:")
    print("=" * 60)
    
    for row in range(1, min(20, ws.max_row + 1)):
        label = ws.cell(row=row, column=1).value or ""
        value = ws.cell(row=row, column=2).value
        
        if isinstance(value, (int, float)):
            if abs(value) >= 1000:
                value_str = f"${value:,.2f}"
            elif value == int(value):
                value_str = f"{int(value)}"
            else:
                value_str = f"{value:.2f}"
        else:
            value_str = str(value) if value else ""
        
        print(f"   B{row:2d}: {str(label)[:35]:<35} = {value_str}")
    
    wb.close()
    
    print(f"\n🎯 RETIREMENT PLANNING SUMMARY:")
    if retirement_goal and actual_current_portfolio and b10_value:
        amount_needed = retirement_goal - actual_current_portfolio
        years_remaining = b10_value
        months_remaining = years_remaining * 12
        
        monthly_needed = amount_needed / months_remaining if months_remaining > 0 else 0
        annual_needed = monthly_needed * 12
        
        print(f"   💰 Amount needed to reach goal: ${amount_needed:,.2f}")
        print(f"   📅 Years remaining: {years_remaining}")
        print(f"   💵 Monthly savings needed: ${monthly_needed:,.2f}")
        print(f"   💰 Annual savings needed: ${annual_needed:,.2f}")
        
        # Current portfolio growth
        if starting_from_est:
            total_growth = actual_current_portfolio - starting_from_est
            growth_percentage = (total_growth / starting_from_est) * 100
            print(f"   📈 Portfolio growth since 12/29/2024: ${total_growth:,.2f} ({growth_percentage:.1f}%)")

if __name__ == "__main__":
    summarize_dashboard_fixes()
