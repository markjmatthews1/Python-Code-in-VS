#!/usr/bin/env python3
"""
FIXED Integration - Correct Column Structure
Column 14: Beginning Dividend Yield (preserved)
Column 15: New yield data (inserted with today's date)
"""

def fixed_update_dividend_sheet_integration():
    """
    Fixed version that uses correct column positioning
    """
    
    from etrade_auth import get_etrade_session
    import openpyxl
    from openpyxl.styles import PatternFill
    from datetime import date
    
    # Get E*TRADE session
    session, base_url = get_etrade_session()
    
    # Account selection logic here...
    file_path = "dividend_stocks.xlsx"
    
    try:
        workbook = openpyxl.load_workbook(file_path)
        sheet = workbook.active

        # FIXED: Insert new column at position 15 (after Beginning Dividend Yield in column 14)
        today = date.today().strftime("%m-%d-%Y")
        sheet.insert_cols(15)  # CORRECTED: Insert at 15, not 17
        sheet.cell(row=1, column=15).value = today

        # Iterate through rows and update with market quote data
        for index, row in enumerate(sheet.iter_rows(min_row=2, values_only=False), start=2):
            symbol = row[0].value
            if symbol is None or not isinstance(symbol, str) or symbol.strip() == "":
                break

            try:
                # Get quote data from E*TRADE
                quote_url = f"{base_url}/v1/market/quote/{symbol}.json"
                quote_response = session.get(quote_url)
                quote_json = quote_response.json()
                
                if 'QuoteResponse' in quote_json and 'QuoteData' in quote_json['QuoteResponse']:
                    quote_data = quote_json['QuoteResponse']['QuoteData'][0]
                    
                    # Update last trade price (column 4)
                    last_trade = quote_data['All'].get('lastTrade', '')
                    if last_trade:
                        sheet.cell(row=index, column=4).value = round(float(last_trade), 2)

                    # FIXED: Put current yield in column 15 (newly inserted)
                    current_yield = quote_data['All'].get('yield', '')
                    if current_yield:
                        sheet.cell(row=index, column=15).value = round(float(current_yield), 2)
                    else:
                        sheet.cell(row=index, column=15).value = current_yield

                    # FIXED: Compare column 15 (new) vs column 14 (beginning)
                    col_15_value = sheet.cell(row=index, column=15).value  # New yield
                    col_14_value = sheet.cell(row=index, column=14).value  # Beginning yield
                    
                    try:
                        col_15_value = float(col_15_value) if col_15_value else 0
                        col_14_value = float(col_14_value) if col_14_value else 0
                        
                        if col_15_value < col_14_value:
                            fill_color = "FF0000"  # Red
                        elif col_15_value > col_14_value:
                            fill_color = "00FF00"  # Green
                        else:
                            fill_color = "FFFF00"  # Yellow
                            
                        fill = PatternFill(start_color=fill_color, end_color=fill_color, fill_type="solid")
                        sheet.cell(row=index, column=15).fill = fill
                        
                    except (TypeError, ValueError):
                        print(f"Skipping row {index} due to invalid data types for comparison.")

            except Exception as e:
                print(f"Error processing {symbol}: {e}")

        workbook.save(file_path)
        print(f"File processed and saved at: {file_path}")
        
    except Exception as e:
        print(f"An error occurred: {e}")
