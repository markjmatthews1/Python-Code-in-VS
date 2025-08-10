#!/usr/bin/env python3
"""
Fix the Estimated Income 2025 sheet structure
Separate dividend estimates (top 1-6 rows) from portfolio values (bottom section)
"""

import sys
import os
import pandas as pd
import openpyxl
from openpyxl.styles import Font, Border, Side, PatternFill
from openpyxl.styles.numbers import FORMAT_CURRENCY_USD_SIMPLE
from datetime import datetime

# Add modules to path
sys.path.append(os.path.join(os.path.dirname(__file__), 'modules'))

OUTPUT_FILE = os.path.join(os.path.dirname(__file__), "outputs", "Dividends_2025.xlsx")

# Styling
HEADER_FONT = Font(bold=True, name="Arial", size=12)
SECTION_HEADER_FONT = Font(bold=True, name="Arial", size=14)
NORMAL_FONT = Font(name="Arial", size=12)
BOX_BORDER = Border(
    left=Side(style='thin'), right=Side(style='thin'),
    top=Side(style='thin'), bottom=Side(style='thin')
)
GREEN_FILL = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
RED_FILL = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")

def fix_estimated_income_sheet():
    """Fix the Estimated Income 2025 sheet structure"""
    
    if not os.path.exists(OUTPUT_FILE):
        print(f"❌ File not found: {OUTPUT_FILE}")
        return
    
    wb = openpyxl.load_workbook(OUTPUT_FILE)
    
    # Remove existing problematic sheet
    if "Estimated Income 2025" in wb.sheetnames:
        del wb["Estimated Income 2025"]
        print("🗑️ Removed existing Estimated Income 2025 sheet")
    
    # Create new sheet with proper structure
    ws = wb.create_sheet(title="Estimated Income 2025", index=0)
    
    # Sample weeks (you can add more as needed)
    weeks = [
        "12/29/2024", "01/05/2025", "01/12/2025", "01/19/2025", 
        "01/26/2025", "02/02/2025", "02/09/2025", "02/16/2025"
    ]
    
    current_row = 1
    
    # === SECTION 1: ESTIMATED MONTHLY DIVIDEND INCOME (Rows 1-6) ===
    ws.cell(row=current_row, column=1, value="ESTIMATED MONTHLY DIVIDEND INCOME")
    ws.cell(row=current_row, column=1).font = SECTION_HEADER_FONT
    current_row += 2
    
    # Headers for dividend section
    ws.cell(row=current_row, column=1, value="Account Type")
    ws.cell(row=current_row, column=1).font = HEADER_FONT
    
    for col_idx, week in enumerate(weeks, start=2):
        ws.cell(row=current_row, column=col_idx, value=week)
        ws.cell(row=current_row, column=col_idx).font = HEADER_FONT
        ws.cell(row=current_row, column=col_idx).border = BOX_BORDER
    
    current_row += 1
    
    # Dividend estimate rows (rows 4-5)
    dividend_accounts = ["E*TRADE IRA", "E*TRADE Taxable"]
    
    for account in dividend_accounts:
        ws.cell(row=current_row, column=1, value=account)
        ws.cell(row=current_row, column=1).font = NORMAL_FONT
        
        # Sample dividend values (you'll replace these with real data)
        sample_dividends = [1850.00, 1875.00, 1900.00, 1925.00, 1950.00, 1975.00, 2000.00, 2025.00]
        
        for col_idx, dividend in enumerate(sample_dividends, start=2):
            if col_idx <= len(weeks) + 1:
                cell = ws.cell(row=current_row, column=col_idx, value=dividend)
                cell.number_format = FORMAT_CURRENCY_USD_SIMPLE
                cell.font = NORMAL_FONT
                cell.border = BOX_BORDER
        
        current_row += 1
    
    # Total dividend row (row 6)
    ws.cell(row=current_row, column=1, value="TOTAL DIVIDEND INCOME")
    ws.cell(row=current_row, column=1).font = Font(bold=True, name="Arial", size=12)
    
    for col_idx in range(2, len(weeks) + 2):
        # Calculate total from above rows
        formula = f"=SUM({ws.cell(row=current_row-2, column=col_idx).coordinate}:{ws.cell(row=current_row-1, column=col_idx).coordinate})"
        total_cell = ws.cell(row=current_row, column=col_idx, value=formula)
        total_cell.number_format = FORMAT_CURRENCY_USD_SIMPLE
        total_cell.font = Font(bold=True, name="Arial", size=12)
        total_cell.border = BOX_BORDER
    
    current_row += 4  # Leave space between sections
    
    # === SECTION 2: PORTFOLIO VALUE TRACKING ===
    ws.cell(row=current_row, column=1, value="PORTFOLIO VALUE TRACKING")
    ws.cell(row=current_row, column=1).font = SECTION_HEADER_FONT
    current_row += 2
    
    # Headers for portfolio section
    ws.cell(row=current_row, column=1, value="Account Type")
    ws.cell(row=current_row, column=1).font = HEADER_FONT
    
    for col_idx, week in enumerate(weeks, start=2):
        ws.cell(row=current_row, column=col_idx, value=week)
        ws.cell(row=current_row, column=col_idx).font = HEADER_FONT
        ws.cell(row=current_row, column=col_idx).border = BOX_BORDER
    
    current_row += 1
    
    # Portfolio account rows
    portfolio_accounts = [
        ("E*TRADE IRA", [278418.00, 280000.00, 282000.00, 285000.00, 287000.00, 289000.00, 291000.00, 293000.00]),
        ("E*TRADE Taxable", [62110.00, 62500.00, 63000.00, 63500.00, 64000.00, 64500.00, 65000.00, 65500.00]),
        ("Schwab Individual", [1605.60, 1610.00, 1615.00, 1620.00, 1625.00, 1630.00, 1635.00, 1640.00]),
        ("Schwab IRA", [49951.53, 50100.00, 50250.00, 50400.00, 50550.00, 50700.00, 50850.00, 51000.00]),
        ("401K Retirement", [125000.00, 125500.00, 126000.00, 126500.00, 127000.00, 127500.00, 128000.00, 128500.00])
    ]
    
    for account, values in portfolio_accounts:
        ws.cell(row=current_row, column=1, value=account)
        ws.cell(row=current_row, column=1).font = NORMAL_FONT
        
        for col_idx, value in enumerate(values, start=2):
            if col_idx <= len(weeks) + 1:
                cell = ws.cell(row=current_row, column=col_idx, value=value)
                cell.number_format = FORMAT_CURRENCY_USD_SIMPLE
                cell.font = NORMAL_FONT
                cell.border = BOX_BORDER
        
        current_row += 1
    
    # Total portfolio row
    ws.cell(row=current_row, column=1, value="TOTAL PORTFOLIO")
    ws.cell(row=current_row, column=1).font = Font(bold=True, name="Arial", size=12)
    
    for col_idx in range(2, len(weeks) + 2):
        # Calculate total from above portfolio rows
        start_row = current_row - len(portfolio_accounts)
        end_row = current_row - 1
        formula = f"=SUM({ws.cell(row=start_row, column=col_idx).coordinate}:{ws.cell(row=end_row, column=col_idx).coordinate})"
        total_cell = ws.cell(row=current_row, column=col_idx, value=formula)
        total_cell.number_format = FORMAT_CURRENCY_USD_SIMPLE
        total_cell.font = Font(bold=True, name="Arial", size=12)
        total_cell.border = BOX_BORDER
    
    # Auto-size columns
    for col in range(1, len(weeks) + 2):
        ws.column_dimensions[ws.cell(row=1, column=col).column_letter].width = 18
    
    # Save workbook
    wb.save(OUTPUT_FILE)
    print(f"✅ Fixed Estimated Income 2025 sheet structure in: {OUTPUT_FILE}")
    print("📋 Sheet now has:")
    print("   Rows 1-6: Estimated Monthly Dividend Income section")
    print("   Rows 10+: Portfolio Value Tracking section")

if __name__ == "__main__":
    fix_estimated_income_sheet()
