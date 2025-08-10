#!/usr/bin/env python3
"""
Fix dividend totals to show monthly values (divide by 12)
"""

import openpyxl
from openpyxl.styles import Font, PatternFill
from openpyxl.styles.numbers import FORMAT_CURRENCY_USD_SIMPLE
import os

OUTPUT_FILE = os.path.join(os.path.dirname(__file__), "outputs", "Dividends_2025.xlsx")

TOTAL_FONT = Font(bold=True, name="Arial", size=12)
GREEN_FILL = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
RED_FILL = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")

def fix_monthly_dividend_totals():
    """Fix dividend totals to show monthly values (annual ÷ 12)"""
    
    wb = openpyxl.load_workbook(OUTPUT_FILE, data_only=True)  # Get calculated values
    ws = wb["Estimated Income 2025"]
    
    print("📊 Converting dividend totals from annual to monthly values...")
    
    # Load original workbook for editing
    wb_edit = openpyxl.load_workbook(OUTPUT_FILE)
    ws_edit = wb_edit["Estimated Income 2025"]
    
    # === FIX DIVIDEND TOTALS TO MONTHLY ===
    print("💰 Converting dividend totals to monthly (Row 8)...")
    
    for col in range(2, ws.max_column + 1):
        # Get the current annual total
        annual_total = 0
        for row in range(4, 8):  # Rows 4-7
            value = ws.cell(row=row, column=col).value
            if isinstance(value, (int, float)):
                annual_total += value
        
        # Convert to monthly (divide by 12)
        monthly_total = annual_total / 12 if annual_total > 0 else 0
        
        # Set the monthly value
        total_cell = ws_edit.cell(row=8, column=col, value=monthly_total)
        total_cell.number_format = FORMAT_CURRENCY_USD_SIMPLE
        total_cell.font = TOTAL_FONT
        
        if monthly_total > 0:
            col_letter = ws.cell(row=1, column=col).column_letter
            print(f"   Column {col_letter}: ${annual_total:,.2f}/year → ${monthly_total:,.2f}/month")
    
    # Save the workbook
    wb_edit.save(OUTPUT_FILE)
    wb.close()
    
    print("✅ Converted to monthly values!")
    
    # === NOW ADD COLOR CODING FOR MONTHLY VALUES ===
    print("🎨 Adding color coding to monthly dividend totals...")
    
    # Reload with new calculated values
    wb_calc = openpyxl.load_workbook(OUTPUT_FILE, data_only=True)
    ws_calc = wb_calc["Estimated Income 2025"]
    
    # Reload for editing
    wb_edit = openpyxl.load_workbook(OUTPUT_FILE)
    ws_edit = wb_edit["Estimated Income 2025"]
    
    # Color code dividend totals (monthly values)
    for col in range(3, ws_calc.max_column + 1):
        current_value = ws_calc.cell(row=8, column=col).value
        previous_value = ws_calc.cell(row=8, column=col-1).value
        
        if (isinstance(current_value, (int, float)) and 
            isinstance(previous_value, (int, float)) and
            current_value > 0 and previous_value > 0):
            
            cell = ws_edit.cell(row=8, column=col)
            if current_value > previous_value:
                cell.fill = GREEN_FILL
                print(f"   Column {col}: GREEN (${current_value:.2f} > ${previous_value:.2f})")
            elif current_value < previous_value:
                cell.fill = RED_FILL
                print(f"   Column {col}: RED (${current_value:.2f} < ${previous_value:.2f})")
    
    # Final save
    wb_edit.save(OUTPUT_FILE)
    wb_calc.close()
    
    print("🎨 Color coding complete!")
    print("✅ Dividend totals now show MONTHLY values with proper color coding!")

def verify_monthly_totals():
    """Verify the monthly totals are correct"""
    
    wb = openpyxl.load_workbook(OUTPUT_FILE, data_only=True)
    ws = wb["Estimated Income 2025"]
    
    print("\n🔍 Verifying monthly dividend totals...")
    print("=" * 50)
    
    # Check a few sample columns
    sample_columns = [2, 10, 20, 30]
    
    for col in sample_columns:
        if col <= ws.max_column:
            col_letter = ws.cell(row=1, column=col).column_letter
            
            # Get annual total from individual accounts
            annual_total = 0
            for row in range(4, 8):
                value = ws.cell(row=row, column=col).value
                if isinstance(value, (int, float)):
                    annual_total += value
            
            # Get the monthly total
            monthly_total = ws.cell(row=8, column=col).value
            expected_monthly = annual_total / 12
            
            print(f"Column {col_letter}:")
            print(f"  Annual total: ${annual_total:,.2f}")
            print(f"  Expected monthly: ${expected_monthly:,.2f}")
            print(f"  Actual monthly: ${monthly_total:,.2f}" if isinstance(monthly_total, (int, float)) else f"  Actual monthly: {monthly_total}")
            print(f"  ✅ Correct" if abs(expected_monthly - (monthly_total or 0)) < 0.01 else "  ❌ Incorrect")
            print()
    
    wb.close()

if __name__ == "__main__":
    fix_monthly_dividend_totals()
    verify_monthly_totals()
