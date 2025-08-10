#!/usr/bin/env python3
"""
Integration Wrapper for Ticker Dividend Tracking
Connects the comprehensive ticker tracker with existing Update_dividend_sheet.py
"""

import os
import sys
import subprocess
from datetime import datetime

# Add modules path
sys.path.append(os.path.join(os.path.dirname(__file__), 'modules'))

def run_integrated_ticker_update():
    """
    Run the complete integrated ticker update workflow
    """
    
    print("🚀 INTEGRATED TICKER DIVIDEND UPDATE WORKFLOW")
    print("=" * 60)
    print(f"Started at: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}")
    
    # Step 1: Run the comprehensive ticker tracker
    print("\n📊 STEP 1: Running Comprehensive Ticker Tracker...")
    print("-" * 50)
    
    try:
        result = subprocess.run([
            sys.executable, 
            "comprehensive_ticker_tracker.py"
        ], capture_output=True, text=True)
        
        if result.returncode == 0:
            print("✅ Comprehensive ticker tracker completed successfully")
            if result.stdout:
                print(result.stdout)
        else:
            print("❌ Error in comprehensive ticker tracker:")
            print(result.stderr)
    except Exception as e:
        print(f"❌ Failed to run comprehensive ticker tracker: {e}")
    
    # Step 2: Check if we can run the existing Update_dividend_sheet.py
    print("\n📈 STEP 2: Checking Update Dividend Sheet Integration...")
    print("-" * 50)
    
    update_sheet_path = os.path.join("modules", "Update_dividend_sheet.py")
    if os.path.exists(update_sheet_path):
        print(f"✅ Found existing Update_dividend_sheet.py")
        print("   This can be integrated to update live prices and yields")
        print("   (Requires E*TRADE authentication)")
    else:
        print("❌ Update_dividend_sheet.py not found in modules/")
    
    # Step 3: Run the main dividend tracker
    print("\n🎯 STEP 3: Running Enhanced Dividend Tracker...")
    print("-" * 50)
    
    enhanced_tracker_path = "enhanced_dividend_tracker.py"
    if os.path.exists(enhanced_tracker_path):
        try:
            result = subprocess.run([
                sys.executable, 
                enhanced_tracker_path
            ], capture_output=True, text=True)
            
            if result.returncode == 0:
                print("✅ Enhanced dividend tracker completed successfully")
            else:
                print("⚠️ Enhanced dividend tracker had issues:")
                print(result.stderr)
        except Exception as e:
            print(f"❌ Failed to run enhanced dividend tracker: {e}")
    else:
        print("❌ enhanced_dividend_tracker.py not found")
    
    # Step 4: Validation report
    print("\n🔍 STEP 4: Validation Report...")
    print("-" * 50)
    
    create_validation_report()
    
    print(f"\n✅ INTEGRATED TICKER UPDATE COMPLETED")
    print(f"Finished at: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}")
    print("=" * 60)

def create_validation_report():
    """
    Create a validation report comparing ticker-level vs portfolio-level data
    """
    
    try:
        import openpyxl
        
        output_file = os.path.join("outputs", "Dividends_2025.xlsx")
        if not os.path.exists(output_file):
            print("❌ Dividends_2025.xlsx not found for validation")
            return
        
        wb = openpyxl.load_workbook(output_file, data_only=True)
        
        # Get portfolio-level dividend data
        portfolio_dividend = 0
        if "Estimated Income 2025" in wb.sheetnames:
            ws_portfolio = wb["Estimated Income 2025"]
            # Look for the most recent week's dividend total
            for row in range(4, 8):  # Dividend rows
                for col in range(2, ws_portfolio.max_column + 1):
                    cell_value = ws_portfolio.cell(row=row, column=col).value
                    if isinstance(cell_value, (int, float)) and cell_value > 0:
                        portfolio_dividend = max(portfolio_dividend, cell_value)
        
        # Get ticker-level dividend data
        ticker_dividend = 0
        if "Ticker Analysis 2025" in wb.sheetnames:
            ws_ticker = wb["Ticker Analysis 2025"]
            # Find the totals row for monthly dividend
            for row in range(2, ws_ticker.max_row + 1):
                if ws_ticker.cell(row=row, column=1).value == "TOTALS":
                    ticker_dividend = ws_ticker.cell(row=row, column=15).value or 0
                    break
        
        # Create validation report
        print(f"📊 VALIDATION REPORT:")
        print(f"   Portfolio-Level Monthly Dividend: ${portfolio_dividend:,.2f}")
        print(f"   Ticker-Level Monthly Dividend: ${ticker_dividend:,.2f}")
        
        if ticker_dividend > 0:
            difference = portfolio_dividend - ticker_dividend
            percentage = (difference / portfolio_dividend * 100) if portfolio_dividend > 0 else 0
            print(f"   Difference: ${difference:,.2f} ({percentage:.1f}%)")
            
            if abs(percentage) < 5:
                print("   ✅ Data validates well (within 5%)")
            elif abs(percentage) < 15:
                print("   ⚠️ Moderate difference (5-15%) - may need investigation")
            else:
                print("   ❌ Significant difference (>15%) - requires investigation")
        else:
            print("   ⚠️ No ticker-level data available for comparison")
        
        wb.close()
        
    except Exception as e:
        print(f"   ❌ Error creating validation report: {e}")

def show_integration_plan():
    """
    Show the complete integration plan
    """
    
    print("\n📋 COMPLETE INTEGRATION PLAN:")
    print("=" * 40)
    
    print("\n🎯 CURRENT CAPABILITIES:")
    print("   ✅ Portfolio-level dividend tracking (Estimated Income 2025)")
    print("   ✅ Portfolio value tracking with historical data")
    print("   ✅ Retirement dashboard for active retiree")
    print("   ✅ Alpha Vantage API integration for dividend estimates")
    print("   ✅ Excel automation with proper formatting")
    
    print("\n🚀 NEW TICKER-LEVEL CAPABILITIES:")
    print("   ✅ Individual stock position tracking")
    print("   ✅ Automated calculation of all ticker metrics")
    print("   ✅ Yield change tracking with color coding")
    print("   ✅ Multi-account support (E*TRADE + Schwab)")
    print("   ✅ Payment cycle management (Monthly/Quarterly/Weekly)")
    print("   ✅ Cross-validation between ticker and portfolio levels")
    
    print("\n⚙️ AUTOMATION FEATURES:")
    print("   🔄 Auto-updates share quantities from APIs")
    print("   🔄 Auto-calculates current values and gains")
    print("   🔄 Auto-tracks dividend yields over time")
    print("   🔄 Auto-applies color coding for changes")
    print("   🔄 Auto-validates against portfolio totals")
    
    print("\n📅 WEEKLY WORKFLOW:")
    print("   1. Run integrated_ticker_update.py")
    print("   2. Review Ticker Analysis 2025 sheet")
    print("   3. Check validation report")
    print("   4. Manual 401K update (if needed)")
    print("   5. Review complete portfolio dashboard")
    
    print("\n📊 OUTPUT SHEETS:")
    print("   • Estimated Income 2025 (existing)")
    print("   • Retirement Dashboard 2025 (existing)")  
    print("   • Ticker Analysis 2025 (new)")
    print("   • All integrated in single Dividends_2025.xlsx file")

if __name__ == "__main__":
    if len(sys.argv) > 1 and sys.argv[1] == "--plan":
        show_integration_plan()
    else:
        run_integrated_ticker_update()
