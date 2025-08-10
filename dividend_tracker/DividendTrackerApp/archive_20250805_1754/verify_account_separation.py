#!/usr/bin/env python3
"""
Verify that accounts and tickers are properly separated in the Ticker Analysis sheet
"""

import openpyxl
import pandas as pd

def verify_ticker_separation():
    """Check that duplicate tickers are properly separated by account"""
    
    print("🔍 Verifying Account and Ticker Separation")
    print("=" * 50)
    
    # Load the workbook
    wb = openpyxl.load_workbook("Dividends_2025.xlsx")
    
    if "Ticker Analysis 2025" not in wb.sheetnames:
        print("❌ Ticker Analysis 2025 sheet not found!")
        return
    
    ws = wb["Ticker Analysis 2025"]
    
    # Get data from the sheet
    data = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        if row[0] and row[1]:  # Ticker and Account columns
            data.append({
                'Ticker': row[0],
                'Account': row[1],
                'Quantity': row[2]
            })
    
    print(f"📊 Found {len(data)} total positions")
    print()
    
    # Group by ticker to find duplicates
    ticker_accounts = {}
    for item in data:
        ticker = item['Ticker']
        account = item['Account']
        
        if ticker not in ticker_accounts:
            ticker_accounts[ticker] = []
        ticker_accounts[ticker].append({
            'account': account,
            'quantity': item['Quantity']
        })
    
    # Find tickers that appear in multiple accounts
    multi_account_tickers = {k: v for k, v in ticker_accounts.items() if len(v) > 1}
    
    print("🎯 Tickers appearing in multiple accounts:")
    print("-" * 40)
    
    if multi_account_tickers:
        for ticker, accounts in multi_account_tickers.items():
            print(f"\n{ticker}:")
            for acc_info in accounts:
                print(f"  └─ {acc_info['account']}: {acc_info['quantity']} shares")
    else:
        print("None found - each ticker appears in only one account")
    
    print(f"\n📈 Summary:")
    print(f"  • Total unique tickers: {len(ticker_accounts)}")
    print(f"  • Total positions: {len(data)}")
    print(f"  • Multi-account tickers: {len(multi_account_tickers)}")
    
    # Check account ordering
    print(f"\n🔄 Account Ordering:")
    print("-" * 20)
    
    accounts_in_order = []
    current_account = None
    
    for item in data:
        if item['Account'] != current_account:
            current_account = item['Account']
            if current_account not in accounts_in_order:
                accounts_in_order.append(current_account)
    
    for i, account in enumerate(accounts_in_order, 1):
        print(f"  {i}. {account}")
    
    expected_order = ['E*TRADE IRA', 'E*TRADE Taxable', 'Schwab Individual', 'Schwab IRA']
    if accounts_in_order[:len(expected_order)] == expected_order:
        print("  ✅ Account ordering is correct!")
    else:
        print("  ❌ Account ordering needs adjustment")
        print(f"  Expected: {expected_order}")
        print(f"  Actual: {accounts_in_order}")

if __name__ == "__main__":
    verify_ticker_separation()
