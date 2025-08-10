#!/usr/bin/env python3
"""
Fix Retirement Dashboard for current retiree - tracking portfolio performance and income
"""

import openpyxl
from openpyxl.styles import Font
from openpyxl.styles.numbers import FORMAT_CURRENCY_USD_SIMPLE, FORMAT_PERCENTAGE
import os

OUTPUT_FILE = os.path.join(os.path.dirname(__file__), "outputs", "Dividends_2025.xlsx")

def fix_retiree_dashboard():
    """Fix dashboard for current retiree tracking portfolio and income"""
    
    wb = openpyxl.load_workbook(OUTPUT_FILE, data_only=True)
    ws_read = wb["Retirement Dashboard 2025"]
    est_ws = wb["Estimated Income 2025"]
    
    # Get current portfolio data
    current_portfolio = None
    starting_portfolio = est_ws.cell(row=19, column=2).value  # 12/29/2024
    
    # Get current portfolio value (latest column)
    for col in range(est_ws.max_column, 1, -1):
        value = est_ws.cell(row=19, column=col).value
        if isinstance(value, (int, float)) and value > 0:
            current_portfolio = value
            break
    
    # Get current monthly dividend income (latest column from dividend totals)
    current_monthly_dividend = None
    for col in range(est_ws.max_column, 1, -1):
        value = est_ws.cell(row=8, column=col).value  # Dividend total row
        if isinstance(value, (int, float)) and value > 0:
            current_monthly_dividend = value
            break
    
    wb.close()
    
    # Load for editing
    wb_edit = openpyxl.load_workbook(OUTPUT_FILE)
    ws = wb_edit["Retirement Dashboard 2025"]
    
    print("🏖️ Updating Retirement Dashboard for Active Retiree...")
    print("=" * 60)
    
    # Set up retiree-focused dashboard
    ws.cell(row=1, column=1, value="RETIREMENT PORTFOLIO TRACKING")
    ws.cell(row=1, column=1).font = Font(bold=True, size=14)
    
    ws.cell(row=2, column=1, value="Current Portfolio Value")
    ws.cell(row=2, column=2, value=current_portfolio)
    ws.cell(row=2, column=2).number_format = FORMAT_CURRENCY_USD_SIMPLE
    
    ws.cell(row=3, column=1, value="Starting Value (12/29/2024)")
    ws.cell(row=3, column=2, value=starting_portfolio)
    ws.cell(row=3, column=2).number_format = FORMAT_CURRENCY_USD_SIMPLE
    
    # B4: Portfolio growth since start of year
    if current_portfolio and starting_portfolio:
        portfolio_growth = current_portfolio - starting_portfolio
        ws.cell(row=4, column=1, value="Portfolio Growth (2025)")
        ws.cell(row=4, column=2, value=portfolio_growth)
        ws.cell(row=4, column=2).number_format = FORMAT_CURRENCY_USD_SIMPLE
        
        # B5: Growth percentage
        growth_percentage = (portfolio_growth / starting_portfolio) * 100
        ws.cell(row=5, column=1, value="Growth Percentage (2025)")
        ws.cell(row=5, column=2, value=growth_percentage / 100)  # As decimal for percentage format
        ws.cell(row=5, column=2).number_format = FORMAT_PERCENTAGE
    
    # B6: Current monthly dividend income
    ws.cell(row=6, column=1, value="Current Monthly Dividend")
    ws.cell(row=6, column=2, value=current_monthly_dividend)
    ws.cell(row=6, column=2).number_format = FORMAT_CURRENCY_USD_SIMPLE
    
    # B7: Annual dividend projection
    if current_monthly_dividend:
        annual_dividend = current_monthly_dividend * 12
        ws.cell(row=7, column=1, value="Annual Dividend Projection")
        ws.cell(row=7, column=2, value=annual_dividend)
        ws.cell(row=7, column=2).number_format = FORMAT_CURRENCY_USD_SIMPLE
    
    # B8: Dividend yield on current portfolio
    if current_monthly_dividend and current_portfolio:
        annual_dividend = current_monthly_dividend * 12
        dividend_yield = (annual_dividend / current_portfolio) * 100
        ws.cell(row=8, column=1, value="Portfolio Dividend Yield")
        ws.cell(row=8, column=2, value=dividend_yield / 100)  # As decimal
        ws.cell(row=8, column=2).number_format = FORMAT_PERCENTAGE
    
    # Empty row
    ws.cell(row=9, column=1, value="")
    
    # B10: Days since retirement (if you want to track time retired)
    from datetime import datetime
    retirement_date = datetime(2024, 5, 16)  # From your data "Retired 5/16/24"
    current_date = datetime.now()
    days_retired = (current_date - retirement_date).days
    
    ws.cell(row=10, column=1, value="Days Since Retirement")
    ws.cell(row=10, column=2, value=days_retired)
    
    # B11: Months since retirement
    months_retired = days_retired / 30.44  # Average days per month
    ws.cell(row=11, column=1, value="Months Since Retirement")
    ws.cell(row=11, column=2, value=round(months_retired, 1))
    
    # B12: Average monthly portfolio growth
    if current_portfolio and starting_portfolio and months_retired > 0:
        avg_monthly_growth = portfolio_growth / months_retired
        ws.cell(row=12, column=1, value="Avg Monthly Portfolio Growth")
        ws.cell(row=12, column=2, value=avg_monthly_growth)
        ws.cell(row=12, column=2).number_format = FORMAT_CURRENCY_USD_SIMPLE
    
    # Empty row
    ws.cell(row=13, column=1, value="")
    
    # Account breakdown (from dashboard)
    ws.cell(row=14, column=1, value="ACCOUNT BREAKDOWN")
    ws.cell(row=14, column=1).font = Font(bold=True)
    
    # These might already be in your dashboard - let's preserve them
    account_rows = [
        ("E*TRADE IRA", 15),
        ("E*TRADE Taxable", 16), 
        ("Schwab IRA", 17),
        ("Schwab Individual", 18),
        ("401K", 19)
    ]
    
    # Get current values from Estimated Income sheet for each account
    for account_name, row_num in account_rows:
        ws.cell(row=row_num, column=1, value=account_name)
        
        # Find the account in the Estimated Income sheet portfolio section
        for est_row in range(14, 19):  # Portfolio rows in Estimated Income
            est_account = est_ws.cell(row=est_row, column=1).value
            if est_account and account_name.replace(" ", "").upper() in str(est_account).replace(" ", "").upper():
                # Get latest value for this account
                for col in range(est_ws.max_column, 1, -1):
                    value = est_ws.cell(row=est_row, column=col).value
                    if isinstance(value, (int, float)) and value > 0:
                        ws.cell(row=row_num, column=2, value=value)
                        ws.cell(row=row_num, column=2).number_format = FORMAT_CURRENCY_USD_SIMPLE
                        break
                break
    
    # Save workbook
    wb_edit.save(OUTPUT_FILE)
    
    print("✅ Updated Retirement Dashboard for Active Retiree!")
    print("\n📊 Dashboard now shows:")
    print(f"   📈 Portfolio tracking and growth")
    print(f"   💰 Current dividend income")
    print(f"   📅 Time since retirement")
    print(f"   🎯 Portfolio performance metrics")
    print(f"   💼 Account breakdown")
    
    # Show summary
    if current_portfolio and starting_portfolio and current_monthly_dividend:
        growth = current_portfolio - starting_portfolio
        growth_pct = (growth / starting_portfolio) * 100
        annual_div = current_monthly_dividend * 12
        div_yield = (annual_div / current_portfolio) * 100
        
        print(f"\n🎯 RETIREE PORTFOLIO SUMMARY:")
        print(f"   💰 Current Portfolio: ${current_portfolio:,.2f}")
        print(f"   📈 Growth since 12/29/24: ${growth:,.2f} ({growth_pct:.1f}%)")
        print(f"   💵 Monthly Dividend: ${current_monthly_dividend:,.2f}")
        print(f"   💰 Annual Dividend: ${annual_div:,.2f}")
        print(f"   📊 Dividend Yield: {div_yield:.2f}%")
        print(f"   📅 Days Retired: {days_retired}")

if __name__ == "__main__":
    fix_retiree_dashboard()
