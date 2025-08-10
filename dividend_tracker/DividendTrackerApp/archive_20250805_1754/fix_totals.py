#!/usr/bin/env python3
"""
Fix the total row formulas to properly calculate and display values
"""

import openpyxl
from openpyxl.styles import Font, PatternFill
from openpyxl.styles.numbers import FORMAT_CURRENCY_USD_SIMPLE
import os

OUTPUT_FILE = os.path.join(os.path.dirname(__file__), "outputs", "Dividends_2025.xlsx")

TOTAL_FONT = Font(bold=True, name="Arial", size=12)
GREEN_FILL = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
RED_FILL = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")

def fix_total_formulas():
    """Fix the total row formulas and ensure they calculate properly"""
    
    if not os.path.exists(OUTPUT_FILE):
        print(f"❌ File not found: {OUTPUT_FILE}")
        return
    
    wb = openpyxl.load_workbook(OUTPUT_FILE)
    
    if "Estimated Income 2025" not in wb.sheetnames:
        print("❌ Estimated Income 2025 sheet not found")
        return
    
    ws = wb["Estimated Income 2025"]
    print(f"📋 Working with sheet: {ws.max_row} rows x {ws.max_column} columns")
    
    # === 1. FIX DIVIDEND TOTAL ROW (Row 8) ===
    print("💰 Fixing dividend total formulas (Row 8)...")
    dividend_total_row = 8
    
    # Ensure row 8 has the correct label
    ws.cell(row=dividend_total_row, column=1, value="TOTAL DIVIDEND INCOME")
    ws.cell(row=dividend_total_row, column=1).font = TOTAL_FONT
    
    # Add formulas for each column starting from column 2
    for col in range(2, ws.max_column + 1):
        col_letter = ws.cell(row=1, column=col).column_letter
        formula = f"=SUM({col_letter}4:{col_letter}7)"
        
        total_cell = ws.cell(row=dividend_total_row, column=col, value=formula)
        total_cell.number_format = FORMAT_CURRENCY_USD_SIMPLE
        total_cell.font = TOTAL_FONT
        
        print(f"   Column {col_letter}: {formula}")
    
    # === 2. FIX PORTFOLIO TOTAL ROW (Row 19) ===
    print("💰 Fixing portfolio total formulas (Row 19)...")
    portfolio_total_row = 19
    
    # Ensure row 19 has the correct label
    ws.cell(row=portfolio_total_row, column=1, value="TOTAL PORTFOLIO")
    ws.cell(row=portfolio_total_row, column=1).font = TOTAL_FONT
    
    # Add formulas for each column starting from column 2
    for col in range(2, ws.max_column + 1):
        col_letter = ws.cell(row=1, column=col).column_letter
        formula = f"=SUM({col_letter}14:{col_letter}18)"
        
        total_cell = ws.cell(row=portfolio_total_row, column=col, value=formula)
        total_cell.number_format = FORMAT_CURRENCY_USD_SIMPLE
        total_cell.font = TOTAL_FONT
        
        print(f"   Column {col_letter}: {formula}")
    
    # Save first so Excel can calculate the formulas
    wb.save(OUTPUT_FILE)
    print("💾 Saved workbook with updated formulas")
    
    # === 3. RELOAD WITH CALCULATED VALUES AND ADD COLOR CODING ===
    print("🎨 Adding color coding based on calculated values...")
    wb_calculated = openpyxl.load_workbook(OUTPUT_FILE, data_only=True)
    ws_calc = wb_calculated["Estimated Income 2025"]
    
    # Reload the original workbook to apply formatting
    wb = openpyxl.load_workbook(OUTPUT_FILE)
    ws = wb["Estimated Income 2025"]
    
    # Color code dividend total row
    print("🎨 Color coding dividend totals (Row 8)...")
    for col in range(3, ws.max_column + 1):  # Start from column 3 (second data column)
        current_value = ws_calc.cell(row=dividend_total_row, column=col).value
        previous_value = ws_calc.cell(row=dividend_total_row, column=col-1).value
        
        current_cell = ws.cell(row=dividend_total_row, column=col)
        
        if (isinstance(current_value, (int, float)) and 
            isinstance(previous_value, (int, float)) and
            current_value != 0 and previous_value != 0):
            
            if current_value > previous_value:
                current_cell.fill = GREEN_FILL
                print(f"   Column {col}: GREEN (${current_value:.2f} > ${previous_value:.2f})")
            elif current_value < previous_value:
                current_cell.fill = RED_FILL
                print(f"   Column {col}: RED (${current_value:.2f} < ${previous_value:.2f})")
    
    # Color code portfolio total row
    print("🎨 Color coding portfolio totals (Row 19)...")
    for col in range(3, ws.max_column + 1):  # Start from column 3 (second data column)
        current_value = ws_calc.cell(row=portfolio_total_row, column=col).value
        previous_value = ws_calc.cell(row=portfolio_total_row, column=col-1).value
        
        current_cell = ws.cell(row=portfolio_total_row, column=col)
        
        if (isinstance(current_value, (int, float)) and 
            isinstance(previous_value, (int, float)) and
            current_value != 0 and previous_value != 0):
            
            if current_value > previous_value:
                current_cell.fill = GREEN_FILL
                print(f"   Column {col}: GREEN (${current_value:,.2f} > ${previous_value:,.2f})")
            elif current_value < previous_value:
                current_cell.fill = RED_FILL
                print(f"   Column {col}: RED (${current_value:,.2f} < ${previous_value:,.2f})")
    
    # Final save
    wb.save(OUTPUT_FILE)
    wb_calculated.close()
    
    print(f"\n✅ Successfully fixed total row formulas and color coding!")
    print(f"💰 Dividend totals: Sum of rows 4-7 for each week")
    print(f"💰 Portfolio totals: Sum of rows 14-18 for each week")
    print(f"🟢 Green = Increase from previous week")
    print(f"🔴 Red = Decrease from previous week")

if __name__ == "__main__":
    fix_total_formulas()
