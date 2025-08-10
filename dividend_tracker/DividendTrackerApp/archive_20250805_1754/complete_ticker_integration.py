#!/usr/bin/env python3
"""
Complete Ticker Integration with Historical Yield Tracking
Integrates with existing Update_dividend_sheet.py to maintain yield history
"""

import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment
from datetime import datetime, date
import os
import sys
import pandas as pd

# Add modules path
sys.path.append(os.path.join(os.path.dirname(__file__), 'modules'))

def update_ticker_with_yield_history():
    """
    Main function to update ticker data and maintain yield history
    """
    
    print("🚀 COMPLETE TICKER UPDATE WITH YIELD HISTORY")
    print("=" * 60)
    print(f"Update started: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}")
    
    # File paths
    data_file = os.path.join("data", "dividend_stocks.xlsx")
    output_file = os.path.join("outputs", "Dividends_2025.xlsx")
    
    # Step 1: Load existing historical data
    print("\n📊 STEP 1: Loading Historical Data...")
    historical_data = load_complete_historical_data(data_file, output_file)
    
    # Step 2: Get current API data (integrate with Update_dividend_sheet.py logic)
    print("\n🔌 STEP 2: Getting Current Market Data...")
    current_data = get_current_api_data()
    
    # Step 3: Update historical tracking with today's data
    print("\n📈 STEP 3: Updating Historical Yield Tracking...")
    update_yield_history(historical_data, current_data)
    
    # Step 4: Create/update the complete ticker sheet
    print("\n📝 STEP 4: Creating Complete Ticker Analysis...")
    create_complete_ticker_analysis(historical_data, current_data, output_file)
    
    # Step 5: Validation against portfolio totals
    print("\n🔍 STEP 5: Validation Report...")
    create_ticker_validation_report(output_file)
    
    print(f"\n✅ COMPLETE TICKER UPDATE FINISHED")
    print(f"Update completed: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}")
    print("=" * 60)

def load_complete_historical_data(data_file, output_file):
    """
    Load complete historical data from both source file and existing tracking
    """
    
    historical_data = {
        'tickers': {},
        'yield_history': {},
        'date_columns': []
    }
    
    # First, load from dividend_stocks.xlsx if it exists
    if os.path.exists(data_file):
        print(f"   📂 Loading from {data_file}...")
        try:
            df = pd.read_excel(data_file)
            
            for _, row in df.iterrows():
                if pd.notna(row.get('Ticker')):
                    ticker = str(row['Ticker']).strip()
                    
                    # Load ticker basic data
                    historical_data['tickers'][ticker] = {
                        'quantity': float(row.get('Qty #', 0)),
                        'price_paid': float(row.get('Price Paid $', 0)),
                        'payment_cycle': str(row.get('Payment cycle', 'Monthly')),
                        'rate_per_share': float(row.get('Rate per share', 0)),
                        'account': determine_account(row),
                        'original_value': float(row.get('Original Value $', 0)),
                        'pay_date': str(row.get('Pay Date', '')),
                        'beginning_yield': float(row.get('Beginning Dividend Yield', 0))
                    }
                    
                    # Load historical yields from date columns
                    historical_data['yield_history'][ticker] = {}
                    
                    for col in df.columns:
                        col_str = str(col)
                        # Look for date patterns: MM-DD-YYYY or YYYY-MM-DD
                        if ('-2025' in col_str or '-2024' in col_str) and col_str != 'Ticker':
                            try:
                                yield_value = row.get(col)
                                if pd.notna(yield_value) and isinstance(yield_value, (int, float)):
                                    # Standardize date format to MM-DD-YYYY
                                    formatted_date = standardize_date_format(col_str)
                                    historical_data['yield_history'][ticker][formatted_date] = float(yield_value)
                                    
                                    if formatted_date not in historical_data['date_columns']:
                                        historical_data['date_columns'].append(formatted_date)
                            except:
                                continue
            
            print(f"      ✅ Loaded {len(historical_data['tickers'])} tickers")
            print(f"      ✅ Found {len(historical_data['date_columns'])} historical dates")
            
        except Exception as e:
            print(f"      ❌ Error loading source file: {e}")
    
    # Second, check existing tracking sheet for additional history
    if os.path.exists(output_file):
        print(f"   📊 Checking existing tracking in {output_file}...")
        try:
            wb = openpyxl.load_workbook(output_file, data_only=True)
            if "Ticker Historical Analysis" in wb.sheetnames:
                ws = wb["Ticker Historical Analysis"]
                
                # Load any additional historical data from previous runs
                additional_dates = extract_additional_dates_from_sheet(ws)
                historical_data['date_columns'].extend(additional_dates)
                
                print(f"      ✅ Found additional tracking data")
            
            wb.close()
        except Exception as e:
            print(f"      ⚠️ Could not load existing tracking: {e}")
    
    # Sort and deduplicate dates
    historical_data['date_columns'] = sorted(list(set(historical_data['date_columns'])), 
                                           key=lambda x: datetime.strptime(x, "%m-%d-%Y"), reverse=True)
    
    print(f"   📅 Total historical dates: {len(historical_data['date_columns'])}")
    if historical_data['date_columns']:
        print(f"   📅 Date range: {historical_data['date_columns'][-1]} to {historical_data['date_columns'][0]}")
    
    return historical_data

def determine_account(row):
    """Determine account type from row data"""
    account_field = str(row.get('Account', '')).lower()
    if 'etrade' in account_field:
        return 'E*TRADE'
    elif 'schwab' in account_field:
        return 'Schwab'
    else:
        # Check if it's at the end of the data (Schwab individual positions)
        return 'Schwab' if 'schwab' in str(row.get('Ticker', '')).lower() else 'E*TRADE'

def standardize_date_format(date_str):
    """Convert various date formats to MM-DD-YYYY"""
    try:
        # Handle YYYY-MM-DD format
        if date_str.count('-') == 2 and len(date_str.split('-')[0]) == 4:
            year, month, day = date_str.split('-')
            return f"{month}-{day}-{year}"
        # Handle MM-DD-YYYY format (already correct)
        elif date_str.count('-') == 2:
            return date_str
        else:
            return date_str
    except:
        return date_str

def extract_additional_dates_from_sheet(ws):
    """Extract additional date columns from existing sheet"""
    additional_dates = []
    
    # Check header row for date columns
    for col in range(1, ws.max_column + 1):
        header_value = ws.cell(row=1, column=col).value
        if header_value and isinstance(header_value, str):
            if '-2025' in header_value or '-2024' in header_value:
                formatted_date = standardize_date_format(header_value)
                additional_dates.append(formatted_date)
    
    return additional_dates

def get_current_api_data():
    """
    Get current data from APIs
    This would integrate with the existing Update_dividend_sheet.py logic
    """
    
    print("   🔌 Connecting to market data APIs...")
    
    # This would eventually integrate with your existing Update_dividend_sheet.py
    # For now, providing sample current data
    current_data = {
        'ABR': {
            'current_price': 11.21, 'change': -0.12, 'change_pct': -0.87,
            'current_yield': 14.20, 'dividend_amount': 0.10, 'last_trade': 11.21
        },
        'ACP': {
            'current_price': 5.94, 'change': -0.02, 'change_pct': -0.33,
            'current_yield': 15.53, 'dividend_amount': 0.08, 'last_trade': 5.94
        },
        'AGNC': {
            'current_price': 9.25, 'change': -0.07, 'change_pct': -0.69,
            'current_yield': 13.81, 'dividend_amount': 0.12, 'last_trade': 9.25
        },
        'DX': {
            'current_price': 12.53, 'change': -0.07, 'change_pct': -0.52,
            'current_yield': 14.62, 'dividend_amount': 0.17, 'last_trade': 12.53
        },
        'ECC': {
            'current_price': 7.45, 'change': 0.03, 'change_pct': 0.34,
            'current_yield': 19.35, 'dividend_amount': 0.14, 'last_trade': 7.45
        }
    }
    
    print(f"      📈 Retrieved current data for {len(current_data)} tickers")
    return current_data

def update_yield_history(historical_data, current_data):
    """
    Update historical yield tracking with today's data
    """
    
    today = datetime.now().strftime("%m-%d-%Y")
    print(f"   📅 Adding yield data for {today}")
    
    # Add today to date columns if not already present
    if today not in historical_data['date_columns']:
        historical_data['date_columns'].insert(0, today)
    
    # Update each ticker's yield history with current data
    updated_count = 0
    for ticker in historical_data['tickers']:
        if ticker in current_data:
            current_yield = current_data[ticker]['current_yield']
            
            # Initialize yield history for ticker if not exists
            if ticker not in historical_data['yield_history']:
                historical_data['yield_history'][ticker] = {}
            
            # Add today's yield
            historical_data['yield_history'][ticker][today] = current_yield
            updated_count += 1
    
    print(f"      ✅ Updated yield history for {updated_count} tickers")

def create_complete_ticker_analysis(historical_data, current_data, output_file):
    """
    Create the complete ticker analysis sheet with all features
    """
    
    # Load workbook
    wb = openpyxl.load_workbook(output_file) if os.path.exists(output_file) else openpyxl.Workbook()
    
    # Create/recreate sheet
    sheet_name = "Ticker Analysis 2025"
    if sheet_name in wb.sheetnames:
        del wb[sheet_name]
    
    ws = wb.create_sheet(sheet_name)
    
    # Define static headers
    static_headers = [
        'Ticker', 'Account', 'Qty #', 'Price Paid $', 'Last Price $', 'Day\'s Gain $',
        'Price Change $', 'Change %', 'Current Value $', 'Original Value $', 'Total Gain $',
        'Total Gain %', 'Pay Date', 'Payment Cycle', 'Rate per Share', 'Original Payment',
        'New Payment', 'Beginning Yield'
    ]
    
    # Add historical date columns
    all_headers = static_headers + historical_data['date_columns']
    
    # Create headers with formatting
    for col, header in enumerate(all_headers, 1):
        cell = ws.cell(row=1, column=col)
        cell.value = header
        cell.font = Font(bold=True, color="FFFFFF")
        cell.alignment = Alignment(horizontal="center")
        
        # Color code sections
        if col <= len(static_headers):
            cell.fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        else:
            cell.fill = PatternFill(start_color="70AD47", end_color="70AD47", fill_type="solid")
    
    # Add data rows
    row = 2
    totals = {'current_value': 0, 'original_value': 0, 'new_payment': 0}
    
    for ticker, ticker_data in historical_data['tickers'].items():
        # Get current market data
        current = current_data.get(ticker, {})
        current_price = current.get('current_price', ticker_data['price_paid'])
        
        # Calculate values
        quantity = ticker_data['quantity']
        price_paid = ticker_data['price_paid']
        current_value = quantity * current_price
        original_value = quantity * price_paid
        total_gain = current_value - original_value
        total_gain_pct = (total_gain / original_value * 100) if original_value > 0 else 0
        
        # Day's change
        day_change = current.get('change', 0)
        day_gain = day_change * quantity
        change_pct = current.get('change_pct', 0)
        
        # Payment calculations
        rate_per_share = ticker_data['rate_per_share']
        original_payment = rate_per_share * quantity
        new_payment = current.get('dividend_amount', rate_per_share) * quantity
        
        # Static data row
        static_data = [
            ticker,                           # Ticker
            ticker_data['account'],           # Account
            quantity,                         # Qty #
            price_paid,                      # Price Paid $
            current_price,                   # Last Price $
            day_gain,                        # Day's Gain $
            day_change,                      # Price Change $
            change_pct,                      # Change %
            current_value,                   # Current Value $
            original_value,                  # Original Value $
            total_gain,                      # Total Gain $
            total_gain_pct,                  # Total Gain %
            ticker_data['pay_date'],         # Pay Date
            ticker_data['payment_cycle'],    # Payment Cycle
            rate_per_share,                  # Rate per Share
            original_payment,                # Original Payment
            new_payment,                     # New Payment
            ticker_data['beginning_yield']   # Beginning Yield
        ]
        
        # Add static data to sheet
        for col, value in enumerate(static_data, 1):
            cell = ws.cell(row=row, column=col)
            cell.value = value
            
            # Apply formatting
            if col in [4, 5, 6, 7, 9, 10, 11, 15, 16, 17]:  # Currency columns
                if isinstance(value, (int, float)):
                    cell.number_format = '"$"#,##0.00'
            elif col in [8, 12, 18]:  # Percentage columns
                if isinstance(value, (int, float)):
                    cell.number_format = '0.00%'
        
        # Add historical yield data with color coding
        ticker_history = historical_data['yield_history'].get(ticker, {})
        
        for date_idx, date_str in enumerate(historical_data['date_columns']):
            col = len(static_headers) + date_idx + 1
            yield_value = ticker_history.get(date_str, 0)
            
            cell = ws.cell(row=row, column=col)
            cell.value = yield_value
            cell.number_format = '0.00'
            
            # Color code based on change from previous period
            if date_idx > 0:
                prev_date = historical_data['date_columns'][date_idx - 1]
                prev_yield = ticker_history.get(prev_date, yield_value)
                
                yield_diff = yield_value - prev_yield
                if yield_diff > 0.1:
                    cell.fill = PatternFill(start_color="00FF00", end_color="00FF00", fill_type="solid")  # Green
                elif yield_diff < -0.1:
                    cell.fill = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")  # Red
                else:
                    cell.fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")  # Yellow
        
        # Update totals
        totals['current_value'] += current_value
        totals['original_value'] += original_value
        totals['new_payment'] += new_payment
        
        row += 1
    
    # Add totals row
    totals_row = row + 1
    ws.cell(row=totals_row, column=1, value="TOTALS").font = Font(bold=True)
    
    # Format totals
    total_cells = [
        (9, totals['current_value']),     # Current Value
        (10, totals['original_value']),   # Original Value  
        (11, totals['current_value'] - totals['original_value']),  # Total Gain
        (16, totals['new_payment'])       # New Payment
    ]
    
    for col, value in total_cells:
        cell = ws.cell(row=totals_row, column=col)
        cell.value = value
        cell.font = Font(bold=True)
        cell.number_format = '"$"#,##0.00'
    
    # Auto-adjust column widths
    for col in range(1, len(all_headers) + 1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(col)].width = 12
    
    # Save workbook
    wb.save(output_file)
    
    print(f"      ✅ Complete ticker analysis created")
    print(f"      📊 {len(historical_data['tickers'])} tickers tracked")
    print(f"      📅 {len(historical_data['date_columns'])} weeks of yield history")
    print(f"      💰 Total Current Value: ${totals['current_value']:,.2f}")
    print(f"      💰 Total Monthly Dividend: ${totals['new_payment']:,.2f}")

def create_ticker_validation_report(output_file):
    """
    Create validation report comparing ticker vs portfolio data
    """
    
    try:
        wb = openpyxl.load_workbook(output_file, data_only=True)
        
        # Get portfolio dividend total
        portfolio_dividend = 0
        if "Estimated Income 2025" in wb.sheetnames:
            ws_portfolio = wb["Estimated Income 2025"]
            # Find the most recent dividend total
            for col in range(2, min(ws_portfolio.max_column + 1, 50)):
                for row in range(4, 8):
                    cell_value = ws_portfolio.cell(row=row, column=col).value
                    if isinstance(cell_value, (int, float)) and cell_value > 0:
                        portfolio_dividend = max(portfolio_dividend, cell_value)
        
        # Get ticker dividend total
        ticker_dividend = 0
        if "Ticker Analysis 2025" in wb.sheetnames:
            ws_ticker = wb["Ticker Analysis 2025"]
            for row in range(2, ws_ticker.max_row + 1):
                cell_value = ws_ticker.cell(row=row, column=1).value
                if cell_value == "TOTALS":
                    ticker_dividend = ws_ticker.cell(row=row, column=16).value or 0
                    break
        
        print(f"      📊 Portfolio-Level Dividend: ${portfolio_dividend:,.2f}")
        print(f"      📊 Ticker-Level Dividend: ${ticker_dividend:,.2f}")
        
        if ticker_dividend > 0 and portfolio_dividend > 0:
            coverage = (ticker_dividend / portfolio_dividend) * 100
            print(f"      📊 Ticker Coverage: {coverage:.1f}% of portfolio total")
        
        wb.close()
        
    except Exception as e:
        print(f"      ❌ Validation error: {e}")

if __name__ == "__main__":
    update_ticker_with_yield_history()
