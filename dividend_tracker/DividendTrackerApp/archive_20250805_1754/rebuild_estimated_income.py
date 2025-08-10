#!/usr/bin/env python3
"""
Examine the Historical_Import and Portfolio_History_import sheets 
and rebuild Estimated Income 2025 correctly
"""

import sys
import os
import pandas as pd
import openpyxl
from openpyxl.styles import Font, Border, Side, PatternFill
from openpyxl.styles.numbers import FORMAT_CURRENCY_USD_SIMPLE
from datetime import datetime

OUTPUT_FILE = os.path.join(os.path.dirname(__file__), "outputs", "Dividends_2025.xlsx")

# Styling
HEADER_FONT = Font(bold=True, name="Arial", size=12)
SECTION_HEADER_FONT = Font(bold=True, name="Arial", size=14)
NORMAL_FONT = Font(name="Arial", size=12)
BOX_BORDER = Border(
    left=Side(style='thin'), right=Side(style='thin'),
    top=Side(style='thin'), bottom=Side(style='thin')
)
GREEN_FILL = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
RED_FILL = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")

def examine_historical_sheets():
    """Examine the Historical_Import and Portfolio_History_import sheets"""
    
    if not os.path.exists(OUTPUT_FILE):
        print(f"❌ File not found: {OUTPUT_FILE}")
        return None, None
    
    wb = openpyxl.load_workbook(OUTPUT_FILE)
    
    print("📊 Available sheets:", wb.sheetnames)
    
    historical_data = None
    portfolio_data = None
    
    # Examine Historical_Import sheet
    if "Historical_Import" in wb.sheetnames:
        print("\n=== HISTORICAL_IMPORT SHEET ===")
        hist_ws = wb["Historical_Import"]
        print(f"Dimensions: {hist_ws.max_row} rows x {hist_ws.max_column} columns")
        
        # Show first few rows and columns
        print("\n📋 Historical Import Data Structure:")
        for row in range(1, min(11, hist_ws.max_row + 1)):
            row_data = []
            for col in range(1, min(11, hist_ws.max_column + 1)):
                cell_value = hist_ws.cell(row=row, column=col).value
                if cell_value is None:
                    cell_value = ""
                row_data.append(str(cell_value)[:15])  # Truncate long values
            print(f"Row {row:2d}: {' | '.join(row_data)}")
        
        # Extract data to pandas DataFrame for analysis
        if hist_ws.max_row > 1 and hist_ws.max_column > 1:
            data_rows = []
            headers = []
            
            # Get headers from first row
            for col in range(1, hist_ws.max_column + 1):
                header = hist_ws.cell(row=1, column=col).value
                headers.append(str(header) if header else f"Col_{col}")
            
            # Get data rows
            for row in range(2, hist_ws.max_row + 1):
                row_data = []
                for col in range(1, hist_ws.max_column + 1):
                    value = hist_ws.cell(row=row, column=col).value
                    row_data.append(value)
                data_rows.append(row_data)
            
            if data_rows:
                historical_data = pd.DataFrame(data_rows, columns=headers)
                print(f"\n📊 Historical data shape: {historical_data.shape}")
                print(f"📊 Historical data columns: {list(historical_data.columns)}")
        else:
            print("⚠️ Historical_Import sheet appears to be empty")
    
    # Examine Portfolio_History_import sheet
    if "Portfolio_History_import" in wb.sheetnames:
        print("\n=== PORTFOLIO_HISTORY_IMPORT SHEET ===")
        port_ws = wb["Portfolio_History_import"]
        print(f"Dimensions: {port_ws.max_row} rows x {port_ws.max_column} columns")
        
        # Show first few rows and columns
        print("\n📋 Portfolio History Data Structure:")
        for row in range(1, min(11, port_ws.max_row + 1)):
            row_data = []
            for col in range(1, min(11, port_ws.max_column + 1)):
                cell_value = port_ws.cell(row=row, column=col).value
                if cell_value is None:
                    cell_value = ""
                row_data.append(str(cell_value)[:15])  # Truncate long values
            print(f"Row {row:2d}: {' | '.join(row_data)}")
        
        # Extract data to pandas DataFrame for analysis
        if port_ws.max_row > 1 and port_ws.max_column > 1:
            data_rows = []
            headers = []
            
            # Get headers from first row
            for col in range(1, port_ws.max_column + 1):
                header = port_ws.cell(row=1, column=col).value
                headers.append(str(header) if header else f"Col_{col}")
            
            # Get data rows
            for row in range(2, port_ws.max_row + 1):
                row_data = []
                for col in range(1, port_ws.max_column + 1):
                    value = port_ws.cell(row=row, column=col).value
                    row_data.append(value)
                data_rows.append(row_data)
            
            if data_rows:
                portfolio_data = pd.DataFrame(data_rows, columns=headers)
                print(f"\n📊 Portfolio data shape: {portfolio_data.shape}")
                print(f"📊 Portfolio data columns: {list(portfolio_data.columns)}")
        else:
            print("⚠️ Portfolio_History_import sheet appears to be empty")
    
    wb.close()
    return historical_data, portfolio_data

def rebuild_estimated_income_sheet():
    """Rebuild the Estimated Income 2025 sheet using the correct historical data"""
    
    # First examine the data
    hist_data, port_data = examine_historical_sheets()
    
    if hist_data is None or port_data is None:
        print("❌ Could not load historical data")
        return
    
    print("\n🔧 Rebuilding Estimated Income 2025 sheet...")
    
    wb = openpyxl.load_workbook(OUTPUT_FILE)
    
    # Remove existing problematic sheet
    if "Estimated Income 2025" in wb.sheetnames:
        del wb["Estimated Income 2025"]
        print("🗑️ Removed existing Estimated Income 2025 sheet")
    
    # Create new sheet with correct data
    ws = wb.create_sheet(title="Estimated Income 2025", index=0)
    
    print("\n📊 Historical dividend data columns:")
    print(hist_data.columns.tolist())
    
    print("\n📊 Portfolio data columns:")
    print(port_data.columns.tolist())
    
    # Get all date columns from historical data (excluding account/symbol columns)
    date_columns = []
    for col in hist_data.columns:
        if '/' in str(col) or col.startswith('20'):  # Date patterns
            date_columns.append(col)
    
    print(f"\n📅 Found {len(date_columns)} date columns in historical data")
    print("First few dates:", date_columns[:5])
    print("Last few dates:", date_columns[-5:])
    
    # Get all date columns from portfolio data
    port_date_columns = []
    for col in port_data.columns:
        if '/' in str(col) or col.startswith('20'):  # Date patterns
            port_date_columns.append(col)
    
    print(f"\n📅 Found {len(port_date_columns)} date columns in portfolio data")
    
    # Build the sheet using the actual data
    current_row = 1
    
    # === SECTION 1: DIVIDEND ESTIMATES ===
    ws.cell(row=current_row, column=1, value="ESTIMATED MONTHLY DIVIDEND INCOME")
    ws.cell(row=current_row, column=1).font = SECTION_HEADER_FONT
    current_row += 2
    
    # Headers - use the actual date columns from historical data
    ws.cell(row=current_row, column=1, value="Account Type")
    ws.cell(row=current_row, column=1).font = HEADER_FONT
    
    for col_idx, date_col in enumerate(date_columns, start=2):
        ws.cell(row=current_row, column=col_idx, value=str(date_col))
        ws.cell(row=current_row, column=col_idx).font = HEADER_FONT
        ws.cell(row=current_row, column=col_idx).border = BOX_BORDER
    
    current_row += 1
    
    # Dividend data rows - copy from Historical_Import
    print("\n📊 Historical data accounts:")
    unique_accounts = hist_data.iloc[:, 0].unique() if len(hist_data) > 0 else []
    print(unique_accounts)
    
    for account in unique_accounts:
        if pd.isna(account) or account == '':
            continue
            
        account_data = hist_data[hist_data.iloc[:, 0] == account]
        if len(account_data) == 0:
            continue
        
        ws.cell(row=current_row, column=1, value=str(account))
        ws.cell(row=current_row, column=1).font = NORMAL_FONT
        
        # Copy the actual values for each date
        for col_idx, date_col in enumerate(date_columns, start=2):
            if date_col in account_data.columns:
                value = account_data[date_col].iloc[0] if len(account_data) > 0 else 0
                if pd.isna(value):
                    value = 0
                
                cell = ws.cell(row=current_row, column=col_idx, value=float(value) if value != 0 else 0)
                if value != 0:
                    cell.number_format = FORMAT_CURRENCY_USD_SIMPLE
                cell.font = NORMAL_FONT
                cell.border = BOX_BORDER
        
        current_row += 1
    
    # Total dividend row
    dividend_start_row = 4  # First data row
    dividend_end_row = current_row - 1  # Last data row
    
    ws.cell(row=current_row, column=1, value="TOTAL DIVIDEND INCOME")
    ws.cell(row=current_row, column=1).font = Font(bold=True, name="Arial", size=12)
    
    for col_idx in range(2, len(date_columns) + 2):
        formula = f"=SUM({ws.cell(row=dividend_start_row, column=col_idx).coordinate}:{ws.cell(row=dividend_end_row, column=col_idx).coordinate})"
        total_cell = ws.cell(row=current_row, column=col_idx, value=formula)
        total_cell.number_format = FORMAT_CURRENCY_USD_SIMPLE
        total_cell.font = Font(bold=True, name="Arial", size=12)
        total_cell.border = BOX_BORDER
    
    current_row += 4
    
    # === SECTION 2: PORTFOLIO VALUES ===
    ws.cell(row=current_row, column=1, value="PORTFOLIO VALUE TRACKING")
    ws.cell(row=current_row, column=1).font = SECTION_HEADER_FONT
    current_row += 2
    
    # Portfolio headers
    ws.cell(row=current_row, column=1, value="Account Type")
    ws.cell(row=current_row, column=1).font = HEADER_FONT
    
    for col_idx, date_col in enumerate(port_date_columns, start=2):
        ws.cell(row=current_row, column=col_idx, value=str(date_col))
        ws.cell(row=current_row, column=col_idx).font = HEADER_FONT
        ws.cell(row=current_row, column=col_idx).border = BOX_BORDER
    
    current_row += 1
    
    # Portfolio data rows - copy from Portfolio_History_import
    print("\n📊 Portfolio data accounts:")
    unique_port_accounts = port_data.iloc[:, 0].unique() if len(port_data) > 0 else []
    print(unique_port_accounts)
    
    portfolio_start_row = current_row
    
    for account in unique_port_accounts:
        if pd.isna(account) or account == '':
            continue
            
        account_data = port_data[port_data.iloc[:, 0] == account]
        if len(account_data) == 0:
            continue
        
        ws.cell(row=current_row, column=1, value=str(account))
        ws.cell(row=current_row, column=1).font = NORMAL_FONT
        
        # Copy the actual portfolio values for each date
        for col_idx, date_col in enumerate(port_date_columns, start=2):
            if date_col in account_data.columns:
                value = account_data[date_col].iloc[0] if len(account_data) > 0 else 0
                if pd.isna(value):
                    value = 0
                
                cell = ws.cell(row=current_row, column=col_idx, value=float(value) if value != 0 else 0)
                if value != 0:
                    cell.number_format = FORMAT_CURRENCY_USD_SIMPLE
                cell.font = NORMAL_FONT
                cell.border = BOX_BORDER
        
        current_row += 1
    
    # Total portfolio row
    portfolio_end_row = current_row - 1
    
    ws.cell(row=current_row, column=1, value="TOTAL PORTFOLIO")
    ws.cell(row=current_row, column=1).font = Font(bold=True, name="Arial", size=12)
    
    for col_idx in range(2, len(port_date_columns) + 2):
        formula = f"=SUM({ws.cell(row=portfolio_start_row, column=col_idx).coordinate}:{ws.cell(row=portfolio_end_row, column=col_idx).coordinate})"
        total_cell = ws.cell(row=current_row, column=col_idx, value=formula)
        total_cell.number_format = FORMAT_CURRENCY_USD_SIMPLE
        total_cell.font = Font(bold=True, name="Arial", size=12)
        total_cell.border = BOX_BORDER
    
    # Auto-size columns
    max_cols = max(len(date_columns), len(port_date_columns)) + 2
    for col in range(1, max_cols):
        ws.column_dimensions[ws.cell(row=1, column=col).column_letter].width = 18
    
    # Save workbook
    wb.save(OUTPUT_FILE)
    print(f"\n✅ Rebuilt Estimated Income 2025 sheet with correct historical data")
    print(f"📊 Dividend data: {len(date_columns)} weeks")
    print(f"📊 Portfolio data: {len(port_date_columns)} weeks")

if __name__ == "__main__":
    rebuild_estimated_income_sheet()
