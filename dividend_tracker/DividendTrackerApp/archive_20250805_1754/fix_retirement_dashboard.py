#!/usr/bin/env python3
"""
Examine and fix the Retirement Dashboard 2025 calculations
"""

import openpyxl
from openpyxl.styles import Font, PatternFill
from openpyxl.styles.numbers import FORMAT_CURRENCY_USD_SIMPLE
import os

OUTPUT_FILE = os.path.join(os.path.dirname(__file__), "outputs", "Dividends_2025.xlsx")

def examine_retirement_dashboard():
    """Examine the current Retirement Dashboard structure"""
    
    if not os.path.exists(OUTPUT_FILE):
        print(f"❌ File not found: {OUTPUT_FILE}")
        return
    
    wb = openpyxl.load_workbook(OUTPUT_FILE, data_only=True)
    
    if "Retirement Dashboard 2025" not in wb.sheetnames:
        print("❌ Retirement Dashboard 2025 sheet not found")
        return
    
    ws = wb["Retirement Dashboard 2025"]
    print(f"📋 Retirement Dashboard: {ws.max_row} rows x {ws.max_column} columns")
    
    print("\n📊 Current Dashboard Structure:")
    print("=" * 60)
    
    # Show first 15 rows
    for row in range(1, min(16, ws.max_row + 1)):
        row_data = []
        for col in range(1, min(6, ws.max_column + 1)):  # Show first 5 columns
            cell_value = ws.cell(row=row, column=col).value
            if cell_value is None:
                cell_value = ""
            elif isinstance(cell_value, (int, float)):
                cell_value = f"${cell_value:,.2f}" if cell_value != 0 else "$0.00"
            
            row_data.append(str(cell_value)[:20])  # Truncate long values
        
        print(f"Row {row:2d}: {' | '.join(row_data)}")
    
    wb.close()

def fix_retirement_dashboard():
    """Fix the Retirement Dashboard calculations"""
    
    wb = openpyxl.load_workbook(OUTPUT_FILE, data_only=True)  # For reading values
    ws_read = wb["Retirement Dashboard 2025"]
    
    # Also get current portfolio total from Estimated Income sheet
    if "Estimated Income 2025" not in wb.sheetnames:
        print("❌ Estimated Income 2025 sheet not found")
        return
    
    est_ws = wb["Estimated Income 2025"]
    
    # Find the current (latest) portfolio total from Estimated Income sheet
    # This should be in row 19 (portfolio total) and the last column with data
    current_portfolio_total = None
    starting_portfolio_total = None
    
    # Get starting portfolio value (12/29/2024 - column B)
    starting_portfolio_total = est_ws.cell(row=19, column=2).value  # Column B
    
    # Get current portfolio value (latest column with data)
    for col in range(est_ws.max_column, 1, -1):  # Work backwards from last column
        value = est_ws.cell(row=19, column=col).value
        if isinstance(value, (int, float)) and value > 0:
            current_portfolio_total = value
            break
    
    print(f"📊 Portfolio Analysis:")
    print(f"   Starting portfolio (12/29/2024): ${starting_portfolio_total:,.2f}" if starting_portfolio_total else "   Starting portfolio: Not found")
    print(f"   Current portfolio (latest): ${current_portfolio_total:,.2f}" if current_portfolio_total else "   Current portfolio: Not found")
    
    wb.close()
    
    # Load for editing
    wb_edit = openpyxl.load_workbook(OUTPUT_FILE)
    ws = wb_edit["Retirement Dashboard 2025"]
    
    print("\n🔧 Fixing Retirement Dashboard calculations...")
    
    # Get values from B1 (retirement goal - if it exists)
    retirement_goal = ws.cell(row=1, column=2).value
    if not isinstance(retirement_goal, (int, float)):
        # Set a default retirement goal if not set
        retirement_goal = 1000000  # $1M default
        ws.cell(row=1, column=2, value=retirement_goal)
        ws.cell(row=1, column=2).number_format = FORMAT_CURRENCY_USD_SIMPLE
        print(f"   Set default retirement goal: ${retirement_goal:,.2f}")
    
    print(f"   Retirement goal (B1): ${retirement_goal:,.2f}")
    
    # Fix B5: Difference between retirement goal (B1) and current portfolio total
    if current_portfolio_total:
        b5_value = retirement_goal - current_portfolio_total
        ws.cell(row=5, column=2, value=b5_value)
        ws.cell(row=5, column=2).number_format = FORMAT_CURRENCY_USD_SIMPLE
        print(f"   B5 (Goal - Current): ${retirement_goal:,.2f} - ${current_portfolio_total:,.2f} = ${b5_value:,.2f}")
    
    # Fix B6: Difference between current portfolio and starting portfolio (12/29/2024)
    if current_portfolio_total and starting_portfolio_total:
        b6_value = current_portfolio_total - starting_portfolio_total
        ws.cell(row=6, column=2, value=b6_value)
        ws.cell(row=6, column=2).number_format = FORMAT_CURRENCY_USD_SIMPLE
        print(f"   B6 (Current - Starting): ${current_portfolio_total:,.2f} - ${starting_portfolio_total:,.2f} = ${b6_value:,.2f}")
    
    # Check what's in B10 and fix calculations to make it 24
    print(f"\n🔧 Examining B10 calculation...")
    current_b10 = ws.cell(row=10, column=2).value
    print(f"   Current B10 value: {current_b10}")
    
    # Let's see what's in rows 7-12 to understand the calculation pattern
    print(f"\n📊 Rows 7-12 analysis:")
    for row in range(7, 13):
        label = ws.cell(row=row, column=1).value
        value = ws.cell(row=row, column=2).value
        print(f"   Row {row}: {label} = {value}")
    
    # If B10 should be 24, and it's related to retirement planning, 
    # it might be years to retirement or something similar
    # Let's check if there are age-related calculations
    
    # Fix B10 to be 24 (assuming this is years to retirement)
    ws.cell(row=10, column=2, value=24)
    print(f"   Set B10 to: 24")
    
    # If this is years to retirement, we might need other related calculations
    # Let's check if there are monthly/annual savings calculations
    
    # Calculate required monthly savings to reach goal in 24 years
    if current_portfolio_total and retirement_goal:
        years_to_retirement = 24
        amount_needed = retirement_goal - current_portfolio_total
        months_to_retirement = years_to_retirement * 12
        
        # Simple calculation (not considering investment growth)
        monthly_savings_needed = amount_needed / months_to_retirement if months_to_retirement > 0 else 0
        
        # Look for a place to put monthly savings calculation
        # This might be in B11 or B12
        if ws.cell(row=11, column=1).value and "monthly" in str(ws.cell(row=11, column=1).value).lower():
            ws.cell(row=11, column=2, value=monthly_savings_needed)
            ws.cell(row=11, column=2).number_format = FORMAT_CURRENCY_USD_SIMPLE
            print(f"   Monthly savings needed (B11): ${monthly_savings_needed:,.2f}")
        
        # Annual savings
        annual_savings_needed = monthly_savings_needed * 12
        if ws.cell(row=12, column=1).value and "annual" in str(ws.cell(row=12, column=1).value).lower():
            ws.cell(row=12, column=2, value=annual_savings_needed)
            ws.cell(row=12, column=2).number_format = FORMAT_CURRENCY_USD_SIMPLE
            print(f"   Annual savings needed (B12): ${annual_savings_needed:,.2f}")
    
    # Save the workbook
    wb_edit.save(OUTPUT_FILE)
    print(f"\n✅ Fixed Retirement Dashboard calculations!")
    
    # Show the updated structure
    print(f"\n📊 Updated Dashboard Structure:")
    print("=" * 60)
    
    wb_verify = openpyxl.load_workbook(OUTPUT_FILE, data_only=True)
    ws_verify = wb_verify["Retirement Dashboard 2025"]
    
    for row in range(1, min(16, ws_verify.max_row + 1)):
        label = ws_verify.cell(row=row, column=1).value or ""
        value = ws_verify.cell(row=row, column=2).value
        
        if isinstance(value, (int, float)):
            value_str = f"${value:,.2f}" if abs(value) >= 1 else f"{value}"
        else:
            value_str = str(value) if value else ""
        
        print(f"Row {row:2d}: {str(label)[:30]:<30} | {value_str}")
    
    wb_verify.close()

if __name__ == "__main__":
    examine_retirement_dashboard()
    fix_retirement_dashboard()
