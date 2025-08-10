#!/usr/bin/env python3
"""
Fix formatting for Days and Months since retirement (remove currency formatting)
"""

import openpyxl
from openpyxl.styles.numbers import FORMAT_NUMBER
import os

OUTPUT_FILE = os.path.join(os.path.dirname(__file__), "outputs", "Dividends_2025.xlsx")

def fix_retirement_time_formatting():
    """Fix the formatting for days and months since retirement"""
    
    wb = openpyxl.load_workbook(OUTPUT_FILE)
    ws = wb["Retirement Dashboard 2025"]
    
    print("🔧 Fixing retirement time formatting...")
    
    # Fix B10: Days Since Retirement (should be just a number)
    days_cell = ws.cell(row=10, column=2)
    current_days = days_cell.value
    print(f"   B10 (Days): Current value = {current_days}")
    
    # Remove currency formatting and set as regular number
    days_cell.number_format = FORMAT_NUMBER
    print(f"   B10: Fixed formatting (was currency, now number)")
    
    # Fix B11: Months Since Retirement (should be just a number with 1 decimal)
    months_cell = ws.cell(row=11, column=2)
    current_months = months_cell.value
    print(f"   B11 (Months): Current value = {current_months}")
    
    # Set as number with 1 decimal place
    months_cell.number_format = "0.0"
    print(f"   B11: Fixed formatting (was currency, now number with 1 decimal)")
    
    # Also fix B12: Average Monthly Portfolio Growth should stay as currency
    growth_cell = ws.cell(row=12, column=2)
    print(f"   B12 (Avg Monthly Growth): Keeping currency format = ${growth_cell.value:,.2f}" if growth_cell.value else "   B12: No value")
    
    # Save the workbook
    wb.save(OUTPUT_FILE)
    print("✅ Fixed retirement time formatting!")
    
    # Verify the fixes
    print("\n📊 Verification:")
    wb_verify = openpyxl.load_workbook(OUTPUT_FILE, data_only=True)
    ws_verify = wb_verify["Retirement Dashboard 2025"]
    
    days_value = ws_verify.cell(row=10, column=2).value
    months_value = ws_verify.cell(row=11, column=2).value
    growth_value = ws_verify.cell(row=12, column=2).value
    
    print(f"   Days Since Retirement: {days_value} (no currency symbol)")
    print(f"   Months Since Retirement: {months_value} (1 decimal place)")
    print(f"   Avg Monthly Growth: ${growth_value:,.2f} (currency format)")
    
    wb_verify.close()

if __name__ == "__main__":
    fix_retirement_time_formatting()
