#!/usr/bin/env python3
"""
Simple check of what's in Dividends_2025.xlsx
"""

import openpyxl
import os

file_path = 'outputs/Dividends_2025.xlsx'

if os.path.exists(file_path):
    print(f"✅ File exists: {file_path}")
    
    # Get file size and modification time
    import os
    stat = os.stat(file_path)
    print(f"📏 File size: {stat.st_size:,} bytes")
    print(f"🕒 Last modified: {stat.st_mtime}")
    
    try:
        wb = openpyxl.load_workbook(file_path)
        print(f"📊 Sheets: {wb.sheetnames}")
        
        if 'Ticker Analysis 2025' in wb.sheetnames:
            ws = wb['Ticker Analysis 2025']
            print(f"📈 Ticker Analysis 2025: {ws.max_row} rows x {ws.max_column} columns")
            
            # Check first 10 rows
            print("\nFirst 10 rows:")
            for row in range(1, min(11, ws.max_row + 1)):
                col1 = ws.cell(row, 1).value
                col2 = ws.cell(row, 2).value
                col3 = ws.cell(row, 3).value
                print(f"  Row {row}: {col1} | {col2} | {col3}")
        else:
            print("❌ No Ticker Analysis 2025 sheet found")
            
    except Exception as e:
        print(f"❌ Error opening file: {e}")
else:
    print(f"❌ File not found: {file_path}")
