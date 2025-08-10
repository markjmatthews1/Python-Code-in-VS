#!/usr/bin/env python3
"""
Verify that the total formulas are calculating correctly
"""

import openpyxl
import os

OUTPUT_FILE = os.path.join(os.path.dirname(__file__), "outputs", "Dividends_2025.xlsx")

def verify_totals():
    """Verify that the total rows are calculating correctly"""
    
    # Load with calculated values
    wb = openpyxl.load_workbook(OUTPUT_FILE, data_only=True)
    ws = wb["Estimated Income 2025"]
    
    print("🔍 Verifying total calculations...")
    print("=" * 60)
    
    # Check a few sample columns
    sample_columns = [2, 3, 4, 10, 20, 30]  # Sample some columns
    
    for col in sample_columns:
        if col <= ws.max_column:
            col_letter = ws.cell(row=1, column=col).column_letter
            
            # Get dividend values for rows 4-7
            div_values = []
            for row in range(4, 8):
                value = ws.cell(row=row, column=col).value
                if isinstance(value, (int, float)):
                    div_values.append(value)
                else:
                    div_values.append(0)
            
            # Get dividend total
            div_total = ws.cell(row=8, column=col).value
            expected_div_total = sum(div_values)
            
            # Get portfolio values for rows 14-18
            port_values = []
            for row in range(14, 19):
                value = ws.cell(row=row, column=col).value
                if isinstance(value, (int, float)):
                    port_values.append(value)
                else:
                    port_values.append(0)
            
            # Get portfolio total
            port_total = ws.cell(row=19, column=col).value
            expected_port_total = sum(port_values)
            
            print(f"\n📊 Column {col_letter}:")
            print(f"   Dividend values: {[f'${v:,.2f}' if v != 0 else '$0.00' for v in div_values]}")
            print(f"   Expected div total: ${expected_div_total:,.2f}")
            print(f"   Actual div total:   ${div_total:,.2f}" if isinstance(div_total, (int, float)) else f"   Actual div total:   {div_total}")
            print(f"   ✅ Dividend total correct" if abs(expected_div_total - (div_total or 0)) < 0.01 else "   ❌ Dividend total incorrect")
            
            print(f"   Portfolio values: {[f'${v:,.2f}' if v != 0 else '$0.00' for v in port_values]}")
            print(f"   Expected port total: ${expected_port_total:,.2f}")
            print(f"   Actual port total:   ${port_total:,.2f}" if isinstance(port_total, (int, float)) else f"   Actual port total:   {port_total}")
            print(f"   ✅ Portfolio total correct" if abs(expected_port_total - (port_total or 0)) < 0.01 else "   ❌ Portfolio total incorrect")
    
    # Show the date headers
    print(f"\n📅 Date headers for sampled columns:")
    for col in sample_columns:
        if col <= ws.max_column:
            col_letter = ws.cell(row=1, column=col).column_letter
            div_date = ws.cell(row=3, column=col).value  # Dividend section date
            port_date = ws.cell(row=13, column=col).value  # Portfolio section date
            print(f"   Column {col_letter}: Div='{div_date}', Port='{port_date}'")
    
    wb.close()
    print(f"\n✅ Verification complete!")

if __name__ == "__main__":
    verify_totals()
