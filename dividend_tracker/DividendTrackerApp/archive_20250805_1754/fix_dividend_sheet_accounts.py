#!/usr/bin/env python3
"""
Fix Dividend Sheet Accounts and Clean Yield Data
Author: Copilot
Date: July 27, 2025
Purpose: Fix messed up yield data and correct account names with proper ticker quantities
"""

import openpyxl
from openpyxl.styles import PatternFill, Font
import pandas as pd
from datetime import datetime
import sys
import os

# Add modules to path
sys.path.append(r'c:\Python_Projects\DividendTrackerApp\modules')

class DividendSheetFixer:
    """Class to fix dividend sheet accounts and clean up yield data"""
    
    def __init__(self, excel_path):
        self.excel_path = excel_path
        self.wb = openpyxl.load_workbook(excel_path)
        
        # Account mapping for proper names
        self.account_names = {
            'E*TRADE IRA': 'E*TRADE IRA',
            'E*TRADE Taxable': 'E*TRADE Taxable', 
            'Schwab IRA': 'Schwab IRA',
            'Schwab Individual': 'Schwab Individual'
        }
        
        # Load ticker data from source files
        self.load_ticker_data()
    
    def load_ticker_data(self):
        """Load ticker data from E*TRADE Excel files to get proper account assignments"""
        self.ticker_accounts = {}
        
        try:
            # Load E*TRADE IRA data
            ira_df = pd.read_excel(r'c:\Python_Projects\DividendTrackerApp\data\etrade_ira_estimates_2025.xlsx', header=None)
            
            # Look for ticker data - starts after "SEARCH CRITERIA" row
            start_row = None
            for idx in range(len(ira_df)):
                if pd.notna(ira_df.iloc[idx, 0]) and 'SEARCH CRITERIA' in str(ira_df.iloc[idx, 0]):
                    start_row = idx + 1
                    break
            
            if start_row:
                for idx in range(start_row, len(ira_df)):
                    row = ira_df.iloc[idx]
                    symbol = str(row.iloc[0]).strip() if pd.notna(row.iloc[0]) else ""
                    
                    # Skip empty rows and invalid symbols
                    if not symbol or symbol == 'nan' or len(symbol) < 1:
                        continue
                    
                    # Stop at summary rows
                    if symbol.upper() in ['TOTAL', 'TOTALS', 'SUMMARY', '']:
                        break
                    
                    # Extract quantity from available columns (try different positions)
                    qty = 0
                    for col_idx in range(1, min(10, len(row))):
                        try:
                            if pd.notna(row.iloc[col_idx]) and isinstance(row.iloc[col_idx], (int, float)):
                                if row.iloc[col_idx] > 0:
                                    qty = row.iloc[col_idx]
                                    break
                        except:
                            continue
                    
                    if qty > 0:
                        self.ticker_accounts[symbol] = {'account': 'E*TRADE IRA', 'qty': qty}
                        
            print(f"✅ Loaded {len([k for k, v in self.ticker_accounts.items() if v['account'] == 'E*TRADE IRA'])} E*TRADE IRA tickers")
            
        except Exception as e:
            print(f"⚠️ Error loading E*TRADE IRA data: {e}")
        
        try:
            # Load E*TRADE Taxable data
            taxable_df = pd.read_excel(r'c:\Python_Projects\DividendTrackerApp\data\etrade_taxable_estimates_2025.xlsx', header=None)
            
            # Look for ticker data - starts after "SEARCH CRITERIA" row
            start_row = None
            for idx in range(len(taxable_df)):
                if pd.notna(taxable_df.iloc[idx, 0]) and 'SEARCH CRITERIA' in str(taxable_df.iloc[idx, 0]):
                    start_row = idx + 1
                    break
            
            if start_row:
                for idx in range(start_row, len(taxable_df)):
                    row = taxable_df.iloc[idx]
                    symbol = str(row.iloc[0]).strip() if pd.notna(row.iloc[0]) else ""
                    
                    # Skip empty rows and invalid symbols
                    if not symbol or symbol == 'nan' or len(symbol) < 1:
                        continue
                    
                    # Stop at summary rows
                    if symbol.upper() in ['TOTAL', 'TOTALS', 'SUMMARY', '']:
                        break
                    
                    # Extract quantity from available columns (try different positions)
                    qty = 0
                    for col_idx in range(1, min(10, len(row))):
                        try:
                            if pd.notna(row.iloc[col_idx]) and isinstance(row.iloc[col_idx], (int, float)):
                                if row.iloc[col_idx] > 0:
                                    qty = row.iloc[col_idx]
                                    break
                        except:
                            continue
                    
                    if qty > 0:
                        self.ticker_accounts[symbol] = {'account': 'E*TRADE Taxable', 'qty': qty}
                        
            print(f"✅ Loaded {len([k for k, v in self.ticker_accounts.items() if v['account'] == 'E*TRADE Taxable'])} E*TRADE Taxable tickers")
            
        except Exception as e:
            print(f"⚠️ Error loading E*TRADE Taxable data: {e}")
    
    def fix_ticker_analysis_sheet(self):
        """Fix the Ticker Analysis 2025 sheet with proper account names and data"""
        if 'Ticker Analysis 2025' not in self.wb.sheetnames:
            print("❌ Ticker Analysis 2025 sheet not found")
            return False
            
        ws = self.wb['Ticker Analysis 2025']
        
        # Clear existing data (but keep headers)
        print("🧹 Clearing existing ticker data...")
        
        # Get the header row (should be row 1)
        headers = [cell.value for cell in ws[1]]
        
        # Clear all data rows (keep headers)
        for row in range(2, ws.max_row + 1):
            for col in range(1, ws.max_column + 1):
                ws.cell(row=row, column=col).value = None
                ws.cell(row=row, column=col).fill = PatternFill()
        
        # Set up proper headers
        headers = [
            'Account', 'Ticker', 'Quantity', 'Current Price', 'Market Value', 
            'Annual Dividend Rate', 'Annual Dividend Income', 'Current Div Yield %', 
            'Quarterly Payout', 'Yield Status', 'Monthly Dividend', 'Reinvest Status',
            'Last Ex-Date', 'Next Ex-Date', 'Beginning Dividend Yield'
        ]
        
        # Write headers
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col)
            cell.value = header
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        
        # Add ticker data with proper account assignments
        current_row = 2
        
        print("📊 Adding ticker data with proper account assignments...")
        
        for ticker, data in self.ticker_accounts.items():
            # Account name
            ws.cell(row=current_row, column=1).value = data['account']
            
            # Ticker symbol
            ws.cell(row=current_row, column=2).value = ticker
            
            # Quantity
            ws.cell(row=current_row, column=3).value = data['qty']
            
            # Placeholder values for other columns (to be filled by yield update)
            ws.cell(row=current_row, column=4).value = 0  # Current Price
            ws.cell(row=current_row, column=5).value = 0  # Market Value
            ws.cell(row=current_row, column=6).value = 0  # Annual Dividend Rate
            ws.cell(row=current_row, column=7).value = 0  # Annual Dividend Income
            ws.cell(row=current_row, column=8).value = 0  # Current Div Yield %
            ws.cell(row=current_row, column=9).value = 0  # Quarterly Payout
            ws.cell(row=current_row, column=10).value = "NEW"  # Yield Status
            ws.cell(row=current_row, column=11).value = 0  # Monthly Dividend
            ws.cell(row=current_row, column=12).value = "No"  # Reinvest Status
            ws.cell(row=current_row, column=13).value = ""  # Last Ex-Date
            ws.cell(row=current_row, column=14).value = ""  # Next Ex-Date
            ws.cell(row=current_row, column=15).value = 0  # Beginning Dividend Yield
            
            current_row += 1
        
        # Add totals row
        ws.cell(row=current_row, column=1).value = "TOTALS"
        ws.cell(row=current_row, column=3).value = f"=SUM(C2:C{current_row-1})"
        ws.cell(row=current_row, column=5).value = f"=SUM(E2:E{current_row-1})"
        ws.cell(row=current_row, column=7).value = f"=SUM(G2:G{current_row-1})"
        ws.cell(row=current_row, column=11).value = f"=SUM(K2:K{current_row-1})"
        
        # Format totals row
        for col in range(1, 16):
            cell = ws.cell(row=current_row, column=col)
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color="D3D3D3", end_color="D3D3D3", fill_type="solid")
        
        print(f"✅ Added {current_row-2} tickers with proper account assignments")
        return True
    
    def clean_yield_columns(self):
        """Clean up any messed up yield data columns"""
        if 'Ticker Analysis 2025' not in self.wb.sheetnames:
            return False
            
        ws = self.wb['Ticker Analysis 2025']
        
        # Look for any extra yield columns beyond column 15 and remove them
        max_col = ws.max_column
        if max_col > 15:
            print(f"🧹 Cleaning up extra yield columns (found {max_col} columns, should be 15)")
            
            # Delete columns beyond 15
            for col in range(max_col, 15, -1):
                ws.delete_cols(col)
                
            print(f"✅ Cleaned up extra columns, now have {ws.max_column} columns")
        
        return True
    
    def save_fixed_sheet(self):
        """Save the fixed Excel sheet"""
        try:
            self.wb.save(self.excel_path)
            print(f"✅ Successfully saved fixed dividend sheet to {self.excel_path}")
            return True
        except Exception as e:
            print(f"❌ Error saving file: {e}")
            return False
    
    def run_fix(self):
        """Run the complete fix process"""
        print("🔧 Starting dividend sheet fix process...")
        print("=" * 60)
        
        # Step 1: Clean up yield columns
        print("\n1️⃣ Cleaning up messed up yield data...")
        if self.clean_yield_columns():
            print("✅ Yield columns cleaned")
        else:
            print("❌ Failed to clean yield columns")
        
        # Step 2: Fix ticker analysis sheet
        print("\n2️⃣ Fixing ticker analysis with proper account names...")
        if self.fix_ticker_analysis_sheet():
            print("✅ Ticker analysis sheet fixed")
        else:
            print("❌ Failed to fix ticker analysis sheet")
            return False
        
        # Step 3: Save the file
        print("\n3️⃣ Saving fixed dividend sheet...")
        if self.save_fixed_sheet():
            print("✅ File saved successfully")
        else:
            print("❌ Failed to save file")
            return False
        
        print("\n🎉 Dividend sheet fix completed successfully!")
        print("=" * 60)
        
        # Print summary
        print(f"\n📊 Summary:")
        print(f"   • Total tickers: {len(self.ticker_accounts)}")
        
        account_counts = {}
        for ticker, data in self.ticker_accounts.items():
            account = data['account']
            if account not in account_counts:
                account_counts[account] = 0
            account_counts[account] += 1
        
        for account, count in account_counts.items():
            print(f"   • {account}: {count} tickers")
        
        return True

def main():
    """Main function to run the dividend sheet fix"""
    excel_path = r'c:\Python_Projects\DividendTrackerApp\outputs\Dividends_2025.xlsx'
    
    if not os.path.exists(excel_path):
        print(f"❌ Excel file not found: {excel_path}")
        return False
    
    # Create fixer and run
    fixer = DividendSheetFixer(excel_path)
    return fixer.run_fix()

if __name__ == "__main__":
    success = main()
    if success:
        print("\n✅ Ready for new yield report update!")
    else:
        print("\n❌ Fix process failed!")
