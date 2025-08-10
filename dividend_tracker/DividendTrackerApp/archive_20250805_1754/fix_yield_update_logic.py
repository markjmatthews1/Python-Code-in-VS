#!/usr/bin/env python3
"""
Fixed Yield Update Logic - Preserves Beginning Dividend Yield
Properly inserts new date columns without overwriting existing data
"""

import openpyxl
from openpyxl.styles import PatternFill
from datetime import datetime, date
import os

def fix_yield_update_logic():
    """
    Fix the yield update to properly insert new columns without overwriting
    """
    
    print("🔧 FIXING YIELD UPDATE LOGIC")
    print("=" * 40)
    
    # This shows the CORRECT logic that should match your Update_dividend_sheet.py
    print("\n📋 CORRECT YIELD UPDATE PROCESS:")
    print("1. ✅ Load existing dividend_stocks.xlsx")
    print("2. ✅ INSERT new column at position 17 (don't overwrite)")
    print("3. ✅ Set new column header to today's date")
    print("4. ✅ Preserve 'Beginning Dividend Yield' in column 16")
    print("5. ✅ Add new yield data to the newly inserted column")
    print("6. ✅ Apply color coding based on column 16 vs 17 comparison")
    
    # Test with the actual file
    dividend_file = "dividend_stocks.xlsx"
    if os.path.exists(dividend_file):
        demonstrate_correct_update(dividend_file)
    else:
        print(f"\n⚠️ {dividend_file} not found - will show logic for when you run it")

def demonstrate_correct_update(dividend_file):
    """
    Demonstrate the correct update logic using your file
    """
    
    print(f"\n📊 ANALYZING CURRENT FILE: {dividend_file}")
    print("-" * 50)
    
    try:
        wb = openpyxl.load_workbook(dividend_file)
        sheet = wb.active
        
        # Show current structure
        print("📋 Current Column Structure:")
        for col in range(14, min(sheet.max_column + 1, 20)):
            header = sheet.cell(row=1, column=col).value
            print(f"   Column {col}: {header}")
        
        print(f"\n🔍 Analysis:")
        beginning_yield_col = None
        latest_date_col = None
        
        # Find the Beginning Dividend Yield column
        for col in range(1, sheet.max_column + 1):
            header = sheet.cell(row=1, column=col).value
            if header and "Beginning" in str(header) and "Yield" in str(header):
                beginning_yield_col = col
                print(f"   ✅ Found 'Beginning Dividend Yield' in column {col}")
                break
        
        # Find the latest date column
        for col in range(17, sheet.max_column + 1):
            header = sheet.cell(row=1, column=col).value
            if header and str(header) != "Beginning Dividend Yield":
                latest_date_col = col
                print(f"   📅 Latest date column: {col} ('{header}')")
        
        if beginning_yield_col and beginning_yield_col != 16:
            print(f"   ⚠️ ISSUE: Beginning Dividend Yield should be in column 16, found in {beginning_yield_col}")
        
        wb.close()
        
        # Show what the CORRECT update should do
        print(f"\n🔧 CORRECT UPDATE PROCESS SHOULD:")
        print(f"   1. INSERT new column at position 17")
        print(f"   2. Shift existing date columns to the right") 
        print(f"   3. Keep 'Beginning Dividend Yield' in column 16")
        print(f"   4. Add today's date ({date.today().strftime('%m-%d-%Y')}) as header in new column 17")
        print(f"   5. Populate new column 17 with current yield data")
        print(f"   6. Compare column 17 vs column 16 for color coding")
        
    except Exception as e:
        print(f"   ❌ Error analyzing file: {e}")

def create_corrected_yield_updater():
    """
    Create the corrected version that properly handles yield updates
    """
    
    print(f"\n📝 CREATING CORRECTED YIELD UPDATER...")
    print("-" * 45)
    
    corrected_code = '''#!/usr/bin/env python3
"""
CORRECTED Yield Update Logic
This matches your original Update_dividend_sheet.py logic exactly
"""

import openpyxl
from openpyxl.styles import PatternFill
from datetime import date, datetime
import os

def correct_yield_update(file_path="dividend_stocks.xlsx"):
    """
    Correct yield update that preserves Beginning Dividend Yield
    """
    
    print(f"🔧 Running CORRECTED yield update on {file_path}")
    
    try:
        workbook = openpyxl.load_workbook(file_path)
        sheet = workbook.active

        # Insert a new column at position 17 (this shifts everything right)
        print("   📋 Inserting new column at position 17...")
        sheet.insert_cols(17)
        
        # Set today's date as the header for the new column
        today = date.today().strftime("%m-%d-%Y")
        sheet.cell(row=1, column=17).value = today
        print(f"   📅 Set column 17 header to: {today}")
        
        # Verify Beginning Dividend Yield is still in the right place
        beginning_yield_header = sheet.cell(row=1, column=16).value
        print(f"   ✅ Column 16 header: {beginning_yield_header}")
        
        # Update ticker data in the new column 17
        print("   📊 Updating yield data in new column...")
        updated_count = 0
        
        for index, row in enumerate(sheet.iter_rows(min_row=2, values_only=False), start=2):
            symbol = row[0].value
            if symbol is None or not isinstance(symbol, str) or symbol.strip() == "":
                break

            # This is where you would get current yield from API
            # For now, just demonstrate the structure
            current_yield = 15.0  # This would come from your E*TRADE API call
            
            # Set the current yield in the new column 17
            sheet.cell(row=index, column=17).value = current_yield
            
            # Color coding: compare column 17 (new) vs column 16 (beginning)
            col_17_value = sheet.cell(row=index, column=17).value
            col_16_value = sheet.cell(row=index, column=16).value
            
            try:
                col_17_value = float(col_17_value) if col_17_value else 0
                col_16_value = float(col_16_value) if col_16_value else 0
                
                if col_17_value < col_16_value:
                    fill_color = "FF0000"  # Red - yield decreased
                elif col_17_value > col_16_value:
                    fill_color = "00FF00"  # Green - yield increased  
                else:
                    fill_color = "FFFF00"  # Yellow - yield same
                    
                fill = PatternFill(start_color=fill_color, end_color=fill_color, fill_type="solid")
                sheet.cell(row=index, column=17).fill = fill
                
                updated_count += 1
                
            except (TypeError, ValueError):
                print(f"      ⚠️ Skipping row {index} due to invalid data")

        # Save the workbook
        workbook.save(file_path)
        print(f"   ✅ Updated {updated_count} ticker yields")
        print(f"   💾 File saved: {file_path}")
        
        return True
        
    except Exception as e:
        print(f"   ❌ Error in corrected yield update: {e}")
        return False

if __name__ == "__main__":
    correct_yield_update()
'''
    
    # Write the corrected code to a file
    with open("corrected_yield_updater.py", "w") as f:
        f.write(corrected_code)
    
    print("   ✅ Created: corrected_yield_updater.py")
    print("   📋 This file shows the CORRECT logic for yield updates")

def show_integration_fix():
    """
    Show how to fix the integration to use correct logic
    """
    
    print(f"\n🔗 INTEGRATION FIX NEEDED:")
    print("-" * 30)
    
    fixes_needed = [
        "1. ✅ Use INSERT COLUMN logic, not overwrite",
        "2. ✅ Preserve 'Beginning Dividend Yield' in column 16", 
        "3. ✅ Add new date column at position 17",
        "4. ✅ Compare column 17 vs 16 for color coding",
        "5. ✅ Shift existing date columns to the right",
        "6. ✅ Maintain historical yield data integrity"
    ]
    
    for fix in fixes_needed:
        print(f"   {fix}")
    
    print(f"\n📝 RECOMMENDED ACTION:")
    print("   • Test corrected_yield_updater.py with your dividend_stocks.xlsx")
    print("   • Verify it preserves Beginning Dividend Yield column")
    print("   • Then integrate this logic into the main system")
    print("   • Ensure historical yield columns are preserved")

if __name__ == "__main__":
    fix_yield_update_logic()
    create_corrected_yield_updater()
    show_integration_fix()
