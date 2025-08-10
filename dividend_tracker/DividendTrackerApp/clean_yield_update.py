#!/usr/bin/env python3
"""
Clean Yield Update Without Adding Data Week
Author: Copilot
Date: July 27, 2025
Purpose: Update yield data without adding another dividend tracker data week
"""

import sys
sys.path.append(r'c:\Python_Projects\DividendTrackerApp\modules')

from etrade_auth import get_etrade_session
import openpyxl
from openpyxl.styles import PatternFill
from datetime import date, datetime
import time
import json

class CleanYieldUpdater:
    """Class to update yield data without adding duplicate data weeks"""
    
    def __init__(self, excel_path):
        self.excel_path = excel_path
        self.wb = openpyxl.load_workbook(excel_path)
        
        # Get E*TRADE session
        self.session, self.base_url = get_etrade_session()
        print("✅ Got E*TRADE session")
        
    def update_ticker_yields(self):
        """Update only the yield data for existing tickers without adding columns"""
        if 'Ticker Analysis 2025' not in self.wb.sheetnames:
            print("❌ Ticker Analysis 2025 sheet not found")
            return False
            
        ws = self.wb['Ticker Analysis 2025']
        today_str = datetime.now().strftime('%m/%d/%Y')
        
        print(f"📊 Updating yield data for {today_str}...")
        
        # Process each ticker row
        updated_count = 0
        for row in range(2, ws.max_row + 1):
            ticker_cell = ws.cell(row=row, column=2)  # Column B = Ticker
            
            if not ticker_cell.value or ticker_cell.value == "TOTALS":
                continue
                
            ticker = str(ticker_cell.value).strip()
            if not ticker:
                continue
                
            try:
                # Get current market data using direct API call
                print(f"  Fetching data for {ticker}...")
                quote_url = f"{self.base_url}/v1/market/quote/{ticker}.json"
                quote_response = self.session.get(quote_url)
                
                if quote_response.status_code == 200:
                    quote_data = quote_response.json()
                    
                    if 'quoteResponse' in quote_data and quote_data['quoteResponse']:
                        quote_info = quote_data['quoteResponse'][0]
                        
                        # Extract key data
                        current_price = float(quote_info.get('lastTrade', 0))
                        annual_dividend = float(quote_info.get('annualDividend', 0))
                        dividend_yield = float(quote_info.get('divYield', 0))
                        
                        # Get quantity from sheet
                        qty = ws.cell(row=row, column=3).value or 0
                        
                        # Calculate values
                        market_value = current_price * qty
                        annual_income = annual_dividend * qty
                        quarterly_payout = annual_dividend / 4
                        monthly_dividend = annual_income / 12
                        
                        # Update the sheet data (overwrite, don't add columns)
                        ws.cell(row=row, column=4).value = current_price  # Current Price
                        ws.cell(row=row, column=5).value = market_value   # Market Value
                        ws.cell(row=row, column=6).value = annual_dividend  # Annual Dividend Rate
                        ws.cell(row=row, column=7).value = annual_income  # Annual Dividend Income
                        ws.cell(row=row, column=8).value = dividend_yield  # Current Div Yield %
                        ws.cell(row=row, column=9).value = quarterly_payout  # Quarterly Payout
                        ws.cell(row=row, column=11).value = monthly_dividend  # Monthly Dividend
                        
                        # Compare with beginning yield (column 15) for status
                        beginning_yield = ws.cell(row=row, column=15).value or 0
                        
                        if beginning_yield == 0:
                            # First time - set beginning yield and status as NEW
                            ws.cell(row=row, column=15).value = dividend_yield
                            ws.cell(row=row, column=10).value = "NEW"
                            ws.cell(row=row, column=10).fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
                        else:
                            # Compare yields for status
                            if dividend_yield > beginning_yield:
                                ws.cell(row=row, column=10).value = "INCREASED"
                                ws.cell(row=row, column=10).fill = PatternFill(start_color="00FF00", end_color="00FF00", fill_type="solid")
                            elif dividend_yield < beginning_yield:
                                ws.cell(row=row, column=10).value = "DECREASED"
                                ws.cell(row=row, column=10).fill = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")
                            else:
                                ws.cell(row=row, column=10).value = "UNCHANGED"
                                ws.cell(row=row, column=10).fill = PatternFill()
                        
                        updated_count += 1
                        print(f"    ✅ {ticker}: ${current_price:.2f}, {dividend_yield:.2f}% yield")
                    else:
                        print(f"    ⚠️ {ticker}: No quote data in response")
                else:
                    print(f"    ❌ {ticker}: API error {quote_response.status_code}")
                    
                # Rate limiting
                time.sleep(0.5)
                    
            except Exception as e:
                print(f"    ❌ {ticker}: Error - {e}")
                continue
        
        print(f"\n✅ Updated {updated_count} tickers with current yield data")
        return True
    
    def save_updated_sheet(self):
        """Save the updated Excel sheet"""
        try:
            self.wb.save(self.excel_path)
            print(f"✅ Successfully saved updated yield data to {self.excel_path}")
            return True
        except Exception as e:
            print(f"❌ Error saving file: {e}")
            return False
    
    def run_clean_update(self):
        """Run the clean yield update process"""
        print("🔄 Starting clean yield update (no new data week)...")
        print("=" * 60)
        
        # Step 1: Update ticker yields
        print("\n1️⃣ Updating ticker yield data...")
        if self.update_ticker_yields():
            print("✅ Yield data updated")
        else:
            print("❌ Failed to update yield data")
            return False
        
        # Step 2: Save the file
        print("\n2️⃣ Saving updated yield data...")
        if self.save_updated_sheet():
            print("✅ File saved successfully")
        else:
            print("❌ Failed to save file")
            return False
        
        print("\n🎉 Clean yield update completed successfully!")
        print("=" * 60)
        print("📝 Note: No new data week was added - only yield data was refreshed")
        
        return True

def main():
    """Main function to run the clean yield update"""
    excel_path = r'c:\Python_Projects\DividendTrackerApp\outputs\Dividends_2025.xlsx'
    
    # Create updater and run
    updater = CleanYieldUpdater(excel_path)
    return updater.run_clean_update()

if __name__ == "__main__":
    success = main()
    if success:
        print("\n✅ Yield data successfully updated without adding data week!")
    else:
        print("\n❌ Clean yield update failed!")
