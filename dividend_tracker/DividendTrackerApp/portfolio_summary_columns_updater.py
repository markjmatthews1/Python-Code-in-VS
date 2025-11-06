#!/usr/bin/env python3
"""
Portfolio Summary Columns A & B Updater
=======================================

Updates columns A & B of the Portfolio Summary sheet with:
- Column A: Account Names/Labels
- Column B: Current Portfolio Values from real-time API data

Integrates with the enhanced portfolio updater to get real-time values
from E*TRADE and Schwab APIs for all accounts.

Author: Assistant (GitHub Copilot)
Created: September 1, 2025
Purpose: Real-time Portfolio Summary column updates
"""

import os
import sys
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.styles.numbers import FORMAT_CURRENCY_USD_SIMPLE
from datetime import datetime
import traceback

# Add modules to path
sys.path.append(os.path.join(os.path.dirname(__file__), 'modules'))

class PortfolioSummaryColumnsUpdater:
    """
    Updates Portfolio Summary sheet columns A & B with account names and real-time values
    
    Column A: Account names and labels
    Column B: Current portfolio values from real-time API data
    
    Integration with enhanced portfolio updater for real-time data sources.
    """
    
    def __init__(self):
        """Initialize the Portfolio Summary columns updater"""
        self.target_file = os.path.join(os.path.dirname(__file__), "outputs", "Dividends_2025.xlsx")
        self.sheet_name = "Portfolio Summary"
        
    def get_realtime_portfolio_values(self):
        """Get real-time portfolio values from enhanced portfolio updater"""
        try:
            # Import the enhanced portfolio updater to get real-time values
            from enhanced_portfolio_updater_with_schwab import EnhancedPortfolioUpdater
            
            updater = EnhancedPortfolioUpdater()
            
            print("📊 Getting real-time E*TRADE values...")
            etrade_values = updater.get_etrade_values()
            
            print("📊 Getting real-time Schwab values...")  
            schwab_values = updater.get_schwab_values()
            
            if not etrade_values or not schwab_values:
                print("❌ Could not get real-time portfolio values")
                return None
            
            # Combine all values
            portfolio_values = {
                **etrade_values,
                **schwab_values
            }
            
            print("✅ Retrieved real-time portfolio values:")
            for account, value in portfolio_values.items():
                print(f"   {account}: ${value:,.2f}")
                
            return portfolio_values
            
        except Exception as e:
            print(f"❌ Error getting real-time portfolio values: {e}")
            return None
    
    def update_columns_a_and_b(self):
        """Update columns A & B of Portfolio Summary sheet"""
        try:
            print(f"\n📋 UPDATING PORTFOLIO SUMMARY COLUMNS A & B")
            print("=" * 55)
            
            if not os.path.exists(self.target_file):
                print(f"❌ Target file not found: {self.target_file}")
                return False
                
            # Get real-time portfolio values
            portfolio_values = self.get_realtime_portfolio_values()
            if not portfolio_values:
                print("❌ Cannot proceed without real-time portfolio values")
                return False
            
            # Load workbook
            workbook = openpyxl.load_workbook(self.target_file)
            
            if self.sheet_name not in workbook.sheetnames:
                print(f"❌ '{self.sheet_name}' sheet not found")
                return False
                
            sheet = workbook[self.sheet_name]
            print(f"✅ Found '{self.sheet_name}' sheet")
            
            # First, unmerge any merged cells in columns A & B (rows 1-15)
            print("🔧 Handling merged cells...")
            merged_ranges = list(sheet.merged_cells.ranges)
            for merged_range in merged_ranges:
                # Check if the merged range intersects with columns A or B (rows 1-15)
                if (merged_range.min_col <= 2 and merged_range.max_col >= 1 and 
                    merged_range.min_row <= 15 and merged_range.max_row >= 1):
                    print(f"   Unmerging: {merged_range}")
                    sheet.unmerge_cells(str(merged_range))
            
            # Define portfolio structure for columns A & B
            portfolio_structure = [
                {"row": 1, "label": "📊 PORTFOLIO SUMMARY", "value": ""},  # Header
                {"row": 2, "label": "Updated:", "value": datetime.now().strftime("%m/%d/%Y %H:%M")},  # Timestamp
                {"row": 3, "label": "", "value": ""},  # Blank row
                {"row": 4, "label": "💼 INVESTMENT ACCOUNTS", "value": "CURRENT VALUES"},  # Section header
                {"row": 5, "label": "E*TRADE IRA", "value": portfolio_values.get('E*TRADE IRA', 0)},
                {"row": 6, "label": "E*TRADE Taxable", "value": portfolio_values.get('E*TRADE Taxable', 0)},
                {"row": 7, "label": "Schwab IRA", "value": portfolio_values.get('Schwab IRA', 0)},
                {"row": 8, "label": "Schwab Individual", "value": portfolio_values.get('Schwab Individual', 0)},
                {"row": 9, "label": "", "value": ""},  # Blank row
                {"row": 10, "label": "📊 TOTAL PORTFOLIO", "value": "=SUM(B5:B8)"},  # Total formula
            ]
            
            print(f"\n📝 Updating columns A & B...")
            
            # Clear existing data in columns A & B (rows 1-15)
            for row in range(1, 16):
                try:
                    sheet.cell(row=row, column=1).value = None  # Column A
                    sheet.cell(row=row, column=2).value = None  # Column B
                except Exception as e:
                    print(f"   ⚠️ Could not clear row {row}: {e}")
            
            # Write new data
            for item in portfolio_structure:
                row = item["row"]
                label = item["label"]
                value = item["value"]
                
                try:
                    # Column A: Labels
                    cell_a = sheet.cell(row=row, column=1)
                    cell_a.value = label
                    
                    # Column B: Values  
                    cell_b = sheet.cell(row=row, column=2)
                    cell_b.value = value
                    
                    # Apply formatting based on content
                    if "📊" in label or "💼" in label:
                        # Header styling
                        cell_a.font = Font(name='Arial', size=12, bold=True, color='FFFFFF')
                        cell_a.fill = PatternFill(start_color='4F81BD', end_color='4F81BD', fill_type='solid')
                        cell_b.font = Font(name='Arial', size=12, bold=True, color='FFFFFF') 
                        cell_b.fill = PatternFill(start_color='4F81BD', end_color='4F81BD', fill_type='solid')
                    elif label in ["E*TRADE IRA", "E*TRADE Taxable", "Schwab IRA", "Schwab Individual"]:
                        # Account rows
                        cell_a.font = Font(name='Arial', size=12)
                        cell_b.font = Font(name='Arial', size=12)
                        cell_b.number_format = FORMAT_CURRENCY_USD_SIMPLE
                    elif "TOTAL" in label:
                        # Total row
                        cell_a.font = Font(name='Arial', size=12, bold=True)
                        cell_b.font = Font(name='Arial', size=12, bold=True)
                        cell_b.number_format = FORMAT_CURRENCY_USD_SIMPLE
                    else:
                        # Default formatting
                        cell_a.font = Font(name='Arial', size=12)
                        cell_b.font = Font(name='Arial', size=12)
                        
                    print(f"   Row {row}: {label} | {value}")
                    
                except Exception as e:
                    print(f"   ⚠️ Error updating row {row}: {e}")
                    continue
            
            # Set column widths
            sheet.column_dimensions['A'].width = 20  # Account labels
            sheet.column_dimensions['B'].width = 15  # Values
            
            print(f"\n🎨 Applied formatting and column widths")
            print(f"   Column A width: 20 (Account Labels)")
            print(f"   Column B width: 15 (Values)")
            
            # Calculate totals for display
            total_portfolio = sum([
                portfolio_values.get('E*TRADE IRA', 0),
                portfolio_values.get('E*TRADE Taxable', 0), 
                portfolio_values.get('Schwab IRA', 0),
                portfolio_values.get('Schwab Individual', 0)
            ])
            
            print(f"\n📊 PORTFOLIO SUMMARY UPDATED:")
            print(f"   E*TRADE IRA: ${portfolio_values.get('E*TRADE IRA', 0):,.2f}")
            print(f"   E*TRADE Taxable: ${portfolio_values.get('E*TRADE Taxable', 0):,.2f}")
            print(f"   Schwab IRA: ${portfolio_values.get('Schwab IRA', 0):,.2f}")
            print(f"   Schwab Individual: ${portfolio_values.get('Schwab Individual', 0):,.2f}")
            print(f"   📊 TOTAL PORTFOLIO: ${total_portfolio:,.2f}")
            
            # Save workbook
            workbook.save(self.target_file)
            print(f"\n✅ Portfolio Summary columns A & B updated successfully!")
            
            return True
            
        except Exception as e:
            print(f"❌ Error updating Portfolio Summary columns: {e}")
            print(f"Traceback: {traceback.format_exc()}")
            return False

def main():
    """Main execution function"""
    print("Portfolio Summary Columns A & B Updater")
    print("=" * 45)
    
    updater = PortfolioSummaryColumnsUpdater()
    
    success = updater.update_columns_a_and_b()
    
    if success:
        print(f"\n🎯 SUCCESS! Portfolio Summary columns A & B updated with real-time data")
    else:
        print(f"\n❌ Failed to update Portfolio Summary columns")

if __name__ == "__main__":
    main()
    try:
        input("\nPress Enter to continue...")
    except (EOFError, KeyboardInterrupt):
        print("\nExiting...")
