#!/usr/bin/env python3
"""
Clean up the historical yield sheet by removing incorrectly added tickers
and restoring it to the proper dividend-only structure
"""

import openpyxl
from openpyxl.styles import Font, PatternFill
import json
import os

class HistoricalYieldSheetCleanup:
    def __init__(self):
        self.script_dir = os.path.dirname(os.path.abspath(__file__))
        self.outputs_dir = os.path.join(self.script_dir, "outputs")
        self.excel_file = os.path.join(self.outputs_dir, "Dividends_2025.xlsx")
        self.cache_file = os.path.join(self.script_dir, "portfolio_data_cache.json")
        
        # Expected dividend tickers per account (>4% yield only)
        self.expected_structure = {
            "E*TRADE IRA": ["ABR", "PDI", "OFS", "NHS", "MORT", "QDTE", "QQQI", "SVOL", "RYLD", "BITO", "AGNC", "ACP", "BRSP", "CHMI", "DSL", "EIC", "ECC", "DX"],
            "E*TRADE Taxable": ["ABR", "PDI", "QDTE", "QQQI", "RYLD", "OFS", "MORT", "ACP", "AGNC", "BITO", "EIC", "SVOL"],
            "Schwab IRA": ["QDTE", "QQQI", "DX", "ECC", "AGNC"],  # This should include QQQI
            "Schwab Individual": ["QDTE", "DX"]
        }
        
        self.min_yield_threshold = 4.0

    def clean_and_restore_sheet(self):
        """Clean up incorrectly added tickers and restore proper structure"""
        print("🧹 HISTORICAL YIELD SHEET CLEANUP")
        print("=" * 50)
        
        # Load cache to get yield data
        with open(self.cache_file, 'r') as f:
            cache_data = json.load(f)
        
        ticker_yields = cache_data.get('ticker_yields', {})
        
        # Load workbook
        wb = openpyxl.load_workbook(self.excel_file)
        if 'Accounts Div historical yield' not in wb.sheetnames:
            print("ERROR: 'Accounts Div historical yield' sheet not found")
            return False
        
        ws = wb['Accounts Div historical yield']
        print("✅ Found historical yield sheet")
        
        # Find account groups
        account_rows = self.find_account_headers(ws)
        print(f"✅ Found {len(account_rows)} account groups")
        
        # Clean each account section
        for account_name, expected_tickers in self.expected_structure.items():
            if account_name not in account_rows:
                print(f"⚠️ {account_name} not found in sheet")
                continue
                
            print(f"\n🧹 Cleaning {account_name}...")
            self.clean_account_section(ws, account_name, account_rows[account_name], expected_tickers, ticker_yields)
        
        # Save cleaned workbook
        wb.save(self.excel_file)
        print(f"\n✅ Cleaned sheet saved to: {self.excel_file}")
        print("🎉 Historical yield sheet cleanup completed!")
        
    def find_account_headers(self, ws):
        """Find account header rows"""
        account_rows = {}
        
        for row in range(1, min(100, ws.max_row + 1)):
            cell_value = ws.cell(row=row, column=1).value
            if not cell_value:
                continue
                
            cell_text = str(cell_value).strip().upper()
            
            if 'ETRADE' in cell_text and 'IRA' in cell_text and 'TAXABLE' not in cell_text:
                account_rows["E*TRADE IRA"] = row
            elif 'ETRADE' in cell_text and 'TAXABLE' in cell_text:
                account_rows["E*TRADE Taxable"] = row
            elif 'SCHWAB' in cell_text and 'IRA' in cell_text:
                account_rows["Schwab IRA"] = row
            elif 'SCHWAB' in cell_text and 'INDIVIDUAL' in cell_text:
                account_rows["Schwab Individual"] = row
        
        return account_rows
    
    def clean_account_section(self, ws, account_name, header_row, expected_tickers, ticker_yields):
        """Clean and restore an account section with only expected dividend tickers"""
        
        # Find the end of this account section
        start_data_row = header_row + 2  # Skip header and column labels
        end_row = start_data_row
        
        # Find where this section ends
        while end_row <= ws.max_row:
            cell_value = ws.cell(row=end_row, column=1).value
            if cell_value:
                cell_text = str(cell_value).strip().upper()
                # Check if we hit another account header
                if any(keyword in cell_text for keyword in ['ETRADE', 'SCHWAB']) and end_row != header_row:
                    break
            end_row += 1
        
        end_row -= 1  # Last actual data row
        
        print(f"  📍 Data rows: {start_data_row} to {end_row}")
        
        # Clear all existing ticker rows in this section
        for row in range(start_data_row, end_row + 1):
            for col in range(1, ws.max_column + 1):
                ws.cell(row=row, column=col).value = None
        
        # Delete extra rows if any
        rows_to_delete = end_row - start_data_row + 1 - len(expected_tickers)
        if rows_to_delete > 0:
            ws.delete_rows(start_data_row, rows_to_delete)
            print(f"  🗑️ Deleted {rows_to_delete} extra rows")
        
        # Insert rows if needed
        elif rows_to_delete < 0:
            ws.insert_rows(start_data_row, abs(rows_to_delete))
            print(f"  ➕ Added {abs(rows_to_delete)} rows")
        
        # Add expected dividend tickers
        current_row = start_data_row
        for ticker in expected_tickers:
            # Verify this ticker has good yield
            yield_pct = ticker_yields.get(ticker, {}).get('yield', 0)
            if yield_pct < self.min_yield_threshold:
                print(f"    ⚠️ Skipping {ticker} - yield {yield_pct:.2f}% < {self.min_yield_threshold}%")
                continue
            
            # Add ticker symbol
            ws.cell(row=current_row, column=1).value = ticker
            ws.cell(row=current_row, column=1).font = Font(name="Arial", size=12, bold=True, color="3072C2")
            
            print(f"    ✅ Added {ticker} at row {current_row} (Yield: {yield_pct:.2f}%)")
            current_row += 1
        
        print(f"  ✅ {account_name} restored with {current_row - start_data_row} dividend tickers")

if __name__ == "__main__":
    cleanup = HistoricalYieldSheetCleanup()
    cleanup.clean_and_restore_sheet()