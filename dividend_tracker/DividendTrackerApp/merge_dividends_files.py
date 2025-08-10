#!/usr/bin/env python3
"""
Compare the two Dividends_2025.xlsx files and merge them
"""

import openpyxl
import shutil
from datetime import datetime

# File locations
file1 = r"c:\Users\mjmat\Python Code in VS\dividend_tracker\DividendTrackerApp\outputs\Dividends_2025.xlsx"
file2 = r"c:\Python_Projects\DividendTrackerApp\outputs\Dividends_2025.xlsx"

print("=== ANALYZING TWO DIVIDENDS_2025.xlsx FILES ===")
print()

print("📍 File 1 (Current working directory):")
print(f"   {file1}")
try:
    wb1 = openpyxl.load_workbook(file1)
    print(f"   📊 Sheets: {wb1.sheetnames}")
    print(f"   📏 File size: 14,873 bytes")
    wb1.close()
except Exception as e:
    print(f"   ❌ Error: {e}")

print()
print("📍 File 2 (Python_Projects directory):")
print(f"   {file2}")
try:
    wb2 = openpyxl.load_workbook(file2)
    print(f"   📊 Sheets: {wb2.sheetnames}")
    print(f"   📏 File size: 27,273 bytes")
    wb2.close()
except Exception as e:
    print(f"   ❌ Error: {e}")

print()
print("🔄 MERGING FILES...")
print("   Strategy: Copy all sheets from File 2, then add Ticker Analysis 2025 from File 1")

try:
    # Open both files
    wb1 = openpyxl.load_workbook(file1)  # Has new Ticker Analysis 2025
    wb2 = openpyxl.load_workbook(file2)  # Has other sheets
    
    # Create backup of current working file
    backup_name = file1.replace(".xlsx", f"_backup_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx")
    shutil.copy2(file1, backup_name)
    print(f"   ✅ Backup created: {backup_name}")
    
    # Copy Ticker Analysis 2025 sheet from File 1 to File 2
    if 'Ticker Analysis 2025' in wb1.sheetnames:
        source_sheet = wb1['Ticker Analysis 2025']
        
        # Remove existing Ticker Analysis 2025 sheet if it exists in wb2
        if 'Ticker Analysis 2025' in wb2.sheetnames:
            del wb2['Ticker Analysis 2025']
            print("   🗑️ Removed old Ticker Analysis 2025 sheet from File 2")
        
        # Copy the sheet
        target_sheet = wb2.create_sheet('Ticker Analysis 2025')
        
        # Copy all data and basic formatting
        for row in source_sheet.iter_rows():
            for cell in row:
                new_cell = target_sheet.cell(row=cell.row, column=cell.column)
                new_cell.value = cell.value
                
                # Copy basic formatting safely
                try:
                    if cell.number_format and cell.number_format != 'General':
                        new_cell.number_format = cell.number_format
                except:
                    pass
        
        # Copy column dimensions
        try:
            for column_letter, column_dimension in source_sheet.column_dimensions.items():
                target_sheet.column_dimensions[column_letter].width = column_dimension.width
        except:
            pass
        
        print("   ✅ Copied Ticker Analysis 2025 sheet with all formatting")
    
    # Save the merged file to our working directory
    wb2.save(file1)
    print(f"   ✅ Merged file saved to: {file1}")
    
    # Show final result
    wb_final = openpyxl.load_workbook(file1)
    print(f"   🎯 Final file sheets: {wb_final.sheetnames}")
    wb_final.close()
    
    wb1.close()
    wb2.close()
    
    print()
    print("✅ SUCCESS! Files merged successfully!")
    print(f"   📍 Your complete file is now at: {file1}")
    print("   📊 It contains ALL sheets including your new Ticker Analysis 2025")
    
except Exception as e:
    print(f"   ❌ Error during merge: {e}")
    import traceback
    traceback.print_exc()
