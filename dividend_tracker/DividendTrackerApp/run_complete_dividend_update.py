#!/usr/bin/env python3
"""
SINGLE COMMAND INTEGRATION - Your Complete Dividend Tracking System

Run this ONE script and it will:
1. Update your original dividend_stocks.xlsx with live E*TRADE data
2. Automatically update the ticker yield tracking with historical data
3. Update the comprehensive portfolio tracking
4. Generate all reports and analysis

USAGE: python run_complete_dividend_update.py
"""

import os
import sys
import subprocess
from datetime import datetime

def run_complete_dividend_update():
    """
    Single command that runs your complete dividend tracking system
    """
    
    print("ğŸš€ COMPLETE DIVIDEND TRACKING SYSTEM UPDATE")
    print("=" * 60)
    print(f"ğŸ• Started: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}")
    print("\nThis will update:")
    print("   âœ… Original dividend_stocks.xlsx (E*TRADE API)")
    print("   âœ… Ticker yield tracking with historical data")
    print("   âœ… Portfolio-level dividend estimates") 
    print("   âœ… Retirement dashboard")
    print("   âœ… All calculations and color coding")
    print("-" * 60)
    
    # Step 1: Update original dividend sheet
    print("\nğŸ“Š STEP 1: Updating Original Dividend Sheet...")
    print("   ğŸ”Œ Connecting to E*TRADE API and updating prices/yields...")
    
    try:
        # Run your original Update_dividend_sheet.py
        result = subprocess.run([
            sys.executable, 
            os.path.join("modules", "Update_dividend_sheet.py")
        ], cwd=os.getcwd(), capture_output=True, text=True, timeout=300)
        
        if result.returncode == 0:
            print("   âœ… Original dividend sheet updated successfully")
            if "File processed and saved" in result.stdout:
                print("   âœ… dividend_stocks.xlsx saved with latest data")
        else:
            print("   âš ï¸ Original update had some issues:")
            if result.stderr:
                print(f"      {result.stderr}")
            print("   ğŸ”„ Continuing with available data...")
            
    except subprocess.TimeoutExpired:
        print("   âš ï¸ Original update took too long - continuing with available data")
    except Exception as e:
        print(f"   âš ï¸ Could not run original update: {e}")
        print("   ğŸ“ Make sure Update_dividend_sheet.py is in modules/ folder")
    
    # Step 2: Update ticker tracking with historical yields
    print("\nğŸ“ˆ STEP 2: Updating Ticker Tracking with Historical Yields...")
    print("   ğŸ“… Preserving yield history and adding today's data...")
    
    try:
        result = subprocess.run([
            sys.executable, 
            "complete_ticker_integration.py"
        ], cwd=os.getcwd(), capture_output=True, text=True)
        
        if result.returncode == 0:
            print("   âœ… Ticker tracking updated with historical yields")
            print("   ğŸ¨ Color coding applied for yield changes")
        else:
            print("   âš ï¸ Ticker tracking had issues - check complete_ticker_integration.py")
            
    except Exception as e:
        print(f"   âŒ Error updating ticker tracking: {e}")
    
    # Step 3: Update portfolio-level tracking  
    print("\nğŸ¯ STEP 3: Updating Portfolio-Level Tracking...")
    print("   ğŸ“Š Refreshing dividend estimates and portfolio values...")
    
    try:
        result = subprocess.run([
            sys.executable, 
            "enhanced_dividend_tracker.py"
        ], cwd=os.getcwd(), capture_output=True, text=True)
        
        if result.returncode == 0:
            print("   âœ… Portfolio tracking updated")
            print("   ğŸ“ˆ Retirement dashboard refreshed")
        else:
            print("   âš ï¸ Portfolio tracking had issues - check enhanced_dividend_tracker.py")
            
    except Exception as e:
        print(f"   âŒ Error updating portfolio tracking: {e}")
    
    # Step 4: Generate summary report
    print("\nğŸ“‹ STEP 4: Generating Summary Report...")
    generate_summary_report()
    
    print(f"\nâœ… COMPLETE DIVIDEND UPDATE FINISHED")
    print(f"ğŸ• Completed: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}")
    print("\nğŸ“‚ Updated Files:")
    print("   â€¢ dividend_stocks.xlsx (original with latest prices/yields)")
    print("   â€¢ outputs/Dividends_2025.xlsx (comprehensive tracking)")
    print("     â†³ Estimated Income 2025 (portfolio level)")
    print("     â†³ Retirement Dashboard 2025 (retiree tracking)")
    print("     â†³ Ticker Analysis 2025 (individual stocks with yield history)")
    print("=" * 60)
    
    # Final validation
    show_final_validation()

def generate_summary_report():
    """
    Generate a summary of what was updated
    """
    
    try:
        import openpyxl
        
        # Check outputs
        output_file = os.path.join("outputs", "Dividends_2025.xlsx")
        dividend_file = "dividend_stocks.xlsx"
        
        print("   ğŸ“Š System Status Report:")
        
        # Check original dividend file
        if os.path.exists(dividend_file):
            mod_time = os.path.getmtime(dividend_file)
            mod_datetime = datetime.fromtimestamp(mod_time)
            print(f"      âœ… dividend_stocks.xlsx: Last updated {mod_datetime.strftime('%H:%M:%S')}")
        else:
            print("      âŒ dividend_stocks.xlsx: Not found")
        
        # Check comprehensive tracking
        if os.path.exists(output_file):
            mod_time = os.path.getmtime(output_file)
            mod_datetime = datetime.fromtimestamp(mod_time)
            print(f"      âœ… Dividends_2025.xlsx: Last updated {mod_datetime.strftime('%H:%M:%S')}")
            
            # Check sheet contents
            wb = openpyxl.load_workbook(output_file, data_only=True)
            sheets = wb.sheetnames
            print(f"      ğŸ“‹ Contains {len(sheets)} sheets: {', '.join(sheets)}")
            
            # Count tickers if ticker sheet exists
            if "Ticker Analysis 2025" in sheets:
                ws = wb["Ticker Analysis 2025"]
                ticker_count = 0
                for row in range(2, min(ws.max_row + 1, 50)):
                    if ws.cell(row=row, column=1).value and ws.cell(row=row, column=1).value != "TOTALS":
                        ticker_count += 1
                print(f"      ğŸ“ˆ Tracking {ticker_count} individual tickers")
            
            wb.close()
        else:
            print("      âŒ Dividends_2025.xlsx: Not found")
            
    except Exception as e:
        print(f"      âš ï¸ Error generating report: {e}")

def show_final_validation():
    """
    Show final validation of the integrated system
    """
    
    print("\nğŸ” FINAL VALIDATION:")
    print("-" * 25)
    
    try:
        import openpyxl
        
        output_file = os.path.join("outputs", "Dividends_2025.xlsx")
        if not os.path.exists(output_file):
            print("   âš ï¸ Cannot validate - Dividends_2025.xlsx not found")
            return
        
        wb = openpyxl.load_workbook(output_file, data_only=True)
        
        # Portfolio-level dividend
        portfolio_dividend = 0
        if "Estimated Income 2025" in wb.sheetnames:
            ws = wb["Estimated Income 2025"]
            # Find latest dividend total
            for col in range(2, min(ws.max_column + 1, 20)):
                for row in range(4, 8):
                    value = ws.cell(row=row, column=col).value
                    if isinstance(value, (int, float)) and value > 0:
                        portfolio_dividend = max(portfolio_dividend, value)
        
        # Ticker-level dividend
        ticker_dividend = 0
        ticker_count = 0
        if "Ticker Analysis 2025" in wb.sheetnames:
            ws = wb["Ticker Analysis 2025"]
            for row in range(2, ws.max_row + 1):
                cell_value = ws.cell(row=row, column=1).value
                if cell_value == "TOTALS":
                    ticker_dividend = ws.cell(row=row, column=15).value or 0  # New Payment column
                    break
                elif cell_value and cell_value != "TOTALS":
                    ticker_count += 1
        
        # Portfolio value
        portfolio_value = 0
        if "Retirement Dashboard 2025" in wb.sheetnames:
            ws = wb["Retirement Dashboard 2025"]
            portfolio_value = ws.cell(row=4, column=2).value or 0
        
        # Display results
        print(f"   ğŸ“Š Portfolio Value: ${portfolio_value:,.2f}")
        print(f"   ğŸ’° Portfolio Monthly Dividend: ${portfolio_dividend:,.2f}")
        print(f"   ğŸ“ˆ Ticker Monthly Dividend: ${ticker_dividend:,.2f}")
        print(f"   ğŸ¯ Individual Tickers Tracked: {ticker_count}")
        
        if ticker_dividend > 0 and portfolio_dividend > 0:
            coverage = (ticker_dividend / portfolio_dividend) * 100
            print(f"   ğŸ“Š Ticker Coverage: {coverage:.1f}% of portfolio dividends")
            
            if coverage > 80:
                print("   âœ… Excellent! Ticker tracking covers most of your dividends")
            elif coverage > 50:
                print("   âœ… Good coverage - ticker data represents majority of dividends")
            else:
                print("   ğŸ“‹ Partial coverage - consider adding more tickers to tracking")
        
        # Retirement metrics
        if "Retirement Dashboard 2025" in wb.sheetnames:
            ws = wb["Retirement Dashboard 2025"]
            days_retired = ws.cell(row=10, column=2).value or 0
            months_retired = ws.cell(row=11, column=2).value or 0
            print(f"   ğŸ–ï¸ Retirement Status: {int(days_retired)} days ({months_retired:.1f} months)")
        
        wb.close()
        
        print("\nğŸ‰ SUCCESS: Your complete dividend tracking system is updated!")
        print("   ğŸ“± Single command âœ…")
        print("   ğŸ”„ Auto-calculations âœ…") 
        print("   ğŸ“… Historical yield tracking âœ…")
        print("   ğŸ¨ Color coding âœ…")
        print("   ğŸ“Š Multi-level analysis âœ…")
        
    except Exception as e:
        print(f"   âŒ Validation error: {e}")

if __name__ == "__main__":
    run_complete_dividend_update()
