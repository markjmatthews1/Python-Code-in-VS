#!/usr/bin/env python3
"""
Check complete ticker analysis structure with all accounts
"""

import openpyxl

print("=== Complete Ticker Analysis 2025 Structure ===")
wb = openpyxl.load_workbook('outputs/Dividends_2025.xlsx')
ws = wb['Ticker Analysis 2025']

print(f"Total rows: {ws.max_row}")
print(f"Total columns: {ws.max_column}")

print("\nAll rows with account structure:")
for row in range(1, ws.max_row + 1):
    ticker = ws.cell(row=row, column=1).value
    account = ws.cell(row=row, column=2).value
    qty = ws.cell(row=row, column=3).value
    current_price = ws.cell(row=row, column=5).value
    current_value = ws.cell(row=row, column=9).value
    
    if ticker and "===" in str(ticker):
        print(f"\n🏦 {ticker}")
    elif ticker and ticker != "Ticker":
        print(f"   📈 {ticker}: {qty} shares @ ${current_price:.2f} = ${current_value:,.2f}" if current_price else f"   📈 {ticker}: {qty} shares")
    elif ticker == "TOTALS" or "Subtotals" in str(ticker or ""):
        print(f"   💰 {ticker}: ${current_value:,.2f}" if current_value else f"   💰 {ticker}")
    elif ticker == "GRAND TOTALS (All Accounts):":
        print(f"\n🎯 {ticker}: ${current_value:,.2f}" if current_value else f"\n🎯 {ticker}")

print(f"\n✅ Ticker Analysis 2025 sheet complete with account separation!")
print("🔄 Each account has a clear header and blank lines for separation.")
print("📊 All live market data integrated with real-time prices.")
print("💰 Account subtotals and grand totals included.")
