#!/usr/bin/env python3
"""
Fix Account Headers in Historical Yield Sheet
"""

import openpyxl
from openpyxl.styles import Font, PatternFill

def fix_account_headers():
    excel_file = r"C:\Users\mjmat\Python Code in VS\dividend_tracker\DividendTrackerApp\outputs\Dividends_2025.xlsx"
    
    wb = openpyxl.load_workbook(excel_file)
    ws = wb['Accounts Div historical yield']
    
    print("🔍 Current sheet structure:")
    for row in range(1, min(50, ws.max_row + 1)):
        cell_a = ws.cell(row=row, column=1).value
        if cell_a:
            print(f"Row {row}: {cell_a}")
    
    # Fix account headers
    print("\n🔧 Adding missing E*TRADE Taxable header...")
    
    # Find where E*TRADE Taxable should go (after E*TRADE IRA section)
    etrade_ira_end = 20  # Approximate end of E*TRADE IRA section
    
    # Insert E*TRADE Taxable header
    ws.insert_rows(etrade_ira_end + 1, 2)  # Insert 2 rows for header and spacing
    
    # Add E*TRADE Taxable header
    ws.cell(row=etrade_ira_end + 1, column=1).value = "E*TRADE Taxable"
    ws.cell(row=etrade_ira_end + 1, column=1).font = Font(bold=True, size=14)
    ws.cell(row=etrade_ira_end + 1, column=1).fill = PatternFill(start_color="FFA500", end_color="FFA500", fill_type="solid")
    
    # Add column headers
    headers = ["Ticker", "Qty", "Price", "Current Price", "Market Value"]
    for i, header in enumerate(headers, 1):
        ws.cell(row=etrade_ira_end + 2, column=i).value = header
        ws.cell(row=etrade_ira_end + 2, column=i).font = Font(bold=True)
    
    # Save
    wb.save(excel_file)
    print("✅ Fixed account headers")

if __name__ == "__main__":
    fix_account_headers()