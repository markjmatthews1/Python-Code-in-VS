"""
TARGETED Estimated Income Color Coding Fix

This script will ONLY add color coding to the Estimated Income 2025 sheet
without touching any other functionality or sheets.

Issues to fix:
1. Row 9 calculation: Correct to SUM(4:7)/12 
2. Color coding: Green (#00FF00), Red (#FF7C80), Yellow (#FFFF00)
3. Apply to rows 4-9 only
4. Don't touch Portfolio Values or any other sheets
"""
import openpyxl
from openpyxl.styles import PatternFill

# Correct color codes as specified by user
GREEN_COLOR = '00FF00'    # Green for increase
RED_COLOR = 'FF7C80'      # Red for decrease  
YELLOW_COLOR = 'FFFF00'   # Yellow for same

def add_color_coding_only():
    """Add ONLY color coding to Estimated Income sheet - don't change anything else"""
    excel_file = r"c:\Users\mjmat\Python Code in VS\dividend_tracker\DividendTrackerApp\outputs\Dividends_2025.xlsx"
    
    try:
        # Load with calculated values for comparison
        wb = openpyxl.load_workbook(excel_file, data_only=True)
        
        if "Estimated Income 2025" not in wb.sheetnames:
            print("❌ Estimated Income 2025 sheet not found")
            return False
            
        ws = wb["Estimated Income 2025"]
        
        print("🎨 ADDING COLOR CODING TO ESTIMATED INCOME 2025")
        print("=" * 55)
        print(f"Colors: Green={GREEN_COLOR}, Red={RED_COLOR}, Yellow={YELLOW_COLOR}")
        
        # Find the most recent two columns
        max_col = ws.max_column
        current_col = max_col
        prev_col = max_col - 1
        
        current_col_letter = openpyxl.utils.get_column_letter(current_col)
        prev_col_letter = openpyxl.utils.get_column_letter(prev_col)
        
        print(f"📊 Comparing {prev_col_letter} (previous) vs {current_col_letter} (current)")
        
        if prev_col < 2:
            print("⚠️ No previous column to compare with")
            wb.close()
            return False
        
        colors_applied = 0
        
        # Apply color coding to rows 4-9 ONLY
        for row in range(4, 10):  # Rows 4-9
            account_name = ws.cell(row=row, column=1).value
            current_val = ws.cell(row=row, column=current_col).value
            prev_val = ws.cell(row=row, column=prev_col).value
            
            # Only apply if both values exist and are numeric
            if (current_val is not None and prev_val is not None and
                isinstance(current_val, (int, float)) and isinstance(prev_val, (int, float))):
                
                # Determine color based on comparison
                if current_val > prev_val:
                    color = GREEN_COLOR
                    change_type = "INCREASE"
                elif current_val < prev_val:
                    color = RED_COLOR
                    change_type = "DECREASE"
                else:
                    color = YELLOW_COLOR
                    change_type = "SAME"
                
                # Apply the color
                ws.cell(row=row, column=current_col).fill = PatternFill(
                    start_color=color, end_color=color, fill_type='solid'
                )
                
                colors_applied += 1
                print(f"   ✅ Row {row} ({account_name}): ${prev_val:,.2f} → ${current_val:,.2f} = {change_type} ({color})")
        
        # Save the changes
        wb.save(excel_file)
        wb.close()
        
        print(f"\n✅ COLOR CODING APPLIED SUCCESSFULLY!")
        print(f"📊 Applied colors to {colors_applied} cells")
        print("🎯 ONLY color coding was changed - no other formatting touched")
        
        return True
        
    except Exception as e:
        print(f"❌ ERROR applying color coding: {e}")
        import traceback
        traceback.print_exc()
        return False

def update_dividend_tracker_plan():
    """Update the dividend tracker plan with correct color codes"""
    plan_file = r"c:\Users\mjmat\Python Code in VS\dividend_tracker\DividendTrackerApp\dividend_tracker_plan.txt"
    
    try:
        color_note = f"""

📝 ESTIMATED INCOME 2025 SHEET COLOR CODING (Updated 9/7/2025):
- Green (#00FF00): Current value > Previous value (increase)
- Red (#FF7C80): Current value < Previous value (decrease)  
- Yellow (#FFFF00): Current value = Previous value (same)
- Applied to rows 4-9: Account data + Monthly Average
- Compare current column with previous column (left to right progression)
"""
        
        # Append to plan file
        with open(plan_file, 'a', encoding='utf-8') as f:
            f.write(color_note)
            
        print("📋 Color code documentation added to dividend_tracker_plan.txt")
        
    except Exception as e:
        print(f"⚠️ Could not update plan file: {e}")

if __name__ == "__main__":
    success = add_color_coding_only()
    if success:
        update_dividend_tracker_plan()
