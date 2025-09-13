#!/usr/bin/env python3
"""
Cache-Based Historical Yield Updater
====================================

Updates the "Accounts Div historical yield" sheet with current dividend yield data
from portfolio_data_cache.json without making additional API calls.

Features:
- Reads yield data from cache (no API calls)
- Handles dynamic ticker counts per account
- Inserts new column after column O (column 15)
- Applies color coding: Green (increase), Red (decrease), Yellow (same)
- Processes 4 account groups: E*TRADE IRA, E*TRADE Taxable, Schwab IRA, Schwab Individual
"""

import os
import sys
import json
import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment
from openpyxl.styles.numbers import FORMAT_PERCENTAGE
from datetime import datetime
import traceback

class CacheHistoricalYieldUpdater:
    """Updates historical yield sheet using cached data"""
    
    def __init__(self):
        self.script_dir = os.path.dirname(os.path.abspath(__file__))
        self.outputs_dir = os.path.join(self.script_dir, "outputs")
        self.excel_file = os.path.join(self.outputs_dir, "Dividends_2025.xlsx")
        self.cache_file = os.path.join(self.script_dir, "portfolio_data_cache.json")
        self.today = datetime.now()
        self.today_str = self.today.strftime("%m-%d-%Y")
        
        # Account group mapping
        self.account_groups = {
            "E*TRADE IRA": "etrade_ira",
            "E*TRADE Taxable": "etrade_taxable", 
            "Schwab IRA": "schwab_ira",
            "Schwab Individual": "schwab_individual"
        }
        
    def load_cache_data(self):
        """Load yield data from cache"""
        try:
            if not os.path.exists(self.cache_file):
                print(f"❌ Cache file not found: {self.cache_file}")
                return None
                
            with open(self.cache_file, 'r') as f:
                cache_data = json.load(f)
                
            print(f"SUCCESS: Cache loaded: {cache_data.get('timestamp', 'No timestamp')}")
            
            # Extract ticker yields
            ticker_yields = cache_data.get('ticker_yields', {})
            print(f"Found yields for {len(ticker_yields)} tickers")
            
            return ticker_yields
            
        except Exception as e:
            print(f"ERROR: Error loading cache: {e}")
            traceback.print_exc()
            return None
    
    def find_account_groups(self, ws):
        """Find the starting rows for each account group"""
        account_rows = {}
        
        print("[SCAN] Scanning for account groups...")
        
        for row in range(1, min(60, ws.max_row + 1)):
            cell_value = ws.cell(row=row, column=1).value
            if not cell_value:
                continue
                
            cell_text = str(cell_value).strip().upper()
            
            # Look for account group headers
            if 'ETRADE' in cell_text and 'IRA' in cell_text and 'TAXABLE' not in cell_text:
                account_rows["E*TRADE IRA"] = row
                print(f"  SUCCESS: E*TRADE IRA found at row {row}")
            elif 'ETRADE' in cell_text and 'TAXABLE' in cell_text:
                account_rows["E*TRADE Taxable"] = row
                print(f"  SUCCESS: E*TRADE Taxable found at row {row}")
            elif 'SCHWAB' in cell_text and 'IRA' in cell_text:
                account_rows["Schwab IRA"] = row
                print(f"  SUCCESS: Schwab IRA found at row {row}")
            elif 'SCHWAB' in cell_text and 'INDIVIDUAL' in cell_text:
                account_rows["Schwab Individual"] = row
                print(f"  SUCCESS: Schwab Individual found at row {row}")
        
        return account_rows
    
    def get_tickers_for_group(self, ws, start_row):
        """Get all tickers for an account group"""
        tickers = []
        
        # Start from the row after the header (group divider + 1 for header row)
        current_row = start_row + 2  # Skip group divider and header row
        
        while current_row <= ws.max_row:
            ticker = ws.cell(row=current_row, column=1).value
            
            if not ticker:
                current_row += 1
                continue
                
            ticker_str = str(ticker).strip().upper()
            
            # Stop if we hit another account group divider
            if any(keyword in ticker_str for keyword in ['ETRADE', 'SCHWAB']) and any(keyword2 in ticker_str for keyword2 in ['IRA', 'TAXABLE', 'INDIVIDUAL']):
                break
                
            # Skip if it looks like a header row
            if any(keyword in ticker_str for keyword in ['TICKER', 'TOTAL', 'SUMMARY', 'QTY', 'PRICE']):
                current_row += 1
                continue
            
            # Valid ticker found - check if it's a reasonable ticker symbol
            if ticker_str and 2 <= len(ticker_str) <= 6 and ticker_str.isalpha():
                tickers.append((current_row, ticker_str))
            
            current_row += 1
            
            # Safety check - if we have more than 25 tickers, something's wrong
            if len(tickers) > 25:
                break
        
        return tickers
    
    def add_group_averages(self, ws, group_data, insert_col, account_rows):
        """Add AVERAGE calculations for each group at the bottom of each section"""
        # Sort groups by divider row to process in order
        sorted_groups = sorted(account_rows.items(), key=lambda x: x[1])
        
        for i, (group_name, divider_row) in enumerate(sorted_groups):
            if group_name not in group_data:
                continue
                
            data = group_data[group_name]
            first_row = data['first_row']
            last_row = data['last_row']
            
            # Calculate the average row (just before next group or at end)
            if i < len(sorted_groups) - 1:
                # Not the last group - average row is before next group divider
                next_divider_row = sorted_groups[i + 1][1]
                average_row = next_divider_row - 1
            else:
                # Last group - average row is after last ticker
                average_row = last_row + 2
            
            # Create AVERAGE formula using Excel column letter
            from openpyxl.utils import get_column_letter
            col_letter = get_column_letter(insert_col)
            formula = f"=AVERAGE({col_letter}{first_row}:{col_letter}{last_row})"
            
            # Set the average formula
            avg_cell = ws.cell(row=average_row, column=insert_col)
            avg_cell.value = formula
            avg_cell.number_format = '0.00'  # Two decimal places
            
            # Get previous average value for color coding
            old_avg_cell = ws.cell(row=average_row, column=insert_col - 1)
            old_avg = old_avg_cell.value
            
            # Handle formula values - extract the range and calculate manually
            old_avg_numeric = None
            if old_avg and isinstance(old_avg, str) and old_avg.startswith('=AVERAGE('):
                try:
                    # Extract the range from formula like "=AVERAGE(O3:O20)/100"
                    import re
                    match = re.search(r'=AVERAGE\(([^)]+)\)', old_avg)
                    if match:
                        range_str = match.group(1)  # e.g., "O3:O20"
                        # Parse range
                        start_cell, end_cell = range_str.split(':')
                        # Get column and rows
                        start_col = ord(start_cell[0]) - ord('A') + 1
                        start_row_num = int(start_cell[1:])
                        end_row_num = int(end_cell[1:])
                        
                        # Calculate average from the old column
                        total = 0
                        count = 0
                        for r in range(start_row_num, end_row_num + 1):
                            val = ws.cell(row=r, column=start_col).value
                            if val and isinstance(val, (int, float)):
                                total += val
                                count += 1
                        
                        if count > 0:
                            old_avg_numeric = total / count
                            # Check if formula has /100 at the end
                            if '/100' in old_avg:
                                old_avg_numeric = old_avg_numeric / 100
                except Exception as e:
                    print(f"    WARNING: Could not parse old average formula: {e}")
                    old_avg_numeric = None
            elif old_avg and isinstance(old_avg, (int, float)):
                old_avg_numeric = old_avg
            else:
                old_avg_numeric = None
            
            # Calculate current average for color coding comparison
            try:
                current_avg = 0
                count = 0
                for row in range(first_row, last_row + 1):
                    val = ws.cell(row=row, column=insert_col).value
                    if val is not None and isinstance(val, (int, float)):
                        current_avg += val
                        count += 1
                
                if count > 0:
                    current_avg = current_avg / count
                else:
                    current_avg = 0
                    
                # Apply color coding to average cell
                self.apply_color_coding(avg_cell, current_avg, old_avg_numeric)
                
                old_avg_display = f"{old_avg_numeric*100:.2f}%" if old_avg_numeric else "N/A"
                print(f"  SUCCESS: {group_name} average at row {average_row}: {current_avg:.2f}% (was {old_avg_display})")
                
            except Exception as e:
                print(f"  WARNING: Error calculating average for {group_name}: {e}")

    def apply_group_divider_coloring(self, ws, account_rows):
        """Apply orange coloring to group divider rows"""
        orange_fill = PatternFill(start_color='FFC000', end_color='FFC000', fill_type='solid')
        
        for group_name, row_num in account_rows.items():
            try:
                # Color the entire row orange for group dividers
                for col in range(1, ws.max_column + 1):
                    cell = ws.cell(row=row_num, column=col)
                    cell.fill = orange_fill
                print(f"  SUCCESS: {group_name} divider row {row_num} colored orange")
            except Exception as e:
                print(f"WARNING: Error coloring divider row {row_num}: {e}")

    def apply_color_coding(self, cell, new_yield, old_yield):
        """Apply color coding based on yield comparison"""
        try:
            if old_yield is None or old_yield == 0:
                # No previous value, use default formatting
                cell.fill = PatternFill(fill_type=None)
            else:
                old_val = float(old_yield) if old_yield else 0
                new_val = float(new_yield) if new_yield else 0
                
                if new_val > old_val:
                    # Increase - Green #90EE90
                    cell.fill = PatternFill(start_color='90EE90', end_color='90EE90', fill_type='solid')
                elif new_val < old_val:
                    # Decrease - Red #FF7C80
                    cell.fill = PatternFill(start_color='FF7C80', end_color='FF7C80', fill_type='solid')
                else:
                    # Same value - Yellow #FFFF00
                    cell.fill = PatternFill(start_color='FFFF00', end_color='FFFF00', fill_type='solid')
                    
        except Exception as e:
            print(f"      WARNING: Color coding error for {new_yield}: {e}")
            cell.fill = PatternFill(fill_type=None)
    
    def update_historical_yield_sheet(self):
        """Main update function"""
        print("\n[UPDATE] CACHE-BASED HISTORICAL YIELD UPDATER")
        print("=" * 50)
        print(f"Date: {self.today_str}")
        print(f"Excel: {os.path.basename(self.excel_file)}")
        
        # Load cache data
        ticker_yields = self.load_cache_data()
        if not ticker_yields:
            return False
        
        try:
            # Open workbook
            if not os.path.exists(self.excel_file):
                print(f"ERROR: Excel file not found: {self.excel_file}")
                return False
                
            wb = openpyxl.load_workbook(self.excel_file)
            
            # Find the historical yield sheet
            sheet_name = "Accounts Div historical yield"
            if sheet_name not in wb.sheetnames:
                print(f"ERROR: Sheet '{sheet_name}' not found")
                print(f"Available sheets: {wb.sheetnames}")
                return False
            
            ws = wb[sheet_name]
            print(f"SUCCESS: Found sheet: '{sheet_name}'")
            
            # Find account groups
            account_rows = self.find_account_groups(ws)
            if not account_rows:
                print("ERROR: No account groups found")
                return False
            
            # Find insertion column (after column O = column 15)
            insert_col = 16  # Column P
            
            # Insert new column
            ws.insert_cols(insert_col)
            print(f"SUCCESS: Inserted new column at position {insert_col}")
            
            # Add headers for new column in all group sections
            # Format date as mm/dd/yyyy
            from datetime import datetime
            date_formatted = datetime.now().strftime("%m/%d/%Y")
            
            # Determine header rows for each group (group divider + 1)
            header_rows = {}
            for group_name, divider_row in account_rows.items():
                header_row = divider_row + 1
                header_rows[group_name] = header_row
                
                # Add date header
                header_cell = ws.cell(row=header_row, column=insert_col)
                header_cell.value = date_formatted
                header_cell.font = Font(bold=True)
                header_cell.alignment = Alignment(horizontal='center')
                print(f"  SUCCESS: Added date header '{date_formatted}' to row {header_row} for {group_name}")
            
            total_updated = 0
            group_data = {}  # Store ticker ranges for average calculations
            
            # Process each account group
            for group_name, start_row in account_rows.items():
                print(f"\n[PROCESSING] {group_name} (starting row {start_row})...")
                
                # Get tickers for this group
                tickers = self.get_tickers_for_group(ws, start_row)
                print(f"   Found {len(tickers)} tickers")
                
                # Store first and last ticker rows for average calculation
                if tickers:
                    first_ticker_row = tickers[0][0]
                    last_ticker_row = tickers[-1][0]
                    group_data[group_name] = {
                        'first_row': first_ticker_row,
                        'last_row': last_ticker_row,
                        'divider_row': start_row
                    }
                
                for ticker_row, ticker in tickers:
                    # Get current yield from cache
                    current_yield = None
                    if ticker in ticker_yields:
                        current_yield = ticker_yields[ticker].get('yield', 0)
                    
                    if current_yield is None:
                        print(f"      WARNING: {ticker}: No yield data in cache")
                        continue
                    
                    # Get previous yield from column O (now column 15)
                    old_yield_cell = ws.cell(row=ticker_row, column=15)
                    old_yield = old_yield_cell.value
                    
                    # Set new yield value
                    new_cell = ws.cell(row=ticker_row, column=insert_col)
                    new_cell.value = float(current_yield)
                    new_cell.number_format = '0.00'  # Two decimal places
                    
                    # Apply color coding
                    self.apply_color_coding(new_cell, current_yield, old_yield)
                    
                    print(f"      SUCCESS: {ticker}: {current_yield:.2f}% (was {old_yield if old_yield else 'N/A'})")
                    total_updated += 1
            
            # Add group averages
            print("\n[CALCULATIONS] Adding group averages...")
            self.add_group_averages(ws, group_data, insert_col, account_rows)
            
            # Apply orange coloring to group divider rows
            print("\n[FORMATTING] Applying orange coloring to group dividers...")
            self.apply_group_divider_coloring(ws, account_rows)
            
            # Save workbook
            wb.save(self.excel_file)
            wb.close()
            
            print(f"\nSUCCESS! Updated {total_updated} ticker yields")
            print(f"New column inserted after column O with color coding")
            print(f"Saved to: {os.path.basename(self.excel_file)}")
            
            return True
            
        except Exception as e:
            print(f"ERROR: Update failed: {e}")
            traceback.print_exc()
            return False

def main():
    """Main execution"""
    updater = CacheHistoricalYieldUpdater()
    
    try:
        success = updater.update_historical_yield_sheet()
        
        if success:
            print("\nSUCCESS: Historical yield sheet updated successfully!")
        else:
            print("\nWARNING: Update completed with some issues")
            
    except KeyboardInterrupt:
        print("\nCANCELLED: Update cancelled by user")
    except Exception as e:
        print(f"\nERROR: Critical error: {e}")
        traceback.print_exc()

if __name__ == "__main__":
    main()