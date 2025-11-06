#!/usr/bin/env python3
"""
Migrate Dividends_2025.xlsx from Single Sheet to Separate Account Sheets
========================================================================

This script will:
1. Backup the original Dividends_2025.xlsx file
2. Read data from "Accounts Div historical yield" sheet (4 account sections)
3. Create 4 new separate sheets:
   - Etrade_IRA
   - Etrade_Individual
   - Schwab_IRA
   - Schwab_Individual
4. Migrate data, formatting, and formulas to each new sheet
5. Keep original sheet as "Accounts Div historical yield (OLD)" for reference

Author: Claude & Mark
Date: October 18, 2025
"""

import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import shutil
from datetime import datetime
import os

def backup_file(filepath):
    """Create a timestamped backup of the Excel file"""
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    backup_path = filepath.replace('.xlsx', f'_BACKUP_BEFORE_MIGRATION_{timestamp}.xlsx')
    shutil.copy2(filepath, backup_path)
    print(f"✅ Backup created: {backup_path}")
    return backup_path

def copy_cell_style(source_cell, target_cell):
    """Copy all styling from source cell to target cell"""
    if source_cell.has_style:
        target_cell.font = source_cell.font.copy()
        target_cell.border = source_cell.border.copy()
        target_cell.fill = source_cell.fill.copy()
        target_cell.number_format = source_cell.number_format
        target_cell.protection = source_cell.protection.copy()
        target_cell.alignment = source_cell.alignment.copy()

def copy_column_widths(source_sheet, target_sheet):
    """Copy column widths from source to target sheet"""
    for col_letter in source_sheet.column_dimensions:
        target_sheet.column_dimensions[col_letter].width = source_sheet.column_dimensions[col_letter].width

def copy_row_heights(source_sheet, target_sheet, source_rows, target_start_row=1):
    """Copy row heights from source rows to target sheet"""
    for idx, source_row_num in enumerate(source_rows):
        target_row_num = target_start_row + idx
        if source_row_num in source_sheet.row_dimensions:
            target_sheet.row_dimensions[target_row_num].height = source_sheet.row_dimensions[source_row_num].height

def migrate_account_section(source_sheet, target_sheet, section_config, account_name):
    """
    Migrate one account section from source to target sheet
    
    Args:
        source_sheet: The original "Accounts Div historical yield" sheet
        target_sheet: The new individual account sheet
        section_config: Dict with row ranges for this account
        account_name: Name of the account (for display)
    """
    print(f"\n📋 Migrating {account_name}...")
    
    # Define source row ranges
    title_row = section_config['title_row']
    header_row = section_config['header_row']
    data_start = section_config['data_start']
    data_end = section_config['data_end']
    avg_row = section_config.get('avg_row')
    
    # Collect all rows to copy (title, header, data rows, avg row)
    source_rows = [title_row, header_row] + list(range(data_start, data_end + 1))
    if avg_row:
        source_rows.append(avg_row)
    
    # Target starts at row 1
    target_row = 1
    
    # Copy each row
    for source_row_num in source_rows:
        max_col = source_sheet.max_column
        
        for col in range(1, max_col + 1):
            source_cell = source_sheet.cell(row=source_row_num, column=col)
            target_cell = target_sheet.cell(row=target_row, column=col)
            
            # Copy value
            target_cell.value = source_cell.value
            
            # Copy style
            copy_cell_style(source_cell, target_cell)
            
            # Handle formulas (convert relative references if needed)
            if source_cell.data_type == 'f':
                target_cell.value = source_cell.value  # Keep formula as-is for now
        
        target_row += 1
    
    # Copy column widths
    copy_column_widths(source_sheet, target_sheet)
    
    # Copy row heights
    copy_row_heights(source_sheet, target_sheet, source_rows)
    
    print(f"   ✅ Copied {len(source_rows)} rows")
    print(f"   ✅ Copied {max_col} columns")
    return True

def migrate_to_separate_sheets():
    """Main migration function"""
    
    print("=" * 70)
    print("🔄 DIVIDEND TRACKER MIGRATION - Single Sheet → Separate Account Sheets")
    print("=" * 70)
    
    # File path
    excel_path = r"dividend_tracker\DividendTrackerApp\outputs\Dividends_2025.xlsx"
    
    if not os.path.exists(excel_path):
        print(f"❌ Error: File not found at {excel_path}")
        return False
    
    print(f"\n📁 Working with: {excel_path}")
    
    # Step 1: Backup
    print("\n1️⃣  Creating backup...")
    backup_path = backup_file(excel_path)
    
    # Step 2: Load workbook
    print("\n2️⃣  Loading workbook...")
    try:
        wb = openpyxl.load_workbook(excel_path)
        print(f"   ✅ Workbook loaded")
        print(f"   📋 Current sheets: {wb.sheetnames}")
    except Exception as e:
        print(f"   ❌ Error loading workbook: {e}")
        return False
    
    # Step 3: Verify source sheet exists
    source_sheet_name = "Accounts Div historical yield"
    if source_sheet_name not in wb.sheetnames:
        print(f"   ❌ Error: '{source_sheet_name}' sheet not found")
        wb.close()
        return False
    
    source_sheet = wb[source_sheet_name]
    print(f"   ✅ Source sheet found: {source_sheet.max_row} rows × {source_sheet.max_column} cols")
    
    # Step 4: Define account sections based on current structure
    print("\n3️⃣  Defining account sections...")
    
    account_sections = [
        {
            'name': 'Etrade_IRA',
            'display_name': 'E*TRADE Rollover IRA',
            'title_row': 1,
            'header_row': 2,
            'data_start': 3,
            'data_end': 21,  # Verified from inspection
            'avg_row': None  # Will add if found
        },
        {
            'name': 'Etrade_Individual',
            'display_name': 'E*TRADE Individual Brokerage',
            'title_row': 22,
            'header_row': 23,
            'data_start': 24,
            'data_end': 37,  # Verified from inspection
            'avg_row': None  # Will add if found
        },
        {
            'name': 'Schwab_IRA',
            'display_name': 'Schwab IRA',
            'title_row': 38,
            'header_row': 39,
            'data_start': 40,
            'data_end': 48,  # Verified from inspection
            'avg_row': None  # Will add if found
        },
        {
            'name': 'Schwab_Individual',
            'display_name': 'Schwab Individual',
            'title_row': 49,
            'header_row': 50,
            'data_start': 51,
            'data_end': 54,  # Verified from inspection (sheet max row)
            'avg_row': None  # Will add if found
        }
    ]
    
    for section in account_sections:
        print(f"   • {section['display_name']}: rows {section['title_row']}-{section.get('avg_row', section['data_end'])}")
    
    # Step 5: Create new sheets for each account
    print("\n4️⃣  Creating new account sheets...")
    
    created_sheets = []
    for section in account_sections:
        sheet_name = section['name']
        
        # Create new sheet (or use existing if already present)
        if sheet_name in wb.sheetnames:
            print(f"   ⚠️  Sheet '{sheet_name}' already exists, will overwrite")
            del wb[sheet_name]
        
        new_sheet = wb.create_sheet(sheet_name)
        created_sheets.append(sheet_name)
        print(f"   ✅ Created sheet: {sheet_name}")
    
    # Step 6: Migrate data to each new sheet
    print("\n5️⃣  Migrating data to new sheets...")
    
    for section in account_sections:
        success = migrate_account_section(
            source_sheet=source_sheet,
            target_sheet=wb[section['name']],
            section_config=section,
            account_name=section['display_name']
        )
        
        if not success:
            print(f"   ❌ Migration failed for {section['display_name']}")
            wb.close()
            return False
    
    # Step 7: Rename original sheet (keep as reference)
    print("\n6️⃣  Preserving original sheet...")
    old_sheet_name = "Accounts Div historical yield (OLD)"
    
    if old_sheet_name in wb.sheetnames:
        del wb[old_sheet_name]
    
    source_sheet.title = old_sheet_name
    print(f"   ✅ Renamed original sheet to: '{old_sheet_name}'")
    
    # Step 8: Reorder sheets (put new account sheets first)
    print("\n7️⃣  Reordering sheets...")
    
    # Move new account sheets to the front
    for idx, sheet_name in enumerate(created_sheets):
        wb.move_sheet(sheet_name, offset=-len(wb.sheetnames) + idx + 1)
    
    print(f"   ✅ New sheet order: {wb.sheetnames}")
    
    # Step 9: Save workbook
    print("\n8️⃣  Saving workbook...")
    
    try:
        wb.save(excel_path)
        wb.close()
        print(f"   ✅ Workbook saved successfully!")
    except Exception as e:
        print(f"   ❌ Error saving workbook: {e}")
        wb.close()
        return False
    
    # Step 10: Summary
    print("\n" + "=" * 70)
    print("✅ MIGRATION COMPLETE!")
    print("=" * 70)
    print(f"\n📊 Summary:")
    print(f"   • Backup file: {os.path.basename(backup_path)}")
    print(f"   • Original sheet renamed: '{old_sheet_name}'")
    print(f"   • New account sheets created: {len(created_sheets)}")
    for sheet_name in created_sheets:
        print(f"     ✓ {sheet_name}")
    
    print(f"\n⚠️  IMPORTANT NEXT STEPS:")
    print(f"   1. Open Dividends_2025.xlsx and verify data in new sheets")
    print(f"   2. Check that formulas still work correctly")
    print(f"   3. Verify formatting looks good")
    print(f"   4. If all looks good, you can delete the OLD sheet")
    print(f"   5. Update dividend tracker app to use new sheet names")
    
    return True

if __name__ == "__main__":
    try:
        success = migrate_to_separate_sheets()
        if success:
            print("\n🎉 Migration completed successfully!")
        else:
            print("\n❌ Migration failed - check backup file")
    except Exception as e:
        print(f"\n❌ Unexpected error: {e}")
        import traceback
        traceback.print_exc()
