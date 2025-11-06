#!/usr/bin/env python3
"""
Examine Portfolio Summary Sheet Structure
"""
import openpyxl

def examine_portfolio_summary():
    excel_file = 'outputs/Dividends_2025.xlsx'
    
    try:
        wb = openpyxl.load_workbook(excel_file, data_only=True)
        
        if 'Portfolio Summary' in wb.sheetnames:
            ws = wb['Portfolio Summary']
            print('Portfolio Summary Sheet Structure:')
            print(f'Max row: {ws.max_row}, Max col: {ws.max_column}')
            
            print('\nRows 30-40 content:')
            for row in range(30, min(41, ws.max_row + 1)):
                row_data = []
                for col in range(1, min(6, ws.max_column + 1)):
                    cell_value = ws.cell(row=row, column=col).value
                    if cell_value:
                        row_data.append(f'Col {col}: {cell_value}')
                if row_data:
                    print(f'Row {row}: {row_data}')
            
            print('\nWithdrawal data (Column B, rows 32-34):')
            for row in range(32, 35):
                value = ws.cell(row=row, column=2).value
                label = ws.cell(row=row, column=1).value
                print(f'Row {row}: {label} = {value}')
                
        else:
            print('Portfolio Summary sheet not found')
            
        wb.close()
        
    except Exception as e:
        print(f"Error: {e}")

if __name__ == "__main__":
    examine_portfolio_summary()
