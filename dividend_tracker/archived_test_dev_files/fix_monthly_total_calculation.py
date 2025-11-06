#!/usr/bin/env python3
"""
Fix Monthly Total Calculation in Estimated Income 2025 Sheet
- Row 9 should sum rows 4-7 (all account annual totals) and divide by 12
- Apply proper formatting and color coding
- Fix current calculation issue showing $799.84 instead of ~$3000
"""

import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment
from datetime import datetime
import os

def fix_monthly_total_calculation():
    """Fix the monthly total calculation in Estimated Income 2025 sheet"""
    
    excel_file = r'c:\Users\mjmat\Python Code in VS\dividend_tracker\DividendTrackerApp\outputs\Dividends_2025.xlsx'
    
    if not os.path.exists(excel_file):
        print(f"❌ Excel file not found: {excel_file}")
        return False
    
    try:
        print("🔄 Fixing monthly total calculation in Estimated Income 2025 sheet...")
        
        # Load the workbook
        wb = openpyxl.load_workbook(excel_file, data_only=False)
        
        if 'Estimated Income 2025' not in wb.sheetnames:
            print("❌ 'Estimated Income 2025' sheet not found!")
            return False
        
        ws = wb['Estimated Income 2025']
        
        # Define styles
        bold_font = Font(name='Arial', size=12, bold=True)
        currency_format = '$#,##0.00'
        
        # Color fills for change indicators
        green_fill = PatternFill(start_color='00B050', end_color='00B050', fill_type='solid')  # Green for increase
        red_fill = PatternFill(start_color='FF7C80', end_color='FF7C80', fill_type='solid')   # Light red for decrease
        yellow_fill = PatternFill(start_color='FFFF00', end_color='FFFF00', fill_type='solid') # Yellow for no change
        blue_fill = PatternFill(start_color='D9E2F3', end_color='D9E2F3', fill_type='solid')  # Light blue default
        
        print("📊 Analyzing current data structure...")
        
        # Find the data range
        max_row = ws.max_row
        max_col = ws.max_column
        
        print(f"   Sheet dimensions: {max_row} rows x {max_col} columns")
        
        # Set the Monthly Total label if not already there
        monthly_row = 9
        label_cell = ws.cell(row=monthly_row, column=1)
        if not label_cell.value or 'Monthly' not in str(label_cell.value):
            label_cell.value = "Monthly Total"
        label_cell.font = bold_font
        
        print("💰 Fixing monthly total calculations...")
        
        # Process each column (week) from column 2 onwards
        previous_monthly_value = None
        fixed_columns = 0
        
        for col in range(2, max_col + 1):
            # Calculate the sum of annual amounts from rows 4-7 (account rows)
            annual_total = 0
            account_values = []
            
            for account_row in range(4, 8):  # Rows 4-7 for individual accounts
                cell_value = ws.cell(row=account_row, column=col).value
                if isinstance(cell_value, (int, float)):
                    annual_total += cell_value
                    account_values.append(cell_value)
                    
            # Calculate monthly total (annual / 12)
            monthly_total = annual_total / 12 if annual_total > 0 else 0
            
            # Update the monthly total cell
            monthly_cell = ws.cell(row=monthly_row, column=col)
            old_value = monthly_cell.value if isinstance(monthly_cell.value, (int, float)) else 0
            
            monthly_cell.value = monthly_total
            monthly_cell.number_format = currency_format
            monthly_cell.font = bold_font
            
            # Apply color coding based on change from previous column
            if col > 2 and previous_monthly_value is not None:
                if monthly_total > previous_monthly_value:
                    monthly_cell.fill = green_fill  # Increase
                    change_indicator = "📈"
                elif monthly_total < previous_monthly_value:
                    monthly_cell.fill = red_fill   # Decrease  
                    change_indicator = "📉"
                else:
                    monthly_cell.fill = yellow_fill # No change
                    change_indicator = "➡️"
                    
                change_amount = monthly_total - previous_monthly_value
                print(f"   Column {col}: ${monthly_total:,.2f} {change_indicator} (${change_amount:+,.2f}) [was ${old_value:,.2f}]")
            else:
                # First column or no previous value - use light blue
                monthly_cell.fill = blue_fill
                print(f"   Column {col}: ${monthly_total:,.2f} [was ${old_value:,.2f}]")
            
            # Show the account breakdown for this column
            if account_values:
                print(f"      Annual breakdown: {[f'${v:,.2f}' for v in account_values]} = ${annual_total:,.2f} total")
                
            previous_monthly_value = monthly_total
            fixed_columns += 1
        
        print(f"\n✅ Fixed {fixed_columns} monthly total calculations!")
        print(f"📊 Row 9 now correctly sums rows 4-7 and divides by 12")
        
        # Save the workbook
        wb.save(excel_file)
        wb.close()
        
        print(f"💾 Changes saved to: {excel_file}")
        
        return True
        
    except Exception as e:
        print(f"❌ Error fixing monthly totals: {e}")
        return False

def verify_calculation():
    """Verify that the monthly calculation is now correct"""
    
    excel_file = r'c:\Users\mjmat\Python Code in VS\dividend_tracker\DividendTrackerApp\outputs\Dividends_2025.xlsx'
    
    try:
        print("\n🔍 Verifying the fixed calculation...")
        
        wb = openpyxl.load_workbook(excel_file, data_only=True)  # Load with calculated values
        ws = wb['Estimated Income 2025']
        
        # Check the last few columns to verify
        max_col = ws.max_column
        sample_cols = [max_col-2, max_col-1, max_col] if max_col >= 4 else [max_col]
        
        for col in sample_cols:
            if col >= 2:
                # Get account totals
                annual_total = 0
                account_breakdown = []
                
                for row in range(4, 8):
                    value = ws.cell(row=row, column=col).value
                    if isinstance(value, (int, float)):
                        annual_total += value
                        account_breakdown.append(f"${value:,.2f}")
                
                # Get monthly total
                monthly_value = ws.cell(row=9, column=col).value
                expected_monthly = annual_total / 12
                
                print(f"   Column {col}:")
                print(f"      Account totals: {' + '.join(account_breakdown)} = ${annual_total:,.2f} annual")
                print(f"      Monthly total: ${monthly_value:,.2f}")
                print(f"      Expected: ${expected_monthly:,.2f}")
                
                if abs(monthly_value - expected_monthly) < 0.01:
                    print(f"      ✅ Calculation CORRECT")
                else:
                    print(f"      ❌ Calculation ERROR - difference: ${abs(monthly_value - expected_monthly):,.2f}")
        
        wb.close()
        
    except Exception as e:
        print(f"❌ Error during verification: {e}")

if __name__ == "__main__":
    print("🚀 FIXING MONTHLY TOTAL CALCULATION")
    print("=" * 60)
    print("Issue: Row 9 showing $799.84 instead of ~$3000")
    print("Fix: Sum rows 4-7 (account totals) ÷ 12 for monthly amount")
    print("=" * 60)
    
    success = fix_monthly_total_calculation()
    
    if success:
        verify_calculation()
        
        print("\n🎯 Summary:")
        print("✅ Monthly total calculation fixed in row 9")
        print("✅ Proper formatting and color coding applied")
        print("✅ Formula now correctly sums account rows 4-7 and divides by 12")
        print("\n💡 The monthly total should now show ~$3000 instead of $799.84")
    else:
        print("\n❌ Failed to fix monthly total calculation")
