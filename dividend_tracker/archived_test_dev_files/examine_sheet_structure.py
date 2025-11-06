#!/usr/bin/env python3
"""
Quick utility to examine the Accounts Div historical yield sheet structure
"""

import openpyxl
import os

def examine_historical_yield_sheet():
    """Examine the current sheet structure"""
    script_dir = os.path.dirname(os.path.abspath(__file__))
    outputs_dir = os.path.join(script_dir, "outputs")
    excel_file = os.path.join(outputs_dir, "Dividends_2025.xlsx")
    
    if not os.path.exists(excel_file):
        print(f"❌ Excel file not found: {excel_file}")
        return
    
    try:
        wb = openpyxl.load_workbook(excel_file)
        
        # Look for the historical yield sheet
        sheet_names = wb.sheetnames
        print(f"📋 Available sheets: {sheet_names}")
        
        # Find the historical yield sheet (might have different names)
        yield_sheet_name = None
        for name in sheet_names:
            if 'historical' in name.lower() and 'yield' in name.lower():
                yield_sheet_name = name
                break
            elif 'div' in name.lower() and ('historical' in name.lower() or 'yield' in name.lower()):
                yield_sheet_name = name
                break
        
        if not yield_sheet_name:
            print("❌ Historical yield sheet not found")
            print("Available sheets:")
            for i, name in enumerate(sheet_names, 1):
                print(f"  {i}. {name}")
            return
        
        print(f"✅ Found yield sheet: '{yield_sheet_name}'")
        ws = wb[yield_sheet_name]
        
        print(f"\n📊 Sheet dimensions: {ws.max_row} rows x {ws.max_column} columns")
        
        # Examine headers and structure
        print("\n📝 First 10 rows, first 20 columns:")
        for row in range(1, min(11, ws.max_row + 1)):
            row_data = []
            for col in range(1, min(21, ws.max_column + 1)):
                cell_value = ws.cell(row=row, column=col).value
                if cell_value is None:
                    row_data.append("")
                else:
                    row_data.append(str(cell_value)[:15])  # Truncate long values
            print(f"Row {row:2d}: {row_data}")
        
        # Look for column O specifically
        print(f"\n🎯 Column O (column 15) content:")
        for row in range(1, min(21, ws.max_row + 1)):
            cell_value = ws.cell(row=row, column=15).value
            if cell_value:
                print(f"  Row {row}: {cell_value}")
        
        # Look for account groups
        print(f"\n🏦 Looking for account group patterns:")
        for row in range(1, min(50, ws.max_row + 1)):
            cell_value = ws.cell(row=row, column=1).value
            if cell_value and any(keyword in str(cell_value).upper() for keyword in ['ETRADE', 'SCHWAB', 'IRA', 'TAXABLE', 'INDIVIDUAL']):
                print(f"  Row {row}: {cell_value}")
        
        wb.close()
        
    except Exception as e:
        print(f"❌ Error examining sheet: {e}")
        import traceback
        traceback.print_exc()

if __name__ == "__main__":
    examine_historical_yield_sheet()