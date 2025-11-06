import os
import sys
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.styles.numbers import FORMAT_CURRENCY_USD_SIMPLE
from datetime import datetime

def test_monthly_average_formatting():
    """Test the Monthly Average formatting with bold and color coding"""
    
    excel_file = "outputs/Dividends_2025.xlsx"
    if not os.path.exists(excel_file):
        print(f"ERROR: Excel file not found: {excel_file}")
        return
    
    try:
        wb = openpyxl.load_workbook(excel_file)
        
        if "Estimated Income 2025" not in wb.sheetnames:
            print("ERROR: Estimated Income 2025 sheet not found")
            print(f"Available sheets: {wb.sheetnames}")
            return
            
        ws = wb["Estimated Income 2025"]
        
        print("ESTIMATED INCOME 2025 - MONTHLY AVERAGE ANALYSIS")
        print("=" * 60)
        
        # Find Monthly Average row (usually around row 9)
        monthly_row = None
        for row in range(8, 12):
            cell_value = ws.cell(row=row, column=1).value
            if cell_value and 'monthly' in str(cell_value).lower():
                monthly_row = row
                print(f"Found Monthly Average row: {row}")
                break
        
        if monthly_row:
            # Check the most recent columns
            print(f"\nMONTHLY AVERAGE ROW {monthly_row} - RECENT COLUMNS:")
            for col in range(max(1, ws.max_column - 3), ws.max_column + 1):
                cell = ws.cell(row=monthly_row, column=col)
                col_letter = openpyxl.utils.get_column_letter(col)
                
                # Check formatting
                is_bold = cell.font.bold if cell.font else False
                has_fill = cell.fill.start_color.rgb if cell.fill and cell.fill.start_color else None
                
                if hasattr(cell, 'data_type') and cell.data_type == 'f':
                    print(f"  Column {col_letter}: FORMULA = {cell.value} | Bold: {is_bold} | Fill: {has_fill}")
                else:
                    print(f"  Column {col_letter}: VALUE = {cell.value} | Bold: {is_bold} | Fill: {has_fill}")
            
            # Get current and previous values for color coding analysis
            current_value = ws.cell(row=monthly_row, column=ws.max_column).value
            previous_value = ws.cell(row=monthly_row, column=ws.max_column - 1).value if ws.max_column > 1 else None
            
            print(f"\nCOLOR CODING ANALYSIS:")
            print(f"  Current value: {current_value}")
            print(f"  Previous value: {previous_value}")
            
            if isinstance(current_value, (int, float)) and isinstance(previous_value, (int, float)):
                if current_value > previous_value:
                    expected_color = "Green (#90EE90) - Increase"
                elif current_value < previous_value:
                    expected_color = "Red (#FF7C80) - Decrease"  
                else:
                    expected_color = "Yellow (#FFFF00) - Same"
                print(f"  Expected color: {expected_color}")
        else:
            print("ERROR: Monthly Average row not found")
        
        wb.close()
        
    except Exception as e:
        print(f"ERROR: {e}")
        import traceback
        traceback.print_exc()

if __name__ == "__main__":
    test_monthly_average_formatting()