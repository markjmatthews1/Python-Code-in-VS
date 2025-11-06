#!/usr/bin/env python3
"""
Create Weekly Dividend Statistics Column for Portfolio Summary
Incorporates dividend totals by week, month, yearly and withdrawal data
"""
import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from datetime import datetime
import calendar

def create_dividend_stats_column():
    """Create a new column with comprehensive dividend statistics"""
    
    excel_file = 'outputs/Dividends_2025.xlsx'
    
    try:
        print("📊 Creating weekly dividend statistics column...")
        
        wb = openpyxl.load_workbook(excel_file, data_only=True)
        
        # Get dividend data from Estimated Income sheet
        income_ws = wb['Estimated Income 2025']
        portfolio_ws = wb['Portfolio Summary']
        
        # Define styles
        header_font = Font(name='Arial', size=12, bold=True, color='FFFFFF')
        header_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
        bold_font = Font(name='Arial', size=12, bold=True)
        normal_font = Font(name='Arial', size=12)
        currency_format = '$#,##0.00'
        light_blue_fill = PatternFill(start_color='D9E2F3', end_color='D9E2F3', fill_type='solid')
        
        # Get current withdrawal data
        monthly_etrade_ira = portfolio_ws.cell(row=32, column=2).value or 0
        quarterly_etrade_ira = portfolio_ws.cell(row=33, column=2).value or 0
        monthly_etrade_taxable = portfolio_ws.cell(row=34, column=2).value or 0
        
        print(f"📋 Withdrawal data:")
        print(f"   Monthly E*TRADE IRA: ${monthly_etrade_ira}")
        print(f"   Quarterly E*TRADE IRA: ${quarterly_etrade_ira}")
        print(f"   Monthly E*TRADE Taxable: ${monthly_etrade_taxable}")
        
        # Get latest dividend data from Estimated Income sheet
        max_col = income_ws.max_column
        latest_monthly_estimate = income_ws.cell(row=9, column=max_col).value or 0
        
        # Calculate annual estimate
        latest_annual_estimate = latest_monthly_estimate * 12
        
        # Calculate weekly estimate
        weekly_estimate = latest_monthly_estimate / 4.33  # Average weeks per month
        
        # Get account breakdowns from latest column
        etrade_ira_annual = income_ws.cell(row=4, column=max_col).value or 0
        etrade_taxable_annual = income_ws.cell(row=5, column=max_col).value or 0
        schwab_ira_annual = income_ws.cell(row=6, column=max_col).value or 0
        schwab_individual_annual = income_ws.cell(row=7, column=max_col).value or 0
        
        print(f"\n💰 Latest dividend estimates:")
        print(f"   Weekly: ${weekly_estimate:,.2f}")
        print(f"   Monthly: ${latest_monthly_estimate:,.2f}")
        print(f"   Annual: ${latest_annual_estimate:,.2f}")
        
        # Calculate net income (after withdrawals)
        total_monthly_withdrawals = monthly_etrade_ira + monthly_etrade_taxable + (quarterly_etrade_ira / 3)
        net_monthly = latest_monthly_estimate - total_monthly_withdrawals
        net_annual = net_monthly * 12
        
        print(f"\n📊 Net income calculations:")
        print(f"   Total monthly withdrawals: ${total_monthly_withdrawals:,.2f}")
        print(f"   Net monthly (after withdrawals): ${net_monthly:,.2f}")
        print(f"   Net annual (after withdrawals): ${net_annual:,.2f}")
        
        # Start adding dividend statistics in column D (after existing content)
        current_row = 36  # Start after existing content
        col = 4  # Column D
        
        # === DIVIDEND STATISTICS HEADER ===
        cell = portfolio_ws.cell(row=current_row, column=col)
        cell.value = "WEEKLY DIVIDEND STATISTICS"
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center')
        
        # Merge cells for header
        portfolio_ws.merge_cells(f'{chr(64 + col)}{current_row}:{chr(64 + col + 1)}{current_row}')
        
        current_row += 2
        
        # === CURRENT ESTIMATES SECTION ===
        sections = [
            ("CURRENT ESTIMATES", [
                ("Weekly Estimate:", weekly_estimate),
                ("Monthly Estimate:", latest_monthly_estimate),
                ("Annual Estimate:", latest_annual_estimate),
            ]),
            ("", []),  # Spacer
            ("ACCOUNT BREAKDOWN (Annual)", [
                ("E*TRADE IRA:", etrade_ira_annual),
                ("E*TRADE Taxable:", etrade_taxable_annual),
                ("Schwab IRA:", schwab_ira_annual),
                ("Schwab Individual:", schwab_individual_annual),
            ]),
            ("", []),  # Spacer
            ("WITHDRAWAL DATA", [
                ("Monthly E*TRADE IRA:", monthly_etrade_ira),
                ("Quarterly E*TRADE IRA:", quarterly_etrade_ira),
                ("Monthly E*TRADE Taxable:", monthly_etrade_taxable),
                ("Total Monthly Withdrawals:", total_monthly_withdrawals),
            ]),
            ("", []),  # Spacer
            ("NET INCOME (After Withdrawals)", [
                ("Net Monthly Income:", net_monthly),
                ("Net Annual Income:", net_annual),
                ("Reinvestment Rate:", f"{(net_annual / latest_annual_estimate * 100):.1f}%" if latest_annual_estimate > 0 else "0%"),
            ]),
            ("", []),  # Spacer
            ("DIVIDEND METRICS", [
                ("Current Yield:", f"{(latest_annual_estimate / 519000 * 100):.2f}%" if True else "Calculate from portfolio"),  # Approximate portfolio value
                ("Monthly Dividend Coverage:", f"{latest_monthly_estimate / total_monthly_withdrawals:.1f}x" if total_monthly_withdrawals > 0 else "N/A"),
                ("Annual Div Growth Target:", "5-7%"),
                ("Next Update:", "Weekly (Automated)"),
            ])
        ]
        
        for section_title, items in sections:
            if section_title:  # Not a spacer
                # Section header
                cell = portfolio_ws.cell(row=current_row, column=col)
                cell.value = section_title
                cell.font = Font(name='Arial', size=11, bold=True)
                cell.fill = light_blue_fill
                current_row += 1
                
                # Section items
                for label, value in items:
                    # Label
                    label_cell = portfolio_ws.cell(row=current_row, column=col)
                    label_cell.value = label
                    label_cell.font = normal_font
                    
                    # Value
                    value_cell = portfolio_ws.cell(row=current_row, column=col + 1)
                    
                    if isinstance(value, (int, float)):
                        value_cell.value = value
                        value_cell.number_format = currency_format
                    else:
                        value_cell.value = value
                    
                    value_cell.font = normal_font
                    
                    current_row += 1
            
            current_row += 1  # Extra spacing between sections
        
        print(f"\n✅ Dividend statistics added to Portfolio Summary!")
        print(f"   Added {current_row - 36} rows of dividend data")
        print(f"   Location: Column D, starting at row 36")
        
        # Save the workbook
        wb.save(excel_file)
        wb.close()
        
        return True
        
    except Exception as e:
        print(f"❌ Error creating dividend statistics: {e}")
        return False

def update_dividend_stats():
    """Update the dividend statistics with latest data"""
    print("🚀 CREATING WEEKLY DIVIDEND STATISTICS COLUMN")
    print("=" * 60)
    
    success = create_dividend_stats_column()
    
    if success:
        print("\n🎯 Summary:")
        print("✅ Added comprehensive dividend statistics to Portfolio Summary")
        print("✅ Includes weekly, monthly, and annual estimates")
        print("✅ Incorporates withdrawal data from rows 32-34")
        print("✅ Shows net income after withdrawals")
        print("✅ Provides dividend coverage metrics")
        print("\n📍 Location: Column D in Portfolio Summary sheet")
    else:
        print("\n❌ Failed to create dividend statistics column")

if __name__ == "__main__":
    update_dividend_stats()
