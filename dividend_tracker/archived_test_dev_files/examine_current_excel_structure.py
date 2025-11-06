#!/usr/bin/env python3
"""
Examine Current Excel Structure
==============================
Let's see exactly what we have in the restored Excel file
so we can build proper append-only updates.
"""

import openpyxl
import os

def examine_excel_structure():
    """Examine the current Excel file structure"""
    
    excel_file = r"c:\Users\mjmat\Python Code in VS\dividend_tracker\DividendTrackerApp\outputs\Dividends_2025.xlsx"
    
    if not os.path.exists(excel_file):
        print(f"❌ Excel file not found: {excel_file}")
        return
    
    print("📊 EXAMINING CURRENT EXCEL STRUCTURE")
    print("=" * 50)
    print(f"📁 File: {os.path.basename(excel_file)}")
    
    try:
        wb = openpyxl.load_workbook(excel_file)
        
        print(f"📋 Total Sheets: {len(wb.sheetnames)}")
        print(f"📋 Sheet Names: {wb.sheetnames}")
        
        for sheet_name in wb.sheetnames:
            print(f"\n📄 SHEET: {sheet_name}")
            print("-" * 30)
            
            ws = wb[sheet_name]
            
            print(f"   📏 Dimensions: {ws.max_row} rows × {ws.max_column} columns")
            
            # Show headers (first row)
            if ws.max_row >= 1:
                headers = []
                for col in range(1, min(ws.max_column + 1, 11)):  # Max 10 columns
                    cell_value = ws.cell(row=1, column=col).value
                    if cell_value is not None:
                        headers.append(str(cell_value))
                    else:
                        headers.append("")
                
                print(f"   📊 Headers: {headers}")
            
            # Show first few data rows
            if ws.max_row >= 2:
                print(f"   📈 Sample Data (rows 2-{min(ws.max_row, 4)}):")
                for row in range(2, min(ws.max_row + 1, 5)):  # Show max 3 data rows
                    row_data = []
                    for col in range(1, min(ws.max_column + 1, 6)):  # Max 5 columns
                        cell_value = ws.cell(row=row, column=col).value
                        if cell_value is not None:
                            if isinstance(cell_value, (int, float)):
                                row_data.append(f"{cell_value:,.2f}" if cell_value > 1000 else str(cell_value))
                            else:
                                row_data.append(str(cell_value)[:20])  # Truncate long strings
                        else:
                            row_data.append("")
                    
                    print(f"      Row {row}: {row_data}")
            
            # Check for any data patterns
            if ws.max_row > 1:
                # Look for date patterns in first column
                first_col_values = []
                for row in range(2, min(ws.max_row + 1, 10)):
                    cell_value = ws.cell(row=row, column=1).value
                    if cell_value:
                        first_col_values.append(str(cell_value))
                
                if first_col_values:
                    print(f"   📅 First Column Pattern: {first_col_values[:3]}...")
        
        wb.close()
        
    except Exception as e:
        print(f"❌ Error examining Excel file: {e}")
        import traceback
        traceback.print_exc()

if __name__ == "__main__":
    examine_excel_structure()
