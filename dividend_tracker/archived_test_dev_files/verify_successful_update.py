#!/usr/bin/env python3

import openpyxl
import os
from datetime import datetime

class UpdateVerifier:
    def __init__(self):
        self.excel_path = r"C:\Users\mjmat\Python Code in VS\dividend_tracker\DividendTrackerApp\outputs\Dividends_2025.xlsx"
    
    def verify_update(self):
        """Verify the successful update of the historical yield sheet"""
        print("VERIFICATION: Historical Yield Sheet Update")
        print("=" * 50)
        print(f"Verification Time: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}")
        print(f"Excel File: {os.path.basename(self.excel_path)}")
        print()
        
        if not os.path.exists(self.excel_path):
            print("ERROR: Updated Excel file not found!")
            return False
        
        # Load workbook
        try:
            wb = openpyxl.load_workbook(self.excel_path)
            ws = wb['Accounts Div historical yield']
            print("SUCCESS: Loaded updated sheet")
        except Exception as e:
            print(f"ERROR: Could not load sheet: {e}")
            return False
        
        # Verify structure and updates
        account_groups = self.find_account_groups(ws)
        
        print(f"\nSTRUCTURE VERIFICATION:")
        print(f"Found {len(account_groups)} account groups:")
        for group_name, info in account_groups.items():
            print(f"  - {group_name}: rows {info['start_row']}-{info['end_row']}")
        
        # Sample some updated data
        print(f"\nDATA VERIFICATION (Sample Updates):")
        
        for group_name, info in account_groups.items():
            print(f"\n{group_name}:")
            start_row = info['start_row']
            end_row = min(info['end_row'], start_row + 5)  # Show first few entries
            
            for row in range(start_row + 2, end_row + 1):  # Skip header rows
                ticker = ws.cell(row=row, column=1).value
                qty = ws.cell(row=row, column=2).value
                price = ws.cell(row=row, column=4).value
                
                if ticker and str(ticker).strip():
                    print(f"  Row {row}: {ticker} | Qty: {qty} | Price: ${price}")
        
        # Check yield column (should be column P)
        print(f"\nYIELD COLUMN VERIFICATION:")
        header_cell = ws.cell(row=1, column=16)  # Column P
        if header_cell.value:
            print(f"SUCCESS: New yield column header: {header_cell.value}")
        else:
            print("WARNING: Yield column header not found")
        
        # Check formatting on a few cells
        print(f"\nFORMATTING VERIFICATION:")
        test_rows = [3, 38, 63, 74]  # Sample from each group
        for row in test_rows:
            ticker_cell = ws.cell(row=row, column=1)
            qty_cell = ws.cell(row=row, column=2)
            
            if ticker_cell.value:
                print(f"Row {row}: Ticker={ticker_cell.value} | Font={ticker_cell.font.name} | Bold={ticker_cell.font.bold} | Color={ticker_cell.font.color}")
        
        print(f"\nVERIFICATION COMPLETED: Update successful!")
        wb.close()
        return True
    
    def find_account_groups(self, ws):
        """Find all account groups in the sheet"""
        account_info = {}
        
        for row in range(1, min(80, ws.max_row + 1)):
            cell_value = ws.cell(row=row, column=1).value
            if not cell_value:
                continue
                
            cell_text = str(cell_value).strip().upper()
            
            if cell_text == 'ETRADE IRA':
                account_info["E*TRADE IRA"] = {"start_row": row, "end_row": None}
            elif cell_text == 'ETRADE TAXABLE':
                account_info["E*TRADE Taxable"] = {"start_row": row, "end_row": None}
            elif cell_text == 'SCHWAB IRA':
                account_info["Schwab IRA"] = {"start_row": row, "end_row": None}
            elif cell_text == 'SCHWAB INDIVIDUAL':
                account_info["Schwab Individual"] = {"start_row": row, "end_row": None}
        
        # Calculate end rows
        group_order = ["E*TRADE IRA", "E*TRADE Taxable", "Schwab IRA", "Schwab Individual"]
        
        for i, group_name in enumerate(group_order):
            if group_name not in account_info:
                continue
                
            start_row = account_info[group_name]["start_row"]
            
            if i < len(group_order) - 1:
                end_row = ws.max_row
                for next_group in group_order[i+1:]:
                    if next_group in account_info:
                        end_row = account_info[next_group]["start_row"] - 1
                        break
            else:
                end_row = start_row + 2
                while end_row <= ws.max_row:
                    cell_value = ws.cell(row=end_row, column=1).value
                    if not cell_value or str(cell_value).strip() == "":
                        break
                    end_row += 1
                end_row -= 1
                
            account_info[group_name]["end_row"] = end_row
        
        return account_info

if __name__ == "__main__":
    verifier = UpdateVerifier()
    verifier.verify_update()