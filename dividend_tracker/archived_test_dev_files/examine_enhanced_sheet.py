#!/usr/bin/env python3
"""
Examine the enhanced "Etrade IRA historic yield" sheet layout
to understand the new multi-account structure.
"""

import os
import openpyxl

def examine_enhanced_sheet():
    """Examine the updated sheet structure with multiple accounts"""
    
    workbook_path = r"c:\Users\mjmat\Python Code in VS\dividend_tracker\DividendTrackerApp\outputs\Dividends_2025.xlsx"
    
    try:
        print("🔍 EXAMINING ENHANCED SHEET LAYOUT")
        print("=" * 60)
        
        # Load workbook
        workbook = openpyxl.load_workbook(workbook_path)
        
        # Check if sheet name changed
        sheet_names = workbook.sheetnames
        print(f"📋 Available sheets: {sheet_names}")
        
        # Try to find the historic yield sheet
        target_sheet = None
        if "Accounts dividend historical yield" in sheet_names:
            target_sheet = workbook["Accounts dividend historical yield"]
            print(f"✅ Found renamed sheet: 'Accounts dividend historical yield'")
        elif "Etrade IRA historic yield" in sheet_names:
            target_sheet = workbook["Etrade IRA historic yield"]
            print(f"✅ Found original sheet: 'Etrade IRA historic yield'")
        else:
            print("❌ Could not find historic yield sheet")
            return
            
        sheet = target_sheet
        
        print(f"📏 Sheet dimensions: {sheet.max_row} rows × {sheet.max_column} columns")
        print("\n🔍 EXAMINING SHEET STRUCTURE:")
        print("-" * 60)
        
        # Examine first 50 rows to understand the new layout
        for row in range(1, min(51, sheet.max_row + 1)):
            row_data = []
            has_content = False
            
            # Check first 20 columns for content
            for col in range(1, min(21, sheet.max_column + 1)):
                cell = sheet.cell(row=row, column=col)
                value = cell.value
                
                if value is not None:
                    has_content = True
                    # Show cell formatting info
                    fill_color = "No Fill"
                    if cell.fill and hasattr(cell.fill, 'start_color') and cell.fill.start_color.rgb:
                        fill_color = cell.fill.start_color.rgb
                    
                    row_data.append(f"Col{col}: '{value}' ({fill_color})")
                else:
                    if col <= 15:  # Only show first 15 columns if empty
                        row_data.append(f"Col{col}: [empty]")
            
            if has_content:
                print(f"Row {row:2d}: {' | '.join(row_data[:10])}")  # Show first 10 columns
                if len(row_data) > 10:
                    print(f"      {''.join([' '] * 8)}{' | '.join(row_data[10:])}")
        
        # Look for account sections and patterns
        print(f"\n📊 LOOKING FOR ACCOUNT SECTIONS:")
        print("-" * 40)
        
        account_sections = []
        for row in range(1, sheet.max_row + 1):
            cell_a = sheet.cell(row=row, column=1).value
            if cell_a and isinstance(cell_a, str):
                if any(keyword in cell_a.upper() for keyword in ['ETRADE', 'SCHWAB', 'ACCOUNT', 'IRA', 'ROTH']):
                    account_sections.append(f"Row {row}: {cell_a}")
        
        for section in account_sections:
            print(f"  🏦 {section}")
            
        workbook.close()
        
    except Exception as e:
        print(f"❌ Error examining sheet: {e}")

if __name__ == "__main__":
    examine_enhanced_sheet()
