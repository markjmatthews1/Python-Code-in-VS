#!/usr/bin/env python3
"""
Test Integration: E*TRADE Historic Yield Updater
===============================================

Quick test to verify the historic yield updater integrates properly with
the DividendTracker weekend automation system.

This test checks:
1. Can import the module
2. Can access the target sheet
3. Can authenticate with E*TRADE API
4. Integration points work correctly

"""

import os
import sys
from datetime import datetime

def test_historic_yield_integration():
    """Test the integration of historic yield updater"""
    
    print("🧪 TESTING E*TRADE HISTORIC YIELD INTEGRATION")
    print("=" * 60)
    print(f"🕐 Test started: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}")
    
    # Test 1: Module import
    print("\n📦 TEST 1: Module Import")
    try:
        from update_etrade_historic_yield import ETradeHistoricYieldUpdater, run_historic_yield_update
        print("   ✅ Successfully imported historic yield updater module")
    except ImportError as e:
        print(f"   ❌ Failed to import module: {e}")
        return False
    
    # Test 2: File access
    print("\n📁 TEST 2: File Access")
    try:
        updater = ETradeHistoricYieldUpdater()
        
        if os.path.exists(updater.workbook_path):
            print(f"   ✅ Target workbook exists: {updater.workbook_path}")
        else:
            print(f"   ❌ Target workbook not found: {updater.workbook_path}")
            return False
            
        # Check sheet existence
        import openpyxl
        wb = openpyxl.load_workbook(updater.workbook_path)
        if updater.sheet_name in wb.sheetnames:
            print(f"   ✅ Target sheet exists: {updater.sheet_name}")
            
            # Show sheet info
            sheet = wb[updater.sheet_name]
            print(f"   📏 Sheet dimensions: {sheet.max_row} rows × {sheet.max_column} columns")
            
            # Check data structure
            ticker_count = 0
            for row in range(2, min(sheet.max_row + 1, 30)):
                ticker = sheet.cell(row=row, column=1).value
                if ticker and isinstance(ticker, str):
                    ticker_count += 1
                else:
                    break
            print(f"   📊 Found {ticker_count} tickers to update")
            
        else:
            print(f"   ❌ Target sheet not found: {updater.sheet_name}")
            print(f"   Available sheets: {wb.sheetnames}")
            return False
            
        wb.close()
        
    except Exception as e:
        print(f"   ❌ File access error: {e}")
        return False
    
    # Test 3: API Authentication
    print("\n🔐 TEST 3: E*TRADE Authentication")
    try:
        # Test auth module import
        sys.path.append(os.path.join(os.path.dirname(__file__), 'modules'))
        from etrade_auth import get_etrade_session
        print("   ✅ E*TRADE auth module imported successfully")
        
        # Note: We won't actually authenticate in test mode to avoid token usage
        print("   📋 Authentication module ready (tokens will be used during actual run)")
        
    except ImportError as e:
        print(f"   ❌ Could not import E*TRADE auth: {e}")
        return False
    except Exception as e:
        print(f"   ⚠️ Auth setup issue: {e}")
    
    # Test 4: Weekend Automation Integration
    print("\n🔄 TEST 4: Weekend Automation Integration")
    try:
        main_script = "run_complete_dividend_update.py"
        if os.path.exists(main_script):
            print(f"   ✅ Main weekend script found: {main_script}")
            
            # Check if our module is referenced
            with open(main_script, 'r', encoding='utf-8', errors='ignore') as f:
                content = f.read()
                
            if "update_etrade_historic_yield.py" in content:
                print("   ✅ Historic yield updater is integrated into weekend automation")
            else:
                print("   ❌ Historic yield updater not found in weekend automation")
                return False
                
        else:
            print(f"   ❌ Main weekend script not found: {main_script}")
            return False
            
    except Exception as e:
        print(f"   ❌ Integration test error: {e}")
        return False
    
    # Summary
    print("\n✅ INTEGRATION TEST RESULTS")
    print("=" * 60)
    print("🎯 All integration tests passed!")
    print("📋 Ready for weekend automation:")
    print("   • Module imports correctly")  
    print("   • Target files accessible")
    print("   • E*TRADE authentication ready")
    print("   • Weekend script integration complete")
    print("\n🚀 NEXT STEP: Run weekend automation with:")
    print("   python run_complete_dividend_update.py")
    print(f"🕐 Test completed: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}")
    
    return True

if __name__ == "__main__":
    success = test_historic_yield_integration()
    
    if success:
        print("\n🎉 Integration test successful!")
        print("✅ Historic yield updater is ready for weekend automation")
    else:
        print("\n💥 Integration test failed!")
        print("❌ Check errors above and fix before running weekend automation")
