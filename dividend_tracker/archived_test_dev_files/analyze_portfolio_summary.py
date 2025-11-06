import openpyxl
from openpyxl.styles import Font, PatternFill
import json

def analyze_portfolio_summary():
    """Analyze Portfolio Summary sheet structure, data, and formatting"""
    
    file_path = "outputs/Dividends_2025.xlsx"
    
    try:
        wb = openpyxl.load_workbook(file_path)
        ws = wb["Portfolio Summary"]
        
        print("PORTFOLIO SUMMARY SHEET ANALYSIS")
        print("=" * 50)
        
        # Get sheet dimensions
        print(f"Sheet dimensions: {ws.max_row} rows x {ws.max_column} columns")
        print()
        
        # Analyze first 30 rows to understand structure
        print("SHEET STRUCTURE (First 30 rows):")
        print("-" * 40)
        
        summary_data = []
        
        for row in range(1, min(31, ws.max_row + 1)):
            row_data = []
            has_content = False
            
            for col in range(1, min(6, ws.max_column + 1)):  # Check first 5 columns
                cell = ws.cell(row=row, column=col)
                value = cell.value
                
                # Get formatting info
                font_color = cell.font.color.rgb if cell.font.color else "None"
                bg_color = cell.fill.start_color.rgb if cell.fill.start_color else "None"
                is_bold = cell.font.bold if cell.font.bold else False
                
                if value is not None:
                    has_content = True
                    cell_info = {
                        'value': value,
                        'font_color': font_color,
                        'bg_color': bg_color,
                        'bold': is_bold,
                        'column': col
                    }
                    row_data.append(cell_info)
            
            if has_content:
                summary_data.append({
                    'row': row,
                    'cells': row_data
                })
                
                # Print row summary
                row_display = f"Row {row:2d}: "
                for cell_info in row_data:
                    col_letter = chr(64 + cell_info['column'])  # Convert to A, B, C, etc.
                    value_str = str(cell_info['value'])[:20]  # Truncate long values
                    
                    formatting = ""
                    if cell_info['bold']:
                        formatting += "B"
                    if cell_info['bg_color'] != "None":
                        formatting += f"BG:{cell_info['bg_color'][:6]}"
                    if cell_info['font_color'] != "None":
                        formatting += f"FC:{cell_info['font_color'][:6]}"
                    
                    if formatting:
                        row_display += f"{col_letter}:{value_str}({formatting}) | "
                    else:
                        row_display += f"{col_letter}:{value_str} | "
                
                print(row_display)
        
        print("\nCOLUMN ANALYSIS:")
        print("-" * 20)
        
        # Analyze each column
        for col in range(1, 6):
            col_letter = chr(64 + col)
            print(f"\nColumn {col_letter}:")
            
            non_empty_count = 0
            sample_values = []
            
            for row in range(1, min(31, ws.max_row + 1)):
                cell = ws.cell(row=row, column=col)
                if cell.value is not None:
                    non_empty_count += 1
                    if len(sample_values) < 5:
                        sample_values.append((row, cell.value))
            
            print(f"  Non-empty cells: {non_empty_count}")
            print(f"  Sample values: {sample_values}")
        
        # Look for calculation patterns
        print("\nNUMERIC DATA ANALYSIS:")
        print("-" * 25)
        
        numeric_data = []
        for row in range(1, min(31, ws.max_row + 1)):
            for col in range(1, 6):
                cell = ws.cell(row=row, column=col)
                if isinstance(cell.value, (int, float)) and cell.value != 0:
                    numeric_data.append({
                        'row': row,
                        'col': chr(64 + col),
                        'value': cell.value,
                        'formatted_value': f"{cell.value:,.2f}" if isinstance(cell.value, float) else str(cell.value)
                    })
        
        for data in numeric_data[:15]:  # Show first 15 numeric values
            print(f"  {data['col']}{data['row']}: {data['formatted_value']}")
        
        wb.close()
        
        # Save analysis for reference
        with open('portfolio_summary_analysis.json', 'w') as f:
            json.dump({
                'structure': summary_data,
                'numeric_data': numeric_data
            }, f, indent=2, default=str)
        
        print(f"\nAnalysis saved to portfolio_summary_analysis.json")
        print("This will help create an updater that preserves all formatting!")
        
    except Exception as e:
        print(f"ERROR: {e}")

if __name__ == "__main__":
    analyze_portfolio_summary()