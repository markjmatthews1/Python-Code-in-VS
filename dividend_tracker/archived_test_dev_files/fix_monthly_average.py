import os
import sys
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.styles.numbers import FORMAT_CURRENCY_USD_SIMPLE

def fix_monthly_average_formatting():
    """Fix the Monthly Average formatting to be bold and color coded"""
    
    excel_file = "outputs/Dividends_2025.xlsx"
    if not os.path.exists(excel_file):
        print(f"ERROR: Excel file not found: {excel_file}")
        return
    
    try:
        wb = openpyxl.load_workbook(excel_file)
        ws = wb["Estimated Income 2025"]
        
        # Find Monthly Average row
        monthly_row = None
        for row in range(8, 12):
            cell_value = ws.cell(row=row, column=1).value
            if cell_value and 'monthly' in str(cell_value).lower():
                monthly_row = row
                break
        
        if not monthly_row:
            print("ERROR: Monthly Average row not found")
            return
        
        print(f"FIXING MONTHLY AVERAGE FORMATTING - ROW {monthly_row}")
        print("=" * 50)
        
        # Get the last column (current data)
        last_col = ws.max_column
        current_cell = ws.cell(row=monthly_row, column=last_col)
        
        # Get previous column for comparison
        previous_cell = ws.cell(row=monthly_row, column=last_col - 1) if last_col > 1 else None
        
        print(f"Current column: {openpyxl.utils.get_column_letter(last_col)}")
        print(f"Current formula: {current_cell.value}")
        
        # Calculate the values for color comparison
        # We need to evaluate the formulas to get actual values
        wb.save(excel_file)  # Save first to ensure formulas are calculated
        wb.close()
        wb = openpyxl.load_workbook(excel_file)
        ws = wb["Estimated Income 2025"]
        
        current_cell = ws.cell(row=monthly_row, column=last_col)
        previous_cell = ws.cell(row=monthly_row, column=last_col - 1) if last_col > 1 else None
        
        current_value = current_cell.value
        previous_value = previous_cell.value if previous_cell else None
        
        print(f"Current calculated value: ${current_value:,.2f}" if isinstance(current_value, (int, float)) else f"Current value: {current_value}")
        print(f"Previous calculated value: ${previous_value:,.2f}" if isinstance(previous_value, (int, float)) else f"Previous value: {previous_value}")
        
        # Apply formatting: Bold + Color coding
        if isinstance(current_value, (int, float)):
            # Keep the formula but apply bold formatting
            formula = f"=SUM({openpyxl.utils.get_column_letter(last_col)}4:{openpyxl.utils.get_column_letter(last_col)}7)/12"
            current_cell.value = formula
            current_cell.number_format = FORMAT_CURRENCY_USD_SIMPLE
            current_cell.font = Font(name='Arial', size=12, bold=True)
            
            # Apply color coding based on comparison
            if previous_value is not None and isinstance(previous_value, (int, float)):
                if current_value > previous_value:
                    # Increase - Green
                    current_cell.fill = PatternFill(start_color='90EE90', end_color='90EE90', fill_type='solid')
                    color_desc = "Green (Increase)"
                elif current_value < previous_value:
                    # Decrease - Red
                    current_cell.fill = PatternFill(start_color='FF7C80', end_color='FF7C80', fill_type='solid')
                    color_desc = "Red (Decrease)"
                else:
                    # Same - Yellow
                    current_cell.fill = PatternFill(start_color='FFFF00', end_color='FFFF00', fill_type='solid')
                    color_desc = "Yellow (Same)"
            else:
                color_desc = "No color (No previous value)"
                
            print(f"✅ Applied: Bold font + {color_desc}")
            
        # Save the changes
        wb.save(excel_file)
        wb.close()
        print(f"✅ Monthly Average formatting updated in {excel_file}")
        
    except Exception as e:
        print(f"ERROR: {e}")
        if 'PermissionError' in str(e):
            print("NOTE: Please close the Excel file and try again")
        import traceback
        traceback.print_exc()

if __name__ == "__main__":
    fix_monthly_average_formatting()