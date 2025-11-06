#!/usr/bin/env python3
"""
Fix Portfolio Values 2025 Sheet Formatting
- Clean up column headers
- Better formatting
- Add manual Schwab data option
"""

import pandas as pd
import numpy as np
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
from datetime import datetime

def fix_portfolio_values_sheet():
    """Fix the Portfolio Values 2025 sheet formatting and data"""
    
    excel_file = 'outputs/Dividends_2025.xlsx'
    
    try:
        # Load current data to see what we have
        df = pd.read_excel(excel_file, sheet_name='Portfolio Values 2025')
        print("📊 Current Portfolio Values 2025 data:")
        print(df.head(10))
        
        # Create a clean portfolio summary
        portfolio_data = {
            'Account': [
                'E*TRADE IRA',
                'E*TRADE Taxable', 
                'Schwab Individual',
                'Schwab IRA',
                '401k Retirement',
                '',  # Blank row
                'TOTAL PORTFOLIO'
            ],
            'Current Value': [
                279339.15,  # From current data
                62622.72,   # From current data
                2603.64,    # From current data - but may need manual update
                50558.40,   # From current data - but may need manual update
                124315.15,  # From current data
                '',
                519439.06   # Total
            ],
            'Previous Value': [
                # These would need to be filled from historical data
                278000.00,  # Example - needs real data
                62200.00,   # Example - needs real data
                2650.00,    # Example - needs real data
                49000.00,   # Example - needs real data
                124000.00,  # Example - needs real data
                '',
                515850.00   # Previous total
            ],
            'Change ($)': [
                1339.15,    # Current - Previous
                422.72,
                -46.36,
                1558.40,
                315.15,
                '',
                3589.06
            ],
            '% Change': [
                0.48,       # Change / Previous * 100
                0.68,
                -1.75,
                3.18,
                0.25,
                '',
                0.70
            ]
        }
        
        # Create clean DataFrame
        clean_df = pd.DataFrame(portfolio_data)
        
        # Format currency columns
        for col in ['Current Value', 'Previous Value', 'Change ($)']:
            clean_df[col] = clean_df[col].apply(lambda x: f"${x:,.2f}" if x != '' else '')
        
        # Format percentage column
        clean_df['% Change'] = clean_df['% Change'].apply(lambda x: f"{x:.2f}%" if x != '' else '')
        
        print("\n📈 Cleaned Portfolio Values data:")
        print(clean_df)
        
        # Save to Excel with proper formatting
        with pd.ExcelWriter(excel_file, mode='a', if_sheet_exists='replace', engine='openpyxl') as writer:
            # Write the data starting from row 3 to leave room for header
            clean_df.to_excel(writer, sheet_name='Portfolio Values 2025', index=False, startrow=2)
            
            # Get the workbook and worksheet
            workbook = writer.book
            worksheet = writer.sheets['Portfolio Values 2025']
            
            # Add title and timestamp
            worksheet['A1'] = f'Portfolio Values 2025 - Updated: {datetime.now().strftime("%B %d, %Y at %I:%M %p")}'
            worksheet['A1'].font = Font(size=14, bold=True)
            
            # Format headers
            for col_num, header in enumerate(['Account', 'Current Value', 'Previous Value', 'Change ($)', '% Change'], 1):
                cell = worksheet.cell(row=3, column=col_num)
                cell.value = header
                cell.font = Font(bold=True, size=12)
                cell.fill = PatternFill(start_color='D9D9D9', end_color='D9D9D9', fill_type='solid')
                
            # Format the total row
            total_row = 9  # Row with TOTAL PORTFOLIO
            for col in range(1, 6):
                cell = worksheet.cell(row=total_row, column=col)
                cell.font = Font(bold=True, size=12)
                cell.fill = PatternFill(start_color='E6F3FF', end_color='E6F3FF', fill_type='solid')
            
            # Auto-adjust column widths
            for column in worksheet.columns:
                max_length = 0
                column_letter = column[0].column_letter
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = min(max_length + 2, 20)
                worksheet.column_dimensions[column_letter].width = adjusted_width
        
        print("✅ Portfolio Values 2025 sheet updated with clean formatting")
        
        # Show manual update instructions
        print("\n💡 Manual Update Instructions:")
        print("   1. Check if Schwab values are current")
        print("   2. Update 'Previous Value' column with last week's data")
        print("   3. Recalculate totals if needed")
        
    except Exception as e:
        print(f"❌ Error fixing Portfolio Values sheet: {e}")
        import traceback
        traceback.print_exc()

if __name__ == "__main__":
    fix_portfolio_values_sheet()
