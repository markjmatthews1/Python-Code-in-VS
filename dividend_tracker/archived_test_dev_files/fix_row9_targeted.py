"""
Final Row 9 Fix - Targeted

Fix ONLY the Row 9 calculation and add color coding to it.
Don't touch anything else.
"""
import openpyxl
from openpyxl.styles import Font, PatternFill
from openpyxl.styles.numbers import FORMAT_CURRENCY_USD_SIMPLE

GREEN_COLOR = '00FF00'    # Green for increase
RED_COLOR = 'FF7C80'      # Red for decrease  
YELLOW_COLOR = 'FFFF00'   # Yellow for same

def fix_row9_targeted():
    """Fix Row 9 calculation and color coding only"""
    excel_file = r"c:\Users\mjmat\Python Code in VS\dividend_tracker\DividendTrackerApp\outputs\Dividends_2025.xlsx"
    
    try:
        wb = openpyxl.load_workbook(excel_file)
        ws = wb["Estimated Income 2025"]
        
        max_col = ws.max_column
        current_col = max_col
        prev_col = max_col - 1
        
        current_col_letter = openpyxl.utils.get_column_letter(current_col)
        
        print("🧮 FIXING ROW 9 CALCULATION ONLY")
        print("=" * 35)
        
        # 1. Set correct formula for Row 9
        formula = f"=SUM({current_col_letter}4:{current_col_letter}7)/12"
        
        row9_cell = ws.cell(row=9, column=current_col)
        row9_cell.value = formula
        row9_cell.number_format = FORMAT_CURRENCY_USD_SIMPLE
        row9_cell.font = Font(name='Arial', size=12)
        
        print(f"✅ Row 9 formula: {formula}")
        
        # Save to calculate formula
        wb.save(excel_file)
        wb.close()
        
        # 2. Reopen with calculated values for color coding
        wb = openpyxl.load_workbook(excel_file, data_only=True)
        ws = wb["Estimated Income 2025"]
        
        # Get calculated values for color comparison
        current_val = ws.cell(row=9, column=current_col).value
        prev_val = ws.cell(row=9, column=prev_col).value
        
        print(f"📊 Row 9 values: Previous=${prev_val}, Current=${current_val}")
        
        # Apply color coding to Row 9
        if (current_val is not None and prev_val is not None and
            isinstance(current_val, (int, float)) and isinstance(prev_val, (int, float))):
            
            if current_val > prev_val:
                color = GREEN_COLOR
                change_type = "INCREASE (Green)"
            elif current_val < prev_val:
                color = RED_COLOR
                change_type = "DECREASE (Red)"
            else:
                color = YELLOW_COLOR
                change_type = "SAME (Yellow)"
            
            # Apply the color
            ws.cell(row=9, column=current_col).fill = PatternFill(
                start_color=color, end_color=color, fill_type='solid'
            )
            
            print(f"🎨 Row 9 color applied: {change_type}")
        else:
            print("⚠️ Row 9 color not applied - insufficient data for comparison")
        
        # Save the final result
        wb.save(excel_file)
        wb.close()
        
        print("\n✅ ROW 9 TARGETED FIX COMPLETE!")
        
        return True
        
    except Exception as e:
        print(f"❌ ERROR: {e}")
        import traceback
        traceback.print_exc()
        return False

if __name__ == "__main__":
    fix_row9_targeted()
