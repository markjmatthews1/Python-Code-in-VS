#!/usr/bin/env python3
"""
Add New Column to Portfolio Values 2025 Sheet
- Add column AK with 08/23/2025 data
- Follow existing row structure
- Update totals
"""

import pandas as pd
import numpy as np
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
from datetime import datetime

def add_portfolio_column():
    """Add new column AK with today's portfolio data"""
    
    excel_file = 'outputs/Dividends_2025.xlsx'
    
    try:
        # Load the workbook to work with the existing structure
        wb = load_workbook(excel_file)
        ws = wb['Portfolio Values 2025']
        
        print("📊 Current Portfolio Values 2025 sheet structure:")
        print(f"   Max column: {ws.max_column}")
        print(f"   Max row: {ws.max_row}")
        
        # Find the next available column (should be AK = column 37)
        next_col = ws.max_column + 1
        next_col_letter = chr(64 + (next_col - 1) // 26) + chr(65 + (next_col - 1) % 26) if next_col > 26 else chr(64 + next_col)
        
        print(f"   Adding data to column {next_col_letter} (column {next_col})")
        
        # Today's date for the header
        today_date = "08/23/2025"
        
        # Current portfolio values (from your Schwab data and API data)
        etrade_ira = 279339.15
        etrade_taxable = 62622.72
        schwab_ira = 50558.40      # Your provided data
        schwab_individual = 2603.64 # Your provided data  
        retirement_401k = 124315.15
        
        # Calculate total
        total_portfolio = etrade_ira + etrade_taxable + schwab_ira + schwab_individual + retirement_401k
        
        print(f"\n💰 Adding portfolio values for {today_date}:")
        print(f"   Row 3 (Date): {today_date}")
        print(f"   Row 4 (E*TRADE IRA): ${etrade_ira:,.2f}")
        print(f"   Row 5 (E*TRADE Taxable): ${etrade_taxable:,.2f}")
        print(f"   Row 6 (Schwab IRA): ${schwab_ira:,.2f}")
        print(f"   Row 7 (Schwab Individual): ${schwab_individual:,.2f}")
        print(f"   Row 8 (401k): ${retirement_401k:,.2f}")
        print(f"   Row 10 (Total): ${total_portfolio:,.2f}")
        
        # Add the data to the appropriate rows and column
        ws.cell(row=3, column=next_col, value=today_date)  # Date header
        ws.cell(row=4, column=next_col, value=etrade_ira)  # E*TRADE IRA
        ws.cell(row=5, column=next_col, value=etrade_taxable)  # E*TRADE Taxable
        ws.cell(row=6, column=next_col, value=schwab_ira)  # Schwab IRA
        ws.cell(row=7, column=next_col, value=schwab_individual)  # Schwab Individual
        ws.cell(row=8, column=next_col, value=retirement_401k)  # 401k
        ws.cell(row=10, column=next_col, value=total_portfolio)  # Total
        
        # Format the new column to match existing formatting
        # Copy formatting from previous column
        prev_col = next_col - 1
        for row in [3, 4, 5, 6, 7, 8, 10]:
            source_cell = ws.cell(row=row, column=prev_col)
            target_cell = ws.cell(row=row, column=next_col)
            
            # Copy formatting
            if source_cell.font:
                target_cell.font = Font(
                    name=source_cell.font.name,
                    size=source_cell.font.size,
                    bold=source_cell.font.bold,
                    color=source_cell.font.color
                )
            
            if source_cell.fill:
                target_cell.fill = PatternFill(
                    fill_type=source_cell.fill.fill_type,
                    start_color=source_cell.fill.start_color,
                    end_color=source_cell.fill.end_color
                )
            
            # Apply currency formatting for value rows
            if row in [4, 5, 6, 7, 8, 10]:
                target_cell.number_format = '$#,##0.00'
        
        # Save the workbook
        wb.save(excel_file)
        wb.close()
        
        print(f"✅ Added column {next_col_letter} to Portfolio Values 2025 sheet")
        print(f"📈 Total Portfolio Value: ${total_portfolio:,.2f}")
        
        return True
        
    except Exception as e:
        print(f"❌ Error adding portfolio column: {e}")
        import traceback
        traceback.print_exc()
        return False

def check_for_summary_gui():
    """Look for existing portfolio summary GUI or sheet"""
    
    print("\n🔍 Checking for portfolio summary GUI/sheet...")
    
    # Check for GUI files
    import os
    gui_files = [
        'portfolio_summary_report.py',
        'dividend_dashboard.py', 
        'dividend_focused_dashboard.py',
        'simple_dividend_dashboard.py',
        'flask_dashboard.py'
    ]
    
    found_guis = []
    for gui_file in gui_files:
        if os.path.exists(gui_file):
            found_guis.append(gui_file)
            print(f"   📊 Found GUI: {gui_file}")
    
    # Check for summary sheets in Excel
    try:
        xl_file = pd.ExcelFile('outputs/Dividends_2025.xlsx')
        summary_sheets = []
        for sheet in xl_file.sheet_names:
            if any(word in sheet.lower() for word in ['summary', 'dashboard', 'overview', 'totals']):
                summary_sheets.append(sheet)
                print(f"   📋 Found summary sheet: {sheet}")
        
        if not found_guis and not summary_sheets:
            print("   ❌ No portfolio summary GUI or sheet found")
            print("   💡 Should create a portfolio summary dashboard")
        
        return found_guis, summary_sheets
        
    except Exception as e:
        print(f"   ❌ Error checking sheets: {e}")
        return found_guis, []

if __name__ == "__main__":
    print("🎯 Adding Portfolio Values Column & Checking Summary")
    
    # Add the new portfolio column
    if add_portfolio_column():
        print("\n" + "="*50)
        
        # Check for existing summary systems
        guis, sheets = check_for_summary_gui()
        
        print("\n📋 SUMMARY:")
        print(f"✅ Portfolio Values 2025 updated with column AK (08/23/2025)")
        print(f"📊 Found {len(guis)} GUI files and {len(sheets)} summary sheets")
        
        if not guis and not sheets:
            print("💡 RECOMMENDATION: Create a portfolio summary dashboard")
            print("   This could show:")
            print("   • Overall portfolio value trends")
            print("   • Current dividend summary")
            print("   • Account breakdowns")
            print("   • Weekly/monthly changes")
    else:
        print("❌ Failed to add portfolio column")
