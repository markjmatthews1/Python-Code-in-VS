"""
Check Formula State

Check if Row 9 actually has the formula
"""
import openpyxl

def check_formula_state():
    """Check Row 9 formula state"""
    excel_file = r"c:\Users\mjmat\Python Code in VS\dividend_tracker\DividendTrackerApp\outputs\Dividends_2025.xlsx"
    
    try:
        # Check with formulas showing
        wb = openpyxl.load_workbook(excel_file, data_only=False)
        ws = wb["Estimated Income 2025"]
        
        max_col = ws.max_column
        current_col_letter = openpyxl.utils.get_column_letter(max_col)
        
        print(f"🔍 CHECKING ROW 9 FORMULA STATE - Column {current_col_letter}")
        print("=" * 50)
        
        row9_cell = ws.cell(row=9, column=max_col)
        formula = row9_cell.value
        
        print(f"Row 9 cell content: {formula}")
        print(f"Cell type: {type(formula)}")
        
        # Check if it's a formula
        if isinstance(formula, str) and formula.startswith('='):
            print("✅ Row 9 contains a formula")
        else:
            print("❌ Row 9 does not contain a formula")
            
        # Check adjacent cells for comparison
        for row in range(4, 10):
            account = ws.cell(row=row, column=1).value
            cell_value = ws.cell(row=row, column=max_col).value
            print(f"   Row {row} ({account}): {cell_value}")
        
        wb.close()
        
    except Exception as e:
        print(f"❌ ERROR: {e}")

if __name__ == "__main__":
    check_formula_state()
