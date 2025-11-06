import openpyxl
import json
import os
from datetime import datetime
from openpyxl.styles import Font, PatternFill

class TickerOrderPreservingUpdater:
    """Updater that preserves exact ticker order and only updates quantities/prices"""
    
    def __init__(self):
        self.excel_file = "outputs/Dividends_2025.xlsx"
        self.cache_file = "portfolio_data_cache.json"
    
    def run_update(self):
        """Main update that preserves ticker positions"""
        print("TICKER ORDER PRESERVING UPDATER")
        print("=" * 40)
        
        # Load cache
        cache_data = self.load_cache_data()
        if not cache_data:
            return False
            
        positions_data = cache_data.get('positions', {})
        yields_data = cache_data.get('yields', {})
        
        # Open Excel
        wb = openpyxl.load_workbook(self.excel_file)
        ws = wb["Accounts Div historical yield"]
        
        # Update each account group
        account_mapping = {
            "ETRADE IRA": "etrade_ira",
            "ETRADE Taxable": "etrade_taxable", 
            "Schwab IRA": "schwab_ira",
            "Schwab Individual": "schwab_individual"
        }
        
        account_info = self.find_account_groups(ws)
        
        for group_name, cache_key in account_mapping.items():
            if group_name not in account_info:
                continue
                
            print(f"\n{group_name}:")
            print("-" * 20)
            
            group_info = account_info[group_name]
            account_positions = positions_data.get(cache_key, [])
            
            # Create position lookup by symbol
            position_lookup = {pos.get('symbol', '').strip().upper(): pos for pos in account_positions}
            
            # Update existing tickers in place
            start_row = group_info["start_row"] + 2
            end_row = group_info["end_row"]
            
            for row in range(start_row, end_row + 1):
                # Get existing ticker from Column A
                existing_ticker_cell = ws.cell(row=row, column=1)
                existing_ticker = existing_ticker_cell.value
                
                if existing_ticker:
                    ticker_symbol = str(existing_ticker).strip().upper()
                    
                    # Find matching position data
                    if ticker_symbol in position_lookup:
                        position = position_lookup[ticker_symbol]
                        quantity = position.get('quantity', 0)
                        market_value = position.get('market_value', 0)
                        price = round(market_value / quantity, 2) if quantity > 0 else 0
                        
                        # Update ONLY Quantity (Column B) and Price (Column D) 
                        ws.cell(row=row, column=2).value = quantity
                        ws.cell(row=row, column=4).value = price
                        
                        print(f"Row {row}: {ticker_symbol} - Qty: {quantity}, Price: ${price}")
                    else:
                        print(f"Row {row}: {ticker_symbol} - NO POSITION DATA FOUND")
        
        # Save
        wb.save(self.excel_file)
        wb.close()
        
        print("\nSUCCESS: Ticker order preserved, only quantities and prices updated!")
        return True
    
    def find_account_groups(self, ws):
        """Find account group boundaries"""
        account_info = {}
        
        for row in range(1, 60):
            cell_value = ws.cell(row=row, column=1).value
            if cell_value:
                cell_str = str(cell_value).strip().upper()
                
                if "ETRADE IRA" in cell_str:
                    account_info["ETRADE IRA"] = {"start_row": row, "end_row": row + 20}
                elif "ETRADE TAXABLE" in cell_str:
                    account_info["ETRADE Taxable"] = {"start_row": row, "end_row": row + 15}
                elif "SCHWAB IRA" in cell_str:
                    account_info["Schwab IRA"] = {"start_row": row, "end_row": row + 10}
                elif "SCHWAB INDIVIDUAL" in cell_str:
                    account_info["Schwab Individual"] = {"start_row": row, "end_row": row + 10}
        
        return account_info
    
    def load_cache_data(self):
        """Load cache data"""
        try:
            with open(self.cache_file, 'r') as f:
                return json.load(f)
        except Exception as e:
            print(f"ERROR loading cache: {e}")
            return None

if __name__ == "__main__":
    updater = TickerOrderPreservingUpdater()
    updater.run_update()