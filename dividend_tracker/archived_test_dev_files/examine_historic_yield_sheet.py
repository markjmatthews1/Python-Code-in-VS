#!/usr/bin/env python3
"""
Examine the Etrade IRA historic yield sheet structure
"""
import pandas as pd
import openpyxl
import os

def examine_historic_yield_sheet():
    """Examine the structure of the Etrade IRA historic yield sheet"""
    workbook_path = r"c:\Users\mjmat\Python Code in VS\dividend_tracker\DividendTrackerApp\outputs\Dividends_2025.xlsx"
    
    try:
        # Load workbook to see available sheets
        wb = openpyxl.load_workbook(workbook_path, data_only=True)
        print("📊 Available sheets in Dividends_2025.xlsx:")
        for i, sheet_name in enumerate(wb.sheetnames, 1):
            print(f"   {i}. {sheet_name}")
            
        # Check if "Etrade IRA historic yield" exists
        target_sheet = "Etrade IRA historic yield"
        if target_sheet in wb.sheetnames:
            print(f"\n✅ Found target sheet: '{target_sheet}'")
            
            sheet = wb[target_sheet]
            print(f"   📏 Sheet dimensions: {sheet.max_row} rows × {sheet.max_column} columns")
            
            # Show headers (first row)
            print(f"\n📋 Headers (Row 1):")
            for col in range(1, min(sheet.max_column + 1, 20)):  # Show first 20 columns
                header_value = sheet.cell(row=1, column=col).value
                if header_value:
                    print(f"   Column {col}: '{header_value}'")
            
            # Show first few rows of data
            print(f"\n📊 Sample Data (First 5 rows):")
            for row in range(1, min(6, sheet.max_row + 1)):
                row_data = []
                for col in range(1, min(sheet.max_column + 1, 10)):  # Show first 10 columns
                    cell_value = sheet.cell(row=row, column=col).value
                    row_data.append(str(cell_value) if cell_value else "")
                print(f"   Row {row}: {row_data}")
                
        else:
            print(f"\n❌ Sheet '{target_sheet}' not found")
            print("   Available sheets:", wb.sheetnames)
            
        wb.close()
        
    except Exception as e:
        print(f"❌ Error examining workbook: {e}")

if __name__ == "__main__":
    examine_historic_yield_sheet()
