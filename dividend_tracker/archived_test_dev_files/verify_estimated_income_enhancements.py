"""
Verify Enhanced Estimated Income Formatting

Test the corrected:
1. Row 9 calculation: SUM(rows 4:7)/12 
2. Color coding: Green/Red/Yellow for all rows 4-9
3. Proper formatting maintained
"""
import openpyxl

def verify_estimated_income_enhancements():
    """Verify the enhanced Estimated Income formatting"""
    excel_file = r"c:\Users\mjmat\Python Code in VS\dividend_tracker\DividendTrackerApp\outputs\Dividends_2025.xlsx"
    
    try:
        # Load with formulas
        wb = openpyxl.load_workbook(excel_file, data_only=False)
        ws = wb["Estimated Income 2025"]
        
        print("🔍 VERIFYING ESTIMATED INCOME 2025 ENHANCEMENTS")
        print("=" * 60)
        
        max_col = ws.max_column
        current_col = max_col
        prev_col = max_col - 1
        
        current_col_letter = openpyxl.utils.get_column_letter(current_col)
        prev_col_letter = openpyxl.utils.get_column_letter(prev_col)
        
        print(f"📊 Checking columns: {prev_col_letter} (previous) vs {current_col_letter} (current)")
        
        # 1. Verify Row 9 Calculation
        print(f"\n🧮 ROW 9 CALCULATION VERIFICATION:")
        monthly_row = None
        for row in range(8, 12):
            cell_value = ws.cell(row=row, column=1).value
            if cell_value and 'monthly' in str(cell_value).lower():
                monthly_row = row
                break
        
        if monthly_row:
            formula_cell = ws.cell(row=monthly_row, column=current_col)
            formula = formula_cell.value
            expected_formula = f"=SUM({current_col_letter}4:{current_col_letter}7)/12"
            
            print(f"   Row {monthly_row} formula: {formula}")
            print(f"   Expected formula: {expected_formula}")
            
            if str(formula) == expected_formula:
                print("   ✅ Row 9 calculation CORRECT!")
            else:
                print("   ❌ Row 9 calculation INCORRECT!")
        
        # 2. Verify Color Coding
        print(f"\n🎨 COLOR CODING VERIFICATION (Rows 4-9):")
        
        wb.close()
        
        # Reload with calculated values for color verification
        wb = openpyxl.load_workbook(excel_file, data_only=True)
        ws = wb["Estimated Income 2025"]
        
        for row in range(4, 10):  # Rows 4-9
            account_name = ws.cell(row=row, column=1).value
            current_cell = ws.cell(row=row, column=current_col)
            prev_cell = ws.cell(row=row, column=prev_col)
            
            current_val = current_cell.value
            prev_val = prev_cell.value
            
            # Check if cell has color fill
            fill_color = current_cell.fill.start_color.index if current_cell.fill.start_color else None
            
            if current_val is not None and prev_val is not None:
                if isinstance(current_val, (int, float)) and isinstance(prev_val, (int, float)):
                    if current_val > prev_val:
                        expected_color = "Green (90EE90)"
                        color_status = "✅" if fill_color and '90EE90' in str(fill_color) else "❌"
                    elif current_val < prev_val:
                        expected_color = "Red (FFB6C1)"
                        color_status = "✅" if fill_color and 'FFB6C1' in str(fill_color) else "❌"
                    else:
                        expected_color = "Yellow (FFFF99)"
                        color_status = "✅" if fill_color and 'FFFF99' in str(fill_color) else "❌"
                    
                    print(f"   Row {row} ({account_name}): ${prev_val:,.2f} → ${current_val:,.2f}")
                    print(f"      Expected: {expected_color} | Actual: {fill_color} {color_status}")
            else:
                print(f"   Row {row} ({account_name}): Missing data for comparison")
        
        # 3. Verify Formatting
        print(f"\n📋 FORMATTING VERIFICATION:")
        sample_cell = ws.cell(row=5, column=current_col)
        font = sample_cell.font
        number_format = sample_cell.number_format
        
        print(f"   Font: {font.name} {font.size}pt {'Bold' if font.bold else ''}")
        print(f"   Number Format: {number_format}")
        print(f"   Column Width: {ws.column_dimensions[current_col_letter].width}")
        
        wb.close()
        
        print(f"\n✅ VERIFICATION COMPLETE")
        
        return True
        
    except Exception as e:
        print(f"❌ ERROR during verification: {e}")
        import traceback
        traceback.print_exc()
        return False

if __name__ == "__main__":
    verify_estimated_income_enhancements()
