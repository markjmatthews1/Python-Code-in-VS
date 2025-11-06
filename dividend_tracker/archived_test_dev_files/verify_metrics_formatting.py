import openpyxl

def verify_metrics_formatting():
    """Verify the dividend metrics are now properly formatted as numbers with Arial 12"""
    
    workbook_path = "outputs/Dividends_2025.xlsx"
    wb = openpyxl.load_workbook(workbook_path)
    portfolio_sheet = wb["Portfolio Summary"]
    
    print("METRICS FORMATTING VERIFICATION")
    print("=" * 50)
    
    # Find the dividend metrics section
    metrics_found = False
    for row in range(1, 50):
        cell_value = portfolio_sheet.cell(row=row, column=7).value
        if cell_value == "DIVIDEND METRICS":
            print(f"\nFound DIVIDEND METRICS at row {row}")
            metrics_found = True
            
            # Check the next few rows for metrics
            for i in range(1, 6):
                metric_row = row + i
                name_cell = portfolio_sheet.cell(row=metric_row, column=7)
                value_cell = portfolio_sheet.cell(row=metric_row, column=8)
                
                if name_cell.value:
                    print(f"\nRow {metric_row}:")
                    print(f"  Metric: '{name_cell.value}'")
                    print(f"  Value: {value_cell.value} (Type: {type(value_cell.value).__name__})")
                    print(f"  Font: {name_cell.font.name} {name_cell.font.size}pt")
                    
                    if value_cell.number_format:
                        print(f"  Number Format: {value_cell.number_format}")
                    
                    # Special check for Cut Rate percentage
                    if "Cut Rate" in str(name_cell.value):
                        if isinstance(value_cell.value, (int, float)):
                            print(f"  ✅ Cut Rate is numeric: {value_cell.value} ({value_cell.value:.1%})")
                        else:
                            print(f"  ❌ Cut Rate is still text: {value_cell.value}")
            break
    
    if not metrics_found:
        print("❌ DIVIDEND METRICS section not found")
    
    wb.close()

if __name__ == "__main__":
    verify_metrics_formatting()
