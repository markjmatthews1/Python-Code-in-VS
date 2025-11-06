import os
import openpyxl
from openpyxl.styles import Font, PatternFill
from openpyxl.styles.numbers import FORMAT_CURRENCY_USD_SIMPLE

def force_fix_monthly_average():
    """Force fix the Monthly Average formatting with direct cell manipulation"""
    
    excel_file = "outputs/Dividends_2025.xlsx"
    
    try:
        wb = openpyxl.load_workbook(excel_file, data_only=False)
        ws = wb["Estimated Income 2025"]
        
        # Find Monthly Average row
        monthly_row = 9  # Based on previous test, it's row 9
        last_col = ws.max_column  # Column AL based on previous test
        
        print(f"FORCE FIXING MONTHLY AVERAGE")
        print(f"Row: {monthly_row}, Column: {openpyxl.utils.get_column_letter(last_col)}")
        
        # Get the current cell
        cell = ws.cell(row=monthly_row, column=last_col)
        
        # Store current formula
        current_formula = cell.value
        print(f"Current formula: {current_formula}")
        
        # Apply bold formatting and preserve the formula
        cell.font = Font(name='Arial', size=12, bold=True)
        cell.number_format = FORMAT_CURRENCY_USD_SIMPLE
        
        # For color coding, we need to get the calculated values
        # First save to force calculation
        wb.save(excel_file)
        wb.close()
        
        # Reopen with data_only=True to get calculated values
        wb = openpyxl.load_workbook(excel_file, data_only=True)
        ws = wb["Estimated Income 2025"]
        
        current_value = ws.cell(row=monthly_row, column=last_col).value
        previous_value = ws.cell(row=monthly_row, column=last_col - 1).value if last_col > 1 else None
        
        wb.close()
        
        # Reopen without data_only to preserve formulas
        wb = openpyxl.load_workbook(excel_file)
        ws = wb["Estimated Income 2025"]
        cell = ws.cell(row=monthly_row, column=last_col)
        
        # Restore formula and apply formatting
        cell.value = current_formula
        cell.font = Font(name='Arial', size=12, bold=True)
        cell.number_format = FORMAT_CURRENCY_USD_SIMPLE
        
        # Apply color coding if we have values
        if isinstance(current_value, (int, float)) and isinstance(previous_value, (int, float)):
            print(f"Current: ${current_value:,.2f}, Previous: ${previous_value:,.2f}")
            
            if current_value > previous_value:
                cell.fill = PatternFill(start_color='90EE90', end_color='90EE90', fill_type='solid')
                color = "Green (Increase)"
            elif current_value < previous_value:
                cell.fill = PatternFill(start_color='FF7C80', end_color='FF7C80', fill_type='solid')
                color = "Red (Decrease)"
            else:
                cell.fill = PatternFill(start_color='FFFF00', end_color='FFFF00', fill_type='solid')
                color = "Yellow (Same)"
                
            print(f"Applied color: {color}")
        else:
            print("Could not determine color coding - values not numeric")
        
        # Save final result
        wb.save(excel_file)
        wb.close()
        
        print("✅ Monthly Average formatting force-fixed!")
        
    except Exception as e:
        print(f"ERROR: {e}")
        if 'Permission' in str(e):
            print("Please close Excel and try again")
        import traceback
        traceback.print_exc()

if __name__ == "__main__":
    force_fix_monthly_average()