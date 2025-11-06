"""
Enhanced proper_excel_updater.py with color coding and 401K handling

FIXES APPLIED:
1. Correct color codes: Green #90EE90, Red #FF7C80, Yellow #FFFF00
2. Color coding for Portfolio Values rows 4-9 (including 401K)
3. Color coding for Estimated Income rows 4-9 (including Monthly Average)
4. Proper 401K integration in Portfolio Values
5. Row 9 formula correction: =SUM(rows 4:7)/12

This replaces proper_excel_updater.py with working color coding.
"""

import os
import datetime
from datetime import datetime
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils.cell import get_column_letter
from openpyxl.styles.numbers import FORMAT_CURRENCY_USD_SIMPLE
import traceback

class ProperExcelUpdater:
    def __init__(self, excel_file):
        self.excel_file = excel_file
        self.today_str = datetime.now().strftime("%m/%d/%Y")
        
    def run_complete_update(self):
        """Main update function - this is called from Etrade_menu.py"""
        print("\n🔄 COMPLETE SYSTEM UPDATE - APPEND ONLY (Enhanced with Color Coding)")
        print("="*60)
        
        # Create backup first
        print("\n💾 STEP 0: Creating backup...")
        if self.create_backup():
            print("   ✅ Backup created successfully")
        
        # Step 1: Get 401K value (working popup)
        print("\n💰 STEP 1: Getting 401K value...")
        k401_value = self.get_401k_value()
        if not k401_value:
            return False
            
        # Step 2: Get fresh API data
        print("\n📊 STEP 2: Collecting API data...")
        fresh_data = self.get_fresh_api_data(k401_value)
        if not fresh_data:
            return False
            
        # Step 3: Update sheets with color coding
        success_count = 0
        
        # Update Portfolio Values 2025 (time-series) with color coding
        if self.update_portfolio_values_timeseries(fresh_data, k401_value):
            print("   ✅ Portfolio Values 2025: New column added with color coding")
            success_count += 1
        else:
            print("   ❌ Portfolio Values 2025: Failed")
            
        # Update Estimated Income 2025 (time-series) with color coding
        if self.update_estimated_income_timeseries(fresh_data):
            print("   ✅ Estimated Income 2025: New column added with color coding")
            success_count += 1
        else:
            print("   ❌ Estimated Income 2025: Failed")
            
        # Update other sheets safely
        if self.update_other_sheets_safely(fresh_data):
            print("   ✅ Other sheets: Updated safely")
            success_count += 1
        else:
            print("   ❌ Other sheets: Issues occurred")
            
        # Final status
        print(f"\n📈 RESULTS: {success_count}/3 sheets updated successfully")
        if success_count >= 2:
            print("✅ Complete System Update: SUCCESS")
            return True
        else:
            print("❌ Complete System Update: PARTIAL FAILURE")
            return False
    
    def get_401k_value(self):
        """Get 401K value using the working popup"""
        try:
            from gui_prompts import get_k401_value
            print("   📱 Opening 401K dialog (working version)...")
            k401_value = get_k401_value()
            
            if k401_value and k401_value > 0:
                print(f"   ✅ 401K Value: ${k401_value:,.2f}")
                return k401_value
            else:
                print("   ❌ No valid 401K value provided")
                return None
        except Exception as e:
            print(f"   ❌ 401K popup error: {e}")
            return None
    
    def get_fresh_api_data(self, k401_value):
        """Get fresh data from APIs with fallback"""
        try:
            # Import the working collector
            import sys
            if 'c:\\Users\\mjmat\\Python Code in VS\\dividend_tracker\\DividendTrackerApp' not in sys.path:
                sys.path.append('c:\\Users\\mjmat\\Python Code in VS\\dividend_tracker\\DividendTrackerApp')
            
            from portfolio_data_collector import PortfolioDataCollector
            collector = PortfolioDataCollector()
            
            print("   🔄 Collecting from working APIs...")
            
            # Use the working collect method with fallback
            fresh_data = collector.collect_all_data_with_fallback(k401_value)
            
            if fresh_data and fresh_data.get('totals', {}).get('total_portfolio', 0) > 0:
                totals = fresh_data['totals']
                print(f"   ✅ Portfolio: ${totals['total_portfolio']:,.2f}")
                print(f"   ✅ Annual Dividends: ${totals['total_yearly_dividends']:,.2f}")
                return fresh_data
            else:
                print("   ❌ No valid API data collected")
                return None
                
        except Exception as e:
            print(f"   ❌ API collection error: {e}")
            return None
            
    def create_backup(self):
        """Create backup before making changes"""
        try:
            if os.path.exists(self.excel_file):
                timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
                backup_file = self.excel_file.replace('.xlsx', f'_backup_enhanced_{timestamp}.xlsx')
                
                import shutil
                shutil.copy2(self.excel_file, backup_file)
                print(f"   💾 Backup: {os.path.basename(backup_file)}")
                return True
        except Exception as e:
            print(f"   ⚠️ Backup error: {e}")
        return False
    
    def get_previous_column_values(self, ws, row_range):
        """Get values from the previous column for color comparison"""
        previous_values = {}
        
        # Find the last column with data in row 3 (dates row)
        last_col = 1
        for col in range(1, ws.max_column + 1):
            if ws.cell(row=3, column=col).value:
                last_col = col
        
        # Get values from the previous column
        if last_col > 1:
            prev_col = last_col
            for row in row_range:
                account_name = ws.cell(row=row, column=1).value
                if account_name:
                    account_key = str(account_name).strip()
                    prev_value = ws.cell(row=row, column=prev_col).value
                    if prev_value and isinstance(prev_value, (int, float)):
                        previous_values[account_key] = prev_value
        
        return previous_values
    
    def apply_color_coding(self, cell, new_value, old_value):
        """Apply color coding based on value comparison"""
        if old_value is None or old_value == 0:
            # No previous value, use default formatting
            cell.fill = PatternFill(fill_type=None)
        elif new_value > old_value:
            # Increase - Green #90EE90
            cell.fill = PatternFill(start_color='90EE90', end_color='90EE90', fill_type='solid')
        elif new_value < old_value:
            # Decrease - Red #FF7C80
            cell.fill = PatternFill(start_color='FF7C80', end_color='FF7C80', fill_type='solid')
        else:
            # Same value - Yellow #FFFF00
            cell.fill = PatternFill(start_color='FFFF00', end_color='FFFF00', fill_type='solid')
        
    def update_portfolio_values_timeseries(self, fresh_data, k401_value):
        """Update Portfolio Values 2025 by adding new date column with color coding"""
        try:
            wb = openpyxl.load_workbook(self.excel_file)
            
            if "Portfolio Values 2025" not in wb.sheetnames:
                print("      ❌ Portfolio Values 2025 sheet not found")
                return False
                
            ws = wb["Portfolio Values 2025"]
            
            # Find the structure: dates should be in row 3, accounts in column A starting row 4
            print("      📊 Analyzing existing Portfolio Values structure...")
            
            # Get previous values for color comparison (rows 4-9)
            previous_values = self.get_previous_column_values(ws, range(4, 10))
            
            # Find last column with data in row 3 (dates row)
            last_col = 1
            for col in range(1, ws.max_column + 1):
                if ws.cell(row=3, column=col).value:
                    last_col = col
            
            # Add new date column
            new_col = last_col + 1
            
            # Add today's date to the header row (row 3) with proper formatting
            date_cell = ws.cell(row=3, column=new_col, value=self.today_str)
            date_cell.font = Font(name='Arial', size=12, bold=True, color='FFFFFF')
            date_cell.fill = PatternFill(start_color='4F81BD', end_color='4F81BD', fill_type='solid')
            date_cell.alignment = Alignment(horizontal='right')
            date_cell.number_format = 'm/d/yyyy'
            
            # Add account values to the new column
            portfolio_values = fresh_data.get('portfolio_values', {})
            
            # Map the account names to what's in the Excel (including 401K)
            account_mapping = {
                'E*TRADE IRA': portfolio_values.get('E*TRADE IRA', 0),
                'E*TRADE Taxable': portfolio_values.get('E*TRADE Taxable', 0),
                'Schwab IRA': portfolio_values.get('Schwab IRA', 0),
                'Schwab Individual': portfolio_values.get('Schwab Individual', 0),
                '401K': k401_value  # Include 401K value
            }
            
            # Find and update account rows with proper formatting and color coding
            for row in range(4, 10):  # Rows 4-9 (accounts + total)
                account_name = ws.cell(row=row, column=1).value
                if account_name:
                    account_key = str(account_name).strip()
                    
                    # Handle exact matches and variations
                    value = None
                    if account_key in account_mapping:
                        value = account_mapping[account_key]
                    elif 'total' in account_key.lower():
                        # Calculate total for Total row
                        value = sum(portfolio_values.values()) + k401_value
                    elif any(partial in account_key for partial in ['E*TRADE', 'Etrade', 'Schwab', '401K']):
                        # Handle variations in account names
                        for map_key, map_value in account_mapping.items():
                            if map_key.replace('*', '') in account_key or map_key in account_key:
                                value = map_value
                                break
                    
                    if value is not None:
                        cell = ws.cell(row=row, column=new_col, value=value)
                        cell.number_format = FORMAT_CURRENCY_USD_SIMPLE
                        cell.font = Font(name='Arial', size=12, bold=('total' in account_key.lower()))
                        
                        # Apply color coding
                        old_value = previous_values.get(account_key, None)
                        self.apply_color_coding(cell, value, old_value)
                        
                        print(f"      💰 {account_key}: ${value:,.2f} (Color coded)")
            
            # Set column width for proper display
            ws.column_dimensions[openpyxl.utils.get_column_letter(new_col)].width = 15
            
            wb.save(self.excel_file)
            wb.close()
            return True
            
        except Exception as e:
            print(f"      ❌ Portfolio Values update error: {e}")
            traceback.print_exc()
            return False
            
    def update_estimated_income_timeseries(self, fresh_data):
        """Update Estimated Income 2025 by adding new date column with color coding"""
        try:
            wb = openpyxl.load_workbook(self.excel_file)
            
            if "Estimated Income 2025" not in wb.sheetnames:
                print("      ❌ Estimated Income 2025 sheet not found")
                return False
                
            ws = wb["Estimated Income 2025"]
            
            # Find the structure: dates should be in row 3, accounts in column A starting row 4
            print("      📊 Analyzing existing Estimated Income structure...")
            
            # Get previous values for color comparison (rows 4-9)
            previous_values = self.get_previous_column_values(ws, range(4, 10))
            
            # Find last column with data in row 3 (dates row) 
            last_col = 1
            for col in range(1, ws.max_column + 1):
                if ws.cell(row=3, column=col).value:
                    last_col = col
            
            # Add new date column
            new_col = last_col + 1
            
            # Add today's date to the header row (row 3) with proper formatting
            date_cell = ws.cell(row=3, column=new_col, value=self.today_str)
            date_cell.font = Font(name='Arial', size=12, bold=True, color='FFFFFF')
            date_cell.fill = PatternFill(start_color='4F81BD', end_color='4F81BD', fill_type='solid')
            date_cell.alignment = Alignment(horizontal='right')
            date_cell.number_format = 'm/d/yyyy'
            
            # Add dividend estimates to the new column
            dividend_estimates = fresh_data.get('dividend_estimates', {})
            
            # Map the account names
            account_mapping = {
                'E*TRADE IRA': dividend_estimates.get('E*TRADE IRA', 0),
                'E*TRADE Taxable': dividend_estimates.get('E*TRADE Taxable', 0), 
                'Schwab IRA': dividend_estimates.get('Schwab IRA', 0),
                'Schwab Individual': dividend_estimates.get('Schwab Individual', 0)
            }
            
            # Find and update account rows with proper formatting and color coding (rows 4-7)
            for row in range(4, 8):  # Dividend account rows
                account_name = ws.cell(row=row, column=1).value
                if account_name:
                    account_str = str(account_name).strip()
                    
                    # Try exact match first
                    value = None
                    if account_str in account_mapping:
                        value = account_mapping[account_str]
                    else:
                        # Try partial matches for account names
                        for map_key, map_value in account_mapping.items():
                            if map_key.replace('*', '') in account_str or map_key in account_str:
                                value = map_value
                                break
                    
                    if value is not None:
                        cell = ws.cell(row=row, column=new_col, value=value)
                        cell.number_format = FORMAT_CURRENCY_USD_SIMPLE
                        cell.font = Font(name='Arial', size=12)
                        
                        # Apply color coding
                        old_value = previous_values.get(account_str, None)
                        self.apply_color_coding(cell, value, old_value)
                        
                        print(f"      📈 {account_str}: ${value:,.2f} (Color coded)")
            
            # CRITICAL: Add Row 9 Monthly Average Calculation with color coding
            # Row 9 should calculate: =SUM(rows 4:7)/12 (ALL dividend accounts monthly average)
            monthly_row = None
            for row in range(8, 12):  # Look for Monthly Average row
                cell_value = ws.cell(row=row, column=1).value
                if cell_value and 'monthly' in str(cell_value).lower():
                    monthly_row = row
                    break
            
            if monthly_row:
                # Add the formula: =SUM(E4:E7)/12 format (rows 4-7, not 5-7)
                col_letter = openpyxl.utils.get_column_letter(new_col)
                formula = f"=SUM({col_letter}4:{col_letter}7)/12"
                
                cell = ws.cell(row=monthly_row, column=new_col)
                cell.value = formula
                cell.number_format = FORMAT_CURRENCY_USD_SIMPLE
                cell.font = Font(name='Arial', size=12)
                
                # For formulas, get the calculated value for color coding
                wb.save(self.excel_file)  # Save to calculate formula
                wb.close()
                wb = openpyxl.load_workbook(self.excel_file)  # Reload to get calculated value
                ws = wb["Estimated Income 2025"]
                
                calculated_value = ws.cell(row=monthly_row, column=new_col).value
                if isinstance(calculated_value, (int, float)):
                    account_name = ws.cell(row=monthly_row, column=1).value
                    account_key = str(account_name).strip() if account_name else "Monthly"
                    old_value = previous_values.get(account_key, None)
                    
                    # Reapply formatting with color coding
                    cell = ws.cell(row=monthly_row, column=new_col)
                    cell.value = formula
                    cell.number_format = FORMAT_CURRENCY_USD_SIMPLE
                    cell.font = Font(name='Arial', size=12)
                    self.apply_color_coding(cell, calculated_value, old_value)
                    
                    print(f"      🧮 Monthly Average: {formula} = ${calculated_value:,.2f} (Color coded)")
                else:
                    print(f"      🧮 Monthly Average: {formula}")
            
            # Add total dividends if found with proper formatting and color coding
            total_dividends = fresh_data.get('totals', {}).get('total_yearly_dividends', 0)
            for row in range(4, ws.max_row + 1):
                account_name = ws.cell(row=row, column=1).value
                if account_name and 'total' in str(account_name).lower():
                    cell = ws.cell(row=row, column=new_col, value=total_dividends)
                    cell.number_format = FORMAT_CURRENCY_USD_SIMPLE
                    cell.font = Font(name='Arial', size=12, bold=True)
                    
                    # Apply color coding
                    account_key = str(account_name).strip()
                    old_value = previous_values.get(account_key, None)
                    self.apply_color_coding(cell, total_dividends, old_value)
                    
                    print(f"      📈 Total Dividends: ${total_dividends:,.2f} (Color coded)")
                    break
            
            # Set column width for proper display
            ws.column_dimensions[openpyxl.utils.get_column_letter(new_col)].width = 15
            
            wb.save(self.excel_file)
            wb.close()
            return True
            
        except Exception as e:
            print(f"      ❌ Estimated Income update error: {e}")
            traceback.print_exc()
            return False
            
    def update_other_sheets_safely(self, fresh_data):
        """Update other sheets without destroying existing data"""
        try:
            # Just update the timestamp on Portfolio Summary to show it was touched
            wb = openpyxl.load_workbook(self.excel_file)
            
            if "Portfolio Summary" in wb.sheetnames:
                ws = wb["Portfolio Summary"]
                
                # Find a safe place to put an update timestamp
                # Look for an empty area or existing timestamp
                for row in range(1, min(10, ws.max_row + 1)):
                    for col in range(1, min(5, ws.max_column + 1)):
                        cell_value = ws.cell(row=row, column=col).value
                        if cell_value and "updated" in str(cell_value).lower():
                            # Update existing timestamp
                            ws.cell(row=row, column=col, value=f"Last Updated: {self.today_str}")
                            break
                else:
                    # Add timestamp in a safe location (bottom of sheet)
                    safe_row = ws.max_row + 2
                    ws.cell(row=safe_row, column=1, value=f"API Data Updated: {self.today_str}")
            
            wb.save(self.excel_file)
            wb.close()
            return True
            
        except Exception as e:
            print(f"      ❌ Other sheets update error: {e}")
            return False

if __name__ == "__main__":
    print("Enhanced Proper Excel Updater - Testing")
    excel_file = "c:\\Users\\mjmat\\Python Code in VS\\dividend_tracker\\DividendTracker\\Dividends_2025.xlsx"
    
    if os.path.exists(excel_file):
        updater = ProperExcelUpdater(excel_file)
        updater.run_complete_update()
    else:
        print(f"Excel file not found: {excel_file}")
