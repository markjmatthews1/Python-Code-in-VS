#!/usr/bin/env python3
"""
Analyze Current Sheet Formatting Issues
======================================
"""

import openpyxl
from openpyxl.styles import PatternFill
import os

def analyze_formatting_issues():
    """Analyze the current sheet formatting issues"""
    
    excel_path = r"c:\Users\mjmat\Python Code in VS\dividend_tracker\DividendTrackerApp\outputs\Dividends_2025.xlsx"
    
    try:
        workbook = openpyxl.load_workbook(excel_path)
        sheet = workbook["Accounts Div historical yield"]
        
        print(f"📊 FORMATTING ANALYSIS")
        print("=" * 60)
        print(f"📏 Sheet dimensions: {sheet.max_row} rows × {sheet.max_column} columns")
        
        # Define account sections
        sections = [
            {"name": "Etrade IRA", "title_row": 1, "header_row": 2, "data_start": 3, "data_end": 20, "avg_row": 25},
            {"name": "Etrade Taxable", "title_row": 27, "header_row": 28, "data_start": 29, "data_end": 40, "avg_row": 41},
            {"name": "Schwab IRA", "title_row": 43, "header_row": 44, "data_start": 44, "data_end": 47, "avg_row": 49},
            {"name": "Schwab Individual", "title_row": 50, "header_row": 51, "data_start": 52, "data_end": 53, "avg_row": 55}
        ]
        
        # Analyze each section
        for section in sections:
            print(f"\n🏦 {section['name']} Analysis:")
            print("-" * 40)
            
            # Check title row
            title_cell = sheet.cell(row=section["title_row"], column=1)
            title_value = title_cell.value
            title_fill = title_cell.fill
            print(f"📋 Title Row {section['title_row']}: '{title_value}'")
            print(f"   Color: {title_fill.start_color.rgb if title_fill.start_color else 'None'}")
            
            # Check inserted column in title row (column 16)
            title_insert_cell = sheet.cell(row=section["title_row"], column=16)
            title_insert_fill = title_insert_cell.fill
            title_insert_value = title_insert_cell.value
            print(f"   Column 16 value: '{title_insert_value}'")
            print(f"   Column 16 color: {title_insert_fill.start_color.rgb if title_insert_fill.start_color else 'None'}")
            
            # Check header row
            if section["header_row"] <= sheet.max_row:
                header_cell = sheet.cell(row=section["header_row"], column=1)
                header_value = header_cell.value
                header_insert_cell = sheet.cell(row=section["header_row"], column=16)
                header_insert_value = header_insert_cell.value
                print(f"📝 Header Row {section['header_row']}: '{header_value}' | Column 16: '{header_insert_value}'")
            
            # Check average row
            if section["avg_row"] <= sheet.max_row:
                avg_cell = sheet.cell(row=section["avg_row"], column=16)
                avg_value = avg_cell.value
                avg_fill = avg_cell.fill
                print(f"🧮 Average Row {section['avg_row']}: '{avg_value}'")
                print(f"   Average color: {avg_fill.start_color.rgb if avg_fill.start_color else 'None'}")
                
                # Check row above average
                above_avg_cell = sheet.cell(row=section["avg_row"]-1, column=16)
                above_avg_value = above_avg_cell.value
                if above_avg_value:
                    print(f"⚠️  Row {section['avg_row']-1} (above avg): '{above_avg_value}' - Should be empty!")
        
        workbook.close()
        
    except Exception as e:
        print(f"❌ Error: {e}")

if __name__ == "__main__":
    analyze_formatting_issues()
