import openpyxl
import json
import os
from datetime import datetime
from openpyxl.styles import Font, PatternFill

class SurgicalHistoricalYieldUpdater:
    """SURGICAL updater that preserves ticker order and only updates A, B, D columns"""
    
    def __init__(self):
        self.excel_file = r"C:\Users\mjmat\Python Code in VS\dividend_tracker\DividendTrackerApp\outputs\Dividends_2025.xlsx"
        self.cache_file = r"C:\Users\mjmat\Python Code in VS\dividend_tracker\DividendTrackerApp\portfolio_data_cache.json"
        self.account_mapping = {
            "E*TRADE IRA": "etrade_ira",
            "E*TRADE Taxable": "etrade_taxable", 
            "Schwab IRA": "schwab_ira",
            "Schwab Individual": "schwab_individual"
        }
    
    def run_update(self):
        """Main update method that preserves existing ticker order"""
        print("SURGICAL HISTORICAL YIELD UPDATER")
        print("=" * 50)
        
        # Load data
        cache_data = self.load_cache_data()
        if not cache_data:
            return False
            
        positions_data = cache_data.get('positions', {})
        yields_data = cache_data.get('yields', {})
        
        # Open Excel file
        if not os.path.exists(self.excel_file):
            print(f"ERROR: Excel file not found: {self.excel_file}")
            return False
            
        wb = openpyxl.load_workbook(self.excel_file)
        ws = wb["Accounts Div historical yield"]
        
        # Find account groups
        account_info = self.find_account_groups(ws)
        
        # Surgical update for each group
        for group_name in ["E*TRADE IRA", "E*TRADE Taxable", "Schwab IRA", "Schwab Individual"]:
            if group_name not in account_info:
                continue
                
            print(f"\nUPDATING {group_name}:")
            print("-" * 30)
            
            cache_key = self.account_mapping.get(group_name)
            if not cache_key:
                continue
                
            account_positions = positions_data.get(cache_key, [])
            filtered_positions = self.filter_high_yield_positions(account_positions, yields_data)
            
            # Create lookup by symbol
            position_lookup = {pos.get('symbol', '').strip().upper(): pos for pos in filtered_positions}
            
            # Surgically update existing rows
            self.surgical_update_group(ws, account_info[group_name], position_lookup)
        
        # Update yield data with new column
        date_str = datetime.now().strftime("%m/%d/%Y")
        self.insert_yield_column(ws, date_str)
        self.update_yield_data(ws, account_info, yields_data)
        
        # Apply formatting
        self.apply_group_formatting(ws, account_info)
        
        # Save file
        wb.save(self.excel_file)
        wb.close()
        
        print("\nSUCCESS: Surgical update completed!")
        return True
    
    def surgical_update_group(self, ws, group_info, position_lookup):
        """Surgically update only columns A, B, D while preserving ticker order"""
        start_row = group_info["start_row"] + 2  # Skip headers
        end_row = group_info["end_row"]
        
        ticker_font = Font(name="Arial", size=12, bold=True, color="3072C2")
        qty_font = Font(name="Arial", size=12, bold=True, color="3072C2")
        
        updated_count = 0
        
        for row in range(start_row, end_row + 1):
            # Check if this row has existing ticker data (look at other columns)
            has_data = any(ws.cell(row=row, column=col).value for col in [3, 5, 6, 7, 8, 9, 10, 11, 12, 13, 14, 15])
            
            if has_data:
                # Find best matching position for this row
                # Use row position to match with position index
                available_positions = [symbol for symbol in position_lookup.keys() if symbol not in getattr(self, '_used_symbols', set())]
                
                if available_positions and updated_count < len(available_positions):
                    symbol = available_positions[updated_count]
                    position = position_lookup[symbol]
                    
                    quantity = position.get('quantity', 0)
                    market_value = position.get('market_value', 0)
                    price = round(market_value / quantity, 2) if quantity > 0 else 0
                    
                    # Update ONLY columns A, B, D
                    ws.cell(row=row, column=1).value = symbol
                    ws.cell(row=row, column=1).font = ticker_font
                    ws.cell(row=row, column=2).value = quantity
                    ws.cell(row=row, column=2).font = qty_font  
                    ws.cell(row=row, column=4).value = price
                    
                    # Track used symbols
                    if not hasattr(self, '_used_symbols'):
                        self._used_symbols = set()
                    self._used_symbols.add(symbol)
                    
                    updated_count += 1
                    print(f"Row {row}: {symbol} qty={quantity} price=${price}")
        
        print(f"Updated {updated_count} existing positions")
    
    def find_account_groups(self, ws):
        """Find account group positions"""
        account_info = {}
        
        for row in range(1, min(60, ws.max_row + 1)):
            cell_value = ws.cell(row=row, column=1).value
            if cell_value:
                cell_str = str(cell_value).strip().upper()
                for group_name in self.account_mapping.keys():
                    if group_name.upper() in cell_str:
                        account_info[group_name] = {"start_row": row}
                        break
        
        # Calculate end rows
        group_order = ["E*TRADE IRA", "E*TRADE Taxable", "Schwab IRA", "Schwab Individual"]
        for i, group_name in enumerate(group_order):
            if group_name not in account_info:
                continue
                
            if i < len(group_order) - 1:
                # End is 3 rows before next group
                end_row = ws.max_row
                for next_group in group_order[i+1:]:
                    if next_group in account_info:
                        end_row = account_info[next_group]["start_row"] - 3
                        break
            else:
                # Last group
                end_row = account_info[group_name]["start_row"] + 10  # Reasonable default
                
            account_info[group_name]["end_row"] = end_row
        
        return account_info
    
    def filter_high_yield_positions(self, positions, yields_data):
        """Filter positions with yield > 4%"""
        filtered = []
        for position in positions:
            symbol = position.get('symbol', '').strip().upper()
            yield_info = yields_data.get(symbol, {})
            has_dividend = yield_info.get('has_dividend', False)
            current_yield = yield_info.get('yield', 0.0)
            
            if has_dividend and current_yield > 4.0:
                filtered.append(position)
                
        return filtered
    
    def insert_yield_column(self, ws, date_str):
        """Insert new yield column P with date headers"""
        # Check if column P already exists
        if ws.cell(row=2, column=16).value:
            print("INFO: Column P already exists, skipping insertion")
            return
            
        ws.insert_cols(16)
        
        # Add headers
        for row in range(1, 60):
            cell_value = ws.cell(row=row, column=1).value
            if cell_value and any(keyword in str(cell_value).upper() for keyword in ['ETRADE', 'SCHWAB']):
                header_row = row + 1
                ws.cell(row=header_row, column=16).value = date_str
                ws.cell(row=header_row, column=16).font = Font(bold=True)
    
    def update_yield_data(self, ws, account_info, yields_data):
        """Update yield data with correct background colors"""
        for group_name, group_info in account_info.items():
            start_row = group_info["start_row"] + 2
            end_row = group_info["end_row"]
            
            for row in range(start_row, end_row + 1):
                ticker_cell = ws.cell(row=row, column=1)
                if ticker_cell.value:
                    symbol = str(ticker_cell.value).strip().upper()
                    yield_info = yields_data.get(symbol, {})
                    current_yield = yield_info.get('yield', 0.0)
                    
                    yield_cell = ws.cell(row=row, column=16)  # Column P
                    yield_cell.value = current_yield / 100.0  # Convert to decimal
                    yield_cell.number_format = '0.00%'
                    
                    # Compare with Column O for background color
                    previous_cell = ws.cell(row=row, column=15)  # Column O
                    previous_yield = 0.0
                    
                    if previous_cell.value:
                        try:
                            if isinstance(previous_cell.value, str):
                                prev_str = str(previous_cell.value).replace('%', '').strip()
                                previous_yield = float(prev_str)
                            elif isinstance(previous_cell.value, (int, float)):
                                prev_val = float(previous_cell.value)
                                previous_yield = prev_val * 100 if prev_val < 1.0 else prev_val
                        except (ValueError, TypeError):
                            previous_yield = 0.0
                    
                    # Apply background colors based on comparison
                    if previous_yield == 0.0:
                        yield_cell.fill = PatternFill()  # No background
                    elif current_yield > previous_yield:
                        yield_cell.fill = PatternFill(start_color="90EE90", end_color="90EE90", fill_type="solid")  # Green
                    elif current_yield < previous_yield:
                        yield_cell.fill = PatternFill(start_color="FF7C80", end_color="FF7C80", fill_type="solid")  # Red
                    else:
                        yield_cell.fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")  # Yellow
    
    def apply_group_formatting(self, ws, account_info):
        """Apply orange formatting to group dividers"""
        orange_fill = PatternFill(start_color="FFC000", end_color="FFC000", fill_type="solid")
        
        for group_name, group_info in account_info.items():
            divider_row = group_info["start_row"]
            ws.cell(row=divider_row, column=1).fill = orange_fill
            ws.cell(row=divider_row, column=16).fill = orange_fill  # Column P
    
    def load_cache_data(self):
        """Load cache data"""
        try:
            if not os.path.exists(self.cache_file):
                print(f"ERROR: Cache file not found: {self.cache_file}")
                return None
            
            with open(self.cache_file, 'r') as f:
                return json.load(f)
        except Exception as e:
            print(f"ERROR: Failed to load cache: {e}")
            return None

if __name__ == "__main__":
    updater = SurgicalHistoricalYieldUpdater()
    success = updater.run_update()
    
    if success:
        print("\nSUCCESS: Surgical historical yield update completed!")
    else:
        print("\nERROR: Update failed!")