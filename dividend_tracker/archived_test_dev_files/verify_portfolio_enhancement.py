import openpyxl
import pandas as pd

def verify_portfolio_enhancement():
    """Verify the Portfolio Summary dividend enhancement results"""
    
    workbook_path = "outputs/Dividends_2025.xlsx"
    
    try:
        wb = openpyxl.load_workbook(workbook_path)
        
        # Check Portfolio Summary sheet
        portfolio_sheet = wb["Portfolio Summary"]
        
        print("PORTFOLIO SUMMARY DIVIDEND STATUS VERIFICATION")
        print("=" * 55)
        print("\nNew columns G-H content:")
        
        # Read the dividend status columns
        for row in range(1, 20):  # Check first 20 rows
            col_g = portfolio_sheet.cell(row=row, column=7).value
            col_h = portfolio_sheet.cell(row=row, column=8).value
            
            if col_g or col_h:
                print(f"Row {row:2d}: G='{col_g}' | H='{col_h}'")
        
        # Also check the dividend yield data structure
        print("\n" + "=" * 55)
        print("DIVIDEND YIELD DATA STRUCTURE SAMPLE:")
        
        div_sheet = wb["Accounts Div historical yield"]
        print(f"\nSheet dimensions: {div_sheet.max_row} rows x {div_sheet.max_column} columns")
        
        # Show a few sample rows from each account
        sample_rows = [3, 25, 41, 49]  # First row of each account
        for row in sample_rows:
            ticker = div_sheet.cell(row=row, column=1).value
            beginning_yield = div_sheet.cell(row=row, column=15).value  # Column O
            current_yield = div_sheet.cell(row=row, column=16).value    # Column P
            
            print(f"Row {row:2d}: {ticker} | Beginning: {beginning_yield} | Current: {current_yield}")
        
        wb.close()
        
    except Exception as e:
        print(f"Error during verification: {e}")

if __name__ == "__main__":
    verify_portfolio_enhancement()
