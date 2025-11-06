#!/usr/bin/env python3
"""
CORRECTED Enhanced Cache-Based Historical Yield Updater
=======================================================

Fixes:
1. Properly inserts rows at bottom of groups instead of overwriting headers
2. Applies correct formatting: Arial 12, bold, light blue (#3072C2) for tickers and quantities
3. Fixes yield calculation to show actual percentages instead of 0%
4. Preserves group structure and headers
"""

import os
import sys
import json
import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment
from openpyxl.styles.numbers import FORMAT_PERCENTAGE, FORMAT_CURRENCY_USD_SIMPLE
from datetime import datetime
import traceback

class CorrectedEnhancedHistoricalYieldUpdater:
    """Corrected updater that properly handles row insertion and formatting"""
    
    def __init__(self):
        self.script_dir = os.path.dirname(os.path.abspath(__file__))
        self.outputs_dir = os.path.join(self.script_dir, "outputs")
        self.excel_file = os.path.join(self.outputs_dir, "Dividends_2025.xlsx")
        self.cache_file = os.path.join(self.script_dir, "portfolio_data_cache.json")
        self.today = datetime.now()
        self.today_str = self.today.strftime("%m/%d/%Y")
        
        # Account group mapping
        self.account_groups = {
            "E*TRADE IRA": "etrade_ira",
            "E*TRADE Taxable": "etrade_taxable", 
            "Schwab IRA": "schwab_ira",
            "Schwab Individual": "schwab_individual"
        }
        
        # Define colors and fonts
        self.colors = {
            'green': PatternFill(start_color='90EE90', end_color='90EE90', fill_type='solid'),
            'red': PatternFill(start_color='FF7C80', end_color='FF7C80', fill_type='solid'),
            'yellow': PatternFill(start_color='FFFF00', end_color='FFFF00', fill_type='solid'),
            'orange': PatternFill(start_color='FFC000', end_color='FFC000', fill_type='solid')
        }
        
        # Light blue font for tickers and quantities
        self.ticker_font = Font(name="Arial", size=12, bold=True, color="3072C2")
        
    def load_cache_data(self):
        """Load position and yield data from cache"""
        try:
            if not os.path.exists(self.cache_file):
                print(f"❌ Cache file not found: {self.cache_file}")
                return None, None
                
            with open(self.cache_file, 'r') as f:
                cache_data = json.load(f)
                
            print(f"✅ Cache loaded: {cache_data.get('timestamp', 'No timestamp')}")
            
            # Extract ticker yields and positions
            ticker_yields = cache_data.get('ticker_yields', {})
            positions = cache_data.get('positions', {})
            
            print(f"📊 Found yields for {len(ticker_yields)} tickers")
            print(f"📊 Found positions for {len(positions)} account groups")
            
            return ticker_yields, positions
            
        except Exception as e:
            print(f"❌ Error loading cache: {e}")
            traceback.print_exc()
            return None, None
    
    def find_account_groups_and_boundaries(self, ws):
        """Find account groups and their end boundaries to know where to insert new rows"""
        account_info = {}
        
        print(f"\n🔍 Scanning for account groups and boundaries...")
        
        for row in range(1, min(60, ws.max_row + 1)):
            cell_value = ws.cell(row=row, column=1).value
            if not cell_value:
                continue
                
            cell_text = str(cell_value).strip().upper()
            
            # Look for account group headers
            if 'ETRADE' in cell_text and 'IRA' in cell_text and 'TAXABLE' not in cell_text:
                account_info["E*TRADE IRA"] = {"start_row": row, "end_row": None}
                print(f"  ✅ E*TRADE IRA found at row {row}")
            elif 'ETRADE' in cell_text and 'TAXABLE' in cell_text:
                account_info["E*TRADE Taxable"] = {"start_row": row, "end_row": None}
                print(f"  ✅ E*TRADE Taxable found at row {row}")
            elif 'SCHWAB' in cell_text and 'IRA' in cell_text:
                account_info["Schwab IRA"] = {"start_row": row, "end_row": None}
                print(f"  ✅ Schwab IRA found at row {row}")
            elif 'SCHWAB' in cell_text and 'INDIVIDUAL' in cell_text:
                account_info["Schwab Individual"] = {"start_row": row, "end_row": None}
                print(f"  ✅ Schwab Individual found at row {row}")
        
        # Find end boundaries for each group
        group_names = list(account_info.keys())
        for i, group_name in enumerate(group_names):
            start_row = account_info[group_name]["start_row"]
            
            # Look for the next group header or end of data
            if i < len(group_names) - 1:
                next_group = group_names[i + 1]
                end_row = account_info[next_group]["start_row"] - 1
            else:
                # Last group - find actual end of data
                end_row = start_row + 2  # Start with header + 2
                while end_row <= ws.max_row:
                    cell_value = ws.cell(row=end_row, column=1).value
                    if not cell_value or str(cell_value).strip() == "":
                        break
                    end_row += 1
                end_row -= 1  # Back up to last data row
                
            account_info[group_name]["end_row"] = end_row
            print(f"    📍 {group_name}: rows {start_row}-{end_row}")
        
        return account_info
    
    def get_positions_for_account(self, positions_data, account_key):
        """Get positions for a specific account from cache data"""
        account_positions = positions_data.get(account_key, [])
        
        position_dict = {}
        for position in account_positions:
            symbol = position.get('symbol', '').strip().upper()
            quantity = position.get('quantity', 0)
            market_value = position.get('market_value', 0)
            
            # Calculate current price
            current_price = market_value / quantity if quantity > 0 else 0
            
            if symbol:
                position_dict[symbol] = {
                    'quantity': quantity,
                    'current_price': current_price,
                    'market_value': market_value
                }
        
        return position_dict
    
    def get_existing_tickers_in_group(self, ws, start_row, end_row):
        """Get existing tickers in a group"""
        existing_tickers = []
        
        for row in range(start_row + 2, end_row + 1):  # Skip group header and column headers
            cell_value = ws.cell(row=row, column=1).value
            if cell_value and str(cell_value).strip() and len(str(cell_value).strip()) <= 6:
                ticker = str(cell_value).strip().upper()
                if not any(keyword in ticker for keyword in ['ETRADE', 'SCHWAB', 'TAXABLE', 'INDIVIDUAL', 'IRA', 'TICKER']):
                    existing_tickers.append(ticker)
        
        return existing_tickers
    
    def update_position_data_corrected(self, ws, account_group, account_info, positions_data, ticker_yields):
        """CORRECTED: Update position data with proper row insertion"""
        account_key = self.account_groups.get(account_group)
        if not account_key or account_group not in account_info:
            print(f"❌ Unknown account group: {account_group}")
            return
        
        print(f"\n📝 Updating position data for {account_group}...")
        
        # Get current positions from cache
        account_positions = self.get_positions_for_account(positions_data, account_key)
        
        if not account_positions:
            print(f"  ⚠️ No positions found for {account_group}")
            return
        
        print(f"  📊 Found {len(account_positions)} positions")
        
        start_row = account_info[account_group]["start_row"]
        end_row = account_info[account_group]["end_row"]
        
        # Get existing tickers
        existing_tickers = self.get_existing_tickers_in_group(ws, start_row, end_row)
        print(f"  📋 Found {len(existing_tickers)} existing tickers")
        
        # Update existing ticker data (preserve row positions)
        updated_count = 0
        for row in range(start_row + 2, end_row + 1):
            cell_value = ws.cell(row=row, column=1).value
            if cell_value:
                ticker = str(cell_value).strip().upper()
                if ticker in account_positions:
                    position_data = account_positions[ticker]
                    
                    # Column A: Apply ticker formatting
                    ticker_cell = ws.cell(row=row, column=1, value=ticker)
                    ticker_cell.font = self.ticker_font
                    
                    # Column B: Update quantity with formatting
                    quantity_cell = ws.cell(row=row, column=2, value=position_data['quantity'])
                    quantity_cell.font = self.ticker_font
                    quantity_cell.number_format = '#,##0'
                    
                    # Column D: Update price
                    price_cell = ws.cell(row=row, column=4, value=position_data['current_price'])
                    price_cell.number_format = FORMAT_CURRENCY_USD_SIMPLE
                    
                    print(f"    Row {row}: {ticker} | Qty: {position_data['quantity']} | Price: ${position_data['current_price']:.2f}")
                    updated_count += 1
        
        # Add new positions by inserting rows at the bottom of the group
        existing_tickers_set = set(existing_tickers)
        new_positions = {k: v for k, v in account_positions.items() if k not in existing_tickers_set}
        
        if new_positions:
            print(f"  📈 Adding {len(new_positions)} new positions by inserting rows...")
            
            # Insert rows at the end of the group (before next group header)
            insert_at_row = end_row + 1
            
            for ticker, position_data in new_positions.items():
                # Insert a new row
                ws.insert_rows(insert_at_row)
                
                # Column A: Ticker with formatting
                ticker_cell = ws.cell(row=insert_at_row, column=1, value=ticker)
                ticker_cell.font = self.ticker_font
                
                # Column B: Quantity with formatting
                quantity_cell = ws.cell(row=insert_at_row, column=2, value=position_data['quantity'])
                quantity_cell.font = self.ticker_font
                quantity_cell.number_format = '#,##0'
                
                # Column D: Last Price
                price_cell = ws.cell(row=insert_at_row, column=4, value=position_data['current_price'])
                price_cell.number_format = FORMAT_CURRENCY_USD_SIMPLE
                
                print(f"    Row {insert_at_row}: NEW {ticker} | Qty: {position_data['quantity']} | Price: ${position_data['current_price']:.2f}")
                
                insert_at_row += 1
                updated_count += 1
                
                # Update account_info for subsequent groups
                for group_name, info in account_info.items():
                    if info["start_row"] > end_row:
                        info["start_row"] += 1
                        info["end_row"] += 1
        
        print(f"  ✅ Updated {updated_count} position entries with proper formatting")
    
    def insert_new_yield_column(self, ws):
        """Insert new column after column O (15) for today's yields"""
        print(f"\n📝 Inserting new yield column after column O...")
        
        # Insert column after O (column 15)
        ws.insert_cols(16)  # Insert at column P (16)
        
        # Set header for new column
        new_col = 16
        header_cell = ws.cell(row=1, column=new_col, value=self.today_str)
        header_cell.font = Font(bold=True, name="Arial", size=12)
        header_cell.alignment = Alignment(horizontal='center')
        
        print(f"  ✅ Inserted yield column P with date {self.today_str}")
        return new_col
    
    def update_yield_data_corrected(self, ws, account_group, account_info, new_col, ticker_yields):
        """CORRECTED: Update yield data with actual percentages"""
        if account_group not in account_info:
            return
            
        print(f"\n📊 Updating yields for {account_group}...")
        
        start_row = account_info[account_group]["start_row"]
        end_row = account_info[account_group]["end_row"]
        
        # Add date header to group section
        date_header_row = start_row + 1
        date_cell = ws.cell(row=date_header_row, column=new_col, value=self.today_str)
        date_cell.font = Font(bold=True, name="Arial", size=10)
        date_cell.alignment = Alignment(horizontal='center')
        
        # Update yields for each ticker in the group
        updated_count = 0
        
        for row in range(start_row + 2, end_row + 1):
            ticker_cell = ws.cell(row=row, column=1)
            if ticker_cell.value:
                ticker = str(ticker_cell.value).strip().upper()
                
                if ticker in ticker_yields:
                    yield_data = ticker_yields[ticker]
                    # Get the actual yield value from cache (already a percentage)
                    dividend_yield = float(yield_data.get('yield', 0))
                    
                    # Get previous yield for color coding
                    prev_col = new_col - 1
                    prev_yield = ws.cell(row=row, column=prev_col).value
                    prev_yield = float(prev_yield) if prev_yield and str(prev_yield).replace('.', '').replace('-', '').isdigit() else None
                    
                    # Set yield value (convert percentage to decimal for Excel)
                    yield_cell = ws.cell(row=row, column=new_col, value=dividend_yield/100)
                    yield_cell.number_format = FORMAT_PERCENTAGE
                    
                    # Apply color coding
                    if prev_yield is not None:
                        prev_yield_pct = prev_yield * 100 if prev_yield < 1 else prev_yield
                        if dividend_yield > prev_yield_pct:
                            yield_cell.fill = self.colors['green']
                            color_desc = "📈 GREEN"
                        elif dividend_yield < prev_yield_pct:
                            yield_cell.fill = self.colors['red']
                            color_desc = "📉 RED"
                        else:
                            yield_cell.fill = self.colors['yellow']
                            color_desc = "➡️ YELLOW"
                    else:
                        yield_cell.fill = PatternFill(fill_type=None)
                        color_desc = "⚪ No color"
                    
                    print(f"    Row {row}: {ticker} = {dividend_yield:.2f}% ({color_desc})")
                    updated_count += 1
                else:
                    print(f"    Row {row}: {ticker} - No yield data found")
        
        print(f"  ✅ Updated {updated_count} yields for {account_group}")
    
    def apply_group_divider_color(self, ws, account_info, new_col):
        """Apply orange color to group divider rows"""
        print(f"\n🎨 Applying orange color to group dividers...")
        
        for group_name, info in account_info.items():
            row_num = info["start_row"]
            divider_cell = ws.cell(row=row_num, column=new_col)
            divider_cell.fill = self.colors['orange']
            print(f"  🟠 {group_name} divider at row {row_num}")
    
    def run_update(self):
        """Run the corrected enhanced update"""
        print("🚀 CORRECTED ENHANCED HISTORICAL YIELD UPDATER")
        print("=" * 55)
        print(f"📅 Update Date: {self.today_str}")
        print(f"📁 Excel File: {os.path.basename(self.excel_file)}")
        print(f"📋 Cache File: {os.path.basename(self.cache_file)}")
        
        # Step 1: Load cache data
        print(f"\n🔄 STEP 1: Loading cache data...")
        ticker_yields, positions_data = self.load_cache_data()
        if not ticker_yields or not positions_data:
            print("❌ Failed to load cache data")
            return False
        
        # Step 2: Load workbook
        print(f"\n📖 STEP 2: Loading Excel workbook...")
        try:
            workbook = openpyxl.load_workbook(self.excel_file)
        except Exception as e:
            print(f"❌ Error loading workbook: {e}")
            return False
        
        # Step 3: Find historical yield sheet
        historical_sheet = None
        for sheet_name in workbook.sheetnames:
            if "historical yield" in sheet_name.lower():
                historical_sheet = workbook[sheet_name]
                print(f"✅ Found sheet: {sheet_name}")
                break
        
        if not historical_sheet:
            print("❌ Could not find historical yield sheet")
            return False
        
        # Step 4: Find account groups with boundaries
        print(f"\n🔍 STEP 3: Finding account groups and boundaries...")
        account_info = self.find_account_groups_and_boundaries(historical_sheet)
        
        if len(account_info) != 4:
            print(f"❌ Expected 4 account groups, found {len(account_info)}")
            return False
        
        # Step 5: Insert new yield column first (affects all subsequent row references)
        print(f"\n📊 STEP 4: Inserting new yield column...")
        new_col = self.insert_new_yield_column(historical_sheet)
        
        # Step 6: Update position data for each group with corrected method
        print(f"\n📝 STEP 5: Updating position data with proper formatting...")
        for account_group in account_info.keys():
            self.update_position_data_corrected(historical_sheet, account_group, account_info, positions_data, ticker_yields)
        
        # Step 7: Update yields with corrected percentages
        print(f"\n📈 STEP 6: Updating yield data with actual percentages...")
        for account_group in account_info.keys():
            self.update_yield_data_corrected(historical_sheet, account_group, account_info, new_col, ticker_yields)
        
        # Step 8: Apply group divider colors
        print(f"\n🎨 STEP 7: Applying group divider colors...")
        self.apply_group_divider_color(historical_sheet, account_info, new_col)
        
        # Step 9: Save workbook
        print(f"\n💾 STEP 8: Saving workbook...")
        try:
            workbook.save(self.excel_file)
            print(f"✅ Successfully saved: {self.excel_file}")
        except Exception as e:
            print(f"❌ Error saving workbook: {e}")
            return False
        finally:
            workbook.close()
        
        print(f"\n🎉 CORRECTED UPDATE COMPLETED SUCCESSFULLY!")
        print("📊 Fixed: Proper row insertion, formatting, and actual yield percentages")
        print("🎨 Applied: Light blue formatting to tickers and quantities")
        return True

def main():
    """Main execution function"""
    try:
        updater = CorrectedEnhancedHistoricalYieldUpdater()
        success = updater.run_update()
        
        if success:
            print(f"\n✅ Corrected historical yield update completed successfully!")
        else:
            print(f"\n❌ Corrected historical yield update failed")
            
    except Exception as e:
        print(f"\n💥 Fatal error: {e}")
        traceback.print_exc()

if __name__ == "__main__":
    main()