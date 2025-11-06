"""
Complete Estimated Income 2025 Formatting Fix

This addresses:
1. Row 9 calculation: SUM(rows 4:7)/12 (corrected from 5:7)
2. Color coding for rows 4-9 comparing current vs previous column
3. Proper color scheme: Green (increase), Red (decrease), Yellow (same)
"""
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.styles.numbers import FORMAT_CURRENCY_USD_SIMPLE

def fix_estimated_income_complete():
    """Complete fix for Estimated Income 2025 formatting and calculations"""
    excel_file = r"c:\Users\mjmat\Python Code in VS\dividend_tracker\DividendTrackerApp\outputs\Dividends_2025.xlsx"
    
    try:
        # Load workbook and calculate formulas
        wb = openpyxl.load_workbook(excel_file, data_only=False)
        
        if "Estimated Income 2025" not in wb.sheetnames:
            print("❌ Estimated Income 2025 sheet not found")
            return False
            
        ws = wb["Estimated Income 2025"]
        print("🎯 FIXING ESTIMATED INCOME 2025 - COMPLETE FORMATTING")
        print("=" * 60)
        
        # Find the most recent columns (today's data)
        max_col = ws.max_column
        current_col = max_col
        prev_col = max_col - 1
        
        print(f"📊 Working with columns: Previous={openpyxl.utils.get_column_letter(prev_col)}, Current={openpyxl.utils.get_column_letter(current_col)}")
        
        # 1. FIX ROW 9 CALCULATION - Change from SUM(5:7) to SUM(4:7)
        print("\n🧮 FIXING ROW 9 CALCULATION...")
        
        monthly_row = None
        for row in range(8, 12):
            cell_value = ws.cell(row=row, column=1).value
            if cell_value and 'monthly' in str(cell_value).lower():
                monthly_row = row
                break
        
        if monthly_row:
            # Fix the formula in the current column
            col_letter = openpyxl.utils.get_column_letter(current_col)
            correct_formula = f"=SUM({col_letter}4:{col_letter}7)/12"
            
            cell = ws.cell(row=monthly_row, column=current_col)
            cell.value = correct_formula
            cell.number_format = FORMAT_CURRENCY_USD_SIMPLE
            cell.font = Font(name='Arial', size=12)
            print(f"   ✅ Row {monthly_row}: {correct_formula}")
        
        # Save and reopen to calculate formulas
        wb.save(excel_file)
        wb.close()
        
        # Reopen with calculated values for color coding
        wb = openpyxl.load_workbook(excel_file, data_only=True)
        ws = wb["Estimated Income 2025"]
        
        # 2. APPLY COLOR CODING - Compare current vs previous column
        print(f"\n🎨 APPLYING COLOR CODING (Rows 4-9)...")
        
        if prev_col >= 2:  # Only if there's a previous column to compare
            
            colors_applied = 0
            
            for row in range(4, 10):  # Rows 4-9 (accounts + monthly average)
                current_cell = ws.cell(row=row, column=current_col)
                prev_cell = ws.cell(row=row, column=prev_col)
                
                current_val = current_cell.value
                prev_val = prev_cell.value
                
                # Only apply color coding if both values exist and are numeric
                if (current_val is not None and prev_val is not None and 
                    isinstance(current_val, (int, float)) and isinstance(prev_val, (int, float))):
                    
                    account_name = ws.cell(row=row, column=1).value
                    
                    if current_val > prev_val:
                        # Green for increase
                        current_cell.fill = PatternFill(start_color='90EE90', end_color='90EE90', fill_type='solid')
                        change_type = "📈 INCREASE (Green)"
                        colors_applied += 1
                    elif current_val < prev_val:
                        # Red for decrease
                        current_cell.fill = PatternFill(start_color='FFB6C1', end_color='FFB6C1', fill_type='solid')  
                        change_type = "📉 DECREASE (Red)"
                        colors_applied += 1
                    else:
                        # Yellow for same
                        current_cell.fill = PatternFill(start_color='FFFF99', end_color='FFFF99', fill_type='solid')
                        change_type = "➡️ SAME (Yellow)"
                        colors_applied += 1
                    
                    print(f"   Row {row} ({account_name}): ${prev_val:,.2f} → ${current_val:,.2f} {change_type}")
        
        # 3. ENSURE CONSISTENT FORMATTING
        print(f"\n🎨 ENSURING CONSISTENT FORMATTING...")
        
        # Format the current column properly
        for row in range(4, 10):  # Rows 4-9
            cell = ws.cell(row=row, column=current_col)
            if cell.value is not None:
                cell.font = Font(name='Arial', size=12)
                if isinstance(cell.value, (int, float)):
                    cell.number_format = FORMAT_CURRENCY_USD_SIMPLE
        
        # Set column width
        ws.column_dimensions[openpyxl.utils.get_column_letter(current_col)].width = 15
        
        # Save the final result
        wb.save(excel_file)
        wb.close()
        
        print(f"\n✅ ESTIMATED INCOME 2025 COMPLETE FIX APPLIED!")
        print("📋 Summary of changes:")
        print("   ✅ Row 9 calculation: Fixed to SUM(rows 4:7)/12")
        print(f"   ✅ Color coding: Applied to {colors_applied} cells")
        print("   ✅ Formatting: Arial 12pt, proper currency format")
        print("   ✅ Column width: Set to 15 for proper display")
        
        return True
        
    except Exception as e:
        print(f"❌ ERROR in complete fix: {e}")
        import traceback
        traceback.print_exc()
        return False

if __name__ == "__main__":
    fix_estimated_income_complete()
