import os
import openpyxl

def check_total_row():
    """Check what's in the total row and fix the formula"""
    excel_file = "outputs/Dividends_2025.xlsx"
    
    if not os.path.exists(excel_file):
        print(f"ERROR: Excel file not found: {excel_file}")
        return
    
    try:
        wb = openpyxl.load_workbook(excel_file)
        ws = wb['Portfolio Values 2025']
        
        print("TOTAL ROW ANALYSIS")
        print("=" * 50)
        
        # Check row 10 (total row) across all columns
        for col in range(1, min(ws.max_column + 1, 10)):  # Check first 10 columns
            cell = ws.cell(row=10, column=col)
            col_letter = openpyxl.utils.get_column_letter(col)
            
            if hasattr(cell, 'data_type') and cell.data_type == 'f':
                print(f"Column {col_letter}: FORMULA = {cell.value}")
            else:
                print(f"Column {col_letter}: VALUE = {cell.value}")
        
        # Check the most recent columns
        print(f"\nMOST RECENT COLUMNS (last 5):")
        for col in range(max(1, ws.max_column - 4), ws.max_column + 1):
            cell = ws.cell(row=10, column=col)
            col_letter = openpyxl.utils.get_column_letter(col)
            
            if hasattr(cell, 'data_type') and cell.data_type == 'f':
                print(f"Column {col_letter}: FORMULA = {cell.value}")
            else:
                print(f"Column {col_letter}: VALUE = {cell.value}")
        
        # Check what values are actually in rows 4-8 in the latest column
        print(f"\nLATEST COLUMN ({openpyxl.utils.get_column_letter(ws.max_column)}) VALUES:")
        for row in range(4, 9):
            cell = ws.cell(row=row, column=ws.max_column)
            account_name = ws.cell(row=row, column=1).value
            print(f"Row {row} ({account_name}): {cell.value}")
        
        # Calculate manual sum
        values = []
        for row in range(4, 9):
            cell = ws.cell(row=row, column=ws.max_column)
            if cell.value is not None and isinstance(cell.value, (int, float)):
                values.append(cell.value)
                
        manual_sum = sum(values)
        print(f"\nMANUAL SUM of rows 4-8: ${manual_sum:,.2f}")
        print(f"Current total in row 10: ${ws.cell(row=10, column=ws.max_column).value:,.2f}")
        
        wb.close()
        
    except Exception as e:
        print(f"ERROR: {e}")
        import traceback
        traceback.print_exc()

if __name__ == "__main__":
    check_total_row()