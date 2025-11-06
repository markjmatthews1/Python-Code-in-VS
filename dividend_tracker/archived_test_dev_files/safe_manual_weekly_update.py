"""
Safe Manual Weekly Update Script
Purpose: Add new weekly data WITHOUT destroying existing historical data
Author: Assistant (GitHub Copilot)
Created: August 31, 2025

This script SAFELY adds new weekly data to your existing sheets without
destroying your 33 weeks of historical tracking data.
"""

import os
import sys
import openpyxl
from datetime import datetime
import tkinter as tk
from tkinter import simpledialog

# Configuration
OUTPUT_DIR = os.path.join(os.path.dirname(__file__), "outputs")
TARGET_FILE = os.path.join(OUTPUT_DIR, "Dividends_2025.xlsx")

def backup_workbook():
    """Create a safety backup before any changes"""
    timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
    backup_name = f"Dividends_2025_backup_safe_update_{timestamp}.xlsx"
    backup_path = os.path.join(OUTPUT_DIR, backup_name)
    
    import shutil
    shutil.copy2(TARGET_FILE, backup_path)
    print(f"📋 Backup created: {backup_name}")
    return backup_path

def get_401k_value_simple():
    """Simple console prompt for 401K value"""
    try:
        value_str = input("Enter current 401K value (e.g., 125000): $")
        return float(value_str.replace(',', '').replace('$', ''))
    except:
        print("Invalid input, using 0")
        return 0

def get_monthly_dividend_input(account_name):
    """Get manual input for monthly dividend amount"""
    try:
        value_str = input(f"Enter monthly dividend estimate for {account_name}: $")
        return float(value_str.replace(',', '').replace('$', ''))
    except:
        print("Invalid input, using 0")
        return 0

def get_portfolio_value_input(account_name):
    """Get manual input for portfolio value"""
    try:
        value_str = input(f"Enter current portfolio value for {account_name}: $")
        return float(value_str.replace(',', '').replace('$', ''))
    except:
        print("Invalid input, using 0")
        return 0

def find_next_available_column(worksheet, start_col=2):
    """Find the next empty column for new data"""
    col = start_col
    while worksheet.cell(row=1, column=col).value is not None:
        col += 1
    return col

def add_weekly_dividend_data_manual(wb, current_date):
    """Add new weekly dividend data with manual input"""
    
    if "All account weekly dividends" not in wb.sheetnames:
        print("❌ 'All account weekly dividends' sheet not found")
        return False
    
    ws = wb["All account weekly dividends"]
    next_col = find_next_available_column(ws)
    
    print(f"\n📊 Adding dividend data to column {next_col}")
    
    # Add date header
    ws.cell(row=1, column=next_col, value=current_date)
    
    # Get account names from column A (starting row 2)
    accounts = []
    row = 2
    while ws.cell(row=row, column=1).value:
        account_name = ws.cell(row=row, column=1).value
        accounts.append(account_name)
        row += 1
    
    print(f"Found {len(accounts)} accounts to update:")
    for i, account in enumerate(accounts, 2):
        print(f"  Row {i}: {account}")
    
    # Get dividend values for each account
    print(f"\n💰 Enter monthly dividend estimates for each account:")
    for i, account in enumerate(accounts, 2):
        dividend_value = get_monthly_dividend_input(account)
        ws.cell(row=i, column=next_col, value=dividend_value)
    
    print(f"✅ Added weekly dividend data for {current_date}")
    return True

def add_weekly_portfolio_data_manual(wb, current_date):
    """Add new weekly portfolio data with manual input"""
    
    if "Portfolio Values 2025" not in wb.sheetnames:
        print("❌ 'Portfolio Values 2025' sheet not found")
        return False
        
    ws = wb["Portfolio Values 2025"]
    next_col = find_next_available_column(ws)
    
    print(f"\n💰 Adding portfolio data to column {next_col}")
    
    # Add date header
    ws.cell(row=1, column=next_col, value=current_date)
    
    # Get account names from column A
    accounts = []
    row = 2
    while ws.cell(row=row, column=1).value:
        account_name = ws.cell(row=row, column=1).value
        accounts.append(account_name)
        row += 1
    
    print(f"Found {len(accounts)} portfolio accounts to update:")
    for i, account in enumerate(accounts, 2):
        print(f"  Row {i}: {account}")
    
    # Get portfolio values for each account
    print(f"\n💼 Enter current portfolio values:")
    total_value = 0
    
    for i, account in enumerate(accounts, 2):
        if "401" in account.upper():
            # Special handling for 401K
            portfolio_value = get_401k_value_simple()
        else:
            portfolio_value = get_portfolio_value_input(account)
        
        ws.cell(row=i, column=next_col, value=portfolio_value)
        total_value += portfolio_value
    
    print(f"✅ Added weekly portfolio data for {current_date}")
    print(f"   Total Portfolio Value: ${total_value:,.2f}")
    return True

def update_estimated_income_manual(wb, current_date):
    """Update Estimated Income 2025 sheet with manual input"""
    
    if "Estimated Income 2025" not in wb.sheetnames:
        print("❌ 'Estimated Income 2025' sheet not found")
        return False
        
    ws = wb["Estimated Income 2025"]
    next_col = find_next_available_column(ws)
    
    print(f"\n💵 Adding estimated income data to column {next_col}")
    
    # Add date header
    ws.cell(row=1, column=next_col, value=current_date)
    
    # Get account names from column A
    accounts = []
    row = 2
    while ws.cell(row=row, column=1).value:
        account_name = ws.cell(row=row, column=1).value
        accounts.append(account_name)
        row += 1
    
    print(f"Found {len(accounts)} accounts for estimated income:")
    for i, account in enumerate(accounts, 2):
        print(f"  Row {i}: {account}")
    
    # Get estimated income for each account
    print(f"\n📈 Enter estimated monthly income for each account:")
    for i, account in enumerate(accounts, 2):
        income_value = get_monthly_dividend_input(account)
        ws.cell(row=i, column=next_col, value=income_value)
    
    print(f"✅ Added estimated income data for {current_date}")
    return True

def safe_manual_update():
    """Perform safe manual weekly update"""
    
    print("🔒 === SAFE MANUAL WEEKLY UPDATE ===")
    print("This will ADD new weekly data WITHOUT destroying existing historical data")
    print("You'll be prompted to enter values for each account manually")
    print()
    
    if not os.path.exists(TARGET_FILE):
        print(f"❌ Workbook not found: {TARGET_FILE}")
        return False
    
    # Create backup
    backup_path = backup_workbook()
    
    try:
        # Load existing workbook
        wb = openpyxl.load_workbook(TARGET_FILE)
        current_date = datetime.now().strftime('%m/%d/%Y')
        
        print(f"📅 Adding data for: {current_date}")
        
        # Show available sheets
        print(f"\n📋 Available sheets in workbook:")
        for sheet_name in wb.sheetnames:
            print(f"   - {sheet_name}")
        
        success_count = 0
        
        # 1. Update "All account weekly dividends" sheet
        print(f"\n" + "="*60)
        print("STEP 1: Weekly Dividend Data")
        print("="*60)
        if add_weekly_dividend_data_manual(wb, current_date):
            success_count += 1
            
        # 2. Update "Portfolio Values 2025" sheet
        print(f"\n" + "="*60)
        print("STEP 2: Portfolio Values Data")
        print("="*60)
        if add_weekly_portfolio_data_manual(wb, current_date):
            success_count += 1
            
        # 3. Update "Estimated Income 2025" sheet
        print(f"\n" + "="*60)
        print("STEP 3: Estimated Income Data") 
        print("="*60)
        if update_estimated_income_manual(wb, current_date):
            success_count += 1
        
        # Save changes
        wb.save(TARGET_FILE)
        
        print(f"\n✅ SAFE MANUAL UPDATE COMPLETE")
        print(f"   Successfully updated {success_count}/3 sheets")
        print(f"   Backup saved: {os.path.basename(backup_path)}")
        print(f"   All historical data preserved!")
        print(f"   New data added for: {current_date}")
        
        return True
        
    except Exception as e:
        print(f"❌ SAFE UPDATE FAILED: {e}")
        print(f"   Restore from backup: {os.path.basename(backup_path)}")
        return False

if __name__ == "__main__":
    safe_manual_update()
