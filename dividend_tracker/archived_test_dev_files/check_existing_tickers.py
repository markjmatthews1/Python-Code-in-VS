import openpyxl

def check_existing_tickers():
    file_path = "outputs/Dividends_2025.xlsx"
    
    try:
        wb = openpyxl.load_workbook(file_path)
        ws = wb["Accounts Div historical yield"]
        
        print("EXISTING TICKERS IN EXCEL:")
        print("=" * 30)
        
        # Check E*TRADE IRA section (rows 3-20)
        print("E*TRADE IRA (rows 3-20):")
        for row in range(3, 21):
            ticker_cell = ws.cell(row=row, column=1)
            ticker_value = ticker_cell.value
            if ticker_value:
                print(f"  Row {row}: {ticker_value}")
        
        # Check E*TRADE Taxable section (rows 25-36)
        print("\nE*TRADE Taxable (rows 25-36):")
        for row in range(25, 37):
            ticker_cell = ws.cell(row=row, column=1)
            ticker_value = ticker_cell.value
            if ticker_value:
                print(f"  Row {row}: {ticker_value}")
        
        wb.close()
        
    except Exception as e:
        print(f"Error: {e}")

if __name__ == "__main__":
    check_existing_tickers()