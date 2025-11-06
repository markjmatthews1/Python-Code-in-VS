#!/usr/bin/env python3
"""
Examine Excel Sheet Formatting Issues
====================================
"""

import openpyxl
import os

def examine_formatting_issues():
    """Examine the current formatting state of the sheet"""
    
    excel_path = r"c:\Users\mjmat\Python Code in VS\dividend_tracker\DividendTrackerApp\outputs\Dividends_2025.xlsx"
    
    try:
        workbook = openpyxl.load_workbook(excel_path)
        sheet = workbook["Accounts dividend historical yi"]
        
        print(f"📊 EXAMINING FORMATTING ISSUES")
        print(f"=" * 60)
        print(f"📏 Sheet dimensions: {sheet.max_row} rows × {sheet.max_column} columns\n")
        
        # Check key rows for formatting issues
        key_rows = [1, 2, 25, 26, 27, 28, 41, 42, 43, 44, 49, 50, 51, 55]
        
        for row_num in key_rows:
            if row_num <= sheet.max_row:
                # Check several columns for content and formatting
                cells_info = []
                for col in range(1, min(20, sheet.max_column + 1)):  # Check first 20 columns
                    cell = sheet.cell(row=row_num, column=col)
                    content = str(cell.value) if cell.value else ""
                    
                    # Get background color if any
                    fill_color = "none"
                    if cell.fill.start_color.rgb and cell.fill.start_color.rgb != "00000000":
                        fill_color = str(cell.fill.start_color.rgb)
                    
                    if content or fill_color != "none":
                        cells_info.append(f"Col{col}='{content}'({fill_color[:8]})")
                
                if cells_info:
                    print(f"Row {row_num:2d}: {' | '.join(cells_info[:5])}...")  # Show first 5 non-empty cells
        
        # Look specifically at column 16 (P) which should be our inserted column
        print(f"\n🔍 COLUMN 16 (P) - INSERTED COLUMN:")
        for row_num in [1, 2, 25, 27, 28, 41, 43, 44, 49, 50, 51, 55]:
            if row_num <= sheet.max_row:
                cell = sheet.cell(row=row_num, column=16)
                content = str(cell.value) if cell.value else "EMPTY"
                fill_color = "none"
                if cell.fill.start_color.rgb and cell.fill.start_color.rgb != "00000000":
                    fill_color = str(cell.fill.start_color.rgb)[:8]
                
                print(f"   Row {row_num:2d}: '{content}' (color: {fill_color})")
        
        workbook.close()
        
    except Exception as e:
        print(f"❌ Error: {e}")

if __name__ == "__main__":
    examine_formatting_issues()
