#!/usr/bin/env python3
"""
Verify the new dividend statistics column
"""
import openpyxl

def verify_dividend_stats():
    excel_file = 'outputs/Dividends_2025.xlsx'
    
    try:
        wb = openpyxl.load_workbook(excel_file, data_only=True)
        portfolio_ws = wb['Portfolio Summary']
        
        print("📊 NEW DIVIDEND STATISTICS COLUMN (Column D)")
        print("=" * 60)
        
        # Show the new dividend statistics section
        for row in range(36, min(75, portfolio_ws.max_row + 1)):
            col_d_value = portfolio_ws.cell(row=row, column=4).value
            col_e_value = portfolio_ws.cell(row=row, column=5).value
            
            if col_d_value or col_e_value:
                if col_e_value:
                    print(f"Row {row}: {col_d_value} | {col_e_value}")
                else:
                    print(f"Row {row}: {col_d_value}")
        
        wb.close()
        
    except Exception as e:
        print(f"Error: {e}")

if __name__ == "__main__":
    verify_dividend_stats()
