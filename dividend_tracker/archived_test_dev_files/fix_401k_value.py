"""
Fix 401K Value and Portfolio Total

The 401K row exists but has None value. 
Need to add $125,882.01 to get total of $526,619.21
"""
import openpyxl
from openpyxl.styles import Font
from openpyxl.styles.numbers import FORMAT_CURRENCY_USD_SIMPLE

def fix_401k_value():
    """Add the missing 401K value"""
    excel_file = r"c:\Users\mjmat\Python Code in VS\dividend_tracker\DividendTrackerApp\outputs\Dividends_2025.xlsx"
    
    try:
        wb = openpyxl.load_workbook(excel_file)
        ws = wb["Portfolio Values 2025"]
        
        print("💰 FIXING 401K VALUE")
        print("=" * 25)
        
        max_col = ws.max_column
        current_col_letter = openpyxl.utils.get_column_letter(max_col)
        
        k401_value = 125882.01  # Value you entered in GUI
        
        # Add 401K value to Row 8
        ws.cell(row=8, column=max_col).value = k401_value
        ws.cell(row=8, column=max_col).number_format = FORMAT_CURRENCY_USD_SIMPLE
        ws.cell(row=8, column=max_col).font = Font(name='Arial', size=12)
        
        print(f"✅ Added 401K value: ${k401_value:,.2f} to Row 8, Column {current_col_letter}")
        
        # Recalculate total for row 10
        total = 0
        for row in range(4, 9):  # Rows 4-8
            value = ws.cell(row=row, column=max_col).value
            if value and isinstance(value, (int, float)):
                total += value
        
        # Update row 10 with correct total
        for row in range(9, 15):
            account_name = ws.cell(row=row, column=1).value
            if account_name and 'total' in str(account_name).lower():
                ws.cell(row=row, column=max_col).value = total
                ws.cell(row=row, column=max_col).number_format = FORMAT_CURRENCY_USD_SIMPLE
                ws.cell(row=row, column=max_col).font = Font(name='Arial', size=12, bold=True)
                print(f"✅ Updated Row {row} total: ${total:,.2f}")
                break
        
        print(f"\n📊 Portfolio breakdown:")
        for row in range(4, 9):
            account_name = ws.cell(row=row, column=1).value
            value = ws.cell(row=row, column=max_col).value
            if account_name and value:
                print(f"   {account_name}: ${value:,.2f}")
        
        print(f"📊 Total Portfolio: ${total:,.2f}")
        
        # Verify this matches your expected $526,619.21
        expected = 526619.21
        if abs(total - expected) < 1.0:  # Within $1
            print(f"✅ Total matches expected: ${expected:,.2f}")
        else:
            print(f"⚠️ Total ${total:,.2f} doesn't match expected ${expected:,.2f}")
            print(f"   Difference: ${total - expected:,.2f}")
        
        wb.save(excel_file)
        wb.close()
        
        return True
        
    except Exception as e:
        print(f"❌ ERROR: {e}")
        import traceback
        traceback.print_exc()
        return False

if __name__ == "__main__":
    fix_401k_value()
