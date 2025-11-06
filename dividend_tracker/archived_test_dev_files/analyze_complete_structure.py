#!/usr/bin/env python3
"""
Analyze Complete Enhanced Sheet Structure
========================================
"""

import openpyxl
import os

def analyze_complete_structure():
    """Analyze the complete enhanced sheet structure"""
    
    excel_path = r"c:\Users\mjmat\Python Code in VS\dividend_tracker\DividendTrackerApp\outputs\Dividends_2025.xlsx"
    
    try:
        workbook = openpyxl.load_workbook(excel_path)
        target_sheet = workbook["Etrade IRA historic yield"]
        
        print(f"📊 COMPLETE ENHANCED SHEET ANALYSIS")
        print(f"=" * 60)
        print(f"📏 Dimensions: {target_sheet.max_row} rows × {target_sheet.max_column} columns\n")
        
        # Analyze all rows to identify the account structure
        print(f"🏦 Account Structure:")
        for row_num in range(1, target_sheet.max_row + 1):
            cell_a = target_sheet.cell(row=row_num, column=1).value
            cell_b = target_sheet.cell(row=row_num, column=2).value
            cell_c = target_sheet.cell(row=row_num, column=3).value
            
            # Look for account titles, headers, and ticker patterns
            if cell_a and isinstance(cell_a, str):
                if any(keyword in cell_a.lower() for keyword in ['etrade', 'schwab', 'ira', 'taxable', 'individual']):
                    print(f"🏛️  Row {row_num:2d}: ACCOUNT TITLE - {cell_a}")
                elif cell_a.lower() == 'ticker':
                    print(f"📋 Row {row_num:2d}: HEADER ROW - Ticker | {cell_b} | {cell_c}")
                elif len(cell_a) <= 6 and cell_a.isupper():  # Likely a ticker
                    print(f"📈 Row {row_num:2d}: TICKER - {cell_a} | {cell_b}")
            elif not cell_a and not cell_b and not cell_c:
                # Look for separator rows
                if row_num > 1:
                    prev_cell = target_sheet.cell(row=row_num-1, column=1).value
                    next_cell = target_sheet.cell(row=row_num+1, column=1).value
                    if prev_cell or next_cell:
                        print(f"➖ Row {row_num:2d}: SEPARATOR ROW")
        
        # Check for any existing formulas in average rows
        print(f"\n📊 Average/Formula Rows:")
        for row_num in [25, 41, 49, 55]:  # Potential average rows
            if row_num <= target_sheet.max_row:
                cell_p = target_sheet.cell(row=row_num, column=16)  # Column P
                if cell_p.value:
                    print(f"🧮 Row {row_num:2d}: Column P = {cell_p.value}")
        
        workbook.close()
        
    except Exception as e:
        print(f"❌ Error: {e}")

if __name__ == "__main__":
    analyze_complete_structure()
