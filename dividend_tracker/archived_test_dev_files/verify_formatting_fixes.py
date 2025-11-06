#!/usr/bin/env python3

import openpyxl
import os
from datetime import datetime

class FormattingVerifier:
    def __init__(self):
        self.excel_path = r"C:\Users\mjmat\Python Code in VS\dividend_tracker\DividendTrackerApp\outputs\Dividends_2025.xlsx"
    
    def verify_formatting_fixes(self):
        """Verify the formatting improvements in the historical yield sheet"""
        print("FORMATTING VERIFICATION - HISTORICAL YIELD SHEET")
        print("=" * 55)
        print(f"Verification Time: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}")
        print()
        
        if not os.path.exists(self.excel_path):
            print(f"ERROR: Excel file not found: {self.excel_path}")
            return
        
        wb = openpyxl.load_workbook(self.excel_path)
        ws = wb['Accounts Div historical yield']
        
        # Find account groups
        account_groups = {}
        for row in range(1, 60):
            cell_value = ws.cell(row=row, column=1).value
            if cell_value:
                cell_text = str(cell_value).strip().upper()
                if cell_text == 'ETRADE IRA':
                    account_groups["E*TRADE IRA"] = row
                elif cell_text == 'ETRADE TAXABLE':
                    account_groups["E*TRADE Taxable"] = row
                elif cell_text == 'SCHWAB IRA':
                    account_groups["Schwab IRA"] = row
                elif cell_text == 'SCHWAB INDIVIDUAL':
                    account_groups["Schwab Individual"] = row
        
        print("1. ORANGE DIVIDER BAR CHECK:")
        print("-" * 30)
        for group_name, row in account_groups.items():
            # Check Column A (group header)
            col_a_cell = ws.cell(row=row, column=1)
            col_a_fill = col_a_cell.fill
            col_a_orange = False
            if col_a_fill and col_a_fill.start_color:
                rgb = col_a_fill.start_color.rgb
                col_a_orange = rgb in ['00FFC000', 'FFFFC000', 'FFC000']
            
            # Check Column P (yield divider)  
            col_p_cell = ws.cell(row=row, column=16)
            col_p_fill = col_p_cell.fill
            col_p_orange = False
            if col_p_fill and col_p_fill.start_color:
                rgb = col_p_fill.start_color.rgb
                col_p_orange = rgb in ['00FFC000', 'FFFFC000', 'FFC000']
            
            print(f"  {group_name} (Row {row}):")
            print(f"    Column A Orange: {'✅ YES' if col_a_orange else '❌ NO'} (RGB: {col_a_fill.start_color.rgb if col_a_fill.start_color else 'None'})")
            print(f"    Column P Orange: {'✅ YES' if col_p_orange else '❌ NO'} (RGB: {col_p_fill.start_color.rgb if col_p_fill.start_color else 'None'})")
        
        print("\n2. YIELD PERCENTAGE FORMAT CHECK:")
        print("-" * 35)
        
        # Check a few sample yield cells
        sample_rows = [3, 25, 41, 49]  # One from each group
        
        for row in sample_rows:
            ticker_cell = ws.cell(row=row, column=1)
            yield_cell = ws.cell(row=row, column=16)  # Column P
            
            if ticker_cell.value and yield_cell.value is not None:
                ticker = ticker_cell.value
                yield_value = yield_cell.value
                yield_format = yield_cell.number_format
                yield_bg_color = yield_cell.fill.start_color.rgb if yield_cell.fill.start_color else "None"
                
                # Check if it's numeric vs string
                is_numeric = isinstance(yield_value, (int, float))
                
                print(f"  Row {row}: {ticker}")
                print(f"    Value: {yield_value} ({'NUMERIC' if is_numeric else 'STRING'})")
                print(f"    Format: {yield_format}")
                print(f"    Background Color: {yield_bg_color}")
        
        print("\n3. YIELD COLUMN HEADER CHECK:")
        print("-" * 30)
        
        # Check column P header in each group
        for group_name, row in account_groups.items():
            header_cell = ws.cell(row=row+1, column=16)  # Header row
            header_value = header_cell.value
            header_font = header_cell.font.bold if header_cell.font else False
            
            print(f"  {group_name}: '{header_value}' (Bold: {'✅' if header_font else '❌'})")
        
        print("\n4. COMPARISON COLOR CODING CHECK:")
        print("-" * 35)
        
        # Check if colors are based on comparison with column O
        for row in sample_rows:
            ticker_cell = ws.cell(row=row, column=1)
            old_yield_cell = ws.cell(row=row, column=15)  # Column O
            new_yield_cell = ws.cell(row=row, column=16)  # Column P
            
            if ticker_cell.value:
                ticker = ticker_cell.value
                old_value = old_yield_cell.value
                new_value = new_yield_cell.value
                new_color = new_yield_cell.fill.start_color.rgb if new_yield_cell.fill.start_color else "None"
                
                print(f"  Row {row}: {ticker}")
                print(f"    Old (Col O): {old_value}")
                print(f"    New (Col P): {new_value}")
                print(f"    Background Color: {new_color}")
        
        wb.close()
        
        print("\n✅ FORMATTING VERIFICATION COMPLETED!")
        print("\nExpected Results:")
        print("- Orange divider bars: Column A AND Column P should both be orange")
        print("- Yield values: Should be NUMERIC with '0.00%' format")
        print("- Background colors: Light Green (yield up), Light Red (yield down), Yellow (same value), None (no previous data)")

if __name__ == "__main__":
    verifier = FormattingVerifier()
    verifier.verify_formatting_fixes()