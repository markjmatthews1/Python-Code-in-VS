"""
Final Fix for Row 9

Manually set the formula and apply color coding
"""
import openpyxl
from openpyxl.styles import Font, PatternFill
from openpyxl.styles.numbers import FORMAT_CURRENCY_USD_SIMPLE

def final_fix_row9():
    """Final fix for Row 9 formula and color coding"""
    excel_file = r"c:\Users\mjmat\Python Code in VS\dividend_tracker\DividendTrackerApp\outputs\Dividends_2025.xlsx"
    
    try:
        wb = openpyxl.load_workbook(excel_file)
        ws = wb["Estimated Income 2025"]
        
        max_col = ws.max_column
        prev_col = max_col - 1
        current_col = max_col
        
        current_col_letter = openpyxl.utils.get_column_letter(current_col)
        prev_col_letter = openpyxl.utils.get_column_letter(prev_col)
        
        print(f"🔧 FINAL ROW 9 FIX")
        print("=" * 30)
        
        # 1. Set the correct formula in Row 9
        formula = f"=SUM({current_col_letter}4:{current_col_letter}7)/12"
        
        cell = ws.cell(row=9, column=current_col)
        cell.value = formula
        cell.number_format = FORMAT_CURRENCY_USD_SIMPLE
        cell.font = Font(name='Arial', size=12)
        
        print(f"✅ Row 9 formula set: {formula}")
        
        # 2. Get previous month value from Row 9 for comparison
        prev_cell = ws.cell(row=9, column=prev_col)
        prev_formula_val = prev_cell.value
        
        # Calculate current month value manually for color comparison
        # Sum of rows 4-7 in current column / 12
        sum_val = 0
        for r in range(4, 8):
            val = ws.cell(row=r, column=current_col).value
            if val is not None and isinstance(val, (int, float)):
                sum_val += val
        
        current_monthly = sum_val / 12
        
        print(f"📊 Previous monthly: {prev_formula_val}")
        print(f"📊 Current monthly: {current_monthly:.2f}")
        
        # 3. Apply color coding based on comparison
        if prev_formula_val is not None and isinstance(prev_formula_val, (int, float)):
            if current_monthly > prev_formula_val:
                # Green for increase
                cell.fill = PatternFill(start_color='90EE90', end_color='90EE90', fill_type='solid')
                color_applied = "GREEN (increase)"
            elif current_monthly < prev_formula_val:
                # Red for decrease
                cell.fill = PatternFill(start_color='FFB6C1', end_color='FFB6C1', fill_type='solid')
                color_applied = "RED (decrease)"
            else:
                # Yellow for same
                cell.fill = PatternFill(start_color='FFFF99', end_color='FFFF99', fill_type='solid')
                color_applied = "YELLOW (same)"
            
            print(f"🎨 Color applied: {color_applied}")
        else:
            print("🎨 No color applied (no previous value for comparison)")
        
        # Save the changes
        wb.save(excel_file)
        wb.close()
        
        print(f"\n✅ FINAL ROW 9 FIX COMPLETE!")
        
        return True
        
    except Exception as e:
        print(f"❌ ERROR: {e}")
        import traceback
        traceback.print_exc()
        return False

if __name__ == "__main__":
    final_fix_row9()
