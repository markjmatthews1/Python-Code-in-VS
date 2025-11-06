#!/usr/bin/env python3
"""
Enhanced Accounts Dividend Historical Yield Updater for DividendTracker App
===========================================================================

This module updates the "Accounts dividend historical yield" sheet in Dividends_2025.xlsx 
with current yield data for all account sections:
- Etrade IRA
- Etrade Taxable  
- Schwab IRA
- Schwab Individual

Each account section gets yield updates with proper color coding and average calculations.

Original Reference: Update_dividend_sheet.py (unchanged)
Integration: Part of weekend DividendTracker automation
Target: Dividends_2025.xlsx -> "Accounts dividend historical yield" sheet

Author: DividendTracker Integration (Enhanced Multi-Account)
Date: August 30, 2025
"""

import os
import sys
import json
import time
from datetime import datetime, date
import openpyxl
from openpyxl.styles import PatternFill, Font, Border, Side, Alignment
from openpyxl.utils import get_column_letter

# Add modules directory to path for DividendTracker imports
sys.path.append(os.path.join(os.path.dirname(__file__), 'modules'))

try:
    # Import DividendTracker E*TRADE authentication (from modules directory)
    from etrade_auth import get_etrade_session
except ImportError as e:
    print(f"❌ ERROR: Could not import DividendTracker E*TRADE authentication: {e}")
    print("Ensure etrade_auth.py is in the modules/ directory")
    sys.exit(1)

class AccountsHistoricYieldUpdater:
    """Updates the Accounts dividend historical yield sheet with current yield data from E*TRADE API"""
    
    def __init__(self):
        self.workbook_path = os.path.join(os.path.dirname(__file__), "outputs", "Dividends_2025.xlsx")
        self.sheet_name = "Accounts dividend historical yield"
        self.session = None
        self.base_url = None
        
        # Define account sections based on the enhanced structure
        self.account_sections = [
            {
                "name": "Etrade IRA",
                "ticker_start_row": 3,
                "ticker_end_row": 20,
                "average_row": 25
            },
            {
                "name": "Etrade Taxable", 
                "ticker_start_row": 29,
                "ticker_end_row": 40,
                "average_row": 41
            },
            {
                "name": "Schwab IRA",
                "ticker_start_row": 44,
                "ticker_end_row": 47,
                "average_row": 49
            },
            {
                "name": "Schwab Individual",
                "ticker_start_row": 52,
                "ticker_end_row": 53,
                "average_row": 55
            }
        ]
        
    def initialize_etrade_auth(self):
        """Initialize E*TRADE authentication using DividendTracker auth system"""
        try:
            print("🔐 Initializing E*TRADE authentication...")
            self.session, self.base_url = get_etrade_session()
            
            if self.session and self.base_url:
                print("✅ E*TRADE authentication successful")
                return True
            else:
                print("❌ E*TRADE authentication failed")
                return False
                
        except Exception as e:
            print(f"❌ E*TRADE authentication error: {e}")
            return False
    
    def get_quote_data(self, symbol):
        """Get quote data for a symbol including yield information"""
        try:
            if not self.session or not self.base_url:
                print("❌ E*TRADE session not initialized")
                return None
                
            quote_url = f"{self.base_url}/v1/market/quote/{symbol}.json"
            response = self.session.get(quote_url)
            
            print(f"   🔍 API Response for {symbol}: Status {response.status_code}")
            
            if response.status_code == 401:
                print(f"⚠️ 401 Unauthorized for {symbol}. Refreshing session...")
                # Try to refresh session
                self.session, self.base_url = get_etrade_session()
                if self.session:
                    response = self.session.get(quote_url)
                    print(f"   🔍 Retry API Response for {symbol}: Status {response.status_code}")
                else:
                    print(f"❌ Could not refresh session for {symbol}")
                    return None
            
            if response.status_code == 200:
                data = response.json()
                print(f"   📋 Response keys: {list(data.keys()) if data else 'None'}")
                
                if 'QuoteResponse' in data and 'QuoteData' in data['QuoteResponse']:
                    quote_data = data['QuoteResponse']['QuoteData'][0]
                    
                    # Extract yield information (dividend yield)
                    yield_info = {}
                    if 'dividendYield' in quote_data:
                        yield_info['yield'] = quote_data['dividendYield']
                        print(f"   💰 Found dividendYield: {quote_data['dividendYield']}")
                    elif 'annualDividend' in quote_data and 'lastTrade' in quote_data:
                        # Calculate yield if not provided directly
                        annual_div = float(quote_data.get('annualDividend', 0))
                        price = float(quote_data.get('lastTrade', 0))
                        if price > 0:
                            yield_info['yield'] = (annual_div / price) * 100
                            print(f"   💰 Calculated yield: {yield_info['yield']:.4f}% (${annual_div} / ${price})")
                    else:
                        print(f"   ⚠️ No yield data found. Available keys: {list(quote_data.keys())}")
                    
                    return yield_info
                else:
                    print(f"   ❌ Unexpected response structure: {data}")
            else:
                print(f"   ❌ API Error {response.status_code}: {response.text[:200]}")
                return None
                
        except Exception as e:
            print(f"   ❌ Error getting quote for {symbol}: {e}")
            return None

    def update_historic_yield_sheet(self):
        """Update the Accounts dividend historical yield sheet with current yield data for all accounts"""
        
        try:
            print("🚀 STARTING ACCOUNTS DIVIDEND HISTORICAL YIELD UPDATE")
            print("=" * 60)
            print(f"🕐 Started: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}")
            print(f"📁 Target workbook: {self.workbook_path}")
            print(f"📊 Target sheet: {self.sheet_name}")
            
            # Initialize E*TRADE authentication
            if not self.initialize_etrade_auth():
                print("❌ Failed to initialize E*TRADE authentication")
                return False
            
            # Open the workbook
            print("📖 Loading workbook...")
            workbook = openpyxl.load_workbook(self.workbook_path)
            sheet = workbook[self.sheet_name]
            print(f"✅ Successfully opened sheet: {self.sheet_name}")
            print(f"📏 Sheet dimensions: {sheet.max_row} rows × {sheet.max_column} columns")
            
            # Insert new yield column after 'Beginning Dividend Yield' (column 15/O)
            insert_col = 16  # Column P
            today = date.today().strftime("%m-%d-%Y")
            
            print(f"📅 Inserting new yield column after 'Beginning Dividend Yield' at column {insert_col}")
            sheet.insert_cols(insert_col)
            
            # Add header for new column to all account sections
            print(f"📅 Added new yield column header: {today} (Column {insert_col})")
            for account in self.account_sections:
                header_row = account["ticker_start_row"] - 1  # Header is one row before tickers
                if header_row > 0:
                    header_cell = sheet.cell(row=header_row, column=insert_col)
                    header_cell.value = today
            
            # Process each account section
            total_updated = 0
            total_errors = 0
            
            for account in self.account_sections:
                print(f"\n🏦 Processing {account['name']}...")
                account_updated = 0
                account_errors = 0
                
                # Process tickers in this account
                for row_idx in range(account["ticker_start_row"], account["ticker_end_row"] + 1):
                    ticker_cell = sheet.cell(row=row_idx, column=1)  # Column A contains ticker
                    ticker = ticker_cell.value
                    
                    if ticker and isinstance(ticker, str) and len(ticker.strip()) > 0:
                        ticker = ticker.strip().upper()
                        print(f"   🔍 Processing {ticker}...", end=" ")
                        
                        # Get quote data from E*TRADE
                        quote_data = self.get_quote_data(ticker)
                        
                        if quote_data:
                            # Get current yield from quote
                            current_yield = quote_data.get('yield', 0)
                            
                            if current_yield and current_yield != 0:
                                # Set yield value in the new inserted column
                                yield_cell = sheet.cell(row=row_idx, column=insert_col)
                                # Divide by 100 since API returns 10.15 for 10.15% but Excel percentage format will multiply by 100
                                yield_cell.value = round(float(current_yield) / 100, 4)
                                
                                # Format the cell as percentage with 2 decimal places
                                yield_cell.number_format = '0.00%'
                                
                                # Apply color coding by comparing against "Beginning Dividend Yield" (column 15/O)
                                beginning_yield_cell = sheet.cell(row=row_idx, column=15)  # Column O
                                beginning_yield = beginning_yield_cell.value
                                
                                try:
                                    if beginning_yield and float(beginning_yield) != 0:
                                        current_val = float(current_yield) / 100  # Divide by 100 to match our stored format
                                        beginning_val = float(beginning_yield) / 100  # Assuming beginning yield is also in percentage format
                                        
                                        if current_val > beginning_val:
                                            # Green for increase from beginning yield
                                            fill_color = "00FF00"
                                        elif current_val < beginning_val:
                                            # Red for decrease from beginning yield
                                            fill_color = "FF0000"
                                        else:
                                            # Yellow for same as beginning yield
                                            fill_color = "FFFF00"
                                            
                                        fill = PatternFill(start_color=fill_color, end_color=fill_color, fill_type="solid")
                                        yield_cell.fill = fill
                                        
                                except (ValueError, TypeError):
                                    pass  # Skip color coding if comparison fails
                                
                                print(f"✅ {current_yield}% (vs {beginning_yield}% beginning)")
                                account_updated += 1
                            else:
                                print("❌ No yield data")
                                account_errors += 1
                        else:
                            print("❌ Quote failed")
                            account_errors += 1
                            
                        # Small delay to avoid overwhelming the API
                        time.sleep(0.5)
                
                # Add average calculation for this account
                try:
                    print(f"   📊 Adding average calculation for {account['name']}...")
                    avg_cell = sheet.cell(row=account["average_row"], column=insert_col)
                    # Calculate average of tickers in this account (values are already divided by 100)
                    start_row = account["ticker_start_row"]
                    end_row = account["ticker_end_row"]
                    avg_cell.value = f"=ROUND(AVERAGE({get_column_letter(insert_col)}{start_row}:{get_column_letter(insert_col)}{end_row}),4)"
                    
                    # Format the average cell as percentage with 2 decimal places
                    avg_cell.number_format = '0.00%'
                    
                    print(f"   ✅ Added average formula for {account['name']}")
                        
                except Exception as e:
                    print(f"   ⚠️ Warning: Could not add average calculation for {account['name']}: {str(e)}")
                
                print(f"   📊 {account['name']}: ✅ {account_updated} updated, ❌ {account_errors} errors")
                total_updated += account_updated
                total_errors += account_errors
            
            # Save the workbook
            print(f"\n💾 Saving workbook...")
            workbook.save(self.workbook_path)
            workbook.close()
            
            # Summary
            print("\n✅ ACCOUNTS DIVIDEND HISTORICAL YIELD UPDATE COMPLETE")
            print("=" * 60)
            print(f"📊 Results Summary:")
            print(f"   ✅ Successfully updated: {total_updated} tickers across all accounts")
            print(f"   ❌ Errors encountered: {total_errors} tickers")
            print(f"   📅 New column inserted: {today} (Column {insert_col})")
            print(f"   🏦 Accounts processed: {len(self.account_sections)} accounts")
            print(f"   💾 File saved: {self.workbook_path}")
            print(f"🕐 Completed: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}")
            
            return True
            
        except Exception as e:
            print(f"❌ Critical error during historic yield update: {e}")
            try:
                if 'workbook' in locals():
                    workbook.save(self.workbook_path)
                    workbook.close()
                    print("💾 Emergency save completed")
            except:
                pass
            return False

def run_historic_yield_update():
    """Main entry point for historic yield update"""
    updater = AccountsHistoricYieldUpdater()
    return updater.update_historic_yield_sheet()

if __name__ == "__main__":
    print("🧪 TESTING: Enhanced Accounts Historic Yield Updater")
    success = run_historic_yield_update()
    
    if success:
        print("\n🎉 Historic yield update completed successfully!")
    else:
        print("\n💥 Historic yield update failed!")
    
    print("\n⚠️ NOTE: This module is designed to be called by the main DividendTracker weekend automation.")
    print("For production use, integrate this into your weekend update script.")
