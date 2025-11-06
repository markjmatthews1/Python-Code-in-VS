#!/usr/bin/env python3
"""
Fix 401K Retirement Row - Properly Handle Merged Cells
The row 18 appears to be merged across columns A and B, need to unmerge and set properly
"""
import openpyxl
from openpyxl.styles import Font, Alignment

def fix_401k_row_properly():
    """Fix the 401K row by handling merged cells correctly"""
    
    excel_file = 'outputs/Dividends_2025.xlsx'
    
    try:
        print("🔧 Fixing 401K Retirement row with proper merged cell handling...")
        
        wb = openpyxl.load_workbook(excel_file)
        portfolio_ws = wb['Portfolio Summary']
        
        # Calculate 401K percentage
        k401_value = 124315.15
        total_portfolio = 519439.06
        k401_pct = k401_value / total_portfolio
        
        # Define formatting
        normal_font = Font(name='Arial', size=12)
        right_aligned = Alignment(horizontal='right')
        percentage_format = '0.0%'
        
        print(f"📊 Looking for merged cells affecting row 18...")
        
        # Find and handle merged cells in row 18
        merged_ranges_to_remove = []
        for merged_range in portfolio_ws.merged_cells.ranges:
            if merged_range.min_row <= 18 <= merged_range.max_row:
                print(f"   Found merged range affecting row 18: {merged_range}")
                merged_ranges_to_remove.append(merged_range)
        
        # Remove merged cells that affect row 18
        for merged_range in merged_ranges_to_remove:
            portfolio_ws.unmerge_cells(str(merged_range))
            print(f"   Unmerged range: {merged_range}")
        
        # Now set the values properly
        label_cell = portfolio_ws.cell(row=18, column=1)
        label_cell.value = "401k Retirement:"
        label_cell.font = normal_font
        
        pct_cell = portfolio_ws.cell(row=18, column=2)
        pct_cell.value = k401_pct
        pct_cell.number_format = percentage_format
        pct_cell.font = normal_font
        pct_cell.alignment = right_aligned
        
        print(f"   ✅ Set 401K Retirement label in A18: '401k Retirement:'")
        print(f"   ✅ Set 401K Retirement percentage in B18: {k401_pct:.1%}")
        
        # Save the workbook
        wb.save(excel_file)
        wb.close()
        
        return True
        
    except Exception as e:
        print(f"❌ Error fixing 401K row: {e}")
        return False

def verify_401k_final():
    """Verify the 401K row is now correct"""
    excel_file = 'outputs/Dividends_2025.xlsx'
    
    try:
        wb = openpyxl.load_workbook(excel_file, data_only=True)
        portfolio_ws = wb['Portfolio Summary']
        
        print("\n📊 FINAL VERIFICATION - Account Breakdown Section:")
        print("=" * 70)
        
        # Show account breakdown section with exact row numbers
        for row in range(13, 23):
            col_a = portfolio_ws.cell(row=row, column=1).value or ""
            col_b = portfolio_ws.cell(row=row, column=2).value or ""
            
            if col_a:
                # Format the display
                if isinstance(col_b, (int, float)) and col_b < 1 and col_b > 0:
                    display_text = f"Row {row:2d}: {col_a:<25} {col_b:.1%}"
                elif isinstance(col_b, (int, float)) and col_b > 1000:
                    display_text = f"Row {row:2d}: {col_a:<25} ${col_b:,.2f}"
                else:
                    display_text = f"Row {row:2d}: {col_a:<25} {col_b}"
                
                print(display_text)
                
                # Highlight row 18 specifically
                if row == 18:
                    if isinstance(col_b, (int, float)) and col_b < 1 and col_b > 0:
                        print(f"         ✅ 401K percentage correctly set: {col_b:.1%}")
                    else:
                        print(f"         ❌ 401K percentage still missing or incorrect: '{col_b}'")
        
        wb.close()
        
    except Exception as e:
        print(f"Error: {e}")

if __name__ == "__main__":
    print("🎯 FIXING 401K RETIREMENT ROW - MERGED CELL HANDLING")
    print("=" * 70)
    print("Issue: Row 18 appears to be merged across columns A and B")
    print("Solution: Unmerge cells and set values properly")
    print("Target: Row 18 Column A = '401k Retirement:', Column B = 23.9%")
    print("=" * 70)
    
    success = fix_401k_row_properly()
    
    if success:
        verify_401k_final()
        print("\n🎉 401K Retirement row fix completed!")
        print("✅ Row 18 should now show '401k Retirement:' in A18 and 23.9% in B18")
    else:
        print("\n❌ Could not fix 401K Retirement row")
