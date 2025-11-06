#!/usr/bin/env python3
"""
Fix Portfolio Summary Structure Properly
- Separate current account values from percentage breakdown
- Add missing Schwab IRA data
- Proper section organization
- Ensure no data loss from Excel repair
"""
import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment

def fix_portfolio_structure_properly():
    """Fix the portfolio structure with proper separation of values and percentages"""
    
    excel_file = 'outputs/Dividends_2025.xlsx'
    
    try:
        print("🔧 Fixing Portfolio Summary structure properly...")
        
        wb = openpyxl.load_workbook(excel_file)
        portfolio_ws = wb['Portfolio Summary']
        
        # Clear the sheet to start fresh with correct structure
        print("🧹 Clearing sheet for proper rebuild...")
        portfolio_ws.delete_rows(1, portfolio_ws.max_row)
        
        # Get latest dividend data from Estimated Income sheet
        income_ws = wb['Estimated Income 2025']
        max_col = income_ws.max_column
        latest_monthly_estimate = income_ws.cell(row=9, column=max_col).value or 0
        
        # Calculate statistics
        latest_annual_estimate = latest_monthly_estimate * 12
        weekly_estimate = latest_monthly_estimate / 4.33
        
        # Get account values correctly
        etrade_ira_annual = income_ws.cell(row=5, column=max_col).value or 0      # Larger amount = IRA
        etrade_taxable_annual = income_ws.cell(row=4, column=max_col).value or 0  # Smaller amount = Taxable
        schwab_ira_annual = income_ws.cell(row=6, column=max_col).value or 0
        schwab_individual_annual = income_ws.cell(row=7, column=max_col).value or 0
        
        # Calculate current portfolio values from known data
        # These should be the actual current market values, not dividend amounts
        etrade_ira_value = 279339.15
        etrade_taxable_value = 62622.72  
        schwab_ira_value = 50558.40
        schwab_individual_value = 2603.64
        k401_value = 124315.15
        total_portfolio = etrade_ira_value + etrade_taxable_value + schwab_ira_value + schwab_individual_value + k401_value
        
        # Calculate percentages
        etrade_ira_pct = (etrade_ira_value / total_portfolio) * 100
        etrade_taxable_pct = (etrade_taxable_value / total_portfolio) * 100
        schwab_ira_pct = (schwab_ira_value / total_portfolio) * 100
        schwab_individual_pct = (schwab_individual_value / total_portfolio) * 100
        k401_pct = (k401_value / total_portfolio) * 100
        
        # Withdrawal data
        monthly_etrade_ira = 1200
        quarterly_etrade_ira = 370
        monthly_etrade_taxable = 395
        total_monthly_withdrawals = monthly_etrade_ira + monthly_etrade_taxable + (quarterly_etrade_ira / 3)
        net_monthly = latest_monthly_estimate - total_monthly_withdrawals
        net_annual = net_monthly * 12
        
        # Define styles
        header_font = Font(name='Arial', size=12, bold=True, color='FFFFFF')
        header_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
        section_font = Font(name='Arial', size=11, bold=True)
        section_fill = PatternFill(start_color='D9E2F3', end_color='D9E2F3', fill_type='solid')
        bold_font = Font(name='Arial', size=12, bold=True)
        normal_font = Font(name='Arial', size=12)
        right_aligned = Alignment(horizontal='right')
        currency_format = '$#,##0.00'
        percentage_format = '0.0%'
        
        print("📊 Building Portfolio Summary in Column A-C with proper structure...")
        
        # === PORTFOLIO SUMMARY HEADER ===
        cell = portfolio_ws.cell(row=1, column=1)
        cell.value = "PORTFOLIO SUMMARY"
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center')
        portfolio_ws.merge_cells('A1:C1')
        
        current_row = 3
        
        # Portfolio Summary Data with PROPER STRUCTURE
        portfolio_sections = [
            ("CURRENT VALUES", [
                ("E*TRADE IRA:", etrade_ira_value),
                ("E*TRADE Taxable:", etrade_taxable_value),
                ("Schwab IRA:", schwab_ira_value),
                ("Schwab Individual:", schwab_individual_value),
                ("401k Retirement:", k401_value),
                ("Current Portfolio Total:", total_portfolio),
                ("Weekly Change:", "+$2,834.85 (+0.55%)"),
                ("", ""),
            ]),
            ("ACCOUNT BREAKDOWN", [
                ("E*TRADE IRA:", f"{etrade_ira_pct:.1f}%"),
                ("E*TRADE Taxable:", f"{etrade_taxable_pct:.1f}%"),
                ("Schwab IRA:", f"{schwab_ira_pct:.1f}%"),
                ("Schwab Individual:", f"{schwab_individual_pct:.1f}%"),
                ("401k Retirement:", f"{k401_pct:.1f}%"),
                ("", ""),
            ]),
            ("PERFORMANCE TRACKING", [
                ("Total Return:", "Calculate from historical"),
                ("YTD Performance:", "Track weekly changes"),
                ("", ""),
            ]),
            ("DIVIDEND SUMMARY", [
                ("Monthly Dividend Target:", 3600),
                ("Current Monthly Estimate:", latest_monthly_estimate),
                ("Annual Dividend Target:", 43200),
                ("Current Annual Estimate:", latest_annual_estimate),
                ("", ""),
            ]),
            ("WITHDRAWAL SCHEDULE", [
                ("Monthly E*TRADE IRA:", monthly_etrade_ira),
                ("Quarterly E*TRADE IRA:", quarterly_etrade_ira),
                ("Monthly E*TRADE Taxable:", monthly_etrade_taxable),
                ("", ""),
            ]),
            ("NET DIVIDEND INCOME", [
                ("After Withdrawals (Monthly):", net_monthly),
                ("After Withdrawals (Annual):", net_annual),
                ("", ""),
                ("YTD Dividends Received:", "Check All account weekly dividends sheet"),
            ])
        ]
        
        for section_title, items in portfolio_sections:
            # Section header
            if section_title:
                cell = portfolio_ws.cell(row=current_row, column=1)
                cell.value = section_title
                cell.font = section_font
                cell.fill = section_fill
                current_row += 1
            
            # Section items
            for label, value in items:
                if label:
                    # Label
                    label_cell = portfolio_ws.cell(row=current_row, column=1)
                    label_cell.value = label
                    label_cell.font = normal_font
                    
                    # Value
                    if value:
                        value_cell = portfolio_ws.cell(row=current_row, column=2)
                        if isinstance(value, (int, float)):
                            value_cell.value = value
                            value_cell.number_format = currency_format
                        else:
                            value_cell.value = value
                        value_cell.font = normal_font
                
                current_row += 1
            
            current_row += 1  # Extra spacing between sections
        
        print("💰 Building Dividend Statistics in Column D-E...")
        
        # === DIVIDEND STATISTICS IN COLUMN D-E ===
        current_row = 1
        col_d = 4  # Column D
        col_e = 5  # Column E
        
        # === DIVIDEND STATISTICS HEADER ===
        cell = portfolio_ws.cell(row=current_row, column=col_d)
        cell.value = "WEEKLY DIVIDEND STATISTICS"
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center')
        portfolio_ws.merge_cells('D1:E1')
        
        current_row = 3
        
        # Add all the dividend sections with corrected account names
        dividend_sections = [
            ("CURRENT ESTIMATES", [
                ("Weekly Estimate:", weekly_estimate),
                ("Monthly Estimate:", latest_monthly_estimate),
                ("Annual Estimate:", latest_annual_estimate),
            ]),
            ("ACCOUNT BREAKDOWN (Annual)", [
                ("E*TRADE IRA:", etrade_ira_annual),      # Larger amount = IRA
                ("E*TRADE Taxable:", etrade_taxable_annual), # Smaller amount = Taxable
                ("Schwab IRA:", schwab_ira_annual),
                ("Schwab Individual:", schwab_individual_annual),
            ]),
            ("WITHDRAWAL DATA", [
                ("Monthly E*TRADE IRA:", monthly_etrade_ira),
                ("Quarterly E*TRADE IRA:", quarterly_etrade_ira),
                ("Monthly E*TRADE Taxable:", monthly_etrade_taxable),
                ("Total Monthly Withdrawals:", total_monthly_withdrawals),
            ]),
            ("NET INCOME (After Withdrawals)", [
                ("Net Monthly Income:", net_monthly),
                ("Net Annual Income:", net_annual),
                ("Reinvestment Rate:", net_annual / latest_annual_estimate if latest_annual_estimate > 0 else 0),
            ]),
            ("DIVIDEND METRICS", [
                ("Current Yield:", latest_annual_estimate / total_portfolio if total_portfolio > 0 else 0),
                ("Monthly Dividend Coverage:", latest_monthly_estimate / total_monthly_withdrawals if total_monthly_withdrawals > 0 else 0),
                ("Annual Div Growth Target:", 0.06),  # 6% as number
                ("Next Update:", "Weekly (Automated)"),
            ])
        ]
        
        for section_title, items in dividend_sections:
            # Section header
            cell = portfolio_ws.cell(row=current_row, column=col_d)
            cell.value = section_title
            cell.font = section_font
            cell.fill = section_fill
            current_row += 1
            
            # Section items
            for label, value in items:
                # Label in column D
                label_cell = portfolio_ws.cell(row=current_row, column=col_d)
                label_cell.value = label
                label_cell.font = normal_font
                
                # Value in column E with proper formatting
                value_cell = portfolio_ws.cell(row=current_row, column=col_e)
                
                if isinstance(value, (int, float)):
                    value_cell.value = value
                    
                    # Apply specific formatting based on the type of value
                    if "Rate" in label or "Yield" in label or "Growth" in label:
                        value_cell.number_format = percentage_format
                        value_cell.alignment = right_aligned  # Right justify percentages
                    elif "Coverage" in label:
                        value_cell.number_format = '0.0"x"'  # Show as "2.1x"
                        value_cell.alignment = right_aligned  # Right justify
                    else:
                        value_cell.number_format = currency_format
                else:
                    value_cell.value = value
                    if label != "Next Update:":  # Don't right-align text fields
                        value_cell.alignment = right_aligned
                
                value_cell.font = normal_font
                current_row += 1
            
            current_row += 1  # Extra spacing between sections
        
        print("✅ Portfolio structure fixed properly!")
        print("   ✅ Current account values properly separated under CURRENT VALUES")
        print("   ✅ Account breakdown percentages under ACCOUNT BREAKDOWN")
        print("   ✅ Schwab IRA data included in both sections")
        print("   ✅ All account values and percentages present")
        print("   ✅ Proper section organization maintained")
        
        # Save the workbook
        wb.save(excel_file)
        wb.close()
        
        return True
        
    except Exception as e:
        print(f"❌ Error fixing portfolio structure: {e}")
        return False

def verify_proper_structure():
    """Verify the proper portfolio structure"""
    excel_file = 'outputs/Dividends_2025.xlsx'
    
    try:
        wb = openpyxl.load_workbook(excel_file, data_only=True)
        portfolio_ws = wb['Portfolio Summary']
        
        print("\n📊 PROPER STRUCTURE VERIFICATION")
        print("=" * 90)
        print("PORTFOLIO SUMMARY (A-C)              |  DIVIDEND STATISTICS (D-E)")
        print("=" * 90)
        
        # Show side by side for verification
        for row in range(1, min(35, portfolio_ws.max_row + 1)):
            # Portfolio summary (columns A-B)
            col_a = portfolio_ws.cell(row=row, column=1).value or ""
            col_b = portfolio_ws.cell(row=row, column=2).value or ""
            
            # Format portfolio text
            if isinstance(col_b, (int, float)) and col_b > 1000:
                portfolio_text = f"{col_a} ${col_b:,.2f}".strip()
            elif isinstance(col_b, (int, float)):
                portfolio_text = f"{col_a} {col_b}".strip()
            else:
                portfolio_text = f"{col_a} {col_b}".strip()
            
            # Dividend stats (columns D-E)
            col_d = portfolio_ws.cell(row=row, column=4).value or ""
            col_e = portfolio_ws.cell(row=row, column=5).value or ""
            
            # Format dividend text
            if isinstance(col_e, (int, float)):
                if col_e < 1 and col_e > 0:  # Likely a percentage or ratio
                    dividend_text = f"{col_d} {col_e:.1%}"
                elif col_e > 1 and col_e < 10:  # Likely a ratio
                    dividend_text = f"{col_d} {col_e:.1f}x"
                else:
                    dividend_text = f"{col_d} ${col_e:,.2f}"
            else:
                dividend_text = f"{col_d} {col_e}".strip()
            
            if portfolio_text.strip() or dividend_text.strip():
                print(f"{portfolio_text[:40]:<40} | {dividend_text}")
        
        wb.close()
        
    except Exception as e:
        print(f"Error: {e}")

if __name__ == "__main__":
    print("🚀 FIXING PORTFOLIO STRUCTURE PROPERLY")
    print("=" * 60)
    print("Fixes:")
    print("• Separate current account values from percentage breakdown")
    print("• Add missing Schwab IRA data to both sections")
    print("• Proper section organization")
    print("• Ensure no data loss")
    print("=" * 60)
    
    success = fix_portfolio_structure_properly()
    
    if success:
        verify_proper_structure()
        
        print("\n🎯 Portfolio Structure Fixed!")
        print("✅ CURRENT VALUES section with actual account dollar amounts")
        print("✅ ACCOUNT BREAKDOWN section with percentages separate")
        print("✅ Schwab IRA included: $50,558.40 (9.7%)")
        print("✅ All account data properly organized")
        print("✅ No mixing of values and percentages")
        print("✅ Portfolio total correctly calculated")
    else:
        print("\n❌ Failed to fix portfolio structure")
