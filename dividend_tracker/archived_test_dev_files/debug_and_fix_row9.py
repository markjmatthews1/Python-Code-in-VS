"""
Debug Row Structure and Fix Row 9 Formula

Check the exact structure to find where Monthly Average row is located
and fix the formula.
"""
import openpyxl
from openpyxl.styles import Font
from openpyxl.styles.numbers import FORMAT_CURRENCY_USD_SIMPLE

def debug_and_fix_row9():
    """Debug row structure and fix Row 9 formula"""
    excel_file = r"c:\Users\mjmat\Python Code in VS\dividend_tracker\DividendTrackerApp\outputs\Dividends_2025.xlsx"
    
    try:
        wb = openpyxl.load_workbook(excel_file)
        ws = wb["Estimated Income 2025"]
        
        print("🔍 DEBUGGING ROW STRUCTURE")
        print("=" * 40)
        
        # Check rows 1-15 to find the Monthly Average
        for row in range(1, 16):
            cell_value = ws.cell(row=row, column=1).value
            if cell_value:
                print(f"   Row {row}: {cell_value}")
        
        # Find Monthly Average row specifically
        monthly_row = None
        for row in range(1, 20):  # Expanded search range
            cell_value = ws.cell(row=row, column=1).value
            if cell_value and 'monthly' in str(cell_value).lower():
                monthly_row = row
                print(f"\n🎯 Found Monthly Average at Row {monthly_row}: {cell_value}")
                break
        
        if not monthly_row:
            print("\n⚠️ Monthly Average row not found! Checking for alternative labels...")
            for row in range(1, 20):
                cell_value = ws.cell(row=row, column=1).value
                if cell_value:
                    lower_val = str(cell_value).lower()
                    if any(word in lower_val for word in ['average', 'month', 'monthly']):
                        monthly_row = row
                        print(f"🎯 Found potential monthly row at Row {monthly_row}: {cell_value}")
                        break
        
        # If still not found, look for row 9 specifically
        if not monthly_row:
            row9_value = ws.cell(row=9, column=1).value
            print(f"\n📋 Row 9 contains: {row9_value}")
            if row9_value:  # If row 9 has any content, assume it's the monthly row
                monthly_row = 9
                print(f"🎯 Using Row 9 as monthly average row")
        
        if monthly_row:
            # Fix the formula
            max_col = ws.max_column
            current_col = max_col
            col_letter = openpyxl.utils.get_column_letter(current_col)
            
            # Set correct formula: SUM(rows 4:7)/12
            correct_formula = f"=SUM({col_letter}4:{col_letter}7)/12"
            
            cell = ws.cell(row=monthly_row, column=current_col)
            cell.value = correct_formula
            cell.number_format = FORMAT_CURRENCY_USD_SIMPLE
            cell.font = Font(name='Arial', size=12)
            
            print(f"\n✅ FIXED Row {monthly_row} formula: {correct_formula}")
            
            # Save the fix
            wb.save(excel_file)
            wb.close()
            
            print("✅ Row 9 formula fix applied and saved!")
            
            return True
        else:
            print("\n❌ Could not locate Monthly Average row")
            wb.close()
            return False
        
    except Exception as e:
        print(f"❌ ERROR: {e}")
        import traceback
        traceback.print_exc()
        return False

if __name__ == "__main__":
    debug_and_fix_row9()
