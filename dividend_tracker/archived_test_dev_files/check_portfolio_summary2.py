import openpyxl

# Load the workbook and Portfolio Summary sheet
wb = openpyxl.load_workbook('outputs/Dividends_2025.xlsx')
ws = wb['Portfolio Summary']

print("PORTFOLIO SUMMARY SHEET ANALYSIS - PART 2")
print("=" * 50)

print("\nRows 20-35, columns A-F:")
for i in range(20, 36):
    row_data = []
    for j in range(1, 7):
        cell_value = ws.cell(i, j).value
        if cell_value is None:
            row_data.append("None")
        elif isinstance(cell_value, (int, float)):
            row_data.append(f"{cell_value:.2f}")
        else:
            row_data.append(str(cell_value)[:30])
    print(f"Row {i:2d}: {row_data}")

# Check for any errors in specific areas
print(f"\nSpecific checks:")
print(f"Total Portfolio Value (B9): {ws.cell(9, 2).value}")
print(f"Annual Dividend Estimate (E6): {ws.cell(6, 5).value}")
print(f"Weekly Estimate (E4): {ws.cell(4, 5).value}")
print(f"Monthly Estimate (E5): {ws.cell(5, 5).value}")

wb.close()