import openpyxl

def verify_detailed_reductions():
    """Verify the detailed dividend reduction results"""
    
    workbook_path = "outputs/Dividends_2025.xlsx"
    wb = openpyxl.load_workbook(workbook_path)
    portfolio_sheet = wb["Portfolio Summary"]
    
    print("DETAILED DIVIDEND REDUCTIONS VERIFICATION")
    print("=" * 60)
    print("\nColumns G-H content (Dividend Status):")
    
    # Read all the dividend status columns
    for row in range(1, 40):  # Extended range
        col_g = portfolio_sheet.cell(row=row, column=7).value
        col_h = portfolio_sheet.cell(row=row, column=8).value
        
        if col_g or col_h:
            # Format the output nicely
            g_val = str(col_g) if col_g else ""
            h_val = str(col_h) if col_h else ""
            
            if g_val.startswith("  "):  # Indented ticker lines
                print(f"      {g_val.strip():<12} | {h_val}")
            elif "dividend cuts:" in h_val.lower() or "no reductions" in h_val.lower():
                print(f"📊 {g_val:<15} | {h_val}")
            elif g_val in ["DIVIDEND REDUCTIONS", "TOP DIVIDEND PAYERS", "DIVIDEND METRICS"]:
                print(f"\n🔶 {g_val}")
            elif g_val == "DIVIDEND STATUS":
                print(f"📋 {g_val}")
            elif col_g and col_h:
                print(f"   {g_val:<15} | {h_val}")
    
    wb.close()

if __name__ == "__main__":
    verify_detailed_reductions()
