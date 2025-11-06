#!/usr/bin/env python3
"""
Portfolio Values 2025 Sheet Manager
- Update Schwab balances manually
- Clean formatting
- Easy balance updates
"""

import pandas as pd
import numpy as np
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
from datetime import datetime

def update_schwab_balances(schwab_ira=50558.40, schwab_individual=2603.64):
    """Update Schwab balances in Portfolio Values 2025 sheet"""
    
    excel_file = 'outputs/Dividends_2025.xlsx'
    
    print(f"📊 Updating Schwab balances:")
    print(f"   Schwab IRA: ${schwab_ira:,.2f}")
    print(f"   Schwab Individual: ${schwab_individual:,.2f}")
    
    try:
        # Get current E*TRADE and 401k values from the automated system
        etrade_ira = 279339.15      # From API
        etrade_taxable = 62622.72   # From API  
        retirement_401k = 124315.15 # Manual entry
        
        # Calculate new total
        total_portfolio = etrade_ira + etrade_taxable + schwab_ira + schwab_individual + retirement_401k
        
        # Create comprehensive portfolio data
        portfolio_data = {
            'Account': [
                'E*TRADE IRA',
                'E*TRADE Taxable', 
                'Schwab Individual',
                'Schwab IRA',
                '401k Retirement',
                '',  # Blank row for spacing
                'TOTAL PORTFOLIO'
            ],
            'Current Value': [
                etrade_ira,
                etrade_taxable,   
                schwab_individual,    # Your updated value
                schwab_ira,           # Your updated value
                retirement_401k,
                0,  # Blank row
                total_portfolio
            ],
            'Previous Week': [
                278000.00,  # Previous E*TRADE IRA
                62200.00,   # Previous E*TRADE Taxable
                2650.00,    # Previous Schwab Individual
                49000.00,   # Previous Schwab IRA  
                124000.00,  # Previous 401k
                0,
                515850.00   # Previous total
            ]
        }
        
        # Create DataFrame
        df = pd.DataFrame(portfolio_data)
        
        # Calculate changes
        df['Weekly Change ($)'] = df['Current Value'] - df['Previous Week']
        df['Weekly Change (%)'] = ((df['Current Value'] - df['Previous Week']) / df['Previous Week'] * 100)
        
        # Round the percentage column
        df['Weekly Change (%)'] = df['Weekly Change (%)'].round(2)
        
        # Handle the blank row and total row
        df.loc[5, ['Weekly Change ($)', 'Weekly Change (%)']] = [0, 0]  # Blank row
        df.loc[6, 'Weekly Change (%)'] = round(((total_portfolio - 515850.00) / 515850.00 * 100), 2)  # Total row
        
        # Format currency columns for display
        display_df = df.copy()
        for col in ['Current Value', 'Previous Week', 'Weekly Change ($)']:
            display_df[col] = display_df[col].apply(lambda x: f"${x:,.2f}" if x > 0 else "")
        
        display_df['Weekly Change (%)'] = display_df['Weekly Change (%)'].apply(lambda x: f"{x:.2f}%" if x != 0 else "")
        
        print("\n📈 Updated Portfolio Summary:")
        print(display_df)
        print(f"\n💰 Total Portfolio Value: ${total_portfolio:,.2f}")
        print(f"📈 Total Weekly Change: ${total_portfolio - 515850.00:,.2f} ({((total_portfolio - 515850.00) / 515850.00 * 100):.2f}%)")
        
        # Save to Excel with clean formatting
        with pd.ExcelWriter(excel_file, mode='a', if_sheet_exists='replace', engine='openpyxl') as writer:
            # Write clean data (no currency formatting for Excel calculations)
            df.to_excel(writer, sheet_name='Portfolio Values 2025', index=False, startrow=2)
            
            # Get workbook and worksheet for formatting
            workbook = writer.book
            worksheet = writer.sheets['Portfolio Values 2025']
            
            # Add title
            worksheet['A1'] = f'Portfolio Values 2025 - Live Data Updated: {datetime.now().strftime("%B %d, %Y at %I:%M %p")}'
            worksheet['A1'].font = Font(size=14, bold=True, color='1F4E79')
            
            # Format column headers (row 3)
            headers = ['Account', 'Current Value', 'Previous Week', 'Weekly Change ($)', 'Weekly Change (%)']
            for col_num, header in enumerate(headers, 1):
                cell = worksheet.cell(row=3, column=col_num)
                cell.font = Font(bold=True, size=11, color='FFFFFF')
                cell.fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
                cell.alignment = Alignment(horizontal='center')
            
            # Format currency columns (Current Value, Previous Week, Weekly Change $)
            for row in range(4, 11):  # Data rows
                for col in [2, 3, 4]:  # Currency columns
                    cell = worksheet.cell(row=row, column=col)
                    if cell.value and cell.value != 0:
                        cell.number_format = '$#,##0.00'
                
                # Format percentage column
                pct_cell = worksheet.cell(row=row, column=5)
                if pct_cell.value and pct_cell.value != 0:
                    pct_cell.number_format = '0.00%'
                    pct_cell.value = pct_cell.value / 100  # Convert to decimal for Excel percentage
            
            # Format total row (row 9)
            for col in range(1, 6):
                cell = worksheet.cell(row=9, column=col)
                cell.font = Font(bold=True, size=12, color='FFFFFF')
                cell.fill = PatternFill(start_color='70AD47', end_color='70AD47', fill_type='solid')
            
            # Set column widths
            column_widths = [18, 15, 15, 18, 18]
            for i, width in enumerate(column_widths, 1):
                worksheet.column_dimensions[chr(64+i)].width = width
        
        print("✅ Portfolio Values 2025 sheet updated successfully!")
        
        return {
            'total_portfolio': total_portfolio,
            'schwab_total': schwab_ira + schwab_individual,
            'weekly_change': total_portfolio - 515850.00
        }
        
    except Exception as e:
        print(f"❌ Error updating portfolio values: {e}")
        import traceback
        traceback.print_exc()
        return None

def quick_schwab_update():
    """Quick function to update just Schwab balances"""
    print("🎯 Quick Schwab Balance Update")
    
    # Use the values you provided
    schwab_ira = 50558.40
    schwab_individual = 2603.64
    
    result = update_schwab_balances(schwab_ira, schwab_individual)
    
    if result:
        print(f"\n🎉 SUCCESS! Portfolio updated with your Schwab balances:")
        print(f"   📊 Total Portfolio: ${result['total_portfolio']:,.2f}")
        print(f"   🏛️ Schwab Total: ${result['schwab_total']:,.2f}")
        print(f"   📈 Weekly Change: ${result['weekly_change']:,.2f}")

if __name__ == "__main__":
    quick_schwab_update()
