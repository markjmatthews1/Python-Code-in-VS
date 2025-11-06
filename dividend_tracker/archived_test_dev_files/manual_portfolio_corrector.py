#!/usr/bin/env python3
"""
Manual Portfolio Values Corrector
================================

Since the API connections need debugging, let's manually input
the correct current portfolio values and update the sheet properly.
"""

import os
import openpyxl
from datetime import datetime

# Configuration
OUTPUT_DIR = os.path.join(os.path.dirname(__file__), "outputs")
TARGET_FILE = os.path.join(OUTPUT_DIR, "Dividends_2025.xlsx")

def get_manual_portfolio_values():
    """Get the REAL current portfolio values manually"""
    
    print("💼 MANUAL PORTFOLIO VALUE INPUT")
    print("=" * 40)
    print("Please enter the ACTUAL current values from your accounts:")
    
    try:
        # E*TRADE IRA
        etrade_ira_str = input("E*TRADE IRA current value: $")
        etrade_ira = float(etrade_ira_str.replace(',', '').replace('$', ''))
        
        # E*TRADE Taxable
        etrade_taxable_str = input("E*TRADE Taxable current value: $")
        etrade_taxable = float(etrade_taxable_str.replace(',', '').replace('$', ''))
        
        # Schwab IRA
        schwab_ira_str = input("Schwab IRA current value: $")
        schwab_ira = float(schwab_ira_str.replace(',', '').replace('$', ''))
        
        # Schwab Individual  
        schwab_ind_str = input("Schwab Individual current value: $")
        schwab_ind = float(schwab_ind_str.replace(',', '').replace('$', ''))
        
        # 401K
        k401_str = input("401K current value: $")
        k401 = float(k401_str.replace(',', '').replace('$', ''))
        
        values = {
            'E*TRADE IRA': etrade_ira,
            'E*TRADE Taxable': etrade_taxable,
            'Schwab IRA': schwab_ira,
            'Schwab Individual': schwab_ind,
            '401K': k401
        }
        
        print(f"\n📊 VALUES ENTERED:")
        total = 0
        for account, value in values.items():
            print(f"   {account}: ${value:,.2f}")
            total += value
        print(f"   TOTAL: ${total:,.2f}")
        
        confirm = input(f"\nAre these values correct? (y/n): ")
        if confirm.lower() == 'y':
            return values
        else:
            print("Please run the script again to re-enter values.")
            return None
            
    except Exception as e:
        print(f"❌ Error getting manual values: {e}")
        return None

def update_with_manual_values():
    """Update column AL with manually entered REAL values"""
    try:
        print(f"\n📊 UPDATING WITH MANUAL REAL VALUES")
        print("=" * 45)
        
        # Get manual values
        values = get_manual_portfolio_values()
        if not values:
            return False
        
        # Create backup
        timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
        backup_name = f"Dividends_2025_manual_real_values_{timestamp}.xlsx"
        backup_path = os.path.join(OUTPUT_DIR, backup_name)
        
        import shutil
        shutil.copy2(TARGET_FILE, backup_path)
        print(f"\n📋 Backup created: {backup_name}")
        
        # Load workbook
        wb = openpyxl.load_workbook(TARGET_FILE)
        ws = wb["Portfolio Values 2025"]
        
        # Target column AL = 38
        new_col = 38
        current_date = datetime.now().strftime('%m/%d/%Y')
        
        print(f"\n📝 UPDATING SHEET:")
        print(f"   Column: AL ({new_col})")
        print(f"   Date: {current_date}")
        
        # AL3: Date header
        ws.cell(row=3, column=new_col, value=current_date)
        print(f"   ✅ AL3: {current_date}")
        
        # AL4: E*TRADE IRA (REAL MANUAL VALUE)
        ws.cell(row=4, column=new_col, value=values['E*TRADE IRA'])
        print(f"   ✅ AL4: E*TRADE IRA = ${values['E*TRADE IRA']:,.2f}")
        
        # AL5: E*TRADE Taxable (REAL MANUAL VALUE)
        ws.cell(row=5, column=new_col, value=values['E*TRADE Taxable'])
        print(f"   ✅ AL5: E*TRADE Taxable = ${values['E*TRADE Taxable']:,.2f}")
        
        # AL6: Schwab IRA (REAL MANUAL VALUE)
        ws.cell(row=6, column=new_col, value=values['Schwab IRA'])
        print(f"   ✅ AL6: Schwab IRA = ${values['Schwab IRA']:,.2f}")
        
        # AL7: Schwab Individual (REAL MANUAL VALUE)
        ws.cell(row=7, column=new_col, value=values['Schwab Individual'])
        print(f"   ✅ AL7: Schwab Individual = ${values['Schwab Individual']:,.2f}")
        
        # AL8: 401K (REAL MANUAL VALUE)
        ws.cell(row=8, column=new_col, value=values['401K'])
        print(f"   ✅ AL8: 401k = ${values['401K']:,.2f}")
        
        # AL10: Total (CALCULATED FROM REAL MANUAL VALUES)
        total_value = sum(values.values())
        ws.cell(row=10, column=new_col, value=total_value)
        print(f"   ✅ AL10: TOTAL = ${total_value:,.2f}")
        
        # Apply proper formatting
        from openpyxl.styles import Font, Border, Side, Alignment
        from openpyxl.styles.numbers import FORMAT_CURRENCY_USD_SIMPLE
        
        # Format cells
        header_font = Font(bold=True, name="Arial", size=12)
        normal_font = Font(name="Arial", size=12)
        border = Border(
            left=Side(style='thin'), right=Side(style='thin'),
            top=Side(style='thin'), bottom=Side(style='thin')
        )
        
        # AL3: Date header
        ws.cell(row=3, column=new_col).font = header_font
        ws.cell(row=3, column=new_col).border = border
        ws.cell(row=3, column=new_col).alignment = Alignment(horizontal='center')
        
        # AL4-AL8: Account values
        for row in range(4, 9):
            cell = ws.cell(row=row, column=new_col)
            cell.font = normal_font
            cell.border = border
            cell.number_format = FORMAT_CURRENCY_USD_SIMPLE
            cell.alignment = Alignment(horizontal='right')
        
        # AL10: Total (bold)
        total_cell = ws.cell(row=10, column=new_col)
        total_cell.font = Font(bold=True, name="Arial", size=12)
        total_cell.border = border
        total_cell.number_format = FORMAT_CURRENCY_USD_SIMPLE
        total_cell.alignment = Alignment(horizontal='right')
        
        # Save workbook
        wb.save(TARGET_FILE)
        print(f"\n✅ WORKBOOK UPDATED WITH REAL VALUES!")
        print(f"   📋 Backup: {backup_name}")
        
        return True
        
    except Exception as e:
        print(f"❌ Error updating with manual values: {e}")
        import traceback
        traceback.print_exc()
        return False

if __name__ == "__main__":
    print("🎯 MANUAL PORTFOLIO VALUES CORRECTOR")
    print("=" * 50)
    print("This will update column AL with the REAL current values")
    print("from your actual E*TRADE and Schwab accounts.")
    print("=" * 50)
    
    success = update_with_manual_values()
    
    if success:
        print(f"\n🎉 SUCCESS!")
        print(f"   ✅ Portfolio Values 2025 now has REAL current values")
        print(f"   ✅ Column AL updated with authentic data")
        print(f"   ✅ Proper formatting applied")
        print(f"\n📊 Portfolio Values Module: COMPLETE with REAL DATA!")
    else:
        print(f"\n❌ Update failed - please check the errors above")
    
    input("\nPress Enter to close...")
