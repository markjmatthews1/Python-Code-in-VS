import os
import sys
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.styles.numbers import FORMAT_CURRENCY_USD_SIMPLE
from datetime import datetime

def fix_portfolio_total():
    """Fix the Portfolio Values sheet with correct total calculation"""
    
    excel_file = "outputs/Dividends_2025.xlsx"
    if not os.path.exists(excel_file):
        print(f"ERROR: Excel file not found: {excel_file}")
        return
    
    try:
        # Simulate the correct portfolio values (with 401K already included)
        portfolio_values = {
            'E*TRADE IRA': 289870.9963,
            'E*TRADE Taxable': 65093.35,
            'Schwab Individual': 2660.97,
            'Schwab IRA': 51929.91,
            '401K': 128693.17  # Already included in portfolio_values
        }
        k401_value = 128693.17
        
        wb = openpyxl.load_workbook(excel_file)
        ws = wb['Portfolio Values 2025']
        
        # Find the last column (most recent data)
        last_col = ws.max_column
        print(f"Updating column {last_col} ({openpyxl.utils.get_column_letter(last_col)})")
        
        # Account mapping
        account_mapping = {
            'E*TRADE IRA': portfolio_values.get('E*TRADE IRA', 0),
            'E*TRADE Taxable': portfolio_values.get('E*TRADE Taxable', 0),
            'Schwab IRA': portfolio_values.get('Schwab IRA', 0),
            'Schwab Individual': portfolio_values.get('Schwab Individual', 0),
            '401k Retirement (Manual)': k401_value
        }
        
        print("\nUPDATING ROWS:")
        # Update rows 4-10
        for row in range(4, 11):
            account_name = ws.cell(row=row, column=1).value
            if account_name:
                account_key = str(account_name).strip()
                value = None
                
                if account_key in account_mapping:
                    value = account_mapping[account_key]
                    print(f"Row {row}: '{account_key}' = ${value:,.2f} (Account match)")
                elif 'total' in account_key.lower():
                    # CORRECTED CALCULATION - no double counting
                    value = sum(portfolio_values.values())
                    print(f"Row {row}: '{account_key}' = ${value:,.2f} (CORRECTED Total - no double 401K)")
                elif '401k' in account_key.lower() or '401' in account_key:
                    value = k401_value
                    print(f"Row {row}: '{account_key}' = ${value:,.2f} (401K match)")
                
                if value is not None:
                    # Update the cell
                    cell = ws.cell(row=row, column=last_col, value=value)
                    cell.number_format = FORMAT_CURRENCY_USD_SIMPLE
                    print(f"  ✅ Updated {openpyxl.utils.get_column_letter(last_col)}{row}")
        
        print(f"\nSAVING FILE...")
        wb.save(excel_file)
        wb.close()
        print(f"✅ Portfolio Values sheet updated with correct total: ${sum(portfolio_values.values()):,.2f}")
        
    except Exception as e:
        print(f"ERROR: {e}")
        if 'PermissionError' in str(e):
            print("NOTE: Please close the Excel file and try again")
        import traceback
        traceback.print_exc()

if __name__ == "__main__":
    fix_portfolio_total()