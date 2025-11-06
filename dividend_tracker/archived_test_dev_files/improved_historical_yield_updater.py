#!/usr/bin/env python3

import openpyxl
from openpyxl.styles import Font, PatternFill
import json
import os
from datetime import datetime

class ImprovedHistoricalYieldUpdater:
    def __init__(self):
        self.excel_path = r"C:\Users\mjmat\Python Code in VS\dividend_tracker\DividendTrackerApp\outputs\Dividends_2025.xlsx"
        self.output_path = r"C:\Users\mjmat\Python Code in VS\dividend_tracker\DividendTrackerApp\outputs\Dividends_2025_improved.xlsx"
        self.cache_file = "portfolio_data_cache.json"
        
        # Define the account mapping from cache to Excel groups
        self.account_mapping = {
            "E*TRADE IRA": "etrade_ira",
            "E*TRADE Taxable": "etrade_taxable", 
            "Schwab IRA": "schwab_ira",
            "Schwab Individual": "schwab_individual"
        }
        
        # Minimum yield threshold - only tickers above 4% yield
        self.min_yield_threshold = 4.0
    
    def run_update(self):
        """Main update process with improved filtering and positioning"""
        print("IMPROVED HISTORICAL YIELD UPDATER")
        print("=" * 50)
        print(f"Update Date: {datetime.now().strftime('%m/%d/%Y')}")
        print(f"Excel File: {os.path.basename(self.excel_path)}")
        print(f"Cache File: {self.cache_file}")
        print(f"Minimum Yield Threshold: {self.min_yield_threshold}%")
        print()
        
        # Step 1: Load cache data
        print("STEP 1: Loading cache data...")
        cache_data = self.load_cache_data()
        if not cache_data:
            return False
        
        positions_data = cache_data.get("positions", {})
        yields_data = cache_data.get("ticker_yields", {})
        print(f"SUCCESS: Cache loaded: {cache_data.get('timestamp')}")
        print(f"INFO: Found yields for {len(yields_data)} tickers")
        print(f"INFO: Found positions for {len(positions_data)} account groups")
        print()
        
        # Step 2: Load Excel workbook
        print("STEP 2: Loading Excel workbook...")
        if not os.path.exists(self.excel_path):
            print(f"ERROR: Excel file not found: {self.excel_path}")
            return False
        
        wb = openpyxl.load_workbook(self.excel_path)
        if 'Accounts Div historical yield' not in wb.sheetnames:
            print("ERROR: 'Accounts Div historical yield' sheet not found")
            return False
        
        ws = wb['Accounts Div historical yield']
        print("SUCCESS: Found sheet: Accounts Div historical yield")
        print()
        
        # Step 3: Find account groups and boundaries  
        print("STEP 3: Finding account groups and boundaries...")
        account_info = self.find_account_groups_and_boundaries(ws)
        
        if len(account_info) != 4:
            print(f"ERROR: Expected 4 account groups, found {len(account_info)}")
            print("Available groups:", list(account_info.keys()))
            return False
        print()
        
        # Step 4: Insert new yield column
        print("STEP 4: Inserting new yield column...")
        date_str = datetime.now().strftime('%m/%d/%Y')
        self.insert_yield_column(ws, date_str)
        print(f"SUCCESS: Inserted yield column P with date {date_str}")
        print()
        
        # Step 5: Update each account group with proper filtering
        print("STEP 5: Updating position data with dividend filtering...")
        
        for group_name, group_info in account_info.items():
            print(f"INFO: Processing {group_name}...")
            cache_key = self.account_mapping.get(group_name)
            if not cache_key:
                print(f"WARNING: No cache mapping for {group_name}")
                continue
            
            account_positions = positions_data.get(cache_key, [])
            
            # Filter positions to only high-yield dividend stocks
            filtered_positions = self.filter_high_yield_positions(account_positions, yields_data)
            
            print(f"INFO: Found {len(account_positions)} total positions")
            print(f"INFO: Filtered to {len(filtered_positions)} high-yield (>{self.min_yield_threshold}%) dividend stocks")
            
            if filtered_positions:
                # Clear existing ticker data (but preserve headers)
                self.clear_existing_tickers(ws, group_info)
                
                # Insert new position data with proper positioning
                self.insert_position_data(ws, group_info, filtered_positions, yields_data)
                
                print(f"SUCCESS: Updated {len(filtered_positions)} high-yield positions for {group_name}")
            else:
                print(f"INFO: No high-yield positions found for {group_name}")
            print()
        
        # Step 6: Update yield data
        print("STEP 6: Updating yield percentages...")
        self.update_yield_data(ws, account_info, yields_data)
        print()
        
        # Step 7: Apply formatting
        print("STEP 7: Applying formatting...")
        self.apply_group_formatting(ws, account_info)
        print()
        
        # Step 8: Save workbook
        print("STEP 8: Saving workbook...")
        os.makedirs(os.path.dirname(self.output_path), exist_ok=True)
        wb.save(self.output_path)
        print(f"SUCCESS: Successfully saved: {self.output_path}")
        print()
        
        print("SUCCESS: IMPROVED UPDATE COMPLETED!")
        print("Updated: Only high-yield dividend stocks (>4%) in correct account groups")
        print("Applied: Proper row insertion accounting for calculation rows")
        wb.close()
        return True
    
    def filter_high_yield_positions(self, positions, yields_data):
        """Filter positions to only include high-yield dividend stocks"""
        filtered = []
        
        for position in positions:
            symbol = position.get('symbol', '').strip().upper()
            yield_info = yields_data.get(symbol, {})
            
            # Check if ticker has dividend data and meets yield threshold
            ticker_yield = yield_info.get('yield', 0.0)
            has_dividend = yield_info.get('has_dividend', False)
            
            if has_dividend and ticker_yield > self.min_yield_threshold:
                filtered.append(position)
                print(f"INCLUDE: {symbol} - Yield: {ticker_yield:.2f}%")
            else:
                reason = "No dividend" if not has_dividend else f"Low yield: {ticker_yield:.2f}%"
                print(f"EXCLUDE: {symbol} - {reason}")
        
        return filtered
    
    def find_account_groups_and_boundaries(self, ws):
        """Find account groups with improved boundary detection"""
        account_info = {}
        
        print("INFO: Scanning for account groups and boundaries...")
        
        for row in range(1, min(80, ws.max_row + 1)):
            cell_value = ws.cell(row=row, column=1).value
            if not cell_value:
                continue
                
            cell_text = str(cell_value).strip().upper()
            
            # Look for account group headers
            if cell_text == 'ETRADE IRA':
                account_info["E*TRADE IRA"] = {"start_row": row, "end_row": None}
                print(f"SUCCESS: E*TRADE IRA found at row {row}")
            elif cell_text == 'ETRADE TAXABLE':
                account_info["E*TRADE Taxable"] = {"start_row": row, "end_row": None}
                print(f"SUCCESS: E*TRADE Taxable found at row {row}")
            elif cell_text == 'SCHWAB IRA':
                account_info["Schwab IRA"] = {"start_row": row, "end_row": None}
                print(f"SUCCESS: Schwab IRA found at row {row}")
            elif cell_text == 'SCHWAB INDIVIDUAL':
                account_info["Schwab Individual"] = {"start_row": row, "end_row": None}
                print(f"SUCCESS: Schwab Individual found at row {row}")
        
        # Calculate end boundaries accounting for calculation rows
        group_order = ["E*TRADE IRA", "E*TRADE Taxable", "Schwab IRA", "Schwab Individual"]
        
        for i, group_name in enumerate(group_order):
            if group_name not in account_info:
                continue
                
            start_row = account_info[group_name]["start_row"]
            
            if i < len(group_order) - 1:
                # For non-final groups: end is 3 rows before next group
                # (accounts for: calculation row + blank row + next group header)
                end_row = ws.max_row
                for next_group in group_order[i+1:]:
                    if next_group in account_info:
                        end_row = account_info[next_group]["start_row"] - 3
                        break
            else:
                # Last group - find actual end of data
                end_row = start_row + 2
                while end_row <= ws.max_row:
                    cell_value = ws.cell(row=end_row, column=1).value
                    if not cell_value or str(cell_value).strip() == "":
                        break
                    if any(keyword in str(cell_value).strip().upper() for keyword in ['ETRADE', 'SCHWAB']):
                        break
                    end_row += 1
                end_row -= 1
                
            account_info[group_name]["end_row"] = end_row
            account_info[group_name]["insertion_row"] = end_row + 1  # Where to insert new positions
            print(f"INFO: {group_name}: rows {start_row}-{end_row}, insert at row {end_row + 1}")
        
        return account_info
    
    def clear_existing_tickers(self, ws, group_info):
        """Clear existing ticker data but preserve headers and calculation rows"""
        start_row = group_info["start_row"] + 2  # Skip group header and column headers
        end_row = group_info["end_row"]
        
        for row in range(start_row, end_row + 1):
            # Only clear if there's a ticker in column A
            ticker_cell = ws.cell(row=row, column=1)
            if ticker_cell.value and str(ticker_cell.value).strip():
                # Clear columns A, B, D only
                ws.cell(row=row, column=1).value = None  # Ticker
                ws.cell(row=row, column=2).value = None  # Quantity
                ws.cell(row=row, column=4).value = None  # Price
                
                # Clear formatting
                ws.cell(row=row, column=1).font = Font()
                ws.cell(row=row, column=2).font = Font()
                ws.cell(row=row, column=4).font = Font()
    
    def insert_position_data(self, ws, group_info, positions, yields_data):
        """Insert position data with proper formatting"""
        insert_row = group_info["start_row"] + 2  # Start after headers
        
        # Define formatting
        ticker_font = Font(name="Arial", size=12, bold=True, color="3072C2")
        qty_font = Font(name="Arial", size=12, bold=True, color="3072C2")
        
        for i, position in enumerate(positions):
            current_row = insert_row + i
            symbol = position.get('symbol', '').strip().upper()
            quantity = position.get('quantity', 0)
            
            # Calculate price from market_value / quantity
            market_value = position.get('market_value', 0)
            price = round(market_value / quantity, 2) if quantity > 0 else 0
            
            # Insert data
            ticker_cell = ws.cell(row=current_row, column=1)
            qty_cell = ws.cell(row=current_row, column=2)
            price_cell = ws.cell(row=current_row, column=4)
            
            ticker_cell.value = symbol
            qty_cell.value = quantity
            price_cell.value = price
            
            # Apply formatting
            ticker_cell.font = ticker_font
            qty_cell.font = qty_font
            
            yield_info = yields_data.get(symbol, {})
            ticker_yield = yield_info.get('yield', 0.0)
            print(f"INSERT: Row {current_row}: {symbol} | Qty: {quantity} | Price: ${price} | Yield: {ticker_yield:.2f}%")
    
    def insert_yield_column(self, ws, date_str):
        """Insert new yield column after column O"""
        # Insert column P (after O)
        ws.insert_cols(16)  # Column P = 16
        
        # Add date header to each group
        for row in range(1, min(80, ws.max_row + 1)):
            cell_value = ws.cell(row=row, column=1).value
            if cell_value and any(keyword in str(cell_value).upper() for keyword in ['ETRADE', 'SCHWAB']):
                header_row = row + 1
                header_cell = ws.cell(row=header_row, column=16)
                header_cell.value = date_str
                header_cell.font = Font(bold=True)
    
    def update_yield_data(self, ws, account_info, yields_data):
        """Update yield percentages in the new column"""
        for group_name, group_info in account_info.items():
            print(f"INFO: Updating yields for {group_name}...")
            
            start_row = group_info["start_row"] + 2
            end_row = group_info["end_row"]
            yield_count = 0
            
            for row in range(start_row, end_row + 1):
                ticker_cell = ws.cell(row=row, column=1)
                if ticker_cell.value:
                    symbol = str(ticker_cell.value).strip().upper()
                    yield_info = yields_data.get(symbol, {})
                    ticker_yield = yield_info.get('yield', 0.0)
                    
                    yield_cell = ws.cell(row=row, column=16)  # Column P
                    yield_cell.value = f"{ticker_yield:.2f}%"
                    
                    # Color code yields
                    if ticker_yield >= 15.0:
                        yield_cell.font = Font(color="00AA00")  # Green
                    elif ticker_yield >= 10.0:
                        yield_cell.font = Font(color="0000AA")  # Blue
                    else:
                        yield_cell.font = Font(color="000000")  # Black
                    
                    yield_count += 1
                    print(f"YIELD: Row {row}: {symbol} = {ticker_yield:.2f}%")
            
            print(f"SUCCESS: Updated {yield_count} yields for {group_name}")
    
    def apply_group_formatting(self, ws, account_info):
        """Apply orange formatting to group divider headers"""
        orange_fill = PatternFill(start_color="FFA500", end_color="FFA500", fill_type="solid")
        
        print("INFO: Applying orange color to group dividers...")
        for group_name, group_info in account_info.items():
            divider_row = group_info["start_row"]
            divider_cell = ws.cell(row=divider_row, column=1)
            divider_cell.fill = orange_fill
            print(f"DIVIDER: {group_name} divider at row {divider_row}")
    
    def load_cache_data(self):
        """Load position and yield data from cache"""
        try:
            if not os.path.exists(self.cache_file):
                print(f"ERROR: Cache file not found: {self.cache_file}")
                return None
            
            with open(self.cache_file, 'r') as f:
                return json.load(f)
        except Exception as e:
            print(f"ERROR: Failed to load cache: {e}")
            return None

if __name__ == "__main__":
    updater = ImprovedHistoricalYieldUpdater()
    success = updater.run_update()
    
    if success:
        print("\nSUCCESS: Improved historical yield update completed!")
    else:
        print("\nERROR: Update failed!")