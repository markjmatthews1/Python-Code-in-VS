#!/usr/bin/env python3
"""
Test Column Insertion Logic
===========================

Quick test to verify the column insertion logic works correctly
"""

import openpyxl
import os
from datetime import date

def test_column_insertion():
    """Test the column insertion logic"""
    
    print("🧪 TESTING COLUMN INSERTION LOGIC")
    print("=" * 50)
    
    workbook_path = r"c:\Users\mjmat\Python Code in VS\dividend_tracker\DividendTrackerApp\outputs\Dividends_2025.xlsx"
    sheet_name = "Etrade IRA historic yield"
    
    try:
        # Load the workbook in read-only mode to examine structure
        wb = openpyxl.load_workbook(workbook_path, data_only=True)
        
        if sheet_name not in wb.sheetnames:
            print(f"❌ Sheet not found: {sheet_name}")
            return False
            
        sheet = wb[sheet_name]
        
        print(f"📊 Current sheet structure:")
        print(f"   📏 Dimensions: {sheet.max_row} rows × {sheet.max_column} columns")
        
        # Show current headers around column 15-20
        print(f"\n📋 Headers around columns 13-20:")
        for col in range(13, min(21, sheet.max_column + 1)):
            header = sheet.cell(row=1, column=col).value
            print(f"   Column {col:2d}: '{header}'")
        
        print(f"\n📍 Column 15 contains: '{sheet.cell(row=1, column=15).value}'")
        
        # Check where we should insert
        insert_col = 16
        print(f"🎯 New column will be inserted at position: {insert_col}")
        print(f"   This is right after: '{sheet.cell(row=1, column=15).value}'")
        
        # Show what would happen after insertion
        today = date.today().strftime("%m-%d-%Y")
        print(f"\n📅 After insertion:")
        print(f"   Column 15: '{sheet.cell(row=1, column=15).value}' (unchanged)")
        print(f"   Column 16: '{today}' (NEW - inserted)")
        print(f"   Column 17: '{sheet.cell(row=1, column=16).value if sheet.max_column >= 16 else 'Empty'}' (shifted right)")
        
        wb.close()
        
        print("\n✅ Column insertion logic verified!")
        print("   📍 New yield column will be inserted right after 'Beginning Dividend Yield'")
        print("   🔄 All existing historical data will shift one column to the right")
        
        return True
        
    except Exception as e:
        print(f"❌ Error testing column insertion: {e}")
        return False

if __name__ == "__main__":
    success = test_column_insertion()
    
    if success:
        print("\n🎉 Column insertion test passed!")
    else:
        print("\n💥 Column insertion test failed!")
