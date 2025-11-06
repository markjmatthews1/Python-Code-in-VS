import openpyxl

def debug_comprehensive_dividend_data():
    """Comprehensive debug of the dividend data to find the real issue"""
    
    workbook_path = "outputs/Dividends_2025.xlsx"
    wb = openpyxl.load_workbook(workbook_path)
    div_sheet = wb["Accounts Div historical yield"]
    
    print("COMPREHENSIVE DIVIDEND DATA DEBUG")
    print("=" * 60)
    
    # Check what columns actually exist
    print(f"Sheet max columns: {div_sheet.max_column}")
    print(f"Sheet max rows: {div_sheet.max_row}")
    
    # Check column headers
    print("\nColumn headers:")
    for col in range(14, div_sheet.max_column + 1):
        header = div_sheet.cell(row=2, column=col).value
        print(f"Column {col}: '{header}'")
    
    print("\nSample data from multiple accounts:")
    
    # Test rows from each account
    test_rows = [3, 4, 5, 25, 26, 27, 41, 42, 49, 50]
    
    for row in test_rows:
        ticker = div_sheet.cell(row=row, column=1).value
        if ticker:
            beginning = div_sheet.cell(row=row, column=15).value  # Column O
            current = div_sheet.cell(row=row, column=16).value    # Column P
            
            print(f"Row {row:2d} - {ticker:6}: Beginning={beginning} ({type(beginning).__name__}) | Current={current} ({type(current).__name__})")
            
            # Test conversion logic
            if beginning and current:
                try:
                    current_val = float(current)
                    beginning_val = float(beginning)
                    
                    # Apply conversion
                    if current_val < 1.0 and current_val > 0:
                        current_converted = current_val * 100
                    else:
                        current_converted = current_val
                    
                    # Check if it's a reduction (10% threshold)
                    threshold = beginning_val * 0.90
                    is_reduction = current_converted < threshold
                    
                    print(f"         Converted: {beginning_val:.2f}% → {current_converted:.2f}% | Threshold: {threshold:.2f}% | Reduction: {is_reduction}")
                    
                    if is_reduction:
                        reduction_pct = ((beginning_val - current_converted) / beginning_val) * 100
                        print(f"         Reduction amount: {reduction_pct:.1f}%")
                    
                except (ValueError, TypeError) as e:
                    print(f"         Conversion error: {e}")
    
    # Check if there are any null/empty values causing issues
    print("\nChecking for data quality issues:")
    empty_current = 0
    empty_beginning = 0
    total_tickers = 0
    
    for row in range(3, div_sheet.max_row + 1):
        ticker = div_sheet.cell(row=row, column=1).value
        if ticker and str(ticker).strip():
            total_tickers += 1
            beginning = div_sheet.cell(row=row, column=15).value
            current = div_sheet.cell(row=row, column=16).value
            
            if not beginning:
                empty_beginning += 1
            if not current:
                empty_current += 1
    
    print(f"Total tickers found: {total_tickers}")
    print(f"Empty beginning yields: {empty_beginning}")
    print(f"Empty current yields: {empty_current}")
    
    wb.close()

if __name__ == "__main__":
    debug_comprehensive_dividend_data()
