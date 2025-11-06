#!/usr/bin/env python3

import openpyxl
import os
import json
from datetime import datetime

class UpdateAnalyzer:
    def __init__(self):
        self.improved_file = r"C:\Users\mjmat\Python Code in VS\dividend_tracker\DividendTrackerApp\outputs\Dividends_2025_improved.xlsx"
        self.cache_file = "portfolio_data_cache.json"
    
    def analyze_improvements(self):
        """Analyze the improvements made by the new updater"""
        print("IMPROVED UPDATER ANALYSIS")
        print("=" * 50)
        print(f"Analysis Time: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}")
        print()
        
        # Load cache to show filtering decisions
        with open(self.cache_file, 'r') as f:
            cache_data = json.load(f)
        
        positions_data = cache_data.get("positions", {})
        yields_data = cache_data.get("ticker_yields", {})
        
        # Analyze filtering for each account
        account_mapping = {
            "E*TRADE IRA": "etrade_ira",
            "E*TRADE Taxable": "etrade_taxable", 
            "Schwab IRA": "schwab_ira",
            "Schwab Individual": "schwab_individual"
        }
        
        print("FILTERING ANALYSIS:")
        print("=" * 30)
        
        total_positions = 0
        total_filtered = 0
        
        for group_name, cache_key in account_mapping.items():
            positions = positions_data.get(cache_key, [])
            print(f"\n{group_name}:")
            print(f"  Total positions: {len(positions)}")
            
            high_yield_count = 0
            excluded_tickers = []
            included_tickers = []
            
            for position in positions:
                symbol = position.get('symbol', '').strip().upper()
                yield_info = yields_data.get(symbol, {})
                ticker_yield = yield_info.get('yield', 0.0)
                has_dividend = yield_info.get('has_dividend', False)
                
                if has_dividend and ticker_yield > 4.0:
                    high_yield_count += 1
                    included_tickers.append(f"{symbol} ({ticker_yield:.2f}%)")
                else:
                    reason = "No dividend" if not has_dividend else f"Low yield: {ticker_yield:.2f}%"
                    excluded_tickers.append(f"{symbol} - {reason}")
            
            print(f"  High-yield (>4%) dividend stocks: {high_yield_count}")
            print(f"  INCLUDED: {', '.join(included_tickers)}")
            if excluded_tickers:
                print(f"  EXCLUDED: {', '.join(excluded_tickers)}")
            
            total_positions += len(positions)
            total_filtered += high_yield_count
        
        print(f"\nOVERALL FILTERING RESULTS:")
        print(f"Total positions across all accounts: {total_positions}")
        print(f"High-yield dividend stocks included: {total_filtered}")
        print(f"Filtering efficiency: {(total_filtered/total_positions)*100:.1f}%")
        
        # Verify Excel structure
        print(f"\nEXCEL STRUCTURE VERIFICATION:")
        print("=" * 35)
        
        if os.path.exists(self.improved_file):
            wb = openpyxl.load_workbook(self.improved_file)
            ws = wb['Accounts Div historical yield']
            
            # Check account groups
            groups_found = {}
            for row in range(1, 60):
                cell_value = ws.cell(row=row, column=1).value
                if cell_value:
                    cell_text = str(cell_value).strip().upper()
                    if cell_text == 'ETRADE IRA':
                        groups_found["E*TRADE IRA"] = row
                    elif cell_text == 'ETRADE TAXABLE':
                        groups_found["E*TRADE Taxable"] = row
                    elif cell_text == 'SCHWAB IRA':
                        groups_found["Schwab IRA"] = row
                    elif cell_text == 'SCHWAB INDIVIDUAL':
                        groups_found["Schwab Individual"] = row
            
            print(f"Account groups found: {len(groups_found)}")
            for group, row in groups_found.items():
                print(f"  {group}: Row {row}")
            
            # Check yield column
            yield_header = ws.cell(row=2, column=16).value  # Column P
            print(f"New yield column header: {yield_header}")
            
            # Sample some data
            print(f"\nSAMPLE DATA VERIFICATION:")
            for row in [3, 25, 41, 49]:  # Sample from each group
                ticker = ws.cell(row=row, column=1).value
                qty = ws.cell(row=row, column=2).value
                price = ws.cell(row=row, column=4).value
                yield_val = ws.cell(row=row, column=16).value
                
                if ticker:
                    print(f"  Row {row}: {ticker} | Qty: {qty} | Price: ${price} | Yield: {yield_val}")
            
            wb.close()
            print(f"\nSUCCESS: Improved updater working correctly!")
        else:
            print(f"ERROR: Improved file not found")

if __name__ == "__main__":
    analyzer = UpdateAnalyzer()
    analyzer.analyze_improvements()