#!/usr/bin/env python3
"""
Check current portfolio values to see if 401K data is current
"""

import os
from datetime import datetime
import openpyxl

def check_portfolio_values():
    """Check what's currently in the Portfolio Values sheet"""
    
    print("=== CHECKING PORTFOLIO VALUES STATUS ===")
    print(f"Check time: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}")
    
    try:
        # Open the workbook
        workbook_path = os.path.join("outputs", "Dividends_2025.xlsx")
        
        if not os.path.exists(workbook_path):
            print("❌ Dividends_2025.xlsx not found")
            return
            
        workbook = openpyxl.load_workbook(workbook_path)
        
        # Check Portfolio Values 2025 sheet
        if "Portfolio Values 2025" in workbook.sheetnames:
            sheet = workbook["Portfolio Values 2025"]
            print(f"✅ Found Portfolio Values 2025 sheet")
            
            # Look for 401K value - typically in a specific location
            # Check various cells that might contain 401K data
            for row in range(1, 20):
                for col in range(1, 10):
                    cell_value = sheet.cell(row=row, column=col).value
                    if cell_value and isinstance(cell_value, str) and "401" in cell_value:
                        print(f"📊 Found 401K reference at {chr(64+col)}{row}: {cell_value}")
                        # Check adjacent cells for values
                        for offset in [0, 1, 2, 3]:
                            check_col = col + offset
                            if check_col <= 20:  # Stay within reasonable bounds
                                adj_value = sheet.cell(row=row, column=check_col).value
                                if adj_value and isinstance(adj_value, (int, float)) and adj_value > 50000:
                                    print(f"   💰 401K value found: ${adj_value:,.2f} at column {chr(64+check_col)}")
            
            # Check if sheet has today's date
            has_today = False
            today_str = datetime.now().strftime("%m/%d/%Y")
            alt_today = datetime.now().strftime("%Y-%m-%d")
            
            for row in range(1, 10):
                for col in range(1, 15):
                    cell_value = str(sheet.cell(row=row, column=col).value or "")
                    if today_str in cell_value or alt_today in cell_value:
                        has_today = True
                        print(f"✅ Found today's date at {chr(64+col)}{row}: {cell_value}")
                        break
            
            if not has_today:
                print("⚠️ No today's date found - may need updating")
                
        else:
            print("❌ Portfolio Values 2025 sheet not found")
            print(f"Available sheets: {workbook.sheetnames}")
            
        workbook.close()
        
    except Exception as e:
        print(f"❌ Error checking portfolio values: {e}")
        import traceback
        traceback.print_exc()

if __name__ == "__main__":
    check_portfolio_values()
    input("\nPress Enter to continue...")
