import os
import openpyxl
from openpyxl.styles import Font, PatternFill
from openpyxl.styles.numbers import FORMAT_CURRENCY_USD_SIMPLE

def fix_monthly_average_color():
    """Fix the Monthly Average color coding by calculating actual values"""
    
    excel_file = "outputs/Dividends_2025.xlsx"
    
    try:
        # First, open with calculation to get numeric values
        wb = openpyxl.load_workbook(excel_file, data_only=False)
        ws = wb["Estimated Income 2025"]
        
        monthly_row = 9
        last_col = ws.max_column
        
        print(f"FIXING MONTHLY AVERAGE COLOR CODING")
        print(f"Row: {monthly_row}, Column: {openpyxl.utils.get_column_letter(last_col)}")
        
        # Get the formula
        current_cell = ws.cell(row=monthly_row, column=last_col)
        formula = current_cell.value
        print(f"Formula: {formula}")
        
        # Manual calculation of the current value
        # Sum the dividend values in rows 4-7 of the current column, then divide by 12
        current_sum = 0
        for row in range(4, 8):  # Rows 4-7
            cell_value = ws.cell(row=row, column=last_col).value
            if isinstance(cell_value, (int, float)):
                current_sum += cell_value
            else:
                print(f"Row {row}: Non-numeric value {cell_value}")
        
        current_monthly = current_sum / 12 if current_sum > 0 else 0
        
        # Calculate previous value
        if last_col > 1:
            previous_sum = 0
            for row in range(4, 8):  # Rows 4-7
                cell_value = ws.cell(row=row, column=last_col - 1).value
                if isinstance(cell_value, (int, float)):
                    previous_sum += cell_value
            previous_monthly = previous_sum / 12 if previous_sum > 0 else 0
        else:
            previous_monthly = None
            
        print(f"Current monthly average (calculated): ${current_monthly:,.2f}")
        print(f"Previous monthly average (calculated): ${previous_monthly:,.2f}" if previous_monthly else "Previous: None")
        
        # Apply color coding based on comparison
        if previous_monthly is not None:
            if current_monthly > previous_monthly:
                fill_color = PatternFill(start_color='90EE90', end_color='90EE90', fill_type='solid')
                color_desc = "Green (Increase)"
            elif current_monthly < previous_monthly:
                fill_color = PatternFill(start_color='FF7C80', end_color='FF7C80', fill_type='solid')
                color_desc = "Red (Decrease)"
            else:
                fill_color = PatternFill(start_color='FFFF00', end_color='FFFF00', fill_type='solid')
                color_desc = "Yellow (Same)"
                
            # Apply all formatting
            current_cell.value = formula
            current_cell.font = Font(name='Arial', size=12, bold=True)
            current_cell.number_format = FORMAT_CURRENCY_USD_SIMPLE
            current_cell.fill = fill_color
            
            print(f"✅ Applied: Bold + {color_desc}")
        else:
            # Just apply bold if no previous value
            current_cell.font = Font(name='Arial', size=12, bold=True)
            current_cell.number_format = FORMAT_CURRENCY_USD_SIMPLE
            print("✅ Applied: Bold only (no previous value for color)")
        
        # Save changes
        wb.save(excel_file)
        wb.close()
        
        print("✅ Monthly Average color coding applied!")
        
    except Exception as e:
        print(f"ERROR: {e}")
        if 'Permission' in str(e):
            print("Please close Excel and try again")
        import traceback
        traceback.print_exc()

if __name__ == "__main__":
    fix_monthly_average_color()