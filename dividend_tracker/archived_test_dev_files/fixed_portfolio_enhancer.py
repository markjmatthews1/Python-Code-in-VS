"""
Fixed Portfolio Summary Dividend Status Enhancement
==================================================
"""

import pandas as pd
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
import os
from datetime import datetime

class PortfolioSummaryDividendEnhancer:
    """Enhances Portfolio Summary with dividend analysis from historical yield data"""
    
    def __init__(self):
        self.workbook_path = os.path.join(os.path.dirname(__file__), "outputs", "Dividends_2025.xlsx")
        self.portfolio_sheet_name = "Portfolio Summary"
        self.dividend_sheet_name = "Accounts Div historical yield"
        
    def analyze_dividend_changes(self, workbook):
        """Analyze dividend changes from the historical yield sheet"""
        try:
            dividend_sheet = workbook[self.dividend_sheet_name]
            
            dividend_analysis = {
                "reduced": [],
                "top_payers": []
            }
            
            # Account sections row ranges (using current shifted layout)
            account_sections = {
                "Etrade IRA": {"start": 3, "end": 20},
                "Etrade Taxable": {"start": 25, "end": 36}, 
                "Schwab IRA": {"start": 41, "end": 44},
                "Schwab Individual": {"start": 49, "end": 50}
            }
            
            for account, ranges in account_sections.items():
                print(f"Analyzing {account} dividend changes...")
                
                account_reduced = []
                account_yields = []
                
                # Check each ticker in this account
                for row in range(ranges["start"], ranges["end"] + 1):
                    ticker_cell = dividend_sheet.cell(row=row, column=1)
                    if not ticker_cell.value or str(ticker_cell.value).strip() == "":
                        continue
                        
                    ticker = str(ticker_cell.value).strip()
                    
                    # Get beginning yield (column 15) and current yield (column 16)
                    beginning_yield = dividend_sheet.cell(row=row, column=15).value
                    current_yield = dividend_sheet.cell(row=row, column=16).value
                    
                    try:
                        beginning_val = float(beginning_yield) if beginning_yield else 0
                        current_val = float(current_yield) if current_yield else 0
                        
                        if beginning_val <= 0 or current_val <= 0:
                            continue  # Skip if no valid data
                        
                        # Convert current yield from decimal to percentage (0.1015 → 10.15)
                        if current_val < 1.0:
                            current_val *= 100
                        
                        # Beginning yield is already in percentage format
                        
                        # Check for significant dividend reduction (>5% reduction)
                        reduction_threshold = beginning_val * 0.95  # 5% reduction threshold
                        
                        if current_val < reduction_threshold:
                            reduction_pct = ((beginning_val - current_val) / beginning_val) * 100
                            account_reduced.append({
                                "ticker": ticker,
                                "beginning": beginning_val,
                                "current": current_val,
                                "reduction_pct": reduction_pct
                            })
                        
                        # Always add to yields for top payers analysis
                        account_yields.append({
                            "ticker": ticker,
                            "yield": current_val
                        })
                            
                    except (ValueError, TypeError) as e:
                        print(f"  Error processing {ticker}: {e}")
                        continue
                
                # Sort and get top dividend payers for this account
                account_yields.sort(key=lambda x: x["yield"], reverse=True)
                
                dividend_analysis["reduced"].append({
                    "account": account,
                    "reductions": account_reduced
                })
                
                dividend_analysis["top_payers"].append({
                    "account": account,
                    "top_yields": account_yields
                })
                
                print(f"  {account}: {len(account_reduced)} reductions, {len(account_yields)} valid yields")
                if account_yields:
                    print(f"  Top yield: {account_yields[0]['ticker']} at {account_yields[0]['yield']:.2f}%")
                
            return dividend_analysis
            
        except Exception as e:
            print(f"Error analyzing dividend changes: {e}")
            import traceback
            traceback.print_exc()
            return {"reduced": [], "top_payers": []}
    
    def add_dividend_status_columns(self):
        """Add Dividend Status analysis columns starting at G1"""
        try:
            print("Adding Dividend Status columns to Portfolio Summary...")
            
            # Load workbook
            workbook = openpyxl.load_workbook(self.workbook_path)
            
            if self.portfolio_sheet_name not in workbook.sheetnames:
                print(f"Sheet '{self.portfolio_sheet_name}' not found")
                return False
            
            portfolio_sheet = workbook[self.portfolio_sheet_name]
            
            # Unmerge any existing merged cells in columns G-H to avoid conflicts
            merged_ranges_to_remove = []
            for merged_range in portfolio_sheet.merged_cells.ranges:
                if merged_range.min_col <= 8 and merged_range.max_col >= 7:  # Overlaps with G-H
                    merged_ranges_to_remove.append(merged_range)
            
            for merged_range in merged_ranges_to_remove:
                portfolio_sheet.unmerge_cells(str(merged_range))
            
            # Clear existing dividend status columns (G:H) - avoid merged cells
            for row in range(1, 50):  # Extended range to accommodate detailed reduction lists
                try:
                    portfolio_sheet.cell(row=row, column=7).value = None
                    portfolio_sheet.cell(row=row, column=8).value = None
                except AttributeError:
                    # Skip merged cells
                    pass
            
            # Analyze dividend changes
            dividend_analysis = self.analyze_dividend_changes(workbook)
            
            # Header styling - Arial 12
            header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
            header_font = Font(name="Arial", size=12, bold=True, color="FFFFFF")
            section_fill = PatternFill(start_color="B7DEE8", end_color="B7DEE8", fill_type="solid")
            section_font = Font(name="Arial", size=12, bold=True)
            
            # Add main header at G1
            portfolio_sheet.cell(row=1, column=7, value="DIVIDEND STATUS").fill = header_fill
            portfolio_sheet.cell(row=1, column=7).font = header_font
            portfolio_sheet.cell(row=1, column=7).alignment = Alignment(horizontal='center')
            portfolio_sheet.merge_cells('G1:H1')
            
            current_row = 3
            
            # === DIVIDEND REDUCTIONS SECTION ===
            portfolio_sheet.cell(row=current_row, column=7, value="DIVIDEND REDUCTIONS").fill = section_fill
            portfolio_sheet.cell(row=current_row, column=7).font = section_font
            portfolio_sheet.merge_cells(f'G{current_row}:H{current_row}')
            current_row += 1
            
            # Add column headers for reductions
            portfolio_sheet.cell(row=current_row, column=7, value="Account").font = Font(name="Arial", size=12, bold=True)
            portfolio_sheet.cell(row=current_row, column=8, value="Reduced Tickers").font = Font(name="Arial", size=12, bold=True)
            current_row += 1
            
            for account_data in dividend_analysis["reduced"]:
                account = account_data["account"]
                reductions = account_data["reductions"]
                
                if reductions:
                    # Sort by highest reduction percentage
                    reductions.sort(key=lambda x: x["reduction_pct"], reverse=True)
                    
                    # Show account name first
                    portfolio_sheet.cell(row=current_row, column=7, value=account).font = Font(name="Arial", size=12, bold=True)
                    portfolio_sheet.cell(row=current_row, column=8, value=f"{len(reductions)} dividend cuts:")
                    portfolio_sheet.cell(row=current_row, column=8).font = Font(name="Arial", size=12, color="DC143C", bold=True)
                    current_row += 1
                    
                    # Show all reductions for this account
                    for reduction in reductions:
                        ticker_text = f"  {reduction['ticker']}"
                        reduction_text = f"-{reduction['reduction_pct']:.1f}% ({reduction['beginning']:.2f}% → {reduction['current']:.2f}%)"
                        
                        portfolio_sheet.cell(row=current_row, column=7, value=ticker_text)
                        portfolio_sheet.cell(row=current_row, column=7).font = Font(name="Arial", size=12)
                        portfolio_sheet.cell(row=current_row, column=8, value=reduction_text)
                        portfolio_sheet.cell(row=current_row, column=8).font = Font(name="Arial", size=12, color="DC143C")
                        current_row += 1
                    
                    current_row += 1  # Extra space after each account
                else:
                    portfolio_sheet.cell(row=current_row, column=7, value=account).font = Font(name="Arial", size=12, bold=True)
                    portfolio_sheet.cell(row=current_row, column=8, value="No reductions - Good!")
                    portfolio_sheet.cell(row=current_row, column=8).font = Font(name="Arial", size=12, color="228B22", bold=True)
                    current_row += 1
            
            current_row += 1  # Space between sections
            
            # === TOP DIVIDEND PAYERS SECTION ===
            portfolio_sheet.cell(row=current_row, column=7, value="TOP DIVIDEND PAYERS").fill = section_fill
            portfolio_sheet.cell(row=current_row, column=7).font = section_font
            portfolio_sheet.merge_cells(f'G{current_row}:H{current_row}')
            current_row += 1
            
            # Add column headers for top payers
            portfolio_sheet.cell(row=current_row, column=7, value="Account").font = Font(name="Arial", size=12, bold=True)
            portfolio_sheet.cell(row=current_row, column=8, value="Top Yield").font = Font(name="Arial", size=12, bold=True)
            current_row += 1
            
            for account_data in dividend_analysis["top_payers"]:
                account = account_data["account"]
                top_yields = account_data["top_yields"]
                
                if top_yields:
                    # Show top yielder
                    top_yielder = top_yields[0]
                    yield_text = f"{top_yielder['ticker']}: {top_yielder['yield']:.2f}%"
                    
                    if len(top_yields) > 1:
                        yield_text += f" ({len(top_yields)} total)"
                    
                    portfolio_sheet.cell(row=current_row, column=7, value=account)
                    portfolio_sheet.cell(row=current_row, column=7).font = Font(name="Arial", size=12)
                    portfolio_sheet.cell(row=current_row, column=8, value=yield_text)
                    portfolio_sheet.cell(row=current_row, column=8).font = Font(name="Arial", size=12, color="228B22", bold=True)
                else:
                    portfolio_sheet.cell(row=current_row, column=7, value=account)
                    portfolio_sheet.cell(row=current_row, column=7).font = Font(name="Arial", size=12)
                    portfolio_sheet.cell(row=current_row, column=8, value="No yield data")
                    portfolio_sheet.cell(row=current_row, column=8).font = Font(name="Arial", size=12, color="808080")
                
                current_row += 1
            
            current_row += 1  # Space between sections
            
            # === DIVIDEND SUMMARY METRICS ===
            portfolio_sheet.cell(row=current_row, column=7, value="DIVIDEND METRICS").fill = section_fill
            portfolio_sheet.cell(row=current_row, column=7).font = section_font
            portfolio_sheet.merge_cells(f'G{current_row}:H{current_row}')
            current_row += 1
            
            # Calculate overall metrics
            total_reductions = sum(len(acc["reductions"]) for acc in dividend_analysis["reduced"])
            total_positions = sum(len(acc["top_yields"]) for acc in dividend_analysis["top_payers"])
            cut_rate = (total_reductions/total_positions) if total_positions > 0 else 0  # Decimal for Excel percentage formatting
            
            # Add metrics with proper data types and Arial 12 font
            metrics = [
                ("Total Positions:", total_positions),
                ("Dividend Cuts:", total_reductions),
                ("Cut Rate:", cut_rate),
                ("Last Updated:", datetime.now().strftime("%m/%d %H:%M"))
            ]
            
            for metric_name, metric_value in metrics:
                # Metric name with Arial 12 bold
                name_cell = portfolio_sheet.cell(row=current_row, column=7, value=metric_name)
                name_cell.font = Font(name="Arial", size=12, bold=True)
                
                # Metric value with Arial 12
                value_cell = portfolio_sheet.cell(row=current_row, column=8)
                value_cell.font = Font(name="Arial", size=12)
                
                # Set value as number or string appropriately
                if isinstance(metric_value, (int, float)):
                    value_cell.value = metric_value
                    if metric_name == "Cut Rate:":
                        value_cell.number_format = "0.0%"  # Format as percentage
                else:
                    value_cell.value = str(metric_value)
                
                current_row += 1
            
            # Auto-size columns
            portfolio_sheet.column_dimensions['G'].width = 20
            portfolio_sheet.column_dimensions['H'].width = 25
            
            # Save workbook
            workbook.save(self.workbook_path)
            workbook.close()
            
            print("Dividend Status columns added successfully!")
            print(f"   Found {total_reductions} dividend reductions across all accounts")
            print(f"   Analyzed {total_positions} dividend-paying positions")
            print(f"   Overall dividend cut rate: {(total_reductions/total_positions*100):.1f}%" if total_positions > 0 else "   No cut rate calculated")
            
            return True
            
        except Exception as e:
            print(f"Error adding dividend status columns: {e}")
            import traceback
            traceback.print_exc()
            return False

def main():
    """Test the dividend status enhancement"""
    enhancer = PortfolioSummaryDividendEnhancer()
    
    print("Starting Portfolio Summary Dividend Status Enhancement...")
    
    if not os.path.exists(enhancer.workbook_path):
        print(f"Workbook not found: {enhancer.workbook_path}")
        return
    
    success = enhancer.add_dividend_status_columns()
    
    if success:
        print("Portfolio Summary enhancement completed successfully!")
        print("Check columns G-H in the Portfolio Summary sheet for dividend analysis")
    else:
        print("Enhancement failed - check error messages above")

if __name__ == "__main__":
    main()
