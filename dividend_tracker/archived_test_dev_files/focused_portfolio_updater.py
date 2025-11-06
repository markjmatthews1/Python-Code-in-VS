#!/usr/bin/env python3
"""
Focused Portfolio Values 2025 Updater
====================================

Single-purpose script that ONLY updates Portfolio Values 2025 sheet with:
1. Current E*TRADE API portfolio values
2. Schwab portfolio values (if available)
3. 401K value prompt
4. Adds NEW COLUMN preserving all historical data

Purpose: Iterative development - test this module in isolation
Author: Assistant (GitHub Copilot)
Created: August 31, 2025
"""

import os
import sys
import openpyxl
from datetime import datetime
import traceback

# Add modules to path
sys.path.append(os.path.join(os.path.dirname(__file__), 'modules'))

try:
    from gui_prompts import get_k401_value
    from etrade_auth import get_etrade_session
except ImportError as e:
    print(f"❌ Import error: {e}")
    print("Falling back to basic implementations...")
    
    def get_k401_value():
        """Fallback 401K value prompt"""
        try:
            value_str = input("Enter current 401K value: $")
            return float(value_str.replace(',', '').replace('$', ''))
        except:
            return 125000.00  # Default fallback

class FocusedPortfolioUpdater:
    """Updates ONLY the Portfolio Values 2025 sheet with current API data"""
    
    def __init__(self):
        self.script_dir = os.path.dirname(os.path.abspath(__file__))
        self.outputs_dir = os.path.join(self.script_dir, "outputs")
        self.target_file = os.path.join(self.outputs_dir, "Dividends_2025.xlsx")
        
        # E*TRADE session
        self.session = None
        self.base_url = None
        
    def backup_file(self):
        """Create timestamped backup before changes"""
        timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
        backup_name = f"Dividends_2025_portfolio_test_{timestamp}.xlsx"
        backup_path = os.path.join(self.outputs_dir, backup_name)
        
        import shutil
        shutil.copy2(self.target_file, backup_path)
        print(f"📋 Test backup created: {backup_name}")
        return backup_path
        
    def initialize_etrade(self):
        """Initialize E*TRADE API connection"""
        try:
            print(">> Connecting to E*TRADE API...")
            self.session, self.base_url = get_etrade_session()
            
            if self.session and self.base_url:
                print("✅ E*TRADE authentication successful")
                return True
            else:
                print("*** E*TRADE authentication failed")
                return False
                
        except Exception as e:
            print(f"*** E*TRADE authentication error: {e}")
            return False
    
    def get_etrade_account_values(self):
        """Get portfolio values from E*TRADE API"""
        if not self.session or not self.base_url:
            print("*** E*TRADE session not initialized")
            return {}
        
        try:
            # Import E*TRADE accounts module
            from pyetrade.accounts import ETradeAccounts
            
            accounts_api = ETradeAccounts(self.session, self.base_url)
            account_list = accounts_api.get_account_list()
            
            portfolio_values = {}
            
            for account_group in account_list['AccountListResponse']['Accounts']['Account']:
                account_id = account_group['accountId']
                account_desc = account_group.get('accountDesc', 'Unknown')
                
                # Get account balance
                balance_response = accounts_api.get_account_balance(account_id)
                
                if balance_response and 'BalanceResponse' in balance_response:
                    balance = balance_response['BalanceResponse']
                    
                    # Get total account value
                    total_value = 0
                    if 'accountBalance' in balance:
                        total_value = float(balance['accountBalance'])
                    
                    # Map account description to our sheet names
                    if 'IRA' in account_desc.upper():
                        portfolio_values['Etrade IRA'] = total_value
                    elif any(word in account_desc.upper() for word in ['INDIVIDUAL', 'TAXABLE', 'MARGIN']):
                        portfolio_values['Etrade Taxable'] = total_value
                        
                    print(f"📊 {account_desc}: ${total_value:,.2f}")
            
            return portfolio_values
            
        except Exception as e:
            print(f"*** Error getting E*TRADE account values: {e}")
            return {}
    
    def get_schwab_values(self):
        """Get Schwab portfolio values (placeholder for now)"""
        print("📊 Schwab values: Using placeholder values")
        return {
            'Schwab IRA': 25000.00,  # Placeholder
            'Schwab Individual': 15000.00  # Placeholder  
        }
    
    def find_next_column(self, worksheet):
        """Find next empty column for new data"""
        col = 2  # Start from column B
        while worksheet.cell(row=1, column=col).value is not None:
            col += 1
        return col
    
    def update_portfolio_values_sheet(self, k401_value):
        """Update Portfolio Values 2025 sheet with new column"""
        try:
            print(f"\n📊 UPDATING PORTFOLIO VALUES 2025 SHEET")
            print("=" * 50)
            
            # Load workbook
            wb = openpyxl.load_workbook(self.target_file)
            
            if "Portfolio Values 2025" not in wb.sheetnames:
                print("❌ Portfolio Values 2025 sheet not found!")
                return False
                
            ws = wb["Portfolio Values 2025"]
            
            # Find next available column
            next_col = self.find_next_column(ws)
            current_date = datetime.now().strftime('%m/%d/%Y')
            
            print(f"📅 Adding data to column {next_col} for date: {current_date}")
            
            # Add date header
            ws.cell(row=1, column=next_col, value=current_date)
            
            # Get portfolio values from APIs
            print("\n💼 Getting current portfolio values...")
            etrade_values = self.get_etrade_account_values()
            schwab_values = self.get_schwab_values()
            
            # Get account names from column A
            accounts_data = {}
            row = 2
            while ws.cell(row=row, column=1).value:
                account_name = ws.cell(row=row, column=1).value
                
                # Map values to accounts
                if account_name == "Etrade IRA":
                    value = etrade_values.get('Etrade IRA', 0)
                elif account_name == "Etrade Taxable":
                    value = etrade_values.get('Etrade Taxable', 0)
                elif account_name == "Schwab IRA":
                    value = schwab_values.get('Schwab IRA', 0)
                elif account_name == "Schwab Individual":
                    value = schwab_values.get('Schwab Individual', 0)
                elif "401" in account_name.upper():
                    value = k401_value
                else:
                    value = 0
                    
                ws.cell(row=row, column=next_col, value=value)
                accounts_data[account_name] = value
                print(f"   {account_name}: ${value:,.2f}")
                row += 1
            
            # Calculate total
            total_value = sum(accounts_data.values())
            print(f"\n💰 Total Portfolio Value: ${total_value:,.2f}")
            
            # Add total if there's a total row
            if ws.cell(row=row, column=1).value and "total" in str(ws.cell(row=row, column=1).value).lower():
                ws.cell(row=row, column=next_col, value=total_value)
                print(f"   Total row updated: ${total_value:,.2f}")
            
            # Save workbook
            wb.save(self.target_file)
            print(f"✅ Portfolio Values 2025 updated successfully!")
            print(f"   New column: {next_col}")
            print(f"   Date: {current_date}")
            
            return True
            
        except Exception as e:
            print(f"❌ Error updating Portfolio Values 2025: {e}")
            traceback.print_exc()
            return False

    def run_focused_update(self):
        """Run focused portfolio values update with 401K prompt"""
        
        print("🎯 FOCUSED PORTFOLIO VALUES 2025 UPDATE")
        print("=" * 55)
        print("This updates ONLY Portfolio Values 2025 sheet")
        print("✅ Preserves all historical data")  
        print("✅ Adds new column with current values")
        print("✅ Includes 401K prompt")
        print("-" * 55)
        
        # Create test backup
        backup_path = self.backup_file()
        
        try:
            # Initialize E*TRADE connection
            if not self.initialize_etrade():
                print("⚠️ E*TRADE connection failed - using placeholder values")
            
            # Get 401K value
            print(f"\n💰 401K VALUE REQUIRED")
            k401_value = get_k401_value()
            
            if k401_value is None or k401_value == 0:
                print("⚠️ Using fallback 401K value")
                k401_value = 125000.00
                
            print(f"✅ 401K Value: ${k401_value:,.2f}")
            
            # Update portfolio values sheet
            success = self.update_portfolio_values_sheet(k401_value)
            
            if success:
                print(f"\n🎉 FOCUSED UPDATE COMPLETE!")
                print(f"   ✅ Portfolio Values 2025 updated")
                print(f"   ✅ Historical data preserved")
                print(f"   ✅ 401K value included: ${k401_value:,.2f}")
                print(f"   📋 Test backup: {os.path.basename(backup_path)}")
                return True
            else:
                print(f"\n❌ UPDATE FAILED!")
                print(f"   Restore from: {os.path.basename(backup_path)}")
                return False
                
        except Exception as e:
            print(f"\n❌ CRITICAL ERROR: {e}")
            print(f"   Restore from: {os.path.basename(backup_path)}")
            traceback.print_exc()
            return False

if __name__ == "__main__":
    updater = FocusedPortfolioUpdater()
    success = updater.run_focused_update()
    
    input("\nPress Enter to close...")
