"""
Safe Weekly Update Script
Purpose: Add new weekly data WITHOUT destroying existing historical data
Author: Assistant (GitHub Copilot)
Created: August 31, 2025
"""

import os
import sys
import openpyxl
from datetime import datetime
import pandas as pd

# Add modules to path
sys.path.append(os.path.join(os.path.dirname(__file__), 'modules'))

try:
    from gui_prompts import get_401k_value
    from etrade_account_api import ETradeAccountAPI
except ImportError as e:
    print(f"❌ Import error: {e}")
    sys.exit(1)

# Configuration
OUTPUT_DIR = os.path.join(os.path.dirname(__file__), "outputs")
TARGET_FILE = os.path.join(OUTPUT_DIR, "Dividends_2025.xlsx")

def backup_workbook():
    """Create a safety backup before any changes"""
    timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
    backup_name = f"Dividends_2025_backup_safe_update_{timestamp}.xlsx"
    backup_path = os.path.join(OUTPUT_DIR, backup_name)
    
    import shutil
    shutil.copy2(TARGET_FILE, backup_path)
    print(f"📋 Backup created: {backup_name}")
    return backup_path

def find_next_available_column(worksheet, start_col=2):
    """Find the next empty column for new data"""
    col = start_col
    while worksheet.cell(row=1, column=col).value is not None:
        col += 1
    return col

def add_weekly_dividend_data(wb, current_date):
    """Add new weekly dividend data to existing sheets WITHOUT destroying data"""
    
    # Check if "All account weekly dividends" sheet exists
    if "All account weekly dividends" not in wb.sheetnames:
        print("❌ 'All account weekly dividends' sheet not found")
        return False
    
    ws = wb["All account weekly dividends"]
    
    # Find next available column
    next_col = find_next_available_column(ws)
    print(f"📊 Adding new data to column {next_col}")
    
    # Add date header
    ws.cell(row=1, column=next_col, value=current_date)
    
    # Get fresh dividend data
    try:
        data_provider = DataProvider()
        accounts_data = data_provider.get_all_accounts_data()
        
        # Add data for each account row
        row_num = 2
        for account_type, account_data in accounts_data.items():
            monthly_dividend = account_data.get('monthly_dividend', 0)
            ws.cell(row=row_num, column=next_col, value=monthly_dividend)
            row_num += 1
            
        print(f"✅ Added weekly dividend data for {current_date}")
        return True
        
    except Exception as e:
        print(f"❌ Error adding dividend data: {e}")
        return False

def add_weekly_portfolio_data(wb, current_date, k401_value=None):
    """Add new weekly portfolio data WITHOUT destroying existing data"""
    
    # Check if "Portfolio Values 2025" sheet exists
    if "Portfolio Values 2025" not in wb.sheetnames:
        print("❌ 'Portfolio Values 2025' sheet not found")
        return False
        
    ws = wb["Portfolio Values 2025"]
    
    # Find next available column
    next_col = find_next_available_column(ws)
    print(f"💰 Adding portfolio data to column {next_col}")
    
    # Add date header
    ws.cell(row=1, column=next_col, value=current_date)
    
    # Get 401K value if not provided
    if k401_value is None:
        k401_value = get_401k_value()
    
    try:
        data_provider = DataProvider()
        accounts_data = data_provider.get_all_accounts_data()
        
        # Add portfolio values for each account
        row_num = 2  # Start after header
        
        # E*TRADE IRA
        etrade_ira_value = accounts_data.get('Etrade IRA', {}).get('portfolio_value', 0)
        ws.cell(row=row_num, column=next_col, value=etrade_ira_value)
        row_num += 1
        
        # E*TRADE Taxable  
        etrade_taxable_value = accounts_data.get('Etrade Taxable', {}).get('portfolio_value', 0)
        ws.cell(row=row_num, column=next_col, value=etrade_taxable_value)
        row_num += 1
        
        # Schwab IRA
        schwab_ira_value = accounts_data.get('Schwab IRA', {}).get('portfolio_value', 0)
        ws.cell(row=row_num, column=next_col, value=schwab_ira_value)
        row_num += 1
        
        # Schwab Individual
        schwab_ind_value = accounts_data.get('Schwab Individual', {}).get('portfolio_value', 0)
        ws.cell(row=row_num, column=next_col, value=schwab_ind_value)
        row_num += 1
        
        # 401K value
        ws.cell(row=row_num, column=next_col, value=k401_value)
        row_num += 1
        
        # Calculate total
        total_value = etrade_ira_value + etrade_taxable_value + schwab_ira_value + schwab_ind_value + k401_value
        ws.cell(row=row_num, column=next_col, value=total_value)
        
        print(f"✅ Added weekly portfolio data for {current_date}")
        print(f"   401K: ${k401_value:,.2f}")
        print(f"   Total Portfolio: ${total_value:,.2f}")
        return True
        
    except Exception as e:
        print(f"❌ Error adding portfolio data: {e}")
        return False

def update_estimated_income_sheet(wb, current_date):
    """Add new data to Estimated Income 2025 sheet WITHOUT destroying existing data"""
    
    if "Estimated Income 2025" not in wb.sheetnames:
        print("❌ 'Estimated Income 2025' sheet not found")
        return False
        
    ws = wb["Estimated Income 2025"]
    
    # Find next available column
    next_col = find_next_available_column(ws)
    print(f"💵 Adding estimated income data to column {next_col}")
    
    # Add date header
    ws.cell(row=1, column=next_col, value=current_date)
    
    try:
        data_provider = DataProvider()
        accounts_data = data_provider.get_all_accounts_data()
        
        # Calculate monthly totals by account
        row_num = 2
        
        for account_type, account_data in accounts_data.items():
            monthly_dividend = account_data.get('monthly_dividend', 0)
            ws.cell(row=row_num, column=next_col, value=monthly_dividend)
            row_num += 1
            
        print(f"✅ Added estimated income data for {current_date}")
        return True
        
    except Exception as e:
        print(f"❌ Error updating estimated income: {e}")
        return False

def safe_weekly_update():
    """Perform safe weekly update - adds new data without destroying existing data"""
    
    print("🔒 === SAFE WEEKLY UPDATE ===")
    print("This will ADD new weekly data WITHOUT destroying existing historical data")
    print()
    
    # Create backup
    backup_path = backup_workbook()
    
    try:
        # Load existing workbook
        wb = openpyxl.load_workbook(TARGET_FILE)
        current_date = datetime.now().strftime('%m/%d/%Y')
        
        print(f"📅 Adding data for: {current_date}")
        
        # Get 401K value
        print("\n💰 401K Value Required")
        k401_value = get_401k_value()
        
        # Perform safe updates
        success_count = 0
        
        # 1. Add dividend data
        if add_weekly_dividend_data(wb, current_date):
            success_count += 1
            
        # 2. Add portfolio data  
        if add_weekly_portfolio_data(wb, current_date, k401_value):
            success_count += 1
            
        # 3. Update estimated income
        if update_estimated_income_sheet(wb, current_date):
            success_count += 1
        
        # Save changes
        wb.save(TARGET_FILE)
        
        print(f"\n✅ SAFE UPDATE COMPLETE")
        print(f"   Successfully updated {success_count}/3 components")
        print(f"   Backup saved: {os.path.basename(backup_path)}")
        print(f"   All historical data preserved!")
        
        return True
        
    except Exception as e:
        print(f"❌ SAFE UPDATE FAILED: {e}")
        print(f"   Restore from backup: {os.path.basename(backup_path)}")
        return False

if __name__ == "__main__":
    safe_weekly_update()
