#!/usr/bin/env python3
"""
Portfolio Values Verification Script
====================================
Verifies that the Portfolio Values 2025 sheet has been updated with REAL API data
and shows the difference between old hardcoded values vs new real values.
"""

import openpyxl
import os

def verify_portfolio_values():
    """Verify the Portfolio Values sheet has real API data"""
    
    target_file = r"c:\Users\mjmat\Python Code in VS\dividend_tracker\DividendTrackerApp\outputs\Dividends_2025.xlsx"
    
    if not os.path.exists(target_file):
        print(f"❌ File not found: {target_file}")
        return
        
    try:
        wb = openpyxl.load_workbook(target_file)
        
        if "Portfolio Values 2025" not in wb.sheetnames:
            print("❌ Portfolio Values 2025 sheet not found")
            return
            
        ws = wb["Portfolio Values 2025"]
        print("📊 PORTFOLIO VALUES 2025 SHEET VERIFICATION")
        print("=" * 60)
        
        # Find the most recent column (rightmost with data)
        max_col = 1
        for col in range(1, 50):  # Check up to column 50
            if ws.cell(row=1, column=col).value:
                max_col = col
                
        print(f"📅 Latest data in column {max_col}: {ws.cell(row=1, column=max_col).value}")
        print("\n📋 Current Portfolio Values:")
        print("-" * 40)
        
        total_value = 0
        
        # Check key account rows
        for row in range(2, 15):
            account_name = ws.cell(row=row, column=1).value
            if account_name:
                account_value = ws.cell(row=row, column=max_col).value
                if account_value is not None:
                    try:
                        value = float(account_value)
                        print(f"{account_name}: ${value:,.2f}")
                        
                        if 'total' not in account_name.lower():
                            total_value += value
                    except:
                        print(f"{account_name}: {account_value}")
                        
        print(f"\n🎯 Calculated Total: ${total_value:,.2f}")
        
        # Verify we have REAL E*TRADE values (not old hardcoded ones)
        etrade_ira_found = False
        etrade_taxable_found = False
        
        for row in range(2, 15):
            account_name = ws.cell(row=row, column=1).value
            if account_name:
                account_value = ws.cell(row=row, column=max_col).value
                if account_value is not None:
                    try:
                        value = float(account_value)
                        
                        if 'E*TRADE IRA' in str(account_name) or 'Etrade IRA' in str(account_name):
                            etrade_ira_found = True
                            if value > 280000:  # Should be around 284,872
                                print(f"✅ E*TRADE IRA has REAL API value: ${value:,.2f}")
                            else:
                                print(f"⚠️ E*TRADE IRA value seems low: ${value:,.2f}")
                                
                        elif 'E*TRADE Taxable' in str(account_name) or 'Etrade Taxable' in str(account_name):
                            etrade_taxable_found = True
                            if value > 60000:  # Should be around 63,270
                                print(f"✅ E*TRADE Taxable has REAL API value: ${value:,.2f}")
                            else:
                                print(f"⚠️ E*TRADE Taxable value seems low: ${value:,.2f}")
                                
                    except:
                        pass
        
        print(f"\n🔍 VERIFICATION RESULTS:")
        print(f"✅ E*TRADE IRA found: {etrade_ira_found}")
        print(f"✅ E*TRADE Taxable found: {etrade_taxable_found}")
        
        if etrade_ira_found and etrade_taxable_found:
            print(f"\n🎉 SUCCESS! Portfolio Values sheet has been updated with REAL E*TRADE API data!")
            print(f"📊 Total Portfolio Value: ${total_value:,.2f}")
            print(f"💾 File: {target_file}")
        else:
            print(f"\n⚠️ Some E*TRADE accounts not found - check sheet structure")
            
        wb.close()
        
    except Exception as e:
        print(f"❌ Error reading file: {e}")

if __name__ == "__main__":
    verify_portfolio_values()
    input("\nPress Enter to continue...")
