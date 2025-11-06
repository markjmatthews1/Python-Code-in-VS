#!/usr/bin/env python3
"""
Fix Missing 401K Retirement Percentage - Row 18 Column B
Specifically target the 401K Retirement row that's still missing its percentage
"""
import openpyxl
from openpyxl.styles import Font, Alignment

def fix_401k_percentage():
    """Fix the missing 401K percentage in row 18 column B"""
    
    excel_file = 'outputs/Dividends_2025.xlsx'
    
    try:
        print("🔧 Fixing missing 401K Retirement percentage...")
        
        wb = openpyxl.load_workbook(excel_file)
        portfolio_ws = wb['Portfolio Summary']
        
        # Calculate 401K percentage
        k401_value = 124315.15
        total_portfolio = 519439.06
        k401_pct = k401_value / total_portfolio
        
        # Define formatting
        normal_font = Font(name='Arial', size=12)
        right_aligned = Alignment(horizontal='right')
        percentage_format = '0.0%'
        
        print(f"📊 Searching for 401K Retirement row...")
        
        # Search specifically around row 18 and nearby rows for 401K Retirement
        found_401k = False
        for row in range(15, 25):
            try:
                cell_value = portfolio_ws.cell(row=row, column=1).value
                if cell_value and "401k retirement" in str(cell_value).lower():
                    print(f"   Found 401K Retirement at row {row}: '{cell_value}'")
                    
                    # Get the cell in column B
                    pct_cell = portfolio_ws.cell(row=row, column=2)
                    
                    # Check if it's a merged cell
                    is_merged = False
                    for merged_range in portfolio_ws.merged_cells.ranges:
                        if (merged_range.min_row <= row <= merged_range.max_row and
                            merged_range.min_col <= 2 <= merged_range.max_col):
                            # Get the top-left cell of the merged range
                            pct_cell = portfolio_ws.cell(row=merged_range.min_row, column=merged_range.min_col)
                            is_merged = True
                            print(f"   Row {row} is part of merged range - using cell ({merged_range.min_row},{merged_range.min_col})")
                            break
                    
                    # Set the value and formatting
                    current_value = pct_cell.value
                    print(f"   Current value in column B: '{current_value}'")
                    
                    pct_cell.value = k401_pct
                    pct_cell.number_format = percentage_format
                    pct_cell.font = normal_font
                    pct_cell.alignment = right_aligned
                    
                    print(f"   ✅ Set 401K Retirement percentage to {k401_pct:.1%} at row {row}")
                    found_401k = True
                    break
                    
            except Exception as e:
                print(f"   Warning: Error checking row {row}: {e}")
                continue
        
        if not found_401k:
            print("   ❌ 401K Retirement row not found in expected range")
            
            # Try a broader search
            print("   🔍 Doing broader search...")
            for row in range(1, 50):
                try:
                    cell_value = portfolio_ws.cell(row=row, column=1).value
                    if cell_value and ("401" in str(cell_value) or "retirement" in str(cell_value).lower()):
                        print(f"   Found potential match at row {row}: '{cell_value}'")
                except Exception:
                    continue
        
        # Save the workbook
        wb.save(excel_file)
        wb.close()
        
        return found_401k
        
    except Exception as e:
        print(f"❌ Error fixing 401K percentage: {e}")
        return False

def verify_401k_fix():
    """Verify the 401K percentage was fixed"""
    excel_file = 'outputs/Dividends_2025.xlsx'
    
    try:
        wb = openpyxl.load_workbook(excel_file, data_only=True)
        portfolio_ws = wb['Portfolio Summary']
        
        print("\n📊 VERIFICATION - Account Breakdown Section:")
        print("=" * 60)
        
        # Show the account breakdown section
        for row in range(13, 23):
            col_a = portfolio_ws.cell(row=row, column=1).value or ""
            col_b = portfolio_ws.cell(row=row, column=2).value or ""
            
            if col_a and ("account breakdown" in str(col_a).lower() or 
                         "e*trade" in str(col_a).lower() or 
                         "schwab" in str(col_a).lower() or 
                         "401k" in str(col_a).lower() or
                         "retirement" in str(col_a).lower()):
                
                # Format the display
                if isinstance(col_b, (int, float)) and col_b < 1 and col_b > 0:
                    display_text = f"Row {row}: {col_a} {col_b:.1%}"
                else:
                    display_text = f"Row {row}: {col_a} {col_b}"
                
                print(display_text)
        
        wb.close()
        
    except Exception as e:
        print(f"Error: {e}")

if __name__ == "__main__":
    print("🎯 FIXING MISSING 401K RETIREMENT PERCENTAGE")
    print("=" * 60)
    print("Target: Row 18 Column B - 401K Retirement percentage")
    print("Expected: 23.9% (124,315.15 / 519,439.06)")
    print("=" * 60)
    
    success = fix_401k_percentage()
    
    if success:
        verify_401k_fix()
        print("\n✅ 401K Retirement percentage fix completed!")
    else:
        print("\n❌ Could not locate or fix 401K Retirement percentage")
        print("Please check the Portfolio Summary sheet manually")
