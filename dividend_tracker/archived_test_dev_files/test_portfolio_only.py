import os
import sys
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.styles.numbers import FORMAT_CURRENCY_USD_SIMPLE
from datetime import datetime
import json

def test_portfolio_update():
    """Test just the Portfolio Values update with the correct 401k value"""
    
    # Test data - your actual values
    k401_value = 128693.17
    portfolio_values = {
        'E*TRADE IRA': 289870.9963,
        'E*TRADE Taxable': 65093.35,
        'Schwab IRA': 51929.91,
        'Schwab Individual': 2660.97
    }
    
    excel_file = "outputs/Dividends_2025.xlsx"
    if not os.path.exists(excel_file):
        print(f"ERROR: Excel file not found: {excel_file}")
        return
    
    try:
        wb = openpyxl.load_workbook(excel_file)
        ws = wb['Portfolio Values 2025']
        
        # Find the last column and add new column
        new_col = ws.max_column + 1
        today_str = datetime.now().strftime("%m/%d/%Y")
        
        print(f"Adding data to column {new_col} with date {today_str}")
        
        # Add date header
        date_cell = ws.cell(row=3, column=new_col, value=today_str)
        date_cell.font = Font(name='Arial', size=12, bold=True)
        date_cell.alignment = Alignment(horizontal='center')
        
        # Account mapping
        account_mapping = {
            'E*TRADE IRA': portfolio_values.get('E*TRADE IRA', 0),
            'E*TRADE Taxable': portfolio_values.get('E*TRADE Taxable', 0),
            'Schwab IRA': portfolio_values.get('Schwab IRA', 0),
            'Schwab Individual': portfolio_values.get('Schwab Individual', 0),
            '401k Retirement (Manual)': k401_value
        }
        
        # Update rows 4-10
        for row in range(4, 11):
            account_name = ws.cell(row=row, column=1).value
            if account_name:
                account_key = str(account_name).strip()
                value = None
                
                if account_key in account_mapping:
                    value = account_mapping[account_key]
                    print(f"Row {row}: '{account_key}' = ${value:,.2f} (Exact match)")
                elif 'total' in account_key.lower():
                    value = sum(portfolio_values.values()) + k401_value
                    print(f"Row {row}: '{account_key}' = ${value:,.2f} (Calculated total)")
                elif '401k' in account_key.lower() or '401' in account_key:
                    value = k401_value
                    print(f"Row {row}: '{account_key}' = ${value:,.2f} (401k match)")
                
                if value is not None:
                    cell = ws.cell(row=row, column=new_col, value=value)
                    cell.number_format = FORMAT_CURRENCY_USD_SIMPLE
                    print(f"  Updated cell {openpyxl.utils.get_column_letter(new_col)}{row} = ${value:,.2f}")
                else:
                    print(f"Row {row}: '{account_key}' - NO MATCH")
        
        # Save the file
        wb.save(excel_file)
        wb.close()
        
        print(f"\nSUCCESS: Portfolio Values updated with 401k value ${k401_value:,.2f}")
        print(f"Expected total: ${sum(portfolio_values.values()) + k401_value:,.2f}")
        
    except Exception as e:
        print(f"ERROR: {e}")
        import traceback
        traceback.print_exc()

if __name__ == "__main__":
    test_portfolio_update()