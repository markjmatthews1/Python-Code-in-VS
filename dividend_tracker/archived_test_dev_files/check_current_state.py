"""
Simple Current State Check

Check what's actually in the Excel file right now.
"""
import openpyxl

def check_current_state():
    """Check the current state of the Excel file"""
    excel_file = r"c:\Users\mjmat\Python Code in VS\dividend_tracker\DividendTrackerApp\outputs\Dividends_2025.xlsx"
    
    try:
        # Load with formulas showing
        wb = openpyxl.load_workbook(excel_file, data_only=False)
        ws = wb["Estimated Income 2025"]
        
        max_col = ws.max_column
        current_col_letter = openpyxl.utils.get_column_letter(max_col)
        
        print(f"🔍 CURRENT STATE CHECK - Column {current_col_letter}")
        print("=" * 50)
        
        # Check rows 4-9
        for row in range(4, 10):
            account_name = ws.cell(row=row, column=1).value
            cell_value = ws.cell(row=row, column=max_col).value
            
            # Check if cell has fill color
            fill = ws.cell(row=row, column=max_col).fill
            fill_color = fill.start_color.index if fill.start_color else 'None'
            
            print(f"   Row {row} ({account_name}): {cell_value} | Color: {fill_color}")
        
        wb.close()
        
        # Now load with calculated values
        print(f"\n📊 CALCULATED VALUES:")
        wb = openpyxl.load_workbook(excel_file, data_only=True)
        ws = wb["Estimated Income 2025"]
        
        for row in range(4, 10):
            account_name = ws.cell(row=row, column=1).value
            calc_value = ws.cell(row=row, column=max_col).value
            
            print(f"   Row {row} ({account_name}): {calc_value}")
        
        wb.close()
        
    except Exception as e:
        print(f"❌ ERROR: {e}")

if __name__ == "__main__":
    check_current_state()
