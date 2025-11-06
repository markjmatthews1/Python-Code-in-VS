"""
Fix Estimated Income 2025 Sheet Formatting and Row 9 Calculation

Based on Dividend Tracker Plan requirements:
- Row 9: Calculate monthly average = SUM(rows 5:7)/12  
- Apply proper formatting (Arial 12pt, currency format)
- Don't change data, only fix calculation and formatting
"""
import openpyxl
from openpyxl.styles import Font, Border, Side, Alignment
from openpyxl.styles.numbers import FORMAT_CURRENCY_USD_SIMPLE

def fix_estimated_income_formatting():
    """Fix formatting and row 9 calculation in Estimated Income 2025 sheet"""
    excel_file = r"c:\Users\mjmat\Python Code in VS\dividend_tracker\DividendTrackerApp\outputs\Dividends_2025.xlsx"
    
    try:
        wb = openpyxl.load_workbook(excel_file)
        
        if "Estimated Income 2025" not in wb.sheetnames:
            print("❌ Estimated Income 2025 sheet not found")
            return False
            
        ws = wb["Estimated Income 2025"]
        print("🎯 FIXING ESTIMATED INCOME 2025 FORMATTING")
        print("=" * 50)
        
        # Define proper formatting based on plan
        arial_font = Font(name='Arial', size=12)
        arial_bold = Font(name='Arial', size=12, bold=True)
        header_font = Font(name='Arial', size=12, bold=True, color='FFFFFF')  # White text for headers
        currency_format = FORMAT_CURRENCY_USD_SIMPLE
        
        # Find the data range - dates should be in row 3
        print("📊 Analyzing sheet structure...")
        
        # Apply formatting to all data columns (starting from column with dates)
        date_row = 3
        max_col = ws.max_column
        max_row = ws.max_row
        
        print(f"📋 Sheet dimensions: {max_row} rows x {max_col} columns")
        
        # Format row 3 (dates) with proper header formatting
        print("🎨 Formatting date headers (row 3)...")
        for col in range(1, max_col + 1):
            cell = ws.cell(row=3, column=col)
            if cell.value:  # Only format cells with data
                cell.font = header_font
                cell.alignment = Alignment(horizontal='right')
                
        # Format account rows (4-8) with proper fonts and currency format
        print("💰 Formatting account rows (4-8)...")
        account_rows = [4, 5, 6, 7, 8]  # E*TRADE IRA, E*TRADE Taxable, Schwab IRA, Schwab Individual, 401K
        
        for row in account_rows:
            # Column A (account names) - Arial 12
            ws.cell(row=row, column=1).font = arial_font
            
            # All data columns - currency format + Arial 12
            for col in range(2, max_col + 1):
                cell = ws.cell(row=row, column=col)
                if cell.value is not None:
                    cell.font = arial_font
                    if isinstance(cell.value, (int, float)):
                        cell.number_format = currency_format
        
        # Fix Row 9 Calculation - Monthly Average (SUM of dividend accounts / 12)
        print("🧮 Fixing Row 9 monthly calculation...")
        
        # Row 9 should calculate monthly average from dividend account rows (5, 6, 7)
        # Rows 5-7 are: E*TRADE Taxable, Schwab IRA, Schwab Individual (dividend accounts)
        # Based on plan: "sum row 5: row 7 /12"
        
        # Set row 9 label
        ws.cell(row=9, column=1, value="Monthly Average")
        ws.cell(row=9, column=1).font = arial_font
        
        # Add the formula for each data column
        for col in range(2, max_col + 1):
            if ws.cell(row=5, column=col).value is not None:  # Only if there's data above
                # Formula: =SUM(row5:row7)/12
                col_letter = openpyxl.utils.get_column_letter(col)
                formula = f"=SUM({col_letter}5:{col_letter}7)/12"
                
                cell = ws.cell(row=9, column=col)
                cell.value = formula
                cell.font = arial_font
                cell.number_format = currency_format
        
        # Format row 10 (Total) if it exists
        print("📊 Formatting total row...")
        for row in range(9, max_row + 1):
            cell_value = ws.cell(row=row, column=1).value
            if cell_value and 'total' in str(cell_value).lower():
                # This is the total row
                ws.cell(row=row, column=1).font = arial_bold
                
                for col in range(2, max_col + 1):
                    cell = ws.cell(row=row, column=col)
                    if cell.value is not None:
                        cell.font = arial_bold
                        if isinstance(cell.value, (int, float)):
                            cell.number_format = currency_format
                break
        
        # Set column widths for proper display
        print("📏 Adjusting column widths...")
        for col in range(1, max_col + 1):
            ws.column_dimensions[openpyxl.utils.get_column_letter(col)].width = 15
        
        # Save the workbook
        wb.save(excel_file)
        wb.close()
        
        print("✅ SUCCESS: Estimated Income 2025 formatting and calculation fixed!")
        print("📋 Changes made:")
        print("   • Row 3: Date headers formatted with proper styling")
        print("   • Rows 4-8: Account values formatted as currency with Arial 12pt")  
        print("   • Row 9: Monthly average calculation fixed (=SUM(rows5:7)/12)")
        print("   • Column widths set to 15 for proper display")
        print("   • Total row formatted with bold font")
        
        return True
        
    except Exception as e:
        print(f"❌ ERROR fixing formatting: {e}")
        import traceback
        traceback.print_exc()
        return False

if __name__ == "__main__":
    fix_estimated_income_formatting()
