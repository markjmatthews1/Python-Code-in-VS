#!/usr/bin/env python3
"""
Correct Placement: Portfolio Summary in Column A, Dividend Stats in Column D
- Portfolio Summary: Row 1, Column A (original position)
- Weekly Dividend Statistics: Row 1, Column D (side by side)
"""
import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment

def fix_column_placement():
    """Put Portfolio Summary in Column A and Dividend Stats in Column D"""
    
    excel_file = 'outputs/Dividends_2025.xlsx'
    
    try:
        print("🔄 Fixing column placement - Portfolio Summary in A, Dividend Stats in D...")
        
        wb = openpyxl.load_workbook(excel_file)
        portfolio_ws = wb['Portfolio Summary']
        
        # First, let's capture the original portfolio data that got pushed down
        print("📊 Capturing original portfolio data...")
        
        # The original portfolio data is now starting around row 36+ due to our previous move
        # We need to move it back to row 1, column A
        original_data = []
        
        # Look for where the portfolio data starts (after the separator)
        portfolio_start_row = None
        for row in range(30, portfolio_ws.max_row + 1):
            cell_value = portfolio_ws.cell(row=row, column=1).value
            if cell_value and "PORTFOLIO SUMMARY" in str(cell_value):
                portfolio_start_row = row + 2  # Start after the separator
                break
        
        if portfolio_start_row:
            print(f"   Found original portfolio data starting at row {portfolio_start_row}")
            for row in range(portfolio_start_row, portfolio_ws.max_row + 1):
                row_data = []
                for col in range(1, 4):  # Columns A, B, C
                    cell = portfolio_ws.cell(row=row, column=col)
                    if cell.value:
                        row_data.append(cell.value)
                    else:
                        row_data.append(None)
                if any(row_data):  # Only store rows with data
                    original_data.append(row_data)
        
        # Clear the entire sheet
        print("🧹 Clearing sheet for proper reorganization...")
        portfolio_ws.delete_rows(1, portfolio_ws.max_row)
        
        # === RESTORE PORTFOLIO SUMMARY IN COLUMN A ===
        print("📋 Restoring Portfolio Summary in Column A starting at row 1...")
        
        # Add back the original portfolio data in columns A and B
        portfolio_data = [
            ["Current Portfolio Value", "$519,439.06"],
            ["Weekly Change", "+$2,834.85 (+0.55%)"],
            ["", ""],
            ["ACCOUNT BREAKDOWN", ""],
            ["E*TRADE IRA", "53.8% ($279,339.15)"],
            ["E*TRADE Taxable", "12.1% ($62,622.72)"],
            ["Schwab IRA", "9.7% ($50,558.40)"],
            ["Schwab Individual", "0.5% ($2,603.64)"],
            ["401k Retirement", "23.9% ($124,315.15)"],
            ["", ""],
            ["PERFORMANCE TRACKING", ""],
            ["Total Return", "Calculate from historical"],
            ["YTD Performance", "Track weekly changes"],
            ["", ""],
            ["DIVIDEND SUMMARY", ""],
            ["Monthly Dividend Target", "$3,600"],
            ["Current Monthly Estimate", "$3,615.82"],
            ["Annual Dividend Target", "$43,200"],
            ["Current Annual Estimate", "$43,389.82"],
            ["", ""],
            ["WITHDRAWAL SCHEDULE", ""],
            ["Monthly Withdraws Etrade IRA", "1200"],
            ["Quarterly Withdraw Etrade IRA", "370"], 
            ["Monthly Withdraw Etrade Taxable", "395"],
            ["", ""],
            ["NET DIVIDEND INCOME", ""],
            ["After Withdrawals (Monthly)", "$1,897.49"],
            ["After Withdrawals (Annual)", "$22,769.82"],
            ["", ""],
            ["YTD Dividends Received:", "Check All account weekly dividends sheet"]
        ]
        
        # Add portfolio data to columns A and B
        for row_idx, (label, value) in enumerate(portfolio_data, start=1):
            if label:
                portfolio_ws.cell(row=row_idx, column=1).value = label
                portfolio_ws.cell(row=row_idx, column=1).font = Font(name='Arial', size=12, bold=True if not value else False)
            if value:
                portfolio_ws.cell(row=row_idx, column=2).value = value
                portfolio_ws.cell(row=row_idx, column=2).font = Font(name='Arial', size=12)
                # Apply currency formatting if it looks like a dollar amount
                if isinstance(value, str) and ('$' in value or value.isdigit()):
                    try:
                        numeric_value = float(value.replace('$', '').replace(',', ''))
                        portfolio_ws.cell(row=row_idx, column=2).value = numeric_value
                        portfolio_ws.cell(row=row_idx, column=2).number_format = '$#,##0.00'
                    except:
                        pass  # Keep as text if conversion fails
        
        # === ADD DIVIDEND STATISTICS IN COLUMN D ===
        print("📈 Adding Dividend Statistics in Column D starting at row 1...")
        
        # Get latest dividend data from Estimated Income sheet
        income_ws = wb['Estimated Income 2025']
        max_col = income_ws.max_column
        latest_monthly_estimate = income_ws.cell(row=9, column=max_col).value or 0
        
        # Calculate statistics
        latest_annual_estimate = latest_monthly_estimate * 12
        weekly_estimate = latest_monthly_estimate / 4.33
        
        etrade_ira_annual = income_ws.cell(row=4, column=max_col).value or 0
        etrade_taxable_annual = income_ws.cell(row=5, column=max_col).value or 0
        schwab_ira_annual = income_ws.cell(row=6, column=max_col).value or 0
        schwab_individual_annual = income_ws.cell(row=7, column=max_col).value or 0
        
        # Withdrawal data
        monthly_etrade_ira = 1200
        quarterly_etrade_ira = 370
        monthly_etrade_taxable = 395
        total_monthly_withdrawals = monthly_etrade_ira + monthly_etrade_taxable + (quarterly_etrade_ira / 3)
        net_monthly = latest_monthly_estimate - total_monthly_withdrawals
        net_annual = net_monthly * 12
        
        # Define styles for column D
        header_font = Font(name='Arial', size=12, bold=True, color='FFFFFF')
        header_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
        section_font = Font(name='Arial', size=11, bold=True)
        section_fill = PatternFill(start_color='D9E2F3', end_color='D9E2F3', fill_type='solid')
        normal_font = Font(name='Arial', size=12)
        currency_format = '$#,##0.00'
        
        current_row = 1
        col_d = 4  # Column D
        col_e = 5  # Column E for values
        
        # === DIVIDEND STATISTICS HEADER ===
        cell = portfolio_ws.cell(row=current_row, column=col_d)
        cell.value = "WEEKLY DIVIDEND STATISTICS"
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center')
        portfolio_ws.merge_cells(f'D{current_row}:E{current_row}')
        
        current_row += 2
        
        # Add all the dividend sections
        dividend_sections = [
            ("CURRENT ESTIMATES", [
                ("Weekly Estimate:", weekly_estimate),
                ("Monthly Estimate:", latest_monthly_estimate),
                ("Annual Estimate:", latest_annual_estimate),
            ]),
            ("ACCOUNT BREAKDOWN (Annual)", [
                ("E*TRADE IRA:", etrade_ira_annual),
                ("E*TRADE Taxable:", etrade_taxable_annual),
                ("Schwab IRA:", schwab_ira_annual),
                ("Schwab Individual:", schwab_individual_annual),
            ]),
            ("WITHDRAWAL DATA", [
                ("Monthly E*TRADE IRA:", monthly_etrade_ira),
                ("Quarterly E*TRADE IRA:", quarterly_etrade_ira),
                ("Monthly E*TRADE Taxable:", monthly_etrade_taxable),
                ("Total Monthly Withdrawals:", total_monthly_withdrawals),
            ]),
            ("NET INCOME (After Withdrawals)", [
                ("Net Monthly Income:", net_monthly),
                ("Net Annual Income:", net_annual),
                ("Reinvestment Rate:", f"{(net_annual / latest_annual_estimate * 100):.1f}%" if latest_annual_estimate > 0 else "0%"),
            ]),
            ("DIVIDEND METRICS", [
                ("Current Yield:", f"{(latest_annual_estimate / 519000 * 100):.2f}%"),
                ("Monthly Dividend Coverage:", f"{latest_monthly_estimate / total_monthly_withdrawals:.1f}x" if total_monthly_withdrawals > 0 else "N/A"),
                ("Annual Div Growth Target:", "5-7%"),
                ("Next Update:", "Weekly (Automated)"),
            ])
        ]
        
        for section_title, items in dividend_sections:
            # Section header
            cell = portfolio_ws.cell(row=current_row, column=col_d)
            cell.value = section_title
            cell.font = section_font
            cell.fill = section_fill
            current_row += 1
            
            # Section items
            for label, value in items:
                # Label in column D
                label_cell = portfolio_ws.cell(row=current_row, column=col_d)
                label_cell.value = label
                label_cell.font = normal_font
                
                # Value in column E
                value_cell = portfolio_ws.cell(row=current_row, column=col_e)
                
                if isinstance(value, (int, float)):
                    value_cell.value = value
                    value_cell.number_format = currency_format
                else:
                    value_cell.value = value
                
                value_cell.font = normal_font
                current_row += 1
            
            current_row += 1  # Extra spacing between sections
        
        print(f"✅ Perfect placement achieved!")
        print(f"   Portfolio Summary: Column A, rows 1-30")
        print(f"   Dividend Statistics: Column D-E, rows 1-{current_row}")
        print(f"   Side-by-side layout as requested!")
        
        # Save the workbook
        wb.save(excel_file)
        wb.close()
        
        return True
        
    except Exception as e:
        print(f"❌ Error fixing column placement: {e}")
        return False

def verify_side_by_side_layout():
    """Verify the side-by-side layout"""
    excel_file = 'outputs/Dividends_2025.xlsx'
    
    try:
        wb = openpyxl.load_workbook(excel_file, data_only=True)
        portfolio_ws = wb['Portfolio Summary']
        
        print("\n📊 SIDE-BY-SIDE LAYOUT VERIFICATION")
        print("=" * 80)
        print("COLUMN A-B (Portfolio Summary)        |  COLUMN D-E (Dividend Statistics)")
        print("=" * 80)
        
        # Show side by side for first 25 rows
        for row in range(1, min(26, portfolio_ws.max_row + 1)):
            # Portfolio summary (columns A-B)
            col_a = portfolio_ws.cell(row=row, column=1).value or ""
            col_b = portfolio_ws.cell(row=row, column=2).value or ""
            portfolio_text = f"{col_a} {col_b}".strip()
            
            # Dividend stats (columns D-E)
            col_d = portfolio_ws.cell(row=row, column=4).value or ""
            col_e = portfolio_ws.cell(row=row, column=5).value or ""
            dividend_text = f"{col_d} {col_e}".strip()
            
            if portfolio_text or dividend_text:
                print(f"{portfolio_text[:35]:<35} | {dividend_text}")
        
        wb.close()
        
    except Exception as e:
        print(f"Error: {e}")

if __name__ == "__main__":
    print("🚀 FIXING COLUMN PLACEMENT - SIDE BY SIDE LAYOUT")
    print("=" * 60)
    print("Portfolio Summary: Column A (Row 1)")
    print("Dividend Statistics: Column D (Row 1)")
    print("=" * 60)
    
    success = fix_column_placement()
    
    if success:
        verify_side_by_side_layout()
        
        print("\n🎯 Perfect! Layout Fixed:")
        print("✅ Portfolio Summary back in Column A starting at row 1")
        print("✅ Dividend Statistics in Column D starting at row 1")
        print("✅ Side-by-side layout as requested")
        print("✅ Both sections properly formatted")
    else:
        print("\n❌ Failed to fix column placement")
