"""
Fix Row 9 (Monthly Average) Color Coding

The monthly average row contains a formula, so it needs special handling
for color coding comparison.
"""
import openpyxl
from openpyxl.styles import PatternFill

def fix_row9_color_coding():
    """Fix color coding for Row 9 (Monthly Average) formula cell"""
    excel_file = r"c:\Users\mjmat\Python Code in VS\dividend_tracker\DividendTrackerApp\outputs\Dividends_2025.xlsx"
    
    try:
        # Load with calculated values
        wb = openpyxl.load_workbook(excel_file, data_only=True)
        ws = wb["Estimated Income 2025"]
        
        print("🎯 FIXING ROW 9 MONTHLY AVERAGE COLOR CODING")
        print("=" * 50)
        
        # Find the most recent columns
        max_col = ws.max_column
        current_col = max_col
        prev_col = max_col - 1
        
        # Find Monthly Average row
        monthly_row = None
        for row in range(8, 12):
            cell_value = ws.cell(row=row, column=1).value
            if cell_value and 'monthly' in str(cell_value).lower():
                monthly_row = row
                break
        
        if monthly_row and prev_col >= 2:
            current_cell = ws.cell(row=monthly_row, column=current_col)
            prev_cell = ws.cell(row=monthly_row, column=prev_col)
            
            current_val = current_cell.value
            prev_val = prev_cell.value
            
            print(f"📊 Monthly Average comparison:")
            print(f"   Previous: {prev_val}")
            print(f"   Current: {current_val}")
            
            if (current_val is not None and prev_val is not None and 
                isinstance(current_val, (int, float)) and isinstance(prev_val, (int, float))):
                
                if current_val > prev_val:
                    # Green for increase
                    current_cell.fill = PatternFill(start_color='90EE90', end_color='90EE90', fill_type='solid')
                    change_type = "📈 INCREASE (Green)"
                elif current_val < prev_val:
                    # Red for decrease
                    current_cell.fill = PatternFill(start_color='FFB6C1', end_color='FFB6C1', fill_type='solid')
                    change_type = "📉 DECREASE (Red)"
                else:
                    # Yellow for same
                    current_cell.fill = PatternFill(start_color='FFFF99', end_color='FFFF99', fill_type='solid')
                    change_type = "➡️ SAME (Yellow)"
                
                print(f"   ✅ Row {monthly_row} (Monthly Average): ${prev_val:,.2f} → ${current_val:,.2f} {change_type}")
                
            else:
                print("   ⚠️ Could not compare values - one or both are not numeric")
        
        # Save the result
        wb.save(excel_file)
        wb.close()
        
        print("\n✅ Row 9 Monthly Average color coding applied!")
        
        return True
        
    except Exception as e:
        print(f"❌ ERROR fixing Row 9 color coding: {e}")
        import traceback
        traceback.print_exc()
        return False

if __name__ == "__main__":
    fix_row9_color_coding()
