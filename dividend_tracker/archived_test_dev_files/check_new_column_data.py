#!/usr/bin/env python3
"""
Check Portfolio Values Data in New Column
==========================================

Check if the focused updater actually populated the data values
in the new 08/31/2025 column.
"""

import os
import openpyxl

# Configuration
OUTPUT_DIR = os.path.join(os.path.dirname(__file__), "outputs")
TARGET_FILE = os.path.join(OUTPUT_DIR, "Dividends_2025.xlsx")

def check_new_column_data():
    """Check the data in the newly added 08/31/2025 column"""
    
    try:
        print("📊 CHECKING NEW COLUMN DATA (08/31/2025)")
        print("=" * 50)
        
        # Load workbook
        wb = openpyxl.load_workbook(TARGET_FILE)
        ws = wb["Portfolio Values 2025"]
        
        # Column 2 should have the new 08/31/2025 data
        new_col = 2
        
        print(f"📅 Header: {ws.cell(row=1, column=new_col).value}")
        print(f"\n📊 Account Values in Column {new_col}:")
        
        total_value = 0
        
        # Check each account row
        for row in range(3, ws.max_row + 1):
            account_name = ws.cell(row=row, column=1).value
            account_value = ws.cell(row=row, column=new_col).value
            
            if account_name and "TOTAL" not in str(account_name).upper():
                print(f"   {account_name}: {account_value}")
                if account_value and isinstance(account_value, (int, float)):
                    total_value += account_value
        
        print(f"\n💰 Calculated Total: ${total_value:,.2f}")
        
        # Check if there's a total row
        total_row_value = ws.cell(row=10, column=new_col).value  # Row 10 is TOTAL PORTFOLIO
        print(f"💰 Total Row Value: {total_row_value}")
        
        # Compare with previous column if it exists
        if ws.max_column >= 3:
            print(f"\n📈 COMPARISON WITH PREVIOUS DATA:")
            print("(This will show if the backup was older than expected)")
            
        return True
        
    except Exception as e:
        print(f"❌ Error checking column data: {e}")
        import traceback
        traceback.print_exc()
        return False

if __name__ == "__main__":
    check_new_column_data()
    input("\nPress Enter to close...")
