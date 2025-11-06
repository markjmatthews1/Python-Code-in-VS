import openpyxl
import json
import os
from datetime import datetime
from openpyxl.styles import Font, PatternFill

class SimpleSurgicalUpdater:
    """Simple surgical updater that matches tickers by index position within each group"""
    
    def __init__(self):
        self.excel_file = r"C:\Users\mjmat\Python Code in VS\dividend_tracker\DividendTrackerApp\outputs\Dividends_2025.xlsx"
        self.cache_file = r"C:\Users\mjmat\Python Code in VS\dividend_tracker\DividendTrackerApp\portfolio_data_cache.json"
        self.account_mapping = {
            "Etrade IRA": "etrade_ira",
            "Etrade Taxable": "etrade_taxable", 
            "Schwab IRA": "schwab_ira",
            "Schwab Individual": "schwab_individual"
        }
    
    def run_update(self):
        """Main surgical update that preserves ticker order"""
        print("SIMPLE SURGICAL HISTORICAL YIELD UPDATER")
        print("=" * 55)
        
        # Load data
        cache_data = self.load_cache_data()
        if not cache_data:
            return False
            
        positions_data = cache_data.get('positions', {})
        yields_data = cache_data.get('yields', {})
        
        # Open Excel file
        wb = openpyxl.load_workbook(self.excel_file)
        ws = wb["Accounts Div historical yield"]
        
        # Find all account groups
        account_info = self.find_account_groups(ws)
        print(f"Found {len(account_info)} account groups")
        
        # Process each group surgically
        total_updated = 0
        for group_name in ["Etrade IRA", "Etrade Taxable", "Schwab IRA", "Schwab Individual"]:
            if group_name not in account_info:
                print(f"WARNING: {group_name} not found in Excel")
                continue
                
            cache_key = self.account_mapping.get(group_name)
            if not cache_key:
                continue
                
            account_positions = positions_data.get(cache_key, [])
            # Skip yield filtering for surgical update - update ALL positions
            # filtered_positions = self.filter_high_yield_positions(account_positions, yields_data)
            
            print(f"\n{group_name}:")
            print(f"  Cache positions: {len(account_positions)}")
            
            if account_positions:
                updated = self.surgical_update_by_index(ws, account_info[group_name], account_positions)
                total_updated += updated
                print(f"  Updated: {updated} positions")
        
        print(f"\nTotal positions updated: {total_updated}")
        
        # Update yield column
        date_str = datetime.now().strftime("%m/%d/%Y")
        self.update_yield_data(ws, account_info, yields_data)
        
        # Apply formatting
        self.apply_group_formatting(ws, account_info)
        
        # Save and close
        wb.save(self.excel_file)
        wb.close()
        
        print("\nSUCCESS: Simple surgical update completed!")
        return True
    
    def surgical_update_by_index(self, ws, group_info, positions):
        """Update tickers by matching index position within the group"""
        start_row = group_info["start_row"] + 2  # Skip group header and column headers
        
        ticker_font = Font(name="Arial", size=12, bold=True, color="3072C2")
        qty_font = Font(name="Arial", size=12, bold=True, color="3072C2")
        
        updated_count = 0
        
        # Simply match positions by index order
        for i, position in enumerate(positions):
            row = start_row + i
            
            symbol = position.get('symbol', '').strip().upper()
            quantity = position.get('quantity', 0)
            market_value = position.get('market_value', 0)
            price = round(market_value / quantity, 2) if quantity > 0 else 0
            
            # Update ONLY columns A, B, D - preserve everything else
            ws.cell(row=row, column=1).value = symbol  # Ticker
            ws.cell(row=row, column=1).font = ticker_font
            
            ws.cell(row=row, column=2).value = quantity  # Quantity
            ws.cell(row=row, column=2).font = qty_font
            
            ws.cell(row=row, column=4).value = price  # Price
            
            updated_count += 1
            print(f"    Row {row}: {symbol} qty={quantity} price=${price:.2f}")
        
        return updated_count
    
    def find_account_groups(self, ws):
        """Find account group start positions"""
        account_info = {}
        
        print("Scanning for account groups...")
        for row in range(1, min(80, ws.max_row + 1)):
            cell_value = ws.cell(row=row, column=1).value
            if cell_value:
                cell_str = str(cell_value).strip()
                print(f"  Row {row}: '{cell_str}'")
                
                # Check for each account group
                for group_name in self.account_mapping.keys():
                    if group_name.upper() in cell_str.upper():
                        account_info[group_name] = {"start_row": row}
                        print(f"    --> Found {group_name}")
                        break
        
        return account_info
    
    def filter_high_yield_positions(self, positions, yields_data):
        """Filter for dividend-paying stocks with yield > 4%"""
        filtered = []
        print("    Checking yields:")
        for position in positions:
            symbol = position.get('symbol', '').strip().upper()
            yield_info = yields_data.get(symbol, {})
            has_dividend = yield_info.get('has_dividend', False)
            current_yield = yield_info.get('yield', 0.0)
            
            print(f"      {symbol}: yield={current_yield:.2f}% dividend={has_dividend}")
            
            if has_dividend and current_yield > 4.0:
                filtered.append(position)
                
        return filtered
    
    def update_yield_data(self, ws, account_info, yields_data):
        """Update yield data in Column P with proper colors"""
        for group_name, group_info in account_info.items():
            start_row = group_info["start_row"] + 2
            
            # Process up to 10 rows per group (reasonable limit)
            for row in range(start_row, start_row + 10):
                ticker_cell = ws.cell(row=row, column=1)
                if ticker_cell.value:
                    symbol = str(ticker_cell.value).strip().upper()
                    yield_info = yields_data.get(symbol, {})
                    current_yield = yield_info.get('yield', 0.0)
                    
                    if current_yield > 0:
                        yield_cell = ws.cell(row=row, column=16)  # Column P
                        yield_cell.value = current_yield / 100.0  # Convert to decimal
                        yield_cell.number_format = '0.00%'
                        
                        # Compare with Column O for background color
                        previous_cell = ws.cell(row=row, column=15)  # Column O
                        previous_yield = 0.0
                        
                        if previous_cell.value:
                            try:
                                if isinstance(previous_cell.value, str):
                                    prev_str = str(previous_cell.value).replace('%', '').strip()
                                    previous_yield = float(prev_str)
                                elif isinstance(previous_cell.value, (int, float)):
                                    prev_val = float(previous_cell.value)
                                    previous_yield = prev_val * 100 if prev_val < 1.0 else prev_val
                            except (ValueError, TypeError):
                                previous_yield = 0.0
                        
                        # Apply background colors per plan
                        if previous_yield == 0.0:
                            yield_cell.fill = PatternFill()  # No background
                        elif current_yield > previous_yield:
                            yield_cell.fill = PatternFill(start_color="90EE90", end_color="90EE90", fill_type="solid")  # Green
                        elif current_yield < previous_yield:
                            yield_cell.fill = PatternFill(start_color="FF7C80", end_color="FF7C80", fill_type="solid")  # Red  
                        else:
                            yield_cell.fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")  # Yellow
    
    def apply_group_formatting(self, ws, account_info):
        """Apply orange formatting to group headers"""
        orange_fill = PatternFill(start_color="FFC000", end_color="FFC000", fill_type="solid")
        
        for group_name, group_info in account_info.items():
            divider_row = group_info["start_row"]
            # Orange headers in columns A and P
            ws.cell(row=divider_row, column=1).fill = orange_fill
            ws.cell(row=divider_row, column=16).fill = orange_fill  # Column P
    
    def load_cache_data(self):
        """Load cache data"""
        try:
            if not os.path.exists(self.cache_file):
                print(f"ERROR: Cache file not found: {self.cache_file}")
                return None
            
            with open(self.cache_file, 'r') as f:
                return json.load(f)
        except Exception as e:
            print(f"ERROR: Failed to load cache: {e}")
            return None

if __name__ == "__main__":
    updater = SimpleSurgicalUpdater()
    success = updater.run_update()
    
    if success:
        print("\nSUCCESS: Simple surgical update completed!")
    else:
        print("\nERROR: Update failed!")