import openpyxl
from openpyxl.styles import Font, PatternFill
import sys
import os

def examine_historical_yield_sheet():
    """Examine the structure of the Accounts Div historical yield sheet"""
    
    excel_file = r'C:\Users\mjmat\Python Code in VS\dividend_tracker\DividendTrackerApp\outputs\Dividends_2025.xlsx'
    
    if not os.path.exists(excel_file):
        print(f"❌ Excel file not found: {excel_file}")
        return
    
    try:
        workbook = openpyxl.load_workbook(excel_file)
        
        # Find the historical yield sheet
        historical_sheet = None
        for sheet_name in workbook.sheetnames:
            if "historical yield" in sheet_name.lower() or "accounts div" in sheet_name.lower():
                historical_sheet = workbook[sheet_name]
                print(f"📊 Found sheet: {sheet_name}")
                break
        
        if not historical_sheet:
            print("❌ Could not find historical yield sheet")
            print("Available sheets:", workbook.sheetnames)
            return
        
        print("\n🔍 SHEET STRUCTURE ANALYSIS")
        print("="*50)
        
        # Check first 10 rows for headers and structure
        print("\n📋 First 10 rows:")
        for row in range(1, 11):
            row_data = []
            for col in range(1, 8):  # Check columns A-G
                cell_value = historical_sheet.cell(row=row, column=col).value
                if cell_value:
                    row_data.append(f"Col {chr(64+col)}: {cell_value}")
            if row_data:
                print(f"Row {row:2d}: {' | '.join(row_data)}")
        
        # Look for account group headers
        print("\n🏢 ACCOUNT GROUP DETECTION:")
        account_groups = []
        for row in range(1, 60):  # Check first 60 rows
            cell_a = historical_sheet.cell(row=row, column=1).value
            if cell_a and isinstance(cell_a, str):
                if any(keyword in cell_a.lower() for keyword in ['etrade', 'schwab', 'ira', 'taxable', 'individual']):
                    account_groups.append((row, cell_a))
                    print(f"  Row {row:2d}: {cell_a}")
        
        # Check column headers
        print("\n📝 COLUMN HEADERS:")
        for col in range(1, 10):  # Check columns A-I
            header_cell = historical_sheet.cell(row=1, column=col).value
            if header_cell:
                print(f"  Column {chr(64+col)}: {header_cell}")
        
        # Look for ticker data in column A
        print("\n🎯 TICKER DATA SAMPLE (Column A):")
        ticker_count = 0
        for row in range(1, 60):
            cell_a = historical_sheet.cell(row=row, column=1).value
            if cell_a and isinstance(cell_a, str) and len(cell_a) <= 6 and cell_a.isupper():
                # Likely a ticker symbol
                qty = historical_sheet.cell(row=row, column=2).value
                last_price = historical_sheet.cell(row=row, column=4).value
                print(f"  Row {row:2d}: {cell_a} | Qty: {qty} | Last Price: {last_price}")
                ticker_count += 1
                if ticker_count >= 10:  # Limit output
                    break
        
        print(f"\n📊 Summary:")
        print(f"  - Found {len(account_groups)} potential account groups")
        print(f"  - Found {ticker_count}+ ticker entries")
        print(f"  - Sheet has {historical_sheet.max_row} rows, {historical_sheet.max_column} columns")
        
        workbook.close()
        
    except Exception as e:
        print(f"❌ Error examining sheet: {str(e)}")
        import traceback
        traceback.print_exc()

if __name__ == "__main__":
    examine_historical_yield_sheet()