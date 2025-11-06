#!/usr/bin/env python3

import openpyxl
from openpyxl.styles import PatternFill

def test_orange_fill():
    """Test the orange fill detection"""
    print("TESTING ORANGE FILL DETECTION")
    print("=" * 35)
    
    excel_path = r"C:\Users\mjmat\Python Code in VS\dividend_tracker\DividendTrackerApp\outputs\Dividends_2025.xlsx"
    wb = openpyxl.load_workbook(excel_path)
    ws = wb['Accounts Div historical yield']
    
    # Check row 1 (E*TRADE IRA)
    test_cells = [
        (1, 1, "E*TRADE IRA - Column A"),
        (1, 16, "E*TRADE IRA - Column P"),
        (23, 1, "E*TRADE Taxable - Column A"), 
        (23, 16, "E*TRADE Taxable - Column P")
    ]
    
    for row, col, desc in test_cells:
        cell = ws.cell(row=row, column=col)
        fill = cell.fill
        
        print(f"{desc}:")
        print(f"  Fill Type: {fill.fill_type}")
        print(f"  Start Color: {fill.start_color}")
        if hasattr(fill.start_color, 'rgb'):
            print(f"  RGB: {fill.start_color.rgb}")
        print(f"  Value: '{cell.value}'")
        print()
    
    # Test creating our orange fill
    orange_fill = PatternFill(start_color="FFA500", end_color="FFA500", fill_type="solid")
    print("Our orange fill:")
    print(f"  Start Color: {orange_fill.start_color}")
    print(f"  RGB: {orange_fill.start_color.rgb}")
    
    wb.close()

if __name__ == "__main__":
    test_orange_fill()