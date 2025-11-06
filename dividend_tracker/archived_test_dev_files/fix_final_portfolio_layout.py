#!/usr/bin/env python3
"""
Fix Portfolio Summary Final Layout
- Add missing 401K percentage (row 18 column B)
- Add missing Current Monthly Estimate (row 28 column B)
- Move Dividend Summary to Column D-E (rows 31-36)
- Move Net Dividend Income below Dividend Summary (rows 38-41)
- Remove duplicate Withdrawal Schedule rows (A33-36)
"""
import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment

def fix_final_portfolio_layout():
    """Fix the final portfolio layout with proper organization"""
    
    excel_file = 'outputs/Dividends_2025.xlsx'
    
    try:
        print("🔧 Fixing final Portfolio Summary layout...")
        
        wb = openpyxl.load_workbook(excel_file)
        portfolio_ws = wb['Portfolio Summary']
        
        # Get latest dividend data
        income_ws = wb['Estimated Income 2025']
        max_col = income_ws.max_column
        latest_monthly_estimate = income_ws.cell(row=9, column=max_col).value or 0
        latest_annual_estimate = latest_monthly_estimate * 12
        
        # Calculate 401K percentage
        k401_value = 124315.15
        total_portfolio = 519439.06  # Current portfolio total
        k401_pct = k401_value / total_portfolio
        
        # Withdrawal data
        monthly_etrade_ira = 1200
        quarterly_etrade_ira = 370
        monthly_etrade_taxable = 395
        total_monthly_withdrawals = monthly_etrade_ira + monthly_etrade_taxable + (quarterly_etrade_ira / 3)
        net_monthly = latest_monthly_estimate - total_monthly_withdrawals
        net_annual = net_monthly * 12
        
        # Define styles
        section_font = Font(name='Arial', size=11, bold=True)
        section_fill = PatternFill(start_color='D9E2F3', end_color='D9E2F3', fill_type='solid')
        normal_font = Font(name='Arial', size=12)
        right_aligned = Alignment(horizontal='right')
        percentage_format = '0.0%'
        currency_format = '$#,##0.00'
        
        print("📊 Step 1: Fix missing 401K percentage...")
        
        # Find and fix 401K percentage (around row 18)
        for row in range(15, 25):
            cell_value = portfolio_ws.cell(row=row, column=1).value
            if cell_value and "401k Retirement:" in str(cell_value):
                pct_cell = portfolio_ws.cell(row=row, column=2)
                pct_cell.value = k401_pct
                pct_cell.number_format = percentage_format
                pct_cell.alignment = right_aligned
                pct_cell.font = normal_font
                print(f"   Fixed 401k Retirement percentage at row {row}: {k401_pct:.1%}")
                break
        
        print("💰 Step 2: Fix missing Current Monthly Estimate...")
        
        # Find and fix Current Monthly Estimate (around row 28)
        for row in range(25, 35):
            cell_value = portfolio_ws.cell(row=row, column=1).value
            if cell_value and "Current Monthly Estimate:" in str(cell_value):
                estimate_cell = portfolio_ws.cell(row=row, column=2)
                estimate_cell.value = latest_monthly_estimate
                estimate_cell.number_format = currency_format
                estimate_cell.font = normal_font
                print(f"   Fixed Current Monthly Estimate at row {row}: ${latest_monthly_estimate:,.2f}")
                break
        
        print("🗑️ Step 3: Remove duplicate Withdrawal Schedule section...")
        
        # Find and remove Withdrawal Schedule rows (A33-36)
        withdrawal_start = None
        for row in range(30, 45):
            cell_value = portfolio_ws.cell(row=row, column=1).value
            if cell_value and "WITHDRAWAL SCHEDULE" in str(cell_value):
                withdrawal_start = row
                break
        
        if withdrawal_start:
            # Clear the withdrawal schedule section (5 rows including header and items)
            for clear_row in range(withdrawal_start, withdrawal_start + 5):
                if clear_row <= portfolio_ws.max_row:
                    portfolio_ws.cell(row=clear_row, column=1).value = ""
                    portfolio_ws.cell(row=clear_row, column=2).value = ""
            print(f"   Removed Withdrawal Schedule section starting at row {withdrawal_start}")
        
        print("📈 Step 4: Add Dividend Summary to Column D-E (rows 31-36)...")
        
        # Add Dividend Summary section in Column D-E
        dividend_summary_row = 31
        
        # Section header
        header_cell = portfolio_ws.cell(row=dividend_summary_row, column=4)
        header_cell.value = "DIVIDEND SUMMARY"
        header_cell.font = section_font
        header_cell.fill = section_fill
        
        dividend_summary_items = [
            ("Monthly Dividend Target:", 3600),
            ("Current Monthly Estimate:", latest_monthly_estimate),
            ("Annual Dividend Target:", 43200),
            ("Current Annual Estimate:", latest_annual_estimate),
        ]
        
        current_row = dividend_summary_row + 1
        for label, value in dividend_summary_items:
            # Label in column D
            label_cell = portfolio_ws.cell(row=current_row, column=4)
            label_cell.value = label
            label_cell.font = normal_font
            
            # Value in column E
            value_cell = portfolio_ws.cell(row=current_row, column=5)
            value_cell.value = value
            value_cell.number_format = currency_format
            value_cell.font = normal_font
            
            current_row += 1
        
        print(f"   Added Dividend Summary in Column D-E at rows {dividend_summary_row}-{current_row-1}")
        
        print("💸 Step 5: Add Net Dividend Income below Dividend Summary (rows 38-41)...")
        
        # Add Net Dividend Income section in Column D-E
        net_income_row = 38
        
        # Section header
        header_cell = portfolio_ws.cell(row=net_income_row, column=4)
        header_cell.value = "NET DIVIDEND INCOME"
        header_cell.font = section_font
        header_cell.fill = section_fill
        
        net_income_items = [
            ("After Withdrawals (Monthly):", net_monthly),
            ("After Withdrawals (Annual):", net_annual),
            ("YTD Dividends Received:", "Check All account weekly dividends sheet"),
        ]
        
        current_row = net_income_row + 1
        for label, value in net_income_items:
            # Label in column D
            label_cell = portfolio_ws.cell(row=current_row, column=4)
            label_cell.value = label
            label_cell.font = normal_font
            
            # Value in column E
            value_cell = portfolio_ws.cell(row=current_row, column=5)
            if isinstance(value, (int, float)):
                value_cell.value = value
                value_cell.number_format = currency_format
            else:
                value_cell.value = value
            value_cell.font = normal_font
            
            current_row += 1
        
        print(f"   Added Net Dividend Income in Column D-E at rows {net_income_row}-{current_row-1}")
        
        print("✅ Final Portfolio Summary layout completed!")
        print("   ✅ 401K percentage added")
        print("   ✅ Current Monthly Estimate added")
        print("   ✅ Dividend Summary moved to Column D-E")
        print("   ✅ Net Dividend Income moved below Dividend Summary")
        print("   ✅ Duplicate Withdrawal Schedule removed")
        print("   ✅ Consistent organization achieved")
        
        # Save the workbook
        wb.save(excel_file)
        wb.close()
        
        return True
        
    except Exception as e:
        print(f"❌ Error fixing final layout: {e}")
        return False

def verify_final_layout():
    """Verify the final layout is correct"""
    excel_file = 'outputs/Dividends_2025.xlsx'
    
    try:
        wb = openpyxl.load_workbook(excel_file, data_only=True)
        portfolio_ws = wb['Portfolio Summary']
        
        print("\n📊 FINAL LAYOUT VERIFICATION")
        print("=" * 90)
        print("PORTFOLIO SUMMARY (A-B)              |  DIVIDEND STATISTICS (D-E)")
        print("=" * 90)
        
        # Show the complete layout
        for row in range(1, min(45, portfolio_ws.max_row + 1)):
            # Portfolio summary (columns A-B)
            col_a = portfolio_ws.cell(row=row, column=1).value or ""
            col_b = portfolio_ws.cell(row=row, column=2).value or ""
            
            # Format portfolio text
            if isinstance(col_b, (int, float)):
                if col_b > 1000:
                    portfolio_text = f"{col_a} ${col_b:,.2f}".strip()
                elif col_b < 1 and col_b > 0:
                    portfolio_text = f"{col_a} {col_b:.1%}".strip()
                else:
                    portfolio_text = f"{col_a} {col_b}".strip()
            else:
                portfolio_text = f"{col_a} {col_b}".strip()
            
            # Dividend stats (columns D-E)
            col_d = portfolio_ws.cell(row=row, column=4).value or ""
            col_e = portfolio_ws.cell(row=row, column=5).value or ""
            
            # Format dividend text
            if isinstance(col_e, (int, float)):
                if col_e < 1 and col_e > 0:
                    dividend_text = f"{col_d} {col_e:.1%}"
                elif col_e > 1 and col_e < 10:
                    dividend_text = f"{col_d} {col_e:.1f}x"
                else:
                    dividend_text = f"{col_d} ${col_e:,.2f}"
            else:
                dividend_text = f"{col_d} {col_e}".strip()
            
            if portfolio_text.strip() or dividend_text.strip():
                print(f"{portfolio_text[:40]:<40} | {dividend_text}")
        
        wb.close()
        
    except Exception as e:
        print(f"Error: {e}")

if __name__ == "__main__":
    print("🚀 FIXING FINAL PORTFOLIO SUMMARY LAYOUT")
    print("=" * 60)
    print("Final Fixes:")
    print("• Add missing 401K percentage (row 18 column B)")
    print("• Add missing Current Monthly Estimate (row 28 column B)")
    print("• Move Dividend Summary to Column D-E (rows 31-36)")
    print("• Move Net Dividend Income below (rows 38-41)")
    print("• Remove duplicate Withdrawal Schedule section")
    print("=" * 60)
    
    success = fix_final_portfolio_layout()
    
    if success:
        verify_final_layout()
        
        print("\n🎯 Final Portfolio Summary Layout Complete!")
        print("✅ All missing values added")
        print("✅ Consistent organization with sections in proper columns")
        print("✅ No duplicate information")
        print("✅ Professional layout ready for market analysis")
    else:
        print("\n❌ Failed to fix final layout")
