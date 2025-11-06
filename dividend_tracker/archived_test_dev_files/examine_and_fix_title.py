#!/usr/bin/env python3
"""
Examine Enhanced Sheet Layout and Fix Row 50 Title
=================================================
"""

import openpyxl
import os

def examine_and_fix_sheet():
    """Examine the enhanced sheet structure and fix row 50 title"""
    
    # Path to the Excel file
    excel_path = r"c:\Users\mjmat\Python Code in VS\dividend_tracker\DividendTrackerApp\outputs\Dividends_2025.xlsx"
    
    if not os.path.exists(excel_path):
        print(f"❌ File not found: {excel_path}")
        return
    
    try:
        # Open workbook
        workbook = openpyxl.load_workbook(excel_path)
        sheet_names = workbook.sheetnames
        print(f"📋 Available sheets: {sheet_names}")
        
        # Check if sheet name was changed
        target_sheet = None
        if "Accounts dividend historical yield" in sheet_names:
            target_sheet = workbook["Accounts dividend historical yield"]
            print(f"✅ Found renamed sheet: 'Accounts dividend historical yield'")
        elif "Etrade IRA historic yield" in sheet_names:
            target_sheet = workbook["Etrade IRA historic yield"]
            print(f"✅ Found original sheet: 'Etrade IRA historic yield'")
        else:
            print(f"❌ Could not find target sheet")
            return
        
        # Examine sheet structure
        print(f"\n📊 Sheet Structure Analysis:")
        print(f"📏 Dimensions: {target_sheet.max_row} rows × {target_sheet.max_column} columns")
        
        # Look at key rows to understand the layout
        key_rows = [1, 25, 27, 42, 43, 50, 66, 67]
        for row_num in key_rows:
            if row_num <= target_sheet.max_row:
                cell_a = target_sheet.cell(row=row_num, column=1).value
                cell_b = target_sheet.cell(row=row_num, column=2).value
                print(f"Row {row_num:2d}: A={cell_a} | B={cell_b}")
        
        # Fix row 50 title
        print(f"\n🔧 Fixing Row 50 Title...")
        current_title = target_sheet.cell(row=50, column=1).value
        print(f"Current title in row 50: '{current_title}'")
        
        # Update the title
        target_sheet.cell(row=50, column=1).value = "Schwab Individual"
        new_title = target_sheet.cell(row=50, column=1).value
        print(f"Updated title in row 50: '{new_title}'")
        
        # Save the changes
        workbook.save(excel_path)
        workbook.close()
        
        print(f"\n✅ Successfully updated row 50 title to 'Schwab Individual'")
        print(f"💾 File saved: {excel_path}")
        
    except Exception as e:
        print(f"❌ Error: {e}")

if __name__ == "__main__":
    examine_and_fix_sheet()
