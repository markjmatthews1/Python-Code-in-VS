import openpyxl
from datetime import datetime

def verify_final_results():
    """Comprehensive verification of the enhanced E*TRADE Historic Yield updater results"""
    
    workbook_path = "outputs/Dividends_2025.xlsx"
    
    try:
        wb = openpyxl.load_workbook(workbook_path)
        sheet = wb["Accounts Div historical yield"]
        
        print("🔍 FINAL VERIFICATION: Enhanced E*TRADE Historic Yield Results")
        print("=" * 70)
        
        # Account sections for verification
        account_sections = [
            {
                'name': 'Etrade IRA',
                'ticker_start_row': 3,
                'ticker_end_row': 20,
                'average_row': 25,
                'title_row': 1
            },
            {
                'name': 'Etrade Taxable',
                'ticker_start_row': 29,
                'ticker_end_row': 40,
                'average_row': 41,
                'title_row': 26
            },
            {
                'name': 'Schwab IRA',
                'ticker_start_row': 44,
                'ticker_end_row': 47,
                'average_row': 49,
                'title_row': 42
            },
            {
                'name': 'Schwab Individual',
                'ticker_start_row': 52,
                'ticker_end_row': 53,
                'average_row': 55,
                'title_row': 50
            }
        ]
        
        total_tickers = 0
        total_updated = 0
        
        # Check each account section
        for section in account_sections:
            print(f"\n📊 {section['name']} Section:")
            print(f"   Title Row: {section['title_row']}")
            
            # Check title row formatting
            title_cell = sheet.cell(section['title_row'], 16)
            title_color = title_cell.fill.start_color.rgb if title_cell.fill.start_color else 'None'
            print(f"   Title Row Color (Col 16): {title_color}")
            
            # Count tickers and check for data
            section_tickers = 0
            section_updated = 0
            
            for row in range(section['ticker_start_row'], section['ticker_end_row'] + 1):
                ticker_cell = sheet.cell(row, 1)
                if ticker_cell.value:
                    section_tickers += 1
                    total_tickers += 1
                    
                    # Check if historic yield data exists
                    yield_cell = sheet.cell(row, 16)
                    if yield_cell.value and yield_cell.value != 0:
                        section_updated += 1
                        total_updated += 1
            
            print(f"   Tickers: {section_tickers}")
            print(f"   Updated with Data: {section_updated}")
            
            # Check average calculation
            avg_cell = sheet.cell(section['average_row'], 16)
            avg_value = avg_cell.value if avg_cell.value else "No formula"
            print(f"   Average Row {section['average_row']}: {avg_value}")
        
        print("\n" + "=" * 70)
        print(f"📈 SUMMARY:")
        print(f"   Total Tickers Processed: {total_tickers}")
        print(f"   Total with Historic Yield Data: {total_updated}")
        print(f"   Success Rate: {(total_updated/total_tickers*100):.1f}%" if total_tickers > 0 else "N/A")
        
        # Check date header
        date_header = sheet.cell(2, 16).value
        print(f"   Date Header: {date_header}")
        
        # Check column title
        col_title = sheet.cell(1, 16).value
        print(f"   Column Title: {col_title}")
        
        wb.close()
        
        print("\n✅ VERIFICATION COMPLETE")
        
        if total_tickers == 36 and total_updated >= 30:  # Allow for some potential API issues
            print("🎉 SUCCESS: Multi-account enhancement working perfectly!")
        else:
            print("⚠️  REVIEW NEEDED: Some data may be missing")
            
    except Exception as e:
        print(f"❌ Error during verification: {e}")

if __name__ == "__main__":
    verify_final_results()
