import os
import openpyxl

def check_portfolio_sheet():
    """Check the Portfolio Values 2025 sheet structure"""
    excel_file = "outputs/Dividends_2025.xlsx"
    
    if not os.path.exists(excel_file):
        print(f"ERROR: Excel file not found: {excel_file}")
        return
    
    try:
        wb = openpyxl.load_workbook(excel_file)
        sheet_name = "Portfolio Values 2025"
        
        if sheet_name not in wb.sheetnames:
            print(f"ERROR: Sheet '{sheet_name}' not found")
            print(f"Available sheets: {wb.sheetnames}")
            return
            
        ws = wb[sheet_name]
        
        print("PORTFOLIO VALUES 2025 SHEET ANALYSIS")
        print("=" * 50)
        
        # Check rows 1-15 to understand structure
        for row in range(1, 16):
            col_a = ws.cell(row=row, column=1).value
            # Get last column with data for current values
            last_col = 1
            for col in range(1, ws.max_column + 1):
                if ws.cell(row=row, column=col).value is not None:
                    last_col = col
            
            current_value = ws.cell(row=row, column=last_col).value if last_col > 1 else None
            
            print(f"Row {row:2d}: '{col_a}' | Current Value: {current_value}")
        
        # Check the sum formula in the total row
        for row in range(8, 12):
            col_a = ws.cell(row=row, column=1).value
            if col_a and 'total' in str(col_a).lower():
                for col in range(2, ws.max_column + 1):
                    cell = ws.cell(row=row, column=col)
                    if hasattr(cell, 'data_type') and cell.data_type == 'f':
                        print(f"\nFORMULA in Row {row}, Col {col}: {cell.value}")
                        break
        
        wb.close()
        
    except Exception as e:
        print(f"ERROR: {e}")
        import traceback
        traceback.print_exc()

if __name__ == "__main__":
    check_portfolio_sheet()