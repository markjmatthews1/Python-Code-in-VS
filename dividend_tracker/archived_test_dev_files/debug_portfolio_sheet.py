#!/usr/bin/env python3
"""
Debug Portfolio Values Sheet Structure
=====================================

Examine the current Portfolio Values 2025 sheet to understand why
the focused updater didn't add the new column.
"""

import os
import openpyxl
from datetime import datetime

# Configuration
OUTPUT_DIR = os.path.join(os.path.dirname(__file__), "outputs")
TARGET_FILE = os.path.join(OUTPUT_DIR, "Dividends_2025.xlsx")

def debug_portfolio_sheet():
    """Debug the Portfolio Values 2025 sheet structure"""
    
    try:
        print("🔍 DEBUGGING PORTFOLIO VALUES 2025 SHEET")
        print("=" * 50)
        
        # Load workbook
        wb = openpyxl.load_workbook(TARGET_FILE)
        
        if "Portfolio Values 2025" not in wb.sheetnames:
            print("❌ Portfolio Values 2025 sheet not found!")
            print("Available sheets:")
            for sheet in wb.sheetnames:
                print(f"  - {sheet}")
            return
            
        ws = wb["Portfolio Values 2025"]
        
        print(f"📋 Sheet found: Portfolio Values 2025")
        print(f"   Max Row: {ws.max_row}")
        print(f"   Max Column: {ws.max_column}")
        
        # Check header row (row 1) to see all dates
        print(f"\n📅 HEADER ROW (Dates):")
        for col in range(1, ws.max_column + 1):
            cell_value = ws.cell(row=1, column=col).value
            if cell_value:
                print(f"   Column {col}: {cell_value}")
            else:
                print(f"   Column {col}: [EMPTY]")
                break  # Stop at first empty column
        
        # Check account names in column A
        print(f"\n📊 ACCOUNT NAMES (Column A):")
        for row in range(1, ws.max_row + 1):
            cell_value = ws.cell(row=row, column=1).value
            if cell_value:
                print(f"   Row {row}: {cell_value}")
        
        # Find next available column using our logic
        next_col = 2  # Start from column B
        while ws.cell(row=1, column=next_col).value is not None:
            next_col += 1
        
        print(f"\n🔍 COLUMN ANALYSIS:")
        print(f"   Next available column should be: {next_col}")
        print(f"   Last column with data: {next_col - 1}")
        
        # Check a few specific columns around the expected area
        print(f"\n🔍 COLUMN CONTENT CHECK:")
        for col in range(max(1, next_col - 3), next_col + 3):
            header_val = ws.cell(row=1, column=col).value
            print(f"   Column {col}: Header = '{header_val}'")
        
        print(f"\n📊 SAMPLE DATA (first few accounts, last few columns):")
        for row in range(2, min(7, ws.max_row + 1)):  # First 5 accounts
            account_name = ws.cell(row=row, column=1).value
            if account_name:
                print(f"   {account_name}:")
                for col in range(max(2, ws.max_column - 2), ws.max_column + 3):
                    cell_val = ws.cell(row=row, column=col).value
                    header_val = ws.cell(row=1, column=col).value
                    if header_val:
                        print(f"     Column {col} ({header_val}): {cell_val}")
        
    except Exception as e:
        print(f"❌ Error debugging sheet: {e}")
        import traceback
        traceback.print_exc()

if __name__ == "__main__":
    debug_portfolio_sheet()
    input("\nPress Enter to close...")
