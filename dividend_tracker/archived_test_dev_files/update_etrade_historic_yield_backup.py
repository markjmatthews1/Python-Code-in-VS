#!/usr/bin/env python3
"""
E*TRADE IRA Historic Yield Updater for DividendTracker App
==========================================================

This module is a DividendTracker-integrated version of Update_dividend_sheet.py
Updates the "Etrade IRA historic yield" sheet in Dividends_2025.xlsx with current yield data.

Original Reference: Update_dividend_sheet.py (unchanged)
Integration: Part of weekend DividendTracker automation
Target: Dividends_2025.xlsx -> "Etrade IRA historic yield" sheet

Author: DividendTracker Integration
Date: August 30, 2025
"""

import os
import sys
import json
import time
from datetime import datetime, date
import openpyxl
from openpyxl.styles import PatternFill, Font, Border, Side, Alignment
from openpyxl.utils import get_column_letter

# Add modules directory to path for DividendTracker imports
sys.path.append(os.path.join(os.path.dirname(__file__), 'modules'))

try:
    # Import DividendTracker E*TRADE authentication (from modules directory)
    from etrade_auth import get_etrade_session
except ImportError as e:
    print(f"❌ ERROR: Could not import DividendTracker E*TRADE authentication: {e}")
    print("Ensure etrade_auth.py is in the modules/ directory")
    sys.exit(1)

class ETradeHistoricYieldUpdater:
    """Updates the Accounts dividend historical yield sheet with current yield data from E*TRADE API"""
    
    def __init__(self):
        self.workbook_path = os.path.join(os.path.dirname(__file__), "outputs", "Dividends_2025.xlsx")
        self.sheet_name = "Accounts dividend historical yield"
        self.session = None
        self.base_url = None
        
    def initialize_etrade_auth(self):
        """Initialize E*TRADE authentication using DividendTracker auth system"""
        try:
            print("🔐 Initializing E*TRADE authentication...")
            self.session, self.base_url = get_etrade_session()
            
            if self.session and self.base_url:
                print("✅ E*TRADE authentication successful")
                return True
            else:
                print("❌ E*TRADE authentication failed")
                return False
                
        except Exception as e:
            print(f"❌ E*TRADE authentication error: {e}")
            return False
    
    def get_quote_data(self, symbol):
        """Get quote data for a symbol including yield information"""
        try:
            if not self.session or not self.base_url:
                print("❌ E*TRADE session not initialized")
                return None
                
            quote_url = f"{self.base_url}/v1/market/quote/{symbol}.json"
            response = self.session.get(quote_url)
            
            if response.status_code == 401:
                print(f"⚠️ 401 Unauthorized for {symbol}. Refreshing session...")
                # Try to refresh session
                self.session, self.base_url = get_etrade_session()
                if self.session:
                    response = self.session.get(quote_url)
                else:
                    print(f"❌ Could not refresh session for {symbol}")
                    return None
            
            if response.status_code == 200:
                quote_json = response.json()
                
                if ('QuoteResponse' in quote_json and 
                    'QuoteData' in quote_json['QuoteResponse'] and 
                    isinstance(quote_json['QuoteResponse']['QuoteData'], list) and
                    len(quote_json['QuoteResponse']['QuoteData']) > 0):
                    
                    quote_data = quote_json['QuoteResponse']['QuoteData'][0]
                    all_data = quote_data.get('All', {})
                    
                    return {
                        'symbol': symbol,
                        'lastTrade': all_data.get('lastTrade', 0),
                        'yield': all_data.get('yield', 0),
                        'dividendAmount': all_data.get('dividendAmount', 0),
                        'dividendYield': all_data.get('dividendYield', 0),
                        'dividendPayableDate': all_data.get('dividendPayableDate', '')
                    }
                else:
                    print(f"⚠️ Unexpected quote response format for {symbol}")
                    return None
                    
            else:
                print(f"⚠️ Quote API error for {symbol}: HTTP {response.status_code}")
                return None
                
        except Exception as e:
            print(f"❌ Error getting quote for {symbol}: {e}")
            return None
    
    def update_historic_yield_sheet(self):
        """Main function to update the Etrade IRA historic yield sheet"""
        try:
            print("🚀 STARTING E*TRADE IRA HISTORIC YIELD UPDATE")
            print("=" * 60)
            print(f"🕐 Started: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}")
            print(f"📁 Target workbook: {self.workbook_path}")
            print(f"📊 Target sheet: {self.sheet_name}")
            
            # Initialize E*TRADE authentication
            if not self.initialize_etrade_auth():
                return False
                
            # Load the workbook
            if not os.path.exists(self.workbook_path):
                print(f"❌ Workbook not found: {self.workbook_path}")
                return False
                
            print("📖 Loading workbook...")
            workbook = openpyxl.load_workbook(self.workbook_path)
            
            # Check if sheet exists
            if self.sheet_name not in workbook.sheetnames:
                print(f"❌ Sheet '{self.sheet_name}' not found in workbook")
                print(f"Available sheets: {workbook.sheetnames}")
                return False
                
            sheet = workbook[self.sheet_name]
            print(f"✅ Successfully opened sheet: {self.sheet_name}")
            print(f"📏 Sheet dimensions: {sheet.max_row} rows × {sheet.max_column} columns")
            
            # Insert new column right after column O (column 15 - "Beginning Dividend Yield")
            # Column 15 = "Beginning Dividend Yield", so insert at column 16
            insert_col = 16
            
            print(f"📅 Inserting new yield column after 'Beginning Dividend Yield' at column {insert_col}")
            
            # Insert the new column in the sheet
            sheet.insert_cols(insert_col)
            
            # Add header for new date in the inserted column
            today = date.today().strftime("%m-%d-%Y")
            header_cell = sheet.cell(row=1, column=insert_col, value=today)
            header_cell.font = Font(bold=True, color="FFFFFF")
            header_cell.fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
            header_cell.alignment = Alignment(horizontal="center")
            
            print(f"📅 Added new yield column header: {today} (Column {insert_col})")
            
            # Process each ticker row
            updated_count = 0
            error_count = 0
            
            for row_idx in range(2, sheet.max_row + 1):
                ticker_cell = sheet.cell(row=row_idx, column=1)  # Column A = Ticker
                ticker = ticker_cell.value
                
                if not ticker or not isinstance(ticker, str) or ticker.strip() == "":
                    break  # End of data
                    
                ticker = ticker.strip().upper()
                print(f"   🔍 Processing {ticker}...", end=" ")
                
                # Get quote data for this ticker
                quote_data = self.get_quote_data(ticker)
                
                if quote_data:
                    # Get current yield from quote
                    current_yield = quote_data.get('yield', 0)
                    
                    if current_yield and current_yield != 0:
                        # Set yield value in the new inserted column
                        yield_cell = sheet.cell(row=row_idx, column=insert_col)
                        # Divide by 100 since API returns 10.15 for 10.15% but Excel percentage format will multiply by 100
                        yield_cell.value = round(float(current_yield) / 100, 4)
                        
                        # Format the cell as percentage with 2 decimal places
                        yield_cell.number_format = '0.00%'
                        
                        # Apply color coding by comparing against "Beginning Dividend Yield" (column 15/O)
                        beginning_yield_cell = sheet.cell(row=row_idx, column=15)  # Column O
                        beginning_yield = beginning_yield_cell.value
                        
                        try:
                            if beginning_yield and float(beginning_yield) != 0:
                                current_val = float(current_yield) / 100  # Divide by 100 to match our stored format
                                beginning_val = float(beginning_yield) / 100  # Assuming beginning yield is also in percentage format
                                
                                if current_val > beginning_val:
                                    # Green for increase from beginning yield
                                    fill_color = "00FF00"
                                elif current_val < beginning_val:
                                    # Red for decrease from beginning yield
                                    fill_color = "FF0000"
                                else:
                                    # Yellow for same as beginning yield
                                    fill_color = "FFFF00"
                                    
                                fill = PatternFill(start_color=fill_color, end_color=fill_color, fill_type="solid")
                                yield_cell.fill = fill
                                
                        except (ValueError, TypeError):
                            pass  # Skip color coding if comparison fails
                        
                        print(f"✅ {current_yield}% (vs {beginning_yield}% beginning)")
                        updated_count += 1
                        
                    else:
                        yield_cell = sheet.cell(row=row_idx, column=insert_col)
                        yield_cell.value = ""
                        print("⚠️ No yield data")
                        
                else:
                    print("❌ Quote failed")
                    error_count += 1
                    
                # Small delay to avoid overwhelming the API
                time.sleep(0.5)
            
            # Add average calculation in row 25 for the new column
            try:
                print(f"\n📊 Adding average calculation...")
                avg_cell = sheet.cell(row=25, column=insert_col)
                # Calculate average of P2:P23 (values are already divided by 100)
                avg_cell.value = f"=ROUND(AVERAGE({get_column_letter(insert_col)}2:{get_column_letter(insert_col)}23),4)"
                
                # Format the average cell as percentage with 2 decimal places
                avg_cell.number_format = '0.00%'
                
                # Apply same color coding logic by comparing average against beginning yield average
                beginning_avg_cell = sheet.cell(row=25, column=15)  # Column O row 25
                beginning_avg_val = beginning_avg_cell.value
                
                # Check if beginning average is a formula or value
                if beginning_avg_val and isinstance(beginning_avg_val, str) and beginning_avg_val.startswith('='):
                    # It's a formula, skip color comparison for now
                    print(f"✅ Added average formula (beginning column has formula, skipping color coding)")
                elif beginning_avg_val and isinstance(beginning_avg_val, (int, float)):
                    # It's a numeric value, we can compare
                    # The average formula will calculate when Excel opens, 
                    # for now we'll manually calculate for color coding
                    yields_sum = 0
                    yields_count = 0
                    
                    for r in range(2, 24):  # rows 2-23
                        cell_val = sheet.cell(row=r, column=insert_col).value
                        if cell_val and isinstance(cell_val, (int, float)):
                            yields_sum += float(cell_val)
                            yields_count += 1
                    
                    if yields_count > 0:
                        current_avg = yields_sum / yields_count  # Values are already divided by 100
                        beginning_avg = float(beginning_avg_val) / 100  # Assuming beginning is in percentage format
                        
                        if beginning_avg > 0:
                            if current_avg > beginning_avg:
                                fill_color = "00FF00"  # Green
                            elif current_avg < beginning_avg:
                                fill_color = "FF0000"  # Red
                            else:
                                fill_color = "FFFF00"  # Yellow
                                
                            fill = PatternFill(start_color=fill_color, end_color=fill_color, fill_type="solid")
                            avg_cell.fill = fill
                            print(f"✅ Added average calculation with color coding ({current_avg*100:.2f}% vs {beginning_avg*100:.2f}% beginning)")
                        else:
                            print(f"✅ Added average calculation (beginning average is 0, no color coding)")
                    else:
                        print(f"✅ Added average calculation (no valid yields found for averaging)")
                else:
                    print(f"✅ Added average calculation (beginning column empty, no color coding)")
                            
            except Exception as e:
                print(f"⚠️ Warning: Could not add average calculation: {str(e)}")
            
            # Save the workbook
            print(f"\n💾 Saving workbook...")
            workbook.save(self.workbook_path)
            workbook.close()
            
            # Summary
            print("\n✅ E*TRADE IRA HISTORIC YIELD UPDATE COMPLETE")
            print("=" * 60)
            print(f"📊 Results Summary:")
            print(f"   ✅ Successfully updated: {updated_count} tickers")
            print(f"   ❌ Errors encountered: {error_count} tickers")
            print(f"   📅 New column inserted: {today} (Column {insert_col})")
            print(f"   💾 File saved: {self.workbook_path}")
            print(f"🕐 Completed: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}")
            
            return True
            
        except Exception as e:
            print(f"❌ Critical error during historic yield update: {e}")
            try:
                if 'workbook' in locals():
                    workbook.save(self.workbook_path)
                    workbook.close()
                    print("💾 Emergency save completed")
            except:
                pass
            return False

def run_historic_yield_update():
    """Main entry point for historic yield update"""
    updater = ETradeHistoricYieldUpdater()
    return updater.update_historic_yield_sheet()

if __name__ == "__main__":
    """Direct execution for testing"""
    print("🧪 TESTING: E*TRADE Historic Yield Updater")
    success = run_historic_yield_update()
    
    if success:
        print("\n🎉 Historic yield update completed successfully!")
    else:
        print("\n💥 Historic yield update failed!")
        
    print("\n⚠️ NOTE: This module is designed to be called by the main DividendTracker weekend automation.")
    print("For production use, integrate this into your weekend update script.")
