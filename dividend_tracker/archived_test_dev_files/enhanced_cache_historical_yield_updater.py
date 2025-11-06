#!/usr/bin/env python3
"""
Enhanced Cache-Based Historical Yield Updater
==============================================

Updates the "Accounts Div historical y        print(f"\n📝 Updating position data for {account_group}...")
        
        # Get current positions from cache
        account_positions = self.get_positions_for_account(positions_data, account_key)" sheet with:
1. Current dividend yield data from portfolio_data_cache.json
2. Ticker symbols in column A
3. Quantities in column B  
4. Last prices in column D

Features:
- Reads position and yield data from cache (no API calls)
- Updates tickers, quantities, and prices from current positions
- Handles dynamic ticker counts per account
- Inserts new yield column after column O (column 15)
- Applies color coding: Green (increase), Red (decrease), Yellow (same), Orange (dividers)
- Processes 4 account groups: E*TRADE IRA, E*TRADE Taxable, Schwab IRA, Schwab Individual
"""

import os
import sys
import json
import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment
from openpyxl.styles.numbers import FORMAT_PERCENTAGE, FORMAT_CURRENCY_USD_SIMPLE
from datetime import datetime
import traceback

class EnhancedCacheHistoricalYieldUpdater:
    """Enhanced updater for historical yield sheet with position data"""
    
    def __init__(self):
        self.script_dir = os.path.dirname(os.path.abspath(__file__))
        self.outputs_dir = os.path.join(self.script_dir, "outputs")
        self.excel_file = os.path.join(self.outputs_dir, "Dividends_2025.xlsx")
        self.cache_file = os.path.join(self.script_dir, "portfolio_data_cache.json")
        self.today = datetime.now()
        self.today_str = self.today.strftime("%m/%d/%Y")
        
        # Account group mapping
        self.account_groups = {
            "E*TRADE IRA": "etrade_ira",
            "E*TRADE Taxable": "etrade_taxable", 
            "Schwab IRA": "schwab_ira",
            "Schwab Individual": "schwab_individual"
        }
        
        # Define colors
        self.colors = {
            'green': PatternFill(start_color='90EE90', end_color='90EE90', fill_type='solid'),
            'red': PatternFill(start_color='FF7C80', end_color='FF7C80', fill_type='solid'),
            'yellow': PatternFill(start_color='FFFF00', end_color='FFFF00', fill_type='solid'),
            'orange': PatternFill(start_color='FFC000', end_color='FFC000', fill_type='solid')
        }
        
    def load_cache_data(self):
        """Load position and yield data from cache"""
        try:
            if not os.path.exists(self.cache_file):
                print(f"❌ Cache file not found: {self.cache_file}")
                return None, None
                
            with open(self.cache_file, 'r') as f:
                cache_data = json.load(f)
                
            print(f"✅ Cache loaded: {cache_data.get('timestamp', 'No timestamp')}")
            
            # Extract ticker yields and positions
            ticker_yields = cache_data.get('ticker_yields', {})
            positions = cache_data.get('positions', {})
            
            print(f"📊 Found yields for {len(ticker_yields)} tickers")
            print(f"📊 Found positions for {len(positions)} account groups")
            
            return ticker_yields, positions
            
        except Exception as e:
            print(f"❌ Error loading cache: {e}")
            traceback.print_exc()
            return None, None
    
    def find_account_groups(self, ws):
        """Find the starting rows for each account group"""
        account_rows = {}
        
        print("\n🔍 Scanning for account groups...")
        
        for row in range(1, min(60, ws.max_row + 1)):
            cell_value = ws.cell(row=row, column=1).value
            if not cell_value:
                continue
                
            cell_text = str(cell_value).strip().upper()
            
            # Look for account group headers
            if 'ETRADE' in cell_text and 'IRA' in cell_text and 'TAXABLE' not in cell_text:
                account_rows["E*TRADE IRA"] = row
                print(f"  ✅ E*TRADE IRA found at row {row}")
            elif 'ETRADE' in cell_text and 'TAXABLE' in cell_text:
                account_rows["E*TRADE Taxable"] = row
                print(f"  ✅ E*TRADE Taxable found at row {row}")
            elif 'SCHWAB' in cell_text and 'IRA' in cell_text:
                account_rows["Schwab IRA"] = row
                print(f"  ✅ Schwab IRA found at row {row}")
            elif 'SCHWAB' in cell_text and 'INDIVIDUAL' in cell_text:
                account_rows["Schwab Individual"] = row
                print(f"  ✅ Schwab Individual found at row {row}")
        
        return account_rows
    
    def get_current_tickers_for_group(self, ws, start_row):
        """Get current tickers listed in the sheet for an account group"""
        tickers = []
        row = start_row + 2  # Skip header rows
        
        while row <= ws.max_row:
            cell_value = ws.cell(row=row, column=1).value
            if not cell_value:
                break
            
            cell_text = str(cell_value).strip().upper()
            
            # Check if this is another group header or empty
            if any(keyword in cell_text for keyword in ['ETRADE', 'SCHWAB', 'TAXABLE', 'INDIVIDUAL', 'IRA']):
                break
                
            # Valid ticker symbol
            if cell_text and len(cell_text) <= 6:
                tickers.append(cell_text)
                
            row += 1
            
        return tickers
    
    def get_positions_for_account(self, positions_data, account_key):
        """Get positions for a specific account from cache data"""
        account_positions = positions_data.get(account_key, [])
        
        position_dict = {}
        for position in account_positions:
            symbol = position.get('symbol', '').strip().upper()
            quantity = position.get('quantity', 0)
            market_value = position.get('market_value', 0)
            
            # Calculate current price
            current_price = market_value / quantity if quantity > 0 else 0
            
            if symbol:
                position_dict[symbol] = {
                    'quantity': quantity,
                    'current_price': current_price,
                    'market_value': market_value
                }
        
        return position_dict
    
    def update_position_data(self, ws, account_group, start_row, positions_data, ticker_yields):
        """Update ONLY ticker symbols, quantities, and prices - preserve all other calculations"""
        account_key = self.account_groups.get(account_group)
        if not account_key:
            print(f"❌ Unknown account group: {account_group}")
            return
        
        print(f"\n📝 Updating position data for {account_group}...")
        
        # Get current positions from cache
        account_positions = self.get_positions_for_account(positions_data, account_key)
        
        if not account_positions:
            print(f"  ⚠️ No positions found for {account_group}")
            return
        
        print(f"  📊 Found {len(account_positions)} positions")
        
        # DON'T update headers - preserve existing ones
        # Start updating data rows only
        data_row = start_row + 2
        
        # Get current tickers in the sheet to match positions
        current_tickers = self.get_current_tickers_for_group(ws, start_row)
        
        # Update existing ticker rows with current data - PRESERVE OTHER COLUMNS
        for i, ticker in enumerate(current_tickers):
            current_row = data_row + i
            
            if ticker in account_positions:
                position_data = account_positions[ticker]
                
                # Column A: Update ticker only if different
                current_ticker = ws.cell(row=current_row, column=1).value
                if str(current_ticker).strip().upper() != ticker:
                    ws.cell(row=current_row, column=1, value=ticker)
                
                # Column B: Update quantity only - preserve existing formatting
                quantity_cell = ws.cell(row=current_row, column=2)
                old_format = quantity_cell.number_format
                quantity_cell.value = position_data['quantity']
                if old_format and old_format != 'General':
                    quantity_cell.number_format = old_format
                else:
                    quantity_cell.number_format = '#,##0'
                
                # Column D: Update last price only - preserve existing formatting  
                price_cell = ws.cell(row=current_row, column=4)
                old_format = price_cell.number_format
                price_cell.value = position_data['current_price']
                if old_format and old_format != 'General':
                    price_cell.number_format = old_format
                else:
                    price_cell.number_format = FORMAT_CURRENCY_USD_SIMPLE
                
                print(f"    Row {current_row}: {ticker} | Qty: {position_data['quantity']} | Price: ${position_data['current_price']:.2f}")
            else:
                print(f"    Row {current_row}: {ticker} - No position data found (keeping existing)")
        
        # If there are new positions not in the sheet, add them at the end
        existing_tickers = set(current_tickers)
        new_positions = {k: v for k, v in account_positions.items() if k not in existing_tickers}
        
        if new_positions:
            print(f"  📈 Adding {len(new_positions)} new positions:")
            next_row = data_row + len(current_tickers)
            
            for ticker, position_data in new_positions.items():
                # Column A: Ticker
                ws.cell(row=next_row, column=1, value=ticker)
                
                # Column B: Quantity
                quantity_cell = ws.cell(row=next_row, column=2, value=position_data['quantity'])
                quantity_cell.number_format = '#,##0'
                
                # Column D: Last Price (skip column C - preserve existing calculations)
                price_cell = ws.cell(row=next_row, column=4, value=position_data['current_price'])
                price_cell.number_format = FORMAT_CURRENCY_USD_SIMPLE
                
                print(f"    Row {next_row}: NEW {ticker} | Qty: {position_data['quantity']} | Price: ${position_data['current_price']:.2f}")
                next_row += 1
        
        updated_count = len([t for t in current_tickers if t in account_positions]) + len(new_positions)
        print(f"  ✅ Updated {updated_count} position entries (preserved all calculations)")
    
    def insert_new_yield_column(self, ws):
        """Insert new column after column O (15) for today's yields"""
        print(f"\n📝 Inserting new yield column after column O...")
        
        # Insert column after O (column 15)
        ws.insert_cols(16)  # Insert at column P (16)
        
        # Set header for new column
        new_col = 16
        ws.cell(row=1, column=new_col, value=self.today_str)
        ws.cell(row=1, column=new_col).font = Font(bold=True, name="Arial", size=12)
        ws.cell(row=1, column=new_col).alignment = Alignment(horizontal='center')
        
        print(f"  ✅ Inserted yield column P with date {self.today_str}")
        return new_col
    
    def update_yield_data_for_group(self, ws, account_group, start_row, new_col, ticker_yields):
        """Update yield data for an account group - preserve all other columns"""
        print(f"\n📊 Updating yields for {account_group} starting at row {start_row}...")
        
        # Add date header to group section only in the new yield column
        date_header_row = start_row + 1
        date_cell = ws.cell(row=date_header_row, column=new_col, value=self.today_str)
        date_cell.font = Font(bold=True, name="Arial", size=10)
        date_cell.alignment = Alignment(horizontal='center')
        
        # Get current tickers in this group
        current_tickers = self.get_current_tickers_for_group(ws, start_row)
        print(f"  📋 Found {len(current_tickers)} tickers: {current_tickers}")
        
        if not current_tickers:
            print(f"  ⚠️ No tickers found for {account_group}")
            return
        
        # Update yields ONLY in the new column - don't touch other columns
        data_start_row = start_row + 2
        updated_count = 0
        
        for i, ticker in enumerate(current_tickers):
            current_row = data_start_row + i
            
            if ticker in ticker_yields:
                yield_data = ticker_yields[ticker]
                # Cache uses 'yield' key, not 'dividend_yield'
                dividend_yield = yield_data.get('yield', 0)
                
                # Get previous yield for color coding (previous column)
                prev_col = new_col - 1
                prev_yield = ws.cell(row=current_row, column=prev_col).value
                prev_yield = float(prev_yield) if prev_yield and str(prev_yield).replace('.', '').replace('-', '').isdigit() else None
                
                # Set yield value ONLY in the new column
                yield_cell = ws.cell(row=current_row, column=new_col, value=dividend_yield/100)
                yield_cell.number_format = FORMAT_PERCENTAGE
                
                # Apply color coding
                if prev_yield is not None:
                    if dividend_yield > prev_yield * 100:
                        yield_cell.fill = self.colors['green']
                        color_desc = "📈 GREEN"
                    elif dividend_yield < prev_yield * 100:
                        yield_cell.fill = self.colors['red'] 
                        color_desc = "📉 RED"
                    else:
                        yield_cell.fill = self.colors['yellow']
                        color_desc = "➡️ YELLOW"
                else:
                    yield_cell.fill = PatternFill(fill_type=None)
                    color_desc = "⚪ No color"
                
                print(f"    Row {current_row}: {ticker} = {dividend_yield:.2f}% ({color_desc})")
                updated_count += 1
            else:
                print(f"    Row {current_row}: {ticker} - No yield data found")
        
        print(f"  ✅ Updated {updated_count} yields for {account_group} (preserved all other columns)")
    
    def apply_group_divider_color(self, ws, account_rows, new_col):
        """Apply orange color to group divider rows"""
        print(f"\n🎨 Applying orange color to group dividers...")
        
        for group_name, row_num in account_rows.items():
            divider_cell = ws.cell(row=row_num, column=new_col)
            divider_cell.fill = self.colors['orange']
            print(f"  🟠 {group_name} divider at row {row_num}")
    
    def run_update(self):
        """Run the complete enhanced update"""
        print("🚀 ENHANCED CACHE-BASED HISTORICAL YIELD UPDATER")
        print("=" * 55)
        print(f"📅 Update Date: {self.today_str}")
        print(f"📁 Excel File: {os.path.basename(self.excel_file)}")
        print(f"📋 Cache File: {os.path.basename(self.cache_file)}")
        
        # Step 1: Load cache data
        print(f"\n🔄 STEP 1: Loading cache data...")
        ticker_yields, positions_data = self.load_cache_data()
        if not ticker_yields or not positions_data:
            print("❌ Failed to load cache data")
            return False
        
        # Step 2: Load workbook
        print(f"\n📖 STEP 2: Loading Excel workbook...")
        try:
            workbook = openpyxl.load_workbook(self.excel_file)
        except Exception as e:
            print(f"❌ Error loading workbook: {e}")
            return False
        
        # Step 3: Find historical yield sheet
        historical_sheet = None
        for sheet_name in workbook.sheetnames:
            if "historical yield" in sheet_name.lower():
                historical_sheet = workbook[sheet_name]
                print(f"✅ Found sheet: {sheet_name}")
                break
        
        if not historical_sheet:
            print("❌ Could not find historical yield sheet")
            return False
        
        # Step 4: Find account groups
        print(f"\n🔍 STEP 3: Finding account groups...")
        account_rows = self.find_account_groups(historical_sheet)
        
        if len(account_rows) != 4:
            print(f"❌ Expected 4 account groups, found {len(account_rows)}")
            return False
        
        # Step 5: Update position data for each group
        print(f"\n📝 STEP 4: Updating position data...")
        for account_group, start_row in account_rows.items():
            self.update_position_data(historical_sheet, account_group, start_row, positions_data, ticker_yields)
        
        # Step 6: Insert new yield column  
        print(f"\n📊 STEP 5: Inserting new yield column...")
        new_col = self.insert_new_yield_column(historical_sheet)
        
        # Step 7: Update yields for each group
        print(f"\n📈 STEP 6: Updating yield data...")
        for account_group, start_row in account_rows.items():
            self.update_yield_data_for_group(historical_sheet, account_group, start_row, new_col, ticker_yields)
        
        # Step 8: Apply group divider colors
        print(f"\n🎨 STEP 7: Applying group divider colors...")
        self.apply_group_divider_color(historical_sheet, account_rows, new_col)
        
        # Step 9: Save workbook
        print(f"\n💾 STEP 8: Saving workbook...")
        try:
            workbook.save(self.excel_file)
            print(f"✅ Successfully saved: {self.excel_file}")
        except Exception as e:
            print(f"❌ Error saving workbook: {e}")
            return False
        finally:
            workbook.close()
        
        print(f"\n🎉 ENHANCED UPDATE COMPLETED SUCCESSFULLY!")
        print("📊 Updated: Tickers, Quantities, Prices, and Yields")
        print("🎨 Applied: Color coding and group dividers")
        return True

def main():
    """Main execution function"""
    try:
        updater = EnhancedCacheHistoricalYieldUpdater()
        success = updater.run_update()
        
        if success:
            print(f"\n✅ Enhanced historical yield update completed successfully!")
        else:
            print(f"\n❌ Enhanced historical yield update failed")
            
    except Exception as e:
        print(f"\n💥 Fatal error: {e}")
        traceback.print_exc()

if __name__ == "__main__":
    main()