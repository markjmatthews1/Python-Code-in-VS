#!/usr/bin/env python3
"""
Ultra-Detailed Portfolio Values Debugger
=========================================

This version shows EVERY step in detail to identify exactly
where the update process is failing.
"""

import os
import sys
import openpyxl
from datetime import datetime
import traceback

# Add modules to path
sys.path.append(os.path.join(os.path.dirname(__file__), 'modules'))

def get_k401_value_simple():
    """Simple 401K value prompt"""
    try:
        value_str = input("Enter current 401K value: $")
        return float(value_str.replace(',', '').replace('$', ''))
    except:
        return 125000.00

class UltraDebugPortfolioUpdater:
    """Shows every detail of the update process"""
    
    def __init__(self):
        self.script_dir = os.path.dirname(os.path.abspath(__file__))
        self.outputs_dir = os.path.join(self.script_dir, "outputs")
        self.target_file = os.path.join(self.outputs_dir, "Dividends_2025.xlsx")
        
        print(f"🔍 INITIALIZATION:")
        print(f"   Script dir: {self.script_dir}")
        print(f"   Outputs dir: {self.outputs_dir}")
        print(f"   Target file: {self.target_file}")
        print(f"   File exists: {os.path.exists(self.target_file)}")
        
    def ultra_debug_sheet_analysis(self):
        """Show every detail of the sheet structure"""
        try:
            print(f"\n🔍 ULTRA-DETAILED SHEET ANALYSIS")
            print("=" * 60)
            
            # Load workbook
            print(f"📂 Loading workbook...")
            wb = openpyxl.load_workbook(self.target_file)
            print(f"   ✅ Workbook loaded successfully")
            
            # Check sheet names
            print(f"📋 Available sheets:")
            for i, sheet_name in enumerate(wb.sheetnames):
                print(f"   {i+1}. {sheet_name}")
            
            if "Portfolio Values 2025" not in wb.sheetnames:
                print("❌ Portfolio Values 2025 sheet NOT FOUND!")
                return False
                
            ws = wb["Portfolio Values 2025"]
            print(f"✅ Portfolio Values 2025 sheet found")
            print(f"   Max row: {ws.max_row}")
            print(f"   Max column: {ws.max_column}")
            
            # Examine EVERY cell in the first 10 rows and 10 columns
            print(f"\n🔍 COMPLETE CELL ANALYSIS (First 10x10):")
            for row in range(1, min(11, ws.max_row + 1)):
                print(f"   Row {row}:")
                for col in range(1, min(11, ws.max_column + 1)):
                    cell_val = ws.cell(row=row, column=col).value
                    if cell_val is not None and cell_val != "":
                        print(f"     Col {col}: '{cell_val}' (type: {type(cell_val).__name__})")
                    else:
                        print(f"     Col {col}: [EMPTY]")
            
            # Find the actual structure
            print(f"\n🔍 STRUCTURE DETECTION:")
            
            # Look for date headers anywhere
            date_locations = []
            for row in range(1, min(6, ws.max_row + 1)):
                for col in range(1, min(20, ws.max_column + 1)):
                    cell_val = ws.cell(row=row, column=col).value
                    if cell_val and "/" in str(cell_val):
                        date_locations.append((row, col, cell_val))
            
            print(f"📅 Date headers found:")
            for row, col, val in date_locations:
                print(f"   Row {row}, Col {col}: {val}")
            
            # Look for account names
            account_locations = []
            keywords = ['ETRADE', 'SCHWAB', '401', 'TOTAL']
            for row in range(1, ws.max_row + 1):
                for col in range(1, min(5, ws.max_column + 1)):
                    cell_val = ws.cell(row=row, column=col).value
                    if cell_val and any(keyword in str(cell_val).upper() for keyword in keywords):
                        account_locations.append((row, col, cell_val))
            
            print(f"🏢 Account names found:")
            for row, col, val in account_locations:
                print(f"   Row {row}, Col {col}: {val}")
            
            return True
            
        except Exception as e:
            print(f"❌ Error in sheet analysis: {e}")
            traceback.print_exc()
            return False
    
    def test_update_attempt(self):
        """Attempt an ultra-simple update to see what happens"""
        try:
            print(f"\n🧪 TESTING SIMPLE UPDATE")
            print("=" * 40)
            
            # Load workbook
            wb = openpyxl.load_workbook(self.target_file)
            ws = wb["Portfolio Values 2025"]
            
            # Try to find the simplest place to add data
            print(f"🔍 Finding where to add new data...")
            
            # Test adding to cell B1 (simple test)
            current_b1 = ws.cell(row=1, column=2).value
            print(f"   Current B1 value: '{current_b1}'")
            
            # Add test data
            test_date = datetime.now().strftime('%m/%d/%Y')
            print(f"   Attempting to set B1 to: '{test_date}'")
            ws.cell(row=1, column=2, value=test_date)
            
            # Add test account value
            print(f"   Attempting to set B4 to: 999999")
            ws.cell(row=4, column=2, value=999999)
            
            # Save the file
            print(f"💾 Saving workbook...")
            wb.save(self.target_file)
            print(f"   ✅ Workbook saved successfully")
            
            # Verify the changes were saved
            print(f"🔍 Verifying changes...")
            wb2 = openpyxl.load_workbook(self.target_file)
            ws2 = wb2["Portfolio Values 2025"]
            
            new_b1 = ws2.cell(row=1, column=2).value
            new_b4 = ws2.cell(row=4, column=2).value
            
            print(f"   B1 after save: '{new_b1}'")
            print(f"   B4 after save: '{new_b4}'")
            
            if str(new_b1) == test_date:
                print(f"   ✅ Date was saved correctly!")
            else:
                print(f"   ❌ Date was NOT saved correctly")
                
            if new_b4 == 999999:
                print(f"   ✅ Test value was saved correctly!")
                return True
            else:
                print(f"   ❌ Test value was NOT saved correctly")
                return False
                
        except Exception as e:
            print(f"❌ Error in test update: {e}")
            traceback.print_exc()
            return False
    
    def run_ultra_debug(self):
        """Run comprehensive debugging"""
        
        print("🔍 ULTRA-DETAILED PORTFOLIO VALUES DEBUGGING")
        print("=" * 65)
        print("This will show EVERY step in extreme detail")
        print("=" * 65)
        
        # Step 1: Analyze sheet structure
        if not self.ultra_debug_sheet_analysis():
            return False
        
        # Step 2: Test simple update
        print(f"\n" + "="*65)
        success = self.test_update_attempt()
        
        print(f"\n🎯 DEBUGGING RESULTS:")
        print(f"   Sheet analysis: ✅ Completed")
        print(f"   Simple update test: {'✅ SUCCESS' if success else '❌ FAILED'}")
        
        if success:
            print(f"\n✅ The sheet CAN be updated!")
            print(f"   The issue was in the logic, not the file access")
        else:
            print(f"\n❌ The sheet CANNOT be updated!")
            print(f"   There may be a file permission or structure issue")
        
        return success

if __name__ == "__main__":
    print(f"Starting ultra-detailed debugging...")
    debugger = UltraDebugPortfolioUpdater()
    success = debugger.run_ultra_debug()
    
    input("\nPress Enter to close...")
