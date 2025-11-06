import openpyxl

def debug_dividend_data():
    """Debug the dividend data format issue"""
    
    workbook_path = "outputs/Dividends_2025.xlsx"
    wb = openpyxl.load_workbook(workbook_path)
    div_sheet = wb["Accounts Div historical yield"]
    
    print("DEBUGGING DIVIDEND DATA CONVERSION")
    print("=" * 50)
    
    # Test with ABR from Etrade IRA (row 3)
    ticker = div_sheet.cell(row=3, column=1).value
    beginning_yield = div_sheet.cell(row=3, column=15).value
    current_yield = div_sheet.cell(row=3, column=16).value
    
    print(f"Ticker: {ticker}")
    print(f"Beginning yield (col 15): {beginning_yield} ({type(beginning_yield)})")
    print(f"Current yield (col 16): {current_yield} ({type(current_yield)})")
    
    # Apply conversion logic
    current_val = float(current_yield) if current_yield else 0
    beginning_val = float(beginning_yield) if beginning_yield else 0
    
    print(f"\nAfter float conversion:")
    print(f"Current: {current_val}")
    print(f"Beginning: {beginning_val}")
    
    # Convert current to percentage
    if current_val < 1.0 and current_val > 0:
        current_val *= 100
        print(f"Current after *100: {current_val}")
    
    # Check reduction
    reduction_threshold = beginning_val * 0.90
    print(f"\nReduction check:")
    print(f"Beginning: {beginning_val}%")
    print(f"Current: {current_val}%") 
    print(f"90% threshold: {reduction_threshold}%")
    print(f"Is reduction? {current_val < reduction_threshold}")
    
    if beginning_val > 0 and current_val < reduction_threshold:
        reduction_pct = ((beginning_val - current_val) / beginning_val) * 100
        print(f"Reduction percentage: {reduction_pct:.1f}%")
    
    # Check if it would be added to yields list
    print(f"\nWould be added to yields list? {current_val > 0}")
    
    wb.close()

if __name__ == "__main__":
    debug_dividend_data()
