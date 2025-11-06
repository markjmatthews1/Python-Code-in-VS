import sys
import os
from openpyxl import load_workbook

def examine_schwab_section():
    """Debug the Schwab IRA section structure"""
    excel_file = "outputs/Dividends_2025.xlsx"
    
    if not os.path.exists(excel_file):
        print(f"ERROR: Excel file not found: {excel_file}")
        return
    
    try:
        wb = load_workbook(excel_file)
        ws = wb['Accounts Div historical yield']
        
        print("SCHWAB IRA SECTION ANALYSIS")
        print("=" * 50)
        
        # Examine rows 35-50 to understand the structure
        for row in range(35, 51):
            col_a = ws.cell(row=row, column=1).value
            col_b = ws.cell(row=row, column=2).value
            col_c = ws.cell(row=row, column=3).value
            
            print(f"Row {row:2d}: A='{col_a}' | B='{col_b}' | C='{col_c}'")
            
            # Stop if we see too many empty rows
            if not col_a and not col_b and not col_c:
                continue
        
        print("\n" + "=" * 50)
        print("FULL ROW SCAN FOR GROUP DIVIDERS")
        print("=" * 50)
        
        for row in range(1, 60):
            cell_a = ws.cell(row=row, column=1).value
            if cell_a and any(keyword in str(cell_a).upper() for keyword in ['ETRADE', 'SCHWAB']):
                print(f"Row {row:2d}: '{cell_a}'")
        
    except Exception as e:
        print(f"ERROR: {e}")
        import traceback
        traceback.print_exc()

if __name__ == "__main__":
    examine_schwab_section()