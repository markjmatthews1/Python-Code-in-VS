"""
Fix Portfolio Values and Color Issues

Issues to fix:
1. Portfolio Values: Missing 401K value, incorrect row 10 calculation 
2. Color coding: Wrong green (#00FF00 should be #90EE90)
3. Estimated Income: Apply color coding to rows 4-9

Expected Portfolio Total: $526,619.21 (not $657,700.45)
"""
import openpyxl
from openpyxl.styles import PatternFill, Font
from openpyxl.styles.numbers import FORMAT_CURRENCY_USD_SIMPLE

# CORRECT color codes
CORRECT_GREEN = '90EE90'   # Light green for increases
RED_COLOR = 'FF7C80'       # Red for decreases
YELLOW_COLOR = 'FFFF00'    # Yellow for same

def fix_portfolio_and_colors():
    """Fix Portfolio Values 401K issue and correct all color coding"""
    excel_file = r"c:\Users\mjmat\Python Code in VS\dividend_tracker\DividendTrackerApp\outputs\Dividends_2025.xlsx"
    
    try:
        wb = openpyxl.load_workbook(excel_file)
        
        print("🔧 COMPREHENSIVE PORTFOLIO & COLOR FIX")
        print("=" * 50)
        
        # FIX 1: Portfolio Values 2025 Sheet
        if "Portfolio Values 2025" in wb.sheetnames:
            ws_portfolio = wb["Portfolio Values 2025"]
            
            print("\n💰 FIXING PORTFOLIO VALUES 2025:")
            
            max_col = ws_portfolio.max_column
            current_col_letter = openpyxl.utils.get_column_letter(max_col)
            
            print(f"   Working with column {current_col_letter}")
            
            # Check current values in the latest column
            account_values = {}
            total_should_be = 0
            
            for row in range(4, 10):  # Check rows 4-9 for accounts
                account_name = ws_portfolio.cell(row=row, column=1).value
                current_value = ws_portfolio.cell(row=row, column=max_col).value
                
                if account_name:
                    account_values[account_name] = current_value
                    print(f"   Row {row} ({account_name}): {current_value}")
                    
                    if current_value and isinstance(current_value, (int, float)):
                        total_should_be += current_value
            
            # Add 401K value if missing (you mentioned entering $125,882.01)
            k401_value = 125882.01
            k401_found = False
            
            for account_name in account_values:
                if '401' in str(account_name) or 'Retirement' in str(account_name):
                    k401_found = True
                    break
            
            if not k401_found:
                # Find 401K row and add the value
                for row in range(4, 15):  # Extended search
                    account_name = ws_portfolio.cell(row=row, column=1).value
                    if account_name and ('401' in str(account_name) or 'Retirement' in str(account_name)):
                        ws_portfolio.cell(row=row, column=max_col).value = k401_value
                        total_should_be += k401_value
                        print(f"   ✅ Added 401K value: ${k401_value:,.2f} to row {row}")
                        break
            
            print(f"   📊 Expected total: ${total_should_be:,.2f}")
            
            # Fix row 10 total calculation
            for row in range(8, 15):  # Look for Total row
                account_name = ws_portfolio.cell(row=row, column=1).value
                if account_name and 'total' in str(account_name).lower():
                    ws_portfolio.cell(row=row, column=max_col).value = total_should_be
                    ws_portfolio.cell(row=row, column=max_col).font = Font(name='Arial', size=12, bold=True)
                    print(f"   ✅ Fixed row {row} total: ${total_should_be:,.2f}")
                    break
        
        # FIX 2: Correct Color Coding in Estimated Income 2025
        if "Estimated Income 2025" in wb.sheetnames:
            print(f"\n🎨 FIXING COLOR CODING (Using correct green #{CORRECT_GREEN}):")
            
            ws_income = wb["Estimated Income 2025"]
            max_col = ws_income.max_column
            prev_col = max_col - 1
            
            if prev_col >= 2:  # Only if there's a previous column
                colors_applied = 0
                
                # Apply to rows 4-9
                for row in range(4, 10):
                    account_name = ws_income.cell(row=row, column=1).value
                    current_val = ws_income.cell(row=row, column=max_col).value
                    prev_val = ws_income.cell(row=row, column=prev_col).value
                    
                    if (current_val is not None and prev_val is not None and
                        isinstance(current_val, (int, float)) and isinstance(prev_val, (int, float))):
                        
                        # Determine correct color
                        if current_val > prev_val:
                            color = CORRECT_GREEN  # Light green
                            change_type = "INCREASE"
                        elif current_val < prev_val:
                            color = RED_COLOR
                            change_type = "DECREASE"
                        else:
                            color = YELLOW_COLOR
                            change_type = "SAME"
                        
                        # Apply the color
                        ws_income.cell(row=row, column=max_col).fill = PatternFill(
                            start_color=color, end_color=color, fill_type='solid'
                        )
                        
                        colors_applied += 1
                        print(f"   Row {row} ({account_name}): {change_type} (#{color})")
                
                print(f"   ✅ Applied correct colors to {colors_applied} cells")
            else:
                print("   ⚠️ No previous column for color comparison")
        
        # Save all changes
        wb.save(excel_file)
        wb.close()
        
        print(f"\n✅ COMPREHENSIVE FIX COMPLETE!")
        print("📋 Fixed:")
        print("   • Portfolio Values: Added missing 401K, fixed row 10 total")
        print("   • Color coding: Applied correct green (#90EE90) to Estimated Income")
        print("   • All calculations should now be accurate")
        
        return True
        
    except Exception as e:
        print(f"❌ ERROR: {e}")
        import traceback
        traceback.print_exc()
        return False

if __name__ == "__main__":
    fix_portfolio_and_colors()
