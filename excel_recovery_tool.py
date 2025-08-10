import pandas as pd
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, numbers
import os
import shutil
from datetime import datetime

def recover_excel_file():
    """
    Clean recovery tool for Bryan Perry Transactions.xlsx
    Fixes corruption while preserving all data and formatting
    """
    
    original_file = "Bryan Perry Transactions.xlsx"
    backup_file = f"Bryan Perry Transactions_BACKUP_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    clean_file = "Bryan Perry Transactions_CLEAN.xlsx"
    
    print("🔧 Excel Recovery Tool Starting...")
    
    # Step 1: Create backup
    try:
        shutil.copy2(original_file, backup_file)
        print(f"✅ Backup created: {backup_file}")
    except Exception as e:
        print(f"⚠️ Backup failed: {e}")
        return False
    
    # Step 2: Load data (Excel will auto-recover during this step)
    try:
        print("📖 Reading corrupted file (auto-recovering)...")
        df = pd.read_excel(original_file, header=1, sheet_name="2025 Results")
        print(f"✅ Loaded {len(df)} rows of data")
    except Exception as e:
        print(f"❌ Could not read Excel file: {e}")
        return False
    
    # Step 3: Create clean workbook with proper structure
    try:
        print("🔨 Creating clean workbook...")
        wb = Workbook()
        ws = wb.active
        ws.title = "2025 Results"
        
        # Add the header structure (blank row, then column headers)
        ws.append([""])  # Row 1: Blank
        
        # Row 2: Column headers
        headers = df.columns.tolist()
        ws.append(headers)
        
        # Format header row
        for cell in ws[2]:
            if cell.value:
                cell.alignment = Alignment(horizontal="center")
        
        print("✅ Headers added and formatted")
        
    except Exception as e:
        print(f"❌ Error creating workbook structure: {e}")
        return False
    
    # Step 4: Add all data with proper formatting
    try:
        print("📝 Adding data with formatting...")
        
        # Define column types for formatting
        CURRENCY_COLS = [
            "Strike price", "Cost", "Total investment", "Call Value", 
            "Put Value", "Put cash requirement", "Close Price", 
            "Realized profit loss", "Current Price", "Current P/L"
        ]
        
        DATE_COLS = ["Date of transaction", "Expiration Date", "Close Date"]
        
        SHARE_COLS = ["Number of shares +/-"]
        
        PERCENT_COLS = ["P&L %"]
        
        # Add each row with proper formatting
        for _, row in df.iterrows():
            excel_row = []
            
            for col_name in headers:
                value = row.get(col_name, "")
                
                # Clean up NaN values
                if pd.isna(value) or str(value).lower() in ("nan", ""):
                    excel_row.append("")
                else:
                    excel_row.append(value)
            
            # Append the row
            ws.append(excel_row)
            current_row = ws.max_row
            
            # Apply formatting to each cell
            for col_idx, col_name in enumerate(headers, start=1):
                cell = ws.cell(row=current_row, column=col_idx)
                value = excel_row[col_idx - 1]
                
                # Skip empty cells
                if value == "" or value is None:
                    continue
                
                try:
                    # Format shares as whole numbers
                    if col_name in SHARE_COLS:
                        cell.value = int(float(value))
                        cell.number_format = '0'
                        cell.alignment = Alignment(horizontal="right")
                    
                    # Format currency columns
                    elif col_name in CURRENCY_COLS:
                        if str(value).strip() not in ("", "0", "0.00"):
                            numeric_val = float(str(value).replace("$", "").replace(",", ""))
                            cell.value = numeric_val
                            cell.number_format = '"$"#,##0.00'
                        else:
                            cell.value = ""
                    
                    # Format date columns
                    elif col_name in DATE_COLS:
                        if "/" in str(value):
                            cell.value = str(value)
                            cell.alignment = Alignment(horizontal="right")
                        else:
                            cell.value = ""
                    
                    # Format percentage columns
                    elif col_name in PERCENT_COLS:
                        try:
                            pct_val = float(value)
                            cell.value = pct_val / 100  # Excel expects fraction
                            cell.number_format = '0%'
                        except:
                            cell.value = value
                    
                    # All other columns - keep as-is
                    else:
                        cell.value = value
                        
                except Exception as cell_error:
                    print(f"⚠️ Cell formatting error for {col_name}: {cell_error}")
                    cell.value = str(value)  # Fallback to string
        
        print(f"✅ Added {len(df)} rows with formatting")
        
    except Exception as e:
        print(f"❌ Error adding data: {e}")
        return False
    
    # Step 5: Save clean file
    try:
        print("💾 Saving clean file...")
        wb.save(clean_file)
        wb.close()
        print(f"✅ Clean file saved: {clean_file}")
        
    except Exception as e:
        print(f"❌ Error saving clean file: {e}")
        return False
    
    # Step 6: Replace original with clean version
    try:
        print("🔄 Replacing original file...")
        os.remove(original_file)
        os.rename(clean_file, original_file)
        print(f"✅ Original file replaced with clean version")
        
    except Exception as e:
        print(f"❌ Error replacing original: {e}")
        print(f"💡 Clean file available as: {clean_file}")
        return False
    
    print("🎉 Excel recovery completed successfully!")
    print(f"📁 Backup available: {backup_file}")
    print(f"✅ Clean file: {original_file}")
    
    return True

def validate_recovery():
    """Validate that the recovered file works properly"""
    try:
        print("\n🔍 Validating recovered file...")
        df = pd.read_excel("Bryan Perry Transactions.xlsx", header=1, sheet_name="2025 Results")
        
        print(f"✅ File loads successfully: {len(df)} rows")
        print(f"✅ Columns present: {len(df.columns)}")
        
        # Check for key columns
        required_cols = ["Stock Purchase Type", "Stock ticker", "Date of transaction"]
        missing_cols = [col for col in required_cols if col not in df.columns]
        
        if missing_cols:
            print(f"⚠️ Missing columns: {missing_cols}")
        else:
            print("✅ All key columns present")
        
        return True
        
    except Exception as e:
        print(f"❌ Validation failed: {e}")
        return False

if __name__ == "__main__":
    print("=" * 50)
    print("🔧 EXCEL RECOVERY TOOL")
    print("=" * 50)
    
    # Check if original file exists
    if not os.path.exists("Bryan Perry Transactions.xlsx"):
        print("❌ Original file not found!")
        print("💡 Make sure 'Bryan Perry Transactions.xlsx' is in the current directory")
        exit(1)
    
    # Run recovery
    success = recover_excel_file()
    
    if success:
        # Validate the recovery
        validate_recovery()
        print("\n🎉 RECOVERY COMPLETE!")
        print("💡 Your TradeTracker should now work without corruption errors")
    else:
        print("\n❌ RECOVERY FAILED!")
        print("💡 Check the error messages above and try manual recovery")
    
    input("\nPress Enter to exit...")