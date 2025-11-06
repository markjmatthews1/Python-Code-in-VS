"""
Update Estimated Income 2025 with realistic dividend estimates
Uses account values and practical yield estimates based on portfolio composition
"""

import sys
import os
sys.path.append(os.path.dirname(os.path.abspath(__file__)))

from etrade_auth import get_etrade_session
import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment
import requests
import time
from datetime import datetime

def get_etrade_account_values():
    """Get E*TRADE account values for dividend calculations"""
    print("🔄 Getting E*TRADE account values...")
    
    try:
        # Import the dividend tracker's E*TRADE API
        sys.path.append(r'c:\Users\mjmat\Python Code in VS\dividend_tracker\DividendTrackerApp')
        from modules.etrade_account_api import ETRADEAccountAPI
        
        etrade_api = ETRADEAccountAPI()
        accounts = etrade_api.get_account_list()
        
        if not accounts:
            print("❌ No E*TRADE accounts found")
            return {}
        
        account_data = {}
        
        for account in accounts:
            account_id = account.get('accountIdKey', '')
            account_type = account.get('accountType', '')
            account_desc = account.get('accountDesc', '')
            
            # Skip futures account ending in 7650
            if account_id.endswith('7650'):
                print(f"   ⏭️ Skipping futures account: {account_desc}")
                continue
            
            # Determine account name
            if 'IRA' in account_type.upper() or 'ROLLOVER' in account_type.upper():
                account_name = 'E*TRADE IRA'
            else:
                account_name = 'E*TRADE Taxable'
            
            print(f"   📊 Processing {account_name} ({account_desc})")
            
            positions = etrade_api.get_account_positions(account_id)
            if not positions:
                print(f"   ⚠️ No positions found for {account_name}")
                account_data[account_name] = {'total_value': 0.0, 'dividend_stocks': []}
                continue
            
            total_value = 0.0
            dividend_stocks = []
            
            # Analyze positions for dividend stock identification
            for position in positions:
                symbol = position.get('symbolDescription', '').upper().strip()
                quantity = position.get('quantity', 0)
                market_value = position.get('marketValue', 0)
                
                if quantity > 0 and market_value > 0 and symbol:
                    total_value += market_value
                    
                    # Identify known dividend stocks from your portfolio
                    # Based on the positions I saw in the previous run
                    known_dividend_stocks = [
                        'ABR',   # Arbor Realty Trust - REIT
                        'QDTE',  # High dividend ETF
                        'PDI',   # PIMCO Dynamic Income
                        'OFS',   # OFS Capital Corp
                        'NHS',   # Neuberger Berman
                        'QYLD',  # Global X NASDAQ Covered Call
                        'RYLD',  # Global X Russell Covered Call  
                        'AGNC',  # AGNC Investment Corp - REIT
                        'ACP',   # Avenue Capital
                        'DSL',   # DoubleLine Income Solutions
                        'DX',    # Dynex Capital - REIT
                        'EIC',   # Eagle Point Income
                        'ECC',   # Eagle Point Credit
                        'ARI',   # Apollo Commercial Real Estate
                        'MORT',  # VanEck Mortgage REIT Income ETF
                        'CHMI',  # Cherry Hill Mortgage Investment
                    ]
                    
                    if symbol in known_dividend_stocks:
                        dividend_stocks.append({
                            'symbol': symbol,
                            'value': market_value,
                            'quantity': quantity
                        })
                        print(f"      💰 Dividend stock: {symbol} = ${market_value:,.2f}")
                    else:
                        print(f"      📈 Growth stock: {symbol} = ${market_value:,.2f}")
            
            account_data[account_name] = {
                'total_value': total_value,
                'dividend_stocks': dividend_stocks
            }
            
            print(f"   ✅ {account_name} Total Value: ${total_value:,.2f}")
            
        return account_data
        
    except Exception as e:
        print(f"❌ E*TRADE API error: {e}")
        return {}

def get_schwab_account_values():
    """Get Schwab account values for dividend calculations"""  
    print("🔄 Getting Schwab account values...")
    
    try:
        # Import existing Schwab auth
        from Schwab_auth import get_valid_access_token
        
        access_token = get_valid_access_token()
        if not access_token:
            print("❌ No valid Schwab access token")
            return {}
        
        headers = {
            'Authorization': f'Bearer {access_token}',
            'Accept': 'application/json'
        }
        
        # Get account numbers
        accounts_url = 'https://api.schwabapi.com/v1/accounts/accountNumbers'
        accounts_response = requests.get(accounts_url, headers=headers, timeout=30)
        
        if accounts_response.status_code != 200:
            print(f"❌ Schwab accounts error: {accounts_response.status_code}")
            return {}
        
        accounts_data = accounts_response.json()
        account_data = {}
        
        for account in accounts_data:
            account_number = account['accountNumber']
            account_type = account.get('type', 'Unknown')
            
            # Determine account name
            if 'IRA' in account_type.upper():
                account_name = 'Schwab IRA'
            else:
                account_name = 'Schwab Individual'
            
            print(f"   📊 Processing {account_name} ({account_number})")
            
            # Get positions
            positions_url = f'https://api.schwabapi.com/v1/accounts/{account_number}/positions'
            pos_response = requests.get(positions_url, headers=headers, timeout=30)
            
            if pos_response.status_code != 200:
                print(f"   ⚠️ Error getting positions: {pos_response.status_code}")
                account_data[account_name] = {'total_value': 0.0}
                continue
            
            positions_data = pos_response.json()
            positions = positions_data.get('securitiesAccount', {}).get('positions', [])
            
            total_value = 0.0
            for position in positions:
                market_value = position.get('marketValue', 0)
                total_value += market_value
            
            account_data[account_name] = {'total_value': total_value}
            print(f"   ✅ {account_name} Total Value: ${total_value:,.2f}")
        
        return account_data
        
    except Exception as e:
        print(f"❌ Schwab API error: {e}")
        return {}

def calculate_dividend_estimates(etrade_data, schwab_data):
    """Calculate realistic dividend estimates based on portfolio composition"""
    print("🧮 Calculating dividend estimates...")
    
    estimates = {}
    
    # E*TRADE calculations
    for account_name, data in etrade_data.items():
        total_value = data['total_value']
        dividend_stocks = data['dividend_stocks']
        
        if total_value == 0:
            estimates[account_name] = 0.0
            continue
        
        # Calculate dividend stock value
        dividend_stock_value = sum(stock['value'] for stock in dividend_stocks)
        growth_stock_value = total_value - dividend_stock_value
        
        print(f"   📊 {account_name}:")
        print(f"      Total Value: ${total_value:,.2f}")
        print(f"      Dividend Stocks: ${dividend_stock_value:,.2f}")
        print(f"      Growth Stocks: ${growth_stock_value:,.2f}")
        
        # Apply realistic yields based on stock types
        dividend_income = 0.0
        
        # High dividend stocks (REITs, covered calls, etc.) - 8-12% yield
        high_yield_symbols = ['QYLD', 'RYLD', 'ABR', 'AGNC', 'PDI', 'OFS', 'NHS', 'QDTE', 'MORT', 'CHMI', 'DX', 'ARI']
        moderate_yield_symbols = ['ACP', 'DSL', 'EIC', 'ECC']
        
        for stock in dividend_stocks:
            symbol = stock['symbol']
            value = stock['value']
            
            if symbol in high_yield_symbols:
                # High dividend yield stocks - 9% average
                dividend_income += value * 0.09
                print(f"         {symbol}: ${value:,.2f} × 9% = ${value * 0.09:,.2f}")
            elif symbol in moderate_yield_symbols:
                # Moderate dividend yield - 6% average  
                dividend_income += value * 0.06
                print(f"         {symbol}: ${value:,.2f} × 6% = ${value * 0.06:,.2f}")
            else:
                # Other dividend stocks - 4% average
                dividend_income += value * 0.04
                print(f"         {symbol}: ${value:,.2f} × 4% = ${value * 0.04:,.2f}")
        
        # Growth stocks - minimal dividend (0.5%)
        if growth_stock_value > 0:
            growth_dividend = growth_stock_value * 0.005
            dividend_income += growth_dividend
            print(f"         Growth stocks: ${growth_stock_value:,.2f} × 0.5% = ${growth_dividend:,.2f}")
        
        estimates[account_name] = dividend_income
        print(f"      📈 {account_name} Annual Dividend Estimate: ${dividend_income:,.2f}")
    
    # Schwab calculations
    for account_name, data in schwab_data.items():
        total_value = data['total_value']
        
        if total_value == 0:
            estimates[account_name] = 0.0
            continue
        
        # Use account-specific yield estimates
        if 'IRA' in account_name:
            # IRA accounts tend to be more dividend-focused
            yield_rate = 0.045  # 4.5%
        else:
            # Individual accounts tend to be more growth-focused
            yield_rate = 0.035  # 3.5%
        
        dividend_income = total_value * yield_rate
        estimates[account_name] = dividend_income
        
        print(f"   📊 {account_name}: ${total_value:,.2f} × {yield_rate*100}% = ${dividend_income:,.2f}")
    
    return estimates

def update_estimated_income_sheet():
    """Update Estimated Income 2025 sheet with realistic dividend estimates"""
    excel_path = r'c:\Users\mjmat\Python Code in VS\dividend_tracker\DividendTrackerApp\outputs\Dividends_2025.xlsx'
    
    print("📈 Getting account values and calculating realistic dividend estimates...")
    
    # Get account values
    etrade_data = get_etrade_account_values()
    schwab_data = get_schwab_account_values()
    
    if not etrade_data and not schwab_data:
        print("❌ No account data retrieved from APIs")
        return
    
    # Calculate realistic dividend estimates
    estimates = calculate_dividend_estimates(etrade_data, schwab_data)
    
    print("\n📊 REALISTIC DIVIDEND ESTIMATES SUMMARY:")
    print("=" * 55)
    
    total_annual = 0.0
    
    for account, estimate in estimates.items():
        print(f"{account}: ${estimate:,.2f}/year")
        total_annual += estimate
    
    monthly_total = total_annual / 12
    print(f"\nTOTAL ANNUAL: ${total_annual:,.2f}")
    print(f"MONTHLY AVERAGE: ${monthly_total:,.2f}")
    
    # Update Excel sheet
    print(f"\n📝 Updating Excel sheet: {excel_path}")
    
    try:
        wb = openpyxl.load_workbook(excel_path)
        ws = wb['Estimated Income 2025']
        
        # Update account rows with realistic estimates
        if 'E*TRADE IRA' in estimates:
            ws.cell(row=4, column=44, value=estimates['E*TRADE IRA'])
            print(f"   Row 4 (E*TRADE IRA): ${estimates['E*TRADE IRA']:,.2f}")
        
        if 'E*TRADE Taxable' in estimates:
            ws.cell(row=5, column=44, value=estimates['E*TRADE Taxable'])
            print(f"   Row 5 (E*TRADE Taxable): ${estimates['E*TRADE Taxable']:,.2f}")
        
        if 'Schwab IRA' in estimates:
            ws.cell(row=6, column=44, value=estimates['Schwab IRA'])
            print(f"   Row 6 (Schwab IRA): ${estimates['Schwab IRA']:,.2f}")
        
        if 'Schwab Individual' in estimates:
            ws.cell(row=7, column=44, value=estimates['Schwab Individual'])
            print(f"   Row 7 (Schwab Individual): ${estimates['Schwab Individual']:,.2f}")
        
        # Update monthly total
        ws.cell(row=9, column=44, value=monthly_total)
        print(f"   Row 9 (Monthly Total): ${monthly_total:,.2f}")
        
        # Apply formatting
        soft_green = PatternFill(start_color='00FA71', end_color='00FA71', fill_type='solid')
        soft_red = PatternFill(start_color='FF7C80', end_color='FF7C80', fill_type='solid')
        
        for row in [4, 5, 6, 7, 9]:
            cell = ws.cell(row=row, column=44)
            if cell.value and cell.value > 0:
                cell.fill = soft_green
            else:
                cell.fill = soft_red
            
            # Set column width
            ws.column_dimensions[cell.column_letter].width = 15
        
        wb.save(excel_path)
        print(f"✅ Excel updated successfully!")
        
    except Exception as e:
        print(f"❌ Excel update error: {e}")

if __name__ == "__main__":
    print("🚀 REALISTIC DIVIDEND ESTIMATES UPDATE")
    print("📊 Using portfolio composition analysis")
    print("=" * 50)
    update_estimated_income_sheet()
    print("\n✅ Update complete!")
