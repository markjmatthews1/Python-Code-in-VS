#!/usr/bin/env python3
"""
Calculate yearly estimated dividend income for E*TRADE Taxable account
using ticker yield data from IRA account and current positions.
"""

import json
import sys
import os
from datetime import datetime
import openpyxl
from openpyxl.utils import get_column_letter

# Add the dividend tracker path for E*TRADE API
sys.path.append(r'c:\Users\mjmat\Python Code in VS\dividend_tracker\DividendTrackerApp')
from modules.etrade_account_api import ETRADEAccountAPI

def load_ticker_yields():
    """Load ticker yield data from the IRA dividend data file."""
    try:
        with open('actual_ira_dividend_data_20250825.json', 'r') as f:
            data = json.load(f)
        
        ticker_yields = {}
        for symbol, info in data.get('all_current_holdings', {}).items():
            if info.get('has_dividend', False):
                ticker_yields[symbol] = {
                    'yield': info.get('yield', 0.0),
                    'dividend_per_share': info.get('dividend_per_share', 0.0),
                    'annual_dividend': info.get('dividend_per_share', 0.0) * 4  # Assume quarterly
                }
        
        # Add manual entries for tickers not in IRA but known to pay dividends
        ticker_yields['ARI'] = {
            'yield': 11.5,  # Apollo Commercial Real Estate Finance - typical REIT yield
            'dividend_per_share': 0.35,
            'annual_dividend': 1.40
        }
        
        print(f"✅ Loaded yield data for {len(ticker_yields)} dividend-paying tickers (including manual entries)")
        return ticker_yields
    except Exception as e:
        print(f"❌ Error loading ticker yields: {e}")
        return {}

def get_etrade_taxable_positions():
    """Get current positions from E*TRADE Taxable account."""
    try:
        etrade_api = ETRADEAccountAPI()
        accounts = etrade_api.get_account_list()
        
        if not accounts:
            print("❌ No E*TRADE accounts found")
            return []
        
        print(f"📊 Found {len(accounts)} E*TRADE accounts")
        
        # Find the taxable account ending in 744285
        target_account_key = None
        for account in accounts:
            account_id = account.get('accountId', '')
            if account_id.endswith('744285'):
                target_account_key = account.get('accountIdKey', '')
                print(f"✅ Found Taxable account: {account_id}")
                break
        
        if not target_account_key:
            print("❌ Could not find Taxable account ending in 744285")
            return []
        
        # Get positions for this account
        positions = etrade_api.get_account_positions(target_account_key)
        
        if not positions:
            print("❌ No positions found in Taxable account")
            return []
        
        print(f"📈 Found {len(positions)} positions in Taxable account")
        
        return positions
        
    except Exception as e:
        print(f"❌ Error getting E*TRADE positions: {e}")
        return []

def calculate_dividend_income(positions, ticker_yields):
    """Calculate total annual dividend income from positions."""
    total_annual_income = 0.0
    dividend_breakdown = {}
    
    print("\n💰 CALCULATING ANNUAL DIVIDEND INCOME:")
    print("=" * 60)
    
    for position in positions:
        try:
            # Extract position info from E*TRADE API format
            product = position.get('Product', {})
            symbol = product.get('symbol', 'UNKNOWN')
            quantity = float(position.get('quantity', 0))
            current_value = float(position.get('marketValue', 0))
            
            # Skip if no quantity or not a stock
            if quantity <= 0 or product.get('securityType') != 'EQ':
                continue
            
            # Check if we have dividend data for this ticker
            if symbol in ticker_yields:
                yield_data = ticker_yields[symbol]
                yield_percent = yield_data['yield'] / 100.0  # Convert percentage to decimal
                
                # Calculate annual dividend income using market value and yield
                annual_income = current_value * yield_percent
                total_annual_income += annual_income
                
                dividend_breakdown[symbol] = {
                    'quantity': quantity,
                    'current_value': current_value,
                    'yield_percent': yield_data['yield'],
                    'annual_income': annual_income
                }
                
                print(f"📊 {symbol}: ${current_value:,.2f} × {yield_data['yield']:.2f}% = ${annual_income:,.2f}")
            else:
                print(f"⚠️  {symbol}: ${current_value:,.2f} (no dividend data)")
                
        except Exception as e:
            print(f"❌ Error processing position {symbol}: {e}")
            continue
    
    print("=" * 60)
    print(f"🎯 TOTAL ANNUAL DIVIDEND INCOME: ${total_annual_income:,.2f}")
    
    return total_annual_income, dividend_breakdown

def format_dividend_cell(ws, row, col, value):
    """Apply standard formatting to dividend income cells."""
    from openpyxl.styles import PatternFill, Font
    
    # Set the value
    ws.cell(row=row, column=col, value=value)
    
    # Apply Arial 12 font
    arial_font = Font(name='Arial', size=12)
    ws.cell(row=row, column=col).font = arial_font
    
    # Apply currency formatting
    ws.cell(row=row, column=col).number_format = '$#,##0.00'
    
    # Apply light red background color (#FF7C80)
    fill = PatternFill(start_color='FF7C80', end_color='FF7C80', fill_type='solid')
    ws.cell(row=row, column=col).fill = fill
    
    return ws.cell(row=row, column=col)

def update_excel_sheet(annual_income):
    """Update the Estimated Income 2025 sheet with the calculated amount."""
    try:
        excel_path = r'c:\Users\mjmat\Python Code in VS\dividend_tracker\DividendTrackerApp\outputs\Dividends_2025.xlsx'
        wb = openpyxl.load_workbook(excel_path)
        
        if 'Estimated Income 2025' not in wb.sheetnames:
            print("❌ 'Estimated Income 2025' sheet not found")
            return False
        
        ws = wb['Estimated Income 2025']
        
        # Use column AI (35) directly since we know 8/24/2025 is there
        from openpyxl.utils import column_index_from_string
        target_col = column_index_from_string('AI')  # Column 35
        
        # Verify the date is correct
        date_cell = ws.cell(row=3, column=target_col).value
        if str(date_cell).strip() != "8/24/2025":
            print(f"⚠️  Expected 8/24/2025 in column AI, found: {date_cell}")
        
        # Add the annual income with proper formatting
        format_dividend_cell(ws, 4, target_col, annual_income)
        
        wb.save(excel_path)
        print(f"✅ Updated Estimated Income 2025 sheet:")
        print(f"   Date: 8/24/2025 (Column AI)")
        print(f"   Row 4 (E*TRADE Taxable): ${annual_income:,.2f}")
        print(f"   Formatting: Arial 12, Currency ($#,##0.00), Light Red Background (#FF7C80)")
        
        return True
            
    except Exception as e:
        print(f"❌ Error updating Excel sheet: {e}")
        return False

def main():
    """Main function to calculate and update E*TRADE Taxable dividend income."""
    print("💼 CALCULATING E*TRADE TAXABLE ANNUAL DIVIDEND INCOME")
    print("=" * 65)
    
    # Load ticker yields from IRA data
    ticker_yields = load_ticker_yields()
    if not ticker_yields:
        print("❌ No ticker yield data available")
        return
    
    # Get E*TRADE Taxable positions
    positions = get_etrade_taxable_positions()
    if not positions:
        print("❌ No positions found in E*TRADE Taxable account")
        return
    
    # Calculate dividend income
    annual_income, breakdown = calculate_dividend_income(positions, ticker_yields)
    
    if annual_income > 0:
        # Update Excel sheet
        if update_excel_sheet(annual_income):
            print("\n🎉 SUCCESS! E*TRADE Taxable dividend income calculated and updated.")
        else:
            print("\n⚠️  Calculation complete but Excel update failed.")
            print(f"   Manual entry needed: ${annual_income:,.2f} in row 4")
    else:
        print("\n⚠️  No dividend income calculated - check positions and ticker data")

if __name__ == "__main__":
    main()
